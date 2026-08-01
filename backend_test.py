"""Backend API Tests for PT Mitra Karya Sarana ERP
Testing Excel Preview Features (Iteration 20d):
1. Form templates: preview-page-meta and preview-page-image endpoints
2. Excel templates: preview-page-meta and preview-page-image endpoints  
3. Inquiry attachments: page-meta and page-image for Excel files
4. BOM attachments: page-meta and page-image for Excel files
5. Download endpoints return ORIGINAL Excel files (not PDF)
6. Regression: existing drawing page-meta/page-image still work
7. Error handling: non-PDF/non-office files return 400
"""
import requests
import sys
import json
import io
import base64
from datetime import datetime
from openpyxl import Workbook

class BackendTester:
    def __init__(self, base_url="https://error-fix-dev.preview.emergentagent.com/api"):
        self.base_url = base_url
        self.session = requests.Session()  # Use session to maintain cookies
        self.tests_run = 0
        self.tests_passed = 0
        self.tests_failed = 0
        self.failures = []
        self.test_data = {}  # Store created test data for cleanup

    def log(self, msg, level="INFO"):
        """Log test messages"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] [{level}] {msg}")

    def run_test(self, name, method, endpoint, expected_status, data=None, headers=None, files=None, allow_redirects=True):
        """Run a single API test"""
        url = f"{self.base_url}/{endpoint}"
        req_headers = {}
        if headers:
            req_headers.update(headers)
        
        # Only add Content-Type for JSON requests
        if data and not files:
            req_headers['Content-Type'] = 'application/json'
        
        self.tests_run += 1
        self.log(f"Testing {name}...", "TEST")
        
        try:
            if method == 'GET':
                response = self.session.get(url, headers=req_headers, timeout=60, allow_redirects=allow_redirects)
            elif method == 'POST':
                if files:
                    response = self.session.post(url, files=files, data=data, headers=req_headers, timeout=60, allow_redirects=allow_redirects)
                else:
                    response = self.session.post(url, json=data, headers=req_headers, timeout=60, allow_redirects=allow_redirects)
            elif method == 'PUT':
                response = self.session.put(url, json=data, headers=req_headers, timeout=60, allow_redirects=allow_redirects)
            elif method == 'DELETE':
                response = self.session.delete(url, headers=req_headers, timeout=60, allow_redirects=allow_redirects)
            else:
                raise ValueError(f"Unsupported method: {method}")

            success = response.status_code == expected_status
            if success:
                self.tests_passed += 1
                self.log(f"✅ PASSED - {name} (Status: {response.status_code})", "PASS")
                try:
                    # Try to parse JSON, but return raw content if it fails (for images, PDFs, etc.)
                    if 'application/json' in response.headers.get('Content-Type', ''):
                        return True, response.json()
                    else:
                        return True, {'content': response.content, 'headers': dict(response.headers)}
                except:
                    return True, {'content': response.content, 'headers': dict(response.headers)}
            else:
                self.tests_failed += 1
                error_detail = ""
                try:
                    error_detail = response.json().get('detail', '')
                except:
                    error_detail = response.text[:200]
                msg = f"❌ FAILED - {name} - Expected {expected_status}, got {response.status_code}. Detail: {error_detail}"
                self.log(msg, "FAIL")
                self.failures.append(msg)
                return False, {}

        except Exception as e:
            self.tests_failed += 1
            msg = f"❌ FAILED - {name} - Error: {str(e)}"
            self.log(msg, "FAIL")
            self.failures.append(msg)
            return False, {}

    def login(self, username, password):
        """Login and get token"""
        self.log(f"Logging in as {username}...", "AUTH")
        success, response = self.run_test(
            f"Login as {username}",
            "POST",
            "auth/login",
            200,
            data={"username": username, "password": password}
        )
        if success:
            # Check if session has cookies
            if 'access_token' in self.session.cookies:
                self.log(f"✅ Logged in as {username}", "AUTH")
                return True
        self.log(f"❌ Login failed for {username}", "AUTH")
        return False

    def create_sample_xlsx(self, filename="test.xlsx"):
        """Create a simple Excel file for testing"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        ws['A1'] = "Item Name"
        ws['B1'] = "Quantity"
        ws['C1'] = "Unit"
        ws['A2'] = "Test Item 1"
        ws['B2'] = 10
        ws['C2'] = "Pcs"
        ws['A3'] = "Test Item 2"
        ws['B3'] = 25
        ws['C3'] = "Meter"
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def test_form_template_preview(self):
        """Test form template preview endpoints"""
        self.log("Testing form template preview endpoints", "TEST")
        
        # 1. List form templates
        success, data = self.run_test(
            "List form templates",
            "GET",
            "form-templates",
            200
        )
        
        if not success or not data:
            self.log("No form templates found, skipping form template tests", "WARN")
            return False
        
        # Get first template
        template_id = data[0].get('id')
        self.log(f"Using form template: {template_id}", "INFO")
        
        # 2. Test preview-page-meta
        success, meta = self.run_test(
            "Form template preview-page-meta",
            "GET",
            f"form-templates/{template_id}/preview-page-meta",
            200
        )
        
        if success and 'pages' in meta and 'sizes' in meta:
            self.log(f"✅ Form template page-meta: pages={meta['pages']}, sizes={len(meta.get('sizes', []))}", "PASS")
        else:
            self.log("❌ Form template page-meta missing required fields", "FAIL")
            return False
        
        # 3. Test preview-page-image
        success, img_data = self.run_test(
            "Form template preview-page-image",
            "GET",
            f"form-templates/{template_id}/preview-page-image?page=0&scale=2.0",
            200
        )
        
        if success:
            content_type = img_data.get('headers', {}).get('Content-Type', '')
            if 'image/png' in content_type:
                self.log("✅ Form template page-image returned PNG", "PASS")
                return True
            else:
                self.log(f"❌ Form template page-image wrong content type: {content_type}", "FAIL")
                return False
        
        return False

    def test_excel_template_preview(self):
        """Test Excel template preview endpoints"""
        self.log("Testing Excel template preview endpoints", "TEST")
        
        # 1. List excel templates
        success, data = self.run_test(
            "List excel templates",
            "GET",
            "excel-templates",
            200
        )
        
        if not success or not data:
            self.log("No excel templates found, attempting to create one", "WARN")
            # Try to upload a test Excel template
            xlsx_bytes = self.create_sample_xlsx("MCL_test.xlsx")
            files = {
                'file': ('MCL_test.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            form_data = {
                'code': 'MCL',
                'filename': 'MCL_test.xlsx'
            }
            success, upload_result = self.run_test(
                "Upload Excel template",
                "POST",
                "excel-templates/upload",
                200,
                data=form_data,
                files=files
            )
            
            if not success:
                self.log("Failed to create Excel template, skipping tests", "WARN")
                return False
            
            template_id = upload_result.get('id')
            self.test_data['excel_template_id'] = template_id
        else:
            # Get first active template
            template_id = data[0].get('id')
        
        self.log(f"Using Excel template: {template_id}", "INFO")
        
        # 2. Test preview-page-meta
        success, meta = self.run_test(
            "Excel template preview-page-meta",
            "GET",
            f"excel-templates/{template_id}/preview-page-meta",
            200
        )
        
        if success and 'pages' in meta and 'sizes' in meta:
            self.log(f"✅ Excel template page-meta: pages={meta['pages']}, sizes={len(meta.get('sizes', []))}", "PASS")
        else:
            self.log("❌ Excel template page-meta missing required fields", "FAIL")
            return False
        
        # 3. Test preview-page-image
        success, img_data = self.run_test(
            "Excel template preview-page-image",
            "GET",
            f"excel-templates/{template_id}/preview-page-image?page=0&scale=2.0",
            200
        )
        
        if success:
            content_type = img_data.get('headers', {}).get('Content-Type', '')
            if 'image/png' in content_type:
                self.log("✅ Excel template page-image returned PNG", "PASS")
                return True
            else:
                self.log(f"❌ Excel template page-image wrong content type: {content_type}", "FAIL")
                return False
        
        return False

    def test_inquiry_attachment_excel(self):
        """Test inquiry attachment Excel preview"""
        self.log("Testing inquiry attachment Excel preview", "TEST")
        
        # 1. Create an inquiry
        inquiry_data = {
            "title": "Test Inquiry for Excel Preview",
            "customer_name": "Test Customer",
            "project_name": "Excel Preview Test",
            "description": "Testing Excel attachment preview",
            "items": [],
            "save_as_draft": False
        }
        
        success, inquiry = self.run_test(
            "Create inquiry",
            "POST",
            "inquiries",
            200,
            data=inquiry_data
        )
        
        if not success:
            return False
        
        inquiry_id = inquiry.get('id')
        self.test_data['inquiry_id'] = inquiry_id
        self.log(f"Created inquiry: {inquiry_id}", "INFO")
        
        # 2. Upload Excel attachment
        xlsx_bytes = self.create_sample_xlsx("inquiry_test.xlsx")
        files = {
            'file': ('inquiry_test.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        form_data = {
            'slot': 'sales'
        }
        
        success, attachment = self.run_test(
            "Upload Excel attachment to inquiry",
            "POST",
            f"inquiries/{inquiry_id}/attachments",
            200,
            data=form_data,
            files=files
        )
        
        if not success:
            return False
        
        file_id = attachment.get('id')
        self.log(f"Uploaded Excel attachment: {file_id}", "INFO")
        
        # 3. Test page-meta
        success, meta = self.run_test(
            "Inquiry attachment page-meta (Excel)",
            "GET",
            f"inquiries/{inquiry_id}/attachments/{file_id}/page-meta",
            200
        )
        
        if success and 'pages' in meta and 'sizes' in meta:
            self.log(f"✅ Inquiry Excel page-meta: pages={meta['pages']}, sizes={len(meta.get('sizes', []))}", "PASS")
        else:
            self.log("❌ Inquiry Excel page-meta missing required fields", "FAIL")
            return False
        
        # 4. Test page-image
        success, img_data = self.run_test(
            "Inquiry attachment page-image (Excel)",
            "GET",
            f"inquiries/{inquiry_id}/attachments/{file_id}/page-image?page=0&scale=2.0",
            200
        )
        
        if success:
            content_type = img_data.get('headers', {}).get('Content-Type', '')
            if 'image/png' in content_type:
                self.log("✅ Inquiry Excel page-image returned PNG", "PASS")
            else:
                self.log(f"❌ Inquiry Excel page-image wrong content type: {content_type}", "FAIL")
                return False
        
        # 5. Test download returns ORIGINAL Excel
        success, download_data = self.run_test(
            "Download inquiry Excel attachment (original format)",
            "GET",
            f"inquiries/{inquiry_id}/attachments/{file_id}/download",
            200
        )
        
        if success:
            content_type = download_data.get('headers', {}).get('Content-Type', '')
            content_disp = download_data.get('headers', {}).get('Content-Disposition', '')
            if 'spreadsheet' in content_type or 'octet-stream' in content_type:
                self.log(f"✅ Download returned original Excel (Content-Type: {content_type})", "PASS")
                return True
            else:
                self.log(f"❌ Download returned wrong content type: {content_type}", "FAIL")
                return False
        
        return False

    def test_bom_attachment_excel(self):
        """Test BOM attachment Excel preview"""
        self.log("Testing BOM attachment Excel preview", "TEST")
        
        # 1. List BOMs to find one to use
        success, data = self.run_test(
            "List BOMs",
            "GET",
            "bom?limit=10",
            200
        )
        
        if not success or not data.get('items'):
            self.log("No BOMs found, skipping BOM attachment tests", "WARN")
            return False
        
        bom_id = data['items'][0].get('id')
        self.log(f"Using BOM: {bom_id}", "INFO")
        
        # 2. Upload Excel costing attachment
        xlsx_bytes = self.create_sample_xlsx("costing_test.xlsx")
        files = {
            'file': ('costing_test.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        form_data = {
            'category': 'costing',
            'remark': 'Test costing Excel'
        }
        
        success, attachment = self.run_test(
            "Upload Excel costing to BOM",
            "POST",
            f"bom/{bom_id}/attachments",
            200,
            data=form_data,
            files=files
        )
        
        if not success:
            return False
        
        attach_id = attachment.get('attachment', {}).get('id')
        self.test_data['bom_attachment_id'] = attach_id
        self.test_data['bom_id'] = bom_id
        self.log(f"Uploaded BOM Excel attachment: {attach_id}", "INFO")
        
        # 3. Test page-meta
        success, meta = self.run_test(
            "BOM attachment page-meta (Excel)",
            "GET",
            f"bom/{bom_id}/attachments/{attach_id}/page-meta",
            200
        )
        
        if success and 'pages' in meta and 'sizes' in meta:
            self.log(f"✅ BOM Excel page-meta: pages={meta['pages']}, sizes={len(meta.get('sizes', []))}", "PASS")
        else:
            self.log("❌ BOM Excel page-meta missing required fields", "FAIL")
            return False
        
        # 4. Test page-image
        success, img_data = self.run_test(
            "BOM attachment page-image (Excel)",
            "GET",
            f"bom/{bom_id}/attachments/{attach_id}/page-image?page=0&scale=2.0",
            200
        )
        
        if success:
            content_type = img_data.get('headers', {}).get('Content-Type', '')
            if 'image/png' in content_type:
                self.log("✅ BOM Excel page-image returned PNG", "PASS")
            else:
                self.log(f"❌ BOM Excel page-image wrong content type: {content_type}", "FAIL")
                return False
        
        # 5. Test download returns ORIGINAL Excel
        success, download_data = self.run_test(
            "Download BOM Excel attachment (original format)",
            "GET",
            f"bom/{bom_id}/attachments/{attach_id}/download",
            200
        )
        
        if success:
            content_type = download_data.get('headers', {}).get('Content-Type', '')
            if 'spreadsheet' in content_type or 'octet-stream' in content_type:
                self.log(f"✅ Download returned original Excel (Content-Type: {content_type})", "PASS")
                return True
            else:
                self.log(f"❌ Download returned wrong content type: {content_type}", "FAIL")
                return False
        
        return False

    def test_drawing_regression(self):
        """Test that existing drawing page-meta/page-image still work"""
        self.log("Testing drawing endpoints regression", "TEST")
        
        # List drawings to find one
        success, data = self.run_test(
            "List drawings",
            "GET",
            "drawings?limit=10",
            200
        )
        
        if not success or not data.get('items'):
            self.log("No drawings found, skipping regression test", "WARN")
            return False
        
        drawing_id = data['items'][0].get('id')
        self.log(f"Using drawing: {drawing_id}", "INFO")
        
        # Test page-meta
        success, meta = self.run_test(
            "Drawing page-meta (regression)",
            "GET",
            f"drawings/{drawing_id}/page-meta?target=mks",
            200
        )
        
        if success and 'pages' in meta and 'sizes' in meta:
            self.log(f"✅ Drawing page-meta: pages={meta['pages']}", "PASS")
        else:
            self.log("❌ Drawing page-meta failed", "FAIL")
            return False
        
        # Test page-image
        success, img_data = self.run_test(
            "Drawing page-image (regression)",
            "GET",
            f"drawings/{drawing_id}/page-image?target=mks&page=0&scale=2.0",
            200
        )
        
        if success:
            content_type = img_data.get('headers', {}).get('Content-Type', '')
            if 'image/png' in content_type:
                self.log("✅ Drawing page-image returned PNG", "PASS")
                return True
            else:
                self.log(f"❌ Drawing page-image wrong content type: {content_type}", "FAIL")
                return False
        
        return False

    def test_error_handling(self):
        """Test error handling for non-PDF/non-office files"""
        self.log("Testing error handling for invalid file types", "TEST")
        
        # Create inquiry for testing
        inquiry_data = {
            "title": "Test Error Handling",
            "customer_name": "Test Customer",
            "save_as_draft": False
        }
        
        success, inquiry = self.run_test(
            "Create inquiry for error test",
            "POST",
            "inquiries",
            200,
            data=inquiry_data
        )
        
        if not success:
            return False
        
        inquiry_id = inquiry.get('id')
        self.test_data['error_inquiry_id'] = inquiry_id
        
        # Upload a text file (not PDF/Excel)
        txt_content = b"This is a plain text file, not PDF or Excel"
        files = {
            'file': ('test.txt', txt_content, 'text/plain')
        }
        form_data = {
            'slot': 'sales'
        }
        
        success, attachment = self.run_test(
            "Upload text file to inquiry",
            "POST",
            f"inquiries/{inquiry_id}/attachments",
            200,
            data=form_data,
            files=files
        )
        
        if not success:
            return False
        
        file_id = attachment.get('id')
        
        # Try to get page-meta (should return 400)
        success, _ = self.run_test(
            "Get page-meta for text file (expect 400)",
            "GET",
            f"inquiries/{inquiry_id}/attachments/{file_id}/page-meta",
            400
        )
        
        if success:
            self.log("✅ Error handling: non-office file returned 400", "PASS")
            return True
        else:
            self.log("❌ Error handling: should return 400 for non-office files", "FAIL")
            return False

    def cleanup(self):
        """Clean up test data"""
        self.log("Cleaning up test data...", "INFO")
        
        # Note: Inquiries don't have DELETE endpoint, they use soft delete
        # Leaving test inquiries in the system (they're marked as test data)
        
        # Delete BOM attachment
        if 'bom_attachment_id' in self.test_data and 'bom_id' in self.test_data:
            self.run_test(
                "Delete BOM test attachment",
                "DELETE",
                f"bom/{self.test_data['bom_id']}/attachments/{self.test_data['bom_attachment_id']}",
                200
            )
        
        # Delete Excel template
        if 'excel_template_id' in self.test_data:
            self.run_test(
                "Delete test Excel template",
                "DELETE",
                f"excel-templates/{self.test_data['excel_template_id']}",
                200
            )

    def print_summary(self):
        """Print test summary"""
        print("\n" + "="*80)
        print("BACKEND TEST SUMMARY - Excel Preview Features")
        print("="*80)
        print(f"Total Tests Run: {self.tests_run}")
        print(f"Tests Passed: {self.tests_passed} ✅")
        print(f"Tests Failed: {self.tests_failed} ❌")
        print(f"Success Rate: {(self.tests_passed/self.tests_run*100) if self.tests_run > 0 else 0:.1f}%")
        
        if self.failures:
            print("\n" + "="*80)
            print("FAILURES:")
            print("="*80)
            for i, failure in enumerate(self.failures, 1):
                print(f"{i}. {failure}")
        
        print("="*80)

def main():
    tester = BackendTester()
    
    print("="*80)
    print("PT MITRA KARYA SARANA - BACKEND API TESTS")
    print("Testing: Excel Preview Features (LibreOffice conversion)")
    print("="*80 + "\n")
    
    # Test with super_admin (susanto / Subbalo1994)
    if not tester.login("susanto", "Subbalo1994"):
        print("❌ Failed to login as susanto. Cannot proceed with tests.")
        return 1
    
    # Test 1: Form Template Preview
    print("\n" + "-"*80)
    print("TEST 1: Form Template Preview (page-meta & page-image)")
    print("-"*80)
    tester.test_form_template_preview()
    
    # Test 2: Excel Template Preview
    print("\n" + "-"*80)
    print("TEST 2: Excel Template Preview (page-meta & page-image)")
    print("-"*80)
    tester.test_excel_template_preview()
    
    # Test 3: Inquiry Attachment Excel
    print("\n" + "-"*80)
    print("TEST 3: Inquiry Attachment Excel Preview")
    print("-"*80)
    tester.test_inquiry_attachment_excel()
    
    # Test 4: BOM Attachment Excel
    print("\n" + "-"*80)
    print("TEST 4: BOM Attachment Excel Preview")
    print("-"*80)
    tester.test_bom_attachment_excel()
    
    # Test 5: Drawing Regression
    print("\n" + "-"*80)
    print("TEST 5: Drawing Endpoints Regression")
    print("-"*80)
    tester.test_drawing_regression()
    
    # Test 6: Error Handling
    print("\n" + "-"*80)
    print("TEST 6: Error Handling (non-PDF/non-office files)")
    print("-"*80)
    tester.test_error_handling()
    
    # Cleanup
    tester.cleanup()
    
    # Print summary
    tester.print_summary()
    
    # Return exit code
    return 0 if tester.tests_failed == 0 else 1

if __name__ == "__main__":
    sys.exit(main())
