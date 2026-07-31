"""Store module: receive, issue, stock, FIFO, requests/approvals, manual receive, production issue."""
import io
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from db import db
from deps import (
    _now_iso,
    can_see_prices,
    get_current_user,
    is_admin_like,
    log_action,
    require_admin,
    require_approve_perm,
    require_store_access,
    require_store_write,
)
from services.soft_delete import NOT_DELETED_FILTER, merged, soft_delete_one, soft_delete_many
from models import (
    BulkIssueRequest,
    BulkReceiveRequest,
    IncomingGoodsRequest,
    ManualReceiveRequest,
    ProductionIssueRequest,
    StoreIssueRequest,
    StoreReceiveRequest,
    StoreRequestCreate,
    StoreRequestReview,
)

router = APIRouter(tags=["store"])


async def _sum_received_for_tx(tx_id: str) -> float:
    agg = await db.store_receipts.aggregate([
        {"$match": {"transaction_id": tx_id}},
        {"$group": {"_id": None, "total": {"$sum": "$qty_received"}}}
    ]).to_list(length=1)
    return float(agg[0]["total"]) if agg else 0.0


@router.get("/store/pending")
async def store_pending(current: dict = Depends(require_store_access)):
    """List item transaksi yang di-flag 'post_to_store' dan belum full-received."""
    txs = await db.transactions.find(
        {"post_to_store": True},
        {"_id": 0},
    ).sort("invoice_date", -1).to_list(length=5000)

    result = []
    for t in txs:
        received = await _sum_received_for_tx(t["id"])
        remaining = float(t.get("qty", 0)) - received
        if remaining > 0:
            result.append({
                "transaction_id": t["id"],
                "invoice_date": t.get("invoice_date"),
                "po_no": t.get("po_no"),
                "invoice_no": t.get("invoice_no"),
                "vendor_name": t.get("vendor_name"),
                "item_name": t.get("item_name"),
                "unit": t.get("unit"),
                "qty_po": float(t.get("qty", 0)),
                "qty_received": received,
                "qty_remaining": remaining,
                "po_date": t.get("po_date"),
                # Purchasing's upfront decision: True = will enter stock inventory when received,
                # False = log-only (still in Incoming Good report but not counted as stock).
                "should_stock": bool(t.get("should_stock", True)),
            })
    return result


@router.get("/store/pending/grouped")
async def store_pending_grouped(current: dict = Depends(require_store_access)):
    """Grouped view: 1 baris per PO (fallback invoice_no) dengan ringkasan qty."""
    items = await store_pending(current)
    groups: dict = {}
    for it in items:
        key = it.get("po_no") or f"INV:{it.get('invoice_no', '')}" or f"TX:{it.get('transaction_id')}"
        g = groups.setdefault(key, {
            "group_key": key,
            "po_no": it.get("po_no"),
            "invoice_no": it.get("invoice_no"),
            "vendor_name": it.get("vendor_name"),
            "invoice_date": it.get("invoice_date"),
            "po_date": it.get("po_date"),
            "items": [],
            "total_qty_po": 0.0,
            "total_qty_received": 0.0,
            "total_qty_remaining": 0.0,
        })
        g["items"].append(it)
        g["total_qty_po"] += it["qty_po"]
        g["total_qty_received"] += it["qty_received"]
        g["total_qty_remaining"] += it["qty_remaining"]
    return sorted(groups.values(), key=lambda x: (x["invoice_date"] or ""), reverse=True)


@router.post("/store/receive/bulk")
async def store_receive_bulk(payload: BulkReceiveRequest, current: dict = Depends(require_store_access)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    received_docs = []
    touched_tx_ids: set = set()
    for item in payload.items:
        if item.qty_received <= 0:
            continue
        tx = await db.transactions.find_one({"id": item.transaction_id})
        if not tx or not tx.get("post_to_store"):
            raise HTTPException(status_code=400, detail=f"Item {item.transaction_id} tidak valid")
        already = await _sum_received_for_tx(item.transaction_id)
        remaining = float(tx.get("qty", 0)) - already
        if item.qty_received > remaining + 1e-9:
            raise HTTPException(
                status_code=400,
                detail=f"{tx.get('item_name')}: qty terima ({item.qty_received}) > sisa ({remaining})"
            )
        add_to_stock = True if item.add_to_stock is None else bool(item.add_to_stock)
        # Store role cannot override Purchasing's decision. Use tx.should_stock as source of truth.
        # Only super_admin/admin/purchasing may override via explicit payload.
        role = current.get("role", "")
        is_super = bool(current.get("is_super_admin"))
        if not (is_super or role in ("super_admin", "admin", "purchasing")):
            add_to_stock = bool(tx.get("should_stock", True))
        doc = {
            "id": str(uuid.uuid4()),
            "transaction_id": item.transaction_id,
            "po_no": tx.get("po_no", ""),
            "invoice_no": payload.invoice_no or tx.get("invoice_no", ""),
            "vendor_name": tx.get("vendor_name", ""),
            "item_name": tx.get("item_name", ""),
            "unit": tx.get("unit", "Ea"),
            "unit_price": float(tx.get("unit_price", 0)),
            "do_number": payload.do_number or "",
            "qty_received": float(item.qty_received),
            # If not added to stock, qty_remaining = 0 (barang langsung habis pakai)
            "qty_remaining": float(item.qty_received) if add_to_stock else 0.0,
            "add_to_stock": add_to_stock,
            "receive_date": payload.receive_date,
            "note": item.note or "",
            "source": "po",
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        received_docs.append(doc)
        touched_tx_ids.add(item.transaction_id)
    if not received_docs:
        raise HTTPException(status_code=400, detail="Semua qty kosong / 0")
    await db.store_receipts.insert_many([d.copy() for d in received_docs])

    # Auto-update source transactions with invoice_no + receive_date so purchasing masterlist reflects real receive
    tx_updates: dict = {"receive_date": payload.receive_date}
    if payload.invoice_no:
        tx_updates["invoice_no"] = payload.invoice_no
    if touched_tx_ids:
        await db.transactions.update_many(
            {"id": {"$in": list(touched_tx_ids)}},
            {"$set": tx_updates},
        )

    await log_action(current, "store_receive", "store_receipt", "-", {
        "count": len(received_docs), "po_no": received_docs[0].get("po_no"),
        "do_number": payload.do_number, "invoice_no": payload.invoice_no,
        "vendor": received_docs[0].get("vendor_name"),
    })
    return {"received": len(received_docs)}


@router.post("/store/issue/bulk")
async def store_issue_bulk(payload: BulkIssueRequest, current: dict = Depends(require_store_access)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    created = []
    for it in payload.items:
        if it.qty <= 0 or not it.item_name or not it.taker_name.strip():
            continue
        batches = await db.store_receipts.find(
            {"item_name": it.item_name, "qty_remaining": {"$gt": 0}}
        ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)
        avail = sum(b.get("qty_remaining", 0) for b in batches)
        if it.qty > avail + 1e-9:
            raise HTTPException(status_code=400, detail=f"{it.item_name}: stok tidak cukup (tersedia {avail}, diminta {it.qty})")
        remain = float(it.qty)
        allocations = []
        for b in batches:
            if remain <= 1e-9:
                break
            take = min(float(b["qty_remaining"]), remain)
            allocations.append({
                "receipt_id": b["id"],
                "qty": take,
                "unit_price": float(b.get("unit_price", 0)),
                "vendor_name": b.get("vendor_name", ""),
                "receive_date": b.get("receive_date"),
            })
            await db.store_receipts.update_one({"id": b["id"]}, {"$inc": {"qty_remaining": -take}})
            remain -= take
        total_cost = sum(a["qty"] * a["unit_price"] for a in allocations)
        doc = {
            "id": str(uuid.uuid4()),
            "item_name": it.item_name,
            "unit": batches[0].get("unit", "Ea") if batches else "Ea",
            "qty": float(it.qty),
            "issue_date": it.issue_date,
            "taker_name": it.taker_name.strip(),
            "so_number": it.so_number or "",
            "note": it.note or "",
            "allocations": allocations,
            "total_cost": total_cost,
            "avg_unit_price": (total_cost / it.qty) if it.qty else 0,
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        created.append(doc)
    if not created:
        raise HTTPException(status_code=400, detail="Tidak ada item valid")
    await db.store_issuances.insert_many([d.copy() for d in created])
    await log_action(current, "store_issue", "store_issuance", "-", {
        "count": len(created), "first_item": created[0].get("item_name"), "so_number": created[0].get("so_number"),
    })
    return {"issued": len(created)}


@router.post("/store/receive")
async def store_receive(payload: StoreReceiveRequest, current: dict = Depends(require_store_access)):
    tx = await db.transactions.find_one({"id": payload.transaction_id})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaksi asal tidak ditemukan")
    if not tx.get("post_to_store"):
        raise HTTPException(status_code=400, detail="Item ini tidak di-flag ke Store")
    already = await _sum_received_for_tx(payload.transaction_id)
    remaining = float(tx.get("qty", 0)) - already
    if payload.qty_received <= 0:
        raise HTTPException(status_code=400, detail="Qty terima harus > 0")
    if payload.qty_received > remaining + 1e-9:
        raise HTTPException(
            status_code=400,
            detail=f"Qty terima ({payload.qty_received}) melebihi sisa PO ({remaining}). Over-receive tidak diizinkan."
        )

    doc = {
        "id": str(uuid.uuid4()),
        "transaction_id": payload.transaction_id,
        "po_no": tx.get("po_no", ""),
        "invoice_no": tx.get("invoice_no", ""),
        "vendor_name": tx.get("vendor_name", ""),
        "item_name": tx.get("item_name", ""),
        "unit": tx.get("unit", "Ea"),
        "unit_price": float(tx.get("unit_price", 0)),
        "do_number": payload.do_number or "",
        "qty_received": float(payload.qty_received),
        "qty_remaining": float(payload.qty_received),
        "add_to_stock": True,
        "source": "po",
        "receive_date": payload.receive_date,
        "note": payload.note or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_receipts.insert_one(doc.copy())
    await log_action(current, "store_receive", "store_receipt", doc["id"], {
        "item": doc["item_name"], "qty": doc["qty_received"], "po_no": doc["po_no"],
        "do_number": doc["do_number"],
    })
    doc.pop("_id", None)
    if not can_see_prices(current):
        doc.pop("unit_price", None)
    return doc


@router.get("/store/receipts")
async def store_receipts(current: dict = Depends(require_store_access),
                         item_name: Optional[str] = None,
                         transaction_id: Optional[str] = None):
    filt: dict = {}
    if item_name:
        filt["item_name"] = item_name
    if transaction_id:
        filt["transaction_id"] = transaction_id
    docs = await db.store_receipts.find(merged(filt, NOT_DELETED_FILTER), {"_id": 0}).sort("receive_date", -1).to_list(length=1000)
    if not can_see_prices(current):
        for d in docs:
            d.pop("unit_price", None)
    return docs


@router.get("/store/stock")
async def store_stock(current: dict = Depends(require_store_access),
                      customer_only: bool = False,
                      exclude_customer: bool = False):
    match: dict = {"qty_remaining": {"$gt": 0}}
    if customer_only:
        match["is_customer_material"] = True
    elif exclude_customer:
        match["$or"] = [{"is_customer_material": False}, {"is_customer_material": {"$exists": False}}]
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {"item": "$item_name", "customer": {"$ifNull": ["$is_customer_material", False]}},
            "qty": {"$sum": "$qty_remaining"},
            "unit": {"$first": "$unit"},
            "last_receive_date": {"$max": "$receive_date"},
            "vendors": {"$addToSet": "$vendor_name"},
            "batches": {"$sum": 1},
        }},
        {"$sort": {"_id.item": 1}},
    ]
    docs = await db.store_receipts.aggregate(pipeline).to_list(length=5000)
    return [{
        "item_name": d["_id"]["item"], "qty": d["qty"], "unit": d["unit"],
        "last_receive_date": d.get("last_receive_date"),
        "vendors": d.get("vendors", []), "batches": d.get("batches", 0),
        "is_customer_material": bool(d["_id"].get("customer")),
    } for d in docs if d["_id"].get("item")]


@router.get("/store/stock/history")
async def stock_item_history(
    item_name: str,
    is_customer_material: Optional[bool] = None,
    current: dict = Depends(require_store_access),
):
    """Return the full IN/OUT ledger for a single stock item, sorted chronologically.

    Each row:
      - date: transaction date
      - kind: "IN" or "OUT"
      - description: PO/DO/Invoice info for IN, taker note for OUT
      - ref: invoice_no (IN) or taker_name (OUT)
      - so_no: SO number if any
      - qty_in / qty_out: qty numbers
      - balance: running balance (computed client-side or here)
      - unit / vendor_name / created_by_username
    """
    if not item_name:
        raise HTTPException(status_code=400, detail="item_name wajib")
    filt: dict = {"item_name": item_name}
    if is_customer_material is not None:
        filt["is_customer_material"] = is_customer_material

    # 1. IN — from store_receipts (barang masuk)
    receipts = await db.store_receipts.find(
        merged(filt, NOT_DELETED_FILTER), {"_id": 0}
    ).sort("receive_date", 1).to_list(length=5000)

    # 2. OUT — from store_issuances (barang keluar). Issuances don't store
    # is_customer_material directly, so we key only by item_name.
    iss_filt: dict = {"item_name": item_name}
    issuances = await db.store_issuances.find(
        merged(iss_filt, NOT_DELETED_FILTER), {"_id": 0}
    ).sort("issue_date", 1).to_list(length=5000)

    rows = []
    for r in receipts:
        qty_in = float(r.get("qty_received") or 0)
        rows.append({
            "date": r.get("receive_date") or (r.get("created_at") or "")[:10],
            "kind": "IN",
            "description": r.get("vendor_name") or r.get("customer_name") or "",
            "ref": r.get("invoice_no") or r.get("po_no") or r.get("do_number") or "",
            "so_no": r.get("so_no") or "",
            "qty_in": qty_in,
            "qty_out": 0.0,
            "unit": r.get("unit") or "",
            "unit_price": float(r.get("unit_price") or 0),
            "note": r.get("note") or "",
            "added_to_stock": bool(r.get("add_to_stock", True)),
            "created_by_username": r.get("created_by_username") or "",
        })
    for iss in issuances:
        qty_out = float(iss.get("qty") or 0)
        rows.append({
            "date": iss.get("issue_date") or (iss.get("created_at") or "")[:10],
            "kind": "OUT",
            "description": iss.get("note") or "",
            "ref": iss.get("taker_name") or "",
            "so_no": iss.get("so_number") or "",
            "qty_in": 0.0,
            "qty_out": qty_out,
            "unit": iss.get("unit") or "",
            "unit_price": float(iss.get("avg_unit_price") or 0),
            "note": iss.get("note") or "",
            "added_to_stock": False,
            "created_by_username": iss.get("created_by_username") or "",
        })
    # Sort chronologically; when same day, IN first then OUT
    rows.sort(key=lambda x: (x.get("date") or "", 0 if x["kind"] == "IN" else 1))
    # Compute running balance (only IN with added_to_stock=True count toward stock)
    balance = 0.0
    for r in rows:
        if r["kind"] == "IN" and r["added_to_stock"]:
            balance += r["qty_in"]
        elif r["kind"] == "OUT":
            balance -= r["qty_out"]
        r["balance"] = balance
    return {"item_name": item_name, "count": len(rows), "rows": rows, "current_balance": balance}


# ==================== Reorder Points & Low Stock Alerts ====================
async def _current_stock_qty(item_name: str) -> float:
    """Return current total stock (qty_remaining sum) for a given item across all batches."""
    agg = await db.store_receipts.aggregate([
        {"$match": {"item_name": item_name, "qty_remaining": {"$gt": 0}}},
        {"$group": {"_id": None, "total": {"$sum": "$qty_remaining"}}},
    ]).to_list(length=1)
    return float(agg[0]["total"]) if agg else 0.0


@router.get("/store/reorder-points")
async def list_reorder_points(current: dict = Depends(require_store_access)):
    """List all configured minimum stock thresholds."""
    docs = await db.store_reorder_points.find(
        merged({}, NOT_DELETED_FILTER), {"_id": 0}
    ).sort("item_name", 1).to_list(length=5000)
    # Enrich each item with current stock
    enriched = []
    for d in docs:
        current_qty = await _current_stock_qty(d["item_name"])
        d["current_qty"] = current_qty
        d["is_below_min"] = current_qty < float(d.get("min_qty", 0))
        enriched.append(d)
    return enriched


@router.post("/store/reorder-points")
async def create_reorder_point(payload: dict, current: dict = Depends(require_store_write)):
    """Create or upsert a reorder point for an item. Payload: {item_name, min_qty, unit, note?}"""
    item_name = (payload.get("item_name") or "").strip()
    if not item_name:
        raise HTTPException(status_code=400, detail="Nama item wajib")
    try:
        min_qty = float(payload.get("min_qty") or 0)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Minimum qty harus angka")
    if min_qty < 0:
        raise HTTPException(status_code=400, detail="Minimum qty tidak boleh negatif")

    existing = await db.store_reorder_points.find_one({"item_name": item_name, "deleted_at": {"$exists": False}})
    now = _now_iso()
    if existing:
        upd = {
            "min_qty": min_qty,
            "unit": (payload.get("unit") or existing.get("unit") or "").strip(),
            "note": (payload.get("note") or "").strip(),
            "updated_at": now,
            "updated_by": current.get("username"),
        }
        await db.store_reorder_points.update_one({"id": existing["id"]}, {"$set": upd})
        await log_action(current, "update_reorder_point", "reorder_point", existing["id"], {"item": item_name, "min_qty": min_qty})
        doc = await db.store_reorder_points.find_one({"id": existing["id"]}, {"_id": 0})
        return doc
    doc = {
        "id": str(uuid.uuid4()),
        "item_name": item_name,
        "min_qty": min_qty,
        "unit": (payload.get("unit") or "").strip(),
        "note": (payload.get("note") or "").strip(),
        "created_at": now,
        "created_by": current.get("username"),
    }
    await db.store_reorder_points.insert_one(doc.copy())
    await log_action(current, "create_reorder_point", "reorder_point", doc["id"], {"item": item_name, "min_qty": min_qty})
    doc.pop("_id", None)
    return doc


@router.delete("/store/reorder-points/{rp_id}")
async def delete_reorder_point(rp_id: str, current: dict = Depends(require_store_write)):
    ok = await soft_delete_one("store_reorder_points", {"id": rp_id}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="Reorder point tidak ditemukan")
    await log_action(current, "delete_reorder_point", "reorder_point", rp_id, {})
    return {"ok": True}


@router.get("/store/low-stock")
async def low_stock_alerts(current: dict = Depends(require_store_access)):
    """Return list of items whose current stock is below their configured minimum."""
    rps = await db.store_reorder_points.find(
        merged({}, NOT_DELETED_FILTER), {"_id": 0}
    ).to_list(length=5000)
    low = []
    for rp in rps:
        current_qty = await _current_stock_qty(rp["item_name"])
        if current_qty < float(rp.get("min_qty", 0)):
            low.append({
                **rp,
                "current_qty": current_qty,
                "shortage": float(rp.get("min_qty", 0)) - current_qty,
            })
    low.sort(key=lambda x: x["shortage"], reverse=True)
    return {"count": len(low), "items": low}


@router.post("/store/issue")
async def store_issue(payload: StoreIssueRequest, current: dict = Depends(require_store_access)):
    if payload.qty <= 0:
        raise HTTPException(status_code=400, detail="Qty keluar harus > 0")
    if not payload.taker_name.strip():
        raise HTTPException(status_code=400, detail="Nama pengambil wajib diisi")

    batches = await db.store_receipts.find(
        {"item_name": payload.item_name, "qty_remaining": {"$gt": 0}}
    ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)

    total_available = sum(b.get("qty_remaining", 0) for b in batches)
    if payload.qty > total_available + 1e-9:
        raise HTTPException(
            status_code=400,
            detail=f"Stok tidak cukup. Tersedia {total_available}, diminta {payload.qty}."
        )

    remaining_to_take = float(payload.qty)
    allocations = []
    for b in batches:
        if remaining_to_take <= 1e-9:
            break
        take = min(float(b["qty_remaining"]), remaining_to_take)
        allocations.append({
            "receipt_id": b["id"],
            "qty": take,
            "unit_price": float(b.get("unit_price", 0)),
            "vendor_name": b.get("vendor_name", ""),
            "receive_date": b.get("receive_date"),
        })
        await db.store_receipts.update_one({"id": b["id"]}, {"$inc": {"qty_remaining": -take}})
        remaining_to_take -= take

    total_cost = sum(a["qty"] * a["unit_price"] for a in allocations)
    doc = {
        "id": str(uuid.uuid4()),
        "item_name": payload.item_name,
        "unit": batches[0].get("unit", "Ea") if batches else "Ea",
        "qty": float(payload.qty),
        "issue_date": payload.issue_date,
        "taker_name": payload.taker_name.strip(),
        "so_number": payload.so_number or "",
        "note": payload.note or "",
        "allocations": allocations,
        "total_cost": total_cost,
        "avg_unit_price": (total_cost / payload.qty) if payload.qty else 0,
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_issuances.insert_one(doc.copy())
    await log_action(current, "store_issue", "store_issuance", doc["id"], {
        "item": doc["item_name"], "qty": doc["qty"], "so_number": doc["so_number"], "taker": doc["taker_name"],
    })
    doc.pop("_id", None)
    if not can_see_prices(current):
        doc.pop("total_cost", None)
        doc.pop("avg_unit_price", None)
        for a in doc.get("allocations", []):
            a.pop("unit_price", None)
    return doc


@router.get("/store/issuances")
async def list_issuances(
    current: dict = Depends(require_store_access),
    q: Optional[str] = None,
    so_number: Optional[str] = None,
    taker: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"taker_name": {"$regex": q, "$options": "i"}},
            {"so_number": {"$regex": q, "$options": "i"}},
        ]
    if so_number:
        filt["so_number"] = {"$regex": so_number, "$options": "i"}
    if taker:
        filt["taker_name"] = {"$regex": taker, "$options": "i"}
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        filt["issue_date"] = d

    total = await db.store_issuances.count_documents(merged(filt, NOT_DELETED_FILTER))
    cursor = db.store_issuances.find(merged(filt, NOT_DELETED_FILTER), {"_id": 0}).sort("issue_date", -1).skip((page - 1) * page_size).limit(page_size)
    items = await cursor.to_list(length=page_size)
    hide_price = not can_see_prices(current)
    if hide_price:
        for d in items:
            d.pop("total_cost", None)
            d.pop("avg_unit_price", None)
            for a in d.get("allocations", []):
                a.pop("unit_price", None)
    return {"total": total, "page": page, "page_size": page_size, "items": items, "prices_visible": not hide_price}


@router.get("/store/report/xlsx")
async def store_report_xlsx(
    current: dict = Depends(require_store_access),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    if not can_see_prices(current):
        raise HTTPException(status_code=403, detail="Tidak berwenang melihat harga di laporan Store")
    filt: dict = {}
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        filt["issue_date"] = d
    issuances = await db.store_issuances.find(filt, {"_id": 0}).sort("issue_date", 1).to_list(length=100000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pengeluaran Stok"
    headers = ["Nomor SO", "Tgl Keluar", "Nama Barang", "Qty", "Unit", "Unit Price (FIFO)", "Total Price", "Pengambil", "Vendor Asal", "Note"]
    ws.append(headers)
    for iss in issuances:
        for a in iss.get("allocations", []):
            ws.append([
                iss.get("so_number", ""), iss.get("issue_date", ""), iss.get("item_name", ""),
                float(a.get("qty", 0)), iss.get("unit", ""), float(a.get("unit_price", 0)),
                float(a.get("qty", 0)) * float(a.get("unit_price", 0)),
                iss.get("taker_name", ""), a.get("vendor_name", ""), iss.get("note", ""),
            ])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"laporan_pengeluaran_stok_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------- Edit/Delete Request (approval workflow) ----------------
ALLOWED_EDIT_FIELDS = {"qty", "so_number", "taker_name"}


async def _apply_qty_correction(iss: dict, new_qty: float):
    """Scale issuance qty and its FIFO allocations proportionally, syncing receipts."""
    old_qty = float(iss.get("qty") or 0)
    if old_qty <= 0:
        raise HTTPException(status_code=400, detail="Qty lama tidak valid")
    if abs(new_qty - old_qty) < 1e-9:
        return
    allocations = iss.get("allocations") or []
    if not allocations:
        # No FIFO allocations (edge case) — just update the qty field
        await db.store_issuances.update_one({"id": iss["id"]}, {"$set": {"qty": new_qty}})
        return

    # Compute new allocation qty proportionally
    scale = new_qty / old_qty
    updates = []  # list of (receipt_id, delta_to_add_back, new_alloc_qty, alloc_index)
    for idx, a in enumerate(allocations):
        old_a = float(a.get("qty") or 0)
        new_a = old_a * scale
        delta = old_a - new_a  # positive means refund to receipt, negative means take more
        updates.append((a.get("receipt_id"), delta, new_a, idx))

    # If we need MORE from a receipt (delta < 0), verify receipt has enough qty_remaining
    for rid, delta, _new_a, _i in updates:
        if delta < -1e-9:
            rec = await db.store_receipts.find_one({"id": rid})
            if not rec:
                raise HTTPException(status_code=400, detail=f"Receipt sumber (id: {rid}) sudah tidak ada, tidak bisa menambah qty")
            remaining = float(rec.get("qty_remaining") or 0)
            if remaining + 1e-9 < -delta:
                raise HTTPException(status_code=400, detail=f"Stok tersisa pada receipt sumber ('{rec.get('item_name')}' PO {rec.get('po_no')}) hanya {remaining}, tidak cukup untuk koreksi qty ke {new_qty}")

    # Apply changes
    new_allocations = list(allocations)
    total_cost = 0.0
    for rid, delta, new_a, idx in updates:
        # delta > 0 → refund to receipt (add to qty_remaining)
        # delta < 0 → consume more from receipt (subtract qty_remaining)
        await db.store_receipts.update_one({"id": rid}, {"$inc": {"qty_remaining": delta}})
        new_allocations[idx] = {**new_allocations[idx], "qty": new_a}
        total_cost += new_a * float(new_allocations[idx].get("unit_price") or 0)

    await db.store_issuances.update_one(
        {"id": iss["id"]},
        {"$set": {
            "qty": new_qty,
            "allocations": new_allocations,
            "total_cost": total_cost,
            "avg_unit_price": (total_cost / new_qty) if new_qty else 0,
        }}
    )


@router.post("/store/requests")
async def create_store_request(payload: StoreRequestCreate, current: dict = Depends(require_store_access)):
    if payload.target_type not in ("receipt", "issuance"):
        raise HTTPException(status_code=400, detail="target_type harus 'receipt' atau 'issuance'")
    if payload.action_type not in ("edit", "delete"):
        raise HTTPException(status_code=400, detail="action_type harus 'edit' atau 'delete'")
    if not payload.reason.strip():
        raise HTTPException(status_code=400, detail="Alasan wajib diisi")

    coll = db.store_receipts if payload.target_type == "receipt" else db.store_issuances
    target = await coll.find_one({"id": payload.target_id})
    if not target:
        raise HTTPException(status_code=404, detail="Data yang diajukan tidak ditemukan")

    if current.get("role") == "store" and target.get("created_by") != current["id"]:
        raise HTTPException(status_code=403, detail="Hanya bisa mengajukan koreksi untuk data milik sendiri")

    proposed = payload.proposed_changes or {}
    # New structured edit validation (issuance only for now)
    if payload.action_type == "edit":
        field = proposed.get("field")
        if field:
            if payload.target_type != "issuance":
                raise HTTPException(status_code=400, detail="Koreksi terstruktur hanya untuk data Keluar Barang")
            if field not in ALLOWED_EDIT_FIELDS:
                raise HTTPException(status_code=400, detail=f"Field '{field}' tidak dapat dikoreksi. Pilih: qty, so_number, taker_name")
            new_val = proposed.get("new_value")
            if new_val is None or (isinstance(new_val, str) and not new_val.strip()):
                raise HTTPException(status_code=400, detail="Nilai baru wajib diisi")
            if field == "qty":
                try:
                    nq = float(new_val)
                except Exception:
                    raise HTTPException(status_code=400, detail="Qty baru harus angka")
                if nq <= 0:
                    raise HTTPException(status_code=400, detail="Qty baru harus > 0")
                proposed["new_value"] = nq
            else:
                proposed["new_value"] = str(new_val).strip()
            # snapshot old value from target for reliability
            proposed["old_value"] = target.get(field)

    summary = {
        "item_name": target.get("item_name"),
        "qty": target.get("qty") or target.get("qty_received"),
        "issue_date": target.get("issue_date") or target.get("receive_date"),
        "po_no": target.get("po_no"),
        "so_number": target.get("so_number"),
        "do_number": target.get("do_number"),
        "taker_name": target.get("taker_name"),
    }
    doc = {
        "id": str(uuid.uuid4()),
        "target_type": payload.target_type,
        "target_id": payload.target_id,
        "target_summary": summary,
        "action_type": payload.action_type,
        "reason": payload.reason.strip(),
        "proposed_changes": proposed,
        "status": "pending",
        "requested_by": current["id"],
        "requested_by_username": current.get("username", ""),
        "requested_at": _now_iso(),
        "reviewed_by": None,
        "reviewed_by_username": None,
        "reviewed_at": None,
        "review_note": "",
    }
    await db.store_requests.insert_one(doc.copy())
    await log_action(current, "store_request", "store_request", doc["id"], {
        "target_type": payload.target_type, "action": payload.action_type,
        "item": summary.get("item_name"), "reason_preview": payload.reason[:80],
    })
    doc.pop("_id", None)
    return doc


@router.get("/store/requests")
async def list_store_requests(
    current: dict = Depends(get_current_user),
    status: Optional[str] = None,
    mine: bool = False,
):
    filt: dict = {}
    if status:
        filt["status"] = status
    if not is_admin_like(current) or mine:
        filt["requested_by"] = current["id"]
    docs = await db.store_requests.find(merged(filt, NOT_DELETED_FILTER), {"_id": 0}).sort("requested_at", -1).to_list(length=500)
    return docs


@router.get("/store/requests/pending-count")
async def pending_count(current: dict = Depends(require_approve_perm)):
    n = await db.store_requests.count_documents(merged({"status": "pending"}, NOT_DELETED_FILTER))
    return {"count": n}


@router.post("/store/requests/{req_id}/review")
async def review_store_request(req_id: str, payload: StoreRequestReview, current: dict = Depends(require_approve_perm)):
    req = await db.store_requests.find_one({"id": req_id})
    if not req:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail=f"Request sudah di-{req['status']}")

    new_status = "approved" if payload.approve else "rejected"

    if payload.approve and req["action_type"] == "edit" and req["target_type"] == "issuance":
        pc = req.get("proposed_changes") or {}
        field = pc.get("field")
        new_val = pc.get("new_value")
        if field in ALLOWED_EDIT_FIELDS and new_val is not None:
            iss = await db.store_issuances.find_one({"id": req["target_id"]})
            if not iss:
                raise HTTPException(status_code=404, detail="Data issuance sudah tidak ada")
            if field in ("so_number", "taker_name"):
                await db.store_issuances.update_one(
                    {"id": req["target_id"]},
                    {"$set": {field: str(new_val)}}
                )
            elif field == "qty":
                await _apply_qty_correction(iss, float(new_val))

    if payload.approve and req["action_type"] == "delete":
        if req["target_type"] == "issuance":
            iss = await db.store_issuances.find_one({"id": req["target_id"]})
            if iss:
                for a in iss.get("allocations", []):
                    await db.store_receipts.update_one(
                        {"id": a.get("receipt_id")},
                        {"$inc": {"qty_remaining": float(a.get("qty", 0))}}
                    )
                await db.store_issuances.delete_one({"id": req["target_id"]})
        elif req["target_type"] == "receipt":
            rec = await db.store_receipts.find_one({"id": req["target_id"]})
            if rec:
                consumed = float(rec.get("qty_received", 0)) - float(rec.get("qty_remaining", 0))
                if consumed > 1e-9:
                    raise HTTPException(status_code=400, detail=f"Tidak bisa hapus receipt: {consumed} unit sudah dipakai (issuance). Batalkan issuance dulu.")
                await db.store_receipts.delete_one({"id": req["target_id"]})

    await db.store_requests.update_one(
        {"id": req_id},
        {"$set": {
            "status": new_status,
            "reviewed_by": current["id"],
            "reviewed_by_username": current.get("username", ""),
            "reviewed_at": _now_iso(),
            "review_note": payload.review_note or "",
        }}
    )
    await log_action(current, "review_store_request", "store_request", req_id, {
        "decision": new_status, "target_type": req["target_type"], "action": req["action_type"],
        "item": req.get("target_summary", {}).get("item_name"),
    })
    return {"status": new_status}


# ---------------- Manual Store Receipt ----------------
@router.post("/store/receive/manual")
async def store_receive_manual(payload: ManualReceiveRequest, current: dict = Depends(require_store_write)):
    if payload.qty <= 0:
        raise HTTPException(status_code=400, detail="Qty harus > 0")
    if payload.source_type not in ("customer", "supplier"):
        raise HTTPException(status_code=400, detail="source_type harus 'customer' atau 'supplier'")
    if not payload.source_name.strip():
        raise HTTPException(status_code=400, detail="Nama customer/supplier wajib")
    if not payload.item_name.strip():
        raise HTTPException(status_code=400, detail="Nama barang wajib")
    is_customer = payload.source_type == "customer"
    doc = {
        "id": str(uuid.uuid4()),
        "transaction_id": None,
        "source": "manual",
        "source_type": payload.source_type,
        "is_customer_material": is_customer,
        "po_no": payload.po_no or "",
        "invoice_no": "",
        "vendor_name": payload.source_name.strip(),
        "customer_name": payload.source_name.strip() if is_customer else "",
        "item_name": payload.item_name.strip(),
        "unit": payload.unit or "Ea",
        "unit_price": float(payload.unit_price or 0),
        "do_number": payload.do_no or "",
        "so_no": payload.so_no or "",
        "qty_received": float(payload.qty),
        "qty_remaining": float(payload.qty),
        "receive_date": payload.receive_date,
        "mcl_done": bool(payload.mcl_done),
        "mif_done": bool(payload.mif_done),
        "note": payload.remark or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_receipts.insert_one(doc.copy())
    await log_action(current, "store_receive_manual", "store_receipt", doc["id"], {
        "source": payload.source_type, "source_name": doc["vendor_name"],
        "item": doc["item_name"], "qty": doc["qty_received"],
    })
    doc.pop("_id", None)
    if not can_see_prices(current):
        doc.pop("unit_price", None)
    return doc


@router.patch("/store/receipts/{rid}/flags")
async def update_receipt_flags(rid: str, payload: dict, current: dict = Depends(require_store_write)):
    rec = await db.store_receipts.find_one({"id": rid})
    if not rec:
        raise HTTPException(status_code=404, detail="Receipt tidak ditemukan")
    upd: dict = {}
    if "mcl_done" in payload:
        upd["mcl_done"] = bool(payload["mcl_done"])
    if "mif_done" in payload:
        upd["mif_done"] = bool(payload["mif_done"])
    # Admin/store can toggle add_to_stock directly (no request/approval).
    if "add_to_stock" in payload:
        new_val = bool(payload["add_to_stock"])
        consumed = float(rec.get("qty_received", 0)) - float(rec.get("qty_remaining", 0))
        if not new_val:
            # Turning OFF: only allowed if nothing has been consumed (else stock has been used)
            if consumed > 1e-9:
                raise HTTPException(status_code=400, detail=f"Tidak bisa hilangkan dari stok: {consumed} unit sudah dipakai (issuance).")
            upd["add_to_stock"] = False
            upd["qty_remaining"] = 0.0
        else:
            upd["add_to_stock"] = True
            upd["qty_remaining"] = float(rec.get("qty_received", 0))
    if upd:
        await db.store_receipts.update_one({"id": rid}, {"$set": upd})
        await log_action(current, "update_receipt_flags", "store_receipt", rid, upd)
    return {"ok": True, "flags": upd}


# ---------------- Input Incoming Goods (multi-item manual receipt) ----------------
@router.post("/store/incoming")
async def store_incoming(payload: IncomingGoodsRequest, current: dict = Depends(require_store_write)):
    """Multi-item manual receiving. Replaces single-item /store/receive/manual.
    Each item can be flagged add_to_stock=True (masuk stok, tracked via qty_remaining)
    or False (habis pakai, qty_remaining=0 but still logged for Incoming Goods report)."""
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    if payload.source_type not in ("customer", "supplier"):
        raise HTTPException(status_code=400, detail="source_type harus 'customer' atau 'supplier'")
    if not payload.source_name.strip():
        raise HTTPException(status_code=400, detail="Nama customer/supplier wajib")
    is_customer = payload.source_type == "customer"
    docs = []
    for it in payload.items:
        if it.qty <= 0 or not it.item_name.strip():
            continue
        add_stock = True if it.add_to_stock is None else bool(it.add_to_stock)
        docs.append({
            "id": str(uuid.uuid4()),
            "transaction_id": None,
            "source": "manual",
            "source_type": payload.source_type,
            "is_customer_material": is_customer,
            "po_no": payload.po_no or "",
            "invoice_no": "",
            "vendor_name": payload.source_name.strip(),
            "customer_name": payload.source_name.strip() if is_customer else "",
            "item_name": it.item_name.strip(),
            "unit": it.unit or "Ea",
            "unit_price": float(it.unit_price or 0),
            "do_number": payload.do_no or "",
            "so_no": it.so_no or "",
            "qty_received": float(it.qty),
            "qty_remaining": float(it.qty) if add_stock else 0.0,
            "add_to_stock": add_stock,
            "receive_date": payload.receive_date,
            "mcl_done": False,  # set later via /flags (in Incoming Goods report)
            "mif_done": False,
            "note": it.remark or "",
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        })
    if not docs:
        raise HTTPException(status_code=400, detail="Tidak ada item valid")
    await db.store_receipts.insert_many([d.copy() for d in docs])
    await log_action(current, "store_incoming", "store_receipt", "-", {
        "count": len(docs), "source": payload.source_type,
        "source_name": docs[0]["vendor_name"],
    })
    return {"received": len(docs)}


# --- Pending-PO: receipts terima dulu, PO belakangan ---
@router.get("/store/receipts/pending-po")
async def receipts_pending_po(
    vendor: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """List store_receipts yang belum punya PO / transaksi ter-link.
    Kriteria: transaction_id IS NULL AND (po_no IS NULL/'') AND source in ('manual','bulk-direct').
    Digunakan oleh Purchasing untuk konsolidasi PO belakangan."""
    filt: dict = {
        "$and": [
            {"$or": [{"transaction_id": None}, {"transaction_id": {"$exists": False}}]},
            {"$or": [{"po_no": ""}, {"po_no": {"$exists": False}}, {"po_no": None}]},
            {"source": {"$in": ["manual", "bulk-direct"]}},
            # EXCLUDE customer receipts — customer sends goods without PO, so they should
            # NOT appear in the "Tarik DO Belum PO" list. Supplier-only allowed.
            {"source_type": {"$ne": "customer"}},
        ]
    }
    if vendor and vendor.strip():
        filt["vendor_name"] = {"$regex": re.escape(vendor.strip()), "$options": "i"}
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date
        filt["receive_date"] = rng

    docs = await db.store_receipts.find(
        merged(filt, NOT_DELETED_FILTER), {"_id": 0}
    ).sort("receive_date", -1).to_list(length=500)
    return docs


@router.get("/store/incoming-report")
async def incoming_report(
    current: dict = Depends(require_store_access),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    source: Optional[str] = None,  # 'po' | 'manual'
    q: Optional[str] = None,
    add_to_stock: Optional[bool] = None,  # True = hanya yang masuk stok (Accounting need)
    page: int = 1,
    page_size: int = 100,
):
    """Unified report of ALL incoming goods (from PO purchasing + manual)."""
    filt: dict = {}
    if source in ("po", "manual"):
        filt["source"] = source
    if add_to_stock is not None:
        filt["add_to_stock"] = bool(add_to_stock)
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date
        filt["receive_date"] = rng
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"vendor_name": {"$regex": q, "$options": "i"}},
            {"po_no": {"$regex": q, "$options": "i"}},
            {"invoice_no": {"$regex": q, "$options": "i"}},
            {"do_number": {"$regex": q, "$options": "i"}},
        ]
    total = await db.store_receipts.count_documents(merged(filt, NOT_DELETED_FILTER))
    cursor = db.store_receipts.find(merged(filt, NOT_DELETED_FILTER), {"_id": 0}).sort("receive_date", -1).skip((page - 1) * page_size).limit(page_size)
    items = await cursor.to_list(length=page_size)
    if not can_see_prices(current):
        for d in items:
            d.pop("unit_price", None)
    return {"total": total, "page": page, "page_size": page_size, "items": items}


MCL_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "assets" / "mcl_template.xlsx"


def _format_date_id(iso_date: Optional[str]) -> str:
    """Format ISO date → DD-MMM-YYYY (e.g. 27-Jul-2026). English 3-letter month."""
    if not iso_date:
        return ""
    try:
        return datetime.fromisoformat(str(iso_date)[:10]).strftime("%d-%b-%Y")
    except Exception:
        return str(iso_date)


@router.get("/store/incoming/mcl/{receipt_id}")
async def print_mcl(receipt_id: str, current: dict = Depends(require_store_access)):
    """Generate a Material Control Label XLSX (per nota) using the MKS template.
    
    Groups all receipts sharing the same (vendor_name, po_no, do_number, invoice_no, receive_date)
    with the given receipt_id as the anchor. Returns a filled XLSX ready to print or save as PDF.
    """
    if not MCL_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="MCL template not found on server")

    anchor = await db.store_receipts.find_one({"id": receipt_id})
    if not anchor:
        raise HTTPException(status_code=404, detail="Receipt tidak ditemukan")

    # Build group filter: match same nota — key fields together
    group_filter: dict = {
        "vendor_name": anchor.get("vendor_name") or "",
        "receive_date": anchor.get("receive_date") or "",
    }
    for k in ("po_no", "do_number", "invoice_no"):
        v = anchor.get(k) or ""
        group_filter[k] = v

    siblings = await db.store_receipts.find(group_filter, {"_id": 0}).to_list(length=1000)
    if not siblings:
        siblings = [anchor]

    # Load template & fill
    wb = load_workbook(str(MCL_TEMPLATE_PATH))
    ws = wb["RECEIVED MATERIAL"] if "RECEIVED MATERIAL" in wb.sheetnames else wb.active

    # Add PT name box next to logo (top-left area). Merge D1:G3 for the company name box.
    from openpyxl.styles import Alignment, Border, Side, Font as XLFont
    thin = Side(style="thin", color="000000")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    try:
        # Merge D1:G3 if not already merged
        rng = "D1:G3"
        already = any(str(m) == rng for m in ws.merged_cells.ranges)
        if not already:
            ws.merge_cells(rng)
        pt_cell = ws["D1"]
        pt_cell.value = "PT. MITRA KARYA SARANA"
        pt_cell.font = XLFont(name="Arial", size=14, bold=True)
        pt_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Apply border to all cells within the merged range
        for row_i in range(1, 4):
            for col_i in range(4, 8):  # D=4 .. G=7
                ws.cell(row=row_i, column=col_i).border = box_border
    except Exception:
        pass

    # Header values (labels are at B6:C6 / D6=":" / value at E6..)
    ws["E6"] = anchor.get("do_number") or "-"
    ws["E7"] = anchor.get("po_no") or "-"
    ws["E8"] = anchor.get("vendor_name") or "-"

    # Item rows: A12..A26 already contain 1..15
    MAX_ROWS = 15
    if len(siblings) > MAX_ROWS:
        # Only first 15 items fit on the template; note it in a header cell
        pass

    for i, rec in enumerate(siblings[:MAX_ROWS]):
        r = 12 + i
        ws.cell(row=r, column=2).value = rec.get("so_no") or rec.get("so_number") or ""  # B: SO No
        # C:E merged = Material Description
        desc = rec.get("item_name") or ""
        unit = rec.get("unit") or ""
        ws.cell(row=r, column=3).value = f"{desc}" + (f" ({unit})" if unit else "")
        ws.cell(row=r, column=6).value = rec.get("qty_received") or rec.get("qty") or 0  # F: Qty
        ws.cell(row=r, column=7).value = _format_date_id(rec.get("receive_date"))  # G

    # Note if overflow
    if len(siblings) > MAX_ROWS:
        ws["A28"] = f"*Catatan: dokumen ini berisi {len(siblings)} item, hanya {MAX_ROWS} pertama ditampilkan."

    # Output
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_do = (anchor.get("do_number") or anchor.get("invoice_no") or anchor.get("po_no") or "MCL").replace("/", "-").replace(" ", "_")
    filename = f"MCL_{safe_do}_{anchor.get('receive_date','')}.xlsx"

    await log_action(current, "print_mcl", "store_receipt", receipt_id, {"group_size": len(siblings)})

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/store/incoming/mcl/{receipt_id}/pdf")
async def print_mcl_pdf(receipt_id: str, download: bool = False, format: str = "pdf", current: dict = Depends(require_store_access)):
    """Generate MCL as PDF (for preview or print). Returns inline by default; ?download=1 to force download.

    Query params:
    - format=pdf (default) → PDF via LibreOffice
    - format=xlsx → substituted xlsx file (open in Excel & print — WYSIWYG, exact match to design)

    If an active custom MCL template exists in `form_templates`, use it. Otherwise fall back to legacy layout.
    """
    anchor = await db.store_receipts.find_one({"id": receipt_id})
    if not anchor:
        raise HTTPException(status_code=404, detail="Receipt tidak ditemukan")

    group_filter: dict = {
        "vendor_name": anchor.get("vendor_name") or "",
        "receive_date": anchor.get("receive_date") or "",
    }
    for k in ("po_no", "do_number", "invoice_no"):
        group_filter[k] = anchor.get(k) or ""
    siblings = await db.store_receipts.find(group_filter, {"_id": 0}).to_list(length=1000) or [anchor]

    # -------- NEW: use custom Form Template if available --------
    # Priority: (1) Excel-uploaded template, (2) Canvas-based form_templates, (3) legacy layout
    from routers.excel_templates import get_active_xlsx_bytes, render_excel_template
    xlsx_bytes = await get_active_xlsx_bytes("MCL")
    if xlsx_bytes:
        data = {
            "company_name": "PT. MITRA KARYA SARANA",
            "receive_date": anchor.get("receive_date") or "",
            "vendor_name": anchor.get("vendor_name") or "-",
            "po_no": anchor.get("po_no") or "-",
            "do_number": anchor.get("do_number") or "-",
            "invoice_no": anchor.get("invoice_no") or "-",
            "print_date": _now_iso()[:10],
            "printed_by": current.get("username", ""),
            "items": [
                {
                    "so_no": r.get("so_no") or r.get("so_number") or "",
                    "item_name": r.get("item_name") or "",
                    "qty_received": r.get("qty_received") or r.get("qty") or 0,
                    "unit": r.get("unit") or "",
                    "receive_date": _format_date_id(r.get("receive_date")),
                } for r in siblings
            ],
        }
        safe_do = (anchor.get("do_number") or anchor.get("invoice_no") or anchor.get("po_no") or "MCL").replace("/", "-").replace(" ", "_")

        # ?format=xlsx returns substituted xlsx for direct print in Excel (WYSIWYG)
        if format.lower() == "xlsx":
            xlsx_out = render_excel_template(xlsx_bytes, data, as_pdf=False)
            filename = f"MCL_{safe_do}_{anchor.get('receive_date','')}.xlsx"
            await log_action(current, "print_mcl_xlsx", "store_receipt", receipt_id,
                             {"group_size": len(siblings), "engine": "xlsx-direct"})
            return StreamingResponse(
                io.BytesIO(xlsx_out),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"', "Cache-Control": "no-store"},
            )

        # Query flag: ?format=xlsx returns substituted xlsx for direct print in Excel (WYSIWYG),
        # else return PDF (via LibreOffice — may have minor layout differences).
        from fastapi import Request as _Req
        want_xlsx = False
        try:
            import inspect
            # Read the format query from environ dict of the request scope if available
            # Simpler: use starlette-style query params on the outer function; here we look at
            # a global via `download` fallback. Instead just accept ?format=xlsx in URL.
        except Exception:
            pass
        # Read `format` from starlette request
        # (Retrieve from FastAPI's dependency chain via a lightweight parse)
        # The endpoint receives `download` bool but not the raw request; simulate by
        # returning xlsx if requested via new dedicated endpoint below.

        pdf_bytes = render_excel_template(xlsx_bytes, data, as_pdf=True)
        filename = f"MCL_{safe_do}_{anchor.get('receive_date','')}.pdf"
        await log_action(current, "print_mcl_pdf_xlsx", "store_receipt", receipt_id,
                         {"group_size": len(siblings), "engine": "xlsx"})
        disp = "attachment" if download else "inline"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'{disp}; filename="{filename}"', "Cache-Control": "no-store"},
        )

    tpl = await db.form_templates.find_one(
        {"code": "MCL", "is_active": True, "$or": [{"deleted": {"$exists": False}}, {"deleted": False}]},
        {"_id": 0}, sort=[("is_default", -1), ("updated_at", -1)],
    )
    if tpl:
        from routers.form_templates import _render_pdf
        data = {
            "company_name": "PT. MITRA KARYA SARANA",
            "receive_date": anchor.get("receive_date") or "",
            "vendor_name": anchor.get("vendor_name") or "-",
            "po_no": anchor.get("po_no") or "-",
            "do_number": anchor.get("do_number") or "-",
            "invoice_no": anchor.get("invoice_no") or "-",
            "print_date": _now_iso()[:10],
            "printed_by": current.get("username", ""),
            "items": [
                {
                    "so_no": r.get("so_no") or r.get("so_number") or "",
                    "item_name": r.get("item_name") or "",
                    "qty_received": r.get("qty_received") or r.get("qty") or 0,
                    "unit": r.get("unit") or "",
                    "receive_date": _format_date_id(r.get("receive_date")),
                } for r in siblings
            ],
        }
        pdf_bytes = _render_pdf(tpl, data)
        safe_do = (anchor.get("do_number") or anchor.get("invoice_no") or anchor.get("po_no") or "MCL").replace("/", "-").replace(" ", "_")
        filename = f"MCL_{safe_do}_{anchor.get('receive_date','')}.pdf"
        await log_action(current, "print_mcl_pdf_tpl", "store_receipt", receipt_id,
                         {"group_size": len(siblings), "template_id": tpl.get("id")})
        disp = "attachment" if download else "inline"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'{disp}; filename="{filename}"'},
        )
    # -------- END NEW --------

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    pdf_buf = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=8*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle('body', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=12)
    small_bold = ParagraphStyle('sb', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10)

    story = []

    # ---- Header: compact logo (~15x13mm) + small PT name box next to it ----
    letterhead_path = Path(__file__).resolve().parent.parent / "assets" / "letterhead.png"
    logo_img = Image(str(letterhead_path), width=15*mm, height=13*mm) if letterhead_path.exists() else Paragraph("", body_style)
    pt_name_box = Table(
        [[Paragraph("<para align='left'><b>PT MITRA KARYA SARANA</b></para>", small_bold)]],
        colWidths=[45*mm], rowHeights=[6*mm],
    )
    pt_name_box.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.75, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    header_table = Table(
        [[logo_img, pt_name_box, ""]],
        colWidths=[18*mm, 47*mm, 115*mm],
        rowHeights=[13*mm],
    )
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (0, 0), 'TOP'),
        ('VALIGN', (1, 0), (1, 0), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 3*mm))

    # ---- Title: bold underlined 18pt centered ----
    story.append(Paragraph(
        "<para align='center'><b><u><font size=18>MATERIAL CONTROL LABEL</font></u></b></para>",
        body_style,
    ))
    story.append(Spacer(1, 6*mm))

    # ---- Identification block: 3-row bordered compact table, left-aligned ----
    id_block = Table([
        [Paragraph("<b>DO No.</b>", small_bold), Paragraph("<b>:</b>", small_bold), anchor.get("do_number") or ""],
        [Paragraph("<b>PO No.</b>", small_bold), Paragraph("<b>:</b>", small_bold), anchor.get("po_no") or ""],
        [Paragraph("<b>Supplier/Customer</b>", small_bold), Paragraph("<b>:</b>", small_bold), anchor.get("vendor_name") or ""],
    ], colWidths=[35*mm, 5*mm, 60*mm], rowHeights=[6*mm, 6*mm, 6*mm], hAlign='LEFT')
    id_block.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.75, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    story.append(id_block)
    story.append(Spacer(1, 8*mm))

    # ---- Main items table with "FOR QC USE ONLY" as merged sub-header row above cols 6-7 ----
    # Column widths: NO=8mm, SO No=15mm, Material Desc=60mm, Qty=12mm, Received Date=25mm, Insp Result=25mm, Insp Date/Sig=35mm = 180mm
    col_widths = [8*mm, 15*mm, 60*mm, 12*mm, 25*mm, 25*mm, 35*mm]

    # Build rows:
    # Row 0: FOR QC USE ONLY (merged cols 5-6, only top borders on left cells so it looks free-floating)
    # Row 1: main headers
    # Rows 2..16: 15 data rows
    rows = []
    rows.append(["", "", "", "", "",
                 Paragraph("<para align='center'><b><i>FOR QC USE ONLY</i></b></para>", small_bold),
                 ""])
    rows.append(["NO", "SO No",
                 Paragraph("<para align='center'><b>Material Description</b></para>", small_bold),
                 "Qty", "Received Date",
                 Paragraph("<para align='center'><b>Inspection<br/>Result</b></para>", small_bold),
                 Paragraph("<para align='center'><b>Inspection Date /<br/>Signature</b></para>", small_bold)])
    MAX_ROWS = 15
    for i in range(MAX_ROWS):
        if i < len(siblings):
            rec = siblings[i]
            desc = rec.get("item_name") or ""
            unit = rec.get("unit") or ""
            rows.append([
                str(i + 1),
                rec.get("so_no") or rec.get("so_number") or "",
                f"{desc}" + (f" ({unit})" if unit else ""),
                str(rec.get("qty_received") or rec.get("qty") or 0),
                _format_date_id(rec.get("receive_date")),
                Paragraph("<para align='center'><b>OK / NG*</b></para>", small_bold),
                "",
            ])
        else:
            rows.append([
                str(i + 1), "", "", "", "",
                Paragraph("<para align='center'><b>OK / NG*</b></para>", small_bold),
                "",
            ])

    items_table = Table(rows, colWidths=col_widths, rowHeights=[5*mm, 11*mm] + [7*mm]*MAX_ROWS)
    items_table.setStyle(TableStyle([
        # FOR QC USE ONLY row: only cols 5-6 have borders; span them together
        ('SPAN', (5, 0), (6, 0)),
        ('BOX', (5, 0), (6, 0), 0.75, colors.black),
        # Rest of the table (from header row onwards): full grid
        ('GRID', (0, 1), (-1, -1), 0.75, colors.black),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 10),
        ('FONTSIZE', (0, 2), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
        ('ALIGN', (0, 2), (0, -1), 'CENTER'),
        ('ALIGN', (3, 2), (3, -1), 'RIGHT'),
        ('ALIGN', (4, 2), (4, -1), 'CENTER'),
        ('ALIGN', (5, 2), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 3*mm))

    # ---- Footer: reference note (italic, left-aligned) ----
    story.append(Paragraph(
        "<i>* Refer to MKS-F-QAD-002 Material Incoming Inspection</i>",
        ParagraphStyle('foot', parent=styles['Normal'], fontSize=9, textColor=colors.black),
    ))
    story.append(Spacer(1, 14*mm))

    # ---- Received by block (left-indented, not centered) ----
    receiver = Table([
        [Paragraph("<b>Received by,</b>", body_style)],
        [Paragraph("<br/><br/>", body_style)],
        [Paragraph("_______________", body_style)],
        [Paragraph("<b>Khairul</b>", body_style)],
    ], colWidths=[60*mm])
    receiver.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 30),   # indent ~30pt (matches ref ~x=105)
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story.append(receiver)

    # ---- Document control at bottom-left ----
    story.append(Spacer(1, 80*mm))
    story.append(Paragraph(
        "<font size=8>MKS-F-STR-004#Rev.00</font>",
        ParagraphStyle('doc_ctrl', parent=styles['Normal'], fontSize=8),
    ))

    doc.build(story)
    pdf_buf.seek(0)

    safe_do = (anchor.get("do_number") or anchor.get("invoice_no") or anchor.get("po_no") or "MCL").replace("/", "-").replace(" ", "_")
    filename = f"MCL_{safe_do}_{anchor.get('receive_date','')}.pdf"
    disp = "attachment" if download else "inline"

    await log_action(current, "print_mcl_pdf", "store_receipt", receipt_id,
                     {"group_size": len(siblings), "download": download})

    return StreamingResponse(
        pdf_buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'{disp}; filename="{filename}"'},
    )


@router.get("/store/incoming/mif/{receipt_id}/pdf")
async def print_mif_pdf(receipt_id: str, download: bool = False, format: str = "pdf",
                        current: dict = Depends(require_store_access)):
    """Generate MIF (Material Issue Form) via active Excel template.

    Requires a Form Template with code=MIF to be uploaded first.
    Same data model as MCL (grouped by vendor/PO/DO).
    """
    anchor = await db.store_receipts.find_one({"id": receipt_id})
    if not anchor:
        raise HTTPException(status_code=404, detail="Receipt tidak ditemukan")

    group_filter: dict = {
        "vendor_name": anchor.get("vendor_name") or "",
        "receive_date": anchor.get("receive_date") or "",
    }
    for k in ("po_no", "do_number", "invoice_no"):
        group_filter[k] = anchor.get(k) or ""
    siblings = await db.store_receipts.find(group_filter, {"_id": 0}).to_list(length=1000) or [anchor]

    from routers.excel_templates import get_active_xlsx_bytes, render_excel_template, render_excel_template_two_copies
    xlsx_bytes = await get_active_xlsx_bytes("MIF")
    if not xlsx_bytes:
        raise HTTPException(
            status_code=404,
            detail="Template MIF belum di-upload. Silakan admin upload template Excel MIF terlebih dahulu di Admin → Template Form.",
        )

    # Auto-generate MIF form number: NNN/MM/YYYY, counter reset per month.
    # Uses `mongo counters` collection keyed by year+month.
    now = datetime.now()
    counter_key = f"mif_form_no_{now.strftime('%Y%m')}"
    counter_doc = await db.counters.find_one_and_update(
        {"_id": counter_key},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = counter_doc.get("seq", 1) if counter_doc else 1
    # Re-fetch if we got the pre-increment doc
    if not counter_doc or "seq" not in counter_doc:
        seq = 1
    form_no = f"{seq:03d}/{now.strftime('%m')}/{now.strftime('%Y')}"

    data = {
        "company_name": "PT. MITRA KARYA SARANA",
        "receive_date": _format_date_id(anchor.get("receive_date")) or "",
        "issue_date": _format_date_id(anchor.get("receive_date")) or "",
        "vendor_name": anchor.get("vendor_name") or anchor.get("customer_name") or "-",
        "po_no": anchor.get("po_no") or "-",
        "do_number": anchor.get("do_number") or "-",
        "invoice_no": anchor.get("invoice_no") or "-",
        "form_no": form_no,
        "requested_by": "-",
        "print_date": _format_date_id(_now_iso()[:10]),
        "printed_by": current.get("username", ""),
        "items": [
            {
                "so_no": r.get("so_no") or r.get("so_number") or "",
                "item_name": r.get("item_name") or "",
                "qty_received": r.get("qty_received") or r.get("qty") or 0,
                "unit": r.get("unit") or "",
                "receive_date": _format_date_id(r.get("receive_date")),
            } for r in siblings
        ],
    }
    safe_do = (anchor.get("do_number") or anchor.get("invoice_no") or anchor.get("po_no") or "MIF").replace("/", "-").replace(" ", "_")
    if format.lower() == "xlsx":
        xlsx_out = render_excel_template_two_copies(xlsx_bytes, data, as_pdf=False)
        filename = f"MIF_{safe_do}_{anchor.get('receive_date','')}.xlsx"
        return StreamingResponse(
            io.BytesIO(xlsx_out),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"', "Cache-Control": "no-store"},
        )
    pdf_bytes = render_excel_template_two_copies(xlsx_bytes, data, as_pdf=True)
    filename = f"MIF_{safe_do}_{anchor.get('receive_date','')}.pdf"
    await log_action(current, "print_mif_pdf", "store_receipt", receipt_id, {"group_size": len(siblings)})
    disp = "attachment" if download else "inline"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'{disp}; filename="{filename}"', "Cache-Control": "no-store"},
    )



@router.get("/store/issuances/takers")
async def issuance_takers(current: dict = Depends(require_store_access)):
    """Distinct taker names from historical issuances for autocomplete."""
    names = await db.store_issuances.distinct("taker_name")
    return sorted([n for n in names if n])


@router.post("/store/receipts/bulk-delete")
async def bulk_delete_receipts(payload: dict, current: dict = Depends(require_admin)):
    """Admin-only direct bulk delete of store receipts.
    Rejects receipts that have been consumed (qty_remaining < qty_received) unless force=true."""
    ids = payload.get("ids") or []
    force = bool(payload.get("force"))
    if not ids:
        raise HTTPException(status_code=400, detail="Tidak ada ID yang dipilih")
    receipts = await db.store_receipts.find({"id": {"$in": ids}}).to_list(length=len(ids))
    if not receipts:
        raise HTTPException(status_code=404, detail="Receipt tidak ditemukan")
    consumed = [r for r in receipts if float(r.get("qty_received", 0)) - float(r.get("qty_remaining", 0)) > 1e-9]
    if consumed and not force:
        raise HTTPException(
            status_code=400,
            detail=(
                f"{len(consumed)} receipt sudah dipakai (issuance). Batalkan issuance dulu, "
                f"atau kirim ulang dengan force=true untuk paksa hapus."
            ),
        )
    n = await soft_delete_many("store_receipts", {"id": {"$in": ids}}, current)
    await log_action(current, "bulk_delete_receipts", "store_receipt", "-", {
        "count": n, "requested": len(ids), "forced": force,
    })
    return {"deleted": n, "forced_consumed": len(consumed) if force else 0}


@router.get("/store/report/combined-xlsx")
async def combined_stock_report_xlsx(
    current: dict = Depends(require_store_access),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """Export 1 file Excel dengan 2 sheet: Stok Masuk + Stok Keluar (untuk Accounting)."""
    # -------- Stok Masuk (add_to_stock=true) --------
    in_filt: dict = {"add_to_stock": True}
    if start_date or end_date:
        rng: dict = {}
        if start_date: rng["$gte"] = start_date
        if end_date: rng["$lte"] = end_date
        in_filt["receive_date"] = rng
    in_docs = await db.store_receipts.find(merged(in_filt, NOT_DELETED_FILTER), {"_id": 0}).sort("receive_date", -1).to_list(length=100000)

    # -------- Stok Keluar (issuances) --------
    out_filt: dict = {}
    if start_date or end_date:
        rng: dict = {}
        if start_date: rng["$gte"] = start_date
        if end_date: rng["$lte"] = end_date
        out_filt["issue_date"] = rng
    out_docs = await db.store_issuances.find(merged(out_filt, NOT_DELETED_FILTER), {"_id": 0}).sort("issue_date", -1).to_list(length=100000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Masuk & Keluar"
    # Header: kolom sama untuk semua baris (format Stok Keluar), Jenis = penanda
    ws.append([
        "Jenis", "Tanggal", "Nomor SO", "Nama Barang", "Qty", "Unit",
        "Pengambil", "Unit Price", "Total Price", "Vendor",
    ])

    total_in = 0.0
    total_out = 0.0
    row_masuk = 0
    row_keluar = 0

    # --- STOK MASUK rows (Pengambil & SO kosong) ---
    for d in in_docs:
        up = float(d.get("unit_price") or 0)
        qty = float(d.get("qty_received") or 0)
        total = up * qty if up else 0.0
        total_in += total
        row_masuk += 1
        ws.append([
            "STOK MASUK",
            d.get("receive_date", ""),
            "",                                         # Nomor SO kosong
            d.get("item_name", ""),
            qty,
            d.get("unit", ""),
            "",                                         # Pengambil kosong
            up if up else "",
            total if up else "",
            d.get("vendor_name") or d.get("customer_name", ""),
        ])

    # --- STOK KELUAR rows (dengan Pengambil & SO) ---
    for d in out_docs:
        allocs = d.get("allocations") or []
        if allocs:
            for a in allocs:
                up = float(a.get("unit_price") or 0)
                aqty = float(a.get("qty") or 0)
                total = up * aqty
                total_out += total
                row_keluar += 1
                ws.append([
                    "STOK KELUAR",
                    d.get("issue_date", ""),
                    d.get("so_number", ""),
                    d.get("item_name", ""),
                    aqty,
                    d.get("unit", ""),
                    d.get("taker_name", ""),
                    up,
                    total,
                    a.get("vendor_name", ""),
                ])
        else:
            row_keluar += 1
            ws.append([
                "STOK KELUAR",
                d.get("issue_date", ""),
                d.get("so_number", ""),
                d.get("item_name", ""),
                float(d.get("qty") or 0),
                d.get("unit", ""),
                d.get("taker_name", ""),
                "", "", "",
            ])

    # Total baris
    ws.append([])
    ws.append(["", "", "", "", "", "", "", "TOTAL MASUK:", total_in, ""])
    ws.append(["", "", "", "", "", "", "", "TOTAL KELUAR:", total_out, ""])
    ws.append(["", "", "", "", "", "", "", "SELISIH (M−K):", total_in - total_out, ""])
    ws.append([])
    ws.append([f"Periode: {start_date or '(awal)'} s/d {end_date or '(hari ini)'}"])
    ws.append([f"Baris Stok Masuk: {row_masuk}  |  Baris Stok Keluar: {row_keluar}"])
    ws.append([f"Dibuat oleh: {current.get('username','')}  |  {_now_iso()}"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"stok_gabungan_{(start_date or 'awal')}_{(end_date or 'now')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/store/incoming-report/xlsx")
async def incoming_report_xlsx(
    current: dict = Depends(require_store_access),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    source: Optional[str] = None,
    q: Optional[str] = None,
    add_to_stock: Optional[bool] = None,
):
    """Excel export of Incoming Goods report."""
    filt: dict = {}
    if source in ("po", "manual"):
        filt["source"] = source
    if add_to_stock is not None:
        filt["add_to_stock"] = bool(add_to_stock)
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date
        filt["receive_date"] = rng
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"vendor_name": {"$regex": q, "$options": "i"}},
            {"po_no": {"$regex": q, "$options": "i"}},
            {"invoice_no": {"$regex": q, "$options": "i"}},
            {"do_number": {"$regex": q, "$options": "i"}},
        ]
    docs = await db.store_receipts.find(filt, {"_id": 0}).sort("receive_date", -1).to_list(length=100000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Incoming Goods"
    headers = ["Tgl Terima", "Sumber", "Vendor/Customer", "Nama Barang", "Qty", "Unit",
               "Nomor SO", "PO No", "DO No", "Invoice No", "Ke Stok?", "MCL", "MIF", "Catatan"]
    ws.append(headers)
    for d in docs:
        src_label = "PO" if d.get("source") == "po" else ("Customer" if d.get("is_customer_material") else "Supplier")
        ws.append([
            d.get("receive_date", ""), src_label, d.get("vendor_name", ""),
            d.get("item_name", ""), float(d.get("qty_received", 0)), d.get("unit", ""),
            d.get("so_no", "") or d.get("so_number", ""),
            d.get("po_no", ""), d.get("do_number", ""), d.get("invoice_no", ""),
            "Ya" if d.get("add_to_stock", True) else "Tidak",
            "Ya" if d.get("mcl_done") else "Tidak",
            "Ya" if d.get("mif_done") else "Tidak",
            d.get("note", ""),
        ])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"incoming_goods_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/store/issuances/xlsx")
async def issuances_xlsx(
    current: dict = Depends(require_store_access),
    q: Optional[str] = None,
    so_number: Optional[str] = None,
    taker: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """Excel export of Keluar Barang (issuances). Store role: no prices; admin/finance: with FIFO cost."""
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"taker_name": {"$regex": q, "$options": "i"}},
            {"so_number": {"$regex": q, "$options": "i"}},
        ]
    if so_number:
        filt["so_number"] = {"$regex": so_number, "$options": "i"}
    if taker:
        filt["taker_name"] = {"$regex": taker, "$options": "i"}
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date
        filt["issue_date"] = rng
    docs = await db.store_issuances.find(filt, {"_id": 0}).sort("issue_date", -1).to_list(length=100000)

    show_prices = can_see_prices(current)
    wb = Workbook()
    ws = wb.active
    ws.title = "Keluar Barang"
    headers = ["Tgl Keluar", "Nama Barang", "Qty", "Unit", "Nomor SO", "Pengambil", "Catatan"]
    if show_prices:
        headers += ["Avg Unit Price (FIFO)", "Total Cost"]
    ws.append(headers)
    for d in docs:
        row = [
            d.get("issue_date", ""), d.get("item_name", ""),
            float(d.get("qty", 0)), d.get("unit", ""),
            d.get("so_number", ""), d.get("taker_name", ""),
            d.get("note", ""),
        ]
        if show_prices:
            row += [float(d.get("avg_unit_price", 0)), float(d.get("total_cost", 0))]
        ws.append(row)
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"keluar_barang_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------- Production Issue (Customer material → Produksi) ----------------
@router.post("/store/issue/production")
async def store_issue_production(payload: ProductionIssueRequest, current: dict = Depends(require_store_write)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    if not payload.taker_name.strip():
        raise HTTPException(status_code=400, detail="Nama penerima produksi wajib")
    created = []
    for it in payload.items:
        if it.qty <= 0 or not it.item_name:
            continue
        batches = await db.store_receipts.find(
            {"item_name": it.item_name, "is_customer_material": True, "qty_remaining": {"$gt": 0}}
        ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)
        avail = sum(b.get("qty_remaining", 0) for b in batches)
        if it.qty > avail + 1e-9:
            raise HTTPException(status_code=400, detail=f"{it.item_name}: stok Customer tidak cukup (tersedia {avail})")
        remain = float(it.qty)
        allocations = []
        for b in batches:
            if remain <= 1e-9:
                break
            take = min(float(b["qty_remaining"]), remain)
            allocations.append({
                "receipt_id": b["id"], "qty": take,
                "unit_price": float(b.get("unit_price", 0)),
                "vendor_name": b.get("vendor_name", ""),
                "customer_name": b.get("customer_name", ""),
                "receive_date": b.get("receive_date"),
            })
            await db.store_receipts.update_one({"id": b["id"]}, {"$inc": {"qty_remaining": -take}})
            remain -= take
        doc = {
            "id": str(uuid.uuid4()),
            "type": "production",
            "is_customer_material": True,
            "item_name": it.item_name,
            "unit": batches[0].get("unit", "Ea") if batches else "Ea",
            "qty": float(it.qty),
            "issue_date": payload.issue_date,
            "taker_name": payload.taker_name.strip(),
            "so_number": it.so_number or "",
            "note": it.note or "",
            "allocations": allocations,
            "total_cost": 0,
            "avg_unit_price": 0,
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        created.append(doc)
    if not created:
        raise HTTPException(status_code=400, detail="Tidak ada item valid")
    await db.store_issuances.insert_many([d.copy() for d in created])
    await log_action(current, "store_issue_production", "store_issuance", "-", {
        "count": len(created), "so_number": created[0].get("so_number"),
        "first_item": created[0].get("item_name"), "taker": payload.taker_name,
    })
    return {"issued": len(created)}


# ==================== Excel Export: Store Movement ====================
@router.get("/store/movements/export/xlsx")
async def export_store_movements(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current: dict = Depends(require_store_access),
):
    """Excel report: item receipts + issues within a date range."""
    filt_r: dict = {}
    filt_i: dict = {}
    if start_date:
        filt_r["receive_date"] = {"$gte": start_date}
        filt_i["issue_date"] = {"$gte": start_date}
    if end_date:
        filt_r.setdefault("receive_date", {})["$lte"] = end_date
        filt_i.setdefault("issue_date", {})["$lte"] = end_date

    receipts = await db.store_receipts.find(merged(filt_r, NOT_DELETED_FILTER), {"_id": 0}).sort("receive_date", 1).to_list(length=10000)
    issues = await db.store_issues.find(merged(filt_i, NOT_DELETED_FILTER), {"_id": 0}).sort("issue_date", 1).to_list(length=10000)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()

    # Sheet 1: Receipts
    ws1 = wb.active
    ws1.title = "Penerimaan"
    hdr = ["Tanggal Terima", "Item", "Vendor", "PO No", "Qty Terima", "Qty Sisa", "Satuan", "Batch Masuk Stok?", "Diterima Oleh"]
    for i, h in enumerate(hdr, 1):
        c = ws1.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E293B")
        c.alignment = Alignment(horizontal="center")
    for r_idx, r in enumerate(receipts, 2):
        ws1.cell(r_idx, 1, r.get("receive_date", ""))
        ws1.cell(r_idx, 2, r.get("item_name", ""))
        ws1.cell(r_idx, 3, r.get("vendor_name", ""))
        ws1.cell(r_idx, 4, r.get("po_no", ""))
        ws1.cell(r_idx, 5, r.get("qty_received", 0))
        ws1.cell(r_idx, 6, r.get("qty_remaining", 0))
        ws1.cell(r_idx, 7, r.get("unit", ""))
        ws1.cell(r_idx, 8, "Ya" if r.get("add_to_stock", True) else "Tidak")
        ws1.cell(r_idx, 9, r.get("received_by_name", ""))
    for i in range(1, 10):
        ws1.column_dimensions[chr(64 + i)].width = 18

    # Sheet 2: Issues
    ws2 = wb.create_sheet("Pengeluaran")
    hdr2 = ["Tanggal Keluar", "Item", "Qty Keluar", "Satuan", "Tujuan", "Pengambil", "Petugas Store", "Catatan"]
    for i, h in enumerate(hdr2, 1):
        c = ws2.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7F1D1D")
        c.alignment = Alignment(horizontal="center")
    for r_idx, i_doc in enumerate(issues, 2):
        ws2.cell(r_idx, 1, i_doc.get("issue_date", ""))
        ws2.cell(r_idx, 2, i_doc.get("item_name", ""))
        ws2.cell(r_idx, 3, i_doc.get("qty", 0))
        ws2.cell(r_idx, 4, i_doc.get("unit", ""))
        ws2.cell(r_idx, 5, i_doc.get("destination", ""))
        ws2.cell(r_idx, 6, i_doc.get("taker_name", ""))
        ws2.cell(r_idx, 7, i_doc.get("issued_by_name", ""))
        ws2.cell(r_idx, 8, i_doc.get("note", ""))
    for i in range(1, 9):
        ws2.column_dimensions[chr(64 + i)].width = 18

    # Sheet 3: Summary per item (net movement)
    ws3 = wb.create_sheet("Ringkasan")
    hdr3 = ["Item", "Total Masuk", "Total Keluar", "Net (Sisa Stok)", "Satuan"]
    for i, h in enumerate(hdr3, 1):
        c = ws3.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="065F46")
        c.alignment = Alignment(horizontal="center")
    item_agg: dict = {}
    for r in receipts:
        k = r.get("item_name", "")
        e = item_agg.setdefault(k, {"in": 0.0, "out": 0.0, "unit": r.get("unit", "")})
        e["in"] += float(r.get("qty_received", 0) or 0)
    for i_doc in issues:
        k = i_doc.get("item_name", "")
        e = item_agg.setdefault(k, {"in": 0.0, "out": 0.0, "unit": i_doc.get("unit", "")})
        e["out"] += float(i_doc.get("qty", 0) or 0)
    for r_idx, (k, e) in enumerate(sorted(item_agg.items()), 2):
        ws3.cell(r_idx, 1, k)
        ws3.cell(r_idx, 2, e["in"])
        ws3.cell(r_idx, 3, e["out"])
        ws3.cell(r_idx, 4, e["in"] - e["out"])
        ws3.cell(r_idx, 5, e["unit"])
    for i in range(1, 6):
        ws3.column_dimensions[chr(64 + i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"store_movement_{start_date or 'all'}_{end_date or 'now'}.xlsx"
    await log_action(current, "export_store_movements", "store", "-", {
        "start_date": start_date, "end_date": end_date,
        "receipts": len(receipts), "issues": len(issues),
    })
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
