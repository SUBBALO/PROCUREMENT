from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import uuid
import logging
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Any

import bcrypt
import jwt
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from openpyxl import Workbook, load_workbook

# ---------------- Setup ----------------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]

app = FastAPI(title="Laporan Pembelian API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ---------------- Auth helpers ----------------
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "type": "access",
               "exp": datetime.now(timezone.utc) + timedelta(hours=8)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "type": "refresh",
               "exp": datetime.now(timezone.utc) + timedelta(days=7)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=False, samesite="lax", max_age=8 * 3600, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=False, samesite="lax", max_age=7 * 24 * 3600, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        if user.get("active") is False:
            raise HTTPException(status_code=403, detail="Akun user dinonaktifkan")
        user.pop("password_hash", None)
        user.pop("_id", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def require_admin(current: dict = Depends(get_current_user)) -> dict:
    if current.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Hanya admin yang bisa mengakses")
    return current


async def require_approve_perm(current: dict = Depends(get_current_user)) -> dict:
    if current.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Hanya admin yang bisa approve")
    if "approve_store_requests" not in (current.get("perms") or []):
        raise HTTPException(status_code=403, detail="Anda tidak berwenang menyetujui permohonan Store")
    return current


async def require_write(current: dict = Depends(get_current_user)) -> dict:
    if current.get("role") == "finance":
        raise HTTPException(status_code=403, detail="Akun Finance hanya untuk view — tidak bisa mengubah data")
    return current


# ---------------- Models ----------------
class LoginRequest(BaseModel):
    username: str
    password: str

class UserCreate(BaseModel):
    username: str
    password: str
    name: Optional[str] = ""
    role: Optional[str] = "staff"  # 'admin' | 'staff' | 'store'
    perms: Optional[List[str]] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    active: Optional[bool] = None
    password: Optional[str] = None
    perms: Optional[List[str]] = None

class UserOut(BaseModel):
    id: str
    username: str
    name: str
    role: str
    active: bool
    perms: List[str] = []

class TransactionBase(BaseModel):
    invoice_date: str  # ISO date string YYYY-MM-DD
    project_no: Optional[str] = ""
    po_no: Optional[str] = ""
    vendor_name: str
    item_name: str
    qty: float
    unit: Optional[str] = "Ea"
    unit_price: float
    total_price: float
    invoice_no: Optional[str] = ""
    po_date: Optional[str] = None
    receive_date: Optional[str] = None
    notes: Optional[str] = ""
    is_compliant: Optional[bool] = True
    is_completed: Optional[bool] = True
    post_to_store: Optional[bool] = False

class TransactionCreate(TransactionBase):
    pass

class Transaction(TransactionBase):
    id: str
    created_at: str
    updated_at: str

class BulkCreateRequest(BaseModel):
    transactions: List[TransactionCreate]


# ---------------- Auth Routes ----------------
@api_router.post("/auth/login")
async def login(payload: LoginRequest, response: Response):
    username = payload.username.lower().strip()
    user = await db.users.find_one({"username": username})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Username atau password salah")
    if user.get("active") is False:
        raise HTTPException(status_code=403, detail="Akun user dinonaktifkan")
    access = create_access_token(user["id"], username)
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    await _log_action(user, "login", "auth", user["id"], {"username": username})
    return {
        "id": user["id"],
        "username": user["username"],
        "name": user.get("name", ""),
        "role": user["role"],
        "perms": user.get("perms", []),
    }

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    # Best-effort log logout if user was authenticated
    token = request.cookies.get("access_token")
    if token:
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            u = await db.users.find_one({"id": payload.get("sub")})
            if u:
                await _log_action(u, "logout", "auth", u["id"], {})
        except Exception:
            pass
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}

@api_router.get("/auth/me")
async def me(current: dict = Depends(get_current_user)):
    return {
        "id": current["id"],
        "username": current.get("username", ""),
        "name": current.get("name", ""),
        "role": current["role"],
        "perms": current.get("perms", []),
    }

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        access = create_access_token(user["id"], user.get("username", ""))
        response.set_cookie("access_token", access, httponly=True, secure=False, samesite="lax", max_age=8 * 3600, path="/")
        return {"ok": True}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")


# ---------------- Users Management (admin only) ----------------
def _sanitize_user(u: dict) -> dict:
    return {
        "id": u["id"],
        "username": u.get("username", ""),
        "name": u.get("name", ""),
        "role": u.get("role", "staff"),
        "active": u.get("active", True),
        "perms": u.get("perms", []),
        "created_at": u.get("created_at", ""),
    }

@api_router.get("/users")
async def list_users(current: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", 1).to_list(length=500)
    return [_sanitize_user(u) for u in users]

@api_router.post("/users")
async def create_user(payload: UserCreate, current: dict = Depends(require_admin)):
    username = payload.username.lower().strip()
    if not username or len(username) < 3:
        raise HTTPException(status_code=400, detail="Username minimal 3 karakter")
    if not payload.password or len(payload.password) < 6:
        raise HTTPException(status_code=400, detail="Password minimal 6 karakter")
    role = payload.role if payload.role in ("admin", "staff", "store", "finance") else "staff"
    existing = await db.users.find_one({"username": username})
    if existing:
        raise HTTPException(status_code=400, detail="Username sudah dipakai")
    user_doc = {
        "id": str(uuid.uuid4()),
        "username": username,
        "password_hash": hash_password(payload.password),
        "name": (payload.name or username).strip(),
        "role": role,
        "active": True,
        "perms": payload.perms or [],
        "created_at": _now_iso(),
    }
    await db.users.insert_one(user_doc.copy())
    await _log_action(current, "create_user", "user", user_doc["id"], {"username": username, "role": role})
    return _sanitize_user(user_doc)

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, payload: UserUpdate, current: dict = Depends(require_admin)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    updates: dict = {}
    changed: dict = {}
    if payload.name is not None:
        updates["name"] = payload.name.strip()
        changed["name"] = payload.name.strip()
    if payload.role is not None and payload.role in ("admin", "staff", "store", "finance"):
        # prevent demoting yourself
        if user_id == current["id"] and payload.role != "admin":
            raise HTTPException(status_code=400, detail="Tidak bisa demote akun sendiri")
        updates["role"] = payload.role
        changed["role"] = payload.role
    if payload.active is not None:
        # prevent disabling yourself
        if user_id == current["id"] and payload.active is False:
            raise HTTPException(status_code=400, detail="Tidak bisa menonaktifkan akun sendiri")
        updates["active"] = bool(payload.active)
        changed["active"] = bool(payload.active)
    if payload.perms is not None:
        updates["perms"] = list(payload.perms)
        changed["perms"] = list(payload.perms)
    if payload.password:
        if len(payload.password) < 6:
            raise HTTPException(status_code=400, detail="Password minimal 6 karakter")
        updates["password_hash"] = hash_password(payload.password)
        changed["password"] = "***"
    if updates:
        await db.users.update_one({"id": user_id}, {"$set": updates})
        await _log_action(current, "update_user", "user", user_id, {"target": user.get("username"), "changes": changed})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return _sanitize_user(updated)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current: dict = Depends(require_admin)):
    if user_id == current["id"]:
        raise HTTPException(status_code=400, detail="Tidak bisa hapus akun sendiri")
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    await db.users.delete_one({"id": user_id})
    await _log_action(current, "delete_user", "user", user_id, {"username": user.get("username")})
    return {"ok": True}


# ---------------- Activity Log ----------------
async def _log_action(actor: dict, action: str, entity: str, entity_id: str, details: Optional[dict] = None):
    """Fire-and-forget audit log. Errors are swallowed to not disrupt main flow."""
    try:
        await db.activity_logs.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": actor.get("id"),
            "username": actor.get("username", ""),
            "user_name": actor.get("name", ""),
            "action": action,
            "entity": entity,
            "entity_id": entity_id,
            "details": details or {},
            "timestamp": _now_iso(),
        })
    except Exception as e:
        logger.warning(f"Failed to log action {action}: {e}")

@api_router.get("/logs")
async def list_logs(
    current: dict = Depends(require_admin),
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    entity: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
):
    filt: dict = {}
    if user_id:
        filt["user_id"] = user_id
    if action:
        filt["action"] = action
    if entity:
        filt["entity"] = entity
    if start_date or end_date:
        ts: dict = {}
        if start_date:
            ts["$gte"] = start_date
        if end_date:
            ts["$lte"] = end_date + "T23:59:59"
        filt["timestamp"] = ts
    total = await db.activity_logs.count_documents(filt)
    cursor = db.activity_logs.find(filt, {"_id": 0}).sort("timestamp", -1).skip((page - 1) * page_size).limit(page_size)
    items = await cursor.to_list(length=page_size)
    return {"total": total, "page": page, "page_size": page_size, "items": items}


# ---------------- Transaction Routes ----------------
def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def _clean_doc(doc: dict) -> dict:
    doc.pop("_id", None)
    return doc

@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(payload: TransactionCreate, current: dict = Depends(require_write)):
    now = _now_iso()
    tx = payload.model_dump()
    tx["id"] = str(uuid.uuid4())
    tx["created_at"] = now
    tx["updated_at"] = now
    await db.transactions.insert_one(tx.copy())
    await _log_action(current, "create_transaction", "transaction", tx["id"], {
        "vendor": tx.get("vendor_name"), "item": tx.get("item_name"),
        "invoice_no": tx.get("invoice_no"), "total": tx.get("total_price")
    })
    return _clean_doc(tx)

@api_router.post("/transactions/bulk")
async def bulk_create(payload: BulkCreateRequest, current: dict = Depends(require_write)):
    now = _now_iso()
    docs = []
    for t in payload.transactions:
        d = t.model_dump()
        d["id"] = str(uuid.uuid4())
        d["created_at"] = now
        d["updated_at"] = now
        docs.append(d)
    if docs:
        await db.transactions.insert_many([d.copy() for d in docs])
        first = docs[0]
        await _log_action(current, "bulk_create_transaction", "transaction", "-", {
            "count": len(docs), "vendor": first.get("vendor_name"),
            "invoice_no": first.get("invoice_no"),
        })
    return {"inserted": len(docs)}

@api_router.get("/transactions")
async def list_transactions(
    current: dict = Depends(get_current_user),
    q: Optional[str] = None,
    vendor: Optional[str] = None,
    project_no: Optional[str] = None,
    po_no: Optional[str] = None,
    invoice_no: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sort_by: str = "invoice_date",
    sort_dir: str = "desc",
    page: int = 1,
    page_size: int = 50,
):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"vendor_name": {"$regex": q, "$options": "i"}},
            {"invoice_no": {"$regex": q, "$options": "i"}},
            {"project_no": {"$regex": q, "$options": "i"}},
            {"po_no": {"$regex": q, "$options": "i"}},
        ]
    if vendor:
        filt["vendor_name"] = {"$regex": vendor, "$options": "i"}
    if project_no:
        filt["project_no"] = {"$regex": project_no, "$options": "i"}
    if po_no:
        filt["po_no"] = {"$regex": po_no, "$options": "i"}
    if invoice_no:
        filt["invoice_no"] = {"$regex": invoice_no, "$options": "i"}
    if start_date or end_date:
        date_filt = {}
        if start_date:
            date_filt["$gte"] = start_date
        if end_date:
            date_filt["$lte"] = end_date
        filt["invoice_date"] = date_filt

    direction = -1 if sort_dir == "desc" else 1
    total = await db.transactions.count_documents(filt)
    cursor = db.transactions.find(filt, {"_id": 0}).sort(sort_by, direction).skip((page - 1) * page_size).limit(page_size)
    items = await cursor.to_list(length=page_size)
    return {"total": total, "page": page, "page_size": page_size, "items": items}

@api_router.get("/transactions/{tx_id}")
async def get_transaction(tx_id: str, current: dict = Depends(get_current_user)):
    doc = await db.transactions.find_one({"id": tx_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    return doc

@api_router.put("/transactions/{tx_id}", response_model=Transaction)
async def update_transaction(tx_id: str, payload: TransactionCreate, current: dict = Depends(require_write)):
    existing = await db.transactions.find_one({"id": tx_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    upd = payload.model_dump()
    upd["updated_at"] = _now_iso()
    await db.transactions.update_one({"id": tx_id}, {"$set": upd})
    doc = await db.transactions.find_one({"id": tx_id}, {"_id": 0})
    await _log_action(current, "update_transaction", "transaction", tx_id, {
        "vendor": upd.get("vendor_name"), "item": upd.get("item_name"),
        "invoice_no": upd.get("invoice_no"), "total": upd.get("total_price"),
    })
    return doc

@api_router.delete("/transactions/{tx_id}")
async def delete_transaction(tx_id: str, current: dict = Depends(require_write)):
    existing = await db.transactions.find_one({"id": tx_id})
    res = await db.transactions.delete_one({"id": tx_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    if existing:
        await _log_action(current, "delete_transaction", "transaction", tx_id, {
            "vendor": existing.get("vendor_name"), "item": existing.get("item_name"),
            "invoice_no": existing.get("invoice_no"), "total": existing.get("total_price"),
        })
    return {"ok": True}


# ---------------- Master lists ----------------
@api_router.get("/master/vendors")
async def master_vendors(current: dict = Depends(get_current_user)):
    vendors = await db.transactions.distinct("vendor_name")
    return sorted([v for v in vendors if v])

@api_router.get("/master/items")
async def master_items(current: dict = Depends(get_current_user)):
    """Return unique items with latest price and vendor."""
    pipeline = [
        {"$sort": {"invoice_date": -1, "created_at": -1}},
        {"$group": {
            "_id": "$item_name",
            "last_price": {"$first": "$unit_price"},
            "last_vendor": {"$first": "$vendor_name"},
            "last_date": {"$first": "$invoice_date"},
            "unit": {"$first": "$unit"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
        {"$limit": 5000},
    ]
    result = await db.transactions.aggregate(pipeline).to_list(length=5000)
    return [{"item_name": r["_id"], "last_price": r["last_price"], "last_vendor": r["last_vendor"],
             "last_date": r["last_date"], "unit": r.get("unit", "Ea"), "count": r["count"]} for r in result if r["_id"]]


# ---------------- Dashboard Stats ----------------
@api_router.get("/stats/summary")
async def stats_summary(current: dict = Depends(get_current_user), year: Optional[int] = None):
    match: dict = {}
    if year:
        match["invoice_date"] = {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}

    total_count = await db.transactions.count_documents(match)

    # Sum totals + top vendors + monthly totals
    agg_total = await db.transactions.aggregate([
        {"$match": match},
        {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
    ]).to_list(length=1)
    total_amount = agg_total[0]["total"] if agg_total else 0

    top_vendors = await db.transactions.aggregate([
        {"$match": match},
        {"$group": {"_id": "$vendor_name", "total": {"$sum": "$total_price"}, "count": {"$sum": 1}}},
        {"$sort": {"total": -1}},
        {"$limit": 8},
    ]).to_list(length=8)

    monthly = await db.transactions.aggregate([
        {"$match": match},
        {"$group": {"_id": {"$substr": ["$invoice_date", 0, 7]}, "total": {"$sum": "$total_price"}, "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
    ]).to_list(length=200)

    # Distinct vendor & item counts
    vendors = await db.transactions.distinct("vendor_name", match)
    items = await db.transactions.distinct("item_name", match)

    return {
        "total_transactions": total_count,
        "total_amount": total_amount,
        "unique_vendors": len(vendors),
        "unique_items": len(items),
        "top_vendors": [{"vendor": v["_id"] or "-", "total": v["total"], "count": v["count"]} for v in top_vendors],
        "monthly": [{"month": m["_id"], "total": m["total"], "count": m["count"]} for m in monthly],
    }


# ---------------- KPI Purchasing ----------------
@api_router.get("/kpi")
async def kpi_report(
    current: dict = Depends(get_current_user),
    start_date: str = Query(...),
    end_date: str = Query(...),
    ontime_grace_days: int = 7,
):
    """Compute KPI Purchasing for a date range.

    Total PO = jumlah Nomor PO unik dalam periode (baris tanpa po_no dikelompokkan via invoice_no).
    On Time = PO dianggap On Time jika seluruh receive_date <= po_date + grace_days.
    Compliance Quality = PO on-spec jika seluruh item is_compliant=true.
    PO Completion Rate = PO selesai jika seluruh item is_completed=true.
    """
    match = {"invoice_date": {"$gte": start_date, "$lte": end_date}}
    txs = await db.transactions.find(match, {"_id": 0}).to_list(length=200000)

    from collections import defaultdict
    groups: dict = defaultdict(list)
    for t in txs:
        po_no = (t.get("po_no") or "").strip()
        inv_no = (t.get("invoice_no") or "").strip()
        key = po_no if po_no else (f"INV:{inv_no}" if inv_no else f"ID:{t.get('id')}")
        groups[key].append(t)

    total_po = len(groups)
    on_time_po = 0
    compliant_po = 0
    completed_po = 0
    late_details = []

    for key, items in groups.items():
        po_on_time = True
        for it in items:
            pd = it.get("po_date")
            rd = it.get("receive_date")
            if not pd or not rd:
                po_on_time = False
                break
            try:
                pd_d = datetime.strptime(pd, "%Y-%m-%d").date()
                rd_d = datetime.strptime(rd, "%Y-%m-%d").date()
                if rd_d > pd_d + timedelta(days=ontime_grace_days):
                    po_on_time = False
                    break
            except Exception:
                po_on_time = False
                break
        if po_on_time:
            on_time_po += 1
        else:
            first = items[0]
            late_details.append({
                "po_no": key,
                "vendor": first.get("vendor_name", ""),
                "invoice_no": first.get("invoice_no", ""),
                "po_date": first.get("po_date"),
                "receive_date": first.get("receive_date"),
                "item_name": first.get("item_name", ""),
            })

        if all(it.get("is_compliant", True) for it in items):
            compliant_po += 1
        if all(it.get("is_completed", True) for it in items):
            completed_po += 1

    def pct(n, d):
        return round((n / d) * 100, 2) if d else 0.0

    on_time_pct = pct(on_time_po, total_po)
    compliant_pct = pct(compliant_po, total_po)
    completed_pct = pct(completed_po, total_po)

    score_on_time = round(on_time_pct * 0.40, 2)
    score_compliance = round(compliant_pct * 0.35, 2)
    score_completion = round(completed_pct * 0.25, 2)
    total_score = round(score_on_time + score_compliance + score_completion, 2)

    if total_score >= 90:
        category = "SANGAT BAIK"
    elif total_score >= 80:
        category = "BAIK"
    elif total_score >= 71:
        category = "CUKUP"
    else:
        category = "PERLU PERBAIKAN"

    return {
        "period": {"start_date": start_date, "end_date": end_date, "ontime_grace_days": ontime_grace_days},
        "total_po": total_po,
        "kpis": [
            {
                "no": 1, "name": "On Time Delivery",
                "description": "Persentase pengiriman barang dari supplier yang diterima tepat waktu sesuai dengan jadwal (ETA)",
                "formula_num": "Jumlah On Time Shipment", "formula_den": "Total PO",
                "target": "≥ 90%", "weight": 40,
                "numerator": on_time_po, "denominator": total_po,
                "achievement": on_time_pct, "score": score_on_time, "max_score": 40,
            },
            {
                "no": 2, "name": "Compliance Quality",
                "description": "Persentase pengiriman barang dari supplier yang diterima sesuai dengan pemesanan (spesifikasi)",
                "formula_num": "Jumlah Pembelian yang sesuai Spesifikasi", "formula_den": "Total PO",
                "target": "≥ 98%", "weight": 35,
                "numerator": compliant_po, "denominator": total_po,
                "achievement": compliant_pct, "score": score_compliance, "max_score": 35,
            },
            {
                "no": 3, "name": "PO Completion Rate",
                "description": "Persentase Purchase Order yang berhasil diselesaikan dalam periode tertentu sebagai indikator efektivitas proses procurement",
                "formula_num": "Jumlah PO selesai", "formula_den": "Total PO",
                "target": "≥ 90%", "weight": 25,
                "numerator": completed_po, "denominator": total_po,
                "achievement": completed_pct, "score": score_completion, "max_score": 25,
            },
        ],
        "total_score": total_score,
        "category": category,
        "late_details": late_details[:100],
    }




# ---------------- Store Module ----------------
STORE_ACCESS_ROLES = ("admin", "store", "finance")


async def require_store_access(current: dict = Depends(get_current_user)) -> dict:
    if current.get("role") not in STORE_ACCESS_ROLES:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    return current


async def require_store_write(current: dict = Depends(get_current_user)) -> dict:
    if current.get("role") not in ("admin", "store"):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    return current


def _can_see_prices(user: dict) -> bool:
    role = user.get("role")
    if role == "store":
        return False
    if role in ("admin", "finance"):
        return True
    return "view_store_report" in (user.get("perms") or [])


class StoreReceiveRequest(BaseModel):
    transaction_id: str
    do_number: Optional[str] = ""
    qty_received: float
    receive_date: str
    note: Optional[str] = ""


class StoreIssueRequest(BaseModel):
    item_name: str
    qty: float
    issue_date: str
    taker_name: str
    so_number: Optional[str] = ""
    note: Optional[str] = ""


async def _sum_received_for_tx(tx_id: str) -> float:
    agg = await db.store_receipts.aggregate([
        {"$match": {"transaction_id": tx_id}},
        {"$group": {"_id": None, "total": {"$sum": "$qty_received"}}}
    ]).to_list(length=1)
    return float(agg[0]["total"]) if agg else 0.0


@api_router.get("/store/pending")
async def store_pending(current: dict = Depends(require_store_access)):
    """List item transaksi yang di-flag 'post_to_store' dan belum full-received."""
    txs = await db.transactions.find(
        {"post_to_store": True},
        {"_id": 0},
    ).sort("invoice_date", -1).to_list(length=5000)

    result = []
    for t in txs:
        received = await _sum_received_for_tx(t["id"])
        remaining = float(t.get("qty", 0)) - received
        if remaining > 0:
            result.append({
                "transaction_id": t["id"],
                "invoice_date": t.get("invoice_date"),
                "po_no": t.get("po_no"),
                "invoice_no": t.get("invoice_no"),
                "vendor_name": t.get("vendor_name"),
                "item_name": t.get("item_name"),
                "unit": t.get("unit"),
                "qty_po": float(t.get("qty", 0)),
                "qty_received": received,
                "qty_remaining": remaining,
                "po_date": t.get("po_date"),
            })
    return result


@api_router.get("/store/pending/grouped")
async def store_pending_grouped(current: dict = Depends(require_store_access)):
    """Grouped view: 1 baris per PO (fallback invoice_no) dengan ringkasan qty."""
    items = await store_pending(current)
    groups: dict = {}
    for it in items:
        key = it.get("po_no") or f"INV:{it.get('invoice_no', '')}" or f"TX:{it.get('transaction_id')}"
        g = groups.setdefault(key, {
            "group_key": key,
            "po_no": it.get("po_no"),
            "invoice_no": it.get("invoice_no"),
            "vendor_name": it.get("vendor_name"),
            "invoice_date": it.get("invoice_date"),
            "po_date": it.get("po_date"),
            "items": [],
            "total_qty_po": 0.0,
            "total_qty_received": 0.0,
            "total_qty_remaining": 0.0,
        })
        g["items"].append(it)
        g["total_qty_po"] += it["qty_po"]
        g["total_qty_received"] += it["qty_received"]
        g["total_qty_remaining"] += it["qty_remaining"]
    return sorted(groups.values(), key=lambda x: (x["invoice_date"] or ""), reverse=True)


class BulkReceiveItem(BaseModel):
    transaction_id: str
    qty_received: float
    note: Optional[str] = ""


class BulkReceiveRequest(BaseModel):
    do_number: Optional[str] = ""
    receive_date: str
    items: List[BulkReceiveItem]


@api_router.post("/store/receive/bulk")
async def store_receive_bulk(payload: BulkReceiveRequest, current: dict = Depends(require_store_access)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    received_docs = []
    for item in payload.items:
        if item.qty_received <= 0:
            continue
        tx = await db.transactions.find_one({"id": item.transaction_id})
        if not tx or not tx.get("post_to_store"):
            raise HTTPException(status_code=400, detail=f"Item {item.transaction_id} tidak valid")
        already = await _sum_received_for_tx(item.transaction_id)
        remaining = float(tx.get("qty", 0)) - already
        if item.qty_received > remaining + 1e-9:
            raise HTTPException(
                status_code=400,
                detail=f"{tx.get('item_name')}: qty terima ({item.qty_received}) > sisa ({remaining})"
            )
        doc = {
            "id": str(uuid.uuid4()),
            "transaction_id": item.transaction_id,
            "po_no": tx.get("po_no", ""),
            "invoice_no": tx.get("invoice_no", ""),
            "vendor_name": tx.get("vendor_name", ""),
            "item_name": tx.get("item_name", ""),
            "unit": tx.get("unit", "Ea"),
            "unit_price": float(tx.get("unit_price", 0)),
            "do_number": payload.do_number or "",
            "qty_received": float(item.qty_received),
            "qty_remaining": float(item.qty_received),
            "receive_date": payload.receive_date,
            "note": item.note or "",
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        received_docs.append(doc)
    if not received_docs:
        raise HTTPException(status_code=400, detail="Semua qty kosong / 0")
    await db.store_receipts.insert_many([d.copy() for d in received_docs])
    await _log_action(current, "store_receive", "store_receipt", "-", {
        "count": len(received_docs), "po_no": received_docs[0].get("po_no"),
        "do_number": payload.do_number, "vendor": received_docs[0].get("vendor_name"),
    })
    return {"received": len(received_docs)}


class BulkIssueItem(BaseModel):
    item_name: str
    qty: float
    so_number: Optional[str] = ""
    taker_name: str
    issue_date: str
    note: Optional[str] = ""


class BulkIssueRequest(BaseModel):
    items: List[BulkIssueItem]


@api_router.post("/store/issue/bulk")
async def store_issue_bulk(payload: BulkIssueRequest, current: dict = Depends(require_store_access)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    created = []
    for it in payload.items:
        if it.qty <= 0 or not it.item_name or not it.taker_name.strip():
            continue
        # Compute FIFO on the fly per item
        batches = await db.store_receipts.find(
            {"item_name": it.item_name, "qty_remaining": {"$gt": 0}}
        ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)
        avail = sum(b.get("qty_remaining", 0) for b in batches)
        if it.qty > avail + 1e-9:
            raise HTTPException(status_code=400, detail=f"{it.item_name}: stok tidak cukup (tersedia {avail}, diminta {it.qty})")
        remain = float(it.qty)
        allocations = []
        for b in batches:
            if remain <= 1e-9:
                break
            take = min(float(b["qty_remaining"]), remain)
            allocations.append({
                "receipt_id": b["id"],
                "qty": take,
                "unit_price": float(b.get("unit_price", 0)),
                "vendor_name": b.get("vendor_name", ""),
                "receive_date": b.get("receive_date"),
            })
            await db.store_receipts.update_one({"id": b["id"]}, {"$inc": {"qty_remaining": -take}})
            remain -= take
        total_cost = sum(a["qty"] * a["unit_price"] for a in allocations)
        doc = {
            "id": str(uuid.uuid4()),
            "item_name": it.item_name,
            "unit": batches[0].get("unit", "Ea") if batches else "Ea",
            "qty": float(it.qty),
            "issue_date": it.issue_date,
            "taker_name": it.taker_name.strip(),
            "so_number": it.so_number or "",
            "note": it.note or "",
            "allocations": allocations,
            "total_cost": total_cost,
            "avg_unit_price": (total_cost / it.qty) if it.qty else 0,
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        created.append(doc)
    if not created:
        raise HTTPException(status_code=400, detail="Tidak ada item valid")
    await db.store_issuances.insert_many([d.copy() for d in created])
    await _log_action(current, "store_issue", "store_issuance", "-", {
        "count": len(created), "first_item": created[0].get("item_name"), "so_number": created[0].get("so_number"),
    })
    return {"issued": len(created)}



@api_router.post("/store/receive")
async def store_receive(payload: StoreReceiveRequest, current: dict = Depends(require_store_access)):
    tx = await db.transactions.find_one({"id": payload.transaction_id})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaksi asal tidak ditemukan")
    if not tx.get("post_to_store"):
        raise HTTPException(status_code=400, detail="Item ini tidak di-flag ke Store")
    already = await _sum_received_for_tx(payload.transaction_id)
    remaining = float(tx.get("qty", 0)) - already
    if payload.qty_received <= 0:
        raise HTTPException(status_code=400, detail="Qty terima harus > 0")
    if payload.qty_received > remaining + 1e-9:
        raise HTTPException(
            status_code=400,
            detail=f"Qty terima ({payload.qty_received}) melebihi sisa PO ({remaining}). Over-receive tidak diizinkan."
        )

    doc = {
        "id": str(uuid.uuid4()),
        "transaction_id": payload.transaction_id,
        "po_no": tx.get("po_no", ""),
        "invoice_no": tx.get("invoice_no", ""),
        "vendor_name": tx.get("vendor_name", ""),
        "item_name": tx.get("item_name", ""),
        "unit": tx.get("unit", "Ea"),
        "unit_price": float(tx.get("unit_price", 0)),  # cost basis FIFO
        "do_number": payload.do_number or "",
        "qty_received": float(payload.qty_received),
        "qty_remaining": float(payload.qty_received),  # for FIFO tracking
        "receive_date": payload.receive_date,
        "note": payload.note or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_receipts.insert_one(doc.copy())
    await _log_action(current, "store_receive", "store_receipt", doc["id"], {
        "item": doc["item_name"], "qty": doc["qty_received"], "po_no": doc["po_no"],
        "do_number": doc["do_number"],
    })
    doc.pop("_id", None)
    if not _can_see_prices(current):
        doc.pop("unit_price", None)
    return doc


@api_router.get("/store/receipts")
async def store_receipts(current: dict = Depends(require_store_access), item_name: Optional[str] = None, transaction_id: Optional[str] = None):
    filt: dict = {}
    if item_name:
        filt["item_name"] = item_name
    if transaction_id:
        filt["transaction_id"] = transaction_id
    docs = await db.store_receipts.find(filt, {"_id": 0}).sort("receive_date", -1).to_list(length=1000)
    if not _can_see_prices(current):
        for d in docs:
            d.pop("unit_price", None)
    return docs


@api_router.get("/store/stock")
async def store_stock(current: dict = Depends(require_store_access), customer_only: bool = False, exclude_customer: bool = False):
    """Aggregated stok by item_name from receipts.qty_remaining. Bisa filter customer/non-customer."""
    match: dict = {"qty_remaining": {"$gt": 0}}
    if customer_only:
        match["is_customer_material"] = True
    elif exclude_customer:
        match["$or"] = [{"is_customer_material": False}, {"is_customer_material": {"$exists": False}}]
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {"item": "$item_name", "customer": {"$ifNull": ["$is_customer_material", False]}},
            "qty": {"$sum": "$qty_remaining"},
            "unit": {"$first": "$unit"},
            "last_receive_date": {"$max": "$receive_date"},
            "vendors": {"$addToSet": "$vendor_name"},
            "batches": {"$sum": 1},
        }},
        {"$sort": {"_id.item": 1}},
    ]
    docs = await db.store_receipts.aggregate(pipeline).to_list(length=5000)
    return [{
        "item_name": d["_id"]["item"], "qty": d["qty"], "unit": d["unit"],
        "last_receive_date": d.get("last_receive_date"),
        "vendors": d.get("vendors", []), "batches": d.get("batches", 0),
        "is_customer_material": bool(d["_id"].get("customer")),
    } for d in docs if d["_id"].get("item")]


@api_router.post("/store/issue")
async def store_issue(payload: StoreIssueRequest, current: dict = Depends(require_store_access)):
    if payload.qty <= 0:
        raise HTTPException(status_code=400, detail="Qty keluar harus > 0")
    if not payload.taker_name.strip():
        raise HTTPException(status_code=400, detail="Nama pengambil wajib diisi")

    # FIFO: fetch batches with qty_remaining > 0 for this item, ordered by receive_date asc, then created_at
    batches = await db.store_receipts.find(
        {"item_name": payload.item_name, "qty_remaining": {"$gt": 0}}
    ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)

    total_available = sum(b.get("qty_remaining", 0) for b in batches)
    if payload.qty > total_available + 1e-9:
        raise HTTPException(
            status_code=400,
            detail=f"Stok tidak cukup. Tersedia {total_available}, diminta {payload.qty}."
        )

    remaining_to_take = float(payload.qty)
    allocations = []
    for b in batches:
        if remaining_to_take <= 1e-9:
            break
        take = min(float(b["qty_remaining"]), remaining_to_take)
        allocations.append({
            "receipt_id": b["id"],
            "qty": take,
            "unit_price": float(b.get("unit_price", 0)),
            "vendor_name": b.get("vendor_name", ""),
            "receive_date": b.get("receive_date"),
        })
        # decrement batch
        await db.store_receipts.update_one(
            {"id": b["id"]},
            {"$inc": {"qty_remaining": -take}}
        )
        remaining_to_take -= take

    total_cost = sum(a["qty"] * a["unit_price"] for a in allocations)
    doc = {
        "id": str(uuid.uuid4()),
        "item_name": payload.item_name,
        "unit": batches[0].get("unit", "Ea") if batches else "Ea",
        "qty": float(payload.qty),
        "issue_date": payload.issue_date,
        "taker_name": payload.taker_name.strip(),
        "so_number": payload.so_number or "",
        "note": payload.note or "",
        "allocations": allocations,
        "total_cost": total_cost,
        "avg_unit_price": (total_cost / payload.qty) if payload.qty else 0,
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_issuances.insert_one(doc.copy())
    await _log_action(current, "store_issue", "store_issuance", doc["id"], {
        "item": doc["item_name"], "qty": doc["qty"], "so_number": doc["so_number"], "taker": doc["taker_name"],
    })
    doc.pop("_id", None)
    if not _can_see_prices(current):
        doc.pop("total_cost", None)
        doc.pop("avg_unit_price", None)
        for a in doc.get("allocations", []):
            a.pop("unit_price", None)
    return doc


@api_router.get("/store/issuances")
async def list_issuances(
    current: dict = Depends(require_store_access),
    q: Optional[str] = None,
    so_number: Optional[str] = None,
    taker: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"taker_name": {"$regex": q, "$options": "i"}},
            {"so_number": {"$regex": q, "$options": "i"}},
        ]
    if so_number:
        filt["so_number"] = {"$regex": so_number, "$options": "i"}
    if taker:
        filt["taker_name"] = {"$regex": taker, "$options": "i"}
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        filt["issue_date"] = d

    total = await db.store_issuances.count_documents(filt)
    cursor = db.store_issuances.find(filt, {"_id": 0}).sort("issue_date", -1).skip((page - 1) * page_size).limit(page_size)
    items = await cursor.to_list(length=page_size)
    hide_price = not _can_see_prices(current)
    if hide_price:
        for d in items:
            d.pop("total_cost", None)
            d.pop("avg_unit_price", None)
            for a in d.get("allocations", []):
                a.pop("unit_price", None)
    return {"total": total, "page": page, "page_size": page_size, "items": items, "prices_visible": not hide_price}


@api_router.get("/store/report/xlsx")
async def store_report_xlsx(
    current: dict = Depends(require_store_access),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    if not _can_see_prices(current):
        raise HTTPException(status_code=403, detail="Tidak berwenang melihat harga di laporan Store")
    filt: dict = {}
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        filt["issue_date"] = d
    issuances = await db.store_issuances.find(filt, {"_id": 0}).sort("issue_date", 1).to_list(length=100000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pengeluaran Stok"
    headers = ["Nomor SO", "Tgl Keluar", "Nama Barang", "Qty", "Unit", "Unit Price (FIFO)", "Total Price", "Pengambil", "Vendor Asal", "Note"]
    ws.append(headers)
    for iss in issuances:
        # Flatten per allocation (each batch = 1 row) so FIFO price is transparent
        for a in iss.get("allocations", []):
            ws.append([
                iss.get("so_number", ""), iss.get("issue_date", ""), iss.get("item_name", ""),
                float(a.get("qty", 0)), iss.get("unit", ""), float(a.get("unit_price", 0)),
                float(a.get("qty", 0)) * float(a.get("unit_price", 0)),
                iss.get("taker_name", ""), a.get("vendor_name", ""), iss.get("note", ""),
            ])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"laporan_pengeluaran_stok_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




# ---------------- Store Edit/Delete Request (Approval) ----------------
class StoreRequestCreate(BaseModel):
    target_type: str  # 'receipt' | 'issuance'
    target_id: str
    action_type: str  # 'edit' | 'delete'
    reason: str
    proposed_changes: Optional[dict] = None  # free-text via 'description' key


class StoreRequestReview(BaseModel):
    approve: bool
    review_note: Optional[str] = ""


@api_router.post("/store/requests")
async def create_store_request(payload: StoreRequestCreate, current: dict = Depends(require_store_access)):
    if payload.target_type not in ("receipt", "issuance"):
        raise HTTPException(status_code=400, detail="target_type harus 'receipt' atau 'issuance'")
    if payload.action_type not in ("edit", "delete"):
        raise HTTPException(status_code=400, detail="action_type harus 'edit' atau 'delete'")
    if not payload.reason.strip():
        raise HTTPException(status_code=400, detail="Alasan wajib diisi")

    coll = db.store_receipts if payload.target_type == "receipt" else db.store_issuances
    target = await coll.find_one({"id": payload.target_id})
    if not target:
        raise HTTPException(status_code=404, detail="Data yang diajukan tidak ditemukan")

    # Only allow requester to submit against their own entries (or admin can request anything, but admin doesn't need this)
    if current.get("role") == "store" and target.get("created_by") != current["id"]:
        raise HTTPException(status_code=403, detail="Hanya bisa mengajukan koreksi untuk data milik sendiri")

    summary = {
        "item_name": target.get("item_name"),
        "qty": target.get("qty") or target.get("qty_received"),
        "issue_date": target.get("issue_date") or target.get("receive_date"),
        "po_no": target.get("po_no"),
        "so_number": target.get("so_number"),
        "do_number": target.get("do_number"),
        "taker_name": target.get("taker_name"),
    }
    doc = {
        "id": str(uuid.uuid4()),
        "target_type": payload.target_type,
        "target_id": payload.target_id,
        "target_summary": summary,
        "action_type": payload.action_type,
        "reason": payload.reason.strip(),
        "proposed_changes": payload.proposed_changes or {},
        "status": "pending",
        "requested_by": current["id"],
        "requested_by_username": current.get("username", ""),
        "requested_at": _now_iso(),
        "reviewed_by": None,
        "reviewed_by_username": None,
        "reviewed_at": None,
        "review_note": "",
    }
    await db.store_requests.insert_one(doc.copy())
    await _log_action(current, "store_request", "store_request", doc["id"], {
        "target_type": payload.target_type, "action": payload.action_type,
        "item": summary.get("item_name"), "reason_preview": payload.reason[:80],
    })
    doc.pop("_id", None)
    return doc


@api_router.get("/store/requests")
async def list_store_requests(
    current: dict = Depends(get_current_user),
    status: Optional[str] = None,
    mine: bool = False,
):
    """Admin sees all. Store/staff sees own by default."""
    filt: dict = {}
    if status:
        filt["status"] = status
    if current.get("role") != "admin" or mine:
        filt["requested_by"] = current["id"]
    docs = await db.store_requests.find(filt, {"_id": 0}).sort("requested_at", -1).to_list(length=500)
    return docs


@api_router.get("/store/requests/pending-count")
async def pending_count(current: dict = Depends(require_approve_perm)):
    n = await db.store_requests.count_documents({"status": "pending"})
    return {"count": n}


@api_router.post("/store/requests/{req_id}/review")
async def review_store_request(req_id: str, payload: StoreRequestReview, current: dict = Depends(require_approve_perm)):
    req = await db.store_requests.find_one({"id": req_id})
    if not req:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail=f"Request sudah di-{req['status']}")

    new_status = "approved" if payload.approve else "rejected"

    # If approved & action is delete, execute delete with rollback for FIFO consistency
    if payload.approve and req["action_type"] == "delete":
        if req["target_type"] == "issuance":
            iss = await db.store_issuances.find_one({"id": req["target_id"]})
            if iss:
                # Rollback: restore each allocation back to its receipt.qty_remaining
                for a in iss.get("allocations", []):
                    await db.store_receipts.update_one(
                        {"id": a.get("receipt_id")},
                        {"$inc": {"qty_remaining": float(a.get("qty", 0))}}
                    )
                await db.store_issuances.delete_one({"id": req["target_id"]})
        elif req["target_type"] == "receipt":
            rec = await db.store_receipts.find_one({"id": req["target_id"]})
            if rec:
                consumed = float(rec.get("qty_received", 0)) - float(rec.get("qty_remaining", 0))
                if consumed > 1e-9:
                    raise HTTPException(status_code=400, detail=f"Tidak bisa hapus receipt: {consumed} unit sudah dipakai (issuance). Batalkan issuance dulu.")
                await db.store_receipts.delete_one({"id": req["target_id"]})

    await db.store_requests.update_one(
        {"id": req_id},
        {"$set": {
            "status": new_status,
            "reviewed_by": current["id"],
            "reviewed_by_username": current.get("username", ""),
            "reviewed_at": _now_iso(),
            "review_note": payload.review_note or "",
        }}
    )
    await _log_action(current, "review_store_request", "store_request", req_id, {
        "decision": new_status, "target_type": req["target_type"], "action": req["action_type"],
        "item": req.get("target_summary", {}).get("item_name"),
    })
    return {"status": new_status}




# ---------------- Manual Store Receipt (from Customer/Supplier w/o PO) ----------------
class ManualReceiveRequest(BaseModel):
    receive_date: str
    source_type: str  # 'customer' | 'supplier'
    source_name: str
    so_no: Optional[str] = ""
    do_no: Optional[str] = ""
    po_no: Optional[str] = ""
    item_name: str
    qty: float
    unit: Optional[str] = "Ea"
    mcl_done: Optional[bool] = False
    mif_done: Optional[bool] = False
    remark: Optional[str] = ""
    unit_price: Optional[float] = 0.0


@api_router.post("/store/receive/manual")
async def store_receive_manual(payload: ManualReceiveRequest, current: dict = Depends(require_store_write)):
    if payload.qty <= 0:
        raise HTTPException(status_code=400, detail="Qty harus > 0")
    if payload.source_type not in ("customer", "supplier"):
        raise HTTPException(status_code=400, detail="source_type harus 'customer' atau 'supplier'")
    if not payload.source_name.strip():
        raise HTTPException(status_code=400, detail="Nama customer/supplier wajib")
    if not payload.item_name.strip():
        raise HTTPException(status_code=400, detail="Nama barang wajib")
    is_customer = payload.source_type == "customer"
    doc = {
        "id": str(uuid.uuid4()),
        "transaction_id": None,
        "source": "manual",
        "source_type": payload.source_type,
        "is_customer_material": is_customer,
        "po_no": payload.po_no or "",
        "invoice_no": "",
        "vendor_name": payload.source_name.strip(),
        "customer_name": payload.source_name.strip() if is_customer else "",
        "item_name": payload.item_name.strip(),
        "unit": payload.unit or "Ea",
        "unit_price": float(payload.unit_price or 0),
        "do_number": payload.do_no or "",
        "so_no": payload.so_no or "",
        "qty_received": float(payload.qty),
        "qty_remaining": float(payload.qty),
        "receive_date": payload.receive_date,
        "mcl_done": bool(payload.mcl_done),
        "mif_done": bool(payload.mif_done),
        "note": payload.remark or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_receipts.insert_one(doc.copy())
    await _log_action(current, "store_receive_manual", "store_receipt", doc["id"], {
        "source": payload.source_type, "source_name": doc["vendor_name"],
        "item": doc["item_name"], "qty": doc["qty_received"],
    })
    doc.pop("_id", None)
    if not _can_see_prices(current):
        doc.pop("unit_price", None)
    return doc


@api_router.patch("/store/receipts/{rid}/flags")
async def update_receipt_flags(rid: str, payload: dict, current: dict = Depends(require_store_write)):
    rec = await db.store_receipts.find_one({"id": rid})
    if not rec:
        raise HTTPException(status_code=404, detail="Receipt tidak ditemukan")
    upd: dict = {}
    if "mcl_done" in payload:
        upd["mcl_done"] = bool(payload["mcl_done"])
    if "mif_done" in payload:
        upd["mif_done"] = bool(payload["mif_done"])
    if upd:
        await db.store_receipts.update_one({"id": rid}, {"$set": upd})
        await _log_action(current, "update_receipt_flags", "store_receipt", rid, upd)
    return {"ok": True, "flags": upd}


# ---------------- Production Issue (Customer material → Produksi) ----------------
class ProductionIssueItem(BaseModel):
    item_name: str
    qty: float
    so_number: Optional[str] = ""
    note: Optional[str] = ""


class ProductionIssueRequest(BaseModel):
    issue_date: str
    taker_name: str
    items: List[ProductionIssueItem]


@api_router.post("/store/issue/production")
async def store_issue_production(payload: ProductionIssueRequest, current: dict = Depends(require_store_write)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    if not payload.taker_name.strip():
        raise HTTPException(status_code=400, detail="Nama penerima produksi wajib")
    created = []
    for it in payload.items:
        if it.qty <= 0 or not it.item_name:
            continue
        batches = await db.store_receipts.find(
            {"item_name": it.item_name, "is_customer_material": True, "qty_remaining": {"$gt": 0}}
        ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)
        avail = sum(b.get("qty_remaining", 0) for b in batches)
        if it.qty > avail + 1e-9:
            raise HTTPException(status_code=400, detail=f"{it.item_name}: stok Customer tidak cukup (tersedia {avail})")
        remain = float(it.qty)
        allocations = []
        for b in batches:
            if remain <= 1e-9:
                break
            take = min(float(b["qty_remaining"]), remain)
            allocations.append({
                "receipt_id": b["id"], "qty": take,
                "unit_price": float(b.get("unit_price", 0)),
                "vendor_name": b.get("vendor_name", ""),
                "customer_name": b.get("customer_name", ""),
                "receive_date": b.get("receive_date"),
            })
            await db.store_receipts.update_one({"id": b["id"]}, {"$inc": {"qty_remaining": -take}})
            remain -= take
        doc = {
            "id": str(uuid.uuid4()),
            "type": "production",
            "is_customer_material": True,
            "item_name": it.item_name,
            "unit": batches[0].get("unit", "Ea") if batches else "Ea",
            "qty": float(it.qty),
            "issue_date": payload.issue_date,
            "taker_name": payload.taker_name.strip(),
            "so_number": it.so_number or "",
            "note": it.note or "",
            "allocations": allocations,
            "total_cost": 0,
            "avg_unit_price": 0,
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        created.append(doc)
    if not created:
        raise HTTPException(status_code=400, detail="Tidak ada item valid")
    await db.store_issuances.insert_many([d.copy() for d in created])
    await _log_action(current, "store_issue_production", "store_issuance", "-", {
        "count": len(created), "so_number": created[0].get("so_number"),
        "first_item": created[0].get("item_name"), "taker": payload.taker_name,
    })
    return {"issued": len(created)}


# ---------------- Deliveries (Pengiriman Barang - log only) ----------------
class DeliveryItem(BaseModel):
    item_name: str
    qty: float
    unit: Optional[str] = "Ea"


class DeliveryCreate(BaseModel):
    delivery_date: str
    gate_pass_no: Optional[str] = ""
    do_no: Optional[str] = ""
    destination: str
    driver_name: Optional[str] = ""
    items: List[DeliveryItem]
    remark: Optional[str] = ""


@api_router.post("/deliveries")
async def create_delivery(payload: DeliveryCreate, current: dict = Depends(require_store_write)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Minimal 1 item")
    if not payload.destination.strip():
        raise HTTPException(status_code=400, detail="Tujuan wajib diisi")
    doc = {
        "id": str(uuid.uuid4()),
        "delivery_date": payload.delivery_date,
        "gate_pass_no": payload.gate_pass_no or "",
        "do_no": payload.do_no or "",
        "destination": payload.destination.strip(),
        "driver_name": payload.driver_name or "",
        "items": [it.model_dump() for it in payload.items],
        "remark": payload.remark or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.deliveries.insert_one(doc.copy())
    await _log_action(current, "create_delivery", "delivery", doc["id"], {
        "destination": doc["destination"], "gate_pass": doc["gate_pass_no"], "items": len(doc["items"]),
    })
    doc.pop("_id", None)
    return doc


@api_router.get("/deliveries")
async def list_deliveries(
    current: dict = Depends(get_current_user),
    q: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"destination": {"$regex": q, "$options": "i"}},
            {"gate_pass_no": {"$regex": q, "$options": "i"}},
            {"do_no": {"$regex": q, "$options": "i"}},
            {"driver_name": {"$regex": q, "$options": "i"}},
        ]
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        filt["delivery_date"] = d
    total = await db.deliveries.count_documents(filt)
    items = await db.deliveries.find(filt, {"_id": 0}).sort("delivery_date", -1).skip((page - 1) * page_size).limit(page_size).to_list(length=page_size)
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@api_router.delete("/deliveries/{did}")
async def delete_delivery(did: str, current: dict = Depends(require_write)):
    if current.get("role") not in ("admin", "store"):
        raise HTTPException(status_code=403, detail="Tidak berwenang")
    res = await db.deliveries.delete_one({"id": did})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pengiriman tidak ditemukan")
    await _log_action(current, "delete_delivery", "delivery", did, {})
    return {"ok": True}


# ---------------- Master Sales Order ----------------
class SOCreate(BaseModel):
    so_no: str
    so_date: str
    customer: str
    description: Optional[str] = ""


@api_router.get("/sales-orders")
async def list_sales_orders(current: dict = Depends(get_current_user), q: Optional[str] = None):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"so_no": {"$regex": q, "$options": "i"}},
            {"customer": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}},
        ]
    docs = await db.sales_orders.find(filt, {"_id": 0}).sort("so_date", -1).to_list(length=5000)
    return docs


@api_router.post("/sales-orders")
async def create_so(payload: SOCreate, current: dict = Depends(require_write)):
    so_no = payload.so_no.strip()
    if not so_no:
        raise HTTPException(status_code=400, detail="Nomor SO wajib")
    existing = await db.sales_orders.find_one({"so_no": so_no})
    if existing:
        raise HTTPException(status_code=400, detail=f"SO {so_no} sudah ada")
    doc = {
        "id": str(uuid.uuid4()),
        "so_no": so_no,
        "so_date": payload.so_date,
        "customer": payload.customer.strip(),
        "description": payload.description or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.sales_orders.insert_one(doc.copy())
    await _log_action(current, "create_so", "sales_order", doc["id"], {"so_no": so_no, "customer": doc["customer"]})
    doc.pop("_id", None)
    return doc


@api_router.put("/sales-orders/{sid}")
async def update_so(sid: str, payload: SOCreate, current: dict = Depends(require_write)):
    so = await db.sales_orders.find_one({"id": sid})
    if not so:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    upd = {
        "so_no": payload.so_no.strip(),
        "so_date": payload.so_date,
        "customer": payload.customer.strip(),
        "description": payload.description or "",
    }
    await db.sales_orders.update_one({"id": sid}, {"$set": upd})
    await _log_action(current, "update_so", "sales_order", sid, upd)
    updated = await db.sales_orders.find_one({"id": sid}, {"_id": 0})
    return updated


@api_router.delete("/sales-orders/{sid}")
async def delete_so(sid: str, current: dict = Depends(require_write)):
    res = await db.sales_orders.delete_one({"id": sid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    await _log_action(current, "delete_so", "sales_order", sid, {})
    return {"ok": True}


# ---------------- Excel Import/Export ----------------
EXPORT_HEADERS = [
    "Tanggal Invoice", "Nomor Project (SO)", "Nomor PO", "Nama Toko", "Nama Barang",
    "Qty", "Unit", "Unit Price", "Total Price", "Nomor Invoice", "Tanggal PO", "Tanggal Terima", "Catatan"
]

def _to_date_str(val: Any) -> Optional[str]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    # Try parse common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s  # fallback

def _to_float(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(val)
    except Exception:
        try:
            return float(str(val).replace(",", "").replace(".", ""))
        except Exception:
            return 0.0

@api_router.get("/transactions/export/xlsx")
async def export_xlsx(current: dict = Depends(get_current_user)):
    docs = await db.transactions.find({}, {"_id": 0}).sort("invoice_date", 1).to_list(length=100000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pembelian"
    ws.append(EXPORT_HEADERS)
    for d in docs:
        ws.append([
            d.get("invoice_date", ""), d.get("project_no", ""), d.get("po_no", ""),
            d.get("vendor_name", ""), d.get("item_name", ""), d.get("qty", 0),
            d.get("unit", ""), d.get("unit_price", 0), d.get("total_price", 0),
            d.get("invoice_no", ""), d.get("po_date", ""), d.get("receive_date", ""), d.get("notes", ""),
        ])
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        max_len = max([len(str(header))] + [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, min(ws.max_row, 100) + 1)])
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"laporan_pembelian_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@api_router.post("/transactions/import/xlsx")
async def import_xlsx(file: UploadFile = File(...), current: dict = Depends(get_current_user)):
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File Excel tidak valid: {e}")

    inserted = 0
    errors: List[str] = []
    now = _now_iso()

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Find header row - flexible matching for common patterns
        header_row_idx = None
        header_map = {}
        for i, row in enumerate(rows[:5]):
            row_str = [str(c).strip().lower() if c is not None else "" for c in row]
            if any("tanggal" in c or "invoice date" in c for c in row_str) and any("nama" in c or "description" in c or "item" in c for c in row_str):
                header_row_idx = i
                for idx, cell in enumerate(row_str):
                    header_map[idx] = cell
                break
        if header_row_idx is None:
            continue

        def find_col(*keywords) -> Optional[int]:
            for idx, cell in header_map.items():
                for kw in keywords:
                    if kw in cell:
                        return idx
            return None

        col_date = find_col("tanggal invoice", "invoice date", "tanggal")
        col_so = find_col("project", "so")
        col_po = find_col("purchase order", "po no", "po")
        col_vendor = find_col("toko", "vendor", "supplier")
        col_item = find_col("nama barang", "detail description", "description", "item")
        col_qty = find_col("qty", "quantity")
        col_unit = find_col("item unit", "unit", "satuan")
        col_price = find_col("harga", "unit price", "price")
        col_total = find_col("jumlah", "amount", "total")
        col_inv = find_col("nomor invoice", "invoice no")
        col_podate = find_col("po date", "purchase order po date", "tanggal po")
        col_recv = find_col("receive", "terima")

        docs = []
        for row in rows[header_row_idx + 1:]:
            if row is None or all(c is None or c == "" for c in row):
                continue
            try:
                item_name = row[col_item] if col_item is not None and col_item < len(row) else None
                if not item_name:
                    continue
                d = {
                    "id": str(uuid.uuid4()),
                    "invoice_date": _to_date_str(row[col_date]) if col_date is not None and col_date < len(row) else "",
                    "project_no": str(row[col_so]) if col_so is not None and col_so < len(row) and row[col_so] not in (None, "") else "",
                    "po_no": str(row[col_po]) if col_po is not None and col_po < len(row) and row[col_po] not in (None, "") else "",
                    "vendor_name": str(row[col_vendor]).strip() if col_vendor is not None and col_vendor < len(row) and row[col_vendor] else "",
                    "item_name": str(item_name).strip(),
                    "qty": _to_float(row[col_qty]) if col_qty is not None and col_qty < len(row) else 0,
                    "unit": str(row[col_unit]).strip() if col_unit is not None and col_unit < len(row) and row[col_unit] else "Ea",
                    "unit_price": _to_float(row[col_price]) if col_price is not None and col_price < len(row) else 0,
                    "total_price": _to_float(row[col_total]) if col_total is not None and col_total < len(row) else 0,
                    "invoice_no": str(row[col_inv]) if col_inv is not None and col_inv < len(row) and row[col_inv] not in (None, "") else "",
                    "po_date": _to_date_str(row[col_podate]) if col_podate is not None and col_podate < len(row) else None,
                    "receive_date": _to_date_str(row[col_recv]) if col_recv is not None and col_recv < len(row) else None,
                    "notes": "",
                    "created_at": now,
                    "updated_at": now,
                }
                if d["total_price"] == 0 and d["qty"] and d["unit_price"]:
                    d["total_price"] = d["qty"] * d["unit_price"]
                if not d["invoice_date"]:
                    continue
                docs.append(d)
            except Exception as e:
                errors.append(f"Sheet {sn}: {e}")

        if docs:
            await db.transactions.insert_many([d.copy() for d in docs])
            inserted += len(docs)

    return {"inserted": inserted, "errors": errors[:20]}


# ---------------- Root ----------------
@api_router.get("/")
async def root():
    return {"message": "Laporan Pembelian API"}


# ---------------- Startup ----------------
async def seed_admin():
    admin_username = os.environ.get("ADMIN_USERNAME", "admin").lower().strip()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    # Migrate legacy admin: if there's an existing admin without a `username`, upgrade it
    legacy = await db.users.find_one({"role": "admin", "username": {"$exists": False}})
    if legacy:
        await db.users.update_one(
            {"id": legacy["id"]},
            {"$set": {"username": admin_username, "active": True}}
        )
        logger.info(f"Migrated legacy admin to username: {admin_username}")

    existing = await db.users.find_one({"username": admin_username})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "username": admin_username,
            "password_hash": hash_password(admin_password),
            "name": "Admin Utama",
            "role": "admin",
            "active": True,
            "perms": ["approve_store_requests"],
            "created_at": _now_iso(),
        })
        logger.info(f"Seeded admin: {admin_username}")
    else:
        # Ensure admin has approve_store_requests perm & correct password
        updates: dict = {}
        if not verify_password(admin_password, existing["password_hash"]):
            updates["password_hash"] = hash_password(admin_password)
        if "approve_store_requests" not in (existing.get("perms") or []):
            updates["perms"] = list(set((existing.get("perms") or []) + ["approve_store_requests"]))
        updates.setdefault("role", "admin")
        updates.setdefault("active", True)
        if updates:
            await db.users.update_one({"username": admin_username}, {"$set": updates})
            logger.info(f"Updated admin: {admin_username}")

@app.on_event("startup")
async def startup():
    # Drop legacy unique index on email if present, add unique on username
    try:
        info = await db.users.index_information()
        if "email_1" in info:
            await db.users.drop_index("email_1")
    except Exception:
        pass
    await db.users.create_index("username", unique=True, sparse=True)
    await db.transactions.create_index("invoice_date")
    await db.transactions.create_index("vendor_name")
    await db.transactions.create_index("item_name")
    await db.transactions.create_index("invoice_no")
    await db.activity_logs.create_index("timestamp")
    await db.activity_logs.create_index("user_id")
    await db.store_receipts.create_index("item_name")
    await db.store_receipts.create_index("transaction_id")
    await db.store_receipts.create_index("qty_remaining")
    await db.store_issuances.create_index("issue_date")
    await db.store_issuances.create_index("item_name")
    await db.store_requests.create_index("status")
    await db.store_requests.create_index("requested_by")
    await db.deliveries.create_index("delivery_date")
    await db.sales_orders.create_index("so_no", unique=True)
    await db.sales_orders.create_index("so_date")
    await seed_admin()

@app.on_event("shutdown")
async def shutdown():
    client.close()

app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
