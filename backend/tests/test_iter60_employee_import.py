"""
Test iteration 60: Employee Master Excel Import Feature
Tests the bulk upload of employee data via Excel file on Production Attendance page.
"""
import requests
import io
from openpyxl import Workbook, load_workbook

BASE_URL = "https://error-fix-dev.preview.emergentagent.com/api"

class TestEmployeeImport:
    def __init__(self):
        self.session = requests.Session()
        self.token = None
        self.tests_run = 0
        self.tests_passed = 0
        self.created_employee_ids = []

    def log(self, msg, success=None):
        """Log test results"""
        self.tests_run += 1
        if success is True:
            self.tests_passed += 1
            print(f"✅ {msg}")
        elif success is False:
            print(f"❌ {msg}")
        else:
            print(f"ℹ️  {msg}")

    def login(self):
        """Login as admin"""
        print("\n🔐 Logging in as admin...")
        try:
            resp = self.session.post(
                f"{BASE_URL}/auth/login",
                json={"username": "admin", "password": "admin123"}
            )
            if resp.status_code == 200:
                self.log("Login successful", True)
                return True
            else:
                self.log(f"Login failed: {resp.status_code} - {resp.text}", False)
                return False
        except Exception as e:
            self.log(f"Login error: {e}", False)
            return False

    def test_template_download(self):
        """Test GET /production/employees/template.xlsx"""
        print("\n📥 Testing template download...")
        try:
            resp = self.session.get(f"{BASE_URL}/production/employees/template.xlsx")
            
            # Check status code
            if resp.status_code != 200:
                self.log(f"Template download failed: {resp.status_code}", False)
                return False
            
            # Check content type
            content_type = resp.headers.get('Content-Type', '')
            if 'spreadsheet' not in content_type and 'excel' not in content_type:
                self.log(f"Wrong content type: {content_type}", False)
                return False
            
            # Load and validate Excel structure
            wb = load_workbook(io.BytesIO(resp.content))
            
            # Check if 'Karyawan' sheet exists
            if 'Karyawan' not in wb.sheetnames:
                self.log("'Karyawan' sheet not found", False)
                return False
            
            ws = wb['Karyawan']
            
            # Check headers
            headers = [cell.value for cell in ws[1]]
            if headers[0] != 'Nama' or headers[1] != 'Jabatan':
                self.log(f"Wrong headers: {headers}", False)
                return False
            
            # Check example rows
            row2 = [cell.value for cell in ws[2]]
            row3 = [cell.value for cell in ws[3]]
            
            if 'Budi Santoso' not in str(row2[0]) or 'Andi Wijaya' not in str(row3[0]):
                self.log(f"Example rows missing or incorrect", False)
                return False
            
            self.log("Template download successful with correct structure", True)
            return True
            
        except Exception as e:
            self.log(f"Template download error: {e}", False)
            return False

    def create_test_excel(self, employees):
        """Create a test Excel file with employee data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Karyawan"
        
        # Headers
        ws.append(["Nama", "Jabatan"])
        
        # Add employee data
        for emp in employees:
            ws.append([emp['name'], emp.get('designation', '')])
        
        # Save to BytesIO
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_import_valid_excel(self):
        """Test POST /production/employees/import with valid Excel"""
        print("\n📤 Testing valid Excel import...")
        
        # Create test data with unique names
        import time
        timestamp = str(int(time.time()))[-6:]
        test_employees = [
            {'name': f'TESTIMP_One_{timestamp}', 'designation': 'Operator Bubut'},
            {'name': f'TESTIMP_Two_{timestamp}', 'designation': 'Operator Milling'},
            {'name': f'TESTIMP_Three_{timestamp}', 'designation': 'Welder'}
        ]
        
        excel_file = self.create_test_excel(test_employees)
        
        try:
            files = {'file': ('test_employees.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            resp = self.session.post(f"{BASE_URL}/production/employees/import", files=files)
            
            if resp.status_code != 200:
                self.log(f"Import failed: {resp.status_code} - {resp.text}", False)
                return False
            
            data = resp.json()
            
            # Validate response structure
            if 'created' not in data or 'skipped' not in data or 'total_rows' not in data:
                self.log(f"Invalid response structure: {data}", False)
                return False
            
            # Check that all 3 employees were created
            if data['created'] != 3:
                self.log(f"Expected 3 created, got {data['created']}", False)
                return False
            
            if data['skipped'] != 0:
                self.log(f"Expected 0 skipped, got {data['skipped']}", False)
                return False
            
            self.log(f"Import successful: {data['created']} created, {data['skipped']} skipped", True)
            
            # Store test employee names for cleanup
            self.test_employee_names = [emp['name'] for emp in test_employees]
            return True
            
        except Exception as e:
            self.log(f"Import error: {e}", False)
            return False

    def test_import_duplicate(self):
        """Test importing the same file twice - should skip all"""
        print("\n🔄 Testing duplicate import...")
        
        # Use the same names from previous test
        if not hasattr(self, 'test_employee_names'):
            self.log("Skipping duplicate test - no previous import", None)
            return True
        
        test_employees = [{'name': name, 'designation': 'Test'} for name in self.test_employee_names]
        excel_file = self.create_test_excel(test_employees)
        
        try:
            files = {'file': ('test_employees.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            resp = self.session.post(f"{BASE_URL}/production/employees/import", files=files)
            
            if resp.status_code != 200:
                self.log(f"Duplicate import failed: {resp.status_code}", False)
                return False
            
            data = resp.json()
            
            # All should be skipped
            if data['created'] != 0:
                self.log(f"Expected 0 created on duplicate, got {data['created']}", False)
                return False
            
            if data['skipped'] != 3:
                self.log(f"Expected 3 skipped on duplicate, got {data['skipped']}", False)
                return False
            
            self.log(f"Duplicate handling correct: {data['created']} created, {data['skipped']} skipped", True)
            return True
            
        except Exception as e:
            self.log(f"Duplicate import error: {e}", False)
            return False

    def test_import_with_template_examples(self):
        """Test that template example rows (Budi Santoso, Andi Wijaya) are skipped"""
        print("\n🚫 Testing template example row skipping...")
        
        import time
        timestamp = str(int(time.time()))[-6:]
        test_employees = [
            {'name': 'Budi Santoso', 'designation': 'Operator'},  # Should be skipped
            {'name': 'Andi Wijaya', 'designation': 'Operator'},   # Should be skipped
            {'name': f'TESTIMP_Valid_{timestamp}', 'designation': 'Welder'}  # Should be created
        ]
        
        excel_file = self.create_test_excel(test_employees)
        
        try:
            files = {'file': ('test_with_examples.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            resp = self.session.post(f"{BASE_URL}/production/employees/import", files=files)
            
            if resp.status_code != 200:
                self.log(f"Import with examples failed: {resp.status_code}", False)
                return False
            
            data = resp.json()
            
            # Only 1 should be created, 2 skipped (template examples)
            if data['created'] != 1:
                self.log(f"Expected 1 created (template examples skipped), got {data['created']}", False)
                return False
            
            if data['skipped'] != 2:
                self.log(f"Expected 2 skipped (template examples), got {data['skipped']}", False)
                return False
            
            self.log(f"Template example skipping works: {data['created']} created, {data['skipped']} skipped", True)
            return True
            
        except Exception as e:
            self.log(f"Template example test error: {e}", False)
            return False

    def test_import_invalid_file(self):
        """Test POST /production/employees/import with non-xlsx file"""
        print("\n❌ Testing invalid file type...")
        
        try:
            # Create a text file pretending to be xlsx
            fake_file = io.BytesIO(b"This is not an Excel file")
            files = {'file': ('test.txt', fake_file, 'text/plain')}
            
            resp = self.session.post(f"{BASE_URL}/production/employees/import", files=files)
            
            # Should return 400
            if resp.status_code != 400:
                self.log(f"Expected 400 for invalid file, got {resp.status_code}", False)
                return False
            
            self.log("Invalid file type correctly rejected with 400", True)
            return True
            
        except Exception as e:
            self.log(f"Invalid file test error: {e}", False)
            return False

    def test_employees_list(self):
        """Test GET /production/employees to verify imported employees appear"""
        print("\n📋 Testing employee list retrieval...")
        
        if not hasattr(self, 'test_employee_names'):
            self.log("Skipping list test - no employees imported", None)
            return True
        
        try:
            resp = self.session.get(f"{BASE_URL}/production/employees")
            
            if resp.status_code != 200:
                self.log(f"Employee list failed: {resp.status_code}", False)
                return False
            
            data = resp.json()
            
            if 'items' not in data:
                self.log(f"Invalid response structure: {data}", False)
                return False
            
            # Check if our test employees are in the list
            employee_names = [emp['name'] for emp in data['items']]
            
            found_count = 0
            for test_name in self.test_employee_names:
                if test_name in employee_names:
                    found_count += 1
            
            if found_count != len(self.test_employee_names):
                self.log(f"Only {found_count}/{len(self.test_employee_names)} test employees found in list", False)
                return False
            
            self.log(f"All {found_count} imported employees found in list", True)
            
            # Store IDs for cleanup
            for emp in data['items']:
                if emp['name'] in self.test_employee_names:
                    self.created_employee_ids.append(emp['id'])
            
            return True
            
        except Exception as e:
            self.log(f"Employee list error: {e}", False)
            return False

    def cleanup(self):
        """Delete test employees"""
        print("\n🧹 Cleaning up test data...")
        
        if not self.created_employee_ids:
            print("No test employees to clean up")
            return
        
        deleted = 0
        for emp_id in self.created_employee_ids:
            try:
                resp = self.session.delete(f"{BASE_URL}/production/employees/{emp_id}")
                if resp.status_code == 200:
                    deleted += 1
            except Exception as e:
                print(f"Failed to delete employee {emp_id}: {e}")
        
        print(f"Cleaned up {deleted}/{len(self.created_employee_ids)} test employees")

    def run_all_tests(self):
        """Run all tests"""
        print("=" * 60)
        print("🧪 Employee Master Excel Import Feature Tests")
        print("=" * 60)
        
        if not self.login():
            print("\n❌ Login failed - cannot proceed with tests")
            return False
        
        # Run tests in order
        self.test_template_download()
        self.test_import_valid_excel()
        self.test_import_duplicate()
        self.test_import_with_template_examples()
        self.test_import_invalid_file()
        self.test_employees_list()
        
        # Cleanup
        self.cleanup()
        
        # Print summary
        print("\n" + "=" * 60)
        print(f"📊 Test Summary: {self.tests_passed}/{self.tests_run} tests passed")
        print("=" * 60)
        
        return self.tests_passed == self.tests_run


if __name__ == "__main__":
    tester = TestEmployeeImport()
    success = tester.run_all_tests()
    exit(0 if success else 1)
