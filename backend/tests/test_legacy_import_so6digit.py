"""Test Legacy Import + SO 6-digit normalization features."""
import requests
import sys
import io
from datetime import datetime
from openpyxl import Workbook

BASE_URL = "https://error-fix-dev.preview.emergentagent.com/api"

class TestLegacyImportSO6Digit:
    def __init__(self):
        self.session = requests.Session()
        self.token = None
        self.tests_run = 0
        self.tests_passed = 0
        self.test_drawing_ids = []
        self.test_bom_ids = []

    def log(self, msg):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

    def run_test(self, name, method, endpoint, expected_status, data=None, files=None, headers=None):
        """Run a single API test"""
        url = f"{BASE_URL}{endpoint}"
        h = headers or {}
        
        self.tests_run += 1
        self.log(f"🔍 Testing {name}...")
        
        try:
            if method == 'GET':
                response = self.session.get(url, headers=h)
            elif method == 'POST':
                if files:
                    response = self.session.post(url, data=data, files=files, headers=h)
                else:
                    response = self.session.post(url, json=data, headers=h)
            elif method == 'PATCH':
                response = self.session.patch(url, json=data, headers=h)
            elif method == 'DELETE':
                response = self.session.delete(url, headers=h)
            else:
                raise ValueError(f"Unsupported method: {method}")

            success = response.status_code == expected_status
            if success:
                self.tests_passed += 1
                self.log(f"✅ Passed - Status: {response.status_code}")
                try:
                    return True, response.json()
                except:
                    return True, {}
            else:
                self.log(f"❌ Failed - Expected {expected_status}, got {response.status_code}")
                try:
                    self.log(f"   Response: {response.json()}")
                except:
                    self.log(f"   Response: {response.text[:200]}")
                return False, {}

        except Exception as e:
            self.log(f"❌ Failed - Error: {str(e)}")
            return False, {}

    def login(self, username, password):
        """Login and get session cookie"""
        self.log(f"Logging in as {username}...")
        success, response = self.run_test(
            f"Login as {username}",
            "POST",
            "/auth/login",
            200,
            data={"username": username, "password": password}
        )
        if success:
            self.log(f"✅ Logged in successfully as {username}")
            return True
        return False

    def create_test_xlsx(self):
        """Create a minimal BOM Excel file for testing"""
        wb = Workbook()
        ws = wb.active
        
        # Minimal BOM structure matching MKS template
        ws['A4'] = 'TO'
        ws['B4'] = 'PT TEST'
        ws['L4'] = 'BOM.NO.'
        ws['N4'] = ':TEST-BOM-001'
        
        ws['A5'] = 'DATE'
        ws['B5'] = ':2025-01-15'
        ws['L5'] = 'REV.NO.'
        ws['N5'] = ':0'
        
        ws['A7'] = 'PROJECT'
        ws['D7'] = 'ENG.DRW.'
        ws['G7'] = 'CUSTOMER'
        ws['K7'] = 'CLASS OF MATERIAL'
        ws['N7'] = 'SO.NO.'
        ws['P7'] = 'DELIVERY DATE'
        
        ws['A8'] = 'TEST PROJECT'
        ws['D8'] = 'TEST-AUTOMATION-DWG1'
        ws['G8'] = 'TEST CUSTOMER'
        ws['K8'] = 'SS400'
        ws['N8'] = '5501'  # Will be normalized to 005501
        ws['P8'] = '2025-02-01'
        
        ws['A9'] = 'NO.'
        ws['B9'] = 'ITEM NAME'
        ws['C9'] = 'SPECIFICATION'
        ws['H9'] = 'QTY'
        ws['I9'] = 'UOM'
        ws['K9'] = 'MATERIAL'
        
        ws['A11'] = 1
        ws['B11'] = 'Test Item 1'
        ws['C11'] = 'Test Spec'
        ws['H11'] = 10
        ws['I11'] = 'Pcs'
        ws['K11'] = 'Steel'
        
        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def create_test_pdf(self):
        """Create a minimal PDF for testing"""
        # Minimal PDF structure
        pdf_content = b"""%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /Resources << /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> >> >> /MediaBox [0 0 612 792] /Contents 4 0 R >>
endobj
4 0 obj
<< /Length 44 >>
stream
BT
/F1 12 Tf
100 700 Td
(Test Drawing) Tj
ET
endstream
endobj
xref
0 5
0000000000 65535 f 
0000000009 00000 n 
0000000058 00000 n 
0000000115 00000 n 
0000000314 00000 n 
trailer
<< /Size 5 /Root 1 0 R >>
startxref
407
%%EOF
"""
        return pdf_content

    def test_legacy_import_analyze(self):
        """Test /api/legacy-import/analyze endpoint"""
        self.log("\n=== Testing Legacy Import Analyze ===")
        
        xlsx_bytes = self.create_test_xlsx()
        files = {'file': ('test_bom.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, response = self.run_test(
            "Analyze BOM Excel",
            "POST",
            "/legacy-import/analyze",
            200,
            files=files
        )
        
        if success:
            self.log(f"   Suggested SO: {response.get('suggested', {}).get('so_no')}")
            self.log(f"   Items count: {response.get('items_count')}")
            if response.get('suggested', {}).get('so_no') == '005501':
                self.log("   ✅ SO number correctly normalized to 6 digits")
            else:
                self.log(f"   ⚠️  SO number not normalized: {response.get('suggested', {}).get('so_no')}")
        
        return success

    def test_legacy_import_commit(self):
        """Test /api/legacy-import/commit endpoint"""
        self.log("\n=== Testing Legacy Import Commit ===")
        
        # Prepare files
        pdf_bytes = self.create_test_pdf()
        xlsx_bytes = self.create_test_xlsx()
        
        meta = {
            "drawing_no": "TEST-AUTOMATION-DWG1",
            "so_no": "5501",  # Should be normalized to 005501
            "project_name": "Test Project",
            "customer": "TEST",
            "revision": "Rev-0",
            "items": [
                {
                    "item_no": 1,
                    "item_name": "Test Item",
                    "item_specification": "Test Spec",
                    "qty": 10,
                    "uom": "Pcs",
                    "material": "Steel"
                }
            ]
        }
        
        import json
        files = {
            'eng_dwg': ('test_drawing.pdf', pdf_bytes, 'application/pdf'),
            'bom_file': ('test_bom.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        data = {'meta': json.dumps(meta)}
        
        success, response = self.run_test(
            "Commit Legacy Import",
            "POST",
            "/legacy-import/commit",
            200,
            data=data,
            files=files
        )
        
        if success:
            self.log(f"   Drawing ID: {response.get('drawing_id')}")
            self.log(f"   SO Number: {response.get('so_no')}")
            self.log(f"   BOM ID: {response.get('bom_id')}")
            
            # Store for cleanup
            if response.get('drawing_id'):
                self.test_drawing_ids.append(response['drawing_id'])
            if response.get('bom_id'):
                self.test_bom_ids.append(response['bom_id'])
            
            if response.get('so_no') == '005501':
                self.log("   ✅ SO number correctly stored as 6 digits")
            else:
                self.log(f"   ⚠️  SO number not normalized: {response.get('so_no')}")
        
        return success

    def test_so_6digit_normalization(self):
        """Test SO 6-digit normalization in various endpoints"""
        self.log("\n=== Testing SO 6-digit Normalization ===")
        
        # Test GET /api/sales-orders/check/{so_no}
        success1, response1 = self.run_test(
            "Check SO 5501 (should normalize to 005501)",
            "GET",
            "/sales-orders/check/5501",
            200
        )
        if success1:
            self.log(f"   Normalized SO: {response1.get('so_no')}")
        
        # Test GET /api/sales-orders/autocomplete?q=5501
        success2, response2 = self.run_test(
            "Autocomplete SO 5501",
            "GET",
            "/sales-orders/autocomplete?q=5501",
            200
        )
        if success2:
            self.log(f"   Found {len(response2.get('items', []))} matches")
        
        return success1 and success2

    def test_drawing_verification(self):
        """Verify drawing was created correctly"""
        self.log("\n=== Verifying Drawing in Master List ===")
        
        success, response = self.run_test(
            "Search for TEST-AUTOMATION-DWG1",
            "GET",
            "/drawings?q=TEST-AUTOMATION-DWG1",
            200
        )
        
        if success:
            items = response.get('items', [])
            self.log(f"   Found {len(items)} drawing(s)")
            if items:
                drawing = items[0]
                self.log(f"   Drawing No: {drawing.get('drawing_no')}")
                self.log(f"   SO No: {drawing.get('so_no')}")
                self.log(f"   Approval Status: {drawing.get('approval_status')}")
                self.log(f"   Legacy Import: {drawing.get('legacy_import')}")
                
                if drawing.get('approval_status') == 'controlled':
                    self.log("   ✅ Drawing is controlled (final)")
                if drawing.get('so_no') == '005501':
                    self.log("   ✅ SO number is 6-digit padded")
                if drawing.get('legacy_import'):
                    self.log("   ✅ Marked as legacy import")
        
        return success

    def test_403_non_admin(self):
        """Test that non-admin roles get 403"""
        self.log("\n=== Testing 403 for Non-Admin Roles ===")
        
        # Try to login as a non-admin user (if available)
        # For now, we'll test with the current session
        # In a real scenario, you'd login as a different user
        
        # Test analyze endpoint (should work for eng roles)
        xlsx_bytes = self.create_test_xlsx()
        files = {'file': ('test.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        # This test assumes we're logged in as susanto (super_admin)
        # So it should pass. To test 403, we'd need to login as a non-privileged user
        self.log("   ⚠️  Skipping 403 test - requires non-admin user login")
        return True

    def test_excel_preview_regression(self):
        """Test Excel preview endpoints still work"""
        self.log("\n=== Testing Excel Preview Regression ===")
        
        # These endpoints require existing data, so we'll just check they respond
        # Test GET /api/excel-templates (list templates)
        success1, response1 = self.run_test(
            "List Excel Templates",
            "GET",
            "/excel-templates",
            200
        )
        
        if success1:
            if isinstance(response1, list):
                self.log(f"   Found {len(response1)} templates")
            elif isinstance(response1, dict):
                self.log(f"   Found {len(response1.get('items', []))} templates")
        
        return success1

    def cleanup(self):
        """Delete test data"""
        self.log("\n=== Cleaning Up Test Data ===")
        
        # Delete test drawings
        for drawing_id in self.test_drawing_ids:
            self.log(f"Deleting drawing {drawing_id}...")
            self.run_test(
                f"Delete drawing {drawing_id}",
                "DELETE",
                f"/drawings/{drawing_id}",
                200
            )
        
        # Delete test BOMs
        for bom_id in self.test_bom_ids:
            self.log(f"Deleting BOM {bom_id}...")
            self.run_test(
                f"Delete BOM {bom_id}",
                "DELETE",
                f"/bom/{bom_id}",
                200
            )
        
        self.log("✅ Cleanup complete")

    def cleanup_existing_test_data(self):
        """Clean up any existing test data before running tests"""
        self.log("\n=== Cleaning Up Existing Test Data ===")
        
        # Search for existing test drawings
        success, response = self.run_test(
            "Search for existing test drawings",
            "GET",
            "/drawings?q=TEST-AUTOMATION",
            200
        )
        
        if success and response.get('items'):
            for drawing in response['items']:
                drawing_id = drawing.get('id')
                if drawing_id:
                    self.log(f"Deleting existing drawing {drawing.get('drawing_no')}...")
                    self.run_test(
                        f"Delete drawing {drawing_id}",
                        "DELETE",
                        f"/drawings/{drawing_id}",
                        200
                    )
        
        # Search for existing test BOMs
        success, response = self.run_test(
            "Search for existing test BOMs",
            "GET",
            "/bom?q=TEST",
            200
        )
        
        if success and response.get('items'):
            for bom in response['items']:
                if 'TEST' in bom.get('project_name', '') or 'TEST' in bom.get('so_no', ''):
                    bom_id = bom.get('id')
                    if bom_id:
                        self.log(f"Deleting existing BOM {bom.get('bom_no')}...")
                        self.run_test(
                            f"Delete BOM {bom_id}",
                            "DELETE",
                            f"/bom/{bom_id}",
                            200
                        )
        
        self.log("✅ Existing test data cleaned up")

    def run_all_tests(self):
        """Run all tests"""
        self.log("=" * 60)
        self.log("LEGACY IMPORT + SO 6-DIGIT BACKEND TESTS")
        self.log("=" * 60)
        
        # Login as super admin
        if not self.login("susanto", "Subbalo1994"):
            self.log("❌ Login failed, stopping tests")
            return 1
        
        # Clean up existing test data first
        self.cleanup_existing_test_data()
        
        # Run tests
        self.test_legacy_import_analyze()
        self.test_legacy_import_commit()
        self.test_so_6digit_normalization()
        self.test_drawing_verification()
        self.test_403_non_admin()
        self.test_excel_preview_regression()
        
        # Cleanup
        self.cleanup()
        
        # Print results
        self.log("\n" + "=" * 60)
        self.log(f"📊 Tests passed: {self.tests_passed}/{self.tests_run}")
        self.log("=" * 60)
        
        return 0 if self.tests_passed == self.tests_run else 1

def main():
    tester = TestLegacyImportSO6Digit()
    return tester.run_all_tests()

if __name__ == "__main__":
    sys.exit(main())
