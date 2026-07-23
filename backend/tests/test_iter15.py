"""Iteration 15 backend tests: Feature D (Print MCL), E (BOM), Bonus (plan_delivery_date)."""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
API = f"{BASE_URL}/api"


def _session(username, password):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, f"login failed for {username}: {r.status_code} {r.text}"
    # auth via cookie set by response
    return s


@pytest.fixture(scope="module")
def admin_token():
    return _session("susanto", "admin123")


@pytest.fixture(scope="module")
def engineer_token():
    return _session("engineer01", "eng123")


@pytest.fixture(scope="module")
def store_token():
    return _session("store01", "store123")


def H(sess):
    # Backward-compat: returns a helper dict, but we use session directly
    return {}


# ------------------------ Feature D — Print MCL ------------------------
class TestMCL:
    def test_mcl_template_exists_and_endpoint_returns_xlsx(self, admin_token):
        # Find any receipt
        r = admin_token.get(f"{API}/store/incoming-report?page=1&page_size=5")
        assert r.status_code == 200
        items = r.json().get("items", [])
        assert len(items) > 0, "No store_receipts available to test MCL"
        rid = items[0]["id"]
        r2 = admin_token.get(f"{API}/store/incoming/mcl/{rid}")
        assert r2.status_code == 200, f"MCL endpoint failed: {r2.status_code} {r2.text[:200]}"
        assert "spreadsheetml" in r2.headers.get("content-type", "")
        cd = r2.headers.get("content-disposition", "")
        assert "MCL_" in cd and ".xlsx" in cd, f"Bad Content-Disposition: {cd}"

        wb = load_workbook(io.BytesIO(r2.content))
        assert "RECEIVED MATERIAL" in wb.sheetnames, f"sheet not found. sheets={wb.sheetnames}"
        ws = wb["RECEIVED MATERIAL"]
        # A55 must be preserved
        a55 = ws["A55"].value
        assert a55 and "MKS-F-STR-004" in str(a55), f"A55 doc register lost: {a55!r}"
        # A4 title
        assert "MATERIAL CONTROL LABEL" in str(ws["A4"].value or "").upper()
        # E6/E7/E8 headers filled (or '-')
        assert ws["E6"].value is not None
        assert ws["E7"].value is not None
        assert ws["E8"].value is not None
        # Row 12 has item data
        assert ws.cell(row=12, column=6).value is not None  # qty (F12)
        assert ws.cell(row=12, column=3).value  # item description (C12)

    def test_mcl_grouping(self, admin_token):
        """Create 3 manual receipts sharing same vendor/po/do/invoice/date, verify all 3 rows appear."""
        payload = {
            "source_type": "supplier",
            "source_name": "TEST_MCL_Vendor",
            "po_no": "TEST-MCL-PO-1",
            "do_no": "TEST-MCL-DO-1",
            "receive_date": "2026-01-15",
            "items": [
                {"item_name": f"TEST_MCL_Item_{i}", "qty": 10 + i, "unit": "Pcs"} for i in range(3)
            ],
        }
        # /store/incoming uses IncomingGoodsRequest which sets invoice_no=""
        r = admin_token.post(f"{API}/store/incoming", json=payload)
        assert r.status_code == 200, f"incoming create failed: {r.text}"
        # Fetch receipts we just made
        rlist = admin_token.get(f"{API}/store/incoming-report?q=TEST_MCL_Item&page=1&page_size=10")
        items = rlist.json().get("items", [])
        assert len(items) >= 3, f"expected 3, got {len(items)}"
        anchor = items[0]["id"]
        r3 = admin_token.get(f"{API}/store/incoming/mcl/{anchor}")
        assert r3.status_code == 200
        wb = load_workbook(io.BytesIO(r3.content))
        ws = wb["RECEIVED MATERIAL"]
        # Rows 12,13,14 should have C values (item name)
        c12, c13, c14 = ws["C12"].value, ws["C13"].value, ws["C14"].value
        assert c12 and c13 and c14, f"grouping failed: C12={c12} C13={c13} C14={c14}"
        assert "TEST_MCL_Item" in str(c12) and "TEST_MCL_Item" in str(c13)


# ------------------------ Feature E — BOM ------------------------
class TestBOM:
    def test_engineer_can_list(self, engineer_token):
        r = engineer_token.get(f"{API}/bom")
        assert r.status_code == 200
        data = r.json()
        assert "items" in data

    def test_engineer_cannot_access_transactions(self, engineer_token):
        r = engineer_token.post(f"{API}/transactions", json={"invoice_date": "2026-01-01"})
        assert r.status_code == 403
        assert "Engineering" in r.text

    def test_upload_bom_without_reason_when_revision_exists(self, engineer_token):
        with open("/tmp/bom/BOM.xls", "rb") as f:
            files = {"file": ("BOM.xls", f.read(), "application/vnd.ms-excel")}
        # First check if SO 005221 already has revisions
        h = engineer_token.get(f"{API}/bom/history/005221")
        assert h.status_code == 200
        has_existing = h.json().get("count", 0) > 0

        # Reset for upload - re-read file
        with open("/tmp/bom/BOM.xls", "rb") as f:
            files = {"file": ("BOM.xls", f, "application/vnd.ms-excel")}
            r = engineer_token.post(
                f"{API}/bom/upload",
                files=files,
                data={"revision_reason": ""},
            )
        if has_existing:
            assert r.status_code == 400
            assert "revisi" in r.text.lower() or "revision" in r.text.lower()
        else:
            assert r.status_code == 200

    def test_upload_bom_with_reason_creates_next_rev(self, engineer_token):
        with open("/tmp/bom/BOM.xls", "rb") as f:
            files = {"file": ("BOM.xls", f, "application/vnd.ms-excel")}
            r = engineer_token.post(
                f"{API}/bom/upload",
                files=files,
                data={"revision_reason": "test rev by engineer iter15"},
            )
        assert r.status_code == 200, f"upload failed: {r.text}"
        bom = r.json().get("bom", {})
        assert bom.get("so_no") == "005221"
        assert bom.get("rev_no", 0) >= 1
        assert len(bom.get("items", [])) > 0

    def test_search_by_so(self, engineer_token):
        r = engineer_token.get(f"{API}/bom?so_no=005221&rev=latest")
        assert r.status_code == 200
        items = r.json().get("items", [])
        assert len(items) == 1
        assert items[0]["so_no"] == "005221"

    def test_engineer_cannot_annotate(self, engineer_token):
        r = engineer_token.get(f"{API}/bom?so_no=005221&rev=latest")
        bom_id = r.json()["items"][0]["id"]
        payload = {"annotations": [{"item_no": 1, "available_stock": 5, "qty_purchase": 2, "admin_remark": "x"}]}
        r2 = engineer_token.patch(f"{API}/bom/{bom_id}/annotations", json=payload)
        assert r2.status_code == 403
        assert "Admin" in r2.text or "annotasi" in r2.text.lower()

    def test_admin_can_annotate_and_persist(self, admin_token):
        r = admin_token.get(f"{API}/bom?so_no=005221&rev=latest")
        bom_id = r.json()["items"][0]["id"]
        payload = {"annotations": [
            {"item_no": 1, "available_stock": 10, "qty_purchase": 5, "purchase_due_date": "2026-02-01", "admin_remark": "stock ready"},
            {"item_no": 2, "available_stock": 0, "qty_purchase": 2, "admin_remark": "need order"},
        ]}
        r2 = admin_token.patch(f"{API}/bom/{bom_id}/annotations", json=payload)
        assert r2.status_code == 200, r2.text
        # Verify persistence
        r3 = admin_token.get(f"{API}/bom/{bom_id}")
        ann = r3.json().get("annotations", {})
        assert ann.get("1", {}).get("available_stock") == 10
        assert ann.get("2", {}).get("admin_remark") == "need order"

    def test_history(self, engineer_token):
        r = engineer_token.get(f"{API}/bom/history/005221")
        assert r.status_code == 200
        data = r.json()
        assert data["count"] >= 2  # at least Rev.0 and one later
        revs = [x["rev_no"] for x in data["revisions"]]
        assert revs == sorted(revs, reverse=True), "history not newest first"

    def test_store_role_view_only(self, store_token):
        r = store_token.get(f"{API}/bom?so_no=005221")
        assert r.status_code == 200
        # store cannot upload
        with open("/tmp/bom/BOM.xls", "rb") as f:
            files = {"file": ("BOM.xls", f, "application/vnd.ms-excel")}
            r2 = store_token.post(
                f"{API}/bom/upload", files=files, data={"revision_reason": "x"})
        assert r2.status_code == 403

    def test_delete_bom_admin_only(self, admin_token, engineer_token):
        # Upload a throwaway BOM revision, delete as engineer (403), then delete as admin (200)
        with open("/tmp/bom/BOM.xls", "rb") as f:
            files = {"file": ("BOM.xls", f, "application/vnd.ms-excel")}
            r = engineer_token.post(f"{API}/bom/upload", files=files,
                              data={"revision_reason": "throwaway for delete test"})
        assert r.status_code == 200
        bom_id = r.json()["bom"]["id"]
        r2 = engineer_token.delete(f"{API}/bom/{bom_id}")
        assert r2.status_code == 403
        r3 = admin_token.delete(f"{API}/bom/{bom_id}")
        assert r3.status_code == 200


# ------------------------ Bonus — plan_delivery_date ------------------------
class TestPlanDelivery:
    def test_create_transaction_with_plan_delivery_and_export(self, admin_token):
        payload = {
            "invoice_date": "2026-01-10",
            "project_no": "SO-TEST-PDD",
            "po_no": "PO-TEST-PDD-1",
            "vendor_name": "TEST_PDD_Vendor",
            "item_name": "TEST_PDD_Item",
            "qty": 5, "unit": "Pcs", "unit_price": 1000, "total_price": 5000,
            "currency": "IDR", "exchange_rate": 1.0, "total_price_idr": 5000,
            "invoice_no": "INV-PDD-1", "po_date": "2026-01-05",
            "plan_delivery_date": "2026-03-15",
            "receive_date": "2026-01-10",
            "post_to_store": False,
        }
        r = admin_token.post(f"{API}/transactions", json=payload)
        assert r.status_code in (200, 201), r.text
        tx_id = r.json().get("id")
        # GET back
        rg = admin_token.get(f"{API}/transactions/{tx_id}")
        assert rg.status_code == 200
        assert rg.json().get("plan_delivery_date") == "2026-03-15"

        # Update plan_delivery_date
        upd = {**payload, "plan_delivery_date": "2026-04-01"}
        ru = admin_token.put(f"{API}/transactions/{tx_id}", json=upd)
        assert ru.status_code in (200, 201), ru.text
        rg2 = admin_token.get(f"{API}/transactions/{tx_id}")
        assert rg2.json().get("plan_delivery_date") == "2026-04-01"

        # Excel export
        rx = admin_token.get(f"{API}/transactions/export/xlsx")
        assert rx.status_code == 200
        wb = load_workbook(io.BytesIO(rx.content))
        ws = wb.active
        header_row = [c.value for c in ws[1]]
        assert "Plan Delivery" in header_row, f"missing header. got: {header_row}"
        # Verify positioning between 'Tanggal PO' and 'Tanggal Terima'
        idx_po = header_row.index("Tanggal PO")
        idx_pdd = header_row.index("Plan Delivery")
        idx_rcv = header_row.index("Tanggal Terima")
        assert idx_po < idx_pdd < idx_rcv, f"bad column order: {header_row}"

        # Cleanup
        admin_token.delete(f"{API}/transactions/{tx_id}")
