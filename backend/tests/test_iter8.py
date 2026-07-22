"""Iteration 8 tests: Excel exports (incoming-report, issuances, deliveries), SO sort asc, RBAC."""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL').rstrip('/')
API = f"{BASE_URL}/api"


def _login(u, p):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"username": u, "password": p})
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module")
def admin(): return _login("susanto", "admin123")
@pytest.fixture(scope="module")
def store(): return _login("store01", "store123")
@pytest.fixture(scope="module")
def finance(): return _login("finance01", "finance123")
@pytest.fixture(scope="module")
def staff(): return _login("staff01", "staff123")


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _open_wb(resp):
    assert resp.status_code == 200, f"{resp.status_code} {resp.text[:200]}"
    assert XLSX_MIME in resp.headers.get("content-type", "")
    assert len(resp.content) > 500
    return load_workbook(io.BytesIO(resp.content))


# ---------------- Auth 401 ----------------
class TestXlsxAuth:
    def test_incoming_report_xlsx_unauth(self):
        r = requests.get(f"{API}/store/incoming-report/xlsx", allow_redirects=False)
        assert r.status_code in (401, 403), r.status_code

    def test_issuances_xlsx_unauth(self):
        r = requests.get(f"{API}/store/issuances/xlsx", allow_redirects=False)
        assert r.status_code in (401, 403)

    def test_deliveries_xlsx_unauth(self):
        r = requests.get(f"{API}/deliveries/xlsx", allow_redirects=False)
        assert r.status_code in (401, 403)


# ---------------- Incoming Report XLSX ----------------
class TestIncomingReportXlsx:
    def test_admin_download(self, admin):
        r = admin.get(f"{API}/store/incoming-report/xlsx")
        wb = _open_wb(r)
        ws = wb.active
        assert ws.title == "Incoming Goods"
        headers = [c.value for c in ws[1]]
        for h in ["Tgl Terima", "Sumber", "PO No", "DO No", "Ke Stok?", "MCL", "MIF"]:
            assert h in headers, f"missing header {h}, got {headers}"

    def test_store_download(self, store):
        r = store.get(f"{API}/store/incoming-report/xlsx")
        assert r.status_code == 200

    def test_finance_download(self, finance):
        r = finance.get(f"{API}/store/incoming-report/xlsx")
        assert r.status_code == 200

    def test_staff_forbidden(self, staff):
        r = staff.get(f"{API}/store/incoming-report/xlsx")
        assert r.status_code == 403, f"staff should not access store route: {r.status_code}"


# ---------------- Issuances XLSX ----------------
class TestIssuancesXlsx:
    def test_admin_has_price_columns(self, admin):
        r = admin.get(f"{API}/store/issuances/xlsx")
        wb = _open_wb(r)
        headers = [c.value for c in wb.active[1]]
        assert "Avg Unit Price (FIFO)" in headers
        assert "Total Cost" in headers

    def test_finance_has_price_columns(self, finance):
        r = finance.get(f"{API}/store/issuances/xlsx")
        wb = _open_wb(r)
        headers = [c.value for c in wb.active[1]]
        assert "Avg Unit Price (FIFO)" in headers

    def test_store_no_price_columns(self, store):
        r = store.get(f"{API}/store/issuances/xlsx")
        wb = _open_wb(r)
        headers = [c.value for c in wb.active[1]]
        assert "Avg Unit Price (FIFO)" not in headers, f"store should not see prices, headers={headers}"
        assert "Total Cost" not in headers

    def test_staff_forbidden(self, staff):
        r = staff.get(f"{API}/store/issuances/xlsx")
        assert r.status_code == 403


# ---------------- Deliveries XLSX ----------------
class TestDeliveriesXlsx:
    def test_admin_download(self, admin):
        r = admin.get(f"{API}/deliveries/xlsx")
        wb = _open_wb(r)
        ws = wb.active
        assert ws.title == "Pengiriman Barang"
        headers = [c.value for c in ws[1]]
        assert len(headers) == 10, f"expected 10 cols, got {headers}"
        for h in ["Tgl", "No Gatepass", "No DO", "Nama Tujuan", "Nomor SO", "Nama Barang", "Qty", "Unit", "Supir", "Remark"]:
            assert h in headers

    def test_store_can_download(self, store):
        r = store.get(f"{API}/deliveries/xlsx")
        assert r.status_code == 200

    def test_staff_can_download(self, staff):
        # /deliveries/xlsx uses get_current_user (any authenticated)
        r = staff.get(f"{API}/deliveries/xlsx")
        assert r.status_code == 200

    def test_flattened_rows_per_item(self, admin):
        # Get list to compute expected row count
        r = admin.get(f"{API}/deliveries")
        assert r.status_code == 200
        payload = r.json()
        docs = payload.get("items", payload) if isinstance(payload, dict) else payload
        expected_rows = 0
        for d in docs:
            items = d.get("items") or []
            expected_rows += len(items) if items else 1
        # Fetch xlsx
        r2 = admin.get(f"{API}/deliveries/xlsx")
        wb = _open_wb(r2)
        ws = wb.active
        actual_rows = ws.max_row - 1  # exclude header
        # Allow >= because filters differ; but same filter here means equal
        assert actual_rows == expected_rows, f"expected {expected_rows} rows, got {actual_rows}"


# ---------------- SO Sort Asc ----------------
class TestSalesOrderSort:
    def test_sos_sorted_ascending(self, admin):
        r = admin.get(f"{API}/sales-orders")
        assert r.status_code == 200
        sos = r.json()
        so_nos = [s["so_no"] for s in sos]
        assert so_nos == sorted(so_nos), f"SOs not sorted ascending: {so_nos}"
