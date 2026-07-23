"""Iteration 14: Test Feature B (Menu Koreksi Rework) + Feature C (Kategori Transaksi)."""
import os
import uuid
import pytest
import requests
from openpyxl import load_workbook
from io import BytesIO

BASE = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE:
    # fall back to reading frontend/.env
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL"):
                BASE = line.split("=", 1)[1].strip().rstrip("/")
API = f"{BASE}/api"


def _login_session(u, p):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"username": u, "password": p}, timeout=15)
    assert r.status_code == 200, f"login {u} failed: {r.text}"
    return s


@pytest.fixture(scope="module")
def admin_h():
    return _login_session("susanto", "admin123")


@pytest.fixture(scope="module")
def store_h():
    return _login_session("store01", "store123")


# ============= Feature C: Kategori =============

def test_master_categories_endpoint_exists(admin_h):
    r = admin_h.get(f"{API}/master/categories", timeout=10)
    assert r.status_code == 200
    assert isinstance(r.json(), list)


def test_create_tx_with_category_and_verify(admin_h):
    cat = f"TEST_CAT_{uuid.uuid4().hex[:6]}"
    payload = {
        "invoice_date": "2026-01-05",
        "vendor_name": f"TEST_VENDOR_{uuid.uuid4().hex[:4]}",
        "category": cat,
        "item_name": f"TEST_ITEM_{uuid.uuid4().hex[:4]}",
        "qty": 2, "unit": "Ea", "unit_price": 100.0, "total_price": 200.0,
    }
    r = admin_h.post(f"{API}/transactions", json=payload, timeout=10)
    assert r.status_code == 200, r.text
    tx = r.json()
    assert tx["category"] == cat

    # Verify categories endpoint returns it
    r2 = admin_h.get(f"{API}/master/categories", timeout=10)
    assert cat in r2.json()

    # Verify master/items has last_category
    r3 = admin_h.get(f"{API}/master/items", timeout=10)
    items = r3.json()
    match = [i for i in items if i["item_name"] == payload["item_name"]]
    assert match and match[0]["last_category"] == cat

    # cleanup
    admin_h.delete(f"{API}/transactions/{tx['id']}", timeout=10)


def test_create_tx_without_category_defaults_uncategorized(admin_h):
    payload = {
        "invoice_date": "2026-01-05",
        "vendor_name": "TEST_V",
        "item_name": f"TEST_NOCAT_{uuid.uuid4().hex[:4]}",
        "qty": 1, "unit": "Ea", "unit_price": 10.0, "total_price": 10.0,
    }
    r = admin_h.post(f"{API}/transactions", json=payload, timeout=10)
    assert r.status_code == 200, r.text
    tx = r.json()
    assert tx.get("category") == "Uncategorized"
    admin_h.delete(f"{API}/transactions/{tx['id']}", timeout=10)


def test_update_tx_category(admin_h):
    # Create
    payload = {
        "invoice_date": "2026-01-05", "vendor_name": "TEST_V",
        "category": "Direct Material",
        "item_name": f"TEST_UPD_{uuid.uuid4().hex[:4]}",
        "qty": 1, "unit": "Ea", "unit_price": 10.0, "total_price": 10.0,
    }
    r = admin_h.post(f"{API}/transactions", json=payload, timeout=10)
    tx = r.json()
    # Update
    payload["category"] = "Overhead"
    r2 = admin_h.put(f"{API}/transactions/{tx['id']}", json=payload, timeout=10)
    assert r2.status_code == 200
    assert r2.json()["category"] == "Overhead"
    # Verify via GET
    r3 = admin_h.get(f"{API}/transactions/{tx['id']}", timeout=10)
    assert r3.json()["category"] == "Overhead"
    admin_h.delete(f"{API}/transactions/{tx['id']}", timeout=10)


def test_excel_export_has_kategori_column(admin_h):
    r = admin_h.get(f"{API}/transactions/export/xlsx", timeout=30)
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Kategori" in headers, f"'Kategori' not in headers: {headers}"
    # Order check: between Nama Toko and Nama Barang
    idx_toko = headers.index("Nama Toko")
    idx_cat = headers.index("Kategori")
    idx_item = headers.index("Nama Barang")
    assert idx_toko < idx_cat < idx_item


# ============= Feature B: Menu Koreksi Rework =============

def _create_receipt_and_issuance(admin_h):
    """Create a fresh tx → receipt → issuance for correction testing."""
    item = f"TEST_ITEM_B_{uuid.uuid4().hex[:6]}"
    tx_payload = {
        "invoice_date": "2026-01-05", "vendor_name": "TEST_V_B",
        "category": "Direct Material",
        "item_name": item,
        "qty": 10, "unit": "Ea", "unit_price": 1000.0, "total_price": 10000.0,
        "post_to_store": True,
    }
    tx = admin_h.post(f"{API}/transactions", json=tx_payload, timeout=10).json()
    # Receive 10
    rcv_payload = {
        "transaction_id": tx["id"], "qty_received": 10, "receive_date": "2026-01-05",
    }
    rcv = admin_h.post(f"{API}/store/receive", json=rcv_payload, timeout=10).json()
    # Issue 3
    iss_payload = {
        "item_name": item, "qty": 3, "issue_date": "2026-01-06",
        "taker_name": "IT12Tester", "so_number": "SO-OLD",
    }
    iss = admin_h.post(f"{API}/store/issue", json=iss_payload, timeout=10).json()
    return tx, rcv, iss


def test_correction_edit_taker_name_auto_applied(admin_h):
    tx, rcv, iss = _create_receipt_and_issuance(admin_h)
    # Submit correction
    new_taker = f"NEW_TAKER_{uuid.uuid4().hex[:4]}"
    req = admin_h.post(f"{API}/store/requests", json={
        "target_type": "issuance", "target_id": iss["id"],
        "action_type": "edit", "reason": "correcting taker",
        "proposed_changes": {"field": "taker_name", "new_value": new_taker},
    }, timeout=10)
    assert req.status_code == 200, req.text
    rid = req.json()["id"]
    # verify snapshot old_value
    assert req.json()["proposed_changes"]["old_value"] == "IT12Tester"
    # Approve
    rev = admin_h.post(f"{API}/store/requests/{rid}/review", json={
        "approve": True, "review_note": "ok",
    }, timeout=10)
    assert rev.status_code == 200
    # Verify issuance updated
    r = admin_h.get(f"{API}/store/issuances", params={"page": 1, "page_size": 200}, timeout=10)
    match = [i for i in r.json()["items"] if i["id"] == iss["id"]]
    assert match and match[0]["taker_name"] == new_taker


def test_correction_edit_so_number_auto_applied(admin_h):
    tx, rcv, iss = _create_receipt_and_issuance(admin_h)
    new_so = "SO-NEW-99"
    req = admin_h.post(f"{API}/store/requests", json={
        "target_type": "issuance", "target_id": iss["id"],
        "action_type": "edit", "reason": "correcting SO",
        "proposed_changes": {"field": "so_number", "new_value": new_so},
    }, timeout=10).json()
    admin_h.post(f"{API}/store/requests/{req['id']}/review",
                  json={"approve": True}, timeout=10)
    r = admin_h.get(f"{API}/store/issuances", params={"so_number": new_so}, timeout=10)
    match = [i for i in r.json()["items"] if i["id"] == iss["id"]]
    assert match and match[0]["so_number"] == new_so


def test_correction_edit_qty_decrease_refunds_receipt(admin_h):
    tx, rcv, iss = _create_receipt_and_issuance(admin_h)
    # receipt qty_remaining should be 10-3 = 7 after issuance
    receipts = admin_h.get(f"{API}/store/receipts",
                            params={"transaction_id": tx["id"]}, timeout=10).json()
    assert receipts[0]["qty_remaining"] == 7

    # Correct qty from 3 → 2 (refund 1)
    req = admin_h.post(f"{API}/store/requests", json={
        "target_type": "issuance", "target_id": iss["id"],
        "action_type": "edit", "reason": "less used",
        "proposed_changes": {"field": "qty", "new_value": 2},
    }, timeout=10).json()
    rev = admin_h.post(f"{API}/store/requests/{req['id']}/review",
                        json={"approve": True}, timeout=10)
    assert rev.status_code == 200
    # Verify receipt refunded → qty_remaining = 8
    receipts2 = admin_h.get(f"{API}/store/receipts",
                             params={"transaction_id": tx["id"]}, timeout=10).json()
    assert abs(receipts2[0]["qty_remaining"] - 8) < 1e-6, receipts2
    # Verify issuance qty = 2
    r = admin_h.get(f"{API}/store/issuances", params={"page_size": 200}, timeout=10).json()
    match = [i for i in r.json()["items"] if i["id"] == iss["id"]] if False else \
            [i for i in r["items"] if i["id"] == iss["id"]]
    assert match and abs(match[0]["qty"] - 2) < 1e-6
    # allocations rescaled
    assert abs(match[0]["allocations"][0]["qty"] - 2) < 1e-6


def test_correction_qty_increase_insufficient_stock_returns_400(admin_h):
    tx, rcv, iss = _create_receipt_and_issuance(admin_h)
    # Issue rest → receipt.qty_remaining = 0
    item = tx["item_name"]
    iss2 = admin_h.post(f"{API}/store/issue", json={
        "item_name": item, "qty": 7, "issue_date": "2026-01-06",
        "taker_name": "TAKER2", "so_number": "SO-X",
    }, timeout=10).json()
    # Now try to increase qty of first issuance from 3 → 5 (needs +2 but receipt empty)
    req = admin_h.post(f"{API}/store/requests", json={
        "target_type": "issuance", "target_id": iss["id"],
        "action_type": "edit", "reason": "more",
        "proposed_changes": {"field": "qty", "new_value": 5},
    }, timeout=10).json()
    rev = admin_h.post(f"{API}/store/requests/{req['id']}/review",
                        json={"approve": True}, timeout=10)
    assert rev.status_code == 400
    assert "tidak cukup" in rev.text.lower() or "stok" in rev.text.lower()


def test_correction_delete_refunds_allocations(admin_h):
    tx, rcv, iss = _create_receipt_and_issuance(admin_h)
    # Delete request
    req = admin_h.post(f"{API}/store/requests", json={
        "target_type": "issuance", "target_id": iss["id"],
        "action_type": "delete", "reason": "revert",
    }, timeout=10).json()
    rev = admin_h.post(f"{API}/store/requests/{req['id']}/review",
                        json={"approve": True}, timeout=10)
    assert rev.status_code == 200
    # receipt fully refunded to 10
    receipts = admin_h.get(f"{API}/store/receipts",
                            params={"transaction_id": tx["id"]}, timeout=10).json()
    assert abs(receipts[0]["qty_remaining"] - 10) < 1e-6
    # Issuance gone
    r = admin_h.get(f"{API}/store/issuances", params={"page_size": 200}, timeout=10).json()
    match = [i for i in r["items"] if i["id"] == iss["id"]]
    assert not match


def test_correction_invalid_field_rejected(admin_h):
    tx, rcv, iss = _create_receipt_and_issuance(admin_h)
    req = admin_h.post(f"{API}/store/requests", json={
        "target_type": "issuance", "target_id": iss["id"],
        "action_type": "edit", "reason": "hacky",
        "proposed_changes": {"field": "total_cost", "new_value": 999999},
    }, timeout=10)
    assert req.status_code == 400
    assert "field" in req.text.lower() or "koreksi" in req.text.lower()


def test_correction_qty_zero_or_negative_rejected(admin_h):
    tx, rcv, iss = _create_receipt_and_issuance(admin_h)
    req = admin_h.post(f"{API}/store/requests", json={
        "target_type": "issuance", "target_id": iss["id"],
        "action_type": "edit", "reason": "bad",
        "proposed_changes": {"field": "qty", "new_value": 0},
    }, timeout=10)
    assert req.status_code == 400
