"""Iteration 13 backend tests.
Covers:
  - GET /api/store/issuances/takers (new endpoint)
  - Gemini migration (POST /api/transactions/parse-po no longer references EMERGENT_LLM_KEY)
  - Incoming report xlsx contains 'Nomor SO' header
"""
import io
import os
import uuid
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL').rstrip('/')
API = f"{BASE_URL}/api"


def _login(u, p):
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{API}/auth/login", json={"username": u, "password": p})
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="session")
def admin(): return _login("susanto", "admin123")

@pytest.fixture(scope="session")
def store(): return _login("store01", "store123")

@pytest.fixture(scope="session")
def finance(): return _login("finance01", "finance123")


# ---------------- Taker autocomplete ----------------
class TestIssuanceTakers:
    def test_endpoint_returns_200_array(self, admin):
        r = admin.get(f"{API}/store/issuances/takers")
        assert r.status_code == 200, r.text
        d = r.json()
        assert isinstance(d, list)
        # Every element should be a non-empty string
        for name in d:
            assert isinstance(name, str) and name.strip()

    def test_store_can_access(self, store):
        r = store.get(f"{API}/store/issuances/takers")
        assert r.status_code == 200

    def test_finance_can_access(self, finance):
        r = finance.get(f"{API}/store/issuances/takers")
        assert r.status_code == 200

    def test_unauth_denied(self):
        r = requests.get(f"{API}/store/issuances/takers")
        assert r.status_code == 401

    def test_takers_include_created_name(self, admin, store):
        # seed a receipt+issuance to guarantee a taker appears
        item = f"TEST_ITER13_TAKER_ITEM_{uuid.uuid4().hex[:6]}"
        taker = f"TEST_TAKER_{uuid.uuid4().hex[:6]}"
        pay_in = {
            "receive_date": "2025-05-01",
            "source_type": "supplier",
            "source_name": "TEST_ITER13_SUP",
            "items": [{"item_name": item, "qty": 3, "unit": "Ea",
                       "add_to_stock": True, "unit_price": 1.0}],
        }
        r = admin.post(f"{API}/store/incoming", json=pay_in)
        assert r.status_code == 200, r.text
        # issue 1 out
        iss = {"items": [{"item_name": item, "qty": 1,
                          "taker_name": taker, "issue_date": "2025-05-02",
                          "so_number": "TEST_SO"}]}
        ir = store.post(f"{API}/store/issue/bulk", json=iss)
        assert ir.status_code == 200, ir.text

        # takers list should include our new taker
        r2 = admin.get(f"{API}/store/issuances/takers")
        assert r2.status_code == 200
        assert taker in r2.json()


# ---------------- Gemini migration ----------------
class TestGeminiMigration:
    def _tiny_png(self):
        # minimal 1x1 PNG bytes
        import base64
        return base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABAQMAAAAl21bKAAAAA1BMVEX///+nxBvIAAAAC0lEQVR4nGNgAAIAAAUAAeImBZsAAAAASUVORK5CYII="
        )

    def test_endpoint_reachable_no_emergent_error(self, admin):
        files = {"file": ("test.png", self._tiny_png(), "image/png")}
        # Strip content-type so multipart is set correctly
        s = requests.Session(); s.cookies.update(admin.cookies)
        r = s.post(f"{API}/transactions/parse-po", files=files)
        # Should NOT be 200 (bad key/image), but MUST NOT mention EMERGENT_LLM_KEY
        body = r.text.lower()
        assert "emergent_llm_key" not in body, f"Old emergent error still present: {r.text}"
        # And endpoint must not be 404 (i.e., it's reachable)
        assert r.status_code != 404, f"Endpoint not found: {r.text}"
        # Also should not be 500 due to ImportError of emergentintegrations
        assert "emergentintegrations" not in body, f"Unexpected emergentintegrations reference: {r.text}"

    def test_unauth_denied(self):
        files = {"file": ("test.png", b"x", "image/png")}
        r = requests.post(f"{API}/transactions/parse-po", files=files)
        assert r.status_code == 401


# ---------------- Incoming xlsx contains Nomor SO ----------------
class TestIncomingReportXlsxSO:
    def test_xlsx_has_nomor_so_header(self, admin):
        r = admin.get(f"{API}/store/incoming-report/xlsx")
        assert r.status_code == 200, r.text
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Nomor SO" in headers, f"Expected 'Nomor SO' in xlsx headers, got {headers}"


# ---------------- Cleanup ----------------
class TestZCleanup13:
    def test_cleanup(self, admin):
        # cleanup TEST_ITER13 receipts by q param
        rep = admin.get(f"{API}/store/incoming-report", params={"q": "TEST_ITER13"}).json()
        # nothing to delete via API directly; just verify no error
        assert isinstance(rep.get("items"), list)
