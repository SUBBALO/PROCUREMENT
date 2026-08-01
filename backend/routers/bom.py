"""Bill of Material (BOM) module.

Engineering / Admin upload BOM Excel (.xls or .xlsx) → parsed & stored as a revision.
Search by SO No. History of all revisions kept indefinitely. Admin can annotate items
with Available Stock, Qty Purchase, Purchase Due Date, and Remark.
"""
import io
import re
import uuid
from datetime import datetime, date, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel

from db import db
from deps import (
    get_current_user,
    log_action,
    require_bom_admin,
    require_bom_upload,
    require_bom_edit,
)
from services.soft_delete import NOT_DELETED_FILTER, merged, soft_delete_one


router = APIRouter(prefix="/bom", tags=["bom"])


# ------------------------------ Models ------------------------------
class BOMItem(BaseModel):
    item_no: int
    item_name: str
    item_specification: str = ""
    qty: float = 0.0
    uom: str = ""
    material: str = ""
    weight_kg: Optional[float] = None
    purchase_due_date: str = ""  # ISO date YYYY-MM-DD, optional
    remark: str = ""  # e.g. "P1", "P2", "P1&P2"


class BOMAnnotation(BaseModel):
    item_no: int
    available_stock: Optional[float] = None
    qty_purchase: Optional[float] = None
    purchase_due_date: Optional[str] = None
    admin_remark: str = ""


class BOMAnnotationsUpdate(BaseModel):
    annotations: List[BOMAnnotation]


# ------------------------------ Helpers ------------------------------
def _excel_serial_to_iso(v) -> Optional[str]:
    """Convert an Excel date serial number to ISO date string. Excel epoch = 1899-12-30."""
    try:
        if v is None or v == "":
            return None
        if isinstance(v, str):
            s = v.strip().replace(":", "").strip()
            if not s:
                return None
            # Try common formats
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(s, fmt).date().isoformat()
                except ValueError:
                    continue
            return s
        n = float(v)
        from datetime import timedelta
        return (datetime(1899, 12, 30) + timedelta(days=n)).date().isoformat()
    except Exception:
        return None


def _clean_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lstrip(":").strip()


def normalize_so_no(v) -> str:
    """Normalise Sales Order number → kanonik 6 digit (zero-padded).

    Rules:
    - Trim whitespace and leading ':' artefacts from Excel imports.
    - Jika murni numerik: zero-pad ke 6 digit (mis. '5251' → '005251', '1234' → '001234').
      Jika sudah > 6 digit, biarkan apa adanya (tidak dipotong).
    - Non-numeric / alfanumerik (mis. 'SO-2026-001') dikembalikan apa adanya.
    - Returns empty string for empty / None input.
    """
    s = _clean_str(v)
    if not s:
        return ""
    if s.isdigit():
        return s.zfill(6)
    return s


def _parse_bom_workbook(rows: List[List]) -> dict:
    """Parse a 2D list of rows (already extracted from either .xls or .xlsx) into BOM header+items.

    Row layout matches the MKS BOM template:
      R4: [TO, ...] ... [BOM.NO., '', ':NNN']
      R5: [DATE, ':dd/mm/yyyy'] ...  [REV.NO., '', ':N']
      R7: labels (PROJECT, ENG.DRW., CUSTOMER, CLASS OF MATERIAL, SO.NO., DELIVERY DATE)
      R8: header values
      R9: item column headers
      R11..: item rows until a blank row or 'NOTES :' row
    """
    def cell(r, c):
        try:
            return rows[r - 1][c - 1]
        except IndexError:
            return ""

    def find_row(startswith_text: str, max_rows: int = 20) -> int:
        for i in range(1, min(max_rows, len(rows)) + 1):
            for j in range(1, min(6, len(rows[i - 1])) + 1):
                v = _clean_str(cell(i, j))
                if v and v.upper().startswith(startswith_text.upper()):
                    return i
        return -1

    header = {}

    # BOM.NO and REV.NO are on R4 and R5 respectively, in columns 12-14
    header["bom_no"] = _clean_str(cell(4, 14) or cell(4, 13) or cell(4, 12))
    header["rev_no_raw"] = _clean_str(cell(5, 14) or cell(5, 13) or cell(5, 12))
    header["to"] = _clean_str(cell(4, 2))
    header["date"] = _excel_serial_to_iso(cell(5, 2)) or _clean_str(cell(5, 2))

    # R7 = labels, R8 = values (columns are shifted per template)
    header["project_name"] = _clean_str(cell(8, 1))
    header["project_dwg"] = _clean_str(cell(8, 4))
    header["customer"] = _clean_str(cell(8, 7))
    header["class_material"] = _clean_str(cell(8, 11))
    header["so_no"] = normalize_so_no(cell(8, 14))
    header["delivery_date"] = _excel_serial_to_iso(cell(8, 16)) or _clean_str(cell(8, 16))

    # Fallback: if SO no still empty, scan header rows for any col with numeric value
    if not header["so_no"]:
        for c in range(13, 17):
            v = _clean_str(cell(8, c))
            if v and v.isdigit():
                header["so_no"] = normalize_so_no(v)
                break

    # Items start at row 11, stop when either blank or "NOTES" prefix
    items: List[dict] = []
    for r in range(11, len(rows) + 1):
        no_raw = cell(r, 1)
        # stop marker: 'NOTES : ...' anywhere on the row's first col
        first_col = _clean_str(no_raw)
        if first_col.upper().startswith("NOTES") or first_col.upper().startswith("TOTAL WEIGHT"):
            break
        # skip completely-blank rows
        if all((_clean_str(cell(r, c)) == "") for c in range(1, min(18, len(rows[r - 1]) + 1) if rows[r - 1] else 1)):
            continue
        try:
            item_no = int(float(no_raw)) if no_raw not in ("", None) else None
        except (ValueError, TypeError):
            item_no = None
        if item_no is None:
            continue
        items.append({
            "item_no": item_no,
            "item_name": _clean_str(cell(r, 2)),
            "item_specification": _clean_str(cell(r, 3)),
            "qty": float(cell(r, 8) or 0),
            "uom": _clean_str(cell(r, 9)),
            "material": _clean_str(cell(r, 11)),
            "weight_kg": (float(cell(r, 12)) if cell(r, 12) not in ("", None) else None),
            "remark": _clean_str(cell(r, 17)),
        })

    return {"header": header, "items": items}


def _read_workbook(content: bytes, filename: str) -> List[List]:
    """Read either .xls (Excel 97-2003) or .xlsx (2007+) bytes and return a 2D list of cell values.

    Falls back to magic-byte detection if the extension is missing / wrong so users can drop
    a `.xls` file even when the browser reports an odd content-type.
    """
    fname = (filename or "").lower()
    is_xlsx = fname.endswith(".xlsx") or fname.endswith(".xlsm")
    is_xls = fname.endswith(".xls")

    # Magic-byte fallback (Excel 97-2003 → OLE2 signature D0CF11E0..., xlsx → PK\x03\x04 ZIP)
    if not (is_xlsx or is_xls) and content:
        if content[:4] == b"PK\x03\x04":
            is_xlsx = True
        elif content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            is_xls = True

    if is_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        return [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]
    if is_xls:
        try:
            import xlrd
        except ImportError as e:
            raise HTTPException(status_code=500, detail=f"Library xlrd tidak terpasang di server: {e}")
        try:
            wb = xlrd.open_workbook(file_contents=content, formatting_info=False)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Gagal membaca file .xls (Excel 97-2003): {e}")
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

    raise HTTPException(status_code=400, detail="Format tidak didukung. Gunakan .xls (Excel 97-2003) atau .xlsx (Excel 2007+).")


# ------------------------------ Endpoints ------------------------------
@router.get("/preparers")
async def bom_preparers(current: dict = Depends(get_current_user)):
    """Distinct list of previously-entered `prepared_by` names for autocomplete."""
    names = await db.boms.distinct("prepared_by")
    return sorted({str(n).strip() for n in names if n and str(n).strip()})


# ============ BOM Number Auto-Generate (Monthly) ============
async def _current_max_bom_seq(mm: str, yyyy: str) -> int:
    """Cari seq tertinggi dari BOMs aktif untuk bulan tsb — sumber kebenaran.
    Format: BOM{seq:03d}-{MM}-{YYYY}. Return 0 kalau belum ada."""
    pattern = rf"^BOM(\d{{3}})-{mm}-{yyyy}$"
    cursor = db.boms.find(
        {"bom_no": {"$regex": pattern}, "deleted_at": {"$exists": False}},
        {"_id": 0, "bom_no": 1},
    )
    max_seq = 0
    async for d in cursor:
        m = re.match(pattern, d.get("bom_no", ""))
        if m:
            try: max_seq = max(max_seq, int(m.group(1)))
            except Exception: pass
    return max_seq


async def _next_bom_no() -> dict:
    """Generate next BOM number: BOM{seq:03d}-{MM}-{YYYY}. Idempotent — hitung dari actual DB max.
    Register loop-in-caller handles concurrent collisions by retrying with +1."""
    now = datetime.utcnow()
    mm = f"{now.month:02d}"
    yyyy = str(now.year)
    max_seq = await _current_max_bom_seq(mm, yyyy)
    seq = max_seq + 1
    bom_no = f"BOM{seq:03d}-{mm}-{yyyy}"
    return {"bom_no": bom_no, "seq": seq, "mm": mm, "yyyy": yyyy}


@router.get("/next-number")
async def preview_next_bom_no(current: dict = Depends(get_current_user)):
    """Preview next BOM number tanpa side-effect. Idempotent — dihitung dari actual max BOM di DB."""
    now = datetime.utcnow()
    mm = f"{now.month:02d}"
    yyyy = str(now.year)
    max_seq = await _current_max_bom_seq(mm, yyyy)
    next_seq = max_seq + 1
    # Find last created BOM (any month) — for context
    last_doc = await db.boms.find_one(
        {"bom_no": {"$regex": r"^BOM\d{3}-\d{2}-\d{4}$"}, "deleted_at": {"$exists": False}},
        {"_id": 0, "bom_no": 1, "uploaded_at": 1},
        sort=[("uploaded_at", -1)],
    )
    # Last BOM this month = the max seq one (bukan yg baru upload, biar match dengan preview)
    last_this_month_no = f"BOM{max_seq:03d}-{mm}-{yyyy}" if max_seq > 0 else None
    return {
        "preview": f"BOM{next_seq:03d}-{mm}-{yyyy}",
        "next_seq": next_seq,
        "mm": mm,
        "yyyy": yyyy,
        "last_bom_no": (last_doc or {}).get("bom_no"),
        "last_bom_no_this_month": last_this_month_no,
    }


class BomRegisterIn(BaseModel):
    bom_no: str = ""              # kosong → auto-generate
    so_no: str = ""
    project_name: str = ""
    project_dwg: str = ""
    customer: str = ""
    class_material: str = ""
    delivery_date: str = ""
    bom_date: str = ""
    prepared_by: str = ""
    remark: str = ""
    drawing_id: str = ""          # link ke drawing (untuk repeat order)
    is_repeat: bool = False
    source_bom_id: str = ""       # copy items dari BOM ini (untuk repeat order)


@router.post("/register")
async def register_bom_manual(payload: BomRegisterIn, current: dict = Depends(require_bom_edit)):
    """Register BOM manual (tanpa Excel upload).
    Items bisa ditambahkan nanti via UI atau upload Excel lanjutan.
    Untuk repeat order: kirim drawing_id + is_repeat=true. Backend copy drawing_no ke project_dwg.
    Jika `source_bom_id` di-set, items dari source BOM akan di-copy ke BOM baru.
    """
    auto = False
    bom_no = (payload.bom_no or "").strip()
    if not bom_no:
        info = await _next_bom_no()
        bom_no = info["bom_no"]
        auto = True
        # Race-safety: kalau BOM dgn nomor ini sudah ada (concurrent create), naikkan seq sampai unique
        _max_retry = 20
        while _max_retry > 0:
            _clash = await db.boms.find_one({"bom_no": bom_no, "deleted_at": {"$exists": False}}, {"_id": 1})
            if not _clash:
                break
            info = await _next_bom_no()
            bom_no = info["bom_no"]
            _max_retry -= 1

    existing = await db.boms.find_one({"bom_no": bom_no, "deleted_at": {"$exists": False}})
    if existing:
        raise HTTPException(status_code=409, detail=f"BOM '{bom_no}' sudah ada")

    # Resolve linked drawing (for repeat orders)
    drawing_ref = None
    project_dwg = payload.project_dwg.strip()
    if payload.drawing_id:
        drawing_ref = await db.drawings.find_one(
            {"id": payload.drawing_id, "deleted_at": {"$exists": False}},
            {"_id": 0, "drawing_no": 1, "project_name": 1, "customer_code": 1, "title": 1},
        )
        if not drawing_ref:
            raise HTTPException(status_code=404, detail="Drawing referensi tidak ditemukan")
        # Inherit drawing info
        if not project_dwg:
            project_dwg = drawing_ref.get("drawing_no", "")

    # Copy items from source BOM (repeat order — items only, tidak ikut bom_no/status)
    copied_items = []
    source_bom_ref = None
    if payload.source_bom_id:
        source_bom_ref = await db.boms.find_one(
            {"id": payload.source_bom_id, "deleted_at": {"$exists": False}},
            {"_id": 0},
        )
        if not source_bom_ref:
            raise HTTPException(status_code=404, detail="Source BOM tidak ditemukan")
        # Deep-copy items (regenerate item id if present, drop server-side fields)
        import copy as _copy
        for it in (source_bom_ref.get("items") or []):
            new_it = _copy.deepcopy(it)
            new_it.pop("_id", None)
            # Reset id per item supaya tidak bentrok
            if "id" in new_it:
                new_it["id"] = str(uuid.uuid4())
            copied_items.append(new_it)

    user_name = current.get("name") or current.get("username")
    doc = {
        "id": str(uuid.uuid4()),
        "so_no": payload.so_no.strip(),
        "rev_no": 0,
        "bom_no": bom_no,
        "bom_no_auto": auto,
        "project_name": (payload.project_name or (source_bom_ref or {}).get("project_name") or (drawing_ref or {}).get("project_name") or "").strip(),
        "project_dwg": project_dwg,
        "customer": (payload.customer or (source_bom_ref or {}).get("customer") or (drawing_ref or {}).get("customer_code") or "").strip(),
        "class_material": (payload.class_material or (source_bom_ref or {}).get("class_material") or "").strip(),
        "delivery_date": payload.delivery_date.strip(),
        "bom_date": payload.bom_date.strip() or datetime.utcnow().date().isoformat(),
        "prepared_by": (payload.prepared_by or user_name or "").strip(),
        "items": copied_items,
        "annotations": (source_bom_ref or {}).get("annotations", {}) if source_bom_ref else {},
        "revision_reason": "",
        "auto_generated": auto,
        "source": "repeat_order" if payload.is_repeat else "manual_register",
        "is_repeat": bool(payload.is_repeat),
        "drawing_id": payload.drawing_id or None,
        "drawing_no": (drawing_ref or {}).get("drawing_no") if drawing_ref else "",
        "source_bom_id": payload.source_bom_id or None,
        "source_bom_no": (source_bom_ref or {}).get("bom_no") if source_bom_ref else None,
        "copied_items_count": len(copied_items),
        "uploaded_by_id": current.get("id"),
        "uploaded_by_name": user_name,
        "uploaded_by_role": current.get("role"),
        "uploaded_at": datetime.utcnow().isoformat(),
        "original_filename": None,
        "remark": payload.remark.strip(),
        # ---- Engineering Approval Workflow (Iter 35) ----
        # draft → pending_review → approved (auto-register to main BOM list)
        "engineering_status": "draft",
        "signatures": {
            "prepared_by": None,        # {name, user_id, at} — set saat submit-review
            "checked_by": None,         # Riskinova / eng_head — set saat approve-review
            "acknowledged_by": None,    # Purchasing (Susanto)
            "approved_by": None,        # Admin (Erwin)
        },
    }
    await db.boms.insert_one(doc.copy())
    await log_action(current, "bom_register_manual", "bom", doc["id"], {"bom_no": bom_no, "so_no": doc["so_no"], "auto": auto, "is_repeat": payload.is_repeat, "source_bom_id": payload.source_bom_id, "copied_items": len(copied_items)})
    doc.pop("_id", None)
    return doc


@router.get("/lookup")
async def lookup_boms(q: str = "", limit: int = 20, current: dict = Depends(get_current_user)):
    """Lightweight BOM lookup for dropdown (returns bom_no + so_no + project_name + items_count)."""
    filt = {"deleted_at": {"$exists": False}}
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [{"bom_no": rx}, {"so_no": rx}, {"project_name": rx}, {"project_dwg": rx}, {"customer": rx}]
    cursor = db.boms.find(filt, {"_id": 0, "id": 1, "bom_no": 1, "so_no": 1, "project_name": 1, "customer": 1, "project_dwg": 1, "items": 1}).sort("uploaded_at", -1).limit(limit)
    docs = await cursor.to_list(length=limit)
    items = []
    for d in docs:
        if not d.get("bom_no"):
            continue
        items.append({
            "id": d.get("id"),
            "bom_no": d.get("bom_no"),
            "so_no": d.get("so_no"),
            "project_name": d.get("project_name"),
            "project_dwg": d.get("project_dwg"),
            "customer": d.get("customer"),
            "items_count": len(d.get("items") or []),
        })
    return {"items": items}


# ============ BOM Item Manual Entry ============
class BOMItemIn(BaseModel):
    item_no: Optional[int] = None  # kosong = auto next
    item_name: str
    item_specification: str = ""
    qty: float = 0
    uom: str = ""
    material: str = ""
    weight_kg: Optional[float] = None
    purchase_due_date: str = ""
    remark: str = ""


@router.post("/{bom_id}/items")
async def add_bom_item(bom_id: str, payload: BOMItemIn, current: dict = Depends(require_bom_edit)):
    """Tambah 1 item ke BOM secara manual."""
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    if not payload.item_name.strip():
        raise HTTPException(status_code=400, detail="item_name wajib")
    items = bom.get("items") or []
    # Auto next item_no
    next_no = payload.item_no
    if next_no is None:
        existing_nos = [int(it.get("item_no") or 0) for it in items]
        next_no = (max(existing_nos, default=0)) + 1
    # Uniqueness within this BOM
    if any(int(it.get("item_no") or 0) == int(next_no) for it in items):
        raise HTTPException(status_code=409, detail=f"item_no {next_no} sudah ada di BOM ini")
    new_item = {
        "item_no": int(next_no),
        "item_name": payload.item_name.strip(),
        "item_specification": payload.item_specification.strip(),
        "qty": float(payload.qty or 0),
        "uom": payload.uom.strip(),
        "material": payload.material.strip(),
        "weight_kg": payload.weight_kg,
        "purchase_due_date": (payload.purchase_due_date or "").strip(),
        "remark": payload.remark.strip(),
    }
    items.append(new_item)
    items.sort(key=lambda x: int(x.get("item_no") or 0))
    await db.boms.update_one({"id": bom_id}, {"$set": {"items": items, "updated_at": datetime.utcnow().isoformat()}})
    await log_action(current, "bom_item_add", "bom", bom_id, {"item_no": next_no, "item_name": new_item["item_name"]})
    return {"success": True, "item": new_item, "total_items": len(items)}


@router.put("/{bom_id}/items/{item_no}")
async def update_bom_item(bom_id: str, item_no: int, payload: BOMItemIn, current: dict = Depends(get_current_user)):
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    items = bom.get("items") or []
    updated = False
    for it in items:
        if int(it.get("item_no") or 0) == int(item_no):
            it["item_name"] = payload.item_name.strip()
            it["item_specification"] = payload.item_specification.strip()
            it["qty"] = float(payload.qty or 0)
            it["uom"] = payload.uom.strip()
            it["material"] = payload.material.strip()
            it["weight_kg"] = payload.weight_kg
            it["purchase_due_date"] = (payload.purchase_due_date or "").strip()
            it["remark"] = payload.remark.strip()
            updated = True
            break
    if not updated:
        raise HTTPException(status_code=404, detail=f"Item no {item_no} tidak ditemukan")
    await db.boms.update_one({"id": bom_id}, {"$set": {"items": items, "updated_at": datetime.utcnow().isoformat()}})
    return {"success": True}


@router.delete("/{bom_id}/items/{item_no}")
async def delete_bom_item(bom_id: str, item_no: int, current: dict = Depends(require_bom_edit)):
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    items = bom.get("items") or []
    new_items = [it for it in items if int(it.get("item_no") or 0) != int(item_no)]
    if len(new_items) == len(items):
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")
    await db.boms.update_one({"id": bom_id}, {"$set": {"items": new_items, "updated_at": datetime.utcnow().isoformat()}})
    await log_action(current, "bom_item_delete", "bom", bom_id, {"item_no": item_no})
    return {"success": True, "remaining": len(new_items)}


# ============ BOM Grid Bulk Save (Excel-like entry) ============
class BOMItemsBulkIn(BaseModel):
    items: List[BOMItemIn]


@router.post("/{bom_id}/items-bulk")
async def replace_bom_items(bom_id: str, payload: BOMItemsBulkIn, current: dict = Depends(require_bom_edit)):
    """Replace **all** items in this BOM with the provided list (Excel-like grid save).
    Item numbers are auto-assigned sequentially (1, 2, 3, ...) — client can omit item_no.
    Only allowed while BOM is in draft or pending_review (engineering can still tweak).
    """
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")

    status_now = bom.get("engineering_status", "approved")
    has_items = bool(bom.get("items") or [])
    # Allow edits in draft/pending_review, OR when BOM is approved but has 0 items (legacy pre-Iter35 grace path)
    if status_now not in ("draft", "pending_review") and has_items:
        raise HTTPException(
            status_code=409,
            detail=f"BOM sudah di-approve (status={status_now}) dan sudah punya {len(bom.get('items') or [])} item. Items tidak bisa diubah langsung — gunakan revisi.",
        )
    # Auto-heal: kalau BOM belum ada status field atau approved-empty, downgrade ke draft supaya masuk workflow benar
    auto_reset_draft = (status_now == "approved" and not has_items) or not bom.get("engineering_status")

    new_items = []
    for idx, it in enumerate(payload.items, start=1):
        name = (it.item_name or "").strip()
        spec = (it.item_specification or "").strip()
        # Skip completely empty rows silently
        if not name and not spec and not (it.qty or 0):
            continue
        new_items.append({
            "item_no": idx,
            "item_name": name or spec,  # fallback: if spec-only row, put spec as name
            "item_specification": spec,
            "qty": float(it.qty or 0),
            "uom": (it.uom or "").strip(),
            "material": (it.material or "").strip(),
            "weight_kg": it.weight_kg,
            "purchase_due_date": (it.purchase_due_date or "").strip(),
            "remark": (it.remark or "").strip(),
        })

    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {"items": new_items, "updated_at": datetime.utcnow().isoformat(),
                  "last_edited_by": current.get("name") or current.get("username"),
                  "last_edited_at": datetime.utcnow().isoformat()}},
    )
    # Auto-heal legacy: kalau BOM tadinya approved-kosong / no engineering_status, set ke draft supaya masuk workflow
    if auto_reset_draft:
        await db.boms.update_one({"id": bom_id}, {"$set": {"engineering_status": "draft"}})
    await log_action(current, "bom_items_bulk_replace", "bom", bom_id, {"count": len(new_items), "auto_reset_draft": auto_reset_draft})
    return {"success": True, "count": len(new_items), "auto_reset_draft": auto_reset_draft}


# ============ BOM Engineering Approval Workflow ============
def _sig_stamp(current: dict, name_override: str = "") -> dict:
    return {
        "name": (name_override or current.get("name") or current.get("username") or "").strip(),
        "user_id": current.get("id"),
        "username": current.get("username"),
        "role": current.get("role"),
        "at": datetime.utcnow().isoformat(),
    }


@router.post("/{bom_id}/submit-review")
async def bom_submit_review(bom_id: str, current: dict = Depends(get_current_user)):
    """Engineer submit BOM draft ke Engineering Leader untuk review.
    Status: draft → pending_review. Stamp Prepared By signature.
    """
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    status_now = bom.get("engineering_status", "approved")
    # Allow submit dari draft, pending_review (re-submit), atau approved-empty (legacy grace)
    if status_now not in ("draft", "pending_review") and (bom.get("items") or []):
        raise HTTPException(status_code=409, detail=f"Status BOM ({status_now}) tidak bisa disubmit ulang (sudah approved dengan items).")
    if not (bom.get("items") or []):
        raise HTTPException(status_code=400, detail="BOM belum ada item — isi minimal 1 item sebelum submit")

    sigs = dict(bom.get("signatures") or {})
    sigs["prepared_by"] = _sig_stamp(current)

    # Riski / Engineering Leader auto-approve → langsung ke status approved, skip pending_review
    is_eng_leader = (current or {}).get("role") in ("eng_leader", "eng_head")
    if is_eng_leader:
        sigs["checked_by"] = _sig_stamp(current)
        await db.boms.update_one(
            {"id": bom_id},
            {"$set": {
                "engineering_status": "approved",
                "signatures": sigs,
                "submitted_at": datetime.utcnow().isoformat(),
                "approved_at": datetime.utcnow().isoformat(),
                "updated_at": datetime.utcnow().isoformat(),
                "auto_approved_by_leader": True,
            }},
        )
        await log_action(current, "bom_auto_approve_by_leader", "bom", bom_id, {"bom_no": bom.get("bom_no")})
        return {"success": True, "engineering_status": "approved", "auto_approved": True}

    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {
            "engineering_status": "pending_review",
            "signatures": sigs,
            "submitted_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }},
    )
    await log_action(current, "bom_submit_review", "bom", bom_id, {"bom_no": bom.get("bom_no")})
    return {"success": True, "engineering_status": "pending_review"}


@router.post("/{bom_id}/approve-review")
async def bom_approve_review(bom_id: str, current: dict = Depends(get_current_user)):
    """Engineering Leader (eng_leader / Riski) approve BOM → auto-register ke BOM Utama.
    Status: pending_review → approved. Stamp Checked By signature.
    """
    # Only eng_leader (canonical), eng_head (legacy), admin-like, or super admin can approve
    role = (current or {}).get("role")
    if role not in ("eng_leader", "eng_head", "engineering", "admin", "super_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Hanya Engineering Leader / Admin yang bisa approve BOM")

    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    if bom.get("engineering_status") != "pending_review":
        raise HTTPException(status_code=409, detail=f"BOM harus di status pending_review (sekarang: {bom.get('engineering_status')})")

    sigs = dict(bom.get("signatures") or {})
    sigs["checked_by"] = _sig_stamp(current)
    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {
            "engineering_status": "approved",
            "signatures": sigs,
            "approved_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }},
    )
    await log_action(current, "bom_approve_review", "bom", bom_id, {"bom_no": bom.get("bom_no")})
    return {"success": True, "engineering_status": "approved"}


@router.post("/{bom_id}/reject-review")
async def bom_reject_review(bom_id: str, payload: dict = None, current: dict = Depends(get_current_user)):
    """Engineering Leader reject BOM → kembali ke draft dengan alasan.
    Status: pending_review → draft.
    """
    role = (current or {}).get("role")
    if role not in ("eng_leader", "eng_head", "engineering", "admin", "super_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Hanya Engineering Leader / Admin yang bisa reject BOM")

    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    if bom.get("engineering_status") != "pending_review":
        raise HTTPException(status_code=409, detail=f"BOM harus di status pending_review (sekarang: {bom.get('engineering_status')})")

    reason = ""
    attach_ids = []
    if isinstance(payload, dict):
        reason = str(payload.get("reason") or "").strip()
        attach_ids = payload.get("attachment_ids") or []
        if not isinstance(attach_ids, list):
            attach_ids = []
    if not reason:
        raise HTTPException(status_code=400, detail="Alasan reject wajib diisi")

    # Auto-append reject as a revision note (so engineer sees consolidated review history)
    note = {
        "id": str(uuid.uuid4()),
        "kind": "reject",
        "by": current.get("name") or current.get("username"),
        "user_id": current.get("id"),
        "role": current.get("role"),
        "comment": reason,
        "attachment_ids": [str(a) for a in attach_ids],
        "at": datetime.utcnow().isoformat(),
    }
    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {
            "engineering_status": "draft",
            "review_rejection_reason": reason,
            "review_rejected_by": current.get("name") or current.get("username"),
            "review_rejected_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }, "$push": {"revision_notes": note}},
    )
    await log_action(current, "bom_reject_review", "bom", bom_id, {"reason": reason, "attach_count": len(attach_ids)})
    return {"success": True, "engineering_status": "draft", "reason": reason, "note": note}


@router.post("/{bom_id}/revision-note")
async def bom_add_revision_note(bom_id: str, payload: dict = None, current: dict = Depends(get_current_user)):
    """Engineering Leader tambah revision note (komentar + optional attachment ids).
    Bisa dilakukan saat pending_review (utk feedback sebelum approve/reject) atau saat reject.
    Payload: {"comment": "...", "attachment_ids": ["..."]}.
    """
    role = (current or {}).get("role")
    if role not in ("eng_leader", "eng_head", "engineering", "admin", "super_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Hanya Engineering Leader / Admin yang bisa tambah revision note")

    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")

    comment = ""
    attach_ids = []
    if isinstance(payload, dict):
        comment = str(payload.get("comment") or "").strip()
        attach_ids = payload.get("attachment_ids") or []
        if not isinstance(attach_ids, list):
            attach_ids = []
    if not comment and not attach_ids:
        raise HTTPException(status_code=400, detail="Isi komentar atau minimal 1 attachment revisi")

    note = {
        "id": str(uuid.uuid4()),
        "by": current.get("name") or current.get("username"),
        "user_id": current.get("id"),
        "role": current.get("role"),
        "comment": comment,
        "attachment_ids": [str(a) for a in attach_ids],
        "at": datetime.utcnow().isoformat(),
    }
    await db.boms.update_one(
        {"id": bom_id},
        {"$push": {"revision_notes": note}, "$set": {"updated_at": datetime.utcnow().isoformat()}},
    )
    await log_action(current, "bom_revision_note", "bom", bom_id, {"comment_preview": comment[:60], "attach_count": len(attach_ids)})
    return {"success": True, "note": note}



async def bom_sign_stage(bom_id: str, payload: dict = None, current: dict = Depends(get_current_user)):
    """Tanda tangan tahap Acknowledge (Purchasing) atau Approved (Admin) SETELAH BOM approved.
    Payload: {"stage": "acknowledged_by" | "approved_by"}.
    """
    stage = ""
    if isinstance(payload, dict):
        stage = str(payload.get("stage") or "").strip()
    if stage not in ("acknowledged_by", "approved_by"):
        raise HTTPException(status_code=400, detail="stage harus 'acknowledged_by' atau 'approved_by'")

    role = (current or {}).get("role")
    if stage == "acknowledged_by" and role not in ("purchasing", "staff", "admin", "super_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Hanya Purchasing / Admin yang bisa acknowledge")
    if stage == "approved_by" and role not in ("admin", "super_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Hanya Admin yang bisa final approve")

    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    if bom.get("engineering_status") != "approved":
        raise HTTPException(status_code=409, detail="BOM harus sudah di-approve Engineering dulu")

    sigs = dict(bom.get("signatures") or {})
    sigs[stage] = _sig_stamp(current)
    await db.boms.update_one({"id": bom_id}, {"$set": {"signatures": sigs, "updated_at": datetime.utcnow().isoformat()}})
    await log_action(current, "bom_sign", "bom", bom_id, {"stage": stage})
    return {"success": True, "signatures": sigs}


@router.post("/upload")
async def upload_bom(
    file: UploadFile = File(...),
    prepared_by: str = Form(...),
    revision_reason: str = Form(""),
    current: dict = Depends(require_bom_upload),
):
    """Upload a BOM Excel. If SO_NO already exists, auto-creates next revision.
    `prepared_by` is REQUIRED — since Engineering shares one login for 7 people,
    the actual creator name is captured here for audit history.
    `revision_reason` is required for revisions beyond the first."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Nama file tidak ada")
    if not prepared_by or not prepared_by.strip():
        raise HTTPException(status_code=400, detail="Nama Pembuat BOM wajib diisi")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File terlalu besar (max 10 MB)")

    try:
        rows = _read_workbook(content, file.filename)
        parsed = _parse_bom_workbook(rows)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal parse file: {e}")

    so_no = normalize_so_no(parsed["header"].get("so_no"))
    if not so_no:
        raise HTTPException(status_code=400, detail="Nomor SO tidak ditemukan di file. Pastikan cell 'PT MKS SO.NO.' terisi.")

    if len(parsed["items"]) == 0:
        raise HTTPException(status_code=400, detail="Tidak ada item terbaca. Periksa format file.")

    # Determine revision
    latest = await db.boms.find_one({"so_no": so_no}, sort=[("rev_no", -1)])
    next_rev = 0 if not latest else int(latest.get("rev_no", 0)) + 1

    if next_rev > 0 and not revision_reason.strip():
        # SO already exists → prompt user for revision reason (frontend catches 409 and shows inline input)
        raise HTTPException(
            status_code=409,
            detail={
                "code": "revision_reason_required",
                "so_no": so_no,
                "latest_rev": int(latest.get("rev_no", 0)),
                "latest_uploaded_by": latest.get("uploaded_by_name") or "",
                "latest_uploaded_at": latest.get("uploaded_at", "")[:19].replace("T", " "),
                "latest_prepared_by": latest.get("prepared_by") or "",
                "message": f"Nomor SO {so_no} sudah ada di database (Rev.{latest.get('rev_no')} diupload oleh {latest.get('uploaded_by_name')} pada {latest.get('uploaded_at','')[:10]}). Silakan isi alasan revisi untuk melanjutkan.",
            },
        )

    # BOM number: pakai dari Excel jika ada, kalau kosong → auto-generate berurutan
    parsed_bom_no = (parsed["header"].get("bom_no") or "").strip()
    if not parsed_bom_no:
        auto_info = await _next_bom_no()
        parsed_bom_no = auto_info["bom_no"]
        bom_auto_gen = True
    else:
        bom_auto_gen = False

    doc = {
        "id": str(uuid.uuid4()),
        "so_no": so_no,
        "rev_no": next_rev,
        "bom_no": parsed_bom_no,
        "bom_no_auto": bom_auto_gen,
        "project_name": parsed["header"].get("project_name") or "",
        "project_dwg": parsed["header"].get("project_dwg") or "",
        "customer": parsed["header"].get("customer") or "",
        "class_material": parsed["header"].get("class_material") or "",
        "delivery_date": parsed["header"].get("delivery_date") or "",
        "bom_date": parsed["header"].get("date") or "",
        "prepared_by": prepared_by.strip(),
        "items": parsed["items"],
        "annotations": {},  # keyed by str(item_no) → {available_stock, qty_purchase, purchase_due_date, admin_remark}
        "revision_reason": revision_reason.strip(),
        "uploaded_by_id": current.get("id"),
        "uploaded_by_name": current.get("name") or current.get("username"),
        "uploaded_by_role": current.get("role"),
        "uploaded_at": datetime.utcnow().isoformat(),
        "original_filename": file.filename,
        # Excel-uploaded BOM is considered pre-approved (legacy path bypasses Engineering Review)
        "engineering_status": "approved",
        "signatures": {
            "prepared_by": {"name": prepared_by.strip(), "user_id": current.get("id"), "at": datetime.utcnow().isoformat()},
            "checked_by": None,
            "acknowledged_by": None,
            "approved_by": None,
        },
    }
    await db.boms.insert_one(doc)
    await log_action(current, "upload_bom", "bom", doc["id"], {"so_no": so_no, "rev_no": next_rev})

    # Auto-create Master SO entry if it doesn't exist yet (idempotent).
    try:
        existing_so = await db.sales_orders.find_one({"so_no": so_no, "deleted_at": {"$exists": False}})
        if not existing_so:
            so_doc = {
                "id": str(uuid.uuid4()),
                "so_no": so_no,
                "so_date": parsed["header"].get("date") or datetime.utcnow().date().isoformat(),
                "customer": parsed["header"].get("customer") or "",
                "description": parsed["header"].get("project_name") or "",
                "created_by": current.get("id"),
                "created_by_username": current.get("username", ""),
                "created_at": datetime.utcnow().isoformat(),
                "source": "bom_upload",
            }
            await db.sales_orders.insert_one(so_doc)
            await log_action(current, "auto_create_so_from_bom", "sales_order", so_doc["id"],
                             {"so_no": so_no, "bom_id": doc["id"]})
    except Exception:
        # Best-effort — never block BOM upload if SO seeding fails
        pass

    # Track the raw imported file in temp storage (for admin cleanup)
    try:
        from routers.storage import record_temp_upload
        await record_temp_upload(current, "bom_import", file.filename, content,
                                  mime=file.content_type or "",
                                  related_entity=f"BOM {so_no} Rev.{next_rev}",
                                  note=f"prepared_by={prepared_by}")
    except Exception:
        pass

    doc.pop("_id", None)
    return {"success": True, "bom": doc, "message": f"BOM tersimpan sebagai Rev.{next_rev}"}


@router.get("")
async def list_or_search_bom(
    so_no: Optional[str] = None,
    q: Optional[str] = None,
    rev: str = "latest",  # 'latest' | 'all'
    engineering_status: Optional[str] = None,  # 'approved' | 'draft' | 'pending_review' | 'all' | 'active' (draft+pending)
    limit: int = 200,
    current: dict = Depends(get_current_user),
):
    """List BOMs. Filters:
      - `so_no`: exact match on SO (backward compat)
      - `q`: fuzzy substring search across so_no / customer / project_name (case-insensitive)
      - `engineering_status`: filter by approval status. Default = 'approved' (main BOM list).
        Use 'active' to see draft+pending_review (Engineering Master List work-in-progress).
        Use 'all' to see everything.
    rev='latest' returns only newest rev per SO. rev='all' returns every revision."""
    import re
    filt: dict = {}
    if so_no:
        filt["so_no"] = normalize_so_no(so_no)
    elif q and q.strip():
        pattern = re.escape(q.strip())
        rx = {"$regex": pattern, "$options": "i"}
        filt["$or"] = [{"so_no": rx}, {"customer": rx}, {"project_name": rx}]

    # Engineering approval filter — default = 'approved' (only registered BOMs in main list)
    # Legacy BOMs without engineering_status field are treated as 'approved' (backward compat)
    if engineering_status is None or engineering_status == "approved":
        filt["$and"] = filt.get("$and", []) + [{
            "$or": [
                {"engineering_status": {"$exists": False}},
                {"engineering_status": "approved"},
            ]
        }]
    elif engineering_status == "active":
        filt["engineering_status"] = {"$in": ["draft", "pending_review"]}
    elif engineering_status == "all":
        pass  # no filter
    else:
        filt["engineering_status"] = engineering_status

    if rev == "latest":
        # Aggregation: group by so_no, take max rev
        pipeline = [{"$match": merged(filt, NOT_DELETED_FILTER)}]
        pipeline.extend([
            {"$sort": {"so_no": 1, "rev_no": -1}},
            {"$group": {"_id": "$so_no", "doc": {"$first": "$$ROOT"}}},
            {"$replaceRoot": {"newRoot": "$doc"}},
            {"$sort": {"uploaded_at": -1}},
            {"$limit": limit},
        ])
        docs = await db.boms.aggregate(pipeline).to_list(length=limit)
    else:
        docs = await db.boms.find(merged(filt, NOT_DELETED_FILTER)).sort([("so_no", 1), ("rev_no", -1)]).limit(limit).to_list(length=limit)

    for d in docs:
        d.pop("_id", None)
        _recompute_bom_purchase_status(d)
    return {"items": docs, "total": len(docs)}


@router.get("/history/{so_no}")
async def bom_history(so_no: str, current: dict = Depends(get_current_user)):
    """Return every revision for a given SO, newest first."""
    docs = await db.boms.find(merged({"so_no": normalize_so_no(so_no)}, NOT_DELETED_FILTER)).sort("rev_no", -1).to_list(length=200)
    for d in docs:
        d.pop("_id", None)
    return {"so_no": so_no, "count": len(docs), "revisions": docs}


@router.get("/{bom_id}")
async def get_bom(bom_id: str, current: dict = Depends(get_current_user)):
    doc = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    # Auto-heal on read: legacy BOM without engineering_status + 0 items → mark as draft
    if not doc.get("engineering_status") and not (doc.get("items") or []):
        await db.boms.update_one({"id": bom_id}, {"$set": {"engineering_status": "draft"}})
        doc["engineering_status"] = "draft"
    return doc


@router.post("/{bom_id}/request-reopen")
async def request_reopen_bom(bom_id: str, payload: dict, current: dict = Depends(get_current_user)):
    """Engineer/Staff request izin edit ulang BOM yg sudah approved.
    Payload: {reason: 'alasan perlu edit'}. Status BOM tidak berubah; hanya buat request record.
    Riski (eng_leader) approve request → panggil /reopen → status ke draft dan buat revision entry.
    """
    reason = ""
    if isinstance(payload, dict):
        reason = str(payload.get("reason") or "").strip()
    if not reason or len(reason) < 8:
        raise HTTPException(status_code=400, detail="Alasan wajib diisi (minimal 8 karakter)")

    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    if bom.get("engineering_status") != "approved":
        raise HTTPException(status_code=409, detail=f"BOM harus di status approved dulu (sekarang: {bom.get('engineering_status')})")

    # Cek apakah sudah ada pending request yg belum di-respons
    existing = await db.bom_reopen_requests.find_one({
        "bom_id": bom_id, "status": "pending", "deleted_at": {"$exists": False}
    })
    if existing:
        raise HTTPException(status_code=409, detail="Sudah ada permintaan reopen pending untuk BOM ini — tunggu response Engineering Leader")

    req = {
        "id": str(uuid.uuid4()),
        "bom_id": bom_id,
        "bom_no": bom.get("bom_no"),
        "so_no": bom.get("so_no"),
        "reason": reason,
        "status": "pending",
        "requested_by_id": current.get("id"),
        "requested_by_name": current.get("name") or current.get("username"),
        "requested_by_role": current.get("role"),
        "created_at": datetime.utcnow().isoformat(),
    }
    await db.bom_reopen_requests.insert_one(req.copy())
    await log_action(current, "bom_reopen_request", "bom", bom_id, {"reason": reason[:60]})
    req.pop("_id", None)
    return {"success": True, "request": req}


@router.get("/_/reopen-requests")
async def list_reopen_requests(status: str = "pending", current: dict = Depends(get_current_user)):
    """List BOM reopen requests. Eng leader/admin sees all pending; engineer sees own."""
    filt = {"deleted_at": {"$exists": False}}
    if status and status != "all":
        filt["status"] = status
    role = (current or {}).get("role")
    if role not in ("eng_leader", "eng_head", "admin", "super_admin", "supervisor"):
        filt["requested_by_id"] = current.get("id")
    docs = await db.bom_reopen_requests.find(filt, {"_id": 0}).sort("created_at", -1).to_list(length=100)
    return {"items": docs, "total": len(docs)}


@router.post("/_/reopen-requests/{req_id}/approve")
async def approve_reopen_request(req_id: str, current: dict = Depends(get_current_user)):
    """Engineering Leader / Admin approve reopen request → BOM back to draft + revision entry."""
    role = (current or {}).get("role")
    if role not in ("eng_leader", "eng_head", "admin", "super_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Hanya Engineering Leader / Admin yang bisa approve")

    req = await db.bom_reopen_requests.find_one({"id": req_id, "deleted_at": {"$exists": False}})
    if not req:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    if req.get("status") != "pending":
        raise HTTPException(status_code=409, detail=f"Request sudah di status {req.get('status')}")

    bom_id = req["bom_id"]
    bom = await db.boms.find_one({"id": bom_id})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM sudah tidak ada")

    # Buat revision entry — snapshot BOM state SEBELUM di-reopen
    current_rev_no = int(bom.get("current_rev_no") or 0) + 1
    revision_entry = {
        "id": str(uuid.uuid4()),
        "rev_no": current_rev_no,
        "reason": req.get("reason") or "",
        "requested_by": req.get("requested_by_name"),
        "approved_by": current.get("name") or current.get("username"),
        "reopened_at": datetime.utcnow().isoformat(),
        "items_before": bom.get("items") or [],
        "signatures_before": bom.get("signatures") or {},
    }

    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {
            "engineering_status": "draft",
            "current_rev_no": current_rev_no,
            "reopened_at": datetime.utcnow().isoformat(),
            "reopened_by": current.get("name"),
            "updated_at": datetime.utcnow().isoformat(),
        }, "$push": {"revisions": revision_entry}},
    )

    await db.bom_reopen_requests.update_one(
        {"id": req_id},
        {"$set": {
            "status": "approved",
            "approved_by": current.get("name") or current.get("username"),
            "approved_at": datetime.utcnow().isoformat(),
        }},
    )
    await log_action(current, "bom_reopen_approve", "bom", bom_id, {"rev_no": current_rev_no, "reason": (req.get("reason") or "")[:60]})
    return {"success": True, "engineering_status": "draft", "rev_no": current_rev_no}


@router.post("/_/reopen-requests/{req_id}/reject")
async def reject_reopen_request(req_id: str, payload: dict = None, current: dict = Depends(get_current_user)):
    role = (current or {}).get("role")
    if role not in ("eng_leader", "eng_head", "admin", "super_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Hanya Engineering Leader / Admin yang bisa reject")
    reason = ""
    if isinstance(payload, dict):
        reason = str(payload.get("reason") or "").strip()
    await db.bom_reopen_requests.update_one(
        {"id": req_id},
        {"$set": {
            "status": "rejected",
            "rejected_by": current.get("name") or current.get("username"),
            "rejected_at": datetime.utcnow().isoformat(),
            "rejection_reason": reason,
        }},
    )
    return {"success": True}


@router.patch("/{bom_id}/meta")
async def update_bom_meta(bom_id: str, payload: dict, current: dict = Depends(require_bom_edit)):
    """Update BOM metadata (class_material, bom_date, delivery_date, project_name, customer, so_no).
    Hanya diizinkan saat BOM status draft atau pending_review.
    """
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    status_now = bom.get("engineering_status", "approved")
    if status_now not in ("draft", "pending_review"):
        raise HTTPException(status_code=409, detail=f"BOM sudah di-approve (status={status_now}) — meta tidak bisa diubah")

    editable_fields = ["class_material", "bom_date", "delivery_date", "project_name", "customer", "so_no"]
    update_set = {}
    for k in editable_fields:
        if k in payload:
            v = payload[k]
            update_set[k] = (str(v).strip() if v is not None else "")
    if not update_set:
        raise HTTPException(status_code=400, detail="Tidak ada field yang di-update")

    update_set["updated_at"] = datetime.utcnow().isoformat()
    update_set["last_edited_by"] = current.get("name") or current.get("username")

    await db.boms.update_one({"id": bom_id}, {"$set": update_set})
    await log_action(current, "bom_meta_update", "bom", bom_id, {"fields": list(update_set.keys())})
    return {"success": True, "updated": update_set}


@router.patch("/{bom_id}/annotations")
async def update_bom_annotations(
    bom_id: str,
    payload: BOMAnnotationsUpdate,
    current: dict = Depends(require_bom_admin),
):
    """Admin fills Available Stock / Qty Purchase / Purchase Due Date / Remark per item.
    Sends full list; server replaces annotations map."""
    bom = await db.boms.find_one({"id": bom_id})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")

    # Build item_no → bom qty lookup for auto-calc qty_purchase
    bom_items = bom.get("items") or []
    qty_by_no = {int(it.get("item_no", -1)): float(it.get("qty") or 0) for it in bom_items}

    ann_map = {}
    for a in payload.annotations:
        # Auto-compute qty_purchase = qty_bom - available_stock (kalau user tidak override)
        bom_qty = qty_by_no.get(int(a.item_no), 0)
        avail = float(a.available_stock or 0)
        auto_qty_purchase = max(0.0, bom_qty - avail)
        # Kalau user submit qty_purchase eksplisit (non-null), pakai itu; else auto
        qty_pur = a.qty_purchase if a.qty_purchase is not None else auto_qty_purchase
        ann_map[str(a.item_no)] = {
            "available_stock": a.available_stock,
            "qty_purchase": qty_pur,
            "purchase_due_date": a.purchase_due_date,
            "admin_remark": a.admin_remark,
            "updated_at": datetime.utcnow().isoformat(),
            "updated_by": current.get("name") or current.get("username"),
        }
    await db.boms.update_one({"id": bom_id}, {"$set": {"annotations": ann_map}})
    await log_action(current, "annotate_bom", "bom", bom_id, {"count": len(ann_map)})
    return {"success": True, "annotations": ann_map}


@router.delete("/{bom_id}")
async def delete_bom(bom_id: str, current: dict = Depends(require_bom_admin)):
    """Admin can delete a specific BOM revision (rarely used, e.g. mistake upload)."""
    ok = await soft_delete_one("boms", {"id": bom_id}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    await log_action(current, "delete_bom", "bom", bom_id, {})
    return {"success": True}


# ---------------------- Purchase History for a BOM's SO ----------------------
@router.post("/{bom_id}/dismiss-purchase-notif")
async def dismiss_bom_purchase_notif(bom_id: str, current: dict = Depends(get_current_user)):
    """Mark a BOM's purchase-notification as dismissed / acknowledged.
    Once dismissed, the notification 'BOM Baru — Butuh Pembelian' will not surface it again."""
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {"purchase_notif_dismissed": True,
                  "purchase_notif_dismissed_at": datetime.utcnow().isoformat(),
                  "purchase_notif_dismissed_by": current.get("username", "")}},
    )
    await log_action(current, "dismiss_bom_purchase_notif", "bom", bom_id, {})
    return {"ok": True, "bom_id": bom_id}


@router.get("/{bom_id}/purchases")
async def bom_purchase_history(bom_id: str, current: dict = Depends(get_current_user)):
    """Cross-reference: list all purchase transactions that reference this BOM's SO number.

    Role rules:
    - super_admin, admin, supervisor, finance : full access (including price)
    - eng_head, eng_staff, engineering, sales : read-only, price visible
    - store : read-only, price HIDDEN (unit_price, total_price, total_price_idr, currency, exchange_rate, totals_by_currency)
    - purchasing : read-only, price visible
    """
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    so_no = (bom.get("so_no") or "").strip()
    role = (current.get("role") or "").lower()

    # Role gate — block roles we don't intend to give any access
    allowed_roles = {
        "super_admin", "admin", "supervisor", "finance",
        "eng_leader", "eng_head", "eng_staff", "engineering",
        "sales", "purchasing", "store",
    }
    if role not in allowed_roles and not current.get("is_super_admin"):
        raise HTTPException(status_code=403, detail="Tidak punya akses ke data pembelian BOM ini")

    if not so_no and bom_id:
        # BOM without SO — only explicit BOM-linked purchases
        pass  # continue with filter below

    # Transactions can match this BOM in 2 ways:
    #   (a) project_no == so_no (loose, by SO number)
    #   (b) bom_item_ref.bom_id == bom_id (explicit link — walau tanpa SO)
    filt = {
        "deleted_at": {"$exists": False},
        "$or": [
            {"project_no": {"$regex": f"^{re.escape(so_no)}$", "$options": "i"}} if so_no else {"_id": None},
            {"bom_item_ref.bom_id": bom_id},
        ],
    }
    cur = db.transactions.find(filt, {"_id": 0}).sort("invoice_date", -1).limit(500)
    docs = await cur.to_list(length=500)

    # Aggregate simple totals per currency
    totals: dict = {}
    for d in docs:
        cur_code = d.get("currency") or "IDR"
        amt = float(d.get("total_price") or 0)
        totals[cur_code] = totals.get(cur_code, 0.0) + amt

    price_hidden = role == "store"
    if price_hidden:
        redacted = []
        for d in docs:
            d2 = {k: v for k, v in d.items()
                  if k not in ("unit_price", "total_price", "total_price_idr", "currency", "exchange_rate")}
            redacted.append(d2)
        docs = redacted
        totals = {}

    return {
        "so_no": so_no,
        "count": len(docs),
        "totals_by_currency": totals,
        "items": docs,
        "price_hidden": price_hidden,
    }



# =============================================================================
# PURCHASE TRACKING — pola sama seperti Consumable Request
# =============================================================================
def _now_iso():
    return datetime.utcnow().isoformat()


def _is_purchasing_role(user: dict) -> bool:
    role = (user or {}).get("role", "").lower()
    return role in ("purchasing", "staff", "admin", "super_admin", "supervisor") or user.get("is_super_admin")


def _clean_bom(doc: dict) -> dict:
    if not doc:
        return doc
    doc.pop("_id", None)
    return doc


def _recompute_bom_purchase_status(bom: dict) -> dict:
    """Compute per-item + overall purchase progress. Considers available_stock annotation:
    - If available_stock >= qty → item marked 'in_stock' (no need to buy)
    - Else needed_qty = qty - available_stock; purchases[] fills against needed_qty
    Modifies bom in-place and returns it."""
    items = bom.get("items") or []
    annotations = bom.get("annotations") or {}
    # Support both list and dict annotation storage
    ann_by_no = {}
    if isinstance(annotations, list):
        for a in annotations:
            ann_by_no[int(a.get("item_no", -1))] = a
    elif isinstance(annotations, dict):
        for k, v in annotations.items():
            try:
                ann_by_no[int(k)] = v if isinstance(v, dict) else {}
            except Exception:
                pass
    fulfilled = 0
    partial = 0
    in_stock = 0
    for it in items:
        purchases = it.get("purchases") or []
        total_bought = sum(float(p.get("qty_bought") or 0) for p in purchases)
        qty = float(it.get("qty") or 0)
        ann = ann_by_no.get(int(it.get("item_no", -1)), {})
        try:
            avail_stock = float(ann.get("available_stock") or 0)
        except (TypeError, ValueError):
            avail_stock = 0
        needed_qty = max(0.0, qty - avail_stock)
        it["total_bought"] = total_bought
        it["needed_qty"] = needed_qty
        it["available_stock_computed"] = avail_stock
        # 1) Stock cukup → tidak perlu beli
        if avail_stock >= qty and qty > 0:
            it["purchase_status"] = "in_stock"
            it["purchased"] = True  # dianggap terpenuhi (via stok)
            in_stock += 1
            continue
        # 2) Butuh beli — evaluasi purchases[] vs needed_qty
        if needed_qty <= 0:
            it["purchase_status"] = "in_stock"
            it["purchased"] = True
            in_stock += 1
        elif total_bought <= 0:
            it["purchase_status"] = "pending"
            it["purchased"] = False
        elif total_bought < needed_qty:
            it["purchase_status"] = "partial"
            it["purchased"] = False
            partial += 1
        elif total_bought >= needed_qty and total_bought <= needed_qty * 1.001:
            it["purchase_status"] = "fulfilled"
            it["purchased"] = True
            fulfilled += 1
        else:
            it["purchase_status"] = "over"
            it["purchased"] = True
            fulfilled += 1
    bom["items"] = items
    total = len(items)
    covered = fulfilled + in_stock
    bom["purchase_progress"] = {
        "total_items": total,
        "fulfilled": fulfilled,
        "in_stock": in_stock,
        "partial": partial,
        "pending": total - covered - partial,
        "percent": round((covered / total * 100) if total else 0, 1),
    }
    return bom


@router.get("/{bom_id}/purchase-status")
async def bom_purchase_status(bom_id: str, current: dict = Depends(get_current_user)):
    """Detailed per-item purchase status + overall progress."""
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    _recompute_bom_purchase_status(bom)
    return {
        "bom_id": bom_id,
        "so_no": bom.get("so_no"),
        "progress": bom.get("purchase_progress"),
        "items": [{
            "item_no": it.get("item_no"),
            "item_name": it.get("item_name"),
            "qty": it.get("qty"),
            "uom": it.get("uom"),
            "available_stock": it.get("available_stock_computed", 0),
            "needed_qty": it.get("needed_qty", it.get("qty", 0)),
            "total_bought": it.get("total_bought", 0),
            "purchase_status": it.get("purchase_status", "pending"),
            "purchases": it.get("purchases") or [],
        } for it in (bom.get("items") or [])],
    }


@router.get("/purchase/open-items")
async def bom_open_items(
    q: Optional[str] = None,
    limit: int = 200,
    current: dict = Depends(get_current_user),
):
    """Flat list of BOM items still needing purchase (across all BOMs) — for 'Tarik Item dari BOM' picker."""
    filt = merged({}, NOT_DELETED_FILTER)
    boms = await db.boms.find(filt, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    out = []
    q_lower = (q or "").strip().lower()
    for b in boms:
        _recompute_bom_purchase_status(b)
        for it in (b.get("items") or []):
            if it.get("purchased"):
                continue
            if it.get("purchase_status") == "in_stock":
                continue  # Stok sudah cukup, tidak perlu beli
            needed = float(it.get("needed_qty") or it.get("qty") or 0)
            remaining = needed - float(it.get("total_bought") or 0)
            if remaining <= 0:
                continue
            row = {
                "bom_id": b.get("id"),
                "so_no": b.get("so_no"),
                "so_customer_name": b.get("customer_name"),
                "revision": b.get("revision"),
                "item_no": it.get("item_no"),
                "item_name": it.get("item_name"),
                "item_specification": it.get("item_specification", ""),
                "qty": it.get("qty"),
                "unit": it.get("uom") or "",
                "material": it.get("material", ""),
                "remark": it.get("remark", ""),
                "total_bought": it.get("total_bought", 0),
                "remaining": remaining,
                "purchase_status": it.get("purchase_status", "pending"),
            }
            if q_lower:
                blob = " ".join(str(v or "").lower() for v in row.values())
                if q_lower not in blob:
                    continue
            out.append(row)
    return out[:limit]


@router.get("/purchase/search-transactions")
async def bom_search_transactions(
    q: Optional[str] = None,
    days: int = 90,
    current: dict = Depends(get_current_user),
):
    """Search existing transactions to link with a BOM item (retroactive)."""
    from datetime import timedelta
    cutoff = (datetime.utcnow() - timedelta(days=int(days))).strftime("%Y-%m-%d")
    filt: dict = {"invoice_date": {"$gte": cutoff}}
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [{"item_name": rx}, {"vendor_name": rx}, {"invoice_no": rx}, {"po_no": rx}, {"project_no": rx}]
    docs = await db.transactions.find(merged(filt, NOT_DELETED_FILTER), {"_id": 0}).sort("invoice_date", -1).to_list(length=200)
    return [
        {
            "id": d.get("id"),
            "invoice_date": d.get("invoice_date"),
            "vendor_name": d.get("vendor_name"),
            "item_name": d.get("item_name"),
            "qty": d.get("qty"),
            "unit": d.get("unit"),
            "unit_price": d.get("unit_price"),
            "po_no": d.get("po_no") or "",
            "invoice_no": d.get("invoice_no") or "",
            "project_no": d.get("project_no") or "",
            "already_linked_bom": bool(d.get("bom_item_ref")),
            "already_linked_cgr": bool(d.get("consumable_request_item_id")),
        }
        for d in docs
    ]


@router.post("/{bom_id}/items/{item_no}/mark-purchased")
async def bom_mark_item_purchased(
    bom_id: str, item_no: int, payload: dict,
    current: dict = Depends(get_current_user),
):
    """Purchasing mark that a BOM item has been purchased.
    Payload: {actual_item_name, vendor_name, qty_bought, unit, purchase_date, po_no, invoice_no, unit_price, transaction_id?, source, note}
    """
    if not _is_purchasing_role(current):
        raise HTTPException(status_code=403, detail="Hanya Purchasing/Admin yang bisa mark-purchased")
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    items = bom.get("items") or []
    target = None
    for it in items:
        if int(it.get("item_no", -1)) == int(item_no):
            target = it
            break
    if not target:
        raise HTTPException(status_code=404, detail=f"Item no.{item_no} tidak ditemukan")

    purchase_entry = {
        "purchased_at": _now_iso(),
        "purchased_by": current.get("username") or current.get("name", ""),
        "actual_item_name": (payload.get("actual_item_name") or target.get("item_name") or "").strip(),
        "vendor_name": (payload.get("vendor_name") or "").strip(),
        "qty_bought": float(payload.get("qty_bought") or 0),
        "unit": (payload.get("unit") or target.get("uom") or "").strip(),
        "purchase_date": payload.get("purchase_date") or _now_iso()[:10],
        "po_no": (payload.get("po_no") or "").strip(),
        "invoice_no": (payload.get("invoice_no") or "").strip(),
        "unit_price": float(payload.get("unit_price") or 0),
        "transaction_id": payload.get("transaction_id"),
        "source": payload.get("source") or "manual",
        "note": (payload.get("note") or "").strip(),
    }
    target.setdefault("purchases", []).append(purchase_entry)
    _recompute_bom_purchase_status(bom)
    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {"items": bom["items"], "purchase_progress": bom["purchase_progress"], "updated_at": _now_iso()}},
    )
    await log_action(current, "bom_mark_item_purchased", "bom", bom_id, {
        "item_no": item_no, "qty_bought": purchase_entry["qty_bought"], "vendor": purchase_entry["vendor_name"],
    })
    fresh = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    return _clean_bom(fresh)


@router.post("/{bom_id}/items/{item_no}/link-transaction")
async def bom_link_transaction(
    bom_id: str, item_no: int, payload: dict,
    current: dict = Depends(get_current_user),
):
    """Retroactively link an existing transaction to a BOM item.
    Payload: {transaction_id: str}"""
    if not _is_purchasing_role(current):
        raise HTTPException(status_code=403, detail="Hanya Purchasing/Admin")
    tx_id = payload.get("transaction_id")
    if not tx_id:
        raise HTTPException(status_code=400, detail="transaction_id wajib")
    tx = await db.transactions.find_one({"id": tx_id, "deleted_at": {"$exists": False}})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    if tx.get("bom_item_ref"):
        raise HTTPException(status_code=400, detail="Transaksi sudah ter-link ke BOM item lain")

    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    items = bom.get("items") or []
    target = next((it for it in items if int(it.get("item_no", -1)) == int(item_no)), None)
    if not target:
        raise HTTPException(status_code=404, detail=f"Item no.{item_no} tidak ditemukan")

    await db.transactions.update_one(
        {"id": tx_id},
        {"$set": {
            "bom_item_ref": {"bom_id": bom_id, "item_no": int(item_no)},
            "updated_at": _now_iso(),
        }},
    )

    purchase_entry = {
        "purchased_at": _now_iso(),
        "purchased_by": current.get("username") or current.get("name", ""),
        "actual_item_name": tx.get("item_name") or target.get("item_name") or "",
        "vendor_name": tx.get("vendor_name") or "",
        "qty_bought": float(tx.get("qty") or 0),
        "unit": tx.get("unit") or target.get("uom") or "",
        "purchase_date": tx.get("invoice_date") or _now_iso()[:10],
        "po_no": tx.get("po_no") or "",
        "invoice_no": tx.get("invoice_no") or "",
        "unit_price": float(tx.get("unit_price") or 0),
        "transaction_id": tx_id,
        "source": "linked",
        "note": "",
    }
    target.setdefault("purchases", []).append(purchase_entry)
    _recompute_bom_purchase_status(bom)
    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {"items": bom["items"], "purchase_progress": bom["purchase_progress"], "updated_at": _now_iso()}},
    )
    await log_action(current, "bom_link_tx", "bom", bom_id, {
        "item_no": item_no, "transaction_id": tx_id, "vendor": tx.get("vendor_name"),
    })
    fresh = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    return _clean_bom(fresh)


@router.post("/{bom_id}/items/{item_no}/unmark-purchased/{purchase_index}")
async def bom_unmark_item_purchased(
    bom_id: str, item_no: int, purchase_index: int,
    current: dict = Depends(get_current_user),
):
    """Remove a purchase entry (undo mark-purchased). Also unlinks transaction if any."""
    if not _is_purchasing_role(current):
        raise HTTPException(status_code=403, detail="Hanya Purchasing/Admin")
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    items = bom.get("items") or []
    target = next((it for it in items if int(it.get("item_no", -1)) == int(item_no)), None)
    if not target:
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")
    purchases = target.get("purchases") or []
    if purchase_index < 0 or purchase_index >= len(purchases):
        raise HTTPException(status_code=404, detail="Purchase entry tidak ditemukan")
    removed = purchases.pop(purchase_index)
    # Unlink transaction if was linked
    tx_id = removed.get("transaction_id")
    if tx_id:
        await db.transactions.update_one(
            {"id": tx_id},
            {"$unset": {"bom_item_ref": ""}, "$set": {"updated_at": _now_iso()}},
        )
    target["purchases"] = purchases
    _recompute_bom_purchase_status(bom)
    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {"items": bom["items"], "purchase_progress": bom["purchase_progress"], "updated_at": _now_iso()}},
    )
    await log_action(current, "bom_unmark_item_purchased", "bom", bom_id,
                     {"item_no": item_no, "purchase_index": purchase_index, "tx_id": tx_id})
    fresh = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    return _clean_bom(fresh)


# =============================================================================
# Helper used by transactions router when creating a NEW purchase transaction
# that references a BOM item (Tarik Item dari BOM flow)
# =============================================================================
async def link_purchase_to_bom(
    bom_id: str,
    item_no: int,
    *,
    actual_item_name: str,
    vendor_name: str,
    qty_bought: float,
    unit: str,
    purchase_date: str,
    po_no: str,
    invoice_no: str = "",
    unit_price: float = 0,
    transaction_id: Optional[str] = None,
    current: dict,
):
    """Called from transactions router when a NEW tx with bom_item_ref is created."""
    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}})
    if not bom:
        return
    items = bom.get("items") or []
    target = next((it for it in items if int(it.get("item_no", -1)) == int(item_no)), None)
    if not target:
        return
    entry = {
        "purchased_at": _now_iso(),
        "purchased_by": (current or {}).get("username") or (current or {}).get("name", ""),
        "actual_item_name": actual_item_name,
        "vendor_name": vendor_name,
        "qty_bought": float(qty_bought or 0),
        "unit": unit or "",
        "purchase_date": purchase_date or _now_iso()[:10],
        "po_no": po_no or "",
        "invoice_no": invoice_no or "",
        "unit_price": float(unit_price or 0),
        "transaction_id": transaction_id,
        "source": "auto",
        "note": "",
    }
    target.setdefault("purchases", []).append(entry)
    _recompute_bom_purchase_status(bom)
    await db.boms.update_one(
        {"id": bom_id},
        {"$set": {"items": bom["items"], "purchase_progress": bom["purchase_progress"], "updated_at": _now_iso()}},
    )



def _normalize_name(s: str) -> str:
    """Normalize for fuzzy item-name matching: lowercase, collapse whitespace, strip punctuation."""
    if not s:
        return ""
    s = str(s).lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _name_match(a: str, b: str) -> bool:
    """Fuzzy match: two names match if one contains the other after normalization,
    OR share ≥ 2 significant tokens."""
    na, nb = _normalize_name(a), _normalize_name(b)
    if not na or not nb:
        return False
    if na in nb or nb in na:
        return True
    ta, tb = set(na.split()), set(nb.split())
    # remove very short tokens (single char, digits)
    ta = {t for t in ta if len(t) >= 3}
    tb = {t for t in tb if len(t) >= 3}
    return len(ta & tb) >= 2


async def auto_link_tx_to_bom_by_so(tx: dict, current: dict) -> None:
    """If a NEW transaction has project_no (SO) but no bom_item_ref set,
    scan BOMs for that SO. If any BOM item matches item_name (fuzzy), auto-link.
    Called from transactions.create_transaction / bulk after insert.
    """
    if not tx or tx.get("bom_item_ref"):
        return  # already explicit link
    so_no = (tx.get("project_no") or "").strip()
    if not so_no:
        return
    item_name = tx.get("item_name") or ""
    if not item_name.strip():
        return
    # Find the latest revision BOM for this SO
    pipeline = [
        {"$match": {"so_no": {"$regex": f"^{re.escape(so_no)}$", "$options": "i"},
                    "deleted_at": {"$exists": False}}},
        {"$sort": {"rev_no": -1}},
        {"$limit": 1},
    ]
    boms = await db.boms.aggregate(pipeline).to_list(length=1)
    if not boms:
        return
    bom = boms[0]
    bom_id = bom.get("id")
    for it in (bom.get("items") or []):
        if _name_match(item_name, it.get("item_name") or ""):
            # Match! Attach bom_item_ref back to tx + append purchase entry
            item_no = int(it.get("item_no", -1))
            await db.transactions.update_one(
                {"id": tx["id"]},
                {"$set": {"bom_item_ref": {"bom_id": bom_id, "item_no": item_no}, "updated_at": _now_iso()}},
            )
            await link_purchase_to_bom(
                bom_id, item_no,
                actual_item_name=item_name,
                vendor_name=tx.get("vendor_name") or "",
                qty_bought=float(tx.get("qty") or 0),
                unit=tx.get("unit") or it.get("uom") or "",
                purchase_date=tx.get("invoice_date") or tx.get("receive_date") or "",
                po_no=tx.get("po_no") or "",
                invoice_no=tx.get("invoice_no") or "",
                unit_price=float(tx.get("unit_price") or 0),
                transaction_id=tx["id"],
                current=current,
            )
            return  # only link to first match



# ============================================================================
# EXPORT BOM → .xlsx (matching MKS template BOM042-07-2026)
# Aktif hanya setelah BOM engineering_status == 'approved' (released)
# ============================================================================
@router.get("/{bom_id}/export/xlsx")
async def export_bom_xlsx(bom_id: str, current: dict = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    bom = await db.boms.find_one({"id": bom_id})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
    if bom.get("engineering_status") != "approved":
        raise HTTPException(
            status_code=409,
            detail=f"BOM belum dirilis (status: {bom.get('engineering_status', 'unknown')}). Export hanya bisa setelah approved/released.",
        )

    # Prepare data context
    approved_iso = bom.get("approved_at") or bom.get("updated_at") or bom.get("created_at") or ""
    approved_date = ""
    if approved_iso:
        try:
            approved_date = datetime.fromisoformat(str(approved_iso).replace("Z", "+00:00")).strftime("%d/%m/%Y")
        except Exception:
            approved_date = str(approved_iso)[:10]
    delivery_date = bom.get("delivery_date") or ""
    if delivery_date:
        try:
            delivery_date = datetime.fromisoformat(str(delivery_date).replace("Z", "+00:00")).strftime("%d/%m/%Y")
        except Exception:
            pass

    items_raw = bom.get("items") or []
    total_weight = 0.0
    items_out = []
    for i, it in enumerate(items_raw):
        w = it.get("weight_kg") or it.get("weight")
        w_val = ""
        if w not in (None, "", 0):
            try:
                total_weight += float(w)
                w_val = float(w)
            except Exception:
                w_val = str(w)
        items_out.append({
            "__index__": i + 1,
            "item_no": it.get("item_no") or (i + 1),
            "item_name": it.get("item_name") or "",
            "specification": it.get("specification") or it.get("item_specification") or "",
            "qty": it.get("qty") or "",
            "uom": it.get("uom") or it.get("unit") or "",
            "material": it.get("material") or "",
            "weight_kg": w_val,
            "available_stock": "",  # KOSONG — diisi manual saat print
            "qty_purchase": "",     # KOSONG — diisi manual saat print
            "purchase_due_date": it.get("purchase_due_date") or "",
            "remark": it.get("remark") or "",
        })

    # Signatures (Prepared By = pembuat BOM, Checked By = eng leader, dst)
    sigs = bom.get("signatures") or {}
    def _sig_name(k):
        s = sigs.get(k) or {}
        if isinstance(s, dict):
            return s.get("name") or s.get("username") or ""
        return str(s or "")

    data = {
        "company_name": "PT. MITRA KARYA SARANA",
        "bom_no": bom.get("bom_no") or "",
        "revision": bom.get("revision") or 0,
        "approved_date": approved_date,
        "project_name": bom.get("project_name") or "",
        "drawing_no": bom.get("drawing_no") or bom.get("project_dwg") or "",
        "customer": bom.get("customer") or "",
        "class_material": bom.get("class_material") or "",
        "so_no": bom.get("so_no") or "",
        "delivery_date": delivery_date,
        "notes": bom.get("notes") or "",
        "total_weight": total_weight if total_weight > 0 else "",
        "print_date": datetime.now(timezone.utc).strftime("%d/%m/%Y"),
        "printed_by": current.get("username", ""),
        # Signature placeholders — nama yang muncul di kolom tanda tangan
        "prepared_by": _sig_name("prepared_by") or bom.get("prepared_by") or bom.get("created_by_name") or "",
        "checked_by": _sig_name("checked_by"),
        "acknowledged_by": _sig_name("acknowledged_by"),
        "approved_by": _sig_name("approved_by"),
        "items": items_out,
    }

    safe_project = (bom.get("project_name") or "BOM").replace("/", "_").replace("\\", "_")[:80]
    filename = f"{bom.get('bom_no', bom_id)} - {safe_project}.xlsx"

    # Prefer USER-UPLOADED template (Admin → Template Excel → BOM)
    try:
        from routers.excel_templates import get_active_xlsx_bytes, render_excel_template
        xlsx_bytes = await get_active_xlsx_bytes("BOM")
    except Exception:
        xlsx_bytes = None

    if xlsx_bytes:
        rendered = render_excel_template(xlsx_bytes, data, as_pdf=False)
        await log_action(current, "export_bom_xlsx", "bom", bom_id,
                         {"filename": filename, "engine": "user-template"})
        return StreamingResponse(
            io.BytesIO(rendered),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ---- Fallback: built-in default layout (kalau user belum upload template) ----
    wb = Workbook()
    ws = wb.active
    ws.title = f"REV {bom.get('revision', 0)}"

    thin = Side(border_style="thin", color="000000")
    med = Side(border_style="medium", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    box_med = Border(left=med, right=med, top=med, bottom=med)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold = Font(bold=True, name="Calibri", size=10)
    title_font = Font(bold=True, name="Calibri", size=16)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    widths = [6, 18, 4, 14, 11, 6, 7, 7, 6, 7, 13, 5, 7, 13, 13, 15, 18, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=18)
    tc = ws.cell(row=2, column=1, value="BILL OF MATERIAL (BOM)")
    tc.font = title_font
    tc.alignment = center
    ws.row_dimensions[2].height = 32

    ws.cell(row=4, column=1, value="TO").font = bold
    ws.cell(row=4, column=2, value=": PUR")
    ws.cell(row=4, column=12, value="BOM.NO.").font = bold
    ws.cell(row=4, column=14, value=f": {bom.get('bom_no', '')}")

    ws.cell(row=5, column=1, value="DATE").font = bold
    ws.cell(row=5, column=2, value=f": {approved_date}")
    ws.cell(row=5, column=12, value="REV.NO.").font = bold
    ws.cell(row=5, column=14, value=f": {bom.get('revision', 0)}")

    meta_labels = [
        ("PROJECT :", 1, 3),
        ("ENG. DRW. PROJECT NO. :", 4, 6),
        ("CUSTOMER :", 7, 10),
        ("CLASS OF MATERIAL :", 11, 13),
        ("PT MKS SO.NO. :", 14, 15),
        ("DELIVERY DATE :", 16, 18),
    ]
    for label, c_start, c_end in meta_labels:
        ws.merge_cells(start_row=7, start_column=c_start, end_row=7, end_column=c_end)
        cell = ws.cell(row=7, column=c_start, value=label)
        cell.font = bold
        cell.alignment = left
        cell.border = box
        cell.fill = header_fill

    meta_values = [
        (data["project_name"], 1, 3),
        (data["drawing_no"], 4, 6),
        (data["customer"], 7, 10),
        (data["class_material"], 11, 13),
        (data["so_no"], 14, 15),
        (data["delivery_date"], 16, 18),
    ]
    for value, c_start, c_end in meta_values:
        ws.merge_cells(start_row=8, start_column=c_start, end_row=8, end_column=c_end)
        cell = ws.cell(row=8, column=c_start, value=value)
        cell.alignment = left
        cell.border = box
    ws.row_dimensions[8].height = 42

    headers = [
        ("ITEM NO", 1, 1),
        ("ITEM NAME", 2, 2),
        ("ITEM SPECIFICATION", 3, 7),
        ("QTY.", 8, 8),
        ("UOM", 9, 10),
        ("MATERIAL", 11, 11),
        ("WEIGHT (KG)", 12, 13),
        ("AVAILABLE STOCK", 14, 14),
        ("QTY PURCHASE", 15, 15),
        ("PURCHASE DUE DATE", 16, 16),
        ("REMARK", 17, 18),
    ]
    for text, c_start, c_end in headers:
        if c_start != c_end:
            ws.merge_cells(start_row=9, start_column=c_start, end_row=9, end_column=c_end)
        cell = ws.cell(row=9, column=c_start, value=text)
        cell.font = bold
        cell.alignment = center
        cell.border = box_med
        cell.fill = header_fill
    ws.row_dimensions[9].height = 30

    for i, it in enumerate(items_out):
        r = 10 + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
        ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=13)
        ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)

        ws.cell(row=r, column=1, value=it["item_no"])
        ws.cell(row=r, column=2, value=it["item_name"])
        ws.cell(row=r, column=3, value=it["specification"])
        ws.cell(row=r, column=8, value=it["qty"])
        ws.cell(row=r, column=9, value=it["uom"])
        ws.cell(row=r, column=11, value=it["material"])
        ws.cell(row=r, column=12, value=it["weight_kg"])
        ws.cell(row=r, column=14, value=it["available_stock"])
        ws.cell(row=r, column=15, value=it["qty_purchase"])
        ws.cell(row=r, column=16, value=it["purchase_due_date"])
        ws.cell(row=r, column=17, value=it["remark"])

        for c in range(1, 19):
            cell = ws.cell(row=r, column=c)
            cell.border = box
            if c in (1, 8):
                cell.alignment = center
            else:
                cell.alignment = left

    last_row = 9 + max(len(items_out), 10)
    if len(items_out) < 10:
        for i in range(len(items_out), 10):
            r = 10 + i
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
            ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
            ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=13)
            ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
            ws.cell(row=r, column=1, value=i + 1)
            for c in range(1, 19):
                ws.cell(row=r, column=c).border = box

    notes_row = last_row + 2
    ws.cell(row=notes_row, column=1, value="NOTES :").font = bold
    ws.merge_cells(start_row=notes_row, start_column=2, end_row=notes_row, end_column=9)
    ws.cell(row=notes_row, column=2, value=data["notes"])
    for k in range(1, 6):
        ws.merge_cells(start_row=notes_row + k, start_column=2, end_row=notes_row + k, end_column=9)

    ws.cell(row=notes_row, column=11, value="TOTAL WEIGHT:").font = bold
    ws.merge_cells(start_row=notes_row, start_column=12, end_row=notes_row, end_column=13)
    tw_cell = ws.cell(row=notes_row, column=12, value=data["total_weight"])
    tw_cell.alignment = center
    tw_cell.font = bold

    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_action(current, "export_bom_xlsx", "bom", bom_id, {"filename": filename, "engine": "default"})
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
