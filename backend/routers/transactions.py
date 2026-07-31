"""Transactions CRUD + master lists + dashboard stats + KPI + Excel I/O."""
import io
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from db import db
from deps import _now_iso, get_current_user, log_action, require_write
from models import BulkCreateRequest, BulkDeleteRequest, Transaction, TransactionCreate
from routers.bom import normalize_so_no
from routers.consumable_requests import link_purchase_to_request, unlink_purchase_from_request
from routers.bom import link_purchase_to_bom, auto_link_tx_to_bom_by_so
from services.soft_delete import NOT_DELETED_FILTER, merged, soft_delete_one, soft_delete_many

router = APIRouter(tags=["transactions"])


def _clean_doc(doc: dict) -> dict:
    doc.pop("_id", None)
    return doc


def _compute_idr(doc: dict) -> None:
    """Compute total_price_idr from total_price × exchange_rate.
    Ensures currency + exchange_rate defaults are present."""
    curr = (doc.get("currency") or "IDR").upper()
    rate = float(doc.get("exchange_rate") or 1.0)
    if curr == "IDR":
        rate = 1.0
    doc["currency"] = curr
    doc["exchange_rate"] = rate
    doc["total_price_idr"] = float(doc.get("total_price") or 0) * rate


async def _link_receipt_to_tx(receipt_id: str, tx: dict) -> None:
    """Close a store_receipt with the PO/transaction info once the tx is created."""
    if not receipt_id:
        return
    await db.store_receipts.update_one(
        {"id": receipt_id, "deleted_at": {"$exists": False}},
        {"$set": {
            "transaction_id": tx["id"],
            "po_no": tx.get("po_no") or "",
            "invoice_no": tx.get("invoice_no") or "",
            "unit_price": float(tx.get("unit_price") or 0),
            "so_no": tx.get("project_no") or "",
        }},
    )


async def _unlink_receipt_from_tx(receipt_id: str) -> None:
    """Reopen a receipt when its tx is deleted."""
    if not receipt_id:
        return
    await db.store_receipts.update_one(
        {"id": receipt_id, "deleted_at": {"$exists": False}},
        {"$set": {"transaction_id": None, "po_no": "", "invoice_no": ""}},
    )


async def _auto_create_log_receipt(tx: dict, current: dict) -> Optional[str]:
    """Auto-create a store_receipt when Purchasing marks a transaction as "Log Only"
    (post_to_store=True, should_stock=False). This makes the item appear in
    Incoming Good report immediately — no Store confirmation needed — while
    NOT counting toward physical stock (qty_remaining=0)."""
    if not tx.get("post_to_store"):
        return None
    if tx.get("should_stock", True):
        # Masuk-Stok items still require Store to confirm physical receipt.
        return None
    if tx.get("linked_receipt_id"):
        # Already linked to an existing receipt (barang datang duluan)
        return None
    now = _now_iso()
    rid = str(uuid.uuid4())
    qty = float(tx.get("qty") or 0)
    receipt = {
        "id": rid,
        "transaction_id": tx["id"],
        "source": "log-only-auto",
        "source_type": "supplier",
        "is_customer_material": False,
        "po_no": tx.get("po_no") or "",
        "invoice_no": tx.get("invoice_no") or "",
        "vendor_name": tx.get("vendor_name") or "",
        "customer_name": "",
        "item_name": tx.get("item_name") or "",
        "unit": tx.get("unit") or "",
        "unit_price": float(tx.get("unit_price") or 0),
        "do_number": "",
        "so_no": tx.get("project_no") or "",
        "qty_received": qty,
        "qty_remaining": 0.0,        # Log Only → does NOT enter stock
        "add_to_stock": False,
        "receive_date": tx.get("receive_date") or tx.get("invoice_date") or now[:10],
        "mcl_done": False,
        "mif_done": False,
        "note": tx.get("notes") or "",
        "created_by": current.get("id"),
        "created_by_username": current.get("username", ""),
        "created_at": now,
    }
    await db.store_receipts.insert_one(receipt.copy())
    await db.transactions.update_one({"id": tx["id"]}, {"$set": {"linked_receipt_id": rid}})
    return rid


# ---------------- Transactions ----------------
@router.post("/transactions", response_model=Transaction)
async def create_transaction(payload: TransactionCreate, current: dict = Depends(require_write)):
    now = _now_iso()
    tx = payload.model_dump()
    tx["id"] = str(uuid.uuid4())
    tx["created_at"] = now
    tx["updated_at"] = now
    tx["project_no"] = normalize_so_no(tx.get("project_no"))
    _compute_idr(tx)
    await db.transactions.insert_one(tx.copy())
    await log_action(current, "create_transaction", "transaction", tx["id"], {
        "vendor": tx.get("vendor_name"), "item": tx.get("item_name"),
        "invoice_no": tx.get("invoice_no"), "total": tx.get("total_price"),
        "currency": tx.get("currency"), "total_idr": tx.get("total_price_idr"),
    })
    # Auto-link to Consumable Request if referenced
    if tx.get("consumable_request_id") and tx.get("consumable_request_item_id"):
        await link_purchase_to_request(
            tx["consumable_request_id"], tx["consumable_request_item_id"],
            actual_item_name=tx.get("item_name") or "",
            vendor_name=tx.get("vendor_name") or "",
            qty_bought=float(tx.get("qty") or 0),
            unit=tx.get("unit") or "",
            purchase_date=tx.get("invoice_date") or tx.get("receive_date") or "",
            po_no=tx.get("po_no") or "",
            transaction_id=tx["id"], current=current,
        )
    # Auto-link to BOM item if referenced
    _bom_ref = tx.get("bom_item_ref") or {}
    if _bom_ref.get("bom_id") and _bom_ref.get("item_no") is not None:
        await link_purchase_to_bom(
            _bom_ref["bom_id"], int(_bom_ref["item_no"]),
            actual_item_name=tx.get("item_name") or "",
            vendor_name=tx.get("vendor_name") or "",
            qty_bought=float(tx.get("qty") or 0),
            unit=tx.get("unit") or "",
            purchase_date=tx.get("invoice_date") or tx.get("receive_date") or "",
            po_no=tx.get("po_no") or "",
            invoice_no=tx.get("invoice_no") or "",
            unit_price=float(tx.get("unit_price") or 0),
            transaction_id=tx["id"], current=current,
        )
    else:
        # No explicit ref → try fuzzy auto-link by project_no (SO) + item_name
        try: await auto_link_tx_to_bom_by_so(tx, current)
        except Exception as e:
            import logging; logging.getLogger(__name__).warning(f"auto-link BOM by SO failed: {e}")
    # Auto-close linked receipt (barang datang duluan, PO baru dibuat)
    if tx.get("linked_receipt_id"):
        await _link_receipt_to_tx(tx["linked_receipt_id"], tx)
    # Log Only items: auto-create a store_receipt so it shows in Incoming Good immediately
    await _auto_create_log_receipt(tx, current)
    return _clean_doc(tx)


@router.post("/transactions/bulk")
async def bulk_create(payload: BulkCreateRequest, current: dict = Depends(require_write)):
    now = _now_iso()
    docs = []
    for t in payload.transactions:
        d = t.model_dump()
        d["id"] = str(uuid.uuid4())
        d["created_at"] = now
        d["updated_at"] = now
        d["project_no"] = normalize_so_no(d.get("project_no"))
        _compute_idr(d)
        docs.append(d)
    if docs:
        await db.transactions.insert_many([d.copy() for d in docs])
        first = docs[0]
        await log_action(current, "bulk_create_transaction", "transaction", "-", {
            "count": len(docs), "vendor": first.get("vendor_name"),
            "invoice_no": first.get("invoice_no"), "currency": first.get("currency"),
        })
        # Auto-link each linked row to its consumable request
        for d in docs:
            if d.get("consumable_request_id") and d.get("consumable_request_item_id"):
                await link_purchase_to_request(
                    d["consumable_request_id"], d["consumable_request_item_id"],
                    actual_item_name=d.get("item_name") or "",
                    vendor_name=d.get("vendor_name") or "",
                    qty_bought=float(d.get("qty") or 0),
                    unit=d.get("unit") or "",
                    purchase_date=d.get("invoice_date") or d.get("receive_date") or "",
                    po_no=d.get("po_no") or "",
                    transaction_id=d["id"], current=current,
                )
            # Auto-link BOM item if referenced
            _br = d.get("bom_item_ref") or {}
            if _br.get("bom_id") and _br.get("item_no") is not None:
                await link_purchase_to_bom(
                    _br["bom_id"], int(_br["item_no"]),
                    actual_item_name=d.get("item_name") or "",
                    vendor_name=d.get("vendor_name") or "",
                    qty_bought=float(d.get("qty") or 0),
                    unit=d.get("unit") or "",
                    purchase_date=d.get("invoice_date") or d.get("receive_date") or "",
                    po_no=d.get("po_no") or "",
                    invoice_no=d.get("invoice_no") or "",
                    unit_price=float(d.get("unit_price") or 0),
                    transaction_id=d["id"], current=current,
                )
            else:
                try: await auto_link_tx_to_bom_by_so(d, current)
                except Exception as e:
                    import logging; logging.getLogger(__name__).warning(f"auto-link BOM by SO failed: {e}")
            # Auto-close linked receipt (barang datang duluan, PO baru dibuat)
            if d.get("linked_receipt_id"):
                await _link_receipt_to_tx(d["linked_receipt_id"], d)
            # Log Only items: auto-create a store_receipt so it shows in Incoming Good
            await _auto_create_log_receipt(d, current)
    return {"inserted": len(docs)}


@router.post("/transactions/bulk-delete")
async def bulk_delete_transactions(payload: BulkDeleteRequest, current: dict = Depends(require_write)):
    if not payload.ids:
        raise HTTPException(status_code=400, detail="Tidak ada ID yang dipilih")
    n = await soft_delete_many("transactions", {"id": {"$in": payload.ids}}, current)
    await log_action(current, "bulk_delete_transaction", "transaction", "-", {
        "count": n, "requested": len(payload.ids),
    })
    return {"deleted": n}


# ---------------- Group Batch Edit (Masterlist) ----------------
@router.get("/transactions/group")
async def get_transaction_group(
    batch_id: Optional[str] = None,
    po_no: Optional[str] = None,
    vendor_name: Optional[str] = None,
    invoice_date: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """Fetch a group of related transactions for batch-edit in Masterlist.

    Grouping priority:
      1. `batch_id` — exact match (bulk upload group)
      2. `po_no` — all rows with same PO number
      3. `vendor_name + invoice_date` — same vendor on same day

    Each row is enriched with `linked_receipt` info (add_to_stock, qty_remaining, id)
    so the UI can offer the "Masuk Stok" bulk toggle.
    """
    filt: dict = {}
    group_key: dict = {}
    if batch_id:
        filt["batch_id"] = batch_id
        group_key = {"type": "batch", "batch_id": batch_id}
    elif po_no:
        filt["po_no"] = po_no
        group_key = {"type": "po", "po_no": po_no}
    elif vendor_name and invoice_date:
        filt["vendor_name"] = vendor_name
        filt["invoice_date"] = invoice_date
        group_key = {"type": "vendor_date", "vendor_name": vendor_name, "invoice_date": invoice_date}
    else:
        raise HTTPException(status_code=400, detail="Wajib isi batch_id, po_no, atau vendor_name+invoice_date")

    filt = merged(filt, NOT_DELETED_FILTER)
    txs = await db.transactions.find(filt, {"_id": 0}).sort("created_at", 1).to_list(length=1000)
    tx_ids = [t["id"] for t in txs]
    receipts = await db.store_receipts.find(
        merged({"transaction_id": {"$in": tx_ids}}, NOT_DELETED_FILTER),
        {"_id": 0}
    ).to_list(length=2000)
    rmap = {r["transaction_id"]: r for r in receipts if r.get("transaction_id")}
    for t in txs:
        r = rmap.get(t["id"])
        t["linked_receipt"] = {
            "id": r["id"],
            "add_to_stock": bool(r.get("add_to_stock")),
            "qty_remaining": float(r.get("qty_remaining") or 0),
        } if r else None
    return {"group": group_key, "count": len(txs), "items": txs}


@router.post("/transactions/bulk-update")
async def bulk_update_transactions(payload: dict, current: dict = Depends(require_write)):
    """Update multiple transactions in one call (used by Group Batch Edit modal).

    Payload:
      { rows: [
          { id, qty?, unit?, unit_price?, notes?, item_name?, category?, po_no?,
            project_no?, invoice_no?, invoice_date?, po_date?, plan_delivery_date?,
            receive_date?, is_compliant?, is_completed?, add_to_stock? }
        ] }

    - `add_to_stock` (if present) flips the linked receipt's add_to_stock flag
      and re-syncs qty_remaining (= qty if true, else 0).
    - qty × unit_price → total_price recomputed if both present.
    """
    rows = (payload or {}).get("rows") or []
    if not rows:
        raise HTTPException(status_code=400, detail="Tidak ada baris untuk di-update")

    updated_tx = 0
    updated_receipts = 0
    now = _now_iso()
    EDITABLE_FIELDS = {
        "qty", "unit", "unit_price", "notes", "item_name", "category",
        "po_no", "project_no", "invoice_no", "invoice_date", "po_date",
        "plan_delivery_date", "receive_date", "is_compliant", "is_completed",
    }
    for r in rows:
        tx_id = r.get("id")
        if not tx_id:
            continue
        existing = await db.transactions.find_one(merged({"id": tx_id}, NOT_DELETED_FILTER))
        if not existing:
            continue

        upd = {k: v for k, v in r.items() if k in EDITABLE_FIELDS and v is not None}
        # Recompute total when qty or unit_price change
        if "qty" in upd or "unit_price" in upd:
            qty = float(upd.get("qty", existing.get("qty") or 0))
            price = float(upd.get("unit_price", existing.get("unit_price") or 0))
            upd["qty"] = qty
            upd["unit_price"] = price
            upd["total_price"] = qty * price
            merged_doc = {**existing, **upd}
            _compute_idr(merged_doc)
            upd["total_price_idr"] = merged_doc["total_price_idr"]
            upd["currency"] = merged_doc["currency"]
            upd["exchange_rate"] = merged_doc["exchange_rate"]
        if "project_no" in upd:
            upd["project_no"] = normalize_so_no(upd["project_no"])
        if upd:
            upd["updated_at"] = now
            await db.transactions.update_one({"id": tx_id}, {"$set": upd})
            updated_tx += 1

        # Sync linked receipt's add_to_stock + qty_remaining
        # Also persist the "should_stock" decision on the transaction itself so that
        # future Store receives use this value (Store cannot override).
        if "add_to_stock" in r and r["add_to_stock"] is not None:
            add_stock = bool(r["add_to_stock"])
            # Persist Purchasing's decision on the transaction
            await db.transactions.update_one(
                {"id": tx_id},
                {"$set": {"should_stock": add_stock, "updated_at": now}},
            )
            receipt = await db.store_receipts.find_one(
                merged({"transaction_id": tx_id}, NOT_DELETED_FILTER)
            )
            new_qty = float(upd.get("qty", existing.get("qty") or 0))
            new_price = float(upd.get("unit_price", existing.get("unit_price") or 0))
            if receipt:
                rupd = {
                    "add_to_stock": add_stock,
                    "qty_received": new_qty,
                    "qty_remaining": new_qty if add_stock else 0.0,
                    "unit_price": new_price,
                    "item_name": upd.get("item_name", existing.get("item_name") or ""),
                    "po_no": upd.get("po_no", existing.get("po_no") or ""),
                    "invoice_no": upd.get("invoice_no", existing.get("invoice_no") or ""),
                }
                await db.store_receipts.update_one({"id": receipt["id"]}, {"$set": rupd})
                updated_receipts += 1
            elif add_stock:
                # Create a new receipt so this row now enters stock (bulk-direct style)
                rid = str(uuid.uuid4())
                receipt_new = {
                    "id": rid,
                    "transaction_id": tx_id,
                    "source": "group-edit",
                    "source_type": "supplier",
                    "is_customer_material": False,
                    "po_no": upd.get("po_no", existing.get("po_no") or ""),
                    "invoice_no": upd.get("invoice_no", existing.get("invoice_no") or ""),
                    "vendor_name": existing.get("vendor_name") or "",
                    "customer_name": "",
                    "item_name": upd.get("item_name", existing.get("item_name") or ""),
                    "unit": upd.get("unit", existing.get("unit") or ""),
                    "unit_price": new_price,
                    "do_number": "",
                    "so_no": upd.get("project_no", existing.get("project_no") or ""),
                    "qty_received": new_qty,
                    "qty_remaining": new_qty,
                    "add_to_stock": True,
                    "receive_date": upd.get("invoice_date", existing.get("invoice_date") or now[:10]),
                    "mcl_done": False,
                    "mif_done": False,
                    "note": upd.get("notes", existing.get("notes") or ""),
                    "created_by": current.get("id"),
                    "created_by_username": current.get("username", ""),
                    "created_at": now,
                }
                await db.store_receipts.insert_one(receipt_new.copy())
                await db.transactions.update_one({"id": tx_id}, {"$set": {"linked_receipt_id": rid}})
                updated_receipts += 1
            else:
                # Log Only: no receipt yet, auto-create one so it shows in Incoming Good
                # report without needing Store to confirm (qty_remaining=0 → not stock).
                merged_tx = {**existing, **upd, "should_stock": False, "post_to_store": True}
                rid = await _auto_create_log_receipt(merged_tx, current)
                if rid:
                    updated_receipts += 1

    await log_action(current, "bulk_update_transaction", "transaction", "-", {
        "updated_tx": updated_tx, "updated_receipts": updated_receipts,
        "requested": len(rows),
    })
    return {"updated_tx": updated_tx, "updated_receipts": updated_receipts}


@router.get("/transactions")
async def list_transactions(
    current: dict = Depends(get_current_user),
    q: Optional[str] = None,
    vendor: Optional[str] = None,
    project_no: Optional[str] = None,
    po_no: Optional[str] = None,
    invoice_no: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sort_by: str = "invoice_date",
    sort_dir: str = "desc",
    page: int = 1,
    page_size: int = 50,
):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"vendor_name": {"$regex": q, "$options": "i"}},
            {"invoice_no": {"$regex": q, "$options": "i"}},
            {"project_no": {"$regex": q, "$options": "i"}},
            {"po_no": {"$regex": q, "$options": "i"}},
        ]
    if vendor:
        filt["vendor_name"] = {"$regex": vendor, "$options": "i"}
    if project_no:
        filt["project_no"] = {"$regex": project_no, "$options": "i"}
    if po_no:
        filt["po_no"] = {"$regex": po_no, "$options": "i"}
    if invoice_no:
        filt["invoice_no"] = {"$regex": invoice_no, "$options": "i"}
    if start_date or end_date:
        date_filt = {}
        if start_date:
            date_filt["$gte"] = start_date
        if end_date:
            date_filt["$lte"] = end_date
        filt["invoice_date"] = date_filt

    direction = -1 if sort_dir == "desc" else 1
    filt = merged(filt, NOT_DELETED_FILTER)
    total = await db.transactions.count_documents(filt)
    cursor = db.transactions.find(filt, {"_id": 0}).sort(sort_by, direction).skip((page - 1) * page_size).limit(page_size)
    items = await cursor.to_list(length=page_size)
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.get("/transactions/{tx_id}")
async def get_transaction(tx_id: str, current: dict = Depends(get_current_user)):
    doc = await db.transactions.find_one({"id": tx_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    return doc


@router.put("/transactions/{tx_id}", response_model=Transaction)
async def update_transaction(tx_id: str, payload: TransactionCreate, current: dict = Depends(require_write)):
    existing = await db.transactions.find_one({"id": tx_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    upd = payload.model_dump()
    upd["updated_at"] = _now_iso()
    upd["project_no"] = normalize_so_no(upd.get("project_no"))
    _compute_idr(upd)
    await db.transactions.update_one({"id": tx_id}, {"$set": upd})
    doc = await db.transactions.find_one({"id": tx_id}, {"_id": 0})
    await log_action(current, "update_transaction", "transaction", tx_id, {
        "vendor": upd.get("vendor_name"), "item": upd.get("item_name"),
        "invoice_no": upd.get("invoice_no"), "total": upd.get("total_price"),
        "currency": upd.get("currency"),
    })
    return doc


@router.delete("/transactions/{tx_id}")
async def delete_transaction(tx_id: str, current: dict = Depends(require_write)):
    existing = await db.transactions.find_one(merged({"id": tx_id}, NOT_DELETED_FILTER))
    if not existing:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    await soft_delete_one("transactions", {"id": tx_id}, current)
    await log_action(current, "delete_transaction", "transaction", tx_id, {
        "vendor": existing.get("vendor_name"), "item": existing.get("item_name"),
        "invoice_no": existing.get("invoice_no"), "total": existing.get("total_price"),
    })
    # Unlink from Consumable Request if this tx was linked
    if existing.get("consumable_request_id") and existing.get("consumable_request_item_id"):
        await unlink_purchase_from_request(
            existing["consumable_request_id"], existing["consumable_request_item_id"],
            transaction_id=tx_id,
        )
    # Reopen linked receipt if any
    if existing.get("linked_receipt_id"):
        await _unlink_receipt_from_tx(existing["linked_receipt_id"])
    return {"ok": True}


@router.post("/transactions/{tx_id}/log-incoming")
async def log_transaction_as_incoming(tx_id: str, payload: dict, current: dict = Depends(require_write)):
    """Log an existing transaction as Incoming Good in Store.

    Params:
    - add_to_stock (bool, default False): False = hanya masuk laporan Incoming Good tanpa mempengaruhi stok
    - source_type (str, "supplier"|"customer", default "supplier")

    Creates a store_receipt entry linked to this transaction.
    Idempotent — if transaction already has linked_receipt_id, returns that.
    """
    from db import db
    tx = await db.transactions.find_one({"id": tx_id})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    if tx.get("linked_receipt_id"):
        return {"ok": True, "already_linked": True, "receipt_id": tx["linked_receipt_id"]}

    add_stock = bool(payload.get("add_to_stock", False))
    source_type = payload.get("source_type", "supplier")

    now = _now_iso()
    receipt = {
        "id": str(uuid.uuid4()),
        "transaction_id": tx_id,
        "source": "manual-log",
        "source_type": source_type,
        "vendor_name": tx.get("vendor_name") or "",
        "item_name": tx.get("item_name") or "",
        "qty_received": tx.get("qty") or 0,
        "unit": tx.get("unit") or "",
        "receive_date": tx.get("invoice_date") or now[:10],
        "po_no": tx.get("po_no") or "",
        "do_number": tx.get("do_number") or "",
        "invoice_no": tx.get("invoice_no") or "",
        "so_no": tx.get("project_no") or tx.get("so_no") or "",
        "add_to_stock": add_stock,
        "created_at": now,
        "created_by": current.get("username", ""),
    }
    await db.store_receipts.insert_one(receipt.copy())
    await db.transactions.update_one({"id": tx_id}, {"$set": {"linked_receipt_id": receipt["id"]}})
    return {"ok": True, "receipt_id": receipt["id"], "add_to_stock": add_stock}


# ---------------- Bulk Transaksi (direct-to-store, no store approval) ----------------
@router.post("/transactions/bulk-direct")
async def bulk_direct_create(payload: dict, current: dict = Depends(require_write)):
    """Bulk transactions with direct incoming goods creation (bypasses store approval).

    Each row must include `masuk_stok` (bool). Row is saved as:
      - a transaction row (post_to_store=False, no approval workflow)
      - a store_receipt row (source='bulk-direct', add_to_stock=<masuk_stok>)
    Tanggal nota = tanggal terima (same day). Nomor PO is optional.
    """
    rows = payload.get("rows") or []
    if not rows:
        raise HTTPException(status_code=400, detail="Tidak ada baris")

    now = _now_iso()
    # One batch_id shared across all rows so we can group-edit them later
    batch_id = str(uuid.uuid4())
    tx_docs = []
    receipt_docs = []
    for r in rows:
        if not r.get("vendor_name") or not str(r.get("vendor_name")).strip():
            raise HTTPException(status_code=400, detail="Nama Supplier wajib pada setiap baris")
        if not r.get("item_name") or not str(r.get("item_name")).strip():
            raise HTTPException(status_code=400, detail="Nama Barang wajib pada setiap baris")
        if r.get("masuk_stok") is None:
            raise HTTPException(status_code=400, detail="Checklist Masuk Stok wajib diisi (ya/tidak)")
        qty = float(r.get("qty") or 0)
        if qty <= 0:
            raise HTTPException(status_code=400, detail="Qty harus > 0")
        unit_price = float(r.get("unit_price") or 0)
        total_price = float(r.get("total_price") or (qty * unit_price))
        invoice_date = r.get("invoice_date") or now[:10]
        tx_id = str(uuid.uuid4())
        add_stock = bool(r.get("masuk_stok"))
        tx_doc = {
            "id": tx_id,
            "invoice_date": invoice_date,
            "project_no": normalize_so_no(r.get("project_no") or ""),
            "po_no": r.get("po_no") or "",
            "vendor_name": str(r.get("vendor_name")).strip(),
            "category": r.get("category") or "Uncategorized",
            "item_name": str(r.get("item_name")).strip(),
            "qty": qty,
            "unit": r.get("unit") or "Ea",
            "unit_price": unit_price,
            "total_price": total_price,
            "currency": (r.get("currency") or "IDR").upper(),
            "exchange_rate": float(r.get("exchange_rate") or 1.0),
            "total_price_idr": 0.0,
            "invoice_no": r.get("invoice_no") or "",
            "po_date": r.get("po_date"),
            "plan_delivery_date": r.get("plan_delivery_date"),
            "receive_date": invoice_date,          # tanggal terima = tanggal nota
            "notes": r.get("notes") or "",
            "is_compliant": True,
            "is_completed": True,
            "post_to_store": False,                # bypass store approval
            "source": "bulk-direct",               # marker for audit
            "batch_id": batch_id,                  # group all rows of this upload
            "should_stock": add_stock,             # Purchasing's decision: masuk stok or log-only
            "created_at": now,
            "updated_at": now,
        }
        _compute_idr(tx_doc)
        # pass through link fields (used after insert)
        tx_doc["consumable_request_id"] = r.get("consumable_request_id") or None
        tx_doc["consumable_request_item_id"] = r.get("consumable_request_item_id") or None
        tx_docs.append(tx_doc)

        receipt_docs.append({
            "id": str(uuid.uuid4()),
            "transaction_id": tx_id,
            "source": "bulk-direct",
            "batch_id": batch_id,
            "source_type": "supplier",
            "is_customer_material": False,
            "po_no": tx_doc["po_no"],
            "invoice_no": tx_doc["invoice_no"],
            "vendor_name": tx_doc["vendor_name"],
            "customer_name": "",
            "item_name": tx_doc["item_name"],
            "unit": tx_doc["unit"],
            "unit_price": unit_price,
            "do_number": "",
            "so_no": tx_doc["project_no"],
            "qty_received": qty,
            "qty_remaining": qty if add_stock else 0.0,
            "add_to_stock": add_stock,
            "receive_date": invoice_date,
            "mcl_done": False,
            "mif_done": False,
            "note": tx_doc["notes"],
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": now,
        })

    await db.transactions.insert_many([d.copy() for d in tx_docs])
    await db.store_receipts.insert_many([d.copy() for d in receipt_docs])
    # Auto-link each row to its consumable request (if referenced)
    for d in tx_docs:
        if d.get("consumable_request_id") and d.get("consumable_request_item_id"):
            await link_purchase_to_request(
                d["consumable_request_id"], d["consumable_request_item_id"],
                actual_item_name=d.get("item_name") or "",
                vendor_name=d.get("vendor_name") or "",
                qty_bought=float(d.get("qty") or 0),
                unit=d.get("unit") or "",
                purchase_date=d.get("invoice_date") or "",
                po_no=d.get("po_no") or "",
                transaction_id=d["id"], current=current,
            )
    await log_action(current, "bulk_direct_transaction", "transaction", "-", {
        "count": len(tx_docs),
        "stock_count": sum(1 for r in receipt_docs if r["add_to_stock"]),
    })
    return {
        "inserted": len(tx_docs),
        "receipts": len(receipt_docs),
        "with_stock": sum(1 for r in receipt_docs if r["add_to_stock"]),
        "tx_ids": [d["id"] for d in tx_docs],
    }


# ---------------- Master lists ----------------
@router.get("/master/vendors")
async def master_vendors(current: dict = Depends(get_current_user)):
    vendors = await db.transactions.distinct("vendor_name")
    return sorted([v for v in vendors if v])


@router.get("/master/categories")
async def master_categories(current: dict = Depends(get_current_user)):
    """Distinct list of transaction item categories for autocomplete."""
    cats = await db.transactions.distinct("category")
    # Filter empty/None then dedupe
    clean = sorted({str(c).strip() for c in cats if c and str(c).strip()})
    return clean


@router.get("/master/items")
async def master_items(current: dict = Depends(get_current_user)):
    pipeline = [
        {"$match": NOT_DELETED_FILTER},
        {"$sort": {"invoice_date": -1, "created_at": -1}},
        {"$group": {
            "_id": "$item_name",
            "last_price": {"$first": "$unit_price"},
            "last_vendor": {"$first": "$vendor_name"},
            "last_date": {"$first": "$invoice_date"},
            "last_category": {"$first": "$category"},
            "unit": {"$first": "$unit"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
        {"$limit": 5000},
    ]
    result = await db.transactions.aggregate(pipeline).to_list(length=5000)
    return [{"item_name": r["_id"], "last_price": r["last_price"], "last_vendor": r["last_vendor"],
             "last_date": r["last_date"], "unit": r.get("unit", "Ea"),
             "last_category": r.get("last_category") or "Uncategorized",
             "count": r["count"]}
            for r in result if r["_id"]]


# ---------------- Dashboard Stats ----------------
@router.get("/stats/monthly")
async def stats_monthly(current: dict = Depends(get_current_user)):
    """Current-month total (from 1st of month to today) + PO count.
    Returns IDR totals so multi-currency is normalized."""
    now = datetime.now()
    first_day = now.replace(day=1).date().isoformat()
    today_str = now.date().isoformat()

    match = {"invoice_date": {"$gte": first_day, "$lte": today_str}}

    agg = await db.transactions.aggregate([
        {"$match": match},
        {"$group": {
            "_id": None,
            "total_idr": {"$sum": {"$ifNull": ["$total_price_idr", "$total_price"]}},
            "tx_count": {"$sum": 1},
        }}
    ]).to_list(length=1)

    po_nos = await db.transactions.distinct("po_no", match)
    po_count = len([p for p in po_nos if p])

    return {
        "period": {"start": first_day, "end": today_str, "month": now.strftime("%B %Y")},
        "total_amount_idr": agg[0]["total_idr"] if agg else 0,
        "transactions": agg[0]["tx_count"] if agg else 0,
        "po_count": po_count,
    }


@router.get("/stats/summary")
async def stats_summary(current: dict = Depends(get_current_user), year: Optional[int] = None):
    match: dict = {}
    if year:
        match["invoice_date"] = {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}

    total_count = await db.transactions.count_documents(match)

    agg_total = await db.transactions.aggregate([
        {"$match": match},
        {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
    ]).to_list(length=1)
    total_amount = agg_total[0]["total"] if agg_total else 0

    top_vendors = await db.transactions.aggregate([
        {"$match": match},
        {"$group": {"_id": "$vendor_name", "total": {"$sum": "$total_price"}, "count": {"$sum": 1}}},
        {"$sort": {"total": -1}},
        {"$limit": 8},
    ]).to_list(length=8)

    monthly = await db.transactions.aggregate([
        {"$match": match},
        {"$group": {"_id": {"$substr": ["$invoice_date", 0, 7]}, "total": {"$sum": "$total_price"}, "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
    ]).to_list(length=200)

    vendors = await db.transactions.distinct("vendor_name", match)
    items = await db.transactions.distinct("item_name", match)

    return {
        "total_transactions": total_count,
        "total_amount": total_amount,
        "unique_vendors": len(vendors),
        "unique_items": len(items),
        "top_vendors": [{"vendor": v["_id"] or "-", "total": v["total"], "count": v["count"]} for v in top_vendors],
        "monthly": [{"month": m["_id"], "total": m["total"], "count": m["count"]} for m in monthly],
    }


# ---------------- KPI Purchasing ----------------
@router.get("/kpi")
async def kpi_report(
    current: dict = Depends(get_current_user),
    start_date: str = Query(...),
    end_date: str = Query(...),
    ontime_grace_days: int = 7,
):
    match = {"invoice_date": {"$gte": start_date, "$lte": end_date}}
    txs = await db.transactions.find(match, {"_id": 0}).to_list(length=200000)

    groups: dict = defaultdict(list)
    for t in txs:
        po_no = (t.get("po_no") or "").strip()
        inv_no = (t.get("invoice_no") or "").strip()
        key = po_no if po_no else (f"INV:{inv_no}" if inv_no else f"ID:{t.get('id')}")
        groups[key].append(t)

    total_po = len(groups)
    on_time_po = 0
    compliant_po = 0
    completed_po = 0
    late_details = []

    for key, items in groups.items():
        po_on_time = True
        for it in items:
            pd = it.get("po_date")
            rd = it.get("receive_date")
            if not pd or not rd:
                po_on_time = False
                break
            try:
                pd_d = datetime.strptime(pd, "%Y-%m-%d").date()
                rd_d = datetime.strptime(rd, "%Y-%m-%d").date()
                if rd_d > pd_d + timedelta(days=ontime_grace_days):
                    po_on_time = False
                    break
            except Exception:
                po_on_time = False
                break
        if po_on_time:
            on_time_po += 1
        else:
            first = items[0]
            late_details.append({
                "po_no": key,
                "vendor": first.get("vendor_name", ""),
                "invoice_no": first.get("invoice_no", ""),
                "po_date": first.get("po_date"),
                "receive_date": first.get("receive_date"),
                "item_name": first.get("item_name", ""),
            })

        if all(it.get("is_compliant", True) for it in items):
            compliant_po += 1
        if all(it.get("is_completed", True) for it in items):
            completed_po += 1

    def pct(n, d):
        return round((n / d) * 100, 2) if d else 0.0

    on_time_pct = pct(on_time_po, total_po)
    compliant_pct = pct(compliant_po, total_po)
    completed_pct = pct(completed_po, total_po)

    score_on_time = round(on_time_pct * 0.40, 2)
    score_compliance = round(compliant_pct * 0.35, 2)
    score_completion = round(completed_pct * 0.25, 2)
    total_score = round(score_on_time + score_compliance + score_completion, 2)

    if total_score >= 90:
        category = "SANGAT BAIK"
    elif total_score >= 80:
        category = "BAIK"
    elif total_score >= 71:
        category = "CUKUP"
    else:
        category = "PERLU PERBAIKAN"

    return {
        "period": {"start_date": start_date, "end_date": end_date, "ontime_grace_days": ontime_grace_days},
        "total_po": total_po,
        "kpis": [
            {
                "no": 1, "name": "On Time Delivery",
                "description": "Persentase pengiriman barang dari supplier yang diterima tepat waktu sesuai dengan jadwal (ETA)",
                "formula_num": "Jumlah On Time Shipment", "formula_den": "Total PO",
                "target": "≥ 90%", "weight": 40,
                "numerator": on_time_po, "denominator": total_po,
                "achievement": on_time_pct, "score": score_on_time, "max_score": 40,
            },
            {
                "no": 2, "name": "Compliance Quality",
                "description": "Persentase pengiriman barang dari supplier yang diterima sesuai dengan pemesanan (spesifikasi)",
                "formula_num": "Jumlah Pembelian yang sesuai Spesifikasi", "formula_den": "Total PO",
                "target": "≥ 98%", "weight": 35,
                "numerator": compliant_po, "denominator": total_po,
                "achievement": compliant_pct, "score": score_compliance, "max_score": 35,
            },
            {
                "no": 3, "name": "PO Completion Rate",
                "description": "Persentase Purchase Order yang berhasil diselesaikan dalam periode tertentu sebagai indikator efektivitas proses procurement",
                "formula_num": "Jumlah PO selesai", "formula_den": "Total PO",
                "target": "≥ 90%", "weight": 25,
                "numerator": completed_po, "denominator": total_po,
                "achievement": completed_pct, "score": score_completion, "max_score": 25,
            },
        ],
        "total_score": total_score,
        "category": category,
        "late_details": late_details[:100],
    }


# ---------------- Excel Import/Export ----------------
EXPORT_HEADERS = [
    "Tanggal Invoice", "Nomor Project (SO)", "Nomor PO", "Nama Toko", "Kategori", "Nama Barang",
    "Qty", "Unit", "Unit Price", "Total Price", "Nomor Invoice", "Tanggal PO", "Plan Delivery", "Tanggal Terima", "Catatan"
]


def _to_date_str(val: Any) -> Optional[str]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s


def _to_float(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(val)
    except Exception:
        try:
            return float(str(val).replace(",", "").replace(".", ""))
        except Exception:
            return 0.0


@router.get("/transactions/export/xlsx")
async def export_xlsx(current: dict = Depends(get_current_user)):
    docs = await db.transactions.find({}, {"_id": 0}).sort("invoice_date", 1).to_list(length=100000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pembelian"
    ws.append(EXPORT_HEADERS)
    for d in docs:
        ws.append([
            d.get("invoice_date", ""), d.get("project_no", ""), d.get("po_no", ""),
            d.get("vendor_name", ""), d.get("category", "Uncategorized"), d.get("item_name", ""), d.get("qty", 0),
            d.get("unit", ""), d.get("unit_price", 0), d.get("total_price", 0),
            d.get("invoice_no", ""), d.get("po_date", ""), d.get("plan_delivery_date", ""), d.get("receive_date", ""), d.get("notes", ""),
        ])
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        max_len = max([len(str(header))] + [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, min(ws.max_row, 100) + 1)])
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"laporan_pembelian_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/transactions/import/xlsx")
async def import_xlsx(file: UploadFile = File(...), current: dict = Depends(get_current_user)):
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File Excel tidak valid: {e}")

    # Track for storage management
    try:
        from routers.storage import record_temp_upload
        await record_temp_upload(current, "transactions_import", file.filename or "transactions.xlsx", content,
                                  mime=file.content_type or "", related_entity="Master Transaksi")
    except Exception:
        pass

    inserted = 0
    errors: List[str] = []
    now = _now_iso()

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_row_idx = None
        header_map = {}
        for i, row in enumerate(rows[:5]):
            row_str = [str(c).strip().lower() if c is not None else "" for c in row]
            if any("tanggal" in c or "invoice date" in c for c in row_str) and any("nama" in c or "description" in c or "item" in c for c in row_str):
                header_row_idx = i
                for idx, cell in enumerate(row_str):
                    header_map[idx] = cell
                break
        if header_row_idx is None:
            continue

        def find_col(*keywords) -> Optional[int]:
            for idx, cell in header_map.items():
                for kw in keywords:
                    if kw in cell:
                        return idx
            return None

        col_date = find_col("tanggal invoice", "invoice date", "tanggal")
        col_so = find_col("project", "so")
        col_po = find_col("purchase order", "po no", "po")
        col_vendor = find_col("toko", "vendor", "supplier")
        col_category = find_col("kategori", "category", "item code")
        col_item = find_col("nama barang", "detail description", "description", "item")
        col_qty = find_col("qty", "quantity")
        col_unit = find_col("item unit", "unit", "satuan")
        col_price = find_col("harga", "unit price", "price")
        col_total = find_col("jumlah", "amount", "total")
        col_inv = find_col("nomor invoice", "invoice no")
        col_podate = find_col("po date", "purchase order po date", "tanggal po")
        col_recv = find_col("receive", "terima")

        docs = []
        for row in rows[header_row_idx + 1:]:
            if row is None or all(c is None or c == "" for c in row):
                continue
            try:
                item_name = row[col_item] if col_item is not None and col_item < len(row) else None
                if not item_name:
                    continue
                d = {
                    "id": str(uuid.uuid4()),
                    "invoice_date": _to_date_str(row[col_date]) if col_date is not None and col_date < len(row) else "",
                    "project_no": normalize_so_no(row[col_so]) if col_so is not None and col_so < len(row) and row[col_so] not in (None, "") else "",
                    "po_no": str(row[col_po]) if col_po is not None and col_po < len(row) and row[col_po] not in (None, "") else "",
                    "vendor_name": str(row[col_vendor]).strip() if col_vendor is not None and col_vendor < len(row) and row[col_vendor] else "",
                    "category": (str(row[col_category]).strip() if col_category is not None and col_category < len(row) and row[col_category] else "") or "Uncategorized",
                    "item_name": str(item_name).strip(),
                    "qty": _to_float(row[col_qty]) if col_qty is not None and col_qty < len(row) else 0,
                    "unit": str(row[col_unit]).strip() if col_unit is not None and col_unit < len(row) and row[col_unit] else "Ea",
                    "unit_price": _to_float(row[col_price]) if col_price is not None and col_price < len(row) else 0,
                    "total_price": _to_float(row[col_total]) if col_total is not None and col_total < len(row) else 0,
                    "currency": "IDR",
                    "exchange_rate": 1.0,
                    "invoice_no": str(row[col_inv]) if col_inv is not None and col_inv < len(row) and row[col_inv] not in (None, "") else "",
                    "po_date": _to_date_str(row[col_podate]) if col_podate is not None and col_podate < len(row) else None,
                    "receive_date": _to_date_str(row[col_recv]) if col_recv is not None and col_recv < len(row) else None,
                    "notes": "",
                    # Import from Excel: NEVER auto-post to store (staff decides later per row)
                    "post_to_store": False,
                    "is_compliant": True,
                    "is_completed": True,
                    "created_at": now,
                    "updated_at": now,
                }
                if d["total_price"] == 0 and d["qty"] and d["unit_price"]:
                    d["total_price"] = d["qty"] * d["unit_price"]
                d["total_price_idr"] = d["total_price"]  # IDR default
                if not d["invoice_date"]:
                    continue
                docs.append(d)
            except Exception as e:
                errors.append(f"Sheet {sn}: {e}")

        if docs:
            await db.transactions.insert_many([d.copy() for d in docs])
            inserted += len(docs)

    return {"inserted": inserted, "errors": errors[:20]}
