"""Excel-as-master form templates.

Admin uploads an XLSX file with placeholder syntax; system substitutes and renders via LibreOffice.

Placeholder syntax (case-sensitive):
- Simple: {{vendor_name}}, {{po_no}}, {{receive_date}}, {{print_date}}, {{printed_by}}, {{company_name}}
- Image: {{IMAGE:company_logo}} — cell content replaced with logo image (fitted to cell)
- Table: any cell containing {{items.item_name}} (or any {{items.KEY}}) marks that row as a table template.
  The entire row is duplicated per item, preserving cell styling/merges.
  Available item keys: __index__, so_no, item_name, qty_received, unit, receive_date

MCL "items" comes from grouping sibling store_receipts.
"""
import base64
import hashlib
import io
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
from copy import copy
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from db import db
from deps import _now_iso, get_current_user, log_action
from services.soft_delete import NOT_DELETED_FILTER, merged, soft_delete_one

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

router = APIRouter(tags=["excel-templates"])
logger = logging.getLogger(__name__)

ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
ADMIN_ROLES = {"admin", "super_admin", "supervisor"}
COLLECTION = "excel_form_templates"

# Supported form codes + available bindings shown to user
FORM_CODES = {
    "MCL": {
        "label": "Material Control Label",
        "fields": [
            "company_name", "receive_date", "vendor_name", "po_no",
            "do_number", "invoice_no", "print_date", "printed_by",
        ],
        "table_key": "items",
        "table_fields": ["__index__", "so_no", "item_name", "qty_received", "unit", "receive_date"],
    },
    "MIF": {
        "label": "Material Issue Form",
        "fields": [
            "company_name", "receive_date", "vendor_name", "po_no",
            "do_number", "invoice_no", "print_date", "printed_by",
            "requested_by", "issue_date", "form_no",
        ],
        "table_key": "items",
        "table_fields": ["__index__", "so_no", "item_name", "qty_received", "unit", "remark", "receive_date"],
    },
    "MII": {
        "label": "Material Incoming Inspection (QC)",
        "fields": [
            "company_name", "source_type", "source_name",
            "supplier_name", "customer_name",
            "do_no", "po_no", "inspection_date", "receive_date",
            "inspector_name", "leader_name", "print_date", "printed_by",
        ],
        "table_key": "items",
        "table_fields": [
            "__index__", "so_no", "batch_grade_heat", "mill_cert_no",
            "description", "qty", "unit",
            "dimension_spec", "dimension_actual", "visual",
            "result_ok", "result_ng", "remark",
        ],
    },
    "BOM": {
        "label": "Bill of Material (BOM Release)",
        "fields": [
            "company_name", "bom_no", "revision", "approved_date",
            "project_name", "drawing_no", "customer", "class_material",
            "so_no", "delivery_date", "notes", "total_weight",
            "prepared_by", "checked_by", "acknowledged_by", "approved_by",
            "print_date", "printed_by",
        ],
        "table_key": "items",
        "table_fields": [
            "__index__", "item_no", "item_name", "specification",
            "qty", "uom", "material", "weight_kg",
            "available_stock", "qty_purchase", "purchase_due_date", "remark",
        ],
    },
}

PLACEHOLDER_RE = re.compile(r"\{\{\s*([\w\.\:]+)\s*\}\}")


def _enable_wrap(cell):
    """Enable wrap_text on cell alignment (preserve other alignment props)."""
    if cell is None:
        return
    try:
        from openpyxl.styles import Alignment
        al = cell.alignment
        cell.alignment = Alignment(
            horizontal=al.horizontal, vertical=al.vertical or "top",
            text_rotation=al.text_rotation, wrap_text=True,
            shrink_to_fit=False, indent=al.indent,
        )
    except Exception:
        pass


def _require_admin(current: dict):
    if current.get("role") not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Hanya admin yang boleh mengelola template Excel")


def _logo_path() -> Optional[Path]:
    for name in ("logo.png", "letterhead.png", "kop_surat.webp"):
        p = ASSETS_DIR / name
        if p.exists():
            return p
    return None


# ---------------- CRUD ----------------
@router.get("/excel-templates")
async def list_excel_templates(current: dict = Depends(get_current_user)):
    docs = await db[COLLECTION].find(merged({}, NOT_DELETED_FILTER),
                                     {"_id": 0, "xlsx_base64": 0}).sort("code", 1).to_list(length=200)
    return docs


@router.get("/excel-templates/codes")
async def list_form_codes(current: dict = Depends(get_current_user)):
    return [{"code": k, **v} for k, v in FORM_CODES.items()]


@router.get("/excel-templates/{code}/active")
async def get_active_excel(code: str, current: dict = Depends(get_current_user)):
    doc = await db[COLLECTION].find_one(
        merged({"code": code.upper(), "is_active": True}, NOT_DELETED_FILTER),
        {"_id": 0, "xlsx_base64": 0},
        sort=[("updated_at", -1)],
    )
    if not doc:
        raise HTTPException(status_code=404, detail=f"Belum ada Excel template aktif untuk {code}")
    return doc


@router.post("/excel-templates/upload")
async def upload_excel_template(
    code: str = Form(...),
    filename: Optional[str] = Form(None),
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    _require_admin(current)
    code = code.upper().strip()
    if code not in FORM_CODES:
        raise HTTPException(status_code=400, detail=f"Kode form tidak dikenali. Pilihan: {list(FORM_CODES.keys())}")
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Hanya file .xlsx / .xlsm yang didukung")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File terlalu besar (max 5 MB)")
    # Try open to validate
    try:
        openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File Excel tidak valid: {e}")

    # Deactivate old templates for this code
    await db[COLLECTION].update_many({"code": code}, {"$set": {"is_active": False}})

    doc = {
        "id": str(uuid.uuid4()),
        "code": code,
        "filename": filename or file.filename or f"{code}.xlsx",
        "xlsx_base64": base64.b64encode(content).decode("ascii"),
        "size_bytes": len(content),
        "is_active": True,
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
        "updated_by": current.get("username"),
    }
    await db[COLLECTION].insert_one(doc.copy())
    await log_action(current, "upload_excel_template", "excel_form_template", doc["id"],
                     {"code": code, "size": len(content)})
    doc.pop("xlsx_base64", None)
    doc.pop("_id", None)
    return doc


@router.get("/excel-templates/{tid}/download")
async def download_excel_template(tid: str, current: dict = Depends(get_current_user)):
    doc = await db[COLLECTION].find_one(merged({"id": tid}, NOT_DELETED_FILTER))
    if not doc:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    data = base64.b64decode(doc["xlsx_base64"])
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{doc.get("filename","template.xlsx")}"'},
    )


@router.delete("/excel-templates/{tid}")
async def delete_excel_template(tid: str, current: dict = Depends(get_current_user)):
    _require_admin(current)
    n = await soft_delete_one(COLLECTION, {"id": tid}, current)
    if not n:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    return {"deleted": 1}


@router.post("/excel-templates/{tid}/activate")
async def activate_excel_template(tid: str, current: dict = Depends(get_current_user)):
    _require_admin(current)
    doc = await db[COLLECTION].find_one(merged({"id": tid}, NOT_DELETED_FILTER))
    if not doc:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    await db[COLLECTION].update_many({"code": doc["code"]}, {"$set": {"is_active": False}})
    await db[COLLECTION].update_one({"id": tid}, {"$set": {"is_active": True, "updated_at": _now_iso()}})
    return {"ok": True}


# ---------------- Substitution engine ----------------
def _cell_value_to_str(val):
    if val is None:
        return ""
    return str(val)


def _substitute_simple(text: str, data: dict) -> str:
    """Replace {{key}} placeholders (excluding items.* which are table rows)."""
    def repl(m):
        key = m.group(1).strip()
        if key.startswith("items."):
            return m.group(0)  # leave table placeholders alone
        if key.startswith("IMAGE:"):
            return m.group(0)
        val = data.get(key, "")
        return "" if val is None else str(val)
    return PLACEHOLDER_RE.sub(repl, text)


def _row_has_table_marker(ws, row_idx: int) -> bool:
    for col in ws.iter_cols(min_row=row_idx, max_row=row_idx, values_only=False):
        for cell in col:
            v = cell.value
            if isinstance(v, str) and "{{items." in v:
                return True
    return False


def _apply_substitution(wb, data: dict):
    """Substitute placeholders in workbook in place.

    Table expansion: any row containing {{items.KEY}} becomes template row.
    For each item in data['items'], row is duplicated (row_idx moves down) with substituted values.

    Page setup:
    - fitToWidth=1: kolom tidak boleh overflow horizontal
    - fitToHeight=0: allow overflow vertikal ke page 2+ (kalau items > capacity template)
    - Header row (yang punya table headers) di-set sebagai print_title_rows → auto-repeat di setiap halaman
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    items = data.get("items") or []
    for ws in wb.worksheets:
        # A4 portrait, allow multi-page vertikal
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 0 = unlimited pages vertically
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)

        # Find table template rows AND their header rows (for print_title_rows)
        table_header_row = None
        template_row_num = None
        max_row = ws.max_row
        for r in range(1, max_row + 1):
            if _row_has_table_marker(ws, r):
                template_row_num = r
                # Header row is usually the row just above the template row
                if r > 1:
                    table_header_row = r - 1
                break
        # Set print titles so header (rows 1..header_row) repeats on every page
        if table_header_row:
            ws.print_title_rows = f"1:{table_header_row}"

        # Find table template rows (scan first)
        max_row = ws.max_row
        max_col = ws.max_column
        template_rows = []
        for r in range(1, max_row + 1):
            if _row_has_table_marker(ws, r):
                template_rows.append(r)

        # Expand each template row (process from bottom to top to preserve row numbers)
        for tr in sorted(template_rows, reverse=True):
            # Capture merged cell ranges that touch the template row so we can duplicate
            # them for each newly-inserted item row (openpyxl doesn't do this automatically).
            from openpyxl.worksheet.merge import MergedCellRange
            tr_merges = []  # list of (start_col, end_col) — assuming single-row merges
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row == tr and mr.max_row == tr:
                    tr_merges.append((mr.min_col, mr.max_col))
            # Detect NO column: any column in the template row whose value is either a bare
            # integer OR a `{{items.__index__}}` placeholder — templates may use either style.
            no_columns = []
            for c in range(1, max_col + 1):
                v = ws.cell(row=tr, column=c).value
                if isinstance(v, int) or (isinstance(v, str) and v.strip().isdigit() and len(v.strip()) <= 2):
                    no_columns.append(c)
                elif isinstance(v, str) and "{{items.__index__}}" in v:
                    no_columns.append(c)
            # Capture the merge pattern of the template row — used to detect which rows
            # below are "static slots" (share the same table structure).
            tr_merge_key = frozenset((mr[0], mr[1]) for mr in tr_merges)
            # Count consecutive "static" empty rows BELOW the marker. A row is a slot if it
            # shares the SAME merged-column pattern as the template row (i.e. it's a pre-drawn
            # extra row of the same table). This is more robust than integer detection since
            # templates commonly use `{{items.__index__}}` placeholder.
            static_rows_below = 0
            probe = tr + 1
            while probe <= ws.max_row and not _row_has_table_marker(ws, probe):
                # Foreign placeholder → stop
                has_other_placeholder = False
                for cc in range(1, max_col + 1):
                    v2 = ws.cell(row=probe, column=cc).value
                    if isinstance(v2, str) and "{{" in v2:
                        has_other_placeholder = True
                        break
                if has_other_placeholder:
                    break
                # Check if the row's merge pattern matches the template row
                probe_merges = frozenset(
                    (mr.min_col, mr.max_col)
                    for mr in ws.merged_cells.ranges
                    if mr.min_row == probe and mr.max_row == probe
                )
                # Match by merge pattern (preferred) OR — if template row has no merges — by
                # NO-column integer detection (fallback for simple templates).
                if tr_merge_key and probe_merges == tr_merge_key:
                    static_rows_below += 1
                    probe += 1
                    continue
                # Fallback (for templates with no merges at all): require NO column to hold int
                if not tr_merge_key and no_columns:
                    is_slot = True
                    for c_no in no_columns:
                        v_no = ws.cell(row=probe, column=c_no).value
                        is_int_like = isinstance(v_no, int) or (
                            isinstance(v_no, str) and v_no.strip().isdigit()
                        )
                        if not is_int_like:
                            is_slot = False
                            break
                    if is_slot:
                        static_rows_below += 1
                        probe += 1
                        continue
                break
            # Capture cells in the template row
            src_cells = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=tr, column=c)
                src_cells.append({
                    "value": cell.value,
                    "font": copy(cell.font) if cell.has_style else None,
                    "fill": copy(cell.fill) if cell.has_style else None,
                    "border": copy(cell.border) if cell.has_style else None,
                    "alignment": copy(cell.alignment) if cell.has_style else None,
                    "number_format": cell.number_format,
                })
            # Capture template row's height BEFORE inserting — new rows in openpyxl have
            # no explicit dimension, which LibreOffice may render as near-zero height.
            template_row_height = None
            if tr in ws.row_dimensions and ws.row_dimensions[tr].height:
                template_row_height = ws.row_dimensions[tr].height

            # Slot planning
            n_needed = max(len(items), 1)
            # Available slots without inserting: 1 (template row) + static_rows_below
            available = 1 + static_rows_below
            extras_to_insert = max(0, n_needed - available)  # only insert if overflow
            if extras_to_insert > 0:
                ws.insert_rows(tr + 1 + static_rows_below, amount=extras_to_insert)
            # Total slots after planning (data + trailing empty)
            total_slots = max(available, n_needed)
            # If items fit within the designer's pre-drawn slots (no overflow), force
            # single-page fit. Prevents Windows LibreOffice from splitting to page 2
            # due to slight row-height rendering differences vs Linux.
            if extras_to_insert == 0:
                ws.page_setup.fitToHeight = 1

            # Fill rows
            from openpyxl.cell.cell import MergedCell
            for idx, item in enumerate(items):
                target_row = tr + idx
                for c in range(1, max_col + 1):
                    src = src_cells[c - 1]
                    cell = ws.cell(row=target_row, column=c)
                    if isinstance(cell, MergedCell):
                        continue  # skip merged secondary cells
                    val = src["value"]
                    if isinstance(val, str):
                        # Substitute item placeholders
                        def item_repl(m, item=item, idx=idx):
                            key = m.group(1).strip()
                            if key.startswith("items."):
                                subkey = key[len("items."):]
                                if subkey == "__index__":
                                    return str(idx + 1)
                                return "" if item.get(subkey) is None else str(item.get(subkey))
                            return m.group(0)
                        val = PLACEHOLDER_RE.sub(item_repl, val)
                        # Also substitute simple placeholders
                        val = _substitute_simple(val, data)
                    # Auto-renumber for NO column (index columns with bare integer template value)
                    if c in no_columns:
                        val = idx + 1
                    cell.value = val
                    if src["font"]: cell.font = copy(src["font"])
                    if src["fill"]: cell.fill = copy(src["fill"])
                    if src["border"]: cell.border = copy(src["border"])
                    if src["alignment"]: cell.alignment = copy(src["alignment"])
                    if src["number_format"]: cell.number_format = src["number_format"]
                    # Enable text wrap so long descriptions flow to next line instead of overflowing
                    _enable_wrap(cell)
            # Auto-renumber the NO column for TRAILING empty slots (rows after data). These
            # are pre-drawn slots from the template — we keep them visible but renumber so
            # the sequence is continuous (e.g. 1..15 for 6 data + 9 empty).
            if no_columns and total_slots > len(items):
                for idx in range(len(items), total_slots):
                    target_row = tr + idx
                    for c in no_columns:
                        cell = ws.cell(row=target_row, column=c)
                        if not isinstance(cell, MergedCell):
                            cell.value = idx + 1

            # Only apply explicit row height + merges to OVERFLOW-inserted rows (rows beyond
            # the original static slot count). Original static slots already carry their own
            # row_dimension + merged_cells from the template designer — don't override.
            row_h = template_row_height or 20.0
            first_overflow_idx = available  # rows [available..n_needed-1] are freshly inserted
            for idx in range(first_overflow_idx, n_needed):
                target_row = tr + idx
                ws.row_dimensions[target_row].height = row_h

            # Re-apply merged cell ranges only to overflow-inserted rows (openpyxl's insert_rows
            # does NOT propagate merges to newly-inserted rows; existing static slots already
            # have their merges intact from the template).
            if tr_merges and first_overflow_idx < n_needed:
                for idx in range(first_overflow_idx, n_needed):
                    tgt = tr + idx
                    for (c_start, c_end) in tr_merges:
                        if c_end > c_start:
                            try:
                                ws.merge_cells(start_row=tgt, end_row=tgt,
                                               start_column=c_start, end_column=c_end)
                            except Exception:
                                pass

        # Substitute simple placeholders in all remaining cells (skip already-processed table area)
        from openpyxl.cell.cell import MergedCell
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue  # cannot write; top-left holds the value
                v = cell.value
                if isinstance(v, str) and "{{" in v:
                    # Handle image placeholder
                    m_img = re.match(r"^\s*\{\{\s*IMAGE\s*:\s*([\w_]+)\s*\}\}\s*$", v)
                    if m_img:
                        img_key = m_img.group(1)
                        if img_key.lower() in ("company_logo", "logo"):
                            lp = _logo_path()
                            if lp and lp.exists():
                                try:
                                    # Convert webp to png via PIL if needed
                                    src = lp
                                    if lp.suffix.lower() == ".webp":
                                        tmp_png = Path(tempfile.gettempdir()) / f"logo_{uuid.uuid4().hex}.png"
                                        PILImage.open(lp).convert("RGBA").save(tmp_png, "PNG")
                                        src = tmp_png
                                    img = XLImage(str(src))
                                    # Fit into cell size roughly
                                    from openpyxl.utils import get_column_letter as gcl
                                    col_letter = gcl(cell.column)
                                    col_w = ws.column_dimensions[col_letter].width or 10
                                    row_h = ws.row_dimensions[cell.row].height or 15
                                    img.width = int(col_w * 7)
                                    img.height = int(row_h * 1.3)
                                    ws.add_image(img, cell.coordinate)
                                    cell.value = None
                                    continue
                                except Exception as e:
                                    logger.warning(f"Failed to insert logo image: {e}")
                        cell.value = ""
                        continue
                    cell.value = _substitute_simple(v, data)
                    _enable_wrap(cell)

        # AFTER all substitution/table expansion: compute actual content bounds
        # and set print_area — prevents empty rows/cols from causing extra blank pages.
        last_row = 1
        last_col = 1
        content_rows = set()
        # Track columns that have EITHER a value OR a visible border (table columns often
        # have empty body cells but borders that outline the printable area — e.g. the
        # "Inspection Date / Signature" column of MCL).
        for r in range(1, ws.max_row + 1):
            row_has = False
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                has_value = v is not None and str(v).strip()
                has_border = False
                if cell.has_style and cell.border:
                    for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom):
                        if side and side.style:
                            has_border = True
                            break
                if has_value or has_border:
                    if r > last_row: last_row = r
                    if c > last_col: last_col = c
                if has_value:
                    row_has = True
            if row_has:
                content_rows.add(r)
        # Extend last_col to cover any merged cell that starts within our detected range.
        for mr in ws.merged_cells.ranges:
            if mr.min_col <= last_col and mr.max_col > last_col and mr.min_row <= last_row:
                last_col = mr.max_col
        if last_row > 1 and last_col > 0:
            ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"

        # NOTE: Do NOT hide empty gap rows. Template designers intentionally leave blank
        # rows between the table, signature block, and document register code footer to
        # create physical space for hand-written signatures when printed. Hiding those
        # rows collapses the layout and destroys the designer's intended A4 print layout.


def _find_soffice() -> Optional[str]:
    """Locate LibreOffice `soffice` binary across Windows, macOS, and Linux.

    Order:
    1. `SOFFICE_BIN` env variable (explicit user override — useful on Windows
       when LibreOffice is installed in a non-standard folder).
    2. `shutil.which()` — checks the system PATH.
    3. Common OS-specific fallback locations.
    """
    override = os.environ.get("SOFFICE_BIN")
    if override and Path(override).exists():
        return override

    # PATH lookup — cross-platform, works everywhere
    for name in ("soffice", "soffice.exe", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found

    # OS-specific fallbacks
    candidates = []
    if sys.platform.startswith("win"):
        # Windows: check Program Files variants
        pf_candidates = [
            os.environ.get("PROGRAMFILES", r"C:\Program Files"),
            os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"),
            os.environ.get("PROGRAMW6432", r"C:\Program Files"),
        ]
        for pf in pf_candidates:
            candidates.append(str(Path(pf) / "LibreOffice" / "program" / "soffice.exe"))
    elif sys.platform == "darwin":
        # macOS
        candidates.extend([
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        ])
    else:
        # Linux
        candidates.extend([
            "/usr/bin/soffice",
            "/usr/lib/libreoffice/program/soffice",
            "/usr/bin/libreoffice",
            "/snap/bin/libreoffice",
        ])

    for p in candidates:
        if p and Path(p).exists():
            return p
    return None



# ------------- LibreOffice PDF cache (Iter 20c) -------------
# Cache PDF hasil konversi keyed by sha256(xlsx). TTL 5 menit, max 32 entry (LRU-ish).
# Cut 3-8 detik LibreOffice call → <10ms cache hit untuk preview yang di-refresh berulang.
_PDF_CACHE: dict = {}
_PDF_CACHE_TTL = 300  # detik
_PDF_CACHE_MAX = 32


def _pdf_cache_get(key: str):
    entry = _PDF_CACHE.get(key)
    if not entry:
        return None
    if time.time() - entry["ts"] > _PDF_CACHE_TTL:
        _PDF_CACHE.pop(key, None)
        return None
    entry["ts"] = time.time()  # bump for LRU
    return entry["pdf"]


def _pdf_cache_set(key: str, pdf: bytes) -> None:
    if len(_PDF_CACHE) >= _PDF_CACHE_MAX:
        oldest = min(_PDF_CACHE.items(), key=lambda kv: kv[1]["ts"])[0]
        _PDF_CACHE.pop(oldest, None)
    _PDF_CACHE[key] = {"pdf": pdf, "ts": time.time()}


def _xlsx_to_pdf(xlsx_bytes: bytes) -> bytes:
    """Convert xlsx bytes to PDF bytes via LibreOffice headless.

    Iter 20c — cache PDF hasil LibreOffice keyed by sha256(xlsx_bytes). TTL 5 menit.
    Untuk data yang sama (mis. user preview drawing yang sama berulang) → hit cache
    instant tanpa panggil LibreOffice (yang butuh 3-8 detik).
    """
    # Check cache first
    cache_key = hashlib.sha256(xlsx_bytes).hexdigest()
    cached = _pdf_cache_get(cache_key)
    if cached is not None:
        return cached

    soffice = _find_soffice()

    # Linux-only: try apt-get auto-install as a last resort (only when running as root
    # inside a Debian-based container — skip on Windows/macOS to avoid noisy errors).
    if not soffice and sys.platform.startswith("linux"):
        try:
            subprocess.run(
                ["apt-get", "install", "-y", "libreoffice-calc", "--no-install-recommends"],
                capture_output=True, timeout=180,
            )
            soffice = _find_soffice()
        except Exception as e:
            logger.error(f"LibreOffice auto-install failed: {e}")

    if not soffice:
        raise HTTPException(
            status_code=500,
            detail=(
                "LibreOffice (soffice) tidak tersedia untuk konversi PDF. "
                "Install LibreOffice, atau set env SOFFICE_BIN ke path soffice.exe (Windows) / soffice (Linux/macOS)."
            ),
        )

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        src = tmp_path / f"in_{uuid.uuid4().hex}.xlsx"
        src.write_bytes(xlsx_bytes)
        # HOME override prevents LibreOffice from polluting the user profile & lets
        # concurrent conversions run without lockfile clashes (works on all OS).
        env = {**os.environ, "HOME": str(tmp_path), "USERPROFILE": str(tmp_path)}
        proc = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(tmp_path), str(src)],
            capture_output=True, timeout=120, env=env,
        )
        if proc.returncode != 0:
            raise HTTPException(
                status_code=500,
                detail=f"LibreOffice convert error: {proc.stderr.decode(errors='ignore')[:400]}",
            )
        pdf = src.with_suffix(".pdf")
        if not pdf.exists():
            raise HTTPException(status_code=500, detail="PDF tidak dihasilkan oleh LibreOffice")
        result = pdf.read_bytes()
        _pdf_cache_set(cache_key, result)
        return result


def _replace_placeholders_in_drawings(xlsx_bytes: bytes, data: dict) -> bytes:
    """Post-process xlsx zip: replace {{key}} placeholders inside drawings/shapes/textboxes XML.

    openpyxl tidak selalu preserve text di dalam Shape/TextBox saat load-save,
    jadi kita edit XML mentah setelah save untuk pastikan placeholder di dalam
    shape (mis. kolom tanda tangan Prepared By/Checked By) tetap ter-fill.
    """
    import zipfile
    import re as _re

    def _sub_all(text: str) -> str:
        def _repl(m):
            key = m.group(1).strip()
            if key.startswith("items."):
                return m.group(0)  # skip table placeholders
            if key == "IMAGE" or key.startswith("IMAGE:"):
                return ""  # image markers not applicable in shapes
            val = data.get(key)
            if val is None:
                return ""
            return str(val)
        return _re.sub(r"\{\{\s*([\w_.:]+?)\s*\}\}", _repl, text)

    src = io.BytesIO(xlsx_bytes)
    dst = io.BytesIO()

    with zipfile.ZipFile(src, "r") as zin:
        with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                # Process drawings & shared strings — shapes text often lives here
                if item.filename.startswith("xl/drawings/") and item.filename.endswith(".xml"):
                    try:
                        text = content.decode("utf-8")
                        if "{{" in text:
                            # Placeholder might be split across multiple <a:t> runs in Excel.
                            # Merge consecutive <a:t>...</a:t> runs within same paragraph first.
                            # Simple approach: strip inter-run tags between {{ and }}.
                            text = _merge_split_placeholders(text)
                            text = _sub_all(text)
                        content = text.encode("utf-8")
                    except Exception:
                        pass
                zout.writestr(item, content)
    return dst.getvalue()


def _merge_split_placeholders(xml_text: str) -> str:
    """Excel sometimes splits a placeholder like `{{printed_by}}` across multiple <a:r>/<a:t> runs
    (due to autocorrect / formatting change). Merge runs so the placeholder becomes a single
    contiguous string before regex substitution.
    """
    import re as _re
    # Pattern: `<a:t...>text1</a:t>[optional whitespace + closing/opening run tags]<a:t...>text2</a:t>`
    # We merge whenever the two texts joined contain `{{` and `}}` split between them (or a `{`/`}` boundary).
    pattern = _re.compile(
        r"(<a:t[^>]*>)([^<]*)(</a:t>)(\s*(?:</a:r>\s*<a:r[^>]*>(?:<a:rPr[^/]*/>|<a:rPr[^>]*>[^<]*</a:rPr>)?)?)(<a:t[^>]*>)([^<]*)(</a:t>)",
        _re.DOTALL,
    )
    prev = None
    txt = xml_text
    max_iter = 20
    while prev != txt and max_iter > 0:
        prev = txt

        def _maybe_merge(m):
            t_open1, t1, t_close1, middle, t_open2, t2, t_close2 = m.groups()
            open_in_t1 = "{{" in t1
            close_in_t1 = "}}" in t1
            open_in_t2 = "{{" in t2
            close_in_t2 = "}}" in t2
            # t1 punya open tanpa close → placeholder mulai di t1 tapi belum tutup → MERGE
            # atau t2 punya close tanpa open → placeholder tutup di t2 tapi mulai sebelumnya → MERGE
            # Boundary: '{' + '{' = '{{', '}' + '}' = '}}'
            boundary_open = t1.endswith("{") and t2.startswith("{")
            boundary_close = t1.endswith("}") and t2.startswith("}")
            should_merge = (
                (open_in_t1 and not close_in_t1)
                or (close_in_t2 and not open_in_t2)
                or boundary_open
                or boundary_close
            )
            if should_merge:
                return f"{t_open1}{t1}{t2}{t_close2}"
            return m.group(0)

        txt = pattern.sub(_maybe_merge, txt)
        max_iter -= 1
    return txt


def render_excel_template(xlsx_bytes: bytes, data: dict, as_pdf: bool = True) -> bytes:
    """Public API: substitute + render. Returns pdf bytes if as_pdf else xlsx bytes."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    _apply_substitution(wb, data)
    out = io.BytesIO()
    wb.save(out)
    xlsx_out = out.getvalue()
    # Additional pass: replace placeholders inside Shape/TextBox XML (drawings)
    try:
        xlsx_out = _replace_placeholders_in_drawings(xlsx_out, data)
    except Exception as e:
        logger.warning(f"drawing placeholder replace failed: {e}")
    if not as_pdf:
        return xlsx_out
    return _xlsx_to_pdf(xlsx_out)


def _duplicate_sheet_content(wb) -> None:
    """Duplicate all content rows in place, creating 2 identical copies on the same sheet.

    - Small forms (content total height <= ~half A4): both copies fit on 1 A4 sheet.
    - Larger forms: LibreOffice will naturally paginate — 1st copy on page 1, 2nd on page 2.
    A blank spacer row (~30pt) is inserted between the copies. Merged ranges, row heights,
    and cell styles are all preserved. Print area is extended to cover both copies.
    """
    from copy import copy as _copy
    from openpyxl.utils import get_column_letter
    for ws in wb.worksheets:
        # Snapshot: find last row/col with content
        max_r, max_c = 1, 1
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip():
                    if r > max_r: max_r = r
                    if c > max_c: max_c = c
        if max_r < 3:
            continue
        # Compute total content height to decide if we insert a page break
        total_h = 0.0
        for r in range(1, max_r + 1):
            h = ws.row_dimensions[r].height if r in ws.row_dimensions else 15.0
            total_h += float(h or 15.0)
        # A4 printable height ~ 800pt at default margins
        half_a4 = 400.0
        full_a4_printable = 780.0
        needs_new_page = total_h > half_a4

        spacer_rows = 1
        base_row = max_r  # last row of copy #1
        offset = base_row + spacer_rows  # copy #2 starts at row (base_row + spacer_rows + 1)

        # If content is large, push copy #2 to start of the next page by making the spacer
        # row tall enough. Add generous headroom so LibreOffice's slight row-height
        # variance doesn't cause bleed into the wrong page.
        if needs_new_page:
            spacer_h = max(150.0, full_a4_printable - total_h + 120.0)
        else:
            spacer_h = 30.0
        ws.row_dimensions[base_row + 1].height = spacer_h

        # Snapshot all cell data + styles for rows 1..max_r
        snapshot = []
        for r in range(1, max_r + 1):
            row_snap = {"h": ws.row_dimensions[r].height if r in ws.row_dimensions else None, "cells": []}
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                row_snap["cells"].append({
                    "value": cell.value,
                    "font": _copy(cell.font) if cell.has_style else None,
                    "fill": _copy(cell.fill) if cell.has_style else None,
                    "border": _copy(cell.border) if cell.has_style else None,
                    "alignment": _copy(cell.alignment) if cell.has_style else None,
                    "number_format": cell.number_format,
                })
            snapshot.append(row_snap)
        # Snapshot merged ranges within the content
        merges_snap = []
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row <= max_r and mr.max_row <= max_r and mr.max_col <= max_c:
                merges_snap.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))

        # Extend sheet: write copy #2 starting at row `offset + 1`
        for idx, row_snap in enumerate(snapshot):
            target_r = offset + 1 + idx
            if row_snap["h"]:
                ws.row_dimensions[target_r].height = row_snap["h"]
            for c_idx, cs in enumerate(row_snap["cells"]):
                cell = ws.cell(row=target_r, column=c_idx + 1)
                cell.value = cs["value"]
                if cs["font"]: cell.font = _copy(cs["font"])
                if cs["fill"]: cell.fill = _copy(cs["fill"])
                if cs["border"]: cell.border = _copy(cs["border"])
                if cs["alignment"]: cell.alignment = _copy(cs["alignment"])
                if cs["number_format"]: cell.number_format = cs["number_format"]
        # Re-apply merged ranges shifted
        for (r1, r2, c1, c2) in merges_snap:
            try:
                ws.merge_cells(start_row=r1 + offset, end_row=r2 + offset,
                               start_column=c1, end_column=c2)
            except Exception:
                pass

        # Extend print area to cover both copies
        new_last_row = offset + max_r
        ws.print_area = f"A1:{get_column_letter(max_c)}{new_last_row}"


def _page_break(row_idx: int):
    """Create a horizontal page break object for openpyxl."""
    from openpyxl.worksheet.pagebreak import Break
    return Break(id=row_idx)


def render_excel_template_two_copies(xlsx_bytes: bytes, data: dict, as_pdf: bool = True) -> bytes:
    """Produce 2 identical copies (for MIF, since Store & Production Dept each keep 1 signed
    rangkap).

    Strategy:
    - Render normally → get single-copy PDF or xlsx.
    - For PDF: if content fits on 1 page AND is ≤ half A4, stack 2 copies on 1 A4 sheet
      via in-sheet duplication. Otherwise concatenate the same PDF twice (each copy fills
      its own dedicated page).
    - For XLSX: always duplicate rows in place (small forms stack, big forms overflow).
    """
    # Detect content size first (small vs large) using a lightweight probe render
    wb_probe = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    _apply_substitution(wb_probe, data)
    ws = wb_probe.active if wb_probe.worksheets else None
    total_h = 0.0
    max_r = 1
    if ws:
        for r in range(1, ws.max_row + 1):
            row_has_content = any(
                (ws.cell(row=r, column=c).value is not None
                 and str(ws.cell(row=r, column=c).value).strip())
                for c in range(1, ws.max_column + 1)
            )
            if row_has_content and r > max_r:
                max_r = r
        for r in range(1, max_r + 1):
            h = ws.row_dimensions[r].height if r in ws.row_dimensions else 15.0
            total_h += float(h or 15.0)
    small_form = total_h <= 400.0  # ≤ half A4

    if not as_pdf:
        # XLSX output: stack in sheet regardless
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        _apply_substitution(wb, data)
        _duplicate_sheet_content(wb)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    if small_form:
        # Small: use in-sheet stacking (spacer stays small so both fit on same A4)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        _apply_substitution(wb, data)
        _duplicate_sheet_content(wb)
        out = io.BytesIO()
        wb.save(out)
        return _xlsx_to_pdf(out.getvalue())

    # Large: render single copy PDF, then concatenate the same PDF twice (each on own page).
    single_pdf = render_excel_template(xlsx_bytes, data, as_pdf=True)
    try:
        from pypdf import PdfReader, PdfWriter
        writer = PdfWriter()
        for _ in range(2):
            reader = PdfReader(io.BytesIO(single_pdf))
            for page in reader.pages:
                writer.add_page(page)
        buf = io.BytesIO()
        writer.write(buf)
        return buf.getvalue()
    except ImportError:
        # pypdf unavailable → fallback to single copy
        return single_pdf


async def get_active_xlsx_bytes(code: str) -> Optional[bytes]:
    doc = await db[COLLECTION].find_one(
        merged({"code": code.upper(), "is_active": True}, NOT_DELETED_FILTER),
        sort=[("updated_at", -1)],
    )
    if not doc:
        return None
    return base64.b64decode(doc["xlsx_base64"])


# ---------------- Preview ----------------
@router.post("/excel-templates/{tid}/preview-raw")
async def preview_raw_excel_template(tid: str, current: dict = Depends(get_current_user)):
    """Convert the raw uploaded xlsx directly to PDF (no substitution).
    Berguna untuk melihat layout Excel Anda apakah sudah benar sebelum data disubstitusi."""
    doc = await db[COLLECTION].find_one(merged({"id": tid}, NOT_DELETED_FILTER))
    if not doc:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    xlsx_bytes = base64.b64decode(doc["xlsx_base64"])
    pdf = _xlsx_to_pdf(xlsx_bytes)
    fname = f"preview_RAW_{doc.get('code','MCL')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"', "Cache-Control": "no-store"},
    )


@router.post("/excel-templates/{tid}/preview-xlsx")
async def preview_xlsx_substituted(tid: str, current: dict = Depends(get_current_user)):
    """Return the SUBSTITUTED xlsx as .xlsx (so user can open in Excel & verify layout)."""
    doc = await db[COLLECTION].find_one(merged({"id": tid}, NOT_DELETED_FILTER))
    if not doc:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    xlsx_bytes = base64.b64decode(doc["xlsx_base64"])
    sample = _sample_data(current)
    xlsx_sub = render_excel_template(xlsx_bytes, sample, as_pdf=False)
    code = doc.get("code", "MCL")
    fname = f"preview_DATA_{code}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_sub),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"', "Cache-Control": "no-store"},
    )


def _sample_data(current):
    return {
        "company_name": "PT. MITRA KARYA SARANA",
        "receive_date": "2026-02-08",
        "vendor_name": "PT VENDOR CONTOH JAYA",
        "po_no": "PO-2026-001",
        "do_number": "SJ-2026-011",
        "invoice_no": "INV-2026-99",
        "print_date": _now_iso()[:10],
        "printed_by": current.get("username", ""),
        "items": [
            {"__index__": 1, "so_no": "SO-4097", "item_name": "Industrial Bearing SKF 6205", "qty_received": 10, "unit": "Pcs", "receive_date": "2026-02-08"},
            {"__index__": 2, "so_no": "SO-4097", "item_name": "Sample Item 2 dengan nama panjang", "qty_received": 25, "unit": "Meter", "receive_date": "2026-02-08"},
            {"__index__": 3, "so_no": "SO-4098", "item_name": "Contoh Item 3", "qty_received": 3, "unit": "Set", "receive_date": "2026-02-08"},
        ],
    }


@router.post("/excel-templates/{tid}/preview")
async def preview_excel_template(tid: str, current: dict = Depends(get_current_user)):
    """Render Excel template with sample data → PDF."""
    doc = await db[COLLECTION].find_one(merged({"id": tid}, NOT_DELETED_FILTER))
    if not doc:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    xlsx_bytes = base64.b64decode(doc["xlsx_base64"])
    sample = _sample_data(current)
    pdf = render_excel_template(xlsx_bytes, sample, as_pdf=True)
    fname = f"preview_{doc.get('code','MCL')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"', "Cache-Control": "no-store"},
    )


@router.post("/excel-templates/{tid}/auto-placeholder")
async def auto_placeholder(tid: str, current: dict = Depends(get_current_user)):
    """Analyze the active Excel template and inject placeholders automatically
    based on common label patterns. Returns the modified xlsx as the new active template."""
    _require_admin(current)
    doc = await db[COLLECTION].find_one(merged({"id": tid}, NOT_DELETED_FILTER))
    if not doc:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    code = doc["code"]
    xlsx_bytes = base64.b64decode(doc["xlsx_base64"])

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    injections = []  # for reporting

    # Label patterns → placeholder (matches label cell, placeholder goes to next non-empty cell OR adjacent)
    LABEL_MAP = [
        (re.compile(r"^\s*(no\.?\s*do|d\.?o\.?\s*no\.?|do\s*no\.?|no\.?\s*surat\s*jalan|s\.?j\.?\s*no)\b", re.I), "do_number"),
        (re.compile(r"^\s*(no\.?\s*po|p\.?o\.?\s*no\.?|po\s*no\.?)\b", re.I), "po_no"),
        (re.compile(r"^\s*(no\.?\s*invoice|invoice\s*no\.?|inv\s*no\.?)\b", re.I), "invoice_no"),
        (re.compile(r"^\s*(form\s*no\.?|no\.?\s*form|nomor\s*form)\b", re.I), "form_no"),
        (re.compile(r"^\s*(supplier|customer|vendor|nama\s*vendor|nama\s*supplier|customer\s*/\s*supplier)\s*[/:]?", re.I), "vendor_name"),
        (re.compile(r"^\s*(received\s*date|receive\s*date|tgl\.?\s*terima|tanggal\s*terima|date|tanggal|issue\s*date)\s*[:]?", re.I), "receive_date"),
        (re.compile(r"^\s*(printed\s*by|dicetak\s*oleh|checked\s*by|admin\s*store|issued\s*by)\s*[:]?", re.I), "printed_by"),
        (re.compile(r"^\s*(pt\.?\s*mitra\s*karya\s*sarana|company)\s*[:]?", re.I), "company_name"),
    ]

    # Table header patterns → column binding
    HEADER_MAP = [
        (re.compile(r"^\s*(no|no\.|#|index|nomor)\s*$", re.I), "__index__"),
        (re.compile(r"^\s*(so\s*no|so\s*no\.|so|so\s*number|nomor\s*so)\s*$", re.I), "so_no"),
        (re.compile(r"^\s*(material\s*description|nama\s*barang|description|descriptions|item\s*name|barang|material|deskripsi)\s*$", re.I), "item_name"),
        (re.compile(r"^\s*(qty|quantity|jumlah)\s*$", re.I), "qty_received"),
        (re.compile(r"^\s*(unit|satuan|uom)\s*$", re.I), "unit"),
        (re.compile(r"^\s*(received\s*date|receive\s*date|tgl\.?\s*terima|tanggal)\s*$", re.I), "receive_date"),
        (re.compile(r"^\s*(remark|remarks|keterangan|note|notes|catatan)\s*$", re.I), "remark"),
    ]

    def is_placeholder(v):
        return isinstance(v, str) and PLACEHOLDER_RE.search(v)

    def enable_wrap(cell):
        """Enable wrap_text on cell alignment (preserve other alignment props)."""
        if cell is None:
            return
        try:
            from openpyxl.styles import Alignment
            al = cell.alignment
            cell.alignment = Alignment(
                horizontal=al.horizontal, vertical=al.vertical or "top",
                text_rotation=al.text_rotation, wrap_text=True,
                shrink_to_fit=al.shrink_to_fit, indent=al.indent,
            )
        except Exception:
            pass

    def writable_cell(ws, cell):
        """Return top-left cell of merged range if cell is MergedCell, else cell itself.
        Return None if not writable."""
        from openpyxl.cell.cell import MergedCell
        if not isinstance(cell, MergedCell):
            return cell
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                return ws.cell(row=mr.min_row, column=mr.min_col)
        return None

    def get_merge_range(ws, cell):
        """Return the merged range containing this cell, or None."""
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                return mr
        return None

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        # 0. FIRST: detect table header row(s) so label scan can skip them
        table_header_rows = set()
        for r in range(1, max_row + 1):
            hits_check = 0
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip():
                    for pat, _ in HEADER_MAP:
                        if pat.match(v.strip()):
                            hits_check += 1
                            break
            if hits_check >= 2:
                table_header_rows.add(r)

        # 1. Label detection (single-value fields) — SKIP table header rows
        for r in range(1, max_row + 1):
            if r in table_header_rows:
                continue
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if not isinstance(v, str) or not v.strip():
                    continue
                if is_placeholder(v):
                    continue
                text = v.strip()
                for pat, key in LABEL_MAP:
                    if pat.search(text):
                        # Determine the label's merged range (so we skip cells within it)
                        label_mr = get_merge_range(ws, cell)
                        skip_after_col = label_mr.max_col if label_mr else c
                        # find placement cell: next non-empty cell to the right OR next cell right if empty
                        target = None
                        for c2 in range(skip_after_col + 1, min(max_col + 1, skip_after_col + 8)):
                            tc = ws.cell(row=r, column=c2)
                            tv = tc.value
                            # Skip cells that are part of the same label merge range (shouldn't happen but safety)
                            if label_mr and tc.coordinate in label_mr:
                                continue
                            if tv is None or (isinstance(tv, str) and tv.strip() == ""):
                                target = tc
                                break
                            if isinstance(tv, str) and tv.strip() in (":", "|", "="):
                                continue  # separator — keep looking
                            if isinstance(tv, str) and is_placeholder(tv):
                                break  # already placeholder — do nothing
                            # Non-empty non-separator non-placeholder — treat as target
                            target = tc
                            break
                        if target is not None:
                            wt = writable_cell(ws, target)
                            # Verify writable cell is NOT the label cell itself
                            if wt is not None and wt.coordinate != cell.coordinate:
                                # Also verify wt.value is not the label text
                                if not (isinstance(wt.value, str) and pat.search(wt.value)):
                                    if wt.value in (None, "") or (isinstance(wt.value, str) and not is_placeholder(wt.value)):
                                        wt.value = f"{{{{{key}}}}}"
                                        injections.append(f"{ws.title}!{wt.coordinate} = {{{{{key}}}}} (label: {text[:30]})")
                        break

        # 2. Table detection: fill row after each header row
        for hr in sorted(table_header_rows):
            hits = []
            for c in range(1, max_col + 1):
                v = ws.cell(row=hr, column=c).value
                if isinstance(v, str) and v.strip():
                    for pat, key in HEADER_MAP:
                        if pat.match(v.strip()):
                            hits.append((c, key))
                            break
            target_row = hr + 1
            if target_row <= max_row:
                for col, key in hits:
                    tc = ws.cell(row=target_row, column=col)
                    wt = writable_cell(ws, tc)
                    if wt is not None and (wt.value in (None, "") or (isinstance(wt.value, str) and not is_placeholder(wt.value))):
                        wt.value = f"{{{{items.{key}}}}}"
                        injections.append(f"{ws.title}!{wt.coordinate} = {{{{items.{key}}}}} (table col: {key})")
            break  # only process first table for now

    out = io.BytesIO()
    wb.save(out)
    new_bytes = out.getvalue()

    # Save as NEW active template (deactivate old ones for this code)
    await db[COLLECTION].update_many({"code": code}, {"$set": {"is_active": False}})
    new_doc = {
        "id": str(uuid.uuid4()),
        "code": code,
        "filename": (doc.get("filename") or f"{code}.xlsx").replace(".xlsx", "_AUTO.xlsx"),
        "xlsx_base64": base64.b64encode(new_bytes).decode("ascii"),
        "size_bytes": len(new_bytes),
        "is_active": True,
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
        "updated_by": current.get("username"),
    }
    await db[COLLECTION].insert_one(new_doc.copy())
    await log_action(current, "auto_placeholder", "excel_form_template", new_doc["id"],
                     {"code": code, "injected": len(injections)})
    return {"ok": True, "injected": len(injections), "details": injections[:30], "new_id": new_doc["id"]}


@router.get("/excel-templates/starter/{code}")
async def download_starter(code: str, current: dict = Depends(get_current_user)):
    """Return a starter Excel with placeholders + header cheatsheet."""
    code = code.upper()
    meta = FORM_CODES.get(code)
    if not meta:
        raise HTTPException(status_code=404, detail="Kode form tidak dikenali")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = code

    # Cheatsheet sheet
    cs = wb.create_sheet("Panduan_Placeholder")
    cs.column_dimensions["A"].width = 32
    cs.column_dimensions["B"].width = 60
    cs["A1"] = "PLACEHOLDER"
    cs["B1"] = "KETERANGAN"
    from openpyxl.styles import Font, PatternFill
    for col in ("A", "B"):
        cs[f"{col}1"].font = Font(bold=True, color="FFFFFF")
        cs[f"{col}1"].fill = PatternFill("solid", fgColor="1F4E79")
    r = 2
    for f in meta["fields"]:
        cs.cell(row=r, column=1).value = f"{{{{{f}}}}}"
        cs.cell(row=r, column=2).value = f"Field header — {f}"
        r += 1
    cs.cell(row=r, column=1).value = "{{IMAGE:company_logo}}"
    cs.cell(row=r, column=2).value = "Sisipkan logo perusahaan (fit ke cell)"
    r += 2
    cs.cell(row=r, column=1).value = "-- BARIS TABEL --"
    cs.cell(row=r, column=1).font = Font(bold=True, italic=True)
    r += 1
    for f in meta["table_fields"]:
        cs.cell(row=r, column=1).value = f"{{{{items.{f}}}}}"
        cs.cell(row=r, column=2).value = f"Kolom tabel — {f}" + (" (auto nomor 1,2,3...)" if f == "__index__" else "")
        r += 1
    r += 1
    cs.cell(row=r, column=1).value = "CATATAN:"
    cs.cell(row=r, column=1).font = Font(bold=True)
    cs.cell(row=r+1, column=1).value = "Cara pakai tabel:"
    cs.cell(row=r+2, column=1).value = "1. Buat 1 baris sebagai template dengan sel-sel berisi {{items.item_name}} dst."
    cs.cell(row=r+3, column=1).value = "2. Sistem akan menduplikasi baris ini otomatis untuk setiap item."

    # ============ Sample layout on main sheet — code-specific ============
    from openpyxl.styles import Alignment, Border, Side
    thin = Side(border_style="thin", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    if code == "MII":
        # Landscape A4, 12 columns matching MKS-F-QAD-002 REV 03
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        widths = [4, 8, 14, 12, 22, 5, 8, 8, 10, 5, 5, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 45

        # Row 1-2: Logo | Company name box | (blank right)
        ws.merge_cells("A1:B2")
        ws["A1"] = "{{IMAGE:company_logo}}"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("C1:H1")
        ws["C1"] = "{{company_name}}"
        ws["C1"].font = Font(bold=True, size=12)
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].border = box
        ws.merge_cells("C2:H2")
        ws["C2"] = "MATERIAL INCOMING INSPECTION"
        ws["C2"].font = Font(bold=True, size=14)
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

        # Header rows: Supplier checkbox / Customer checkbox / DO / Date
        ws["A4"] = "☐ Supplier Name:"; ws["A4"].font = Font(bold=True, size=9)
        ws.merge_cells("B4:F4"); ws["B4"] = "{{supplier_name}}"
        ws["B4"].border = Border(bottom=thin)
        ws["G4"] = "DO. No.:"; ws["G4"].font = Font(bold=True, size=9)
        ws.merge_cells("H4:L4"); ws["H4"] = "{{do_no}}"
        ws["H4"].border = Border(bottom=thin)

        ws["A5"] = "☐ Supplied by Customer:"; ws["A5"].font = Font(bold=True, size=9)
        ws.merge_cells("B5:F5"); ws["B5"] = "{{customer_name}}"
        ws["B5"].border = Border(bottom=thin)
        ws["G5"] = "Date:"; ws["G5"].font = Font(bold=True, size=9)
        ws.merge_cells("H5:L5"); ws["H5"] = "{{inspection_date}}"
        ws["H5"].border = Border(bottom=thin)

        # Table header: row 7-9 (3-row header with merges)
        # Row 7: main headers (some merged down 3 rows)
        header_style = {"font": Font(bold=True, size=8.5), "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True), "border": box, "fill": PatternFill("solid", fgColor="E5E7EB")}

        def _apply(cell, style):
            for k, v in style.items():
                setattr(cell, k, v)

        # rowspan=3 columns: NO(A), SO NO(B), BATCH(C), MILL(D), DESC(E), QTY(F), REMARK(L)
        # colspan=3 (row 7): IQC INSPECTION RESULT covers G-I
        # colspan=2 (row 7): RESULT covers J-K
        # row 8: DIMENTION covers G-H, VISUAL rowspan(2) = I, OK rowspan(2) = J, NG rowspan(2) = K
        # row 9: SPEC = G, ACTUAL = H

        # rowspan=3 headers
        for col, label in [("A", "NO."), ("B", "SO. NO."), ("C", "BATCH No.#/GRADE MAT'L/Heat No.#"),
                           ("D", "MILL CERT/ EDS NO."), ("E", "DESCRIPTION OF PART"),
                           ("F", "QTY"), ("L", "REMARK")]:
            ws.merge_cells(f"{col}7:{col}9")
            _apply(ws[f"{col}7"], header_style)
            ws[f"{col}7"] = label

        # IQC INSPECTION RESULT (G7:I7)
        ws.merge_cells("G7:I7")
        _apply(ws["G7"], header_style)
        ws["G7"] = "IQC INSPECTION RESULT"

        # RESULT (J7:K7)
        ws.merge_cells("J7:K7")
        _apply(ws["J7"], header_style)
        ws["J7"] = "RESULT"

        # DIMENTION (G8:H8)
        ws.merge_cells("G8:H8")
        _apply(ws["G8"], header_style)
        ws["G8"] = "DIMENTION"

        # VISUAL rowspan(2) I8:I9
        ws.merge_cells("I8:I9")
        _apply(ws["I8"], header_style)
        ws["I8"] = "VISUAL"

        # OK rowspan(2) J8:J9
        ws.merge_cells("J8:J9")
        _apply(ws["J8"], header_style)
        ws["J8"] = "OK"

        # NG rowspan(2) K8:K9
        ws.merge_cells("K8:K9")
        _apply(ws["K8"], header_style)
        ws["K8"] = "NG"

        # SPEC (G9), ACTUAL (H9)
        for col, lbl in [("G", "SPEC"), ("H", "ACTUAL")]:
            _apply(ws[f"{col}9"], header_style)
            ws[f"{col}9"] = lbl

        # Template row (row 10) — will be duplicated per item
        row_tpl = [
            "{{items.__index__}}", "{{items.so_no}}", "{{items.batch_grade_heat}}",
            "{{items.mill_cert_no}}", "{{items.description}}", "{{items.qty}}",
            "{{items.dimension_spec}}", "{{items.dimension_actual}}", "{{items.visual}}",
            "{{items.result_ok}}", "{{items.result_ng}}", "{{items.remark}}",
        ]
        for i, v in enumerate(row_tpl, 1):
            c = ws.cell(row=10, column=i, value=v)
            c.border = box
            c.alignment = Alignment(horizontal="center" if i in (1, 6, 7, 8, 10, 11) else "left",
                                    vertical="center", wrap_text=True)
            c.font = Font(size=8)

        # Note row
        ws.merge_cells("A12:L12")
        ws["A12"] = "Note : Visual = Check of Appearance (Dent, Damage, Scratch, Colour)"
        ws["A12"].font = Font(italic=True, size=8)

        # Signatures
        ws["D15"] = "Inspected by,"; ws["D15"].font = Font(size=9)
        ws.merge_cells("D16:F16"); ws["D16"] = "{{inspector_name}}"; ws["D16"].font = Font(bold=True)
        ws.merge_cells("D17:F17"); ws["D17"] = "QC Inspector"; ws["D17"].font = Font(size=8)

        ws["I15"] = "Verified by,"; ws["I15"].font = Font(size=9)
        ws.merge_cells("I16:L16"); ws["I16"] = "{{leader_name}}"; ws["I16"].font = Font(bold=True)
        ws.merge_cells("I17:L17"); ws["I17"] = "QC Leader"; ws["I17"].font = Font(size=8)

        # Doc code
        ws["A19"] = "MKS-F-QAD-002 REV 03"; ws["A19"].font = Font(size=7, color="64748B")
    elif code == "BOM":
        # Landscape A4, 11 columns matching sample BOM042 template
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        widths = [5, 22, 26, 6, 6, 9, 9, 11, 11, 15, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 40

        # Row 1: Logo + company name (edit sesuka Anda)
        ws.merge_cells("A1:B2")
        ws["A1"] = "{{IMAGE:company_logo}}"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = box
        ws.merge_cells("C1:K1")
        ws["C1"] = "{{company_name}}"
        ws["C1"].font = Font(bold=True, size=14)
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].border = box
        ws.merge_cells("C2:K2")
        ws["C2"] = "BILL OF MATERIAL (BOM)"
        ws["C2"].font = Font(bold=True, size=16)
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C2"].border = box

        # Row 4-5: TO / BOM.NO. / DATE / REV.NO.
        ws["A4"] = "TO"; ws["A4"].font = Font(bold=True)
        ws["B4"] = ": PUR"
        ws["H4"] = "BOM.NO."; ws["H4"].font = Font(bold=True)
        ws.merge_cells("I4:K4"); ws["I4"] = ": {{bom_no}}"

        ws["A5"] = "DATE"; ws["A5"].font = Font(bold=True)
        ws["B5"] = ": {{approved_date}}"
        ws["H5"] = "REV.NO."; ws["H5"].font = Font(bold=True)
        ws.merge_cells("I5:K5"); ws["I5"] = ": {{revision}}"

        # Row 7: Meta labels (blue fill)
        blue_fill = PatternFill("solid", fgColor="D9E1F2")
        labels = [("A7", "PROJECT :"), ("C7", "ENG. DRW. PROJECT NO. :"), ("E7", "CUSTOMER :"),
                  ("G7", "CLASS OF MATERIAL :"), ("I7", "PT MKS SO.NO. :"), ("J7", "DELIVERY DATE :")]
        merges = [("A7:B7"), ("C7:D7"), ("E7:F7"), ("G7:H7"), ("I7:I7"), ("J7:K7")]
        for i, (cell_ref, text) in enumerate(labels):
            m = merges[i]
            if ":" in m and m.split(":")[0] != m.split(":")[1]:
                ws.merge_cells(m)
            ws[cell_ref] = text
            ws[cell_ref].font = Font(bold=True, size=9)
            ws[cell_ref].fill = blue_fill
            ws[cell_ref].border = box
            ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Row 8: Values
        vals = [("A8", "{{project_name}}", "A8:B8"), ("C8", "{{drawing_no}}", "C8:D8"),
                ("E8", "{{customer}}", "E8:F8"), ("G8", "{{class_material}}", "G8:H8"),
                ("I8", "{{so_no}}", "I8:I8"), ("J8", "{{delivery_date}}", "J8:K8")]
        for cell_ref, text, mrg in vals:
            if ":" in mrg and mrg.split(":")[0] != mrg.split(":")[1]:
                ws.merge_cells(mrg)
            ws[cell_ref] = text
            ws[cell_ref].border = box
            ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[8].height = 40

        # Row 9: Table header (11 columns)
        headers = ["ITEM NO", "ITEM NAME", "ITEM SPECIFICATION", "QTY.", "UOM",
                   "MATERIAL", "WEIGHT (KG)", "AVAILABLE STOCK", "QTY PURCHASE",
                   "PURCHASE DUE DATE", "REMARK"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=9, column=i, value=h)
            c.font = Font(bold=True, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(left=Side("medium"), right=Side("medium"),
                              top=Side("medium"), bottom=Side("medium"))
            c.fill = blue_fill
        ws.row_dimensions[9].height = 30

        # Row 10: Template baris item (akan diduplikasi per item saat render)
        tpl = ["{{items.__index__}}", "{{items.item_name}}", "{{items.specification}}",
               "{{items.qty}}", "{{items.uom}}", "{{items.material}}", "{{items.weight_kg}}",
               "{{items.available_stock}}", "{{items.qty_purchase}}", "{{items.purchase_due_date}}",
               "{{items.remark}}"]
        for i, v in enumerate(tpl, 1):
            c = ws.cell(row=10, column=i, value=v)
            c.border = box
            c.font = Font(size=9)
            if i in (1, 4, 5):
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[10].height = 25

        # Freeze pane + repeat header on each page (Print Titles)
        ws.freeze_panes = "A10"
        try:
            ws.print_title_rows = "1:9"
        except Exception:
            pass

        # Footer: NOTES + TOTAL WEIGHT (di baris 22)
        ws["A22"] = "NOTES :"; ws["A22"].font = Font(bold=True, size=9)
        ws.merge_cells("B22:F22"); ws["B22"] = "{{notes}}"
        ws["B22"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws["B22"].border = box

        ws["G22"] = "TOTAL WEIGHT:"; ws["G22"].font = Font(bold=True, size=9)
        ws.merge_cells("H22:I22"); ws["H22"] = "{{total_weight}}"
        ws["H22"].font = Font(bold=True, size=10)
        ws["H22"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H22"].border = box

        # Doc code footer
        ws["A24"] = "MKS-F-ENG-BOM-01"
        ws["A24"].font = Font(size=7, color="64748B")
    else:
        # Default MCL/MIF/other — 6-column portrait layout
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 36
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 12
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A1:B2")
        ws["A1"] = "{{IMAGE:company_logo}}"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("C1:F1")
        ws["C1"] = "{{company_name}}"
        ws["C1"].font = Font(bold=True, size=14)
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("C2:F2")
        ws["C2"] = meta["label"].upper()
        ws["C2"].font = Font(bold=True, size=11)
        ws["C2"].alignment = Alignment(horizontal="center")

        ws["A4"] = "Tgl Terima"; ws["A4"].font = Font(bold=True)
        ws["B4"] = "{{receive_date}}"
        ws["A5"] = "Vendor"; ws["A5"].font = Font(bold=True)
        ws.merge_cells("B5:C5"); ws["B5"] = "{{vendor_name}}"

        ws["D4"] = "PO No"; ws["D4"].font = Font(bold=True)
        ws.merge_cells("E4:F4"); ws["E4"] = "{{po_no}}"
        ws["D5"] = "DO/SJ"; ws["D5"].font = Font(bold=True)
        ws.merge_cells("E5:F5"); ws["E5"] = "{{do_number}}"

        # Header table
        hdr_row = 7
        headers = ["No", "SO No", "Nama Barang", "Qty", "Unit", "Tgl"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=hdr_row, column=i, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = box

        # Template row (marked with {{items.*}})
        tr = 8
        tpl = ["{{items.__index__}}", "{{items.so_no}}", "{{items.item_name}}",
               "{{items.qty_received}}", "{{items.unit}}", "{{items.receive_date}}"]
        for i, v in enumerate(tpl, 1):
            c = ws.cell(row=tr, column=i, value=v)
            c.border = box
            if i in (1, 4, 5):
                c.alignment = Alignment(horizontal="center")

        # Footer
        ws["A20"] = "Dicetak Oleh:"; ws["A20"].font = Font(size=9)
        ws.merge_cells("A21:B21"); ws["A21"] = "{{printed_by}}"; ws["A21"].font = Font(bold=True)
        ws["D20"] = "Diperiksa Oleh:"; ws["D20"].font = Font(size=9)
        ws.merge_cells("D21:F21"); ws["D21"] = "(________________)"; ws["D21"].alignment = Alignment(horizontal="center")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="STARTER_{code}.xlsx"'},
    )
