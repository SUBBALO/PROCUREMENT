"""HRD module: Master Karyawan, Slip Gaji (fleksibel), blast email via Gmail SMTP.
Portal dikunci PIN — bahkan admin harus masukkan PIN untuk melihat data gaji."""
import io
import uuid
import smtplib
import ssl
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

import jwt
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Header
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db import db
from deps import get_current_user, log_action, is_super_admin_user
from security import hash_password, verify_password, JWT_SECRET, JWT_ALGORITHM
from services.soft_delete import NOT_DELETED_FILTER, soft_delete_one

router = APIRouter(prefix="/hrd", tags=["hrd"])

ROMAN = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII", 9: "IX", 10: "X", 11: "XI", 12: "XII"}
BULAN_ID = {1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember"}

# HRD menus (Accurate-style granular permission). Each menu supports actions:
# view, create, edit, delete, report
HRD_MENUS = [
    {"key": "hrd_karyawan", "label": "Master Karyawan", "group": "gaji"},
    {"key": "hrd_slip_gaji", "label": "Slip Gaji", "group": "gaji"},
    {"key": "hrd_email", "label": "Kirim Email Slip", "group": "gaji"},
    {"key": "hrd_settings", "label": "Pengaturan Email", "group": "gaji"},
    {"key": "hrd_dokumen", "label": "Dokumen HRD", "group": "dokumen"},
]
HRD_MENU_KEYS = [m["key"] for m in HRD_MENUS]
HRD_ACTIONS = ["view", "create", "edit", "delete", "report"]


def _now():
    return datetime.now(timezone.utc).isoformat()


def _is_super(current: dict) -> bool:
    return is_super_admin_user(current) or current.get("role") == "super_admin" or bool(current.get("is_super_admin"))


def has_perm(current: dict, menu: str, action: str) -> bool:
    if _is_super(current):
        return True
    acc = (current.get("access") or {}).get(menu) or {}
    return bool(acc.get(action))


def _has_any_hrd(current: dict) -> bool:
    if _is_super(current):
        return True
    acc = current.get("access") or {}
    return any((acc.get(k) or {}).get("view") for k in HRD_MENU_KEYS)


# Menu grup "gaji" — area sensitif yang dikunci PIN Gaji
GAJI_GROUP = {m["key"] for m in HRD_MENUS if m["group"] == "gaji"}


def _can_manage_pin(current: dict) -> bool:
    """Super admin ATAU user yang punya akses gaji (mis. Herliana) boleh set/reset PIN Gaji."""
    if _is_super(current):
        return True
    acc = current.get("access") or {}
    return any((acc.get(k) or {}).get("view") for k in GAJI_GROUP)


async def _gaji_pin_is_set() -> bool:
    s = await db.hrd_settings.find_one({"_id": "hrd"})
    return bool(s and s.get("pin_hash"))


def _valid_token(token: str, scope: str, uid: str | None = None) -> bool:
    if not token:
        return False
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("scope") != scope:
            return False
        if uid is not None and payload.get("uid") != uid:
            return False
        return True
    except Exception:
        return False


# ---------------- Portal gate (akses HRD + PIN Portal per-user) ----------------
async def require_hrd(x_hrd_token: str = Header(None), current: dict = Depends(get_current_user)) -> dict:
    # 1) Harus punya akses HRD
    if not _has_any_hrd(current):
        raise HTTPException(status_code=403, detail="Anda tidak punya akses ke Portal HRD")
    # 2) PIN Portal wajib per-user. Super admin dikecualikan (boleh tanpa PIN portal).
    if not _is_super(current):
        if not current.get("hrd_pin_hash"):
            raise HTTPException(status_code=428, detail="PIN Portal HRD belum diatur. Silakan buat PIN portal dulu.")
        if not _valid_token(x_hrd_token, "hrd_portal", current.get("id")):
            raise HTTPException(status_code=401, detail="PIN Portal HRD diperlukan")
    return current


def require_hrd_perm(menu: str, action: str):
    """Portal gate + permission menu/action. Untuk menu grup 'gaji',
    bila PIN Gaji aktif maka wajib token PIN Gaji (header x-hrd-gaji)."""
    async def _dep(x_hrd_gaji: str = Header(None), current: dict = Depends(require_hrd)) -> dict:
        if not has_perm(current, menu, action):
            raise HTTPException(status_code=403, detail="Anda tidak memiliki hak akses untuk aksi ini")
        if menu in GAJI_GROUP and not _is_super(current):
            if await _gaji_pin_is_set() and not _valid_token(x_hrd_gaji, "hrd_gaji"):
                raise HTTPException(status_code=401, detail="PIN Gaji diperlukan")
        return current
    return _dep


class PinIn(BaseModel):
    pin: str
    current_pin: str | None = None


# ---------------- PIN Portal (per-user) ----------------
@router.post("/portal-pin/set")
async def set_portal_pin(payload: PinIn, current: dict = Depends(get_current_user)):
    if not _has_any_hrd(current):
        raise HTTPException(status_code=403, detail="Anda tidak punya akses ke Portal HRD")
    if not payload.pin or len(payload.pin) < 4:
        raise HTTPException(status_code=400, detail="PIN minimal 4 digit")
    existing = current.get("hrd_pin_hash")
    if existing and not (payload.current_pin and verify_password(payload.current_pin, existing)):
        raise HTTPException(status_code=400, detail="PIN portal lama salah")
    await db.users.update_one({"id": current["id"]}, {"$set": {"hrd_pin_hash": hash_password(payload.pin), "hrd_pin_updated_at": _now()}})
    await log_action(current, "hrd_set_portal_pin", "user", current["id"], {})
    return {"success": True}


@router.post("/portal-pin/verify")
async def verify_portal_pin(payload: PinIn, current: dict = Depends(get_current_user)):
    if not current.get("hrd_pin_hash"):
        raise HTTPException(status_code=400, detail="PIN Portal belum diatur.")
    if not verify_password(payload.pin, current["hrd_pin_hash"]):
        await log_action(current, "hrd_access_denied", "hrd", "", {"reason": "PIN portal salah"})
        raise HTTPException(status_code=401, detail="PIN Portal salah")
    token = jwt.encode({"scope": "hrd_portal", "uid": current.get("id"),
                        "exp": datetime.now(timezone.utc) + timedelta(hours=10)}, JWT_SECRET, algorithm=JWT_ALGORITHM)
    await log_action(current, "hrd_access", "hrd", "", {"portal": "HRD"})
    return {"portal_token": token}


# ---------------- PIN Gaji (khusus Herliana / gaji user) ----------------
@router.post("/set-pin")
async def set_pin(payload: PinIn, current: dict = Depends(get_current_user)):
    if not _can_manage_pin(current):
        raise HTTPException(status_code=403, detail="Hanya user Gaji (mis. Herliana) atau Super Admin yang bisa mengatur PIN Gaji")
    if not payload.pin or len(payload.pin) < 4:
        raise HTTPException(status_code=400, detail="PIN minimal 4 digit")
    s = await db.hrd_settings.find_one({"_id": "hrd"})
    if s and s.get("pin_hash") and not (payload.current_pin and verify_password(payload.current_pin, s["pin_hash"])):
        raise HTTPException(status_code=400, detail="PIN Gaji lama salah")
    await db.hrd_settings.update_one({"_id": "hrd"}, {"$set": {"pin_hash": hash_password(payload.pin), "pin_updated_at": _now()}}, upsert=True)
    await log_action(current, "hrd_set_pin", "hrd_settings", "hrd", {})
    return {"success": True}


@router.post("/verify-pin")
async def verify_pin(payload: PinIn, current: dict = Depends(require_hrd)):
    s = await db.hrd_settings.find_one({"_id": "hrd"})
    if not s or not s.get("pin_hash"):
        raise HTTPException(status_code=400, detail="PIN Gaji belum diatur.")
    if not verify_password(payload.pin, s["pin_hash"]):
        await log_action(current, "hrd_access_denied", "hrd", "", {"reason": "PIN gaji salah"})
        raise HTTPException(status_code=401, detail="PIN Gaji salah")
    token = jwt.encode({"scope": "hrd_gaji", "uid": current.get("id"),
                        "exp": datetime.now(timezone.utc) + timedelta(hours=8)}, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return {"gaji_token": token}


@router.get("/pin-status")
async def pin_status(current: dict = Depends(get_current_user)):
    return {
        "portal_pin_set": bool(current.get("hrd_pin_hash")),
        "gaji_pin_set": await _gaji_pin_is_set(),
        "can_manage_gaji_pin": _can_manage_pin(current),
        "is_super": _is_super(current),
    }


@router.get("/menu-defs")
async def menu_defs(current: dict = Depends(get_current_user)):
    """Definisi menu HRD + aksi, untuk editor permission di panel Admin (Super Admin only)."""
    if not _is_super(current):
        raise HTTPException(status_code=403, detail="Hanya Super Admin")
    return {"menus": HRD_MENUS, "actions": HRD_ACTIONS}


@router.get("/my-access")
async def my_access(current: dict = Depends(get_current_user)):
    """Info untuk frontend: akses portal, status PIN (portal & gaji), dan matrix akses efektif."""
    is_super = _is_super(current)
    acc = current.get("access") or {}
    effective = {}
    for k in HRD_MENU_KEYS:
        if is_super:
            effective[k] = {a: True for a in HRD_ACTIONS}
        else:
            m = acc.get(k) or {}
            effective[k] = {a: bool(m.get(a)) for a in HRD_ACTIONS}
    return {
        "is_super": is_super,
        "can_enter": is_super or _has_any_hrd(current),
        "portal_pin_set": bool(current.get("hrd_pin_hash")),
        "gaji_pin_set": await _gaji_pin_is_set(),
        "can_manage_gaji_pin": _can_manage_pin(current),
        "menus": HRD_MENUS,
        "gaji_group": sorted(GAJI_GROUP),
        "access": effective,
    }


# ---------------- Access log ----------------
HRD_LOG_ACTIONS = ["hrd_access", "hrd_access_denied", "hrd_set_pin", "hrd_set_portal_pin", "hrd_import_excel", "hrd_blast"]
ACTION_LABEL = {
    "hrd_access": "Buka Portal HRD",
    "hrd_access_denied": "Gagal masuk (PIN salah)",
    "hrd_set_pin": "Ubah/Set PIN Gaji",
    "hrd_set_portal_pin": "Ubah/Set PIN Portal",
    "hrd_import_excel": "Import Excel slip gaji",
    "hrd_blast": "Kirim email slip gaji",
}


@router.get("/logs")
async def hrd_logs(current: dict = Depends(require_hrd)):
    items = await db.activity_logs.find(
        {"action": {"$in": HRD_LOG_ACTIONS}}, {"_id": 0}
    ).sort("timestamp", -1).to_list(300)
    for it in items:
        it["action_label"] = ACTION_LABEL.get(it.get("action"), it.get("action"))
    return {"items": items}


# ---------------- Settings (Gmail SMTP) ----------------
class SettingsIn(BaseModel):
    gmail_user: str | None = None
    app_password: str | None = None
    sender_name: str | None = None


@router.get("/settings")
async def get_settings(current: dict = Depends(require_hrd_perm("hrd_settings", "view"))):
    s = await db.hrd_settings.find_one({"_id": "hrd"}) or {}
    return {
        "gmail_user": s.get("gmail_user", ""),
        "sender_name": s.get("sender_name", "PT. MITRA KARYA SARANA"),
        "has_app_password": bool(s.get("app_password")),
    }


@router.post("/settings")
async def save_settings(payload: SettingsIn, current: dict = Depends(require_hrd_perm("hrd_settings", "edit"))):
    upd = {}
    if payload.gmail_user is not None:
        upd["gmail_user"] = payload.gmail_user.strip()
    if payload.sender_name is not None:
        upd["sender_name"] = payload.sender_name.strip()
    if payload.app_password:  # only overwrite when provided
        upd["app_password"] = payload.app_password.replace(" ", "").strip()
    upd["settings_updated_at"] = _now()
    await db.hrd_settings.update_one({"_id": "hrd"}, {"$set": upd}, upsert=True)
    return {"success": True}


# ---------------- Employees ----------------
class EmployeeIn(BaseModel):
    nik: str = ""
    nama: str = ""
    email: str = ""
    jabatan: str = ""
    no_rekening: str = ""
    bank: str = ""


@router.get("/employees")
async def list_employees(q: str = "", current: dict = Depends(require_hrd_perm("hrd_karyawan", "view"))):
    flt = dict(NOT_DELETED_FILTER)
    if q:
        flt["$and"] = [{"$or": [{"nama": {"$regex": q, "$options": "i"}}, {"nik": {"$regex": q, "$options": "i"}}, {"jabatan": {"$regex": q, "$options": "i"}}]}]
    items = await db.hrd_employees.find(flt, {"_id": 0}).sort("nama", 1).to_list(1000)
    return {"items": items}


@router.post("/employees")
async def create_employee(payload: EmployeeIn, current: dict = Depends(require_hrd_perm("hrd_karyawan", "create"))):
    if not payload.nama.strip():
        raise HTTPException(status_code=400, detail="Nama wajib diisi")
    doc = payload.dict()
    doc.update({"id": str(uuid.uuid4()), "active": True, "created_at": _now(), "updated_at": _now()})
    await db.hrd_employees.insert_one(dict(doc))
    doc.pop("_id", None)
    return doc


@router.put("/employees/{emp_id}")
async def update_employee(emp_id: str, payload: EmployeeIn, current: dict = Depends(require_hrd_perm("hrd_karyawan", "edit"))):
    r = await db.hrd_employees.update_one({"id": emp_id, **NOT_DELETED_FILTER}, {"$set": {**payload.dict(), "updated_at": _now()}})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    return await db.hrd_employees.find_one({"id": emp_id}, {"_id": 0})


@router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, current: dict = Depends(require_hrd_perm("hrd_karyawan", "delete"))):
    ok = await soft_delete_one("hrd_employees", {"id": emp_id}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    return {"success": True}


# ---------------- Payslips ----------------
class Component(BaseModel):
    label: str = ""
    amount: float = 0.0


class PayslipIn(BaseModel):
    period_month: int
    period_year: int
    employee_id: str | None = None
    nik: str = ""
    nama: str = ""
    email: str = ""
    jabatan: str = ""
    dept: str = "Production"
    no_rekening: str = ""
    bank: str = ""
    earnings: list[Component] = []
    deductions: list[Component] = []
    take_home: float | None = None
    notes: str = ""


def _round_rp(v: float) -> int:
    # Pembulatan ke ribuan terdekat
    return int(round(float(v or 0) / 1000.0) * 1000)


def _compute_slip(d: dict) -> dict:
    gross = round(sum(float(e.get("amount") or 0) for e in d.get("earnings", [])), 2)
    ded = round(sum(float(x.get("amount") or 0) for x in d.get("deductions", [])), 2)
    d["gross"] = gross
    d["total_deduction"] = ded
    d["net"] = round(gross - ded, 2)
    # take_home (PEMBULATAN) editable; default = pembulatan net ke ribuan
    if d.get("take_home") in (None, "", 0):
        d["take_home"] = _round_rp(d["net"])
    else:
        d["take_home"] = float(d["take_home"])
    return d


@router.get("/payslips")
async def list_payslips(month: int = 0, year: int = 0, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "view"))):
    flt = dict(NOT_DELETED_FILTER)
    if month:
        flt["period_month"] = month
    if year:
        flt["period_year"] = year
    items = await db.hrd_payslips.find(flt, {"_id": 0}).sort("nama", 1).to_list(2000)
    return {"items": items}


@router.post("/payslips")
async def create_payslip(payload: PayslipIn, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "create"))):
    d = payload.dict()
    d["earnings"] = [e for e in d["earnings"] if (e.get("label") or e.get("amount"))]
    d["deductions"] = [e for e in d["deductions"] if (e.get("label") or e.get("amount"))]
    _compute_slip(d)
    d.update({"id": str(uuid.uuid4()), "email_status": "belum", "email_error": "", "sent_at": None, "created_at": _now(), "updated_at": _now()})
    await db.hrd_payslips.insert_one(dict(d))
    d.pop("_id", None)
    return d


@router.put("/payslips/{sid}")
async def update_payslip(sid: str, payload: PayslipIn, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "edit"))):
    d = payload.dict()
    d["earnings"] = [e for e in d["earnings"] if (e.get("label") or e.get("amount"))]
    d["deductions"] = [e for e in d["deductions"] if (e.get("label") or e.get("amount"))]
    _compute_slip(d)
    d["updated_at"] = _now()
    r = await db.hrd_payslips.update_one({"id": sid, **NOT_DELETED_FILTER}, {"$set": d})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Slip tidak ditemukan")
    return await db.hrd_payslips.find_one({"id": sid}, {"_id": 0})


@router.delete("/payslips/{sid}")
async def delete_payslip(sid: str, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "delete"))):
    ok = await soft_delete_one("hrd_payslips", {"id": sid}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="Slip tidak ditemukan")
    return {"success": True}


DEDUCT_KEYWORDS = ["potong", "bpjs", "pph", "iuran", "pinjam", "kasbon", "deduct", "denda",
                   "koperasi", "absent", "absen", "jht", "jp", "jkk", "jkm", "jpk"]
KNOWN = {"nik": "nik", "kode": "nik", "kode_karyawan": "nik", "nama": "nama", "name": "nama",
         "email": "email", "jabatan": "jabatan", "posisi": "jabatan", "dept": "dept",
         "departemen": "dept", "department": "dept",
         "no_rekening": "no_rekening", "norekening": "no_rekening", "rekening": "no_rekening",
         "no_rek": "no_rekening", "bank": "bank"}


def _cell(ws, coord):
    try:
        return ws[coord].value
    except Exception:
        return None


def _numify(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(v)
    except Exception:
        return None


def _parse_slip_sheet(ws, month, year):
    """Parse satu sheet slip (format cetak MKS) menjadi dict slip.
    Layout tetap: header di A1-A3, SLIP GAJI di A5, info di baris 8-10,
    PENGHASILAN (kolom A/C/E) & PENGURANGAN (kolom G/I/J/K)."""
    nama = _cell(ws, "C8")
    if not nama or str(nama).strip() == "":
        return None
    nik = _cell(ws, "E8")
    slip = {
        "nama": str(nama).strip(),
        "nik": str(nik).strip() if nik else "",
        "dept": str(_cell(ws, "C9") or "").strip(),
        "jabatan": str(_cell(ws, "C10") or "").strip(),
        "perhari": _numify(_cell(ws, "J8")),
        "lembur_jam": _numify(_cell(ws, "J9")),
        "tkehadiran_rate": _numify(_cell(ws, "J10")),
        "earnings": [],
        "deductions": [],
    }
    max_row = min(ws.max_row, 40)
    # Earnings: kolom A=label, C/D=qty, E=amount. Berhenti di "JUMLAH".
    gross = None
    for row in range(13, max_row + 1):
        lbl = _cell(ws, f"A{row}")
        lbl_s = str(lbl).strip() if lbl else ""
        if not lbl_s:
            continue
        if lbl_s.upper().startswith("JUMLAH"):
            gross = _numify(_cell(ws, f"E{row}"))
            break
        qty = _numify(_cell(ws, f"C{row}"))
        if qty is None:
            qty = _numify(_cell(ws, f"D{row}"))
        amt = _numify(_cell(ws, f"E{row}"))
        slip["earnings"].append({"label": lbl_s, "qty": qty, "amount": amt or 0})
    # Deductions: kolom G=label, I=qty, J=unit, K=amount.
    # Line item hanya dikumpulkan SEBELUM baris "JUMLAH"; setelah itu hanya
    # ambil PENGHASILAN BERSIH & PEMBULATAN lalu berhenti (abaikan footer).
    total_ded = None
    net = None
    take_home = None
    after_jumlah = False
    for row in range(13, max_row + 1):
        lbl = _cell(ws, f"G{row}")
        lbl_s = str(lbl).strip() if lbl else ""
        if not lbl_s:
            continue
        up = lbl_s.upper()
        if up.startswith("JUMLAH"):
            total_ded = _numify(_cell(ws, f"K{row}"))
            after_jumlah = True
            continue
        if "BERSIH" in up:
            net = _numify(_cell(ws, f"K{row}"))
            continue
        if "PEMBULATAN" in up:
            take_home = _numify(_cell(ws, f"K{row}"))
            break
        if after_jumlah:
            continue  # abaikan footer (Batam, Prepared By, HRD, dll)
        qty = _numify(_cell(ws, f"I{row}"))
        unit = _cell(ws, f"J{row}")
        unit_s = str(unit).strip() if unit and not isinstance(unit, (int, float)) else ""
        amt = _numify(_cell(ws, f"K{row}"))
        slip["deductions"].append({"label": lbl_s, "qty": qty, "unit": unit_s, "amount": amt or 0})
    # Terbilang (cari di kolom A/B/C sekitar baris 26)
    terbilang = None
    for row in range(24, max_row + 1):
        a = _cell(ws, f"A{row}")
        if a and "terbilang" in str(a).lower():
            terbilang = _cell(ws, f"C{row}") or _cell(ws, f"D{row}")
            break
    slip["gross"] = gross if gross is not None else round(sum((e.get("amount") or 0) for e in slip["earnings"]), 2)
    slip["total_deduction"] = total_ded if total_ded is not None else round(sum((d.get("amount") or 0) for d in slip["deductions"]), 2)
    slip["net"] = net if net is not None else round(slip["gross"] - slip["total_deduction"], 2)
    slip["take_home"] = take_home if take_home is not None else _round_rp(slip["net"])
    slip["terbilang"] = str(terbilang).strip() if terbilang else ""
    slip["period_month"] = month
    slip["period_year"] = year
    return slip


@router.post("/payslips/import-excel")
async def import_excel(month: int = Form(...), year: int = Form(...), file: UploadFile = File(...), current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "create"))):
    from openpyxl import load_workbook
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File Excel tidak valid (.xlsx)")

    # Deteksi sheet slip: cell A5 == "SLIP GAJI"
    slip_sheets = [n for n in wb.sheetnames if str(_cell(wb[n], "A5") or "").strip().upper() == "SLIP GAJI"]
    if not slip_sheets:
        raise HTTPException(status_code=400,
                            detail="Tidak ditemukan sheet slip gaji (yang memiliki judul 'SLIP GAJI'). Pastikan file berisi sheet slip per karyawan.")

    # Preload master karyawan untuk auto-match email/bank/rekening
    emps = await db.hrd_employees.find(NOT_DELETED_FILTER, {"_id": 0}).to_list(2000)
    by_nik = {(e.get("nik") or "").strip().lower(): e for e in emps if e.get("nik")}
    by_nama = {(e.get("nama") or "").strip().lower(): e for e in emps if e.get("nama")}

    created = 0
    updated = 0
    names = []
    for name in slip_sheets:
        slip = _parse_slip_sheet(wb[name], month, year)
        if not slip:
            continue
        # Auto-match ke Master Karyawan (email/bank/rekening)
        match = by_nik.get((slip.get("nik") or "").strip().lower()) if slip.get("nik") else None
        if not match:
            match = by_nama.get(slip["nama"].strip().lower())
        slip["email"] = ""
        slip["bank"] = ""
        slip["no_rekening"] = ""
        if match:
            slip["employee_id"] = match.get("id")
            slip["email"] = match.get("email", "")
            slip["bank"] = match.get("bank", "")
            slip["no_rekening"] = match.get("no_rekening", "")
            if not slip.get("jabatan") and match.get("jabatan"):
                slip["jabatan"] = match["jabatan"]
        # Upsert berdasarkan (period + nama) agar re-import menimpa, bukan dobel
        existing = await db.hrd_payslips.find_one(
            {"period_month": month, "period_year": year, "nama": slip["nama"], **NOT_DELETED_FILTER})
        slip["updated_at"] = _now()
        if existing:
            slip["email_status"] = existing.get("email_status", "belum")
            slip["email_error"] = existing.get("email_error", "")
            await db.hrd_payslips.update_one({"id": existing["id"]}, {"$set": slip})
            updated += 1
        else:
            slip.update({"id": str(uuid.uuid4()), "email_status": "belum", "email_error": "",
                         "sent_at": None, "notes": "", "created_at": _now()})
            await db.hrd_payslips.insert_one(dict(slip))
            created += 1
        names.append(slip["nama"])

    await log_action(current, "hrd_import_excel", "hrd_payslips", "",
                     {"created": created, "updated": updated, "period": f"{month}-{year}", "sheets": len(slip_sheets)})
    return {"success": True, "created": created, "updated": updated, "names": names, "sheets": slip_sheets}


# ---------------- PDF slip ----------------
COMPANY_ADDRESS = ["Taiwan International Park Blok B No. 117", "Kabil, Nongsa - Batam Island"]


def _rp(v):
    """Angka Rupiah format Indonesia dengan 2 desimal; 0/None => '-' (untuk isi teks biasa)."""
    try:
        f = float(v or 0)
    except Exception:
        return "Rp -"
    if abs(f) < 0.005:
        return "Rp -"
    s = f"{f:,.2f}".replace(",", "#").replace(".", ",").replace("#", ".")
    return "Rp " + s


def _money(v):
    return _rp(v)


def _qty(v):
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else f"{f:g}"
    except Exception:
        return str(v)


_SATUAN = ["", "satu", "dua", "tiga", "empat", "lima", "enam", "tujuh", "delapan", "sembilan",
           "sepuluh", "sebelas"]


def _terbilang(n: int) -> str:
    n = int(abs(n))
    if n < 12:
        return _SATUAN[n]
    if n < 20:
        return _terbilang(n - 10) + " belas"
    if n < 100:
        return _terbilang(n // 10) + " puluh" + ((" " + _terbilang(n % 10)) if n % 10 else "")
    if n < 200:
        return "seratus" + ((" " + _terbilang(n - 100)) if n - 100 else "")
    if n < 1000:
        return _terbilang(n // 100) + " ratus" + ((" " + _terbilang(n % 100)) if n % 100 else "")
    if n < 2000:
        return "seribu" + ((" " + _terbilang(n - 1000)) if n - 1000 else "")
    if n < 1_000_000:
        return _terbilang(n // 1000) + " ribu" + ((" " + _terbilang(n % 1000)) if n % 1000 else "")
    if n < 1_000_000_000:
        return _terbilang(n // 1_000_000) + " juta" + ((" " + _terbilang(n % 1_000_000)) if n % 1_000_000 else "")
    return _terbilang(n // 1_000_000_000) + " milyar" + ((" " + _terbilang(n % 1_000_000_000)) if n % 1_000_000_000 else "")


def _terbilang_rupiah(n) -> str:
    words = _terbilang(int(round(float(n or 0)))).strip()
    if not words:
        words = "nol"
    return " ".join(w.capitalize() for w in words.split()) + " Rupiah"


def _render_slip_pdf(slip: dict, sender_name: str = "PT. MITRA KARYA SARANA", printed_by: str = "") -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=12 * mm, leftMargin=16 * mm, rightMargin=16 * mm)
    styles = getSampleStyleSheet()
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=9)
    tiny = ParagraphStyle("tn", parent=styles["Normal"], fontSize=7.5, textColor=colors.HexColor("#475569"))
    elems = []
    GREY = colors.HexColor("#334155")

    # Header perusahaan
    elems.append(Paragraph(f"<b>{sender_name}</b>", ParagraphStyle("c", parent=styles["Normal"], fontSize=11, fontName="Helvetica-Bold")))
    for line in COMPANY_ADDRESS:
        elems.append(Paragraph(line, ParagraphStyle("addr", parent=styles["Normal"], fontSize=8)))
    elems.append(Spacer(1, 6))
    elems.append(Paragraph("<u>SLIP GAJI</u>", ParagraphStyle("t", parent=styles["Normal"], fontSize=12, alignment=1, fontName="Helvetica-Bold")))
    per = f"{BULAN_ID.get(slip.get('period_month'), slip.get('period_month'))} {slip.get('period_year')}"
    elems.append(Paragraph(f"Periode : {per}", ParagraphStyle("per", parent=small, alignment=1)))
    elems.append(Spacer(1, 8))

    # Info karyawan (kiri) + rate (kanan)
    nik = slip.get("nik", "")
    info = Table([
        ["Nama / NIK", ":", slip.get("nama", ""), nik, "Perhari", ":", _money(slip.get("perhari"))],
        ["Dept", ":", slip.get("dept", "") or "Production", "", "Lembur/Jam", ":", _money(slip.get("lembur_jam"))],
        ["Jabatan", ":", slip.get("jabatan", ""), "", "T. Kehadiran", ":", _money(slip.get("tkehadiran_rate"))],
    ], colWidths=[22 * mm, 4 * mm, 40 * mm, 24 * mm, 26 * mm, 4 * mm, 38 * mm])
    info.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 9), ("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("ALIGN", (6, 0), (6, -1), "RIGHT"),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 2), ("TOPPADDING", (0, 0), (-1, -1), 1)]))
    elems.append(info)
    elems.append(Spacer(1, 6))

    # Tabel PENGHASILAN | PENGURANGAN
    earns = slip.get("earnings", []) or []
    deds = slip.get("deductions", []) or []
    rows = [["PENGHASILAN", "", "", "PENGURANGAN", "", ""]]
    n = max(len(earns), len(deds))
    for i in range(n):
        e = earns[i] if i < len(earns) else None
        d = deds[i] if i < len(deds) else None
        d_qty = ""
        if d is not None:
            q = _qty(d.get("qty"))
            u = (d.get("unit") or "").strip()
            d_qty = (q + (f" {u}" if u else "")).strip()
        rows.append([
            (e.get("label", "") if e else ""), (_qty(e.get("qty")) if e else ""), (_money(e.get("amount")) if e else ""),
            (d.get("label", "") if d else ""), d_qty, (_money(d.get("amount")) if d else ""),
        ])
    rows.append(["JUMLAH", "", _money(slip.get("gross")), "JUMLAH", "", _money(slip.get("total_deduction"))])
    rows.append(["", "", "", "PENGHASILAN BERSIH", "", _money(slip.get("net"))])
    rows.append(["", "", "", "PEMBULATAN", "", _money(slip.get("take_home"))])

    jml_row = len(rows) - 3
    t = Table(rows, colWidths=[34 * mm, 12 * mm, 40 * mm, 34 * mm, 14 * mm, 38 * mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"), ("ALIGN", (5, 0), (5, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "CENTER"), ("ALIGN", (4, 0), (4, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, GREY),           # bawah header
        ("LINEABOVE", (0, jml_row), (-1, jml_row), 0.6, GREY),  # atas JUMLAH
        ("FONTNAME", (0, jml_row), (-1, jml_row), "Helvetica-Bold"),
        ("FONTNAME", (3, jml_row + 1), (5, jml_row + 2), "Helvetica-Bold"),
        ("LINEABOVE", (3, jml_row + 1), (5, jml_row + 1), 0.4, GREY),
        ("BOX", (0, 0), (2, jml_row), 0.8, GREY),            # box kolom penghasilan
        ("BOX", (3, 0), (5, jml_row + 2), 0.8, GREY),        # box kolom pengurangan
        ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 8))

    terb = slip.get("terbilang") or _terbilang_rupiah(slip.get("take_home"))
    elems.append(Paragraph(f"Terbilang : <i>{terb}</i>", small))

    if slip.get("notes"):
        elems.append(Spacer(1, 4))
        elems.append(Paragraph(f"Catatan : {slip['notes']}", small))

    elems.append(Spacer(1, 22))
    tgl = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7)))
    tgl_str = f"Batam, {tgl.day} {BULAN_ID.get(tgl.month)} {tgl.year}"
    right = Table([[tgl_str], ["Prepared By,"], ["HRD"]], colWidths=[70 * mm])
    right.hAlign = "RIGHT"
    right.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 9), ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                               ("BOTTOMPADDING", (0, 0), (-1, -1), 1)]))
    elems.append(right)

    elems.append(Spacer(1, 16))
    stamp = tgl.strftime("%d-%m-%Y %H:%M") + " WIB"
    by = f" oleh {printed_by}" if printed_by else ""
    elems.append(Paragraph(
        f"Dokumen ini dicetak otomatis oleh sistem pada {stamp}{by}. "
        f"Tidak memerlukan tanda tangan basah. Bersifat rahasia — mohon tidak menyebarkan.",
        ParagraphStyle("f", parent=tiny, fontSize=7, textColor=colors.grey)))
    pdf.build(elems)
    buf.seek(0)
    return buf


@router.get("/payslips/{sid}/pdf")
async def payslip_pdf(sid: str, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "report"))):
    slip = await db.hrd_payslips.find_one({"id": sid, **NOT_DELETED_FILTER}, {"_id": 0})
    if not slip:
        raise HTTPException(status_code=404, detail="Slip tidak ditemukan")
    s = await db.hrd_settings.find_one({"_id": "hrd"}) or {}
    buf = _render_slip_pdf(slip, s.get("sender_name") or "PT. MITRA KARYA SARANA",
                           printed_by=current.get("name") or current.get("username", ""))
    fname = f"SlipGaji_{slip.get('nama','')}_{slip.get('period_month')}_{slip.get('period_year')}.pdf".replace(" ", "_")
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="{fname}"'})


# Kolom penghasilan & potongan sesuai template slip gaji MKS
SLIP_EARNINGS = ["Gaji Pokok", "T. Tetap", "T. Kehadiran", "Lembur (1.5)", "Lembur (2)",
                 "Lembur (3)", "Lembur (4)", "Insentive+ 2nd Shift"]
SLIP_DEDUCTIONS = ["Absent", "T. Transport", "PPh 21", "JHT+JP (2%+1%)", "BPJS KESEHATAN 1%", "Pinjaman"]


@router.get("/import-template")
async def import_template(current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "view"))):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Slip Gaji"
    headers = ["NIK", "Nama", "Email", "Jabatan", "Dept", "No Rekening", "Bank"] + SLIP_EARNINGS + SLIP_DEDUCTIONS
    ws.append(headers)
    ws.append(["MKS 0021", "Wawan Munandar", "wawan@email.com", "Supervisor", "Production", "1234567890", "BCA",
               5000000, 570000, 0, 0, 0, 0, 0, 0,  # earnings
               0, 0, 91634, 304920, 101640, 0])     # deductions
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="Template_Import_Slip_Gaji.xlsx"'})


@router.get("/slip-labels")
async def slip_labels(current: dict = Depends(require_hrd)):
    return {"earnings": SLIP_EARNINGS, "deductions": SLIP_DEDUCTIONS}


# ---------------- Blast email ----------------
class BlastIn(BaseModel):
    month: int
    year: int
    ids: list[str] | None = None


@router.post("/blast")
async def blast(payload: BlastIn, current: dict = Depends(require_hrd_perm("hrd_email", "create"))):
    s = await db.hrd_settings.find_one({"_id": "hrd"}) or {}
    gmail_user = s.get("gmail_user")
    app_pw = s.get("app_password")
    sender_name = s.get("sender_name") or "PT. MITRA KARYA SARANA"
    if not gmail_user or not app_pw:
        raise HTTPException(status_code=400, detail="Email Gmail belum dikonfigurasi. Isi di tab Pengaturan.")

    flt = {"period_month": payload.month, "period_year": payload.year, **NOT_DELETED_FILTER}
    if payload.ids:
        flt["id"] = {"$in": payload.ids}
    slips = await db.hrd_payslips.find(flt, {"_id": 0}).to_list(2000)
    if not slips:
        raise HTTPException(status_code=400, detail="Tidak ada slip untuk periode ini")

    per_label = f"{BULAN_ID.get(payload.month, payload.month)} {payload.year}"
    results = []
    try:
        context = ssl.create_default_context()
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context, timeout=30)
        server.login(gmail_user, app_pw)
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=400, detail="Login Gmail gagal. Pastikan email & App Password benar (bukan password biasa).")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal koneksi ke Gmail: {e}")

    try:
        for slip in slips:
            email_to = (slip.get("email") or "").strip()
            status, err = "gagal", ""
            if not email_to:
                err = "Email kosong"
            else:
                try:
                    msg = MIMEMultipart()
                    msg["From"] = f"{sender_name} <{gmail_user}>"
                    msg["To"] = email_to
                    msg["Subject"] = f"Slip Gaji {per_label} - {slip.get('nama','')}"
                    body = (f"Yth. {slip.get('nama','')},\n\n"
                            f"Berikut kami lampirkan slip gaji Anda untuk periode {per_label}.\n"
                            f"Take Home Pay: {_rp(slip.get('take_home'))}.\n\n"
                            f"Dokumen ini bersifat rahasia. Mohon tidak menyebarkan.\n\n"
                            f"Hormat kami,\n{sender_name}")
                    msg.attach(MIMEText(body, "plain"))
                    pdf_buf = _render_slip_pdf(slip, sender_name, printed_by=current.get("name") or current.get("username", ""))
                    part = MIMEApplication(pdf_buf.read(), _subtype="pdf")
                    part.add_header("Content-Disposition", "attachment",
                                    filename=f"SlipGaji_{per_label}_{slip.get('nama','')}.pdf".replace(" ", "_"))
                    msg.attach(part)
                    server.sendmail(gmail_user, [email_to], msg.as_string())
                    status, err = "terkirim", ""
                except Exception as e:
                    status, err = "gagal", str(e)[:200]
            await db.hrd_payslips.update_one({"id": slip["id"]}, {"$set": {"email_status": status, "email_error": err, "sent_at": _now() if status == "terkirim" else None}})
            results.append({"id": slip["id"], "nama": slip.get("nama"), "email": email_to, "status": status, "error": err})
    finally:
        try:
            server.quit()
        except Exception:
            pass

    sent = sum(1 for r in results if r["status"] == "terkirim")
    await log_action(current, "hrd_blast", "hrd_payslips", "", {"period": per_label, "sent": sent, "total": len(results)})
    return {"success": True, "sent": sent, "failed": len(results) - sent, "results": results}
