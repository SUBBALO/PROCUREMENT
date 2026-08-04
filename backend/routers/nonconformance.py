"""NONCONFORMANCE (CAR) — Corrective Action Report terhadap Drawing.

Alur:
  QC / Produksi / Sales menerbitkan NC terhadap 1 atau banyak Drawing.
  Engineering Leader menindaklanjuti: assign ke staff → revisi → terbit ECN.

Status flow: open → assigned → in_progress → closed

Catatan penting:
  - Baseline schema ini SENGAJA dibuat fleksibel. Field detail sesuai template
    NCR resmi perusahaan akan ditambahkan belakangan (extend tanpa breaking).
  - KPI #1 Engineering (Drawing tanpa NC) dihitung dari koleksi ini
    berdasarkan BULAN NC diterbitkan (issued_at) — lihat routers/kpi.py.
  - Auditability: semua perubahan status dicatat di `timeline` + activity_logs.
"""
from __future__ import annotations

import io
import os
import uuid
from datetime import datetime, timezone
from typing import List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from pydantic import BaseModel, Field

from db import db
from deps import (
    get_current_user, log_action, is_admin_like, is_eng_head, is_engineering,
    is_nc_issuer, is_qc, is_production, is_sales, is_mr,
)

router = APIRouter(tags=["nonconformance"])

# ── Status ──────────────────────────────────────────────────────────────────
STATUS_OPEN = "open"
STATUS_ASSIGNED = "assigned"
STATUS_IN_PROGRESS = "in_progress"
STATUS_CLOSED = "closed"
VALID_STATUSES = {STATUS_OPEN, STATUS_ASSIGNED, STATUS_IN_PROGRESS, STATUS_CLOSED}
STATUS_LABELS = {
    STATUS_OPEN: "Open",
    STATUS_ASSIGNED: "Assigned",
    STATUS_IN_PROGRESS: "In Progress",
    STATUS_CLOSED: "Closed",
}

SEVERITY_LEVELS = {"minor", "major", "critical"}
NC_SOURCES = {"in_house", "external"}  # Sesuai form: IN-HOUSE / EXTERNAL

# ── Departemen (CAR berlaku untuk SEMUA dept) ────────────────────────────────
# key → {label, roles}. Dipakai untuk "Issued To" + daftar user yang bisa ditugaskan.
DEPARTMENTS = {
    "engineering": {"label": "Engineering", "roles": ["eng_leader", "eng_head", "engineering", "eng_staff"]},
    "qc": {"label": "Quality Control", "roles": ["qc"]},
    "produksi": {"label": "Produksi", "roles": ["produksi", "production"]},
    "sales": {"label": "Sales", "roles": ["sales"]},
    "purchasing": {"label": "Purchasing", "roles": ["purchasing", "staff"]},
    "store": {"label": "Store", "roles": ["store"]},
    "document_control": {"label": "Document Control", "roles": ["doc_control", "document_control"]},
    "finance": {"label": "Finance", "roles": ["finance"]},
    "management": {"label": "Management", "roles": ["admin", "super_admin", "supervisor"]},
    "other": {"label": "Lainnya", "roles": []},
}
# Kategori objek yang bisa "kena NC". Hanya "drawing" yang memengaruhi KPI Engineering.
LINK_TYPES = {"drawing", "so", "incoming_material", "product_part", "supplier", "process_general"}
LINK_TYPE_LABELS = {
    "drawing": "Drawing", "so": "SO (Sales Order)",
    "incoming_material": "Incoming Material/Goods",
    "product_part": "Produk/Part",
    "supplier": "Supplier/Vendor", "process_general": "Proses/Umum",
}

_ROMAN = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI",
          7: "VII", 8: "VIII", 9: "IX", 10: "X", 11: "XI", 12: "XII"}


def _role_to_dept(role: str) -> str:
    for key, meta in DEPARTMENTS.items():
        if role in meta["roles"]:
            return key
    return "other"


# ── GridFS untuk lampiran bukti NC ────────────────────────────────────────────
_gridfs: Optional[AsyncIOMotorGridFSBucket] = None


def _fs() -> AsyncIOMotorGridFSBucket:
    global _gridfs
    if _gridfs is None:
        _gridfs = AsyncIOMotorGridFSBucket(db, bucket_name="nc_attachments")
    return _gridfs


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ext(name: str) -> str:
    return os.path.splitext(name or "")[1].lower()


def _clean(doc: dict) -> dict:
    if doc:
        doc.pop("_id", None)
    return doc


def _actor(user: dict) -> dict:
    return {"id": user.get("id"), "name": user.get("name") or user.get("username"),
            "role": user.get("role")}


def _issuer_dept_of(user: dict) -> str:
    return _role_to_dept((user or {}).get("role", ""))


def _is_initiator(user: dict, doc: dict) -> bool:
    return (doc.get("issued_by") or {}).get("id") == user.get("id")


def _is_target(user: dict, doc: dict) -> bool:
    """User termasuk dept tujuan / assignee / issued_to_user dari CAR."""
    if _role_to_dept((user or {}).get("role", "")) == doc.get("issued_to_dept"):
        return True
    if (doc.get("assigned_to") or {}).get("id") == user.get("id"):
        return True
    if (doc.get("issued_to_user") or {}).get("id") == user.get("id"):
        return True
    return False


async def _next_nc_no() -> str:
    """Nomor CAR resmi: MKS-QA-CAR-{ROMAN_BULAN}-{YY}-{N}.
    Contoh: MKS-QA-CAR-VII-26-34. Running number RESET setiap bulan (mulai dari 1)."""
    now = datetime.now(timezone.utc)
    yy = f"{now.year % 100:02d}"
    roman = _ROMAN.get(now.month, str(now.month))
    key = f"car_{now.year}_{now.month:02d}"   # counter per BULAN (reset tiap bulan)
    counter = await db.counters.find_one_and_update(
        {"_id": key}, {"$inc": {"value": 1}}, upsert=True, return_document=True,
    )
    seq = (counter or {}).get("value", 1)
    return f"MKS-QA-CAR-{roman}-{yy}-{seq:02d}"


async def _get_nc_or_404(nc_id: str) -> dict:
    doc = await db.nonconformances.find_one({"id": nc_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="NC tidak ditemukan")
    return doc


def _can_view(user: dict) -> bool:
    """CAR berlaku untuk SEMUA dept → semua user terautentikasi boleh melihat."""
    return bool(user)


# ── Payloads ─────────────────────────────────────────────────────────────────
class DrawingRef(BaseModel):
    drawing_id: str = ""
    drawing_no: str = ""


class NonconformanceCreate(BaseModel):
    # ── Header (Completed by CAR Initiator) — bisa dibuat SEMUA user ──
    issued_to_dept: str = ""              # dept tujuan (wajib): engineering/qc/produksi/...
    issued_to_user_id: Optional[str] = "" # user tujuan spesifik (opsional)
    issued_to_user_name: Optional[str] = ""
    issued_to: Optional[str] = ""         # legacy free-text (fallback tampilan)
    expected_reply_date: Optional[str] = ""
    # ── Objek yang kena NC (fleksibel) ──
    link_type: str = "process_general"  # drawing (memengaruhi KPI Eng) | so | product_part | supplier | process_general
    drawings: List[DrawingRef] = Field(default_factory=list)  # dipakai bila link_type=drawing
    object_ref: str = ""                  # teks bebas objek yang kena NC (bila non-drawing)
    incoming_receipt_id: Optional[str] = ""  # id store_receipts (bila link_type=incoming_material)
    so_no: Optional[str] = ""
    customer_name: Optional[str] = ""
    # ── Section 1: NONCONFORMANCE INFORMATION ──
    description: str = ""
    source: str = "in_house"
    title: str = ""
    severity: str = "major"
    extra: dict = Field(default_factory=dict)


class InvestigationIn(BaseModel):
    """Section 2 — INVESTIGATION & ACTION PLANS (oleh Responsible Dept./Assignee)."""
    root_cause: str = ""
    immediate_action: str = ""
    corrective_action: str = ""
    preventive_action: str = ""            # untuk log Internal Eng Process (MKS-F-ENG-006 tab NC)
    completed_by: str = ""                # Actions Completed By (Name)
    completed_date: Optional[str] = ""    # Date
    dept_head_name: str = ""             # Approved by Dept. Head
    dept_head_date: Optional[str] = ""
    ecn_no: str = ""                       # ECN yang diterbitkan (MKS-F-ENG-004)
    set_in_progress: bool = True          # otomatis pindah ke In Progress


class CloseoutIn(BaseModel):
    """Section 3 — CAR CLOSEOUT INFORMATION (oleh Initiator/MR)."""
    initiator_remarks: str = ""
    risk_review: bool = False             # Review of risks & opportunities: Yes/No
    risk_attached: bool = False           # (if yes please attached)
    effectiveness_reviewed_by: str = ""
    effectiveness_date: Optional[str] = ""
    qa_approved_by: str = ""
    qa_date: Optional[str] = ""
    close: bool = False                   # True = sekaligus tutup NC (Closed)


class AssignIn(BaseModel):
    assignee_id: str
    assignee_name: Optional[str] = ""
    notes: str = ""


class StatusIn(BaseModel):
    status: str
    notes: str = ""
    ecn_id: Optional[str] = ""
    ecn_no: Optional[str] = ""


class NoteIn(BaseModel):
    notes: str


# ── CREATE ───────────────────────────────────────────────────────────────────
@router.post("/nonconformance")
async def create_nc(payload: NonconformanceCreate, current: dict = Depends(get_current_user)):
    # CAR berlaku untuk SEMUA departemen → semua user terautentikasi boleh menerbitkan.
    link_type = payload.link_type if payload.link_type in LINK_TYPES else "process_general"

    # Departemen tujuan (Issued To) wajib.
    to_dept = (payload.issued_to_dept or "").strip().lower()
    if to_dept not in DEPARTMENTS:
        raise HTTPException(status_code=400, detail="Departemen tujuan (Issued To) wajib dipilih")

    if not (payload.description or "").strip() and not (payload.title or "").strip():
        raise HTTPException(status_code=400, detail="Deskripsi ketidaksesuaian wajib diisi")

    resolved: List[dict] = []
    so_no = (payload.so_no or "").strip()
    customer_name = (payload.customer_name or "").strip()
    incoming = None

    if link_type == "drawing":
        drawings = [d for d in (payload.drawings or []) if (d.drawing_id or d.drawing_no)]
        if not drawings:
            raise HTTPException(status_code=400, detail="Pilih minimal satu Drawing untuk NC bertipe Drawing")
        for d in drawings:
            q = {"deleted_at": {"$exists": False}}
            if d.drawing_id:
                q["id"] = d.drawing_id
            else:
                q["drawing_no"] = d.drawing_no
            dwg = await db.drawings.find_one(q, {"_id": 0})
            if dwg:
                resolved.append({"drawing_id": dwg.get("id"), "drawing_no": dwg.get("drawing_no"),
                                 "so_no": dwg.get("so_no"), "customer_name": dwg.get("customer_name"),
                                 "project_name": dwg.get("project_name")})
                so_no = so_no or (dwg.get("so_no") or "")
                customer_name = customer_name or (dwg.get("customer_name") or "")
            else:
                resolved.append({"drawing_id": d.drawing_id, "drawing_no": d.drawing_no,
                                 "so_no": "", "customer_name": "", "project_name": ""})
    else:
        # Objek bebas: SO / Incoming Material / Produk-Part / Supplier / Proses-Umum.
        if link_type == "incoming_material" and (payload.incoming_receipt_id or "").strip():
            rc = await db.store_receipts.find_one(
                {"id": payload.incoming_receipt_id.strip(), "deleted_at": {"$exists": False}}, {"_id": 0})
            if not rc:
                raise HTTPException(status_code=404, detail="Data Incoming Goods tidak ditemukan")
            incoming = {
                "receipt_id": rc.get("id"), "item_name": rc.get("item_name"),
                "vendor_name": rc.get("vendor_name"), "receive_date": rc.get("receive_date"),
                "qty_received": rc.get("qty_received"), "unit": rc.get("unit"),
                "po_no": rc.get("po_no"), "invoice_no": rc.get("invoice_no"),
                "do_number": rc.get("do_number"), "so_no": rc.get("so_no"),
            }
            so_no = so_no or (rc.get("so_no") or "")
            customer_name = customer_name or (rc.get("customer_name") or "")
            # object_ref otomatis (mudah dibaca di masterlist)
            auto = f"{rc.get('item_name') or '-'} · {rc.get('vendor_name') or '-'}"
            if rc.get("invoice_no"):
                auto += f" · INV {rc.get('invoice_no')}"
            if not (payload.object_ref or "").strip():
                payload.object_ref = auto
        elif link_type == "so" and not so_no and not (payload.object_ref or "").strip():
            raise HTTPException(status_code=400, detail="Isi No. SO yang kena NC")
        elif link_type not in ("so", "incoming_material") and not (payload.object_ref or "").strip():
            raise HTTPException(status_code=400, detail="Isi objek yang kena NC (mis. part/supplier/proses)")
        if link_type == "incoming_material" and not incoming and not (payload.object_ref or "").strip():
            raise HTTPException(status_code=400, detail="Pilih data Incoming Goods atau isi objeknya")

    sev = payload.severity if payload.severity in SEVERITY_LEVELS else "major"
    dept = _issuer_dept_of(current)   # dept penerbit otomatis dari role
    source = payload.source if payload.source in NC_SOURCES else "in_house"

    # Bila user tujuan spesifik dipilih → langsung Assigned.
    assigned_to = None
    status = STATUS_OPEN
    if (payload.issued_to_user_id or "").strip():
        assigned_to = {"id": payload.issued_to_user_id.strip(),
                       "name": (payload.issued_to_user_name or "").strip(),
                       "role": ""}
        status = STATUS_ASSIGNED

    to_label = DEPARTMENTS.get(to_dept, {}).get("label", to_dept)
    issued_to_display = (payload.issued_to or "").strip() or (
        f"{to_label}" + (f" · {assigned_to['name']}" if assigned_to and assigned_to.get('name') else ""))

    now = _now_iso()
    doc = {
        "id": str(uuid.uuid4()),
        "nc_no": await _next_nc_no(),
        "status": status,
        # ── Header ──
        "issuer_dept": dept,
        "issued_by": _actor(current),
        "issued_at": now,               # Date of Issue + basis bulan KPI #1
        "issued_to_dept": to_dept,
        "issued_to_user": assigned_to.copy() if assigned_to else None,
        "issued_to": issued_to_display,
        "expected_reply_date": (payload.expected_reply_date or "").strip(),
        # ── Objek NC ──
        "link_type": link_type,
        "object_ref": (payload.object_ref or "").strip(),
        "incoming": incoming,
        "drawings": resolved,
        "drawing_ids": [r["drawing_id"] for r in resolved if r.get("drawing_id")],
        "drawing_nos": [r["drawing_no"] for r in resolved if r.get("drawing_no")],
        "so_no": so_no,
        "customer_name": customer_name,
        # ── Section 1 ──
        "title": (payload.title or "").strip(),
        "description": (payload.description or "").strip(),
        "source": source,
        "severity": sev,
        # ── Section 2 & 3 ──
        "investigation": None,
        "closeout": None,
        # ── Follow-up ──
        "assigned_to": assigned_to,
        "ecn_id": "",
        "ecn_no": "",
        "closed_at": None,
        "closed_by": None,
        "extra": payload.extra or {},
        "timeline": [{
            "at": now, "action": "created", "by": _actor(current),
            "notes": f"CAR diterbitkan oleh {DEPARTMENTS.get(dept, {}).get('label', dept)} → ditujukan ke {to_label}"
                     + (f" ({assigned_to['name']})" if assigned_to and assigned_to.get('name') else ""),
        }],
        "created_at": now,
        "updated_at": now,
    }
    await db.nonconformances.insert_one(doc.copy())
    await log_action(current, "nc_create", "nonconformances", doc["id"],
                     {"nc_no": doc["nc_no"], "drawings": doc["drawing_nos"]})
    return _clean(doc)


# ── LIST ─────────────────────────────────────────────────────────────────────
@router.get("/nonconformance")
async def list_nc(
    status: Optional[str] = None,
    issuer_dept: Optional[str] = None,
    issued_to_dept: Optional[str] = None,
    link_type: Optional[str] = None,
    drawing_no: Optional[str] = None,
    assignee_id: Optional[str] = None,
    mine: bool = False,
    q: Optional[str] = None,
    month: Optional[str] = Query(None, description="Filter bulan YYYY-MM (berdasar issued_at)"),
    current: dict = Depends(get_current_user),
):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    filt: dict = {"deleted_at": {"$exists": False}}
    if status:
        filt["status"] = status
    if issuer_dept:
        filt["issuer_dept"] = issuer_dept
    if issued_to_dept:
        filt["issued_to_dept"] = issued_to_dept
    if link_type:
        filt["link_type"] = link_type
    if drawing_no:
        filt["drawing_nos"] = {"$regex": drawing_no.strip(), "$options": "i"}
    if assignee_id:
        filt["assigned_to.id"] = assignee_id
    if mine:
        filt["issued_by.id"] = current.get("id")
    if month and len(month) == 7:
        filt["issued_at"] = {"$regex": f"^{month}"}
    if q and q.strip():
        rx = {"$regex": q.strip(), "$options": "i"}
        filt["$or"] = [{"nc_no": rx}, {"title": rx}, {"description": rx}, {"object_ref": rx},
                       {"so_no": rx}, {"customer_name": rx}, {"drawing_nos": rx}, {"issued_to": rx}]
    docs = await db.nonconformances.find(filt, {"_id": 0}).sort("created_at", -1).limit(500).to_list(length=500)
    return {"items": docs, "total": len(docs)}


# ── DEPARTMENTS + USER TUJUAN (untuk dropdown Issued To / assign) ─────────────
@router.get("/nonconformance/departments")
async def list_departments(current: dict = Depends(get_current_user)):
    return {"departments": [{"key": k, "label": v["label"]} for k, v in DEPARTMENTS.items()]}


@router.get("/nonconformance/assignable-users")
async def assignable_users(dept: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Daftar user (id,name,role) untuk 'Issued To user' / assign. Filter per dept."""
    roles = None
    if dept and dept in DEPARTMENTS:
        roles = DEPARTMENTS[dept]["roles"]
    q = {"active": {"$ne": False}, "deleted_at": {"$exists": False}}
    if roles:
        q["role"] = {"$in": roles}
    users = await db.users.find(q, {"_id": 0, "id": 1, "name": 1, "username": 1, "role": 1}).to_list(length=300)
    users.sort(key=lambda u: (u.get("name") or u.get("username") or "").lower())
    return {"users": [{"id": u["id"], "name": u.get("name") or u.get("username"), "role": u.get("role")} for u in users]}


@router.get("/nonconformance/incoming-goods")
async def incoming_goods(q: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Daftar Incoming Goods (store_receipts) untuk dipilih saat NC bertipe Incoming Material/Goods.
    Dapat diakses semua user agar CAR universal bisa menautkan barang masuk nyata."""
    filt = {"deleted_at": {"$exists": False}}
    if q and q.strip():
        rx = {"$regex": q.strip(), "$options": "i"}
        filt["$or"] = [{"item_name": rx}, {"vendor_name": rx}, {"invoice_no": rx},
                       {"po_no": rx}, {"do_number": rx}, {"so_no": rx}]
    docs = await db.store_receipts.find(
        filt, {"_id": 0, "id": 1, "item_name": 1, "vendor_name": 1, "receive_date": 1,
               "qty_received": 1, "unit": 1, "po_no": 1, "invoice_no": 1, "do_number": 1, "so_no": 1},
    ).sort("receive_date", -1).limit(300).to_list(length=300)
    return {"items": docs, "total": len(docs)}


# ── STATS (untuk badge/queue) ─────────────────────────────────────────────────
@router.get("/nonconformance/stats")
async def nc_stats(current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    base = {"deleted_at": {"$exists": False}}
    out = {"total": await db.nonconformances.count_documents(base)}
    for s in VALID_STATUSES:
        out[s] = await db.nonconformances.count_documents({**base, "status": s})
    # NC belum tuntas (untuk badge Engineering Leader)
    out["open_or_active"] = await db.nonconformances.count_documents(
        {**base, "status": {"$in": [STATUS_OPEN, STATUS_ASSIGNED, STATUS_IN_PROGRESS]}})
    return out


# ── ENG-006 NC LOG (Internal Engineering Process — tab NC) ────────────────────
async def _eng006_rows(month: Optional[str]) -> list:
    filt: dict = {"deleted_at": {"$exists": False}}
    if month and len(month) == 7:
        filt["issued_at"] = {"$regex": f"^{month}"}
    docs = await db.nonconformances.find(filt, {"_id": 0}).sort("issued_at", -1).to_list(length=1000)
    rows = []
    for d in docs:
        inv = d.get("investigation") or {}
        rows.append({
            "nc_no": d.get("nc_no"),
            "so_no": d.get("so_no") or "",
            "date": (d.get("issued_at") or "")[:10],
            "drawing_nos": d.get("drawing_nos") or [],
            "root_cause": inv.get("root_cause") or "",
            "status": d.get("status"),
            "preventive_action": inv.get("preventive_action") or "",
            "corrective_action": inv.get("corrective_action") or "",
            "ecn_no": d.get("ecn_no") or "",
        })
    return rows


@router.get("/nonconformance/eng006-nc-log")
async def eng006_nc_log(month: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Sajikan data NC untuk dicatat ke Form MKS-F-ENG-006 (Internal Engineering Process),
    tab 'NC': SO No | Date | Root Cause | Status | Preventive Action | Corrective Action.
    Alur: NC drawing → Engineer terbit ECN (MKS-F-ENG-004) → input ke ENG-006."""
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering/Admin")
    rows = await _eng006_rows(month)
    return {"rows": rows, "total": len(rows)}


@router.get("/nonconformance/eng006-nc-log/excel")
async def eng006_nc_log_excel(month: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Export log Internal Engineering Process (tab NC) ke Excel (MKS-F-ENG-006)."""
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering/Admin")
    rows = await _eng006_rows(month)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    cols = [
        ("nc_no", "No CAR"), ("so_no", "No SO"), ("date", "Tanggal"),
        ("drawing", "Drawing"), ("root_cause", "Root Cause"),
        ("corrective_action", "Corrective Action"), ("preventive_action", "Preventive Action"),
        ("status", "Status"), ("ecn_no", "No ECN"),
    ]
    status_label = {v: k for k, v in STATUS_LABELS.items()}  # noqa: F841

    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Eng Process - NC"
    ws.cell(row=1, column=1, value="MKS-F-ENG-006 Internal Engineering Process — Nonconformance (NC)").font = \
        Font(bold=True, size=13, color="1E293B")
    meta = f"Periode: {month or 'Semua'} · Dicetak: {datetime.now(timezone.utc).strftime('%d %b %Y')} · Total: {len(rows)} NC"
    ws.cell(row=2, column=1, value=meta).font = Font(size=9, italic=True, color="64748B")

    header_row = 4
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, (_, label) in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=ci, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for ri, row in enumerate(rows, start=header_row + 1):
        values = {
            "nc_no": row["nc_no"], "so_no": row["so_no"], "date": row["date"],
            "drawing": ", ".join(row["drawing_nos"]) if row["drawing_nos"] else "",
            "root_cause": row["root_cause"], "corrective_action": row["corrective_action"],
            "preventive_action": row["preventive_action"],
            "status": STATUS_LABELS.get(row["status"], row["status"]),
            "ecn_no": row["ecn_no"],
        }
        for ci, (key, _) in enumerate(cols, start=1):
            c = ws.cell(row=ri, column=ci, value=values.get(key, ""))
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border

    widths = [20, 12, 12, 24, 34, 34, 34, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    fname = f"MKS-F-ENG-006_NC_{month or 'all'}.xlsx"
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── TEMPLATE WORD (.docx) CAR — dikelola Admin/Management ─────────────────────
# Bila ada template aktif, "Cetak PDF" akan mengisi data ke template lalu
# konversi ke PDF (LibreOffice Writer). Jika tidak ada / gagal → fallback bawaan.
_tpl_gridfs: Optional[AsyncIOMotorGridFSBucket] = None


def _tpl_fs() -> AsyncIOMotorGridFSBucket:
    global _tpl_gridfs
    if _tpl_gridfs is None:
        _tpl_gridfs = AsyncIOMotorGridFSBucket(db, bucket_name="car_templates")
    return _tpl_gridfs


def _require_tpl_admin(current: dict):
    if not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Admin/Management yang boleh mengelola template CAR")


async def _active_car_template_bytes() -> Optional[bytes]:
    t = await db.car_templates.find_one({"active": True, "deleted_at": {"$exists": False}})
    if not t:
        return None
    try:
        stream = await _tpl_fs().open_download_stream(ObjectId(t["file_id"]))
        return await stream.read()
    except Exception:
        return None


@router.get("/nonconformance/car-template")
async def get_car_template(current: dict = Depends(get_current_user)):
    """Meta template CAR yang aktif + daftar semua template (Admin/Management)."""
    _require_tpl_admin(current)
    items = await db.car_templates.find(
        {"deleted_at": {"$exists": False}}, {"_id": 0, "file_id": 0},
    ).sort("uploaded_at", -1).to_list(length=50)
    active = next((x for x in items if x.get("active")), None)
    return {"active": active, "items": items}


@router.get("/nonconformance/car-template/fields")
async def car_template_fields(current: dict = Depends(get_current_user)):
    """Daftar placeholder yang tersedia (untuk cheatsheet UI)."""
    _require_tpl_admin(current)
    from utils.car_word import CAR_FIELDS
    return {"fields": [{"key": k, "desc": d} for k, d in CAR_FIELDS]}


@router.get("/nonconformance/car-template/starter")
async def car_template_starter(current: dict = Depends(get_current_user)):
    """Unduh Starter .docx (replika MKS-F-QAD-004 + placeholder)."""
    _require_tpl_admin(current)
    from utils.car_word import build_car_docx_starter
    data = build_car_docx_starter()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": 'attachment; filename="STARTER_CAR_MKS-F-QAD-004.docx"'},
    )


@router.post("/nonconformance/car-template/upload")
async def car_template_upload(
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    """Upload template .docx CAR → langsung AKTIF menggantikan yang lama."""
    _require_tpl_admin(current)
    ext = _ext(file.filename)
    if ext not in (".docx",):
        raise HTTPException(status_code=400, detail="Hanya file Word .docx yang diizinkan")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File > 25 MB tidak diizinkan")
    # Validasi bisa dibuka python-docx
    try:
        from docx import Document
        Document(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="File .docx tidak valid / rusak")

    file_id = await _tpl_fs().upload_from_stream(file.filename, content,
                                                 metadata={"content_type": file.content_type})
    now = _now_iso()
    doc = {
        "id": str(uuid.uuid4()), "filename": file.filename, "file_id": str(file_id),
        "size_bytes": len(content), "active": True,
        "uploaded_at": now, "uploaded_by": current.get("name") or current.get("username"),
    }
    # Nonaktifkan template lain
    await db.car_templates.update_many({"active": True}, {"$set": {"active": False}})
    await db.car_templates.insert_one(doc.copy())
    await log_action(current, "car_template_upload", "car_templates", doc["id"], {"filename": file.filename})
    doc.pop("file_id", None)
    return {"success": True, "template": doc}


@router.post("/nonconformance/car-template/{tid}/activate")
async def car_template_activate(tid: str, current: dict = Depends(get_current_user)):
    _require_tpl_admin(current)
    t = await db.car_templates.find_one({"id": tid, "deleted_at": {"$exists": False}})
    if not t:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    await db.car_templates.update_many({"active": True}, {"$set": {"active": False}})
    await db.car_templates.update_one({"id": tid}, {"$set": {"active": True}})
    return {"success": True}


@router.delete("/nonconformance/car-template/{tid}")
async def car_template_delete(tid: str, current: dict = Depends(get_current_user)):
    _require_tpl_admin(current)
    t = await db.car_templates.find_one({"id": tid, "deleted_at": {"$exists": False}})
    if not t:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    try:
        await _tpl_fs().delete(ObjectId(t["file_id"]))
    except Exception:
        pass
    await db.car_templates.update_one({"id": tid}, {"$set": {
        "deleted_at": _now_iso(), "deleted_by": current.get("username"), "active": False}})
    return {"success": True}


@router.get("/nonconformance/car-template/{tid}/download")
async def car_template_download(tid: str, current: dict = Depends(get_current_user)):
    _require_tpl_admin(current)
    t = await db.car_templates.find_one({"id": tid, "deleted_at": {"$exists": False}})
    if not t:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    stream = await _tpl_fs().open_download_stream(ObjectId(t["file_id"]))
    raw = await stream.read()
    return StreamingResponse(
        io.BytesIO(raw),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{t["filename"]}"'},
    )


async def _car_template_preview_pdf(tid: str) -> bytes:
    """Render sebuah template (id) dengan data contoh → PDF."""
    t = await db.car_templates.find_one({"id": tid, "deleted_at": {"$exists": False}})
    if not t:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    stream = await _tpl_fs().open_download_stream(ObjectId(t["file_id"]))
    raw = await stream.read()
    from utils.car_word import sample_car_data, substitute_docx
    from utils.office_render import office_to_pdf
    sub = substitute_docx(raw, sample_car_data())
    return office_to_pdf(sub, "docx")


@router.get("/nonconformance/car-template/{tid}/preview-page-meta")
async def car_template_preview_meta(tid: str, current: dict = Depends(get_current_user)):
    _require_tpl_admin(current)
    from utils.pdf_render import pdf_page_meta
    raw = await _car_template_preview_pdf(tid)
    return pdf_page_meta(raw)


@router.get("/nonconformance/car-template/{tid}/preview-page-image")
async def car_template_preview_image(tid: str, page: int = 0, scale: float = 2.0,
                                     current: dict = Depends(get_current_user)):
    _require_tpl_admin(current)
    from utils.pdf_render import pdf_page_png
    raw = await _car_template_preview_pdf(tid)
    try:
        png = pdf_page_png(raw, page, scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ditemukan")
    return StreamingResponse(io.BytesIO(png), media_type="image/png",
                             headers={"Cache-Control": "private, max-age=120"})


# ── DETAIL ───────────────────────────────────────────────────────────────────
@router.get("/nonconformance/{nc_id}/pdf")
async def car_pdf(nc_id: str, attachments: bool = True, current: dict = Depends(get_current_user)):
    """Cetak CAR ke PDF sesuai format resmi ISO MKS-F-QAD-004 Rev.02.

    Bila ada template Word (.docx) aktif → data diisi ke template lalu konversi PDF.
    Jika tidak ada / gagal → fallback ke generator bawaan (reportlab).
    Lampiran bukti (foto + PDF) otomatis di-append sebagai halaman setelah form.
    Set ?attachments=false untuk mencetak form saja tanpa lampiran.
    """
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await _get_nc_or_404(nc_id)
    from utils.car_pdf import build_car_pdf, merge_attachments

    pdf = None
    tpl = await _active_car_template_bytes()
    if tpl:
        try:
            from utils.car_word import render_car_pdf_from_docx
            pdf = render_car_pdf_from_docx(
                tpl, doc, printed_by=current.get("name") or current.get("username"))
        except Exception:
            pdf = None  # fallback ke reportlab bila konversi gagal
    if pdf is None:
        pdf = build_car_pdf(doc)

    if attachments:
        atts = await db.nc_attachments.find(
            {"nc_id": nc_id, "deleted_at": {"$exists": False}},
        ).sort("uploaded_at", 1).to_list(length=200)
        merge_list = []
        for a in atts:
            ext = _ext(a.get("filename", "")).lstrip(".")
            if ext not in {"pdf", "jpg", "jpeg", "png", "webp"}:
                continue  # user: hanya foto + PDF yang digabung
            try:
                stream = await _fs().open_download_stream(ObjectId(a["file_id"]))
                raw = await stream.read()
            except Exception:
                continue
            merge_list.append({"filename": a.get("filename"), "content": raw,
                               "remark": a.get("remark")})
        if merge_list:
            pdf = merge_attachments(pdf, merge_list)

    safe = str(doc.get("nc_no") or nc_id).replace("/", "-")
    return StreamingResponse(
        io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="CAR_{safe}.pdf"'},
    )


@router.get("/nonconformance/{nc_id}")
async def get_nc(nc_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    return _clean(await _get_nc_or_404(nc_id))


# ── ASSIGN (Dept tujuan / Admin) ──────────────────────────────────────────────
@router.post("/nonconformance/{nc_id}/assign")
async def assign_nc(nc_id: str, payload: AssignIn, current: dict = Depends(get_current_user)):
    doc = await _get_nc_or_404(nc_id)
    if doc["status"] == STATUS_CLOSED:
        raise HTTPException(status_code=400, detail="NC sudah Closed")
    if not (is_admin_like(current) or _is_target(current, doc) or _is_initiator(current, doc)):
        raise HTTPException(status_code=403, detail="Hanya Admin, dept tujuan, atau penerbit yang bisa assign NC")

    assignee = await db.users.find_one({"id": payload.assignee_id}, {"_id": 0, "password_hash": 0})
    if not assignee:
        raise HTTPException(status_code=404, detail="User assignee tidak ditemukan")

    now = _now_iso()
    assigned = {"id": assignee["id"], "name": assignee.get("name") or assignee.get("username"),
                "role": assignee.get("role")}
    await db.nonconformances.update_one({"id": nc_id}, {
        "$set": {"assigned_to": assigned, "status": STATUS_ASSIGNED, "updated_at": now},
        "$push": {"timeline": {"at": now, "action": "assigned", "by": _actor(current),
                               "notes": f"Ditugaskan ke {assigned['name']}. {payload.notes}".strip()}},
    })
    await log_action(current, "nc_assign", "nonconformances", nc_id,
                     {"nc_no": doc.get("nc_no"), "assignee": assigned["name"]})
    return {"success": True, "status": STATUS_ASSIGNED, "assigned_to": assigned}


# ── UPDATE STATUS ─────────────────────────────────────────────────────────────
@router.post("/nonconformance/{nc_id}/status")
async def update_status(nc_id: str, payload: StatusIn, current: dict = Depends(get_current_user)):
    new_status = (payload.status or "").strip().lower()
    if new_status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Status tidak valid. Pilih: {sorted(VALID_STATUSES)}")
    doc = await _get_nc_or_404(nc_id)

    is_priv = is_admin_like(current)
    is_init = _is_initiator(current, doc)
    is_tgt = _is_target(current, doc)
    if new_status == STATUS_CLOSED and not (is_priv or is_mr(current)):
        raise HTTPException(status_code=403, detail="Hanya MR / Document Control / Admin yang bisa menutup (Closed) CAR")
    if new_status == STATUS_IN_PROGRESS and not (is_priv or is_tgt):
        raise HTTPException(status_code=403, detail="Hanya dept tujuan/assignee/Admin yang bisa set In Progress")
    if new_status in (STATUS_OPEN, STATUS_ASSIGNED) and not (is_priv or is_tgt or is_init):
        raise HTTPException(status_code=403, detail="Anda tidak berwenang mengubah status ini")

    now = _now_iso()
    updates = {"status": new_status, "updated_at": now}
    if payload.ecn_id or payload.ecn_no:
        updates["ecn_id"] = (payload.ecn_id or "").strip()
        updates["ecn_no"] = (payload.ecn_no or "").strip()
    if new_status == STATUS_CLOSED:
        updates["closed_at"] = now
        updates["closed_by"] = _actor(current)

    note_txt = f"Status → {STATUS_LABELS[new_status]}."
    if payload.ecn_no:
        note_txt += f" ECN: {payload.ecn_no}."
    if payload.notes:
        note_txt += f" {payload.notes}"
    await db.nonconformances.update_one({"id": nc_id}, {
        "$set": updates,
        "$push": {"timeline": {"at": now, "action": f"status_{new_status}",
                               "by": _actor(current), "notes": note_txt.strip()}},
    })
    await log_action(current, "nc_status", "nonconformances", nc_id,
                     {"nc_no": doc.get("nc_no"), "status": new_status, "ecn_no": payload.ecn_no})
    return {"success": True, "status": new_status}


# ── ADD NOTE (timeline) ───────────────────────────────────────────────────────
@router.post("/nonconformance/{nc_id}/note")
async def add_note(nc_id: str, payload: NoteIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    if not (payload.notes or "").strip():
        raise HTTPException(status_code=400, detail="Catatan kosong")
    doc = await _get_nc_or_404(nc_id)
    now = _now_iso()
    await db.nonconformances.update_one({"id": nc_id}, {
        "$set": {"updated_at": now},
        "$push": {"timeline": {"at": now, "action": "note", "by": _actor(current),
                               "notes": payload.notes.strip()}},
    })
    return {"success": True}


# ── SECTION 2: INVESTIGATION & ACTION PLANS (Responsible Dept./Assignee) ──────
@router.post("/nonconformance/{nc_id}/investigation")
async def save_investigation(nc_id: str, payload: InvestigationIn, current: dict = Depends(get_current_user)):
    doc = await _get_nc_or_404(nc_id)
    if doc["status"] == STATUS_CLOSED:
        raise HTTPException(status_code=400, detail="NC sudah Closed, tidak bisa diubah")
    if not (is_admin_like(current) or _is_target(current, doc)):
        raise HTTPException(status_code=403, detail="Hanya dept tujuan/assignee/Admin yang bisa mengisi investigasi")

    now = _now_iso()
    investigation = {
        "root_cause": (payload.root_cause or "").strip(),
        "immediate_action": (payload.immediate_action or "").strip(),
        "corrective_action": (payload.corrective_action or "").strip(),
        "preventive_action": (payload.preventive_action or "").strip(),
        "completed_by": (payload.completed_by or "").strip(),
        "completed_date": (payload.completed_date or "").strip(),
        "dept_head_name": (payload.dept_head_name or "").strip(),
        "dept_head_date": (payload.dept_head_date or "").strip(),
        "saved_by": _actor(current),
        "saved_at": now,
    }
    updates = {"investigation": investigation, "updated_at": now}
    if (payload.ecn_no or "").strip():
        updates["ecn_no"] = payload.ecn_no.strip()
    if payload.set_in_progress and doc["status"] in (STATUS_OPEN, STATUS_ASSIGNED):
        updates["status"] = STATUS_IN_PROGRESS
    await db.nonconformances.update_one({"id": nc_id}, {
        "$set": updates,
        "$push": {"timeline": {"at": now, "action": "investigation", "by": _actor(current),
                               "notes": "Investigasi & rencana tindakan disimpan"}},
    })
    await log_action(current, "nc_investigation", "nonconformances", nc_id, {"nc_no": doc.get("nc_no")})
    return {"success": True, "status": updates.get("status", doc["status"])}


# ── SECTION 3: CAR CLOSEOUT (Initiator / MR / QA) ─────────────────────────────
@router.post("/nonconformance/{nc_id}/closeout")
async def save_closeout(nc_id: str, payload: CloseoutIn, current: dict = Depends(get_current_user)):
    doc = await _get_nc_or_404(nc_id)
    # Section 3 (Closeout) diisi oleh MR / Document Control (mis. salma) / Admin.
    # Form resmi: "Completed by Initiator or MR" → Initiator juga diizinkan.
    is_priv = is_admin_like(current)
    is_initiator = _is_initiator(current, doc)
    if not (is_priv or is_mr(current) or is_initiator):
        raise HTTPException(status_code=403, detail="Section 3 (Closeout) hanya untuk MR / Document Control / Initiator / Admin")

    now = _now_iso()
    closeout = {
        "initiator_remarks": (payload.initiator_remarks or "").strip(),
        "risk_review": bool(payload.risk_review),
        "risk_attached": bool(payload.risk_attached),
        "effectiveness_reviewed_by": (payload.effectiveness_reviewed_by or "").strip(),
        "effectiveness_date": (payload.effectiveness_date or "").strip(),
        "qa_approved_by": (payload.qa_approved_by or "").strip(),
        "qa_date": (payload.qa_date or "").strip(),
        "saved_by": _actor(current),
        "saved_at": now,
    }
    updates = {"closeout": closeout, "updated_at": now}
    action = "closeout"
    if payload.close:
        if not (is_priv or is_mr(current)):
            raise HTTPException(status_code=403, detail="Hanya MR / Document Control / Admin yang bisa menutup (Closed) CAR")
        updates["status"] = STATUS_CLOSED
        updates["closed_at"] = now
        updates["closed_by"] = _actor(current)
        action = "status_closed"
    await db.nonconformances.update_one({"id": nc_id}, {
        "$set": updates,
        "$push": {"timeline": {"at": now, "action": action, "by": _actor(current),
                               "notes": "Closeout disimpan" + (" & NC ditutup (Closed)" if payload.close else "")}},
    })
    await log_action(current, "nc_closeout", "nonconformances", nc_id,
                     {"nc_no": doc.get("nc_no"), "closed": payload.close})
    return {"success": True, "status": updates.get("status", doc["status"])}


# ── DELETE (soft) ─────────────────────────────────────────────────────────────
@router.delete("/nonconformance/{nc_id}")
async def delete_nc(nc_id: str, current: dict = Depends(get_current_user)):
    doc = await _get_nc_or_404(nc_id)
    # Penerbit boleh membatalkan selama masih Open; Admin boleh kapan saja.
    is_owner = (doc.get("issued_by") or {}).get("id") == current.get("id")
    if not (is_admin_like(current) or (is_owner and doc["status"] == STATUS_OPEN)):
        raise HTTPException(status_code=403, detail="Hanya Admin atau penerbit (saat status Open) yang bisa menghapus NC")
    await db.nonconformances.update_one({"id": nc_id}, {"$set": {
        "deleted_at": _now_iso(), "deleted_by": current.get("username")}})
    await log_action(current, "nc_delete", "nonconformances", nc_id, {"nc_no": doc.get("nc_no")})
    return {"success": True}


# ── ATTACHMENTS (bukti foto/pdf) ──────────────────────────────────────────────
ATTACH_ALLOWED_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".xlsx", ".xls", ".doc", ".docx"}


@router.get("/nonconformance/{nc_id}/attachments")
async def list_nc_attachments(nc_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    await _get_nc_or_404(nc_id)
    docs = await db.nc_attachments.find(
        {"nc_id": nc_id, "deleted_at": {"$exists": False}},
        {"_id": 0, "file_id": 0},
    ).sort("uploaded_at", -1).to_list(length=200)
    return {"nc_id": nc_id, "items": docs, "total": len(docs)}


@router.post("/nonconformance/{nc_id}/attachments")
async def upload_nc_attachment(
    nc_id: str,
    file: UploadFile = File(...),
    remark: str = Form(""),
    current: dict = Depends(get_current_user),
):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await _get_nc_or_404(nc_id)
    ext = _ext(file.filename)
    if ext not in ATTACH_ALLOWED_EXT:
        raise HTTPException(status_code=400, detail=f"Ekstensi {ext} tidak diizinkan. Boleh: {sorted(ATTACH_ALLOWED_EXT)}")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File > 50 MB tidak diizinkan")

    fs = _fs()
    file_id = await fs.upload_from_stream(
        file.filename, content,
        metadata={"content_type": file.content_type, "nc_id": nc_id},
    )
    att = {
        "id": str(uuid.uuid4()),
        "nc_id": nc_id,
        "filename": file.filename,
        "file_id": str(file_id),
        "content_type": file.content_type,
        "size_bytes": len(content),
        "remark": remark or "",
        "uploaded_at": _now_iso(),
        "uploaded_by": current.get("name") or current.get("username"),
    }
    await db.nc_attachments.insert_one(att.copy())
    await log_action(current, "nc_attachment_upload", "nonconformances", nc_id,
                     {"nc_no": doc.get("nc_no"), "filename": file.filename})
    att.pop("file_id", None)
    return {"success": True, "attachment": att}


async def _nc_attach_or_404(nc_id: str, attach_id: str) -> dict:
    a = await db.nc_attachments.find_one({"id": attach_id, "nc_id": nc_id, "deleted_at": {"$exists": False}})
    if not a:
        raise HTTPException(status_code=404, detail="Lampiran tidak ditemukan")
    return a


async def _nc_pdf_bytes(nc_id: str, attach_id: str) -> bytes:
    a = await _nc_attach_or_404(nc_id, attach_id)
    ext = _ext(a["filename"]).lstrip(".")
    stream = await _fs().open_download_stream(ObjectId(a["file_id"]))
    raw = await stream.read()
    if ext == "pdf":
        return raw
    from utils.office_render import is_office_ext, office_to_pdf
    if is_office_ext(ext):
        return office_to_pdf(raw, ext)
    # Gambar (jpg/png/webp) → bungkus ke PDF via PyMuPDF agar viewer image-based konsisten
    if ext in ("jpg", "jpeg", "png", "webp"):
        import fitz  # PyMuPDF
        img_doc = fitz.open(stream=raw, filetype=ext)
        pdf_bytes = img_doc.convert_to_pdf()
        return pdf_bytes
    raise HTTPException(status_code=400, detail=f"Preview belum didukung untuk .{ext}")


@router.get("/nonconformance/{nc_id}/attachments/{attach_id}/page-meta")
async def nc_attach_page_meta(nc_id: str, attach_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    from utils.pdf_render import pdf_page_meta
    raw = await _nc_pdf_bytes(nc_id, attach_id)
    try:
        return pdf_page_meta(raw)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="File tidak dapat dirender sebagai halaman gambar")


@router.get("/nonconformance/{nc_id}/attachments/{attach_id}/page-image")
async def nc_attach_page_image(nc_id: str, attach_id: str, page: int = 0, scale: float = 2.0,
                               current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    from utils.pdf_render import pdf_page_png
    raw = await _nc_pdf_bytes(nc_id, attach_id)
    try:
        png = pdf_page_png(raw, page, scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ditemukan")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="File tidak dapat dirender sebagai halaman gambar")
    return StreamingResponse(io.BytesIO(png), media_type="image/png",
                             headers={"Cache-Control": "private, max-age=300"})


@router.get("/nonconformance/{nc_id}/attachments/{attach_id}/download")
async def download_nc_attachment(nc_id: str, attach_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    a = await _nc_attach_or_404(nc_id, attach_id)
    stream = await _fs().open_download_stream(ObjectId(a["file_id"]))
    raw = await stream.read()
    return StreamingResponse(
        io.BytesIO(raw),
        media_type=a.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{a["filename"]}"'},
    )


@router.delete("/nonconformance/{nc_id}/attachments/{attach_id}")
async def delete_nc_attachment(nc_id: str, attach_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    a = await _nc_attach_or_404(nc_id, attach_id)
    try:
        await _fs().delete(ObjectId(a["file_id"]))
    except Exception:
        pass
    await db.nc_attachments.update_one({"id": attach_id}, {"$set": {
        "deleted_at": _now_iso(), "deleted_by": current.get("username")}})
    return {"success": True}
