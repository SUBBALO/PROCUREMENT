"""Deliveries + Sales Orders routes."""
import re
import uuid
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel

from db import db
from deps import _now_iso, get_current_user, log_action, require_store_write, require_write, is_admin_like
from models import DeliveryCreate, SOCreate
from routers.bom import normalize_so_no
from services.soft_delete import NOT_DELETED_FILTER, merged, soft_delete_one

router = APIRouter(tags=["orders"])

# ---- Sales Order (rich) helpers ----
PRICE_ROLES = {"super_admin", "admin", "finance"}
SO_CREATE_ROLES = {"sales", "admin", "super_admin", "finance", "supervisor"}
# Peta status DRF -> status proses SO Engineering
_DRF_TO_SO_STATUS = {
    "submitted": "submitted_eng",
    "accepted": "eng_terima",
    "in_progress": "eng_kerjakan",
    "completed": "selesai_eng",
}
_SO_STATUS_RANK = {
    "belum_drawing_request": 0, "submitted_eng": 1, "eng_terima": 2,
    "eng_kerjakan": 3, "selesai_eng": 4,
}


def _can_see_price(current: dict) -> bool:
    return current.get("role") in PRICE_ROLES


def _strip_price_if_needed(doc: dict, current: dict) -> dict:
    """Sembunyikan harga untuk role selain super_admin/admin/finance."""
    if _can_see_price(current):
        return doc
    doc.pop("total_amount", None)
    doc.pop("currency", None)
    for it in (doc.get("items") or []):
        it.pop("price", None)
        it.pop("line_total", None)
    doc["_price_hidden"] = True
    return doc


async def _so_status_map() -> dict:
    """Map so_no -> status proses eng terbaik (dari drawing_requests)."""
    drfs = await db.drawing_requests.find(
        {"deleted_at": {"$exists": False}},
        {"_id": 0, "so_no": 1, "status": 1},
    ).to_list(length=5000)
    m: dict = {}
    for d in drfs:
        so = d.get("so_no")
        if not so:
            continue
        st = _DRF_TO_SO_STATUS.get(d.get("status"), None)
        if not st:
            continue
        if _SO_STATUS_RANK.get(st, 0) > _SO_STATUS_RANK.get(m.get(so, "belum_drawing_request"), 0):
            m[so] = st
    return m


_DWG_DONE_STATES = {"approved", "controlled", "released"}


async def _so_drawings_map() -> dict:
    """Map so_no -> {list ringkas drawing, count, done} untuk ditampilkan di baris SO."""
    dws = await db.drawings.find(
        {"deleted_at": {"$exists": False}, "so_no": {"$nin": [None, ""]}},
        {"_id": 0, "so_no": 1, "drawing_no": 1, "approval_status": 1},
    ).sort("drawing_no", 1).to_list(length=5000)
    m: dict = {}
    for d in dws:
        so = d.get("so_no")
        if not so:
            continue
        e = m.setdefault(so, {"drawings": [], "count": 0, "done": 0})
        e["count"] += 1
        if d.get("approval_status") in _DWG_DONE_STATES:
            e["done"] += 1
        if len(e["drawings"]) < 8:
            e["drawings"].append({"drawing_no": d.get("drawing_no"), "status": d.get("approval_status")})
    return m


class SOItemIn(BaseModel):
    name: str = ""
    qty: float = 1
    unit: str = "pcs"
    price: float = 0


class SOFullCreate(BaseModel):
    so_no: str
    so_date: str
    customer: str
    customer_address: Optional[str] = ""
    po_customer_no: Optional[str] = ""
    description: Optional[str] = ""
    sales_name: Optional[str] = ""
    currency: Optional[str] = "IDR"
    source_quotation_id: Optional[str] = ""
    source_quotation_no: Optional[str] = ""
    items: List[SOItemIn] = []


@router.get("/sales-users")
async def list_sales_users(current: dict = Depends(get_current_user)):
    """Daftar ringkas user untuk pemilihan 'Nama Sales' (Sales/Admin/Super Admin/Supervisor/Finance)."""
    users = await db.users.find(
        {"role": {"$in": ["sales", "admin", "super_admin", "supervisor", "finance"]},
         "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "name": 1, "username": 1, "role": 1},
    ).sort("name", 1).to_list(length=500)
    items = [{"id": u["id"], "name": u.get("name") or u.get("username"), "username": u.get("username"), "role": u.get("role")} for u in users]
    return {"items": items}



# ---------------- Deliveries (Pengiriman Barang - log only) ----------------
@router.post("/deliveries")
async def create_delivery(payload: DeliveryCreate, current: dict = Depends(require_store_write)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Minimal 1 item")
    if not payload.destination.strip():
        raise HTTPException(status_code=400, detail="Tujuan wajib diisi")
    doc = {
        "id": str(uuid.uuid4()),
        "delivery_date": payload.delivery_date,
        "gate_pass_no": payload.gate_pass_no or "",
        "do_no": payload.do_no or "",
        "destination": payload.destination.strip(),
        "driver_name": payload.driver_name or "",
        "items": [it.model_dump() for it in payload.items],
        "remark": payload.remark or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.deliveries.insert_one(doc.copy())
    await log_action(current, "create_delivery", "delivery", doc["id"], {
        "destination": doc["destination"], "gate_pass": doc["gate_pass_no"], "items": len(doc["items"]),
    })
    doc.pop("_id", None)
    return doc


@router.get("/deliveries")
async def list_deliveries(
    current: dict = Depends(get_current_user),
    q: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"destination": {"$regex": q, "$options": "i"}},
            {"gate_pass_no": {"$regex": q, "$options": "i"}},
            {"do_no": {"$regex": q, "$options": "i"}},
            {"driver_name": {"$regex": q, "$options": "i"}},
        ]
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        filt["delivery_date"] = d
    total = await db.deliveries.count_documents(merged(filt, NOT_DELETED_FILTER))
    items = await db.deliveries.find(merged(filt, NOT_DELETED_FILTER), {"_id": 0}).sort("delivery_date", -1).skip((page - 1) * page_size).limit(page_size).to_list(length=page_size)
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.delete("/deliveries/{did}")
async def delete_delivery(did: str, current: dict = Depends(require_store_write)):
    """Admin + store role only (matches create permission)."""
    ok = await soft_delete_one("deliveries", {"id": did}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="Pengiriman tidak ditemukan")
    await log_action(current, "delete_delivery", "delivery", did, {})
    return {"ok": True}


@router.get("/deliveries/xlsx")
async def deliveries_xlsx(
    current: dict = Depends(get_current_user),
    q: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """Excel export of Pengiriman Barang — flattened, 1 row per item."""
    import io
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    filt: dict = {}
    if q:
        filt["$or"] = [
            {"destination": {"$regex": q, "$options": "i"}},
            {"gate_pass_no": {"$regex": q, "$options": "i"}},
            {"do_no": {"$regex": q, "$options": "i"}},
            {"driver_name": {"$regex": q, "$options": "i"}},
        ]
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date
        filt["delivery_date"] = rng
    docs = await db.deliveries.find(filt, {"_id": 0}).sort("delivery_date", -1).to_list(length=100000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pengiriman Barang"
    headers = ["Tgl", "No Gatepass", "No DO", "Nama Tujuan", "Nomor SO", "Nama Barang", "Qty", "Unit", "Supir", "Remark"]
    ws.append(headers)
    for d in docs:
        items = d.get("items") or []
        if not items:
            ws.append([d.get("delivery_date", ""), d.get("gate_pass_no", ""), d.get("do_no", ""),
                       d.get("destination", ""), "", "", 0, "", d.get("driver_name", ""), d.get("remark", "")])
        else:
            for it in items:
                ws.append([d.get("delivery_date", ""), d.get("gate_pass_no", ""), d.get("do_no", ""),
                           d.get("destination", ""), it.get("so_no", ""), it.get("item_name", ""),
                           float(it.get("qty", 0)), it.get("unit", ""),
                           d.get("driver_name", ""), d.get("remark", "")])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pengiriman_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------- Master Sales Order ----------------
@router.get("/sales-orders")
async def list_sales_orders(current: dict = Depends(get_current_user), q: Optional[str] = None):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"so_no": {"$regex": q, "$options": "i"}},
            {"customer": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}},
        ]
    docs = await db.sales_orders.find(merged(filt, NOT_DELETED_FILTER), {"_id": 0}).sort("so_no", 1).to_list(length=5000)
    status_map = await _so_status_map()
    dwg_map = await _so_drawings_map()
    for d in docs:
        d["drawing_request_status"] = status_map.get(d.get("so_no"), "belum_drawing_request")
        dw = dwg_map.get(d.get("so_no"))
        d["drawings"] = dw["drawings"] if dw else []
        d["drawing_count"] = dw["count"] if dw else 0
        d["drawing_done"] = dw["done"] if dw else 0
        _strip_price_if_needed(d, current)
    return docs


def _compute_items(items: List[SOItemIn]) -> tuple:
    out = []
    total = 0.0
    for it in items:
        qty = float(it.qty or 0)
        price = float(it.price or 0)
        line = qty * price
        total += line
        out.append({"name": (it.name or "").strip(), "qty": qty, "unit": it.unit or "pcs", "price": price, "line_total": line})
    return out, total


@router.post("/sales-orders/full")
async def create_so_full(payload: SOFullCreate, current: dict = Depends(get_current_user)):
    """Buat Sales Order lengkap (dari quotation / manual) dengan item + harga + PO customer.
    Boleh: Sales, Admin, Super Admin, Finance, Supervisor."""
    if current.get("role") not in SO_CREATE_ROLES and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Anda tidak berwenang membuat Sales Order")
    raw = (payload.so_no or "").strip()
    if not re.fullmatch(r"00\d{4}", raw):
        raise HTTPException(status_code=400, detail="Nomor SO wajib 6 digit dan diawali '00' (mis. 005251)")
    so_no = normalize_so_no(raw)
    if await db.sales_orders.find_one({"so_no": so_no, "deleted_at": {"$exists": False}}):
        raise HTTPException(status_code=400, detail=f"Nomor SO {so_no} sudah ada")
    items, total = _compute_items(payload.items)
    doc = {
        "id": str(uuid.uuid4()),
        "so_no": so_no,
        "so_date": payload.so_date or _now_iso()[:10],
        "customer": (payload.customer or "").strip(),
        "customer_address": (payload.customer_address or "").strip(),
        "po_customer_no": (payload.po_customer_no or "").strip(),
        "description": payload.description or "",
        "sales_name": (payload.sales_name or current.get("name") or current.get("username", "")).strip(),
        "currency": payload.currency or "IDR",
        "source_quotation_no": payload.source_quotation_no or "",
        "items": items,
        "total_amount": total,
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_by_name": current.get("name") or current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.sales_orders.insert_one(doc.copy())
    await log_action(current, "create_so_full", "sales_order", doc["id"], {"so_no": so_no, "customer": doc["customer"], "items": len(items)})
    doc.pop("_id", None)
    doc["drawing_request_status"] = "belum_drawing_request"
    return _strip_price_if_needed(doc, current)


@router.put("/sales-orders/{sid}/full")
async def update_so_full(sid: str, payload: SOFullCreate, current: dict = Depends(get_current_user)):
    if current.get("role") not in SO_CREATE_ROLES and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Anda tidak berwenang mengubah Sales Order")
    so = await db.sales_orders.find_one({"id": sid, "deleted_at": {"$exists": False}})
    if not so:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    raw = (payload.so_no or "").strip()
    if not re.fullmatch(r"00\d{4}", raw):
        raise HTTPException(status_code=400, detail="Nomor SO wajib 6 digit dan diawali '00' (mis. 005251)")
    so_no = normalize_so_no(raw)
    dup = await db.sales_orders.find_one({"so_no": so_no, "id": {"$ne": sid}, "deleted_at": {"$exists": False}})
    if dup:
        raise HTTPException(status_code=400, detail=f"Nomor SO {so_no} sudah dipakai SO lain")
    items, total = _compute_items(payload.items)
    upd = {
        "so_no": so_no,
        "so_date": payload.so_date,
        "customer": (payload.customer or "").strip(),
        "customer_address": (payload.customer_address or "").strip(),
        "po_customer_no": (payload.po_customer_no or "").strip(),
        "description": payload.description or "",
        "sales_name": (payload.sales_name or "").strip(),
        "currency": payload.currency or "IDR",
        "items": items,
        "total_amount": total,
        "updated_at": _now_iso(),
        "updated_by_name": current.get("name") or current.get("username", ""),
    }
    await db.sales_orders.update_one({"id": sid}, {"$set": upd})
    await log_action(current, "update_so_full", "sales_order", sid, {"so_no": so_no, "items": len(items)})
    updated = await db.sales_orders.find_one({"id": sid}, {"_id": 0})
    status_map = await _so_status_map()
    updated["drawing_request_status"] = status_map.get(updated.get("so_no"), "belum_drawing_request")
    return _strip_price_if_needed(updated, current)


@router.post("/sales-orders")
async def create_so(payload: SOCreate, current: dict = Depends(require_write)):
    so_no = normalize_so_no(payload.so_no)
    if not so_no:
        raise HTTPException(status_code=400, detail="Nomor SO wajib")
    existing = await db.sales_orders.find_one({"so_no": so_no})
    if existing:
        raise HTTPException(status_code=400, detail=f"SO {so_no} sudah ada")
    doc = {
        "id": str(uuid.uuid4()),
        "so_no": so_no,
        "so_date": payload.so_date,
        "customer": payload.customer.strip(),
        "description": payload.description or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.sales_orders.insert_one(doc.copy())
    await log_action(current, "create_so", "sales_order", doc["id"], {"so_no": so_no, "customer": doc["customer"]})
    doc.pop("_id", None)
    return doc


@router.put("/sales-orders/{sid}")
async def update_so(sid: str, payload: SOCreate, current: dict = Depends(require_write)):
    so = await db.sales_orders.find_one({"id": sid})
    if not so:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    upd = {
        "so_no": normalize_so_no(payload.so_no),
        "so_date": payload.so_date,
        "customer": payload.customer.strip(),
        "description": payload.description or "",
    }
    await db.sales_orders.update_one({"id": sid}, {"$set": upd})
    await log_action(current, "update_so", "sales_order", sid, upd)
    updated = await db.sales_orders.find_one({"id": sid}, {"_id": 0})
    return updated


@router.delete("/sales-orders/{sid}")
async def delete_so(sid: str, current: dict = Depends(require_write)):
    ok = await soft_delete_one("sales_orders", {"id": sid}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    await log_action(current, "delete_so", "sales_order", sid, {})
    return {"ok": True}


class SOBulkDelete(BaseModel):
    ids: List[str]


@router.post("/sales-orders/bulk-delete")
async def bulk_delete_sos(payload: SOBulkDelete, current: dict = Depends(require_write)):
    """Bulk soft-delete a list of Sales Orders. Sales/admin/supervisor allowed."""
    if not payload.ids:
        raise HTTPException(status_code=400, detail="Tidak ada SO yang dipilih")
    deleted = 0
    for sid in payload.ids:
        if await soft_delete_one("sales_orders", {"id": sid}, current):
            deleted += 1
    await log_action(current, "bulk_delete_so", "sales_orders", "-", {"count": deleted, "requested": len(payload.ids)})
    return {"ok": True, "deleted": deleted, "requested": len(payload.ids)}



# ---------------- SO Excel Upload ----------------
@router.post("/sales-orders/import/xlsx")
async def import_sos_xlsx(file: UploadFile = File(...), current: dict = Depends(require_write)):
    """Bulk create SOs from Excel. Expected columns: so_no, so_date, customer, description
    (case-insensitive, flexible header matching)."""
    import io
    from datetime import date, datetime
    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File Excel tidak valid: {e}")

    # Track for storage management
    try:
        from routers.storage import record_temp_upload
        await record_temp_upload(current, "so_import", file.filename or "so.xlsx", content,
                                  mime=file.content_type or "", related_entity="Master SO")
    except Exception:
        pass

    def _to_date(v):
        if v is None or v == "":
            return None
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
        s = str(v).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except Exception:
                pass
        return s

    inserted = 0
    skipped = 0
    errors: List[str] = []
    now = _now_iso()

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_map: dict = {}
        header_row_idx = None
        for i, row in enumerate(rows[:5]):
            row_str = [str(c).strip().lower() if c is not None else "" for c in row]
            if any("so" in c or "sales order" in c or "nomor" in c for c in row_str):
                header_row_idx = i
                for idx, cell in enumerate(row_str):
                    header_map[idx] = cell
                break
        if header_row_idx is None:
            continue

        def find_col(*kws):
            for idx, cell in header_map.items():
                for kw in kws:
                    if kw in cell:
                        return idx
            return None

        col_so = find_col("nomor so", "so_no", "so no", "sales order", "nomor")
        col_date = find_col("tanggal", "date", "so_date")
        col_cust = find_col("customer", "pelanggan")
        col_desc = find_col("description", "deskripsi", "keterangan", "desc")

        docs = []
        for row in rows[header_row_idx + 1:]:
            if row is None or all(c is None or c == "" for c in row):
                continue
            try:
                so_no_raw = row[col_so] if col_so is not None and col_so < len(row) else None
                if not so_no_raw:
                    continue
                so_no = normalize_so_no(so_no_raw)
                # skip duplicates
                if await db.sales_orders.find_one({"so_no": so_no}):
                    skipped += 1
                    continue
                docs.append({
                    "id": str(uuid.uuid4()),
                    "so_no": so_no,
                    "so_date": _to_date(row[col_date]) if col_date is not None and col_date < len(row) else "",
                    "customer": str(row[col_cust]).strip() if col_cust is not None and col_cust < len(row) and row[col_cust] else "",
                    "description": str(row[col_desc]).strip() if col_desc is not None and col_desc < len(row) and row[col_desc] else "",
                    "created_by": current["id"],
                    "created_by_username": current.get("username", ""),
                    "created_at": now,
                })
            except Exception as e:
                errors.append(f"Sheet {sn}: {e}")
        if docs:
            await db.sales_orders.insert_many([d.copy() for d in docs])
            inserted += len(docs)

    await log_action(current, "import_sos", "sales_order", "-", {"inserted": inserted, "skipped": skipped})
    return {"inserted": inserted, "skipped_duplicates": skipped, "errors": errors[:20]}


# ---------------- Delivery Autocomplete ----------------
@router.get("/deliveries/autocomplete")
async def delivery_autocomplete(current: dict = Depends(get_current_user)):
    """Distinct destinations and driver names from historical deliveries."""
    destinations = await db.deliveries.distinct("destination")
    drivers = await db.deliveries.distinct("driver_name")
    return {
        "destinations": sorted([d for d in destinations if d]),
        "drivers": sorted([d for d in drivers if d]),
    }
