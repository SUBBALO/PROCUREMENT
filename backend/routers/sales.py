"""Sales module — Costing Inquiries and Quotations.

Workflow (Inquiries):
  draft → submitted → in_progress (with PIC engineer names) → awaiting_review →
      accepted / revision_requested (loop back) → closed
"""
import io
import re
import uuid
from datetime import datetime
from typing import List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from db import db
from deps import (
    get_current_user, log_action,
    is_admin_like, is_engineering, is_eng_head, is_super_admin_user,
)
from services.soft_delete import NOT_DELETED_FILTER, merged, soft_delete_one

router = APIRouter(tags=["sales"])

# ---------------------------- Storage ----------------------------
_gridfs: Optional[AsyncIOMotorGridFSBucket] = None


def gridfs() -> AsyncIOMotorGridFSBucket:
    global _gridfs
    if _gridfs is None:
        _gridfs = AsyncIOMotorGridFSBucket(db, bucket_name="inquiry_files")
    return _gridfs


# ---------------------------- Models ----------------------------
class InquiryItem(BaseModel):
    item_name: str
    qty: float = 1.0
    unit: str = "Ea"
    specification: str = ""


class InquiryCreate(BaseModel):
    title: str
    customer_name: str
    project_name: str = ""  # nama project (opsional tapi disarankan)
    customer_deadline: Optional[str] = None  # ISO date
    description: str = ""
    items: List[InquiryItem] = []
    save_as_draft: bool = True


class InquiryUpdate(BaseModel):
    title: Optional[str] = None
    customer_name: Optional[str] = None
    project_name: Optional[str] = None
    customer_deadline: Optional[str] = None
    description: Optional[str] = None
    items: Optional[List[InquiryItem]] = None


class InquiryAccept(BaseModel):
    pic_engineer_name: str  # required — the actual engineer person responsible


class InquiryProgress(BaseModel):
    note: str
    status: Optional[str] = "in_progress"  # or "awaiting_review"


class InquiryReview(BaseModel):
    approve: bool  # True → accepted, False → revision_requested
    review_note: str = ""


class InquiryAssign(BaseModel):
    """eng_head assigns inquiry to a specific engineering user (self allowed)."""
    engineer_id: str
    engineer_name: str = ""  # denormalized snapshot; server fills if empty


# ---------------------------- Helpers ----------------------------
ROMAN = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]


async def _next_number(counter_kind: str) -> int:
    """Atomic monthly counter reset. counter_kind: 'inquiry' | 'quotation'."""
    now = datetime.utcnow()
    key = f"{counter_kind}:{now.year}-{now.month:02d}"
    doc = await db.counters.find_one_and_update(
        {"_id": key},
        {"$inc": {"seq": 1}, "$setOnInsert": {"created_at": now.isoformat()}},
        upsert=True,
        return_document=True,
    )
    # After upsert with $inc, seq will be 1 for first, 2 for second, ...
    if not doc:
        doc = await db.counters.find_one({"_id": key})
    return int(doc.get("seq", 1))


async def _new_inquiry_no() -> str:
    now = datetime.utcnow()
    seq = await _next_number("inquiry")
    return f"INQ-{seq:03d}/MKS/{ROMAN[now.month]}/{now.year}"


async def _new_quotation_no() -> str:
    now = datetime.utcnow()
    seq = await _next_number("quotation")
    return f"{seq:03d}/MKS/Q/{ROMAN[now.month]}/{now.year}"


def _clean(doc: dict) -> dict:
    doc.pop("_id", None)
    return doc


# =============================================================================
# INQUIRIES
# =============================================================================
@router.post("/inquiries")
async def create_inquiry(payload: InquiryCreate, current: dict = Depends(get_current_user)):
    if current.get("role") != "sales" and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Sales & Admin yang bisa buat Inquiry")
    if not payload.title.strip():
        raise HTTPException(status_code=400, detail="Judul wajib diisi")
    if not payload.customer_name.strip():
        raise HTTPException(status_code=400, detail="Nama customer wajib diisi")

    now = datetime.utcnow().isoformat()
    status = "draft" if payload.save_as_draft else "submitted"
    doc = {
        "id": str(uuid.uuid4()),
        "inquiry_no": await _new_inquiry_no(),
        "title": payload.title.strip(),
        "customer_name": payload.customer_name.strip(),
        "project_name": (payload.project_name or "").strip(),
        "customer_deadline": payload.customer_deadline or "",
        "description": payload.description.strip(),
        "items": [i.model_dump() for i in payload.items],
        "attachments": [],  # list of {id, filename, size, mime, uploaded_at, uploaded_by}
        "status": status,
        "created_by_id": current.get("id"),
        "created_by_name": current.get("name") or current.get("username"),
        "created_at": now,
        "updated_at": now,
        "submitted_at": now if status == "submitted" else None,
        # Engineering assignment (eng_head assigns to engineer)
        "assigned_to_id": "",
        "assigned_to_name": "",
        "assigned_by_id": "",
        "assigned_by_name": "",
        "assigned_at": None,
        # Engineering side
        "pic_engineer_name": "",       # required at accept time (multi-collab record)
        "accepted_by_id": "",
        "accepted_by_name": "",
        "accepted_at": None,
        "progress_notes": [],           # list of {at, by, note, status}
        "engineer_response_files": [],  # list of attachment ids
        "engineer_response_note": "",
        "completed_at": None,
        # Sales review
        "sales_reviews": [],            # list of {at, by, approve, note}
        "final_status": "",             # accepted / revision_requested (last review)
        "history": [
            {"at": now, "by": current.get("name") or current.get("username"), "action": f"created ({status})"},
        ],
    }
    await db.inquiries.insert_one(doc)
    await log_action(current, "create_inquiry", "inquiry", doc["id"], {"status": status})
    return _clean(doc)


@router.get("/inquiries")
async def list_inquiries(
    status: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[str] = None,  # YYYY-MM-DD (filter by created_at)
    date_to: Optional[str] = None,
    month: Optional[str] = None,  # YYYY-MM
    limit: int = 200,
    current: dict = Depends(get_current_user),
):
    filt: dict = {}
    role = current.get("role")
    # Sales sees ALL inquiries EXCEPT other sales' drafts (own drafts still visible)
    if role == "sales":
        filt["$or"] = [
            {"status": {"$ne": "draft"}},
            {"status": "draft", "created_by_id": current.get("id")},
        ]
    if is_engineering(current):
        # Engineers only see submitted onwards (skip drafts)
        filt["status"] = {"$nin": ["draft"]}
        # eng_staff sees only what's been assigned to them OR still unassigned (so they know pending)
        if role == "eng_staff":
            filt["$or"] = [
                {"assigned_to_id": current.get("id")},
                {"assigned_to_id": ""},
                {"assigned_to_id": {"$exists": False}},
            ]
    if status and status != "all":
        filt["status"] = status
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        search_or = [
            {"inquiry_no": rx},
            {"title": rx},
            {"customer_name": rx},
            {"project_name": rx},
            {"description": rx},
        ]
        if "$or" in filt:
            existing_or = filt.pop("$or")
            filt["$and"] = [{"$or": existing_or}, {"$or": search_or}]
        else:
            filt["$or"] = search_or
    # Month shorthand → derived from created_at
    if month:
        try:
            y, m = month.split("-")
            date_from = f"{int(y):04d}-{int(m):02d}-01"
            nm = int(m) + 1; ny = int(y) + (1 if nm > 12 else 0)
            if nm > 12: nm = 1
            date_to = f"{ny:04d}-{nm:02d}-01"
        except Exception:
            pass
    date_filter = {}
    if date_from: date_filter["$gte"] = date_from
    if date_to: date_filter["$lt" if month else "$lte"] = date_to
    if date_filter: filt["created_at"] = date_filter

    docs = await db.inquiries.find(merged(filt, NOT_DELETED_FILTER)).sort("created_at", -1).limit(limit).to_list(length=limit)
    for d in docs:
        _clean(d)
    return {"items": docs, "total": len(docs)}


@router.get("/inquiries/pending-count")
async def inquiries_pending_count(current: dict = Depends(get_current_user)):
    """Badge count per role:
    - eng_head/legacy engineering: all submitted (unassigned) inquiries awaiting assignment
    - eng_staff: inquiries assigned to them but not yet accepted (still 'submitted')
    - sales: awaiting_review from own
    - admin-like: all active
    """
    role = current.get("role")
    if role == "eng_staff":
        n = await db.inquiries.count_documents({
            "status": "submitted",
            "assigned_to_id": current.get("id"),
        })
        return {"role": role, "count": n, "kind": "assigned_to_me"}
    if is_eng_head(current):
        n = await db.inquiries.count_documents({
            "status": "submitted",
            "$or": [
                {"assigned_to_id": ""},
                {"assigned_to_id": {"$exists": False}},
            ],
        })
        return {"role": role, "count": n, "kind": "pending_assignment"}
    if role == "sales":
        n = await db.inquiries.count_documents({
            "status": "awaiting_review",
            "created_by_id": current.get("id"),
        })
        return {"role": role, "count": n, "kind": "awaiting_review"}
    if is_admin_like(current):
        n = await db.inquiries.count_documents({"status": {"$in": ["submitted", "awaiting_review"]}})
        return {"role": role, "count": n, "kind": "all_active"}
    return {"role": role, "count": 0}


@router.get("/inquiries/engineers")
async def list_engineers(current: dict = Depends(get_current_user)):
    """List of engineering users (for the assign dropdown). Accessible by eng_head, admin-like."""
    if not (is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Tidak berwenang")
    engineers = await db.users.find(
        {"role": {"$in": ["engineering", "eng_head", "eng_staff"]},
         "active": {"$ne": False},
         "deleted_at": {"$exists": False}},
        {"id": 1, "username": 1, "name": 1, "role": 1, "_id": 0},
    ).sort("name", 1).to_list(length=None)
    return {"items": engineers}


@router.get("/inquiries/{inq_id}")
async def get_inquiry(inq_id: str, current: dict = Depends(get_current_user)):
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    # Access: sales (all, but drafts only own), engineering (any), admin-like, finance (read-only)
    role = current.get("role")
    if role == "sales" and d.get("status") == "draft" and d.get("created_by_id") != current.get("id"):
        raise HTTPException(status_code=403, detail="Draft Sales lain tidak bisa dibuka")
    return _clean(d)


@router.put("/inquiries/{inq_id}")
async def update_inquiry_draft(inq_id: str, payload: InquiryUpdate, current: dict = Depends(get_current_user)):
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    if d.get("status") != "draft":
        raise HTTPException(status_code=400, detail="Hanya draft yang bisa diedit")
    if current.get("id") != d.get("created_by_id") and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Bukan Inquiry Anda")

    up = {"updated_at": datetime.utcnow().isoformat()}
    if payload.title is not None: up["title"] = payload.title.strip()
    if payload.customer_name is not None: up["customer_name"] = payload.customer_name.strip()
    if payload.project_name is not None: up["project_name"] = payload.project_name.strip()
    if payload.customer_deadline is not None: up["customer_deadline"] = payload.customer_deadline
    if payload.description is not None: up["description"] = payload.description.strip()
    if payload.items is not None: up["items"] = [i.model_dump() for i in payload.items]
    await db.inquiries.update_one({"id": inq_id}, {"$set": up})
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


@router.post("/inquiries/{inq_id}/submit")
async def submit_inquiry(inq_id: str, current: dict = Depends(get_current_user)):
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    if d.get("status") != "draft":
        raise HTTPException(status_code=400, detail="Hanya draft yang bisa disubmit")
    if current.get("id") != d.get("created_by_id") and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Bukan Inquiry Anda")

    now = datetime.utcnow().isoformat()
    entry = {"at": now, "by": current.get("name") or current.get("username"), "action": "submitted to engineering"}
    await db.inquiries.update_one(
        {"id": inq_id},
        {"$set": {"status": "submitted", "submitted_at": now, "updated_at": now},
         "$push": {"history": entry}},
    )
    await log_action(current, "submit_inquiry", "inquiry", inq_id, {})
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


# ---------------------------- Assignment (Eng Head only) ----------------------------
@router.post("/inquiries/{inq_id}/assign")
async def assign_inquiry(inq_id: str, payload: InquiryAssign, current: dict = Depends(get_current_user)):
    """Engineering Head assigns inquiry to a specific engineer. eng_staff cannot assign.
    Can re-assign to a different engineer while still in submitted/in_progress."""
    if not (is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering Head yang bisa assign inquiry")
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    if d.get("status") not in ("submitted", "in_progress", "revision_requested"):
        raise HTTPException(status_code=400, detail=f"Status '{d.get('status')}' tidak bisa di-assign")

    # Validate the target user is an engineer
    target = await db.users.find_one({"id": payload.engineer_id, "deleted_at": {"$exists": False}})
    if not target:
        raise HTTPException(status_code=404, detail="User target tidak ditemukan")
    if target.get("role") not in ("engineering", "eng_head", "eng_staff"):
        raise HTTPException(status_code=400, detail="User target bukan Engineering")

    engineer_name = payload.engineer_name.strip() or target.get("name") or target.get("username")
    now = datetime.utcnow().isoformat()
    who = current.get("name") or current.get("username")
    hist = {"at": now, "by": who, "action": f"assigned to {engineer_name}"}
    # Auto-transition status to in_progress on first assign (Accept step is now merged into Assign).
    new_status = d.get("status")
    if new_status == "submitted":
        new_status = "in_progress"
        hist_start = {"at": now, "by": who, "action": "engineering started (auto via assign)"}
    else:
        hist_start = None
    set_updates = {
        "assigned_to_id": target.get("id"),
        "assigned_to_name": engineer_name,
        "assigned_by_id": current.get("id"),
        "assigned_by_name": who,
        "assigned_at": now,
        "updated_at": now,
        "status": new_status,
        # Set PIC engineer to the assignee name for backward compatibility
        "pic_engineer_name": engineer_name if not d.get("pic_engineer_name") else d.get("pic_engineer_name"),
    }
    history_push = [hist] + ([hist_start] if hist_start else [])
    await db.inquiries.update_one(
        {"id": inq_id},
        {"$set": set_updates,
         "$push": {"history": {"$each": history_push}}},
    )
    await log_action(current, "assign_inquiry", "inquiry", inq_id,
                     {"engineer_id": target.get("id"), "engineer_name": engineer_name, "status": new_status})
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


@router.post("/inquiries/{inq_id}/accept")
async def accept_inquiry(inq_id: str, payload: InquiryAccept, current: dict = Depends(get_current_user)):
    """Engineering accepts and specifies PIC engineer name.
    - eng_head / admin: can accept any inquiry (self-picks or acts on behalf).
    - eng_staff: can only accept if assigned_to_id == current.id."""
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering yang bisa accept")
    if not payload.pic_engineer_name.strip():
        raise HTTPException(status_code=400, detail="Nama PIC Engineer wajib diisi")
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    if d.get("status") not in ("submitted", "revision_requested"):
        raise HTTPException(status_code=400, detail=f"Status saat ini '{d.get('status')}' tidak bisa di-accept")

    # eng_staff enforcement: must be assigned to them
    if current.get("role") == "eng_staff":
        if d.get("assigned_to_id") != current.get("id"):
            raise HTTPException(
                status_code=403,
                detail="Inquiry ini belum ditugaskan ke Anda. Hubungi Engineering Head untuk assignment.",
            )

    now = datetime.utcnow().isoformat()
    who = current.get("name") or current.get("username")
    entry = {"at": now, "by": who, "action": f"accepted (PIC: {payload.pic_engineer_name.strip()})"}
    upd_set = {
        "status": "in_progress",
        "pic_engineer_name": payload.pic_engineer_name.strip(),
        "accepted_by_id": current.get("id"),
        "accepted_by_name": who,
        "accepted_at": now,
        "updated_at": now,
    }
    # If not previously assigned, auto-assign to accepter (self-pick by eng_head)
    if not d.get("assigned_to_id"):
        upd_set["assigned_to_id"] = current.get("id")
        upd_set["assigned_to_name"] = who
        upd_set["assigned_by_id"] = current.get("id")
        upd_set["assigned_by_name"] = who
        upd_set["assigned_at"] = now
    await db.inquiries.update_one(
        {"id": inq_id},
        {"$set": upd_set, "$push": {"history": entry}},
    )
    await log_action(current, "accept_inquiry", "inquiry", inq_id, {"pic": payload.pic_engineer_name})
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


@router.post("/inquiries/{inq_id}/progress")
async def add_progress(inq_id: str, payload: InquiryProgress, current: dict = Depends(get_current_user)):
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering yang bisa update progress")
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    if d.get("status") not in ("in_progress", "awaiting_review"):
        raise HTTPException(status_code=400, detail=f"Status saat ini '{d.get('status')}' tidak bisa update progress")
    # eng_staff: only assignee can update
    if current.get("role") == "eng_staff" and d.get("assigned_to_id") != current.get("id"):
        raise HTTPException(status_code=403, detail="Anda bukan engineer yang di-assign")

    now = datetime.utcnow().isoformat()
    entry = {"at": now, "by": current.get("name") or current.get("username"), "note": payload.note.strip(), "status": payload.status}
    upd = {"$push": {"progress_notes": entry, "history": {"at": now, "by": entry["by"], "action": f"progress: {payload.note[:60]}"}},
           "$set": {"updated_at": now}}
    if payload.status == "awaiting_review":
        upd["$set"]["status"] = "awaiting_review"
        upd["$set"]["completed_at"] = now
    await db.inquiries.update_one({"id": inq_id}, upd)
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


@router.post("/inquiries/{inq_id}/submit-to-head")
async def submit_to_head(inq_id: str, note: str = Form(""), current: dict = Depends(get_current_user)):
    """Eng Staff (or admin) marks costing done → sends to Eng Head for internal review.
    Status flow: in_progress OR head_revision → pending_head_review.
    Shortcut: if the current user IS an eng_head/admin AND is the assignee → skip head review
    and go directly to awaiting_review (Sales)."""
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering yang bisa submit ke Head")
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    if d.get("status") not in ("in_progress", "head_revision"):
        raise HTTPException(status_code=400, detail="Hanya inquiry in_progress atau head_revision yang bisa dikirim ke Head")
    if current.get("role") == "eng_staff" and d.get("assigned_to_id") != current.get("id"):
        raise HTTPException(status_code=403, detail="Anda bukan engineer yang di-assign")

    now = datetime.utcnow().isoformat()
    who = current.get("name") or current.get("username")

    # If current user is eng_head/admin AND is the assignee → skip head review, go directly to Sales
    is_self_head = (is_eng_head(current) or is_admin_like(current)) and d.get("assigned_to_id") == current.get("id")
    if is_self_head:
        entry = {"at": now, "by": who, "action": "eng_head self-completed, sent directly to sales"}
        await db.inquiries.update_one(
            {"id": inq_id},
            {"$set": {
                "status": "awaiting_review",
                "engineer_response_note": note.strip(),
                "head_review_note": "(auto: engineer adalah Eng Head sendiri)",
                "head_reviewed_at": now,
                "head_reviewed_by_id": current.get("id"),
                "head_reviewed_by_name": who,
                "completed_at": now,
                "submitted_to_head_at": now,
                "updated_at": now,
            }, "$push": {"history": entry}},
        )
        await log_action(current, "submit_to_head_self", "inquiry", inq_id, {"shortcut": True})
    else:
        entry = {"at": now, "by": who, "action": "engineer submitted to head for review"}
        await db.inquiries.update_one(
            {"id": inq_id},
            {"$set": {
                "status": "pending_head_review",
                "engineer_response_note": note.strip(),
                "submitted_to_head_at": now,
                "updated_at": now,
            }, "$push": {"history": entry}},
        )
        await log_action(current, "submit_to_head", "inquiry", inq_id, {})
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


@router.post("/inquiries/{inq_id}/head-review")
async def head_review(inq_id: str, payload: dict, current: dict = Depends(get_current_user)):
    """Eng Head reviews the engineer's costing.
    payload: {approve: bool, note: str}
      approve=True  → status → awaiting_review (Sales)
      approve=False → status → head_revision (back to engineer for fixing)"""
    if not (is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering Head / Admin yang bisa review")
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    if d.get("status") != "pending_head_review":
        raise HTTPException(status_code=400, detail="Hanya inquiry pending_head_review yang bisa direview Head")

    approve = bool(payload.get("approve", False))
    note = (payload.get("note") or "").strip()
    now = datetime.utcnow().isoformat()
    who = current.get("name") or current.get("username")

    if approve:
        new_status = "awaiting_review"  # goes to Sales
        action_label = "head approved & sent to sales"
        set_upd = {
            "status": new_status,
            "head_review_note": note,
            "head_reviewed_at": now,
            "head_reviewed_by_id": current.get("id"),
            "head_reviewed_by_name": who,
            "completed_at": now,  # keep backward compat with old "completed_at" field
            "updated_at": now,
        }
    else:
        new_status = "head_revision"  # back to engineer
        action_label = "head requested revision"
        if not note:
            raise HTTPException(status_code=400, detail="Catatan revisi wajib diisi")
        set_upd = {
            "status": new_status,
            "head_revision_note": note,
            "head_reviewed_at": now,
            "head_reviewed_by_id": current.get("id"),
            "head_reviewed_by_name": who,
            "updated_at": now,
        }
    entry = {"at": now, "by": who, "action": action_label, "note": note}
    await db.inquiries.update_one(
        {"id": inq_id},
        {"$set": set_upd, "$push": {"history": entry}},
    )
    await log_action(current, "head_review_inquiry", "inquiry", inq_id, {"approve": approve})
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


@router.post("/inquiries/{inq_id}/complete")
async def complete_inquiry(inq_id: str, note: str = Form(""), current: dict = Depends(get_current_user)):
    """[LEGACY] Engineering marks work done → awaiting_review. Kept for backward-compat.
    New flow uses submit-to-head → head-review instead."""
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering yang bisa complete")
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    if d.get("status") != "in_progress":
        raise HTTPException(status_code=400, detail="Hanya inquiry yang in_progress bisa di-complete")
    if current.get("role") == "eng_staff" and d.get("assigned_to_id") != current.get("id"):
        raise HTTPException(status_code=403, detail="Anda bukan engineer yang di-assign")

    now = datetime.utcnow().isoformat()
    who = current.get("name") or current.get("username")
    entry = {"at": now, "by": who, "action": "engineering completed"}
    await db.inquiries.update_one(
        {"id": inq_id},
        {"$set": {"status": "awaiting_review", "engineer_response_note": note.strip(),
                  "completed_at": now, "updated_at": now},
         "$push": {"history": entry}},
    )
    await log_action(current, "complete_inquiry", "inquiry", inq_id, {})
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


@router.post("/inquiries/{inq_id}/review")
async def review_inquiry(inq_id: str, payload: InquiryReview, current: dict = Depends(get_current_user)):
    if current.get("role") not in ("sales", "admin", "super_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Hanya Sales/Admin yang bisa review")
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    # Sales team can review any inquiry (open access per policy).
    if d.get("status") != "awaiting_review":
        raise HTTPException(status_code=400, detail="Belum bisa direview (status bukan awaiting_review)")

    now = datetime.utcnow().isoformat()
    who = current.get("name") or current.get("username")
    new_status = "accepted" if payload.approve else "revision_requested"
    entry = {"at": now, "by": who, "approve": payload.approve, "note": payload.review_note.strip()}
    hist = {"at": now, "by": who, "action": f"reviewed → {new_status}: {payload.review_note[:60]}"}
    await db.inquiries.update_one(
        {"id": inq_id},
        {"$set": {"status": new_status, "final_status": new_status, "updated_at": now},
         "$push": {"sales_reviews": entry, "history": hist}},
    )
    await log_action(current, "review_inquiry", "inquiry", inq_id, {"approve": payload.approve})
    updated = await db.inquiries.find_one({"id": inq_id})
    return _clean(updated)


# ---------------------------- Attachments (GridFS) ----------------------------
@router.post("/inquiries/{inq_id}/attachments")
async def upload_attachment(
    inq_id: str,
    file: UploadFile = File(...),
    slot: str = Form("sales"),  # 'sales' (attach ke inquiry) atau 'engineer' (hasil kerja engineer)
    current: dict = Depends(get_current_user),
):
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")

    role = current.get("role")
    if slot == "sales":
        if not ((role == "sales" and d.get("created_by_id") == current.get("id")) or is_admin_like(current)):
            raise HTTPException(status_code=403, detail="Tidak berwenang upload ke sales attachments")
    elif slot == "engineer":
        if not (is_engineering(current) or is_admin_like(current)):
            raise HTTPException(status_code=403, detail="Hanya Engineering yang bisa upload response files")
        if current.get("role") == "eng_staff" and d.get("assigned_to_id") != current.get("id"):
            raise HTTPException(status_code=403, detail="Anda bukan engineer yang di-assign")
    else:
        raise HTTPException(status_code=400, detail="slot harus 'sales' atau 'engineer'")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File terlalu besar (max 25 MB)")

    fs = gridfs()
    file_id = await fs.upload_from_stream(
        file.filename or "attachment",
        io.BytesIO(content),
        metadata={"inquiry_id": inq_id, "slot": slot, "uploaded_by": current.get("name") or current.get("username")},
    )
    meta = {
        "id": str(file_id),
        "filename": file.filename or "attachment",
        "size": len(content),
        "mime": file.content_type or "application/octet-stream",
        "uploaded_at": datetime.utcnow().isoformat(),
        "uploaded_by": current.get("name") or current.get("username"),
        "slot": slot,
    }
    field = "attachments" if slot == "sales" else "engineer_response_files"
    await db.inquiries.update_one({"id": inq_id}, {"$push": {field: meta}})
    return meta


@router.get("/inquiries/{inq_id}/attachments/{file_id}/download")
async def download_attachment(inq_id: str, file_id: str, inline: bool = False, current: dict = Depends(get_current_user)):
    """Download an attachment. Pass ?inline=1 to preview in-browser (PDF/image)."""
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")

    # Locate metadata for filename
    meta = None
    for att in (d.get("attachments") or []) + (d.get("engineer_response_files") or []):
        if att.get("id") == file_id:
            meta = att
            break
    if not meta:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")

    fs = gridfs()
    try:
        stream = await fs.open_download_stream(ObjectId(file_id))
    except Exception:
        raise HTTPException(status_code=404, detail="File tidak ditemukan di storage")

    buf = io.BytesIO()
    async for chunk in stream:
        buf.write(chunk)
    buf.seek(0)
    filename = meta.get("filename") or "download"
    disposition = "inline" if inline else "attachment"
    return StreamingResponse(
        buf,
        media_type=meta.get("mime") or "application/octet-stream",
        headers={"Content-Disposition": f'{disposition}; filename="{filename}"'},
    )


async def _inq_att_bytes(inq_id: str, file_id: str):
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")
    meta = None
    for att in (d.get("attachments") or []) + (d.get("engineer_response_files") or []):
        if att.get("id") == file_id:
            meta = att
            break
    if not meta:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    try:
        stream = await gridfs().open_download_stream(ObjectId(file_id))
        raw = await stream.read()
    except Exception:
        raise HTTPException(status_code=404, detail="File tidak ada di storage")
    return raw, meta


@router.get("/inquiries/{inq_id}/attachments/{file_id}/page-meta")
async def inquiry_attachment_page_meta(inq_id: str, file_id: str, current: dict = Depends(get_current_user)):
    """Metadata halaman untuk viewer image-based (PDF & Excel)."""
    from utils.pdf_render import pdf_page_meta
    from utils.office_render import is_office_ext, office_to_pdf
    raw, meta = await _inq_att_bytes(inq_id, file_id)
    fn = (meta.get("filename") or "").lower()
    ext = fn.rsplit(".", 1)[-1] if "." in fn else ""
    if ext == "pdf":
        return pdf_page_meta(raw)
    if is_office_ext(ext):
        return pdf_page_meta(office_to_pdf(raw, ext))
    raise HTTPException(status_code=400, detail="Preview gambar hanya untuk PDF/Excel")


@router.get("/inquiries/{inq_id}/attachments/{file_id}/page-image")
async def inquiry_attachment_page_image(inq_id: str, file_id: str, page: int = 0, scale: float = 2.0,
                                        current: dict = Depends(get_current_user)):
    """Render satu halaman lampiran inquiry (PDF/Excel) menjadi PNG."""
    from utils.pdf_render import pdf_page_png
    from utils.office_render import is_office_ext, office_to_pdf
    raw, meta = await _inq_att_bytes(inq_id, file_id)
    fn = (meta.get("filename") or "").lower()
    ext = fn.rsplit(".", 1)[-1] if "." in fn else ""
    if ext == "pdf":
        pdf = raw
    elif is_office_ext(ext):
        pdf = office_to_pdf(raw, ext)
    else:
        raise HTTPException(status_code=400, detail="Preview gambar hanya untuk PDF/Excel")
    try:
        png = pdf_page_png(pdf, page, scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ditemukan")
    return StreamingResponse(io.BytesIO(png), media_type="image/png",
                             headers={"Cache-Control": "private, max-age=120"})



@router.delete("/inquiries/{inq_id}/attachments/{file_id}")
async def delete_attachment(inq_id: str, file_id: str, current: dict = Depends(get_current_user)):
    d = await db.inquiries.find_one({"id": inq_id})
    if not d:
        raise HTTPException(status_code=404, detail="Inquiry tidak ditemukan")

    # Determine which slot & permission
    for field in ("attachments", "engineer_response_files"):
        for att in d.get(field, []):
            if att.get("id") == file_id:
                role = current.get("role")
                slot = att.get("slot")
                if slot == "sales" and not (role == "sales" or is_admin_like(current)):
                    raise HTTPException(status_code=403, detail="Tidak berwenang")
                if slot == "engineer" and not (is_engineering(current) or is_admin_like(current)):
                    raise HTTPException(status_code=403, detail="Tidak berwenang")
                await db.inquiries.update_one({"id": inq_id}, {"$pull": {field: {"id": file_id}}})
                try:
                    await gridfs().delete(ObjectId(file_id))
                except Exception:
                    pass
                return {"success": True}
    raise HTTPException(status_code=404, detail="File tidak ditemukan")


# =============================================================================
# QUOTATIONS  (data-only, PDF generation next iteration)
# =============================================================================
class QuotationCreate(BaseModel):
    inquiry_id: Optional[str] = None  # optional link to inquiry
    quotation_no_override: Optional[str] = None  # if set, uses this instead of auto-generated
    customer_name: str
    customer_address: str = ""
    attention: str = ""
    cc: str = ""
    items: List[dict] = []  # [{no, description, qty, unit}]
    notes_lines: List[str] = []  # bullet notes (manual, di bawah items)
    include_standard_note: bool = False  # toggle "full quantity order" clause
    in_words: str = ""
    total_amount: float = 0.0
    currency: str = "IDR"
    payment_term: str = ""
    delivery_time: str = ""
    validity: str = ""
    signature_name: str = ""  # auto-filled from current user if empty
    signature_position: str = "Sales Dept"
    approver_name: str = "Mr. Asiong Lu"
    approver_position: str = "Business Dev. Manager"


class QuotationUpdate(BaseModel):
    """Full-form update — increments revision_no and records revision history."""
    quotation_no_override: Optional[str] = None
    customer_name: Optional[str] = None
    customer_address: Optional[str] = None
    attention: Optional[str] = None
    cc: Optional[str] = None
    items: Optional[List[dict]] = None
    notes_lines: Optional[List[str]] = None
    include_standard_note: Optional[bool] = None
    in_words: Optional[str] = None
    total_amount: Optional[float] = None
    currency: Optional[str] = None
    payment_term: Optional[str] = None
    delivery_time: Optional[str] = None
    validity: Optional[str] = None
    signature_name: Optional[str] = None
    signature_position: Optional[str] = None
    approver_name: Optional[str] = None
    approver_position: Optional[str] = None
    revision_reason: str = ""  # required-ish for audit


def _block_engineering_from_quotation(current: dict):
    """Engineering must not see quotation content (rahasia). Raise 403."""
    if is_engineering(current) and not is_admin_like(current):
        raise HTTPException(
            status_code=403,
            detail="Quotation adalah dokumen konfidensial Sales — Engineering tidak berwenang melihat isi. "
                   "Anda hanya bisa melihat notifikasi bahwa Inquiry telah dibuatkan Quotation.",
        )


@router.post("/quotations")
async def create_quotation(payload: QuotationCreate, current: dict = Depends(get_current_user)):
    if not (current.get("role") == "sales" or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Sales yang bisa buat Quotation")
    if not payload.customer_name.strip():
        raise HTTPException(status_code=400, detail="Customer wajib diisi")

    # Determine quotation number — override or auto
    override = (payload.quotation_no_override or "").strip()
    if override:
        # Check duplicate
        existing = await db.quotations.find_one({"quotation_no": override, "deleted_at": {"$exists": False}})
        if existing:
            raise HTTPException(
                status_code=409,
                detail=f"Nomor Quotation '{override}' sudah dipakai. Silakan pilih nomor lain atau kosongkan untuk auto-generate.",
            )
        q_no = override
    else:
        q_no = await _new_quotation_no()

    # Enrich inquiry reference (denormalize inquiry_no + project + customer for display)
    inquiry_ref = {"inquiry_id": "", "inquiry_no": "", "project_name": ""}
    if payload.inquiry_id:
        inq = await db.inquiries.find_one({"id": payload.inquiry_id, "deleted_at": {"$exists": False}})
        if not inq:
            raise HTTPException(status_code=400, detail="Inquiry referensi tidak ditemukan")
        inquiry_ref = {
            "inquiry_id": inq["id"],
            "inquiry_no": inq.get("inquiry_no", ""),
            "project_name": inq.get("project_name", ""),
        }

    now = datetime.utcnow().isoformat()
    payload_dict = payload.model_dump()
    # Auto-fill signature_name from current user if empty (nama sales = nama user login)
    if not (payload_dict.get("signature_name") or "").strip():
        payload_dict["signature_name"] = current.get("name") or current.get("username") or ""
    doc = {
        "id": str(uuid.uuid4()),
        "quotation_no": q_no,
        **payload_dict,
        **inquiry_ref,
        "so_no": "",  # filled only when status → confirm_order
        "so_confirmed_at": None,
        "so_confirmed_by": "",
        "revision_no": 1,
        "revision_history": [],  # list of snapshots
        "created_by_id": current.get("id"),
        "created_by_name": current.get("name") or current.get("username"),
        "created_at": now,
        "updated_at": now,
        "status": "on_bidding",  # on_bidding | confirm_order | cancel
    }
    doc.pop("quotation_no_override", None)  # not a real field
    await db.quotations.insert_one(doc)

    # If linked to inquiry, add back-reference on inquiry (append to linked_quotations)
    if inquiry_ref["inquiry_id"]:
        await db.inquiries.update_one(
            {"id": inquiry_ref["inquiry_id"]},
            {"$push": {
                "linked_quotations": {
                    "quotation_id": doc["id"],
                    "quotation_no": doc["quotation_no"],
                    "created_at": now,
                    "created_by_name": doc["created_by_name"],
                    "status": "on_bidding",
                },
                "history": {
                    "at": now,
                    "by": doc["created_by_name"],
                    "action": f"quotation {doc['quotation_no']} dibuat",
                },
            }},
        )

    await log_action(current, "create_quotation", "quotation", doc["id"],
                     {"quotation_no": doc["quotation_no"], "inquiry_id": inquiry_ref["inquiry_id"]})
    return _clean(doc)


@router.get("/quotations/next-no")
async def preview_next_quotation_no(current: dict = Depends(get_current_user)):
    """Preview the next quotation number for the current month without incrementing the counter."""
    _block_engineering_from_quotation(current)
    now = datetime.utcnow()
    key = f"quotation:{now.year}-{now.month:02d}"
    doc = await db.counters.find_one({"_id": key})
    seq = int(doc.get("seq", 0)) + 1 if doc else 1
    return {"quotation_no": f"{seq:03d}/MKS/Q/{ROMAN[now.month]}/{now.year}"}


@router.get("/quotations")
async def list_quotations(
    q: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,  # YYYY-MM-DD
    date_to: Optional[str] = None,
    month: Optional[str] = None,  # YYYY-MM shorthand
    limit: int = 200,
    current: dict = Depends(get_current_user),
):
    _block_engineering_from_quotation(current)
    filt: dict = {}
    # Sales sees ALL quotations from all users (open access per policy)
    if status and status != "all":
        filt["status"] = status
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [
            {"quotation_no": rx},
            {"customer_name": rx},
            {"attention": rx},
            {"items.description": rx},
            {"inquiry_no": rx},
            {"project_name": rx},
            {"so_no": rx},
        ]
    # Month shorthand
    if month:
        try:
            y, m = month.split("-")
            date_from = f"{int(y):04d}-{int(m):02d}-01"
            # Last day of month is tricky — just use next month start as upper
            nm = int(m) + 1
            ny = int(y) + (1 if nm > 12 else 0)
            if nm > 12: nm = 1
            date_to = f"{ny:04d}-{nm:02d}-01"
        except Exception:
            pass
    date_filter = {}
    if date_from: date_filter["$gte"] = date_from
    if date_to: date_filter["$lt" if month else "$lte"] = date_to
    if date_filter: filt["created_at"] = date_filter

    docs = await db.quotations.find(merged(filt, NOT_DELETED_FILTER)).sort("created_at", -1).limit(limit).to_list(length=limit)
    for d in docs:
        _clean(d)
    return {"items": docs, "total": len(docs)}


@router.get("/quotations/{qid}")
async def get_quotation(qid: str, current: dict = Depends(get_current_user)):
    _block_engineering_from_quotation(current)
    d = await db.quotations.find_one({"id": qid})
    if not d:
        raise HTTPException(status_code=404, detail="Quotation tidak ditemukan")
    return _clean(d)


@router.patch("/quotations/{qid}")
async def update_quotation(qid: str, payload: QuotationUpdate, current: dict = Depends(get_current_user)):
    """Full-form Quotation edit → creates a revision snapshot + increments revision_no."""
    if not (current.get("role") == "sales" or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Sales yang bisa edit Quotation")
    d = await db.quotations.find_one({"id": qid})
    if not d:
        raise HTTPException(status_code=404, detail="Quotation tidak ditemukan")
    if d.get("status") == "cancel":
        raise HTTPException(status_code=400, detail="Quotation yang dibatalkan tidak bisa di-edit")

    # Handle quotation_no override
    updates = {}
    if payload.quotation_no_override is not None:
        new_no = payload.quotation_no_override.strip()
        if new_no and new_no != d.get("quotation_no"):
            dup = await db.quotations.find_one({"quotation_no": new_no, "id": {"$ne": qid}, "deleted_at": {"$exists": False}})
            if dup:
                raise HTTPException(status_code=409, detail=f"Nomor Quotation '{new_no}' sudah dipakai")
            updates["quotation_no"] = new_no

    # Snapshot BEFORE change → revision history
    snapshot = {
        "revision_no": d.get("revision_no", 1),
        "snapshot_at": datetime.utcnow().isoformat(),
        "snapshot_by": current.get("name") or current.get("username"),
        "reason": (payload.revision_reason or "").strip(),
        "quotation_no": d.get("quotation_no"),
        "customer_name": d.get("customer_name"),
        "items": d.get("items", []),
        "notes_lines": d.get("notes_lines", []),
        "total_amount": d.get("total_amount"),
        "currency": d.get("currency"),
        "payment_term": d.get("payment_term"),
        "delivery_time": d.get("delivery_time"),
        "validity": d.get("validity"),
    }

    # Apply the rest of the fields
    payload_dict = payload.model_dump(exclude_unset=True)
    payload_dict.pop("quotation_no_override", None)
    payload_dict.pop("revision_reason", None)
    updates.update({k: v for k, v in payload_dict.items() if v is not None})

    updates["revision_no"] = d.get("revision_no", 1) + 1
    updates["updated_at"] = datetime.utcnow().isoformat()
    updates["last_revised_by"] = current.get("name") or current.get("username")
    updates["last_revision_reason"] = snapshot["reason"]

    await db.quotations.update_one(
        {"id": qid},
        {"$set": updates, "$push": {"revision_history": snapshot}},
    )
    await log_action(current, "edit_quotation", "quotation", qid,
                     {"revision_no": updates["revision_no"], "reason": snapshot["reason"]})
    updated = await db.quotations.find_one({"id": qid})
    return _clean(updated)


@router.delete("/quotations/{qid}")
async def delete_quotation(qid: str, current: dict = Depends(get_current_user)):
    """Sales & admin can delete quotation directly (soft delete). No approval workflow."""
    if not (current.get("role") == "sales" or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Sales & Admin yang bisa hapus Quotation")
    d = await db.quotations.find_one({"id": qid})
    if not d:
        raise HTTPException(status_code=404, detail="Quotation tidak ditemukan")
    # If linked to inquiry, remove from linked_quotations
    if d.get("inquiry_id"):
        await db.inquiries.update_one(
            {"id": d["inquiry_id"]},
            {"$pull": {"linked_quotations": {"quotation_id": qid}},
             "$push": {"history": {
                 "at": datetime.utcnow().isoformat(),
                 "by": current.get("name") or current.get("username"),
                 "action": f"quotation {d.get('quotation_no', '')} dihapus",
             }}},
        )
    await soft_delete_one("quotations", {"id": qid}, current)
    await log_action(current, "delete_quotation", "quotation", qid, {"quotation_no": d.get("quotation_no")})
    return {"ok": True}


@router.get("/quotations/{qid}/pdf")
async def download_quotation_pdf(qid: str, current: dict = Depends(get_current_user)):
    """Generate & stream Quotation PDF with PT MKS letterhead. Blocked for Engineering."""
    _block_engineering_from_quotation(current)
    from services.quotation_pdf import build_quotation_pdf
    from bson import ObjectId
    from motor.motor_asyncio import AsyncIOMotorGridFSBucket
    d = await db.quotations.find_one({"id": qid})
    if not d:
        raise HTTPException(status_code=404, detail="Quotation tidak ditemukan")
    # Try to enrich address from master customer if not stored on quotation
    if not d.get("customer_address") and d.get("customer_name"):
        cust = await db.customers.find_one({"name": d["customer_name"]})
        if cust and cust.get("address"):
            d["customer_address"] = cust["address"]

    # ---- Inject signature images (sales creator + approver) if available ----
    sig_fs = AsyncIOMotorGridFSBucket(db, bucket_name="signatures")
    async def _load_sig(user_id: str) -> Optional[bytes]:
        if not user_id: return None
        try:
            u = await db.users.find_one({"id": user_id}, {"signature_gridfs_id": 1})
        except Exception:
            return None
        if not u or not u.get("signature_gridfs_id"):
            return None
        try:
            st = await sig_fs.open_download_stream(ObjectId(u["signature_gridfs_id"]))
            buf = io.BytesIO()
            async for chunk in st:
                buf.write(chunk)
            return buf.getvalue()
        except Exception:
            return None

    # Sales signature = created_by user
    sales_bytes = await _load_sig(d.get("created_by_id"))
    if sales_bytes:
        d["_sales_signature_png"] = sales_bytes

    # Approver signature = lookup user by name (default: Asiong Lu)
    approver_name = (d.get("approver_name") or "Mr. Asiong Lu").replace("Mr.", "").strip()
    approver_user = await db.users.find_one(
        {"name": {"$regex": approver_name, "$options": "i"}, "deleted_at": {"$exists": False}},
        {"id": 1, "signature_gridfs_id": 1},
    )
    if approver_user:
        appr_bytes = await _load_sig(approver_user.get("id"))
        if appr_bytes:
            d["_approver_signature_png"] = appr_bytes

    pdf_bytes = build_quotation_pdf(_clean(d))
    await log_action(current, "download_quotation_pdf", "quotations", qid, {"quotation_no": d.get("quotation_no")})
    safe_no = re.sub(r"[^A-Za-z0-9._-]+", "_", d.get("quotation_no") or qid)
    fname = f"Quotation_{safe_no}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



class QuotationStatusUpdate(BaseModel):
    status: str  # on_bidding | confirm_order | cancel
    so_no: Optional[str] = ""  # required (numeric, dinormalisasi ke 6 digit) when status = confirm_order
    force_reuse_so: bool = False  # if so_no already exists, must pass force_reuse_so=True to bind existing SO


# =============================================================================
# CUSTOMERS MASTER
# =============================================================================
class CustomerCreate(BaseModel):
    name: str
    address: str = ""
    pic: str = ""  # Person In Charge / Attention person
    phone: str = ""
    email: str = ""
    notes: str = ""
    customer_code: str = ""  # Kode singkat untuk drawing register (diisi Engineering)


class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    pic: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None
    customer_code: Optional[str] = None


@router.get("/customers")
async def list_customers(q: Optional[str] = None, limit: int = 500, current: dict = Depends(get_current_user)):
    filt: dict = {}
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [{"name": rx}, {"pic": rx}]
    docs = await db.customers.find(merged(filt, NOT_DELETED_FILTER)).sort("name", 1).limit(limit).to_list(length=limit)
    for d in docs:
        _clean(d)
    return {"items": docs, "total": len(docs)}


@router.post("/customers")
async def create_customer(payload: CustomerCreate, current: dict = Depends(get_current_user)):
    if not (current.get("role") == "sales" or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Sales & Admin yang bisa kelola customer")
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Nama customer wajib diisi")
    existing = await db.customers.find_one({"name": {"$regex": f"^{re.escape(payload.name.strip())}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail=f"Customer '{payload.name}' sudah ada")
    doc = {
        "id": str(uuid.uuid4()),
        **{k: v.strip() if isinstance(v, str) else v for k, v in payload.model_dump().items()},
        "name": payload.name.strip(),
        "created_at": datetime.utcnow().isoformat(),
        "created_by_name": current.get("name") or current.get("username"),
    }
    await db.customers.insert_one(doc)
    await log_action(current, "create_customer", "customer", doc["id"], {"name": doc["name"]})
    return _clean(doc)


@router.put("/customers/{cid}")
async def update_customer(cid: str, payload: CustomerUpdate, current: dict = Depends(get_current_user)):
    if not (current.get("role") == "sales" or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Sales & Admin yang bisa edit customer")
    up = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if not up:
        raise HTTPException(status_code=400, detail="Tidak ada perubahan")
    up["updated_at"] = datetime.utcnow().isoformat()
    res = await db.customers.update_one({"id": cid}, {"$set": up})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer tidak ditemukan")
    updated = await db.customers.find_one({"id": cid})
    return _clean(updated)


@router.delete("/customers/{cid}")
async def delete_customer(cid: str, current: dict = Depends(get_current_user)):
    if not (current.get("role") == "sales" or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Tidak berwenang")
    ok = await soft_delete_one("customers", {"id": cid}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="Customer tidak ditemukan")
    await log_action(current, "delete_customer", "customer", cid, {})
    return {"success": True}


class CustomerCodeIn(BaseModel):
    customer_code: str = ""


@router.patch("/customers/{cid}/customer-code")
async def set_customer_code(cid: str, payload: CustomerCodeIn, current: dict = Depends(get_current_user)):
    """Engineering / Admin set customer_code (kode singkat untuk drawing numbering).
    Role sales boleh view saja; role eng_* / admin yang boleh set."""
    role = current.get("role") or ""
    if role not in ("eng_leader", "eng_head", "eng_staff", "engineering") and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Engineering/Admin yang boleh set customer code")
    code = (payload.customer_code or "").upper().strip()
    # basic validation — code hanya alfanumerik & dash, max 10 char
    if code and not re.match(r"^[A-Z0-9\-]{1,10}$", code):
        raise HTTPException(status_code=400, detail="Customer code harus 1-10 karakter alfanumerik/dash (contoh: MKS, YOK, SPM)")
    up = {"customer_code": code, "updated_at": datetime.utcnow().isoformat(),
          "customer_code_updated_by": current.get("name") or current.get("username")}
    res = await db.customers.update_one({"id": cid}, {"$set": up})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer tidak ditemukan")
    await log_action(current, "set_customer_code", "customer", cid, {"customer_code": code})
    updated = await db.customers.find_one({"id": cid})
    return _clean(updated)


@router.post("/customers/upsert-by-name")
async def upsert_customer_by_name(payload: dict, current: dict = Depends(get_current_user)):
    """Upsert customer by name (case-insensitive) — dipakai saat engineering ketik customer code
    manual untuk customer yang belum terdaftar. Auto-save code juga bila disertakan."""
    role = current.get("role") or ""
    if role not in ("eng_leader", "eng_head", "eng_staff", "engineering", "sales") and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Tidak berwenang")
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nama customer wajib diisi")
    code = (payload.get("customer_code") or "").upper().strip()
    if code and not re.match(r"^[A-Z0-9\-]{1,10}$", code):
        raise HTTPException(status_code=400, detail="Customer code harus 1-10 karakter alfanumerik/dash")

    existing = await db.customers.find_one(
        {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}, **NOT_DELETED_FILTER}
    )
    if existing:
        # Update code only if provided and different
        if code and existing.get("customer_code", "") != code:
            await db.customers.update_one(
                {"id": existing["id"]},
                {"$set": {"customer_code": code, "updated_at": datetime.utcnow().isoformat(),
                          "customer_code_updated_by": current.get("name") or current.get("username")}},
            )
            existing["customer_code"] = code
        await log_action(current, "upsert_customer_by_name", "customer", existing["id"], {"code": code, "existed": True})
        return _clean(existing)
    # Create new customer
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "customer_code": code,
        "address": "", "pic": "", "phone": "", "email": "", "notes": "",
        "created_at": datetime.utcnow().isoformat(),
        "created_by_name": current.get("name") or current.get("username"),
    }
    await db.customers.insert_one(doc)
    await log_action(current, "upsert_customer_by_name", "customer", doc["id"], {"code": code, "existed": False})
    return _clean(doc)


@router.patch("/quotations/{qid}/status")
async def update_quotation_status(qid: str, payload: QuotationStatusUpdate, current: dict = Depends(get_current_user)):
    if not (current.get("role") == "sales" or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Sales yang bisa update status")
    if payload.status not in ("on_bidding", "confirm_order", "cancel"):
        raise HTTPException(status_code=400, detail="Status tidak valid")

    quo = await db.quotations.find_one({"id": qid})
    if not quo:
        raise HTTPException(status_code=404, detail="Quotation tidak ditemukan")

    now = datetime.utcnow().isoformat()
    upd = {"status": payload.status, "status_updated_at": now, "updated_at": now}

    # === SO Integration on confirm_order ===
    if payload.status == "confirm_order":
        from routers.bom import normalize_so_no
        raw_so = (payload.so_no or "").strip()
        if not raw_so:
            raise HTTPException(status_code=400, detail="Nomor SO wajib diisi saat konfirmasi order (6 digit)")
        # Validasi: harus numerik & maksimal 6 digit → dinormalisasi ke 6 digit (zero-pad).
        if not (raw_so.isdigit() and len(raw_so) <= 6):
            raise HTTPException(status_code=400, detail="Nomor SO harus angka maksimal 6 digit (mis. 5251 → 005251)")
        so_no = normalize_so_no(raw_so)

        # Check master list SO
        existing_so = await db.sales_orders.find_one({"so_no": so_no, "deleted_at": {"$exists": False}})
        if existing_so and not payload.force_reuse_so:
            # Detect if this SO was already used by ANOTHER active quotation
            other = await db.quotations.find_one({
                "so_no": so_no,
                "id": {"$ne": qid},
                "status": "confirm_order",
                "deleted_at": {"$exists": False},
            })
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "SO_EXISTS",
                    "so_no": so_no,
                    "customer": existing_so.get("customer", ""),
                    "so_date": existing_so.get("so_date", ""),
                    "description": existing_so.get("description", ""),
                    "used_by_quotation": (other or {}).get("quotation_no") if other else None,
                    "message": f"Nomor SO {so_no} sudah tercatat di Master List SO"
                               + (f" — dipakai oleh Quotation {other.get('quotation_no')}" if other else "")
                               + ". Untuk tetap pakai nomor ini, konfirmasi ulang.",
                },
            )

        # Create sales_orders entry if it doesn't exist yet (auto-integrate)
        if not existing_so:
            so_doc = {
                "id": str(uuid.uuid4()),
                "so_no": so_no,
                "so_date": now[:10],
                "customer": quo.get("customer_name") or "",
                "description": f"Auto from Quotation {quo.get('quotation_no', '')}"
                               + (f" · {quo.get('project_name')}" if quo.get('project_name') else ""),
                "created_by": current["id"],
                "created_by_username": current.get("username", ""),
                "created_at": now,
                "source_quotation_id": qid,
                "source_quotation_no": quo.get("quotation_no", ""),
            }
            await db.sales_orders.insert_one(so_doc)
            await log_action(current, "auto_create_so", "sales_order", so_doc["id"],
                             {"so_no": so_no, "from_quotation": qid})

        upd["so_no"] = so_no
        upd["so_confirmed_at"] = now
        upd["so_confirmed_by"] = current.get("name") or current.get("username")

        # Update linked inquiry status → closed (job done)
        if quo.get("inquiry_id"):
            await db.inquiries.update_one(
                {"id": quo["inquiry_id"]},
                {"$set": {
                    "status": "closed",
                    "updated_at": now,
                },
                 "$push": {"history": {
                    "at": now, "by": current.get("name") or current.get("username"),
                    "action": f"quotation {quo.get('quotation_no', '')} confirm_order (SO {so_no}) — inquiry ditutup",
                }}}
            )

    await db.quotations.update_one({"id": qid}, {"$set": upd})

    # Update back-reference on inquiry (linked_quotations[].status)
    if quo.get("inquiry_id"):
        await db.inquiries.update_one(
            {"id": quo["inquiry_id"], "linked_quotations.quotation_id": qid},
            {"$set": {"linked_quotations.$.status": payload.status,
                      "linked_quotations.$.so_no": upd.get("so_no", quo.get("so_no", ""))}},
        )

    await log_action(current, "quotation_status", "quotation", qid,
                     {"status": payload.status, "so_no": upd.get("so_no", "")})
    updated = await db.quotations.find_one({"id": qid})
    return _clean(updated)


# ============ SO Autocomplete/Search — for form integration ============
@router.get("/sales-orders/autocomplete")
async def so_autocomplete(q: Optional[str] = None, limit: int = 20, current: dict = Depends(get_current_user)):
    """Type-ahead search on the Master SO list. Returns matched SOs with customer + date."""
    filt: dict = {}
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [{"so_no": rx}, {"customer": rx}, {"description": rx}]
    docs = await db.sales_orders.find(
        merged(filt, NOT_DELETED_FILTER),
        {"_id": 0, "so_no": 1, "customer": 1, "so_date": 1, "description": 1, "source_quotation_no": 1},
    ).sort("so_no", 1).limit(limit).to_list(length=limit)
    return {"items": docs}


@router.post("/sales-orders/import-list")
async def import_so_list(file: UploadFile = File(...), current: dict = Depends(get_current_user)):
    """Import daftar SO dari file Excel (kolom: SO number, Date, Customer, Description).

    - Nomor SO dinormalisasi ke 6 digit (mis. 4640 -> 004640).
    - Upsert by so_no: yang belum ada dibuat; yang sudah ada dilengkapi (customer/description/date bila kosong).
    """
    if not (is_admin_like(current) or (current.get("role") in ("sales", "supervisor"))):
        raise HTTPException(status_code=403, detail="Hanya Admin/Supervisor/Sales yang boleh import daftar SO")
    from routers.bom import _read_workbook, _clean_str, normalize_so_no, _excel_serial_to_iso

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    try:
        rows = _read_workbook(content, file.filename)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca file: {e}")
    if not rows:
        raise HTTPException(status_code=400, detail="File tidak berisi data")

    def norm(s):
        return _clean_str(s).lower()

    header_idx, col = 0, {"so": 0, "date": None, "customer": None, "desc": None}
    for i, r in enumerate(rows[:10]):
        cells = [norm(c) for c in r]
        joined = " ".join(cells)
        if "so" in joined and ("customer" in joined or "description" in joined or "date" in joined):
            for j, c in enumerate(cells):
                if ("so number" in c) or c in ("so", "so no", "so no.", "so.no", "no so"):
                    col["so"] = j
                elif "date" in c or "tanggal" in c:
                    col["date"] = j
                elif "customer" in c or "pelanggan" in c:
                    col["customer"] = j
                elif "desc" in c or "keterangan" in c or "project" in c:
                    col["desc"] = j
            header_idx = i
            break

    now = datetime.utcnow().isoformat()
    created = updated = skipped = 0
    for r in rows[header_idx + 1:]:
        so_raw = _clean_str(r[col["so"]]) if col["so"] < len(r) else ""
        if not so_raw:
            continue
        so_no = normalize_so_no(so_raw)
        if not so_no:
            skipped += 1
            continue
        customer = _clean_str(r[col["customer"]]) if (col["customer"] is not None and col["customer"] < len(r)) else ""
        desc = _clean_str(r[col["desc"]]) if (col["desc"] is not None and col["desc"] < len(r)) else ""
        date_v = ""
        if col["date"] is not None and col["date"] < len(r):
            date_v = _excel_serial_to_iso(r[col["date"]]) or _clean_str(r[col["date"]])

        existing = await db.sales_orders.find_one(merged({"so_no": so_no}, NOT_DELETED_FILTER))
        if existing:
            patch = {}
            if customer and not existing.get("customer"):
                patch["customer"] = customer
            if desc and not existing.get("description"):
                patch["description"] = desc
            if date_v and not existing.get("so_date"):
                patch["so_date"] = date_v
            if patch:
                await db.sales_orders.update_one({"id": existing["id"]}, {"$set": patch})
                updated += 1
            else:
                skipped += 1
            continue
        await db.sales_orders.insert_one({
            "id": str(uuid.uuid4()), "so_no": so_no, "so_date": date_v,
            "customer": customer, "description": desc,
            "created_by": current.get("id"), "created_by_username": current.get("username", ""),
            "created_at": now, "source": "so_list_import",
        })
        created += 1

    await log_action(current, "import_so_list", "sales_order", "-",
                     {"created": created, "updated": updated, "skipped": skipped, "file": file.filename})
    return {"success": True, "created": created, "updated": updated, "skipped": skipped,
            "total_rows": created + updated + skipped,
            "message": f"Import selesai: {created} SO baru, {updated} dilengkapi, {skipped} dilewati."}


@router.get("/sales-orders/check/{so_no}")
async def check_so(so_no: str, current: dict = Depends(get_current_user)):
    """Pre-check if an SO number already exists in Master List. SO dinormalisasi ke 6 digit.
    Used by forms before submitting."""
    from routers.bom import normalize_so_no
    so_no = normalize_so_no(so_no)
    d = await db.sales_orders.find_one({"so_no": so_no, "deleted_at": {"$exists": False}}, {"_id": 0})
    if not d:
        return {"exists": False, "so_no": so_no}
    # Check other quotation binding
    other_quo = await db.quotations.find_one({
        "so_no": so_no, "status": "confirm_order", "deleted_at": {"$exists": False},
    }, {"_id": 0, "quotation_no": 1, "customer_name": 1, "id": 1})
    return {"exists": True, "so_no": so_no, "master": d, "used_by_quotation": other_quo}


# =============================================================================
# STATS / DASHBOARD
# =============================================================================
INQUIRY_STATUSES = ["draft", "submitted", "in_progress", "pending_head_review", "head_revision", "awaiting_review", "accepted", "revision_requested", "closed"]
QUOTATION_STATUSES = ["on_bidding", "confirm_order", "cancel"]


@router.get("/sales/stats")
async def sales_stats(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """Aggregated dashboard for Sales & Engineering: Inquiry & Quotation counts by status.

    Optional filter by date range (start_date, end_date in YYYY-MM-DD).
    Applied to quotation.created_at and inquiry.created_at.
    """
    role = current.get("role")
    inq_filter: dict = {}
    quo_filter: dict = {}
    if is_engineering(current):
        inq_filter["status"] = {"$nin": ["draft"]}
        if role == "eng_staff":
            inq_filter["$or"] = [
                {"assigned_to_id": current.get("id")},
                {"assigned_to_id": ""},
                {"assigned_to_id": {"$exists": False}},
            ]

    # Date range filter
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = f"{end_date}T23:59:59"
        inq_filter["created_at"] = rng
        quo_filter["created_at"] = rng

    inq_pipeline = [{"$match": merged(inq_filter, NOT_DELETED_FILTER)}, {"$group": {"_id": "$status", "count": {"$sum": 1}}}]
    inq_agg = await db.inquiries.aggregate(inq_pipeline).to_list(length=None)
    inq_counts = {s: 0 for s in INQUIRY_STATUSES}
    for r in inq_agg:
        s = r["_id"] or "draft"
        inq_counts[s] = r["count"]
    inq_total = sum(inq_counts.values())

    quo_pipeline = [{"$match": merged(quo_filter, NOT_DELETED_FILTER)}, {"$group": {"_id": "$status", "count": {"$sum": 1}}}]
    quo_agg = await db.quotations.aggregate(quo_pipeline).to_list(length=None)
    quo_counts = {s: 0 for s in QUOTATION_STATUSES}
    for r in quo_agg:
        s = r["_id"] or "on_bidding"
        quo_counts[s] = r["count"]
    quo_total = sum(quo_counts.values())

    # Quotation values by status + currency
    val_pipeline = [{"$match": merged(quo_filter, NOT_DELETED_FILTER)}, {"$group": {"_id": {"status": "$status", "currency": "$currency"}, "sum": {"$sum": "$total_amount"}}}]
    val_agg = await db.quotations.aggregate(val_pipeline).to_list(length=None)
    quo_values: dict = {}
    for r in val_agg:
        st = r["_id"].get("status") or "on_bidding"
        cur = r["_id"].get("currency") or "IDR"
        quo_values.setdefault(st, {})[cur] = float(r.get("sum") or 0)

    # Unique PT (customer) counts per quotation status
    pt_pipeline = [
        {"$match": merged(quo_filter, NOT_DELETED_FILTER)},
        {"$group": {"_id": {"status": "$status", "customer": "$customer"}}},
        {"$group": {"_id": "$_id.status", "unique_pts": {"$sum": 1}}},
    ]
    pt_agg = await db.quotations.aggregate(pt_pipeline).to_list(length=None)
    unique_pts_by_status = {s: 0 for s in QUOTATION_STATUSES}
    for r in pt_agg:
        unique_pts_by_status[r["_id"] or "on_bidding"] = r["unique_pts"]

    return {
        "inquiries": {"total": inq_total, "by_status": inq_counts},
        "quotations": {
            "total": quo_total,
            "by_status": quo_counts,
            "values_by_status": quo_values,
            "unique_pts_by_status": unique_pts_by_status,
        },
        "date_range": {"start_date": start_date, "end_date": end_date},
        "role": role,
    }


# =============================================================================
# EXCEL EXPORTS
# =============================================================================
def _xl_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="1E293B")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="94A3B8")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _xl_data_border(cell):
    thin = Side(style="thin", color="E2E8F0")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


@router.get("/inquiries/export/excel")
async def export_inquiries_excel(
    status: Optional[str] = None,
    q: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """Export Inquiry list ke Excel (respect same visibility rules as GET /inquiries)."""
    filt: dict = {}
    role = current.get("role")
    if role == "sales":
        filt["created_by_id"] = current.get("id")
    if is_engineering(current):
        filt["status"] = {"$nin": ["draft"]}
        if role == "eng_staff":
            filt["$or"] = [
                {"assigned_to_id": current.get("id")},
                {"assigned_to_id": ""},
                {"assigned_to_id": {"$exists": False}},
            ]
    if status and status != "all":
        filt["status"] = status
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [{"inquiry_no": rx}, {"title": rx}, {"customer_name": rx}]
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = f"{end_date}T23:59:59"
        filt["created_at"] = rng
    docs = await db.inquiries.find(merged(filt, NOT_DELETED_FILTER)).sort("created_at", -1).to_list(length=None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiries"
    headers = ["No", "No Inquiry", "Judul", "Customer", "Deadline", "Status", "PIC Engineer",
               "Jumlah Item", "Dibuat Oleh", "Tanggal Buat", "Diterima Oleh", "Tanggal Selesai"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        _xl_header_style(c)
    ws.row_dimensions[1].height = 26

    for i, d in enumerate(docs, start=2):
        row = [
            i - 1,
            d.get("inquiry_no"),
            d.get("title"),
            d.get("customer_name"),
            d.get("customer_deadline") or "",
            (d.get("status") or "").upper(),
            d.get("pic_engineer_name") or "",
            len(d.get("items") or []),
            d.get("created_by_name") or "",
            (d.get("created_at") or "")[:10],
            d.get("accepted_by_name") or "",
            (d.get("completed_at") or "")[:10] if d.get("completed_at") else "",
        ]
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            _xl_data_border(c)
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [5, 22, 40, 28, 12, 14, 20, 8, 18, 12, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Inquiries_MKS_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    await log_action(current, "export_inquiries_excel", "inquiries", "-", {"rows": len(docs)})
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/quotations/export/excel")
async def export_quotations_excel(
    q: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    _block_engineering_from_quotation(current)
    filt: dict = {}
    if current.get("role") == "sales":
        filt["created_by_id"] = current.get("id")
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [{"quotation_no": rx}, {"customer_name": rx}, {"attention": rx}, {"items.description": rx}]
    if status and status in QUOTATION_STATUSES:
        filt["status"] = status
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = f"{end_date}T23:59:59"
        filt["created_at"] = rng
    docs = await db.quotations.find(merged(filt, NOT_DELETED_FILTER)).sort("created_at", -1).to_list(length=None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotations"
    headers = ["No", "No Quotation", "Tanggal", "Customer", "Attention", "CC",
               "Jumlah Item", "Currency", "Total Amount", "Status", "Payment Term", "Delivery", "Validity", "Sales"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        _xl_header_style(c)
    ws.row_dimensions[1].height = 26

    for i, d in enumerate(docs, start=2):
        row = [
            i - 1,
            d.get("quotation_no"),
            (d.get("created_at") or "")[:10],
            d.get("customer_name"),
            d.get("attention") or "",
            d.get("cc") or "",
            len(d.get("items") or []),
            d.get("currency") or "IDR",
            float(d.get("total_amount") or 0),
            (d.get("status") or "").upper(),
            d.get("payment_term") or "",
            d.get("delivery_time") or "",
            d.get("validity") or "",
            d.get("created_by_name") or "",
        ]
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            _xl_data_border(c)
            c.font = Font(size=10)
            if j == 9:  # Total Amount
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [5, 24, 12, 28, 20, 20, 8, 10, 18, 14, 24, 20, 22, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Quotations_MKS_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    await log_action(current, "export_quotations_excel", "quotations", "-", {"rows": len(docs)})
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )

