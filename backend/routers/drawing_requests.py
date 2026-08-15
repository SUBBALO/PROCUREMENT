"""MKS-F-ENG-001 Drawing Request Form (DRF)

Alur workflow (Iter 19):
  1. SALES buat DRF (New Order atau Repeat Order) di halaman Sales
  2. Sales submit → auto-TTD "Requested By" pakai signature PNG profile
  3. Engineering Leader (Riski) lihat di card "Drawing Request dari Sales" → Accept
  4. Accept → auto-TTD "Received By", navigate ke halaman Register Drawing yg pre-filled
  5. Riski buat drawing normal → approval workflow standard (Eng Head → QC → Sales → DC)
  6. Setelah drawing selesai (approved), status DRF jadi 'completed' → Sales notified

Model DrawingRequest:
    id, form_no (auto: MKS-F-ENG-001/xxx/YYYY)
    request_type: "new_order" | "repeat_order"
    so_no (pilih dari sales_orders)
    ref_so_no (untuk repeat order — SO referensi lama)
    date (default today)
    project_name, customer_code, customer_name
    qty_order: int, unit: str
    material (default "TBA")
    expected_due_date (ISO string)
    attached_files: list of GridFS ids
    referenced_drawings: list of drawing_ids (untuk repeat order)
    status: "draft" | "submitted" | "accepted" | "completed" | "cancelled"
    requested_by: {user_id, name, at}  ← auto saat submit
    received_by: {user_id, name, at}   ← auto saat Eng Leader accept
    linked_drawing_id: str  ← drawing yg dihasilkan Eng dari DRF ini
    created_at, updated_at
"""
from __future__ import annotations
import io
import re
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional, List

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from pydantic import BaseModel, Field

from db import db
from deps import (
    ENGINEERING_HEAD_ROLES, SALES_ROLES,
    get_current_user, is_admin_like, is_eng_head, is_engineering, is_sales_head, log_action,
)

router = APIRouter(tags=["drawing_request"])

_fs: AsyncIOMotorGridFSBucket | None = None


def _bucket() -> AsyncIOMotorGridFSBucket:
    global _fs
    if _fs is None:
        _fs = AsyncIOMotorGridFSBucket(db, bucket_name="drawing_requests")
    return _fs


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]


async def _next_form_no() -> str:
    """Generate MKS-F-ENG-001/nnn/ROMAN/YYYY (increment per month)."""
    now = datetime.now(timezone.utc)
    yy = now.year
    mm = now.month
    roman = ROMAN[mm - 1]
    # Cari nomor terakhir di bulan+tahun ini
    prefix = f"MKS-F-ENG-001/"
    suffix = f"/{roman}/{yy}"
    docs = await db.drawing_requests.find(
        {"form_no": {"$regex": f"^MKS-F-ENG-001/\\d+/{roman}/{yy}$"}},
        {"form_no": 1},
    ).to_list(length=2000)
    max_n = 0
    for d in docs:
        m = re.match(r"^MKS-F-ENG-001/(\d+)/", d.get("form_no", ""))
        if m:
            try: max_n = max(max_n, int(m.group(1)))
            except (TypeError, ValueError): pass
    return f"{prefix}{max_n + 1:03d}{suffix}"


def _sig(user: dict) -> dict:
    return {
        "user_id": user["id"],
        "username": user.get("username", ""),
        "name": user.get("name") or user.get("username", ""),
        "role": user.get("role", ""),
        "at": _now_iso(),
    }


def _clean(d: Optional[dict]) -> Optional[dict]:
    if not d: return d
    d.pop("_id", None)
    return d


def _is_sales(user: dict) -> bool:
    return user.get("role") in SALES_ROLES or is_admin_like(user)


# =========================================================================
# CRUD
# =========================================================================
VALID_REQUEST_TYPES = {"new_order", "repeat_order"}


def _require_request_type(value: Optional[str]) -> str:
    """Validasi jenis permintaan DRF dengan pesan error ramah (bukan 422 teknis)."""
    v = (value or "").strip()
    if v not in VALID_REQUEST_TYPES:
        raise HTTPException(
            status_code=400,
            detail="Jenis permintaan belum dipilih. Silakan pilih 'New Order' atau 'Repeat Order' terlebih dahulu.",
        )
    return v


class DrawingRequestCreate(BaseModel):
    request_type: Optional[str] = ""  # divalidasi manual via _require_request_type (pesan ramah)
    so_no: str
    ref_so_no: Optional[str] = ""  # untuk repeat order
    ref_so_manual: Optional[bool] = False  # True bila SO lama diinput manual (tidak ada di master)
    date: Optional[str] = ""  # kosong = today
    project_name: str = ""
    customer_code: str = ""
    customer_name: str = ""
    po_customer_no: str = ""   # No. PO dari Customer (Sales isi di DRF) → auto-isi stamping
    po_received_date: Optional[str] = ""  # Tanggal PO customer diterima (wajib di UI)
    qty_order: float = 1
    unit: str = "pcs"
    material: str = "TBA"
    items: List[dict] = []   # [{name, qty, unit, material}] — tabel item DRF (ganti qty/unit/material tunggal)
    expected_due_date: Optional[str] = ""  # Deadline Drawing (target selesai gambar)
    delivery_due_date: Optional[str] = ""  # Deadline Pengiriman barang
    notes: str = ""
    referenced_drawings: List[str] = []  # array drawing_ids untuk repeat order


def _stage(status, date="", pic="", extra=None):
    s = {"status": status, "date": date or "", "pic": pic or ""}
    if extra:
        s.update(extra)
    return s


async def _compute_so_progress(q: str = "", limit: int = 60):
    """Progress tiap SO berbasis TAHAP PROSES (Engineering -> DocCon -> Produksi -> QC -> Delivery).
    Departemen kondisional (mis. Purchasing) tidak dijadikan tahap. Fokus SO yang punya drawing."""
    # ---- Kumpulkan daftar SO dari SUMBER YANG SAMA dengan SO Tracker ----
    # SO Tracker (routers/so_tracker.py) bersumber dari drawing_requests berstatus
    # accepted/in_progress/completed. Board lama HANYA membaca sales_orders, sehingga
    # SO yang belum punya record sales_orders tidak muncul. Kini disamakan: gabungan
    # SO dari drawings (ada gambar) DAN drawing_requests (masuk workflow Engineering).
    TRACKER_STATUSES = ["accepted", "received", "in_progress", "completed"]
    dwg_so = set(s for s in await db.drawings.distinct("so_no") if s)
    drf_so = set(s for s in await db.drawing_requests.distinct(
        "so_no", {"status": {"$in": TRACKER_STATUSES}, "deleted_at": {"$exists": False}}) if s)
    # + SEMUA sales_orders aktif → SO baru (belum ada DRF/drawing) langsung tampil di papan
    #   dengan status "SO Baru — Menunggu Drawing Request" (permintaan user: apapun aktivitas SO muncul di dashboard)
    all_so = set(s for s in await db.sales_orders.distinct(
        "so_no", {"deleted_at": {"$exists": False}}) if s)
    universe = list(dwg_so | drf_so | all_so)

    # sales_orders yang ADA → dipakai untuk header (customer/desc/tanggal) bila tersedia
    so_docs = await db.sales_orders.find({"so_no": {"$in": universe}}, {"_id": 0}).to_list(length=5000) if universe else []
    so_map = {s.get("so_no"): s for s in so_docs if s.get("so_no")}

    # DRF untuk semua SO di universe → fallback header + due date + engineer
    drfs = await db.drawing_requests.find(
        {"so_no": {"$in": universe}, "deleted_at": {"$exists": False}},
        {"_id": 0, "so_no": 1, "expected_due_date": 1, "delivery_due_date": 1, "customer_name": 1, "customer": 1,
         "project_name": 1, "notes": 1, "date": 1, "created_at": 1, "assigned_engineer_name": 1, "requested_by": 1},
    ).to_list(length=5000) if universe else []
    drf_latest = {}
    for r in drfs:
        sn = r.get("so_no")
        if not sn:
            continue
        if sn not in drf_latest or (r.get("created_at") or "") >= (drf_latest[sn].get("created_at") or ""):
            drf_latest[sn] = r

    def _so_header(sn):
        s = so_map.get(sn)
        if s:
            return s
        d = drf_latest.get(sn, {})
        return {
            "so_no": sn,
            "customer": (d.get("customer_name") or d.get("customer") or "").strip(),
            "description": (d.get("project_name") or d.get("notes") or "").strip(),
            "so_date": (d.get("date") or (d.get("created_at") or "")[:10] or ""),
        }

    sos = [_so_header(sn) for sn in universe]
    if q and q.strip():
        ql = q.strip().lower()
        sos = [s for s in sos if ql in (s.get("so_no") or "").lower()
               or ql in (s.get("customer") or "").lower()
               or ql in (s.get("description") or "").lower()]
    sos.sort(key=lambda s: (s.get("so_date") or ""), reverse=True)
    sos = sos[:max(1, min(limit, 200))]
    so_nos = [s.get("so_no") for s in sos if s.get("so_no")]

    # Batch data per so_no
    drawings = await db.drawings.find(
        {"so_no": {"$in": so_nos}, "deleted_at": {"$exists": False}},
        {"_id": 0, "so_no": 1, "approval_status": 1, "approvals": 1, "updated_at": 1, "drawing_no": 1,
         "item_name": 1, "item_qty": 1, "file_id": 1, "revision_request": 1,
         "revision_opened_at": 1, "revision_approved_at": 1, "is_revision": 1},
    ).to_list(length=5000)
    deliveries = await db.deliveries.find({"so_no": {"$in": so_nos}}, {"_id": 0, "so_no": 1, "delivery_date": 1}).to_list(length=5000)
    issuances = await db.store_issuances.find({"so_no": {"$in": so_nos}}, {"_id": 0, "so_no": 1, "issue_date": 1, "created_at": 1}).to_list(length=5000)
    qc_insp = await db.qc_inspections.find({"so_no": {"$in": so_nos}, "deleted_at": {"$exists": False}}, {"_id": 0, "so_no": 1, "status": 1, "inspected_at": 1, "created_at": 1}).to_list(length=5000)

    def group(rows, key):
        g = {}
        for r in rows:
            g.setdefault(r.get(key), []).append(r)
        return g

    dwg_by = group(drawings, "so_no")
    del_by = group(deliveries, "so_no")
    iss_by = group(issuances, "so_no")
    qc_by = group(qc_insp, "so_no")
    drf_by = group(drfs, "so_no")

    APPROVED = {"approved", "controlled", "released"}
    DC_DONE = {"controlled", "released"}
    out = []
    for so in sos:
        sono = so.get("so_no")
        dws = dwg_by.get(sono, [])
        total = len(dws)
        approved = sum(1 for d in dws if (d.get("approval_status") or "") in APPROVED)
        controlled = sum(1 for d in dws if (d.get("approval_status") or "") in DC_DONE)
        last_appr = ""
        for d in dws:
            for a in (d.get("approvals") or []):
                if a.get("at", "") > last_appr:
                    last_appr = a.get("at", "")

        # 1) Engineering — semua drawing SO sudah approved
        if total == 0:
            st_eng = _stage("pending")
        elif approved >= total:
            st_eng = _stage("done", last_appr, extra={"progress": f"{approved}/{total}"})
        else:
            st_eng = _stage("in_progress", last_appr, extra={"progress": f"{approved}/{total}"})

        # 2) DocCon — semua drawing di-stamp Document Control (controlled/released)
        if total > 0 and controlled >= total:
            st_doccon = _stage("done", last_appr, extra={"progress": f"{controlled}/{total}"})
        elif controlled > 0:
            st_doccon = _stage("in_progress", "", extra={"progress": f"{controlled}/{total}"})
        else:
            st_doccon = _stage("pending")

        # 3) Produksi — material sudah di-issue dari Store (indikasi produksi berjalan)
        isss = iss_by.get(sono, [])
        if isss:
            _idate = isss[0].get("issue_date") or isss[0].get("created_at") or ""
            st_prod = _stage("done", _idate)
        else:
            st_prod = _stage("pending")

        # 4) QC — ada hasil inspeksi QC (verified) untuk SO
        qcs = qc_by.get(sono, [])
        if any((qq.get("status") or "") == "verified" for qq in qcs):
            _qdate = ""
            for qq in qcs:
                if (qq.get("status") or "") == "verified":
                    _qdate = qq.get("inspected_at") or qq.get("created_at") or ""
                    break
            st_qc = _stage("done", _qdate)
        elif qcs:
            st_qc = _stage("in_progress")
        else:
            st_qc = _stage("pending")

        # 5) Delivery — ada record pengiriman untuk SO
        dels = del_by.get(sono, [])
        st_del = _stage("done", (dels[0].get("delivery_date") if dels else "")) if dels else _stage("pending")

        stages = [
            {"key": "engineering", "label": "Engineering", **st_eng},
            {"key": "doccon", "label": "DocCon", **st_doccon},
            {"key": "produksi", "label": "Produksi", **st_prod},
            {"key": "qc", "label": "QC", **st_qc},
            {"key": "delivery", "label": "Delivery", **st_del},
        ]
        # tahap aktif = stage pertama yang belum done
        current_stage = next((s["label"] for s in stages if s["status"] != "done"), "Delivery")

        # ---- Status detail "apa yang sedang terjadi" (dibaca dari sistem) ----
        def _in_rev(d):
            if d.get("revision_request"):
                return True
            if d.get("revision_opened_at") and not d.get("revision_approved_at"):
                return True
            return False
        draft_n = sum(1 for d in dws if (d.get("approval_status") or "draft") == "draft")
        peh_n = sum(1 for d in dws if (d.get("approval_status") or "") == "pending_eng_head")
        pqc_n = sum(1 for d in dws if (d.get("approval_status") or "") == "pending_qc")
        psl_n = sum(1 for d in dws if (d.get("approval_status") or "") == "pending_sales")
        rev_n = sum(1 for d in dws if _in_rev(d))
        nofile_n = sum(1 for d in dws if not d.get("file_id"))

        if current_stage == "Engineering":
            if total == 0:
                if not drf_by.get(sono):
                    # SO baru dari Sales — belum ada Drawing Request sama sekali
                    status_now, status_kind = "SO Baru — Menunggu Drawing Request", "pending"
                else:
                    status_now, status_kind = "Menunggu Drawing (DRF)", "pending"
            elif rev_n > 0:
                status_now, status_kind = f"Revisi Drawing ({rev_n}/{total})", "revision"
            elif peh_n > 0:
                status_now, status_kind = f"Menunggu Approval Eng Leader ({peh_n}/{total})", "waiting"
            elif pqc_n > 0:
                status_now, status_kind = f"Menunggu Verifikasi QC Drawing ({pqc_n}/{total})", "waiting"
            elif psl_n > 0:
                status_now, status_kind = f"Menunggu Verifikasi Sales ({psl_n}/{total})", "waiting"
            elif draft_n > 0 or nofile_n > 0:
                status_now, status_kind = f"Gambar Drawing ({approved}/{total} selesai)", "progress"
            else:
                status_now, status_kind = f"Engineering proses ({approved}/{total})", "progress"
        elif current_stage == "DocCon":
            status_now, status_kind = f"Menunggu Stamp Document Control ({controlled}/{total})", "waiting"
        elif current_stage == "Produksi":
            status_now, status_kind = "Menunggu / Proses Produksi", "pending"
        elif current_stage == "QC":
            status_now = "Inspeksi QC berjalan" if qcs else "Menunggu QC Final"
            status_kind = "progress" if qcs else "pending"
        elif current_stage == "Delivery":
            if all(s["status"] == "done" for s in stages[:4]):
                status_now, status_kind = "Siap Kirim — Menunggu Pengiriman", "waiting"
            else:
                status_now, status_kind = "Menunggu Pengiriman", "pending"
        else:
            status_now, status_kind = "Selesai — Terkirim", "done"
        if all(s["status"] == "done" for s in stages):
            status_now, status_kind = "Selesai — Terkirim", "done"

        # ---- PIC sesuai tahap/konteks status ----
        _drf = drf_latest.get(sono) or {}
        eng_name = _drf.get("assigned_engineer_name") or ""
        sales_name = ((_drf.get("requested_by") or {}).get("name")
                      or (_drf.get("requested_by") or {}).get("username") or "")
        qc_name = ""
        for qq in qcs:
            qc_name = qq.get("inspector_name") or qq.get("inspected_by") or qq.get("inspector") or ""
            if qc_name:
                break
        if status_kind == "done":
            pic = ""
        elif "SO Baru" in status_now:
            pic = "Sales"
        elif "Verifikasi Sales" in status_now:
            pic = sales_name or "Sales"
        elif "QC" in status_now or current_stage == "QC":
            pic = qc_name or "QC"
        elif "Document Control" in status_now or current_stage == "DocCon":
            pic = "Document Control"
        elif current_stage == "Produksi":
            pic = "Produksi"
        elif current_stage == "Delivery":
            pic = "Store / Pengiriman"
        else:
            pic = eng_name  # tahap Engineering (gambar/approval/revisi)

        # ---- Dua deadline: Drawing (expected_due_date) & Pengiriman (delivery_due_date) ----
        _dl_draw = [r.get("expected_due_date") for r in drf_by.get(sono, []) if r.get("expected_due_date")]
        deadline_drawing = min(_dl_draw) if _dl_draw else ""
        _dl_del = [r.get("delivery_due_date") for r in drf_by.get(sono, []) if r.get("delivery_due_date")]
        deadline_delivery = min(_dl_del) if _dl_del else ""
        # Deadline aktif sesuai tahap: Engineering/DocCon -> Drawing; selebihnya -> Pengiriman
        if current_stage in ("Engineering", "DocCon"):
            deadline = deadline_drawing or deadline_delivery
            deadline_kind = "drawing" if deadline_drawing else ("delivery" if deadline_delivery else "")
        else:
            deadline = deadline_delivery or deadline_drawing
            deadline_kind = "delivery" if deadline_delivery else ("drawing" if deadline_drawing else "")
        # Waktu update terbaru (untuk pengurutan papan monitoring)
        _ups = [last_appr]
        for d in dws:
            _ups.append(d.get("updated_at") or "")
        for r in isss:
            _ups.append(r.get("issue_date") or r.get("created_at") or "")
        for qq in qcs:
            _ups.append(qq.get("inspected_at") or qq.get("created_at") or "")
        for dd in dels:
            _ups.append(dd.get("delivery_date") or "")
        _ups.append(so.get("so_date") or "")
        last_update = max([u for u in _ups if u], default="")
        # Ringkasan nama item (dari drawings SO ini) → biar TV tahu ini project apa
        _item_names = []
        _seen = set()
        for d in dws:
            nm = (d.get("item_name") or "").strip()
            if nm and nm.lower() not in _seen:
                _seen.add(nm.lower())
                _item_names.append(nm)
        item_summary = ", ".join(_item_names[:3])
        if len(_item_names) > 3:
            item_summary += f" +{len(_item_names) - 3} lainnya"
        # Fallback ke project_name/description bila drawing belum punya item_name
        if not item_summary:
            item_summary = (so.get("description") or "").strip()
        out.append({
            "so_no": sono,
            "customer": so.get("customer", ""),
            "description": so.get("description", ""),
            "item_summary": item_summary,
            "engineer": (drf_latest.get(sono) or {}).get("assigned_engineer_name") or "",
            "pic": pic,
            "so_date": so.get("so_date", ""),
            "deadline": deadline,
            "deadline_drawing": deadline_drawing,
            "deadline_delivery": deadline_delivery,
            "deadline_kind": deadline_kind,
            "last_update": last_update,
            "drawings_total": total,
            "drawings_approved": approved,
            "current_stage": current_stage,
            "status_now": status_now,
            "status_kind": status_kind,
            "stages": stages,
        })
    # Urutkan berdasarkan update terbaru (paling baru di atas)
    out.sort(key=lambda x: x.get("last_update") or "", reverse=True)
    return {"items": out, "count": len(out)}


@router.get("/dashboard/so-progress")
async def dashboard_so_progress(q: str = "", limit: int = 60, current: dict = Depends(get_current_user)):
    return await _compute_so_progress(q, limit)


@router.get("/public/so-progress")
async def public_so_progress(q: str = "", limit: int = 80):
    """Endpoint PUBLIK (tanpa login) untuk papan progress SO di Smart TV.
    Read-only; hanya mengembalikan status tahapan proses per SO (tanpa data finansial)."""
    return await _compute_so_progress(q, limit)


def _clean_drf_items(raw) -> list:
    """Normalisasi tabel item DRF: [{name, qty, unit, material}]. Buang baris tanpa nama."""
    out = []
    for it in (raw or []):
        if not isinstance(it, dict):
            continue
        name = str(it.get("name") or "").strip()
        if not name:
            continue
        try:
            qty = float(it.get("qty") or 0)
        except (TypeError, ValueError):
            qty = 0
        out.append({
            "name": name,
            "qty": qty,
            "unit": str(it.get("unit") or "pcs").strip() or "pcs",
            "material": str(it.get("material") or "TBA").strip() or "TBA",
        })
    return out


@router.post("/drawing-requests")
async def create_drawing_request(
    payload: DrawingRequestCreate,
    current: dict = Depends(get_current_user),
):
    """Sales membuat DRF baru (status draft)."""
    if not _is_sales(current):
        raise HTTPException(status_code=403, detail="Hanya Sales yang boleh buat Drawing Request")
    req_type = _require_request_type(payload.request_type)
    if not payload.so_no.strip():
        raise HTTPException(status_code=400, detail="SO wajib dipilih")

    now = _now_iso()
    clean_items = _clean_drf_items(payload.items)
    # qty_order legacy diisi dari total qty item (bila ada) untuk kompatibilitas tampilan lama
    legacy_qty = sum((float(it.get("qty") or 0) for it in clean_items)) if clean_items else payload.qty_order
    doc = {
        "id": str(uuid.uuid4()),
        "form_no": await _next_form_no(),
        "request_type": req_type,
        "so_no": payload.so_no.strip(),
        "ref_so_no": (payload.ref_so_no or "").strip(),
        "ref_so_manual": bool(payload.ref_so_manual),
        "date": payload.date or now[:10],
        "project_name": payload.project_name.strip(),
        "customer_code": payload.customer_code.strip(),
        "customer_name": payload.customer_name.strip(),
        "po_customer_no": (payload.po_customer_no or "").strip(),
        "po_received_date": payload.po_received_date or "",
        "qty_order": legacy_qty,
        "unit": payload.unit,
        "material": payload.material or "TBA",
        "items": clean_items,
        "expected_due_date": payload.expected_due_date or "",
        "delivery_due_date": payload.delivery_due_date or "",
        "notes": payload.notes,
        "referenced_drawings": payload.referenced_drawings or [],
        "attached_files": [],
        "status": "draft",
        "requested_by": None,
        "received_by": None,
        "linked_drawing_id": None,
        "created_by": current["id"],
        "created_at": now,
        "updated_at": now,
    }
    await db.drawing_requests.insert_one(doc.copy())
    await log_action(current, "drf_create", "drawing_requests", doc["id"], {"form_no": doc["form_no"], "so_no": doc["so_no"]})
    return _clean(doc)


@router.get("/drawing-requests")
async def list_drawing_requests(
    status: Optional[str] = None,
    scope: Optional[str] = None,  # "mine" | "for_engineering" | "for_sales_ttd"
    current: dict = Depends(get_current_user),
):
    """List DRF dengan filter."""
    filt: dict = {"deleted_at": {"$exists": False}}
    if status: filt["status"] = status
    if scope == "mine":
        filt["created_by"] = current["id"]
    elif scope == "for_engineering":
        # Engineering Head lihat request yg:
        # - status "submitted" (belum di-accept)
        # - status "accepted" tapi belum ada linked drawing
        # - status "in_progress" (drawing sudah dibuat, masih dikerjakan) — untuk tracking
        filt["$or"] = [
            {"status": "submitted"},
            {"status": "accepted", "linked_drawing_id": {"$in": [None, ""]}},
            {"status": "received", "linked_drawing_id": {"$in": [None, ""]}},
            {"status": "in_progress"},
        ]
    elif scope == "for_sales_ttd":
        # Sales lihat DRF yg drawing-nya sudah completed (butuh TTD approval Sales)
        filt["status"] = "completed"
        filt["created_by"] = current["id"]
    docs = await db.drawing_requests.find(filt, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    return {"items": docs, "total": len(docs)}


@router.get("/drawing-requests/pending-count-for-engineering")
async def pending_count_for_eng(current: dict = Depends(get_current_user)):
    """Untuk badge card di Eng Portal."""
    if not (is_eng_head(current) or is_admin_like(current)):
        return {"count": 0}
    n = await db.drawing_requests.count_documents({
        "$or": [
            {"status": "submitted"},
            {"status": "accepted", "linked_drawing_id": {"$in": [None, ""]}},
        ],
        "deleted_at": {"$exists": False},
    })
    return {"count": n}


@router.get("/engineering/pending-leader-verification")
async def pending_leader_verification(current: dict = Depends(get_current_user)):
    """Antrian 'Menunggu Verifikasi Leader': DRF/SO yang punya drawing
    berstatus pending_eng_head (menunggu review & TTD Eng Leader).
    Dikelompokkan per DRF agar klik item langsung membuka Review Dokumen SO."""
    if not (is_eng_head(current) or is_admin_like(current)):
        return {"items": [], "count": 0}

    drawings = await db.drawings.find(
        {"approval_status": "pending_eng_head", "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "from_drf_id": 1, "bom_id": 1, "bom_no": 1,
         "so_no": 1, "drawing_no": 1, "submitted_at": 1, "updated_at": 1,
         "customer_name": 1, "project_name": 1},
    ).sort("submitted_at", 1).to_list(length=1000)

    groups: dict = {}
    for d in drawings:
        key = d.get("from_drf_id") or f"__no_drf__{d.get('id')}"
        g = groups.get(key)
        if not g:
            g = {
                "drf_id": d.get("from_drf_id") or "",
                "bom_id": d.get("bom_id") or "",
                "bom_no": d.get("bom_no") or "",
                "so_no": d.get("so_no") or "",
                "customer_name": d.get("customer_name") or "",
                "project_name": d.get("project_name") or "",
                "pending_count": 0,
                "oldest_submitted_at": "",
                "drawing_nos": [],
            }
            groups[key] = g
        g["pending_count"] += 1
        if not g["bom_id"] and d.get("bom_id"):
            g["bom_id"] = d["bom_id"]
            g["bom_no"] = d.get("bom_no") or g["bom_no"]
        if d.get("drawing_no"):
            g["drawing_nos"].append(d["drawing_no"])
        sub = d.get("submitted_at") or d.get("updated_at") or ""
        if sub and (not g["oldest_submitted_at"] or sub < g["oldest_submitted_at"]):
            g["oldest_submitted_at"] = sub

    # Tambahan (workflow BOM terpisah): BOM berstatus pending_review juga masuk antrian
    # Leader — walau tidak ada drawing yang pending. Grup via SO/DRF.
    pending_boms = await db.boms.find(
        {"engineering_status": "pending_review", "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "bom_no": 1, "so_no": 1, "submitted_at": 1, "updated_at": 1},
    ).sort("submitted_at", 1).to_list(length=1000)
    bom_so_nos = list({b.get("so_no") for b in pending_boms if b.get("so_no")})
    so_to_drf: dict = {}
    if bom_so_nos:
        drfs_for_bom = await db.drawing_requests.find(
            {"so_no": {"$in": bom_so_nos}},
            {"_id": 0, "id": 1, "so_no": 1, "form_no": 1, "customer_name": 1,
             "project_name": 1, "expected_due_date": 1, "created_at": 1},
        ).sort("created_at", -1).to_list(length=2000)
        for r in drfs_for_bom:
            so_to_drf.setdefault(r.get("so_no"), r)  # ambil DRF terbaru per SO
    bom_key_seen: set = set()
    for b in pending_boms:
        so = b.get("so_no")
        drf = so_to_drf.get(so) or {}
        key = drf.get("id") or f"__bomonly__{so or b.get('id')}"
        g = groups.get(key)
        if not g:
            g = {
                "drf_id": drf.get("id") or "",
                "bom_id": "",
                "bom_no": "",
                "so_no": so or "",
                "customer_name": drf.get("customer_name") or "",
                "project_name": drf.get("project_name") or "",
                "pending_count": 0,
                "oldest_submitted_at": "",
                "drawing_nos": [],
            }
            groups[key] = g
        g["bom_pending_count"] = g.get("bom_pending_count", 0) + 1
        # Prioritaskan BOM pending sebagai target review (BOM pending tertua per SO).
        if key not in bom_key_seen:
            g["bom_id"] = b.get("id")
            g["bom_no"] = b.get("bom_no") or g.get("bom_no") or ""
            bom_key_seen.add(key)
        sub = b.get("submitted_at") or b.get("updated_at") or ""
        if sub and (not g["oldest_submitted_at"] or sub < g["oldest_submitted_at"]):
            g["oldest_submitted_at"] = sub

    drf_ids = [g["drf_id"] for g in groups.values() if g["drf_id"]]
    drf_map: dict = {}
    if drf_ids:
        drfs = await db.drawing_requests.find(
            {"id": {"$in": drf_ids}},
            {"_id": 0, "id": 1, "form_no": 1, "so_no": 1, "customer_name": 1,
             "project_name": 1, "expected_due_date": 1},
        ).to_list(length=len(drf_ids))
        drf_map = {r["id"]: r for r in drfs}

    items = []
    for g in groups.values():
        info = drf_map.get(g["drf_id"], {})
        total = 0
        if g["drf_id"]:
            total = await db.drawings.count_documents(
                {"from_drf_id": g["drf_id"], "deleted_at": {"$exists": False}})
        items.append({
            "drf_id": g["drf_id"],
            "form_no": info.get("form_no") or "",
            "so_no": g["so_no"] or info.get("so_no") or "-",
            "customer_name": g["customer_name"] or info.get("customer_name") or "",
            "project_name": g["project_name"] or info.get("project_name") or "",
            "bom_id": g["bom_id"],
            "bom_no": g["bom_no"],
            "pending_count": g["pending_count"],
            "bom_pending_count": g.get("bom_pending_count", 0),
            "total_drawings": total,
            "drawing_nos": g["drawing_nos"][:6],
            "due_date": info.get("expected_due_date") or "",
            "oldest_submitted_at": g["oldest_submitted_at"],
        })
    items.sort(key=lambda x: (x["oldest_submitted_at"] or "9999"))
    return {"items": items, "count": len(items)}




@router.get("/drawing-requests/engineering-users")
async def drf_engineering_users(current: dict = Depends(get_current_user)):
    """Daftar user Engineering untuk dropdown assign (dipakai Eng Leader saat accept DRF).
    Didefinisikan SEBELUM route /{drf_id} agar tidak tertangkap sebagai id."""
    if not (is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Eng Leader/Admin")
    from deps import ENGINEERING_ROLES
    users = await db.users.find(
        {"role": {"$in": list(ENGINEERING_ROLES)}, "active": {"$ne": False}},
        {"_id": 0, "id": 1, "username": 1, "name": 1, "role": 1},
    ).sort("name", 1).to_list(length=200)
    return {"items": users}


@router.get("/drawing-requests/my-queue")
async def my_job_queue(current: dict = Depends(get_current_user)):
    """Antrian job untuk engineer yang login: DRF yang di-assign ke dia.
    3 tahap:
    - antri     (accepted)    → belum diterima → tombol TERIMA
    - diterima  (received)    → sudah diterima, belum digambar → tombol MULAI KERJAKAN
    - proses    (in_progress) → sedang dikerjakan → buka Work Order
    Didefinisikan SEBELUM route /{drf_id} agar tidak tertangkap sebagai id."""
    uid = current.get("id")
    if not uid:
        raise HTTPException(status_code=401, detail="Unauthorized")
    docs = await db.drawing_requests.find(
        {"assigned_engineer_id": uid,
         "status": {"$in": ["accepted", "received", "in_progress"]},
         "deleted_at": {"$exists": False}},
        {"_id": 0},
    ).sort("assigned_at", -1).to_list(length=200)
    antri, diterima, proses = [], [], []
    for d in docs:
        d = _clean(d)
        st = d.get("status")
        if st == "in_progress" or d.get("work_started_at"):
            proses.append(d)
        elif st == "received" or d.get("work_received_at"):
            diterima.append(d)
        else:
            antri.append(d)
    # Inquiry costing yang di-assign ke engineer ini (belum selesai)
    inq_docs = await db.inquiries.find(
        {"assigned_to_id": uid,
         "status": {"$nin": ["completed", "rejected", "cancelled", "draft", "closed"]},
         "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "inquiry_no": 1, "title": 1, "customer_name": 1, "project_name": 1,
         "status": 1, "assigned_at": 1, "accepted_at": 1, "work_started_at": 1, "customer_deadline": 1},
    ).sort("assigned_at", -1).to_list(length=200)
    inquiry_antri, inquiry_diterima, inquiry_proses = [], [], []
    for iq in inq_docs:
        started = bool(iq.get("work_started_at"))
        accepted = bool(iq.get("accepted_at"))
        stage = "proses" if started else ("diterima" if accepted else "belum_terima")
        row = {
            "id": iq.get("id"), "inquiry_no": iq.get("inquiry_no"), "title": iq.get("title"),
            "customer_name": iq.get("customer_name"), "project_name": iq.get("project_name"),
            "status": iq.get("status"), "assigned_at": iq.get("assigned_at"),
            "accepted_at": iq.get("accepted_at"), "work_started_at": iq.get("work_started_at"),
            "deadline": iq.get("customer_deadline"), "stage": stage,
        }
        if started:
            inquiry_proses.append(row)
        elif accepted:
            inquiry_diterima.append(row)
        else:
            inquiry_antri.append(row)

    return {
        "antri": antri, "diterima": diterima, "proses": proses,
        "antri_count": len(antri), "diterima_count": len(diterima), "proses_count": len(proses),
        # Inquiry costing — 3 tahap sama (Terima → Kerjakan), digabung di satu halaman Tugas Saya
        "inquiry_antri": inquiry_antri, "inquiry_diterima": inquiry_diterima, "inquiry_proses": inquiry_proses,
        "inquiry_antri_count": len(inquiry_antri), "inquiry_diterima_count": len(inquiry_diterima), "inquiry_proses_count": len(inquiry_proses),
        # backward-compat (konsumen lama)
        "pending": antri, "in_progress": proses,
        "pending_count": len(antri), "in_progress_count": len(proses),
    }


@router.get("/drawing-requests/stage-board")
async def drf_stage_board(current: dict = Depends(get_current_user)):
    """Ringkasan beban DRF per engineer dalam 3 tahap (untuk monitor Bos/Leader):
    - antri    (accepted)    : sudah ditugaskan, engineer belum Terima
    - diterima (received)    : sudah diterima, belum digambar
    - proses   (in_progress) : sedang dikerjakan
    Juga mengembalikan jumlah 'submitted' yang masih menunggu di-assign Eng Leader."""
    if not (is_engineering(current) or is_admin_like(current) or is_sales_head(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering / Leader / Direktur")
    from deps import ENGINEERING_ROLES
    users = await db.users.find(
        {"role": {"$in": list(ENGINEERING_ROLES)}, "active": {"$ne": False}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "username": 1, "name": 1, "role": 1},
    ).sort("name", 1).to_list(length=200)
    per = {u["id"]: {
        "user_id": u["id"], "name": u.get("name") or u.get("username"),
        "username": u.get("username"), "role": u.get("role"),
        "antri": 0, "diterima": 0, "proses": 0, "total": 0,
    } for u in users}

    docs = await db.drawing_requests.find(
        {"status": {"$in": ["accepted", "received", "in_progress"]}, "deleted_at": {"$exists": False}},
        {"_id": 0, "assigned_engineer_id": 1, "assigned_engineer_name": 1, "status": 1,
         "work_received_at": 1, "work_started_at": 1},
    ).to_list(length=5000)

    unassigned = {"antri": 0, "diterima": 0, "proses": 0}
    for d in docs:
        st = d.get("status")
        bucket = "proses" if (st == "in_progress" or d.get("work_started_at")) else (
            "diterima" if (st == "received" or d.get("work_received_at")) else "antri")
        uid = d.get("assigned_engineer_id")
        target = per.get(uid)
        if target is None:
            unassigned[bucket] += 1
            continue
        target[bucket] += 1
        target["total"] += 1

    items = [p for p in per.values()]
    items.sort(key=lambda x: (x["total"], x["proses"]), reverse=True)
    submitted_waiting = await db.drawing_requests.count_documents(
        {"status": "submitted", "deleted_at": {"$exists": False}})
    totals = {
        "antri": sum(p["antri"] for p in items) + unassigned["antri"],
        "diterima": sum(p["diterima"] for p in items) + unassigned["diterima"],
        "proses": sum(p["proses"] for p in items) + unassigned["proses"],
        "submitted_waiting": submitted_waiting,
    }
    return {"items": items, "totals": totals, "submitted_waiting": submitted_waiting}



@router.get("/engineering/logwork")
async def engineering_logwork(current: dict = Depends(get_current_user)):
    """Logwork per engineer untuk monitor Direktur/Leader.
    Berisi beban SAAT INI (DRF/SO, Inquiry, Drawing yang sedang dipegang) yang bisa
    diklik ke detail, plus RIWAYAT aktivitas terakhir (terima/mulai/selesai)."""
    if not (is_engineering(current) or is_admin_like(current) or is_sales_head(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering / Leader / Direktur")
    from deps import ENGINEERING_ROLES
    users = await db.users.find(
        {"role": {"$in": list(ENGINEERING_ROLES)}, "active": {"$ne": False}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "username": 1, "name": 1, "role": 1},
    ).sort("name", 1).to_list(length=200)

    def _stage(d):
        st = d.get("status")
        if st == "in_progress" or d.get("work_started_at"):
            return "proses"
        if st == "received" or d.get("work_received_at"):
            return "diterima"
        return "antri"

    per = {}
    for u in users:
        per[u["id"]] = {
            "user_id": u["id"], "name": u.get("name") or u.get("username"),
            "username": u.get("username"), "role": u.get("role"),
            "counts": {"antri": 0, "diterima": 0, "proses": 0},
            "drf": [], "inquiry": [], "drawing": [], "history": [],
        }

    # --- Active DRF ---
    drfs = await db.drawing_requests.find(
        {"status": {"$in": ["accepted", "received", "in_progress"]}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "form_no": 1, "so_no": 1, "project_name": 1, "customer_name": 1,
         "status": 1, "request_type": 1, "assigned_engineer_id": 1, "assigned_at": 1,
         "work_received_at": 1, "work_started_at": 1, "expected_due_date": 1},
    ).to_list(length=5000)
    for d in drfs:
        p = per.get(d.get("assigned_engineer_id"))
        if not p:
            continue
        stg = _stage(d)
        p["counts"][stg] += 1
        p["drf"].append({
            "id": d.get("id"), "form_no": d.get("form_no"), "so_no": d.get("so_no"),
            "project_name": d.get("project_name"), "customer_name": d.get("customer_name"),
            "request_type": d.get("request_type"), "stage": stg,
            "assigned_at": d.get("assigned_at"), "work_received_at": d.get("work_received_at"),
            "work_started_at": d.get("work_started_at"), "due_date": d.get("expected_due_date"),
        })
        # history events dari DRF ini
        if d.get("work_received_at"):
            p["history"].append({"at": d["work_received_at"], "event": "Diterima", "type": "drf",
                                  "ref": d.get("id"), "label": f"{d.get('form_no')} · SO {d.get('so_no') or '-'}"})
        if d.get("work_started_at"):
            p["history"].append({"at": d["work_started_at"], "event": "Mulai Kerjakan", "type": "drf",
                                  "ref": d.get("id"), "label": f"{d.get('form_no')} · SO {d.get('so_no') or '-'}"})

    # --- Active Inquiry (costing) ---
    inqs = await db.inquiries.find(
        {"assigned_to_id": {"$nin": ["", None]},
         "status": {"$nin": ["completed", "rejected", "cancelled", "draft"]},
         "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "inquiry_no": 1, "title": 1, "customer_name": 1, "status": 1,
         "assigned_to_id": 1, "assigned_at": 1},
    ).to_list(length=5000)
    for iq in inqs:
        p = per.get(iq.get("assigned_to_id"))
        if not p:
            continue
        p["inquiry"].append({
            "id": iq.get("id"), "inquiry_no": iq.get("inquiry_no"), "title": iq.get("title"),
            "customer_name": iq.get("customer_name"), "status": iq.get("status"),
            "assigned_at": iq.get("assigned_at"),
        })

    # --- Active Drawing (belum controlled/released) ---
    drws = await db.drawings.find(
        {"assigned_to_user_id": {"$nin": ["", None]},
         "approval_status": {"$nin": ["controlled", "released"]},
         "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "drawing_no": 1, "so_no": 1, "approval_status": 1, "assigned_to_user_id": 1, "title": 1},
    ).to_list(length=20000)
    for dr in drws:
        p = per.get(dr.get("assigned_to_user_id"))
        if not p:
            continue
        p["drawing"].append({
            "id": dr.get("id"), "drawing_no": dr.get("drawing_no"), "so_no": dr.get("so_no"),
            "approval_status": dr.get("approval_status"), "title": dr.get("title"),
        })

    items = []
    for p in per.values():
        p["history"].sort(key=lambda x: x.get("at") or "", reverse=True)
        p["history"] = p["history"][:15]
        p["total_active"] = len(p["drf"]) + len(p["inquiry"]) + len(p["drawing"])
        items.append(p)
    items.sort(key=lambda x: x["total_active"], reverse=True)
    return {"items": items}





async def _compute_workload(start: str = "", end: str = ""):
    """Hitung beban kerja per engineer.
    - Mode AKTIF (tanpa start/end): item yang sedang berjalan sekarang.
    - Mode PERIODE (start & end diisi): item yang DIBUAT/di-assign dalam rentang tanggal (semua status),
      untuk laporan mingguan/bulanan."""
    from deps import ENGINEERING_ROLES
    users = await db.users.find(
        {"role": {"$in": list(ENGINEERING_ROLES)}, "active": {"$ne": False}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "username": 1, "name": 1, "role": 1},
    ).sort("name", 1).to_list(length=200)
    ids = [u["id"] for u in users]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    period = bool(start and end)
    s0, s1 = (start or "")[:10], (end or "")[:10]

    def _overdue(v):
        return bool(v) and str(v)[:10] < today

    def _in_range(v):
        if not v:
            return False
        return s0 <= str(v)[:10] <= s1

    stats = {u["id"]: {
        "user_id": u["id"], "name": u.get("name") or u.get("username"),
        "username": u.get("username"), "role": u.get("role"),
        "drf": 0, "drawing": 0, "inquiry": 0, "ecn": 0, "overdue": 0, "total": 0,
    } for u in users}

    # DRF
    drf_q = ({"assigned_engineer_id": {"$in": ids}, "deleted_at": {"$exists": False}} if period
             else {"status": {"$in": ["accepted", "received", "in_progress"]}, "deleted_at": {"$exists": False}})
    for d in await db.drawing_requests.find(drf_q, {"_id": 0, "assigned_engineer_id": 1, "created_at": 1, "received_at": 1, "expected_due_date": 1, "due_date": 1}).to_list(length=5000):
        uid = d.get("assigned_engineer_id")
        if uid not in stats:
            continue
        if period and not _in_range(d.get("received_at") or d.get("created_at")):
            continue
        stats[uid]["drf"] += 1
        if _overdue(d.get("expected_due_date") or d.get("due_date")):
            stats[uid]["overdue"] += 1

    # Drawing + ECN
    dr_q = ({"assigned_to_user_id": {"$in": ids}, "deleted_at": {"$exists": False}} if period
            else {"approval_status": {"$nin": ["controlled", "released"]}, "deleted_at": {"$exists": False}})
    for d in await db.drawings.find(dr_q, {"_id": 0, "assigned_to_user_id": 1, "created_at": 1, "revision_request": 1}).to_list(length=20000):
        uid = d.get("assigned_to_user_id")
        if uid in stats and (not period or _in_range(d.get("created_at"))):
            stats[uid]["drawing"] += 1
        rr = d.get("revision_request") or {}
        if period:
            if rr and _in_range(rr.get("created_at") or rr.get("requested_at")):
                euid = rr.get("requested_by_id") or d.get("assigned_to_user_id")
                if euid in stats:
                    stats[euid]["ecn"] += 1
        elif rr.get("status") in ("pending", "in_progress"):
            euid = rr.get("requested_by_id") or d.get("assigned_to_user_id")
            if euid in stats:
                stats[euid]["ecn"] += 1

    # Inquiry
    inq_q = ({"assigned_to_id": {"$in": ids}, "deleted_at": {"$exists": False}} if period
             else {"assigned_to_id": {"$nin": ["", None]}, "status": {"$nin": ["completed", "rejected", "cancelled", "draft"]}, "deleted_at": {"$exists": False}})
    for iq in await db.inquiries.find(inq_q, {"_id": 0, "assigned_to_id": 1, "assigned_at": 1, "created_at": 1, "due_date": 1, "target_date": 1}).to_list(length=5000):
        uid = iq.get("assigned_to_id")
        if uid not in stats:
            continue
        if period and not _in_range(iq.get("assigned_at") or iq.get("created_at")):
            continue
        stats[uid]["inquiry"] += 1
        if _overdue(iq.get("due_date") or iq.get("target_date")):
            stats[uid]["overdue"] += 1

    out = []
    for st in stats.values():
        st["total"] = st["drf"] + st["drawing"] + st["inquiry"] + st["ecn"]
        st["level"] = "overload" if st["total"] > 6 else ("busy" if st["total"] >= 4 else "normal")
        out.append(st)
    out.sort(key=lambda x: (x["total"], x["overdue"]), reverse=True)
    summary = {
        "engineers": len(out),
        "total_active": sum(s["total"] for s in out),
        "overload": len([s for s in out if s["level"] == "overload"]),
        "busy": len([s for s in out if s["level"] == "busy"]),
        "normal": len([s for s in out if s["level"] == "normal"]),
        "overdue": sum(s["overdue"] for s in out),
    }
    return {"items": out, "summary": summary, "thresholds": {"busy": 4, "overload": 7},
            "mode": "period" if period else "active", "start": s0, "end": s1}


@router.get("/engineering/workload")
async def engineering_workload(start: str = "", end: str = "", current: dict = Depends(get_current_user)):
    """Monitor beban kerja per engineer. Tanpa start/end = beban aktif sekarang.
    Dengan start & end (YYYY-MM-DD) = laporan periode (mingguan/bulanan)."""
    if not (is_engineering(current) or is_admin_like(current) or is_sales_head(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering / Admin")
    return await _compute_workload(start, end)


@router.get("/engineering/workload/export")
async def engineering_workload_export(format: str = "xlsx", start: str = "", end: str = "",
                                      current: dict = Depends(get_current_user)):
    """Export laporan beban kerja (Excel/PDF) sesuai tampilan monitor + rentang tanggal."""
    if not (is_engineering(current) or is_admin_like(current) or is_sales_head(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering / Admin")
    from fastapi.responses import StreamingResponse
    data = await _compute_workload(start, end)
    rows = data["items"]
    period_txt = f"{data['start']} s/d {data['end']}" if data["mode"] == "period" else "Beban Aktif (Saat Ini)"
    headers = ["Engineer", "Role", "DRF", "Drawing", "Inquiry", "ECN/Revisi", "Terlambat", "Total", "Status"]
    lvl_id = {"overload": "OVERLOAD", "busy": "SIBUK", "normal": "NORMAL"}
    fname_base = f"Beban_Kerja_Engineer_{(data['start'] or 'aktif')}_{(data['end'] or '')}".strip("_")

    if format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        elems = [Paragraph("Laporan Beban Kerja Engineer — PT. Mitra Karya Sarana", styles["Title"]),
                 Paragraph(f"Periode: {period_txt}", styles["Normal"]), Spacer(1, 8)]
        table_data = [headers] + [[r["name"], (r.get("role") or "").replace("_", " "), r["drf"], r["drawing"], r["inquiry"], r["ecn"], r["overdue"], r["total"], lvl_id.get(r["level"], r["level"])] for r in rows]
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#b45309")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("ALIGN", (2, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elems.append(t)
        doc.build(elems)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{fname_base}.pdf"'})

    # Excel (default)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Beban Kerja"
    ws["A1"] = "Laporan Beban Kerja Engineer — PT. Mitra Karya Sarana"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Periode: {period_txt}"
    ws["A2"].font = Font(italic=True, size=10)
    hdr_fill = PatternFill("solid", fgColor="B45309")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for i, r in enumerate(rows, start=5):
        vals = [r["name"], (r.get("role") or "").replace("_", " "), r["drf"], r["drawing"], r["inquiry"], r["ecn"], r["overdue"], r["total"], lvl_id.get(r["level"], r["level"])]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = border
            if c >= 3:
                cell.alignment = Alignment(horizontal="center")
    widths = [22, 16, 8, 10, 10, 12, 11, 8, 12]
    from openpyxl.utils import get_column_letter
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{fname_base}.xlsx"'})


@router.get("/engineering/workload/detail")
async def engineering_workload_detail(user_id: str, current: dict = Depends(get_current_user)):
    """Rincian item beban aktif seorang engineer (view-only): daftar DRF, Drawing, Inquiry, ECN
    yang sedang menjadi beban. Dipakai saat angka breakdown di Monitor diklik."""
    if not (is_engineering(current) or is_admin_like(current) or is_sales_head(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering / Admin")
    u = await db.users.find_one({"id": user_id}, {"_id": 0, "id": 1, "name": 1, "username": 1})
    if not u:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")

    # DRF aktif
    drf = []
    for d in await db.drawing_requests.find(
        {"assigned_engineer_id": user_id, "status": {"$in": ["accepted", "received", "in_progress"]}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "form_no": 1, "so_no": 1, "customer_name": 1, "project_name": 1, "status": 1,
         "expected_due_date": 1, "due_date": 1, "qty_order": 1, "unit": 1, "po_received_date": 1,
         "date": 1, "created_at": 1, "requested_by": 1},
    ).sort("created_at", -1).to_list(length=1000):
        rb = d.get("requested_by") or {}
        drf.append({
            "id": d.get("id"), "no": d.get("form_no", "-"), "so_no": d.get("so_no", "-"),
            "title": " · ".join([x for x in [d.get("customer_name"), d.get("project_name")] if x]) or "-",
            "status": d.get("status", "-"), "due": d.get("expected_due_date") or d.get("due_date") or "",
            "customer_name": d.get("customer_name") or "-",
            "qty": (f"{d.get('qty_order')} {d.get('unit') or ''}".strip() if d.get("qty_order") not in (None, "") else "-"),
            "order_date": (d.get("po_received_date") or d.get("date") or d.get("created_at") or "")[:10],
            "plan_finish": (d.get("expected_due_date") or d.get("due_date") or "")[:10],
            "request_from": rb.get("name") or "-",
        })

    # Drawing aktif + ECN/revisi aktif
    drawing, ecn = [], []
    for d in await db.drawings.find(
        {"assigned_to_user_id": user_id, "approval_status": {"$nin": ["controlled", "released"]}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "drawing_no": 1, "so_no": 1, "part_name": 1, "description": 1, "approval_status": 1},
    ).sort("created_at", -1).to_list(length=5000):
        drawing.append({
            "id": d.get("id"), "no": d.get("drawing_no", "-"), "so_no": d.get("so_no", "-"),
            "title": d.get("part_name") or d.get("description") or "-",
            "status": d.get("approval_status", "-"), "due": "",
        })
    # ECN aktif = revision_request pending/in_progress (untuk user ini)
    for d in await db.drawings.find(
        {"revision_request.status": {"$in": ["pending", "in_progress"]}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "drawing_no": 1, "so_no": 1, "assigned_to_user_id": 1, "revision_request": 1},
    ).to_list(length=5000):
        rr = d.get("revision_request") or {}
        euid = rr.get("requested_by_id") or d.get("assigned_to_user_id")
        if euid == user_id:
            ecn.append({
                "id": d.get("id"), "no": rr.get("ecn_no") or d.get("drawing_no", "-"), "so_no": d.get("so_no", "-"),
                "title": rr.get("reason") or rr.get("notes") or d.get("drawing_no", "-"),
                "status": rr.get("status", "-"), "due": "",
            })

    # Inquiry aktif
    inquiry = []
    for iq in await db.inquiries.find(
        {"assigned_to_id": user_id, "status": {"$nin": ["completed", "rejected", "cancelled", "draft"]}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "inquiry_no": 1, "customer_name": 1, "title": 1, "project_name": 1, "status": 1, "due_date": 1, "target_date": 1, "accepted_at": 1, "work_started_at": 1},
    ).sort("created_at", -1).to_list(length=2000):
        # Status jelas berbasis tahap terima/kerjakan (bukan sekadar 'in_progress')
        if iq.get("work_started_at"):
            stage_label = "Dikerjakan"
        elif iq.get("accepted_at"):
            stage_label = "Diterima (antri)"
        else:
            stage_label = "Belum Diterima"
        inquiry.append({
            "id": iq.get("id"), "no": iq.get("inquiry_no", "-"), "so_no": "-",
            "title": iq.get("title") or iq.get("project_name") or iq.get("customer_name") or "-",
            "status": stage_label, "due": iq.get("due_date") or iq.get("target_date") or "",
        })

    return {
        "user": {"id": u.get("id"), "name": u.get("name") or u.get("username")},
        "drf": drf, "drawing": drawing, "inquiry": inquiry, "ecn": ecn,
        "counts": {"drf": len(drf), "drawing": len(drawing), "inquiry": len(inquiry), "ecn": len(ecn)},
    }


@router.get("/engineering/workload/trend")
async def engineering_workload_trend(weeks: int = 8, current: dict = Depends(get_current_user)):
    """Tren beban mingguan per engineer (N minggu terakhir).
    Menghitung jumlah tugas BARU (DRF + Drawing + Inquiry + ECN) yang di-assign ke tiap engineer
    berdasarkan tanggal dibuat, dikelompokkan per minggu (Senin–Minggu). Untuk lihat siapa yang
    konsisten padat."""
    if not (is_engineering(current) or is_admin_like(current) or is_sales_head(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering / Admin")
    from deps import ENGINEERING_ROLES
    weeks = max(4, min(16, int(weeks or 8)))

    now = datetime.now(timezone.utc)
    monday = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    starts = [monday - timedelta(weeks=(weeks - 1 - i)) for i in range(weeks)]  # lama → baru
    labels = [s.strftime("%d %b") for s in starts]
    first = starts[0]

    def _week_index(v):
        if not v:
            return None
        s = str(v)
        try:
            d = datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            try:
                d = datetime.strptime(s[:10], "%Y-%m-%d")
            except Exception:
                return None
        if d.tzinfo is None:
            d = d.replace(tzinfo=timezone.utc)
        if d < first:
            return None
        for i in range(weeks):
            if starts[i] <= d < starts[i] + timedelta(weeks=1):
                return i
        return None

    users = await db.users.find(
        {"role": {"$in": list(ENGINEERING_ROLES)}, "active": {"$ne": False}, "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "username": 1, "name": 1, "role": 1},
    ).sort("name", 1).to_list(length=200)
    ids = [u["id"] for u in users]
    series = {u["id"]: [0] * weeks for u in users}

    def _bump(uid, when):
        if uid in series:
            idx = _week_index(when)
            if idx is not None:
                series[uid][idx] += 1

    # DRF
    for d in await db.drawing_requests.find(
        {"assigned_engineer_id": {"$in": ids}, "deleted_at": {"$exists": False}},
        {"_id": 0, "assigned_engineer_id": 1, "created_at": 1, "received_at": 1},
    ).to_list(length=5000):
        _bump(d.get("assigned_engineer_id"), d.get("received_at") or d.get("created_at"))

    # Drawing + ECN
    for d in await db.drawings.find(
        {"assigned_to_user_id": {"$in": ids}, "deleted_at": {"$exists": False}},
        {"_id": 0, "assigned_to_user_id": 1, "created_at": 1, "revision_request": 1},
    ).to_list(length=20000):
        _bump(d.get("assigned_to_user_id"), d.get("created_at"))
        rr = d.get("revision_request") or {}
        if rr:
            _bump(rr.get("requested_by_id") or d.get("assigned_to_user_id"),
                  rr.get("created_at") or rr.get("requested_at"))

    # Inquiry
    for iq in await db.inquiries.find(
        {"assigned_to_id": {"$in": ids}, "deleted_at": {"$exists": False}},
        {"_id": 0, "assigned_to_id": 1, "assigned_at": 1, "created_at": 1},
    ).to_list(length=5000):
        _bump(iq.get("assigned_to_id"), iq.get("assigned_at") or iq.get("created_at"))

    items = []
    for u in users:
        s = series[u["id"]]
        items.append({
            "user_id": u["id"], "name": u.get("name") or u.get("username"),
            "username": u.get("username"), "role": u.get("role"),
            "series": s, "total": sum(s), "peak": max(s) if s else 0,
        })
    items.sort(key=lambda x: x["total"], reverse=True)
    return {"weeks": labels, "items": items}


@router.get("/drawing-requests/revision-pending-count")
async def revision_pending_count(current: dict = Depends(get_current_user)):
    """Jumlah DR yang menunggu approval revisi (untuk badge Head Sales/Admin)."""
    n = await db.drawing_requests.count_documents({"status": "revision_requested", "deleted_at": {"$exists": False}})
    return {"count": n}



@router.get("/drawing-requests/{drf_id}")
async def get_drawing_request(drf_id: str, current: dict = Depends(get_current_user)):
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    return doc


@router.get("/drawing-requests/{drf_id}/workgroup-status")
async def get_workgroup_status(drf_id: str, current: dict = Depends(get_current_user)):
    """Status SO-level: jumlah drawing, draft_count, lock, & kelengkapan dokumen.
    Dipakai untuk menentukan lock BOM/Dokumen SO dan checklist submit final."""
    from utils.workgroup import workgroup_status
    return await workgroup_status(drf_id)


@router.put("/drawing-requests/{drf_id}")
async def update_drawing_request(
    drf_id: str,
    payload: DrawingRequestCreate,
    current: dict = Depends(get_current_user),
):
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["status"] not in ("draft",):
        raise HTTPException(status_code=400, detail="DRF sudah submitted, tidak bisa edit")
    if doc["created_by"] != current["id"] and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Bukan pemilik DRF")

    upd = payload.model_dump()
    upd["request_type"] = _require_request_type(payload.request_type)
    upd["items"] = _clean_drf_items(upd.get("items"))
    if upd["items"]:
        upd["qty_order"] = sum((float(it.get("qty") or 0) for it in upd["items"]))
    upd["updated_at"] = _now_iso()
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    out = await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})
    return out


class DeadlineUpdate(BaseModel):
    expected_due_date: Optional[str] = None  # Deadline Drawing
    delivery_due_date: Optional[str] = None  # Deadline Pengiriman


@router.patch("/drawing-requests/{drf_id}/deadlines")
async def update_drf_deadlines(
    drf_id: str,
    payload: DeadlineUpdate,
    current: dict = Depends(get_current_user),
):
    """Ubah Deadline Drawing / Pengiriman KAPAN SAJA (termasuk setelah DRF submitted).
    Hanya menyentuh 2 field deadline, tidak mengubah item/data lain."""
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    upd = {}
    if payload.expected_due_date is not None:
        upd["expected_due_date"] = payload.expected_due_date
    if payload.delivery_due_date is not None:
        upd["delivery_due_date"] = payload.delivery_due_date
    if not upd:
        raise HTTPException(status_code=400, detail="Tidak ada deadline yang diubah")
    upd["updated_at"] = _now_iso()
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    await log_action(current, "drf_update_deadlines", "drawing_requests", drf_id, upd)
    out = await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})
    return out
@router.post("/drawing-requests/{drf_id}/submit")
async def submit_drawing_request(drf_id: str, current: dict = Depends(get_current_user)):
    """Sales submit DRF → auto-TTD Requested By, status = submitted."""
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["status"] != "draft":
        raise HTTPException(status_code=400, detail=f"DRF sudah {doc['status']}")
    if doc["created_by"] != current["id"] and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Bukan pemilik DRF")

    upd = {
        "status": "submitted",
        "requested_by": _sig(current),
        "submitted_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    await log_action(current, "drf_submit", "drawing_requests", drf_id, {"form_no": doc.get("form_no")})
    out = await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})
    return out


@router.post("/drawing-requests/{drf_id}/accept")
async def accept_drawing_request(drf_id: str, current: dict = Depends(get_current_user)):
    """Eng Leader accept DRF → auto-TTD Received By, status = accepted."""
    if not (is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering Head yang boleh accept")
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["status"] != "submitted":
        raise HTTPException(status_code=400, detail=f"DRF status = {doc['status']}, harus 'submitted' dulu")

    upd = {
        "status": "accepted",
        "received_by": _sig(current),
        "accepted_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    await log_action(current, "drf_accept", "drawing_requests", drf_id, {"form_no": doc.get("form_no")})
    out = await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})
    return out


# =========================================================================
# Revisi DR Berjenjang (Sales minta revisi → Head Sales/Admin setujui)
# =========================================================================
class RevisionReqIn(BaseModel):
    reason: str = ""


@router.post("/drawing-requests/{drf_id}/request-revision")
async def request_drf_revision(drf_id: str, payload: RevisionReqIn, current: dict = Depends(get_current_user)):
    """Sales (pembuat) minta revisi DR yang SUDAH diajukan.
    Status → 'revision_requested' (menunggu approval Head Sales/Admin). DR tetap terkunci."""
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if not _is_sales(current) and doc.get("created_by") != current["id"]:
        raise HTTPException(status_code=403, detail="Hanya Sales/pembuat DR yang bisa minta revisi")
    if doc["status"] not in ("submitted", "accepted"):
        raise HTTPException(status_code=400, detail=f"Revisi hanya untuk DR yang sudah diajukan (status saat ini: {doc['status']})")
    if not (payload.reason or "").strip():
        raise HTTPException(status_code=400, detail="Alasan revisi wajib diisi")
    upd = {
        "status": "revision_requested",
        "prev_status": doc["status"],
        "revision_request": {
            "by": current.get("name") or current.get("username"),
            "by_id": current["id"],
            "at": _now_iso(),
            "reason": payload.reason.strip(),
        },
        "updated_at": _now_iso(),
    }
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    await log_action(current, "drf_request_revision", "drawing_requests", drf_id, {"form_no": doc.get("form_no"), "reason": payload.reason.strip()})
    return await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})


@router.post("/drawing-requests/{drf_id}/approve-revision")
async def approve_drf_revision(drf_id: str, current: dict = Depends(get_current_user)):
    """Head Sales (supervisor) / Admin menyetujui revisi → DR dibuka lagi (status 'draft').
    TTD pengajuan & penerimaan Engineering di-reset agar DR mengikuti alur ulang."""
    if not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Head Sales / Admin yang boleh menyetujui revisi")
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["status"] != "revision_requested":
        raise HTTPException(status_code=400, detail="DR tidak sedang menunggu approval revisi")
    upd = {
        "status": "draft",
        "requested_by": None,
        "submitted_at": None,
        "received_by": None,
        "accepted_at": None,
        "assigned_engineer_id": None,
        "assigned_engineer_name": None,
        "revision_round": int(doc.get("revision_round") or 0) + 1,
        "revision_approved": {
            "by": current.get("name") or current.get("username"),
            "at": _now_iso(),
        },
        "updated_at": _now_iso(),
    }
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd, "$unset": {"prev_status": ""}})
    await log_action(current, "drf_approve_revision", "drawing_requests", drf_id, {"form_no": doc.get("form_no")})
    return await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})


@router.post("/drawing-requests/{drf_id}/reject-revision")
async def reject_drf_revision(drf_id: str, payload: RevisionReqIn, current: dict = Depends(get_current_user)):
    """Head Sales / Admin menolak permintaan revisi → status kembali seperti semula."""
    if not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Head Sales / Admin yang boleh menolak revisi")
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["status"] != "revision_requested":
        raise HTTPException(status_code=400, detail="DR tidak sedang menunggu approval revisi")
    restore = doc.get("prev_status") or "submitted"
    upd = {
        "status": restore,
        "revision_rejected": {
            "by": current.get("name") or current.get("username"),
            "at": _now_iso(),
            "reason": (payload.reason or "").strip(),
        },
        "updated_at": _now_iso(),
    }
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd, "$unset": {"prev_status": ""}})
    await log_action(current, "drf_reject_revision", "drawing_requests", drf_id, {"form_no": doc.get("form_no"), "reason": (payload.reason or '').strip()})
    return await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})




class AcceptAssignIn(BaseModel):
    assigned_engineer_id: str
    assigned_engineer_name: Optional[str] = ""


@router.post("/drawing-requests/{drf_id}/accept-assign")
async def accept_and_assign_drf(
    drf_id: str,
    payload: AcceptAssignIn,
    current: dict = Depends(get_current_user),
):
    """Eng Leader accept DRF + langsung tunjuk 1 engineer yang mengerjakan.
    Eng Leader TIDAK mengisi kolom lain — hanya menugaskan siapa yang kerja."""
    if not (is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering Leader yang boleh accept & assign")
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["status"] not in ("submitted", "accepted"):
        raise HTTPException(status_code=400, detail=f"DRF status = {doc['status']}, harus 'submitted' dulu")
    eng = await db.users.find_one({"id": payload.assigned_engineer_id})
    if not eng:
        raise HTTPException(status_code=404, detail="Engineer tidak ditemukan")
    if not is_engineering(eng):
        raise HTTPException(status_code=400, detail="User yang dipilih bukan Engineering")

    upd = {
        "status": "accepted",
        "assigned_engineer_id": eng["id"],
        "assigned_engineer_name": payload.assigned_engineer_name or eng.get("name") or eng.get("username"),
        "assigned_by": current.get("name") or current.get("username"),
        "assigned_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    if not doc.get("received_by"):
        upd["received_by"] = _sig(current)
        upd["accepted_at"] = _now_iso()
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    await log_action(current, "drf_accept_assign", "drawing_requests", drf_id,
                     {"form_no": doc.get("form_no"), "engineer": upd["assigned_engineer_name"]})
    out = await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})
    return _clean(out)


class ReassignIn(BaseModel):
    assigned_engineer_id: str
    assigned_engineer_name: Optional[str] = ""
    reason: str = ""


@router.post("/drawing-requests/{drf_id}/reassign")
async def reassign_drf(
    drf_id: str,
    payload: ReassignIn,
    current: dict = Depends(get_current_user),
):
    """Pindah tugas DRF ke engineer lain (mis. engineer semula berhalangan).
    - Wewenang: Engineering Leader/Admin ATAU engineer yang sedang ditugaskan (handover sendiri).
    - Nama engineer di DRF diperbarui.
    - Nama pada DOKUMEN drawing yang BELUM ditandatangani (approval_status draft) ikut
      berubah otomatis: prepared_by + assigned_to → engineer baru.
    - Drawing yang SUDAH ada TTD tidak diubah 'Prepared By'-nya (arsip TTD dijaga),
      namun kepemilikan (assigned_to) tetap dipindah agar bisa dilanjutkan.
    - Riwayat pemindahan (reassign_history) disimpan lengkap dengan alasan.
    """
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")

    old_id = doc.get("assigned_engineer_id")
    is_assignee = old_id and current.get("id") == old_id
    if not (is_eng_head(current) or is_admin_like(current) or is_assignee):
        raise HTTPException(status_code=403, detail="Hanya Engineering Leader/Admin atau engineer yang ditugaskan yang boleh memindahkan tugas ini")
    if not old_id:
        raise HTTPException(status_code=400, detail="DRF belum pernah di-assign. Gunakan Accept & Assign dulu.")
    if doc.get("status") not in ("accepted", "received", "in_progress"):
        raise HTTPException(status_code=400, detail=f"DRF status = {doc.get('status')} — hanya bisa dipindah saat 'accepted' atau 'in_progress'")

    eng = await db.users.find_one({"id": payload.assigned_engineer_id})
    if not eng:
        raise HTTPException(status_code=404, detail="Engineer tujuan tidak ditemukan")
    if not is_engineering(eng):
        raise HTTPException(status_code=400, detail="User tujuan bukan Engineering")
    if eng["id"] == old_id:
        raise HTTPException(status_code=400, detail="Engineer tujuan sama dengan yang ditugaskan sekarang")

    new_name = (payload.assigned_engineer_name or eng.get("name") or eng.get("username") or "").strip()
    old_name = doc.get("assigned_engineer_name") or ""
    now = _now_iso()

    hist = list(doc.get("reassign_history") or [])
    hist.append({
        "from_id": old_id,
        "from_name": old_name,
        "to_id": eng["id"],
        "to_name": new_name,
        "by": current.get("name") or current.get("username"),
        "by_id": current.get("id"),
        "reason": (payload.reason or "").strip(),
        "at": now,
    })

    await db.drawing_requests.update_one(
        {"id": drf_id},
        {"$set": {
            "assigned_engineer_id": eng["id"],
            "assigned_engineer_name": new_name,
            "reassigned_at": now,
            "reassigned_by": current.get("name") or current.get("username"),
            "reassign_reason": (payload.reason or "").strip(),
            "reassign_history": hist,
            "updated_at": now,
        }},
    )

    # --- Propagasi ke drawing yang di-generate dari DRF ini ---
    linked_q = {"from_drf_id": drf_id, "deleted_at": {"$exists": False}}
    # 1) Semua drawing: pindahkan kepemilikan/edit-rights
    await db.drawings.update_many(
        linked_q,
        {"$set": {"assigned_to_user_id": eng["id"], "assigned_to_name": new_name, "updated_at": now}},
    )
    # 2) Drawing yang BELUM di-TTD (draft): nama 'Prepared By' pada dokumen ikut berubah
    unsigned_q = {**linked_q, "$or": [
        {"approval_status": {"$in": [None, "", "draft"]}},
        {"approval_status": {"$exists": False}},
    ]}
    name_updated = await db.drawings.count_documents(unsigned_q)
    if name_updated:
        await db.drawings.update_many(unsigned_q, {"$set": {"prepared_by": new_name, "updated_at": now}})
    total_linked = await db.drawings.count_documents(linked_q)
    kept_signed = max(0, total_linked - name_updated)

    await log_action(current, "drf_reassign", "drawing_requests", drf_id, {
        "form_no": doc.get("form_no"),
        "from": old_name, "to": new_name,
        "reason": (payload.reason or "").strip(),
        "drawings_name_updated": name_updated, "drawings_kept_signed": kept_signed,
    })

    out = await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})
    return {
        "success": True,
        "drf": _clean(out),
        "drawings_name_updated": name_updated,
        "drawings_kept_signed": kept_signed,
        "message": f"Tugas dipindah ke {new_name}. {name_updated} drawing diperbarui namanya" + (f", {kept_signed} drawing ber-TTD dipertahankan" if kept_signed else ""),
    }



@router.post("/drawing-requests/{drf_id}/accept-work")
async def accept_work_drf(drf_id: str, current: dict = Depends(get_current_user)):
    """Engineer yang ditugaskan KLIK TERIMA (mengakui pekerjaan).
    Status: accepted (Antri) → received (Diterima). Belum mulai menggambar.
    Set work_received_at + work_received_by."""
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    assignee = doc.get("assigned_engineer_id")
    is_assignee = assignee and current.get("id") == assignee
    if not (is_assignee or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya engineer yang ditugaskan yang bisa menerima job ini")
    if not assignee:
        raise HTTPException(status_code=400, detail="DRF belum di-assign ke engineer manapun")
    if doc.get("status") not in ("accepted", "received"):
        raise HTTPException(status_code=400, detail=f"DRF harus sudah di-assign (Antri) dulu (status: {doc.get('status')})")
    # Idempotent
    if doc.get("status") == "received" or doc.get("work_received_at"):
        return _clean(doc)
    now = _now_iso()
    upd = {
        "status": "received",
        "work_received_at": now,
        "work_received_by": current.get("name") or current.get("username"),
        "updated_at": now,
    }
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    await log_action(current, "drf_accept_work", "drawing_requests", drf_id, {"form_no": doc.get("form_no")})
    out = await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})
    return _clean(out)


@router.post("/drawing-requests/{drf_id}/start-work")
async def start_work_drf(drf_id: str, current: dict = Depends(get_current_user)):
    """Engineer KLIK MULAI KERJAKAN → mulai menggambar.
    Status: received/accepted → in_progress (Proses). Set work_started_at.
    Per SO: sekali mulai = start kerja untuk semua drawing di SO ini."""
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    assignee = doc.get("assigned_engineer_id")
    is_assignee = assignee and current.get("id") == assignee
    if not (is_assignee or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya engineer yang ditugaskan yang bisa mulai job ini")
    if not assignee:
        raise HTTPException(status_code=400, detail="DRF belum di-assign ke engineer manapun")
    if doc.get("status") not in ("accepted", "received", "in_progress"):
        raise HTTPException(status_code=400, detail=f"DRF harus sudah diterima dulu (status: {doc.get('status')})")
    # Idempotent: kalau sudah pernah start, jangan overwrite tanggalnya.
    if doc.get("work_started_at"):
        return _clean(doc)
    now = _now_iso()
    upd = {
        "status": "in_progress",
        "work_started_at": now,
        "work_started_by": current.get("name") or current.get("username"),
        "updated_at": now,
    }
    # Kalau engineer skip tombol Terima, catat penerimaan sekaligus saat mulai.
    if not doc.get("work_received_at"):
        upd["work_received_at"] = now
        upd["work_received_by"] = current.get("name") or current.get("username")
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    # Denormalisasi tanggal ke drawing yang sudah ada (jika sudah di-generate sebelumnya)
    await db.drawings.update_many(
        {"from_drf_id": drf_id, "deleted_at": {"$exists": False}},
        {"$set": {"work_started_at": now,
                  "request_received_at": doc.get("accepted_at"),
                  "updated_at": now}},
    )
    await log_action(current, "drf_start_work", "drawing_requests", drf_id, {"form_no": doc.get("form_no")})
    out = await db.drawing_requests.find_one({"id": drf_id}, {"_id": 0})
    return _clean(out)




class GenerateDrawingIn(BaseModel):
    project_initial: str = ""
    drawing_type: str = "Assembly"   # Assembly | Part
    title: str = ""
    discipline: str = "Mechanical"
    customer_drawing_no: str = ""    # Nomor DWG customer (opsional)


class GenerateDrawingsIn(BaseModel):
    drawings: List[GenerateDrawingIn]
    class_material: str = ""


@router.post("/drawing-requests/{drf_id}/generate-drawings")
async def generate_drawings_for_drf(
    drf_id: str,
    payload: GenerateDrawingsIn,
    current: dict = Depends(get_current_user),
):
    """Engineer yang ditugaskan generate 1+ nomor drawing untuk DRF ini.
    Semua drawing berbagi SATU BOM (1 BOM bisa untuk banyak drawing).
    New Order: user tentukan mau berapa drawing → generate nomor sebanyak itu."""
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    # Hanya engineer yang ditugaskan (atau Eng Leader/Admin) yang boleh
    assignee = doc.get("assigned_engineer_id")
    is_assignee = assignee and current.get("id") == assignee
    if not (is_assignee or is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya engineer yang ditugaskan yang boleh mengerjakan DRF ini")
    if doc["status"] not in ("accepted", "received", "in_progress"):
        raise HTTPException(status_code=400, detail=f"DRF harus di-accept & di-assign dulu (status sekarang: {doc['status']})")
    if not payload.drawings:
        raise HTTPException(status_code=400, detail="Minimal 1 drawing")

    # Lazy import untuk hindari circular import
    from routers.drawing_register import create_drawing, DrawingIn

    assigned_name = doc.get("assigned_engineer_name") or (current.get("name") or current.get("username"))
    customer_code = (doc.get("customer_code") or "MKS").upper().strip() or "MKS"
    created = []
    shared_bom_id = ""
    for idx, d in enumerate(payload.drawings):
        if idx == 0:
            bom_mode, bom_id = "create_new", ""
        else:
            bom_mode, bom_id = ("existing", shared_bom_id) if shared_bom_id else ("none", "")
        din = DrawingIn(
            customer_code=customer_code,
            customer_name=doc.get("customer_name") or "",
            po_customer_no=doc.get("po_customer_no") or "",
            project_initial=(d.project_initial or "").strip(),
            drawing_type=d.drawing_type or "Assembly",
            title=d.title or "",
            discipline=d.discipline or "Mechanical",
            customer_drawing_no=(d.customer_drawing_no or "").strip(),
            so_no=doc.get("so_no") or "",
            project_name=doc.get("project_name") or "",
            class_material=payload.class_material or "",
            prepared_by=assigned_name,
            from_drf_id=drf_id,
            assigned_to_user_id=assignee or current.get("id"),
            assigned_to_name=assigned_name,
            bom_link_mode=bom_mode,
            bom_id=bom_id,
        )
        dr = await create_drawing(din, current)
        if idx == 0:
            shared_bom_id = dr.get("bom_id") or ""
        created.append(dr)

    linked_ids = [d["id"] for d in created]
    await db.drawing_requests.update_one(
        {"id": drf_id},
        {"$set": {
            "status": "in_progress",
            "linked_drawing_id": linked_ids[0] if linked_ids else None,
            "linked_drawing_ids": (doc.get("linked_drawing_ids") or []) + linked_ids,
            "shared_bom_id": shared_bom_id,
            "updated_at": _now_iso(),
        }},
    )
    await log_action(current, "drf_generate_drawings", "drawing_requests", drf_id,
                     {"form_no": doc.get("form_no"), "count": len(created)})
    return {"success": True, "drawings": created, "shared_bom_id": shared_bom_id}


class PullRepeatIn(BaseModel):
    source_drawing_ids: List[str]
    class_material: str = ""


@router.post("/drawing-requests/{drf_id}/pull-repeat")
async def pull_repeat_drawings(
    drf_id: str,
    payload: PullRepeatIn,
    current: dict = Depends(get_current_user),
):
    """Repeat Order — tarik-otomatis data drawing lama menjadi drawing baru di DRF ini.
    Untuk tiap source drawing yang dipilih:
      - Buat drawing baru (nomor DWG baru) yang meng-clone metadata drawing lama.
      - Drawing pertama membuat BOM baru dgn meng-clone item + attachment (nesting/costing→costing_prev)
        dari BOM lama (source_bom_id). Drawing berikutnya berbagi BOM yang sama.
      - File level-drawing (MKS drawing, Customer drawing, additional files/nesting) di-clone
        sebagai reference-copy (share GridFS id) → auto attached, editable/replace via Work Order.
    BOM autofilled & editable bila ada perubahan Qty.
    """
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    assignee = doc.get("assigned_engineer_id")
    is_assignee = assignee and current.get("id") == assignee
    if not (is_assignee or is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya engineer yang ditugaskan yang boleh mengerjakan DRF ini")
    if doc["status"] not in ("accepted", "received", "in_progress"):
        raise HTTPException(status_code=400, detail=f"DRF harus di-accept & di-assign dulu (status sekarang: {doc['status']})")
    if not payload.source_drawing_ids:
        raise HTTPException(status_code=400, detail="Pilih minimal 1 drawing lama untuk ditarik")

    from routers.drawing_register import create_drawing, DrawingIn

    assigned_name = doc.get("assigned_engineer_name") or (current.get("name") or current.get("username"))
    customer_code = (doc.get("customer_code") or "MKS").upper().strip() or "MKS"
    created = []
    shared_bom_id = doc.get("shared_bom_id") or ""

    for source_id in payload.source_drawing_ids:
        src = await db.drawings.find_one({"id": source_id, "deleted_at": {"$exists": False}})
        if not src:
            raise HTTPException(status_code=404, detail=f"Drawing sumber {source_id} tidak ditemukan")

        if not shared_bom_id:
            bom_mode = "create_new"
            bom_id = ""
            source_bom_id = src.get("bom_id") or ""
        else:
            bom_mode = "existing"
            bom_id = shared_bom_id
            source_bom_id = ""

        din = DrawingIn(
            customer_code=(src.get("customer_code") or customer_code).upper().strip(),
            customer_name=doc.get("customer_name") or src.get("customer_name") or "",
            project_initial=(src.get("project_initial") or "").strip(),
            drawing_type=src.get("drawing_type") or "Assembly",
            title=src.get("title") or "",
            discipline=src.get("discipline") or "Mechanical",
            customer_drawing_no=(src.get("customer_drawing_no") or "").strip(),
            so_no=doc.get("so_no") or "",
            project_name=doc.get("project_name") or src.get("project_name") or "",
            class_material=payload.class_material or src.get("class_material") or "",
            prepared_by=assigned_name,
            from_drf_id=drf_id,
            assigned_to_user_id=assignee or current.get("id"),
            assigned_to_name=assigned_name,
            bom_link_mode=bom_mode,
            bom_id=bom_id,
            source_bom_id=source_bom_id,
        )
        dr = await create_drawing(din, current)
        if not shared_bom_id:
            shared_bom_id = dr.get("bom_id") or ""

        # Clone file level-drawing dari drawing lama (reference-copy, share GridFS id).
        clone_set = {
            "is_repeat_pulled": True,
            "pulled_from_drawing_id": src.get("id"),
            "pulled_from_drawing_no": src.get("drawing_no"),
            "updated_at": _now_iso(),
        }
        if src.get("file_id"):
            clone_set["file_id"] = src.get("file_id")
            clone_set["filename"] = src.get("filename")
            clone_set["file_uploaded_at"] = _now_iso()
        if src.get("customer_ref_file_id"):
            clone_set["customer_ref_file_id"] = src.get("customer_ref_file_id")
            clone_set["customer_ref_filename"] = src.get("customer_ref_filename")
            clone_set["customer_ref_uploaded_at"] = _now_iso()
        src_extras = src.get("additional_files") or []
        if src_extras:
            cloned_extras = []
            for e in src_extras:
                ne = {**e}
                ne["id"] = str(uuid.uuid4())
                ne["copied_from_drawing"] = src.get("id")
                cloned_extras.append(ne)
            clone_set["additional_files"] = cloned_extras
        await db.drawings.update_one({"id": dr["id"]}, {"$set": clone_set})
        dr.update(clone_set)
        created.append(dr)

    linked_ids = [d["id"] for d in created]
    await db.drawing_requests.update_one(
        {"id": drf_id},
        {"$set": {
            "status": "in_progress",
            "linked_drawing_id": (doc.get("linked_drawing_id") or (linked_ids[0] if linked_ids else None)),
            "linked_drawing_ids": (doc.get("linked_drawing_ids") or []) + linked_ids,
            "shared_bom_id": shared_bom_id,
            "updated_at": _now_iso(),
        }},
    )
    await log_action(current, "drf_pull_repeat", "drawing_requests", drf_id,
                     {"form_no": doc.get("form_no"), "count": len(created)})
    return {"success": True, "drawings": created, "shared_bom_id": shared_bom_id}


@router.post("/drawing-requests/{drf_id}/link-drawing")
async def link_drawing_to_drf(
    drf_id: str,
    payload: dict,
    current: dict = Depends(get_current_user),
):
    """Setelah Eng buat drawing, link drawing_id ke DRF ini."""
    drawing_id = payload.get("drawing_id")
    if not drawing_id:
        raise HTTPException(status_code=400, detail="drawing_id wajib")
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    upd = {
        "linked_drawing_id": drawing_id,
        "status": "in_progress",
        "updated_at": _now_iso(),
    }
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": upd})
    return {"success": True}


@router.post("/drawing-requests/{drf_id}/cancel")
async def cancel_drf(drf_id: str, current: dict = Depends(get_current_user)):
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["created_by"] != current["id"] and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Bukan pemilik")
    if doc["status"] in ("completed",):
        raise HTTPException(status_code=400, detail="DRF sudah selesai, tidak bisa cancel")
    await db.drawing_requests.update_one({"id": drf_id}, {"$set": {"status": "cancelled", "updated_at": _now_iso()}})
    return {"success": True}


# =========================================================================
# Attachments (multi-file upload untuk lampiran DRF)
# =========================================================================
@router.post("/drawing-requests/{drf_id}/attachments")
async def upload_drf_attachment(
    drf_id: str,
    file: UploadFile = File(...),
    category: str = Form("other"),
    current: dict = Depends(get_current_user),
):
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["created_by"] != current["id"] and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Bukan pemilik")
    cat = (category or "other").strip().lower()
    if cat not in ("po_customer", "other"):
        cat = "other"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File > 20 MB")

    fid = await _bucket().upload_from_stream(
        file.filename or "attachment",
        content,
        metadata={"drf_id": drf_id, "content_type": file.content_type, "user_id": current["id"]},
    )
    entry = {
        "file_id": str(fid),
        "filename": file.filename,
        "content_type": file.content_type,
        "category": cat,
        "size": len(content),
        "uploaded_at": _now_iso(),
        "uploaded_by": current.get("name") or current.get("username"),
    }
    await db.drawing_requests.update_one(
        {"id": drf_id},
        {"$push": {"attached_files": entry}, "$set": {"updated_at": _now_iso()}},
    )
    return entry


@router.get("/drawing-requests/{drf_id}/attachments/{file_id}/download")
async def download_drf_attachment(drf_id: str, file_id: str, current: dict = Depends(get_current_user)):
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    entry = next((f for f in (doc.get("attached_files") or []) if f["file_id"] == file_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    try:
        stream = await _bucket().open_download_stream(ObjectId(file_id))
        raw = await stream.read()
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File tidak bisa dibaca: {e}")
    return StreamingResponse(
        io.BytesIO(raw),
        media_type=entry.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{entry.get("filename")}"'},
    )


@router.delete("/drawing-requests/{drf_id}/attachments/{file_id}")
async def delete_drf_attachment(drf_id: str, file_id: str, current: dict = Depends(get_current_user)):
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    if doc["created_by"] != current["id"] and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Bukan pemilik")
    if doc.get("status") not in ("draft",):
        raise HTTPException(status_code=400, detail="Sudah submitted, tidak bisa hapus file")
    try:
        await _bucket().delete(ObjectId(file_id))
    except Exception:
        pass
    await db.drawing_requests.update_one(
        {"id": drf_id},
        {"$pull": {"attached_files": {"file_id": file_id}}},
    )
    return {"success": True}


# =========================================================================
# Attachment preview (image-based, view-only) — dipakai popup Antrian DRF
# =========================================================================
def _entry_is_pdf(entry: dict) -> bool:
    ct = (entry.get("content_type") or "").lower()
    fn = (entry.get("filename") or "").lower()
    return "pdf" in ct or fn.endswith(".pdf")


async def _attachment_raw(drf_id: str, file_id: str) -> tuple:
    doc = await db.drawing_requests.find_one({"id": drf_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="DRF tidak ditemukan")
    entry = next((f for f in (doc.get("attached_files") or []) if f["file_id"] == file_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    try:
        stream = await _bucket().open_download_stream(ObjectId(file_id))
        raw = await stream.read()
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File tidak bisa dibaca: {e}")
    return raw, entry


@router.get("/drawing-requests/{drf_id}/attachments/{file_id}/page-meta")
async def drf_attachment_page_meta(drf_id: str, file_id: str, current: dict = Depends(get_current_user)):
    """Metadata halaman untuk preview image-based lampiran DRF (PDF)."""
    raw, entry = await _attachment_raw(drf_id, file_id)
    if not _entry_is_pdf(entry):
        raise HTTPException(status_code=400, detail="Preview gambar hanya untuk PDF")
    from utils.pdf_render import pdf_page_meta
    return pdf_page_meta(raw)


@router.get("/drawing-requests/{drf_id}/attachments/{file_id}/page-image")
async def drf_attachment_page_image(
    drf_id: str, file_id: str, page: int = 0, scale: float = 2.0,
    current: dict = Depends(get_current_user),
):
    """Render 1 halaman lampiran DRF (PDF) sebagai PNG untuk viewer view-only."""
    raw, entry = await _attachment_raw(drf_id, file_id)
    if not _entry_is_pdf(entry):
        raise HTTPException(status_code=400, detail="Preview gambar hanya untuk PDF")
    from utils.pdf_render import pdf_page_png
    from fastapi.responses import Response
    try:
        png = pdf_page_png(raw, page=page, scale=scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ada")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"File bukan PDF valid: {e}")
    return Response(content=png, media_type="image/png", headers={"Cache-Control": "no-store"})
