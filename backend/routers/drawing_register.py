"""MKS-F-ENG-005 Drawing Master List — register nomor drawing + upload file + verifikasi konten PDF.

Engineering register drawing (drawing_no + metadata) → upload PDF. Sistem extract text
dari PDF dan cek apakah drawing_no register muncul di isi PDF. Jika tidak → warning kuning.
"""
from __future__ import annotations
import io
import os
import re
import uuid
from datetime import datetime, timezone
from typing import Optional, List

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Request
from fastapi.responses import StreamingResponse, Response
from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from pydantic import BaseModel

from db import db
from deps import get_current_user, log_action

router = APIRouter(tags=["drawings"])

VALID_DISCIPLINES = ["Mechanical", "Civil", "Electrical", "Piping", "Structural", "Instrument", "General"]
VALID_STATUS = ["Draft", "Issued", "Superseded", "Cancelled"]

# Default drawing number config
DEFAULT_CONFIG = {
    "default_customer_code": "MKS",   # customer code default
    "assembly_start_seq": 0,           # A.00
    "part_start_seq": 1,               # P.01
}


async def _get_config() -> dict:
    doc = await db.drawing_config.find_one({"_id": "default"}, {"_id": 0})
    if not doc:
        return dict(DEFAULT_CONFIG)
    return {**DEFAULT_CONFIG, **doc}


async def _next_drawing_no(customer_code: str, project_initial: str, drawing_type: str) -> dict:
    """Generate next drawing number in format:
       DWG.YY.MM.NN_CUSTOMER.INITIAL.TYPE.SEQ

    NN = monthly running number, unique per project within a month
    (same customer+initial+YYYYMM = same NN, only TYPE.SEQ differs)
    TYPE = "A" (Assembly) or "P" (Part)
    Assembly starts from 00, Part starts from 01. Both increment.
    """
    cfg = await _get_config()
    customer_code = (customer_code or cfg.get("default_customer_code") or "MKS").upper().strip()
    project_initial = (project_initial or "").upper().strip()
    if not project_initial:
        raise HTTPException(status_code=400, detail="project_initial wajib untuk auto-generate (mis. 'SP' untuk Support Plate)")
    if drawing_type not in ("Assembly", "Part"):
        raise HTTPException(status_code=400, detail="drawing_type harus 'Assembly' atau 'Part'")

    now = datetime.now(timezone.utc)
    yy = f"{now.year % 100:02d}"
    mm = f"{now.month:02d}"
    year_month = f"{yy}.{mm}"

    # 1) Determine monthly_running (NN):
    # Cek apakah project (customer + initial) sudah ada drawing di bulan ini
    existing_project_drawing = await db.drawings.find_one({
        "year_month": year_month,
        "customer_code": customer_code,
        "project_initial": project_initial,
        "deleted_at": {"$exists": False},
    })
    if existing_project_drawing and existing_project_drawing.get("monthly_running") is not None:
        monthly_running = int(existing_project_drawing["monthly_running"])
    else:
        # New project in this month → next monthly counter
        counter_key = f"drawing_monthly_{yy}_{mm}"
        res = await db.counters.find_one_and_update(
            {"_id": counter_key},
            {"$inc": {"value": 1}},
            upsert=True,
            return_document=True,
        )
        monthly_running = (res or {}).get("value") or 1

    # 2) Determine type_seq
    type_letter = "A" if drawing_type == "Assembly" else "P"
    start_seq = int(cfg.get("assembly_start_seq" if drawing_type == "Assembly" else "part_start_seq") or 0)
    # Find max existing seq for this project+type
    existing_type_drawings = await db.drawings.find({
        "year_month": year_month,
        "customer_code": customer_code,
        "project_initial": project_initial,
        "drawing_type": drawing_type,
        "deleted_at": {"$exists": False},
    }, {"type_seq": 1, "_id": 0}).to_list(length=1000)
    max_seq = max([int(d.get("type_seq", -1)) for d in existing_type_drawings], default=None)
    if max_seq is None:
        type_seq = start_seq
    else:
        type_seq = max_seq + 1

    dno = f"DWG.{yy}.{mm}.{monthly_running:02d}_{customer_code}.{project_initial}.{type_letter}.{type_seq:02d}"
    return {
        "drawing_no": dno,
        "year_month": year_month,
        "customer_code": customer_code,
        "project_initial": project_initial,
        "monthly_running": monthly_running,
        "drawing_type": drawing_type,
        "type_letter": type_letter,
        "type_seq": type_seq,
    }

_gridfs: Optional[AsyncIOMotorGridFSBucket] = None


def _fs() -> AsyncIOMotorGridFSBucket:
    global _gridfs
    if _gridfs is None:
        _gridfs = AsyncIOMotorGridFSBucket(db, bucket_name="drawings")
    return _gridfs


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _clean(d):
    if d:
        d.pop("_id", None)
    return d


def _can_edit(user: dict) -> bool:
    role = (user or {}).get("role", "")
    return role in ("admin", "super_admin", "supervisor", "eng_leader", "eng_head", "eng_staff", "engineering")


def _is_eng_head_role(user: dict) -> bool:
    """Eng Head/Leader/Admin bisa assign ke user lain & override permission."""
    role = (user or {}).get("role", "")
    return role in ("admin", "super_admin", "supervisor", "eng_leader", "eng_head")


def _can_modify_drawing(user: dict, drawing: dict) -> bool:
    """Iter 20 — Permission check untuk edit/upload/submit drawing.

    Rules:
    - Kalau drawing belum di-assign ke siapapun → semua eng_role bisa edit (compat lama)
    - Kalau sudah di-assign → hanya assigned user + Eng Head + Admin yang bisa edit
    - Kalau bukan eng_role → deny
    """
    if not _can_edit(user):
        return False
    assigned_id = (drawing or {}).get("assigned_to_user_id")
    if not assigned_id:
        return True  # belum di-assign, siapa saja di eng team boleh
    if _is_eng_head_role(user):
        return True  # Eng Head/Admin selalu boleh
    return user.get("id") == assigned_id


def _can_view(user: dict) -> bool:
    role = (user or {}).get("role", "")
    return role in (
        "admin", "super_admin", "supervisor", "finance", "sales", "purchasing",
        "eng_leader", "eng_head", "eng_staff", "engineering", "production", "produksi", "qc",
        "doc_control", "document_control", "store",
    )


class DrawingIn(BaseModel):
    drawing_no: str = ""  # kosong → auto-generate berdasarkan customer_code + project_initial + drawing_type
    customer_code: str = "MKS"       # default MKS (bisa custom untuk external customer)
    customer_name: str = ""          # opsional — nama customer lengkap (auto-fill dari customer master)
    project_initial: str = ""         # WAJIB kalau auto-generate. mis. "SP" untuk Support Plate
    drawing_type: str = "Assembly"    # Assembly | Part
    title: str = ""
    revision: str = "Rev-0"
    discipline: str = "Mechanical"
    customer_drawing_no: str = ""     # Nomor DWG dari customer (opsional) — dikaitkan dengan DWG MKS
    so_no: str = ""
    project_name: str = ""
    class_material: str = ""          # deskripsi paket order — mis. "RAW MATERIAL FOR QTY 1 PCS"
    prepared_by: str = ""
    request_by_sales: str = ""        # Sales yang request drawing (dropdown nama sales)
    checked_by: str = ""
    drawing_date: str = ""
    status: str = "Draft"
    remark: str = ""
    # BOM linking
    bom_link_mode: str = "none"   # none | create_new | existing
    bom_no: str = ""              # kosong dgn mode=create_new → auto-generate
    bom_id: str = ""              # untuk mode=existing
    source_bom_id: str = ""       # untuk Repeat Order: copy items + attachments dari BOM ini ke BOM baru
    # Iter 19 — Link ke Drawing Request Form (dari Sales)
    from_drf_id: str = ""         # kalau drawing ini dibuat sebagai lanjutan dari DRF, isi ID DRF-nya
    # Iter 20 — Assign engineer yang mengerjakan drawing
    assigned_to_user_id: str = ""
    assigned_to_name: str = ""
    po_customer_no: str = ""      # No. PO Customer (dari DRF) — auto-isi stamping SO
    item_name: str = ""           # Nama item (dari daftar item DRF) untuk drawing ini
    item_qty: float = 0           # Qty item untuk drawing ini (bisa beda/partial per drawing)


class ConfigIn(BaseModel):
    default_customer_code: Optional[str] = None
    assembly_start_seq: Optional[int] = None
    part_start_seq: Optional[int] = None


def _normalize_dno(s: str) -> str:
    """Uppercase + strip non-alphanumeric — for lookup / fuzzy compare in PDF text."""
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


def _ocr_pdf_text(pdf_bytes: bytes, max_pages: int = 3, dpi: int = 220) -> str:
    """OCR fallback untuk PDF hasil scan/gambar (tidak ada teks embedded).
    Render halaman via PyMuPDF → OCR via Tesseract. Butuh binary 'tesseract' terpasang
    (di Windows: install Tesseract-OCR & pastikan ada di PATH). Degrade aman bila tak ada."""
    try:
        import fitz  # PyMuPDF
        import pytesseract
        from PIL import Image
        import io as _io
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        chunks = []
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            try:
                pix = page.get_pixmap(dpi=dpi)
                img = Image.open(_io.BytesIO(pix.tobytes("png")))
                chunks.append(pytesseract.image_to_string(img) or "")
            except Exception:
                pass
        doc.close()
        return "\n".join(chunks)
    except Exception:
        return ""


def _extract_pdf_text_with_source(pdf_bytes: bytes, ocr_fallback: bool = True):
    """Return (text, source). source ∈ {'native','ocr','none'}.
    native = teks embedded (fitz/pypdf, akurat); ocr = hasil Tesseract (bisa kurang akurat)."""
    # Primary: PyMuPDF (fitz)
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        chunks = []
        for i, page in enumerate(doc):
            if i >= 5:  # judul/nomor drawing biasanya di halaman awal
                break
            try:
                chunks.append(page.get_text() or "")
            except Exception:
                pass
        doc.close()
        txt = "\n".join(chunks)
        if txt.strip():
            return txt, "native"
    except Exception:
        pass
    # Fallback: pypdf (opsional)
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        chunks = []
        for i, page in enumerate(reader.pages):
            if i >= 5:
                break
            try:
                chunks.append(page.extract_text() or "")
            except Exception:
                pass
        txt = "\n".join(chunks)
        if txt.strip():
            return txt, "native"
    except Exception:
        pass
    # Fallback terakhir: OCR untuk PDF scan/gambar
    if ocr_fallback:
        octxt = _ocr_pdf_text(pdf_bytes)
        if octxt.strip():
            return octxt, "ocr"
    return "", "none"


def _extract_pdf_text(pdf_bytes: bytes, ocr_fallback: bool = True) -> str:
    """Best-effort text extraction (native → pypdf → OCR). Return '' bila gagal."""
    txt, _src = _extract_pdf_text_with_source(pdf_bytes, ocr_fallback=ocr_fallback)
    return txt


_MKS_DNO_RE = re.compile(
    r"DWG\.\d{2}\.\d{2}\.\d{2}_[A-Za-z0-9]+(?:\.[A-Za-z0-9]+){2,4}",
    re.IGNORECASE,
)
# Loose pattern untuk hasil OCR (tanda baca sering meleset). Hasilnya = SARAN (bisa kurang akurat).
_MKS_DNO_LOOSE_RE = re.compile(
    r"DWG[.\s_]{0,2}\d{2}[.\s_]{0,2}\d{2}[.\s_]{0,2}\d{2}[.\s_]{0,2}[A-Za-z0-9]+(?:[.\s_]{0,2}[A-Za-z0-9]+){2,4}",
    re.IGNORECASE,
)


def _detect_mks_dno(text: str, loose: bool = False) -> str:
    """Deteksi nomor DWG format MKS (mis. 'DWG.26.07.03_THIES.FL.A.03') di dalam teks PDF.
    loose=True → toleran hasil OCR (nomor hasilnya bisa kurang akurat, dipakai sbg saran)."""
    if not text:
        return ""
    cleaned = re.sub(r"\s*([._])\s*", r"\1", text)
    m = _MKS_DNO_RE.search(cleaned)
    if m:
        return m.group(0).upper()
    if loose:
        m2 = _MKS_DNO_LOOSE_RE.search(text)
        if m2:
            return re.sub(r"\s+", "", m2.group(0)).upper()
    return ""


def _check_drawing_no_in_text(drawing_no: str, pdf_text: str, source: str = "native") -> dict:
    """Return {match, extracted_candidates, note, detected_no, detected_source}.

    detected_no = nomor format MKS (DWG.YY.MM.NN_CUST.INIT.TYPE.NN) yang terdeteksi di isi PDF.
    source='ocr' → pakai deteksi loose (hasil bisa kurang akurat → jadi SARAN, bukan auto-apply)."""
    if not drawing_no:
        return {"match": False, "extracted_candidates": [], "note": "drawing_no register kosong", "detected_no": "", "detected_source": source}
    if not pdf_text.strip():
        return {"match": False, "extracted_candidates": [], "note": "Tidak bisa extract teks dari PDF (mungkin scan/gambar)", "detected_no": "", "detected_source": "none"}

    detected_no = _detect_mks_dno(pdf_text, loose=(source == "ocr"))
    target = _normalize_dno(drawing_no)
    haystack = _normalize_dno(pdf_text)
    if target in haystack:
        return {"match": True, "extracted_candidates": [drawing_no], "note": "Nomor drawing ditemukan di PDF", "detected_no": detected_no or drawing_no, "detected_source": source}

    # Try to detect any drawing-number-like patterns in text as suggestions
    candidates = set()
    # Pattern: alphanumeric with dashes/slashes, min 5 chars, at least one letter and one digit
    for m in re.findall(r"[A-Za-z0-9][A-Za-z0-9\-/_.]{4,30}[A-Za-z0-9]", pdf_text):
        norm = _normalize_dno(m)
        if 5 <= len(norm) <= 30 and any(c.isalpha() for c in norm) and any(c.isdigit() for c in norm):
            candidates.add(m.strip())
    # Keep at most 8 unique candidates
    candidates_list = sorted(candidates)[:8]
    # OCR fallback: bila loose gagal, ambil kandidat yang diawali 'DWG' sebagai saran nomor.
    if source == "ocr" and not detected_no:
        for c in candidates_list:
            if c.upper().replace(" ", "").startswith("DWG"):
                detected_no = c.upper().replace(" ", "")
                break
    note = f"Nomor register '{drawing_no}' tidak ditemukan di isi PDF"
    if source == "ocr":
        note += " (dibaca via OCR — mohon verifikasi)"
    return {
        "match": False,
        "extracted_candidates": candidates_list,
        "note": note,
        "detected_no": detected_no,
        "detected_source": source,
    }


# ---- Deteksi "No. Drawing Customer" dari PDF referensi customer ----
# Customer PDF tidak punya format baku, jadi kita cari label umum lalu ambil kode di sebelahnya.
_CUST_DNO_LABELS = [
    r"DRAWING\s*(?:NO|NUMBER|NUM)\.?",
    r"DWG\s*(?:NO|NUMBER)\.?",
    r"DRG\s*(?:NO|NUMBER)\.?",
    r"DOC(?:UMENT)?\s*(?:NO|NUMBER)\.?",
    r"PART\s*(?:NO|NUMBER)\.?",
    r"P\s*/\s*N\.?",
    r"NO\.?\s*GAMBAR",
    r"NOMOR\s*GAMBAR",
]
# token kode: huruf/angka dgn - / _ . , minimal 4 char, wajib ada minimal 1 digit.
_CUST_CODE_TOKEN = r"([A-Za-z0-9][A-Za-z0-9\-/_.]{3,40}[A-Za-z0-9])"
_CUST_LABEL_RE = re.compile(
    r"(?:%s)[\s:._-]*%s" % ("|".join(_CUST_DNO_LABELS), _CUST_CODE_TOKEN),
    re.IGNORECASE,
)


def _detect_customer_dwg_no(text: str):
    """Return (detected_no, candidates[]) dari teks PDF customer.
    detected_no = tebakan terbaik (di sebelah label 'Drawing No' dst) atau '' bila tak ada."""
    if not text or not text.strip():
        return "", []
    detected = ""
    for m in _CUST_LABEL_RE.finditer(text):
        cand = (m.group(1) or "").strip(" .:-_")
        # buang token yang cuma angka murni pendek atau kata umum
        if cand and any(c.isdigit() for c in cand) and len(cand) >= 4:
            detected = cand.upper()
            break
    # Kandidat umum (untuk dropdown/manual autofill) — kode alfanumerik yang "mirip nomor".
    candidates = []
    seen = set()
    for m in re.findall(r"[A-Za-z0-9][A-Za-z0-9\-/_.]{4,30}[A-Za-z0-9]", text):
        c = m.strip()
        norm = c.upper()
        if norm in seen:
            continue
        if any(ch.isalpha() for ch in c) and any(ch.isdigit() for ch in c) and 5 <= len(c) <= 30:
            seen.add(norm)
            candidates.append(c)
        if len(candidates) >= 10:
            break
    return detected, candidates



@router.get("/drawings/config")
async def get_drawing_config(current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    cfg = await _get_config()
    now = datetime.now(timezone.utc)
    return {
        **cfg,
        "format_template": "DWG.YY.MM.NN_CUSTOMER.INITIAL.TYPE.SEQ",
        "format_example": f"DWG.{now.year % 100:02d}.{now.month:02d}.01_MKS.SP.A.00",
    }


@router.put("/drawings/config")
async def update_drawing_config(payload: ConfigIn, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not data:
        raise HTTPException(status_code=400, detail="Tidak ada field valid")
    await db.drawing_config.update_one({"_id": "default"}, {"$set": data}, upsert=True)
    return {"success": True, **data}


@router.get("/drawings/next-number")
async def preview_next_number(
    customer_code: str = "MKS",
    project_initial: str = "",
    drawing_type: str = "Assembly",
    current: dict = Depends(get_current_user),
):
    """Preview drawing_no yang akan digenerate — TIDAK mengubah counter.
    Perlu customer_code + project_initial + drawing_type.
    """
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    if not project_initial:
        # Show placeholder
        now = datetime.now(timezone.utc)
        yy = f"{now.year % 100:02d}"
        mm = f"{now.month:02d}"
        return {"preview": f"DWG.{yy}.{mm}.NN_{customer_code}.[INITIAL].A.00", "note": "Isi Project Initial dulu"}

    cfg = await _get_config()
    customer_code = (customer_code or cfg.get("default_customer_code") or "MKS").upper().strip()
    project_initial = project_initial.upper().strip()
    if drawing_type not in ("Assembly", "Part"):
        drawing_type = "Assembly"

    now = datetime.now(timezone.utc)
    yy = f"{now.year % 100:02d}"
    mm = f"{now.month:02d}"
    year_month = f"{yy}.{mm}"

    # Check existing project
    existing_project = await db.drawings.find_one({
        "year_month": year_month,
        "customer_code": customer_code,
        "project_initial": project_initial,
        "deleted_at": {"$exists": False},
    })
    if existing_project and existing_project.get("monthly_running") is not None:
        monthly_running = int(existing_project["monthly_running"])
        is_new_project = False
    else:
        # Peek next counter (tidak increment)
        counter_key = f"drawing_monthly_{yy}_{mm}"
        counter = await db.counters.find_one({"_id": counter_key}) or {}
        monthly_running = (counter.get("value") or 0) + 1
        is_new_project = True

    type_letter = "A" if drawing_type == "Assembly" else "P"
    start_seq = int(cfg.get("assembly_start_seq" if drawing_type == "Assembly" else "part_start_seq") or 0)
    existing_type = await db.drawings.find({
        "year_month": year_month,
        "customer_code": customer_code,
        "project_initial": project_initial,
        "drawing_type": drawing_type,
        "deleted_at": {"$exists": False},
    }, {"type_seq": 1, "_id": 0}).to_list(length=1000)
    max_seq = max([int(d.get("type_seq", -1)) for d in existing_type], default=None)
    type_seq = start_seq if max_seq is None else max_seq + 1

    preview = f"DWG.{yy}.{mm}.{monthly_running:02d}_{customer_code}.{project_initial}.{type_letter}.{type_seq:02d}"
    return {
        "preview": preview,
        "year_month": year_month,
        "monthly_running": monthly_running,
        "is_new_project": is_new_project,
        "existing_project_drawings": len(existing_type),
        "type_letter": type_letter,
        "type_seq": type_seq,
    }


@router.get("/drawings")
async def list_drawings(
    q: Optional[str] = None,
    discipline: Optional[str] = None,
    status: Optional[str] = None,
    so_no: Optional[str] = None,
    from_drf_id: Optional[str] = None,
    limit: int = 500,
    current: dict = Depends(get_current_user),
):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    filt = {"deleted_at": {"$exists": False}}
    if discipline: filt["discipline"] = discipline
    if status: filt["status"] = status
    if so_no: filt["so_no"] = so_no
    if from_drf_id: filt["from_drf_id"] = from_drf_id
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [
            {"drawing_no": rx}, {"title": rx}, {"project_name": rx},
            {"so_no": rx}, {"prepared_by": rx}, {"remark": rx},
            {"customer_drawing_no": rx}, {"customer_name": rx},
        ]
    docs = await db.drawings.find(filt, {"_id": 0}).sort("updated_at", -1).limit(limit).to_list(length=limit)
    return {"items": docs, "total": len(docs), "disciplines": VALID_DISCIPLINES, "statuses": VALID_STATUS}


_EXPORT_TTD_STAGE = {"submit": "Prepared", "eng_head": "Eng Head", "qc": "QC", "sales": "Sales"}


def _export_rows(docs: list) -> list:
    """Ubah dokumen drawing → baris siap-ekspor (ringkasan status + TTD + DC stamp)."""
    rows = []
    for d in docs:
        approvals = [a for a in (d.get("approvals") or []) if not str(a.get("stage") or "").startswith("reject_")]
        ttd_map = {}
        for a in approvals:
            st = a.get("stage")
            if st in _EXPORT_TTD_STAGE and st not in ttd_map:
                ttd_map[st] = a.get("name") or ""
        ttd_str = "; ".join(f"{_EXPORT_TTD_STAGE[s]}: {n}" for s, n in ttd_map.items()) or "-"
        dc = d.get("dc_stamp") or {}
        rows.append({
            "drawing_no": d.get("drawing_no") or "-",
            "customer_drawing_no": d.get("customer_drawing_no") or "-",
            "title": d.get("title") or "-",
            "revision": d.get("revision") or "-",
            "discipline": d.get("discipline") or "-",
            "so_no": d.get("so_no") or "-",
            "bom_no": d.get("bom_no") or "-",
            "project_name": d.get("project_name") or "-",
            "prepared_by": d.get("prepared_by") or "-",
            "request_by_sales": d.get("request_by_sales") or "-",
            "status": d.get("status") or "-",
            "approval_status": d.get("approval_status") or "draft",
            "dc_stamp_by": dc.get("name") or "-",
            "ttd": ttd_str,
        })
    return rows


_EXPORT_COLS = [
    ("drawing_no", "Drawing No"),
    ("customer_drawing_no", "Cust DWG No"),
    ("title", "Title"),
    ("revision", "Rev"),
    ("discipline", "Discipline"),
    ("so_no", "SO"),
    ("bom_no", "BOM"),
    ("project_name", "Project"),
    ("prepared_by", "Prepared By"),
    ("request_by_sales", "Request By (Sales)"),
    ("status", "Status"),
    ("approval_status", "Approval"),
    ("dc_stamp_by", "DC Stamp By"),
    ("ttd", "TTD"),
]


def _build_export_xlsx(rows: list, meta_line: str) -> bytes:
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Drawing Master List"

    ws.cell(row=1, column=1, value="MKS-F-ENG-005 Drawing Master List").font = Font(bold=True, size=13, color="1E293B")
    ws.cell(row=2, column=1, value=meta_line).font = Font(size=9, italic=True, color="64748B")

    header_row = 4
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (_, label) in enumerate(_EXPORT_COLS, start=1):
        c = ws.cell(row=header_row, column=ci, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for ri, row in enumerate(rows, start=header_row + 1):
        for ci, (key, _) in enumerate(_EXPORT_COLS, start=1):
            c = ws.cell(row=ri, column=ci, value=row.get(key, ""))
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border

    widths = [22, 16, 26, 7, 12, 10, 12, 24, 16, 16, 11, 12, 16, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_export_pdf(rows: list, meta_line: str) -> bytes:
    import io as _io
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=8)
    head_style = ParagraphStyle("head", parent=styles["Normal"], fontSize=7.5, leading=9,
                                textColor=colors.white, fontName="Helvetica-Bold")

    elems = [
        Paragraph("MKS-F-ENG-005 Drawing Master List",
                  ParagraphStyle("t", parent=styles["Title"], fontSize=15, spaceAfter=2)),
        Paragraph(meta_line, ParagraphStyle("m", parent=styles["Normal"], fontSize=8,
                                            textColor=colors.HexColor("#64748B"))),
        Spacer(1, 6),
    ]

    data = [[Paragraph(label, head_style) for _, label in _EXPORT_COLS]]
    for row in rows:
        data.append([Paragraph(str(row.get(key, "") or "-"), cell_style) for key, _ in _EXPORT_COLS])

    col_widths = [w * mm for w in [26, 20, 34, 8, 16, 14, 16, 32, 20, 20, 14, 16, 22, 40]]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elems.append(table)
    doc.build(elems)
    return buf.getvalue()


@router.get("/drawings/export")
async def export_drawings(
    q: Optional[str] = None,
    discipline: Optional[str] = None,
    status: Optional[str] = None,
    so_no: Optional[str] = None,
    format: str = "xlsx",
    current: dict = Depends(get_current_user),
):
    """Ekspor Master Drawing List (mengikuti filter aktif) ke Excel (.xlsx) atau PDF untuk arsip Engineering."""
    from fastapi.responses import StreamingResponse
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")

    fmt = (format or "xlsx").lower()
    if fmt not in ("xlsx", "pdf"):
        raise HTTPException(status_code=400, detail="Format harus 'xlsx' atau 'pdf'")

    filt = {"deleted_at": {"$exists": False}}
    if discipline: filt["discipline"] = discipline
    if status: filt["status"] = status
    if so_no: filt["so_no"] = so_no
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [
            {"drawing_no": rx}, {"title": rx}, {"project_name": rx},
            {"so_no": rx}, {"prepared_by": rx}, {"remark": rx},
            {"customer_drawing_no": rx}, {"customer_name": rx},
        ]
    docs = await db.drawings.find(filt, {"_id": 0}).sort("updated_at", -1).limit(5000).to_list(length=5000)
    rows = _export_rows(docs)

    now = datetime.now(timezone.utc)
    exported_by = current.get("name") or current.get("username") or "-"
    meta_line = (f"Total {len(rows)} entri | Diekspor oleh {exported_by} | "
                 f"{now.strftime('%d %b %Y %H:%M')} UTC")
    ts = now.strftime("%Y%m%d_%H%M")

    try:
        await log_action(current, "drawing_export", "drawings", "-",
                         {"format": fmt, "count": len(rows)})
    except Exception:
        pass

    if fmt == "xlsx":
        data = _build_export_xlsx(rows, meta_line)
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="MasterDrawingList_{ts}.xlsx"'},
        )
    data = _build_export_pdf(rows, meta_line)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="MasterDrawingList_{ts}.pdf"'},
    )



@router.get("/drawings/repeat-search")
async def repeat_search_drawings(
    q: Optional[str] = None,
    limit: int = 30,
    current: dict = Depends(get_current_user),
):
    """Cari drawing lama untuk Repeat Order (by Drawing No / Customer DWG No / SO / Project / Customer).
    Balikkan ringkasan + indikator ketersediaan MKS drawing, Customer drawing, Nesting, & Costing
    sehingga engineer bisa auto-pull data lama ke DRF repeat.
    Didefinisikan SEBELUM route generic /drawings/{drawing_id}.
    """
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    filt = {"deleted_at": {"$exists": False}}
    if q and q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [
            {"drawing_no": rx}, {"customer_drawing_no": rx}, {"so_no": rx},
            {"title": rx}, {"project_name": rx}, {"customer_name": rx},
        ]
    docs = await db.drawings.find(filt, {"_id": 0}).sort("updated_at", -1).limit(limit).to_list(length=limit)

    # Kumpulkan bom_id untuk cek attachment nesting/costing dalam 1 query
    bom_ids = [d.get("bom_id") for d in docs if d.get("bom_id")]
    att_map: dict = {}
    if bom_ids:
        atts = await db.bom_attachments.find(
            {"bom_id": {"$in": bom_ids}, "deleted_at": {"$exists": False}},
            {"_id": 0, "bom_id": 1, "category": 1},
        ).to_list(length=2000)
        for a in atts:
            att_map.setdefault(a.get("bom_id"), set()).add(a.get("category"))

    out = []
    for d in docs:
        cats = att_map.get(d.get("bom_id"), set())
        extras = d.get("additional_files") or d.get("extras") or []
        out.append({
            "id": d.get("id"),
            "drawing_no": d.get("drawing_no"),
            "customer_drawing_no": d.get("customer_drawing_no") or "",
            "so_no": d.get("so_no") or "",
            "title": d.get("title") or d.get("project_name") or "",
            "project_name": d.get("project_name") or "",
            "customer_name": d.get("customer_name") or "",
            "customer_code": d.get("customer_code") or "MKS",
            "project_initial": d.get("project_initial") or "",
            "drawing_type": d.get("drawing_type") or "Assembly",
            "discipline": d.get("discipline") or "Mechanical",
            "class_material": d.get("class_material") or "",
            "bom_id": d.get("bom_id") or "",
            "bom_no": d.get("bom_no") or "",
            "has_mks": bool(d.get("file_id")),
            "has_customer_ref": bool(d.get("customer_ref_file_id")),
            "extras_count": len(extras),
            "has_nesting": ("nesting" in cats) or any((e.get("label") or "").lower().find("nest") >= 0 for e in extras),
            "has_costing": ("costing" in cats) or ("costing_prev" in cats),
            "updated_at": d.get("updated_at"),
        })
    return {"items": out, "total": len(out)}



@router.post("/drawings")
async def create_drawing(payload: DrawingIn, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")

    # Auto-generate if empty
    auto_generated = False
    drawing_no = (payload.drawing_no or "").strip()
    meta = {}
    if not drawing_no:
        meta = await _next_drawing_no(payload.customer_code, payload.project_initial, payload.drawing_type)
        drawing_no = meta["drawing_no"]
        auto_generated = True

    # Uniqueness check on drawing_no + revision
    existing = await db.drawings.find_one({
        "drawing_no": drawing_no,
        "revision": payload.revision,
        "deleted_at": {"$exists": False},
    })
    if existing:
        raise HTTPException(status_code=409, detail=f"Drawing '{drawing_no}' rev {payload.revision} sudah ada")

    user_name = current.get("username") or current.get("name")

    # Handle BOM linking
    bom_id_final = None
    bom_no_final = ""
    if payload.bom_link_mode == "create_new":
        # Auto-generate BOM or use provided bom_no
        bom_no_to_create = (payload.bom_no or "").strip()
        auto_bom = False
        if not bom_no_to_create:
            # Import lazily to avoid circular
            from routers.bom import _next_bom_no
            bom_info = await _next_bom_no()
            bom_no_to_create = bom_info["bom_no"]
            auto_bom = True
        # Check uniqueness
        existing_bom = await db.boms.find_one({"bom_no": bom_no_to_create, "deleted_at": {"$exists": False}})
        if existing_bom:
            raise HTTPException(status_code=409, detail=f"BOM '{bom_no_to_create}' sudah ada — pilih 'existing' atau kosongkan untuk auto")

        # Repeat Order: copy items dari source BOM
        copied_items = []
        source_bom = None
        if payload.source_bom_id:
            source_bom = await db.boms.find_one(
                {"id": payload.source_bom_id, "deleted_at": {"$exists": False}},
                {"_id": 0, "items": 1, "class_material": 1, "bom_no": 1}
            )
            if not source_bom:
                raise HTTPException(status_code=404, detail="Source BOM (untuk Repeat Order) tidak ditemukan")
            for idx, it in enumerate((source_bom.get("items") or []), start=1):
                copied_items.append({
                    "item_no": idx,
                    "item_name": it.get("item_name") or it.get("item_specification") or "",
                    "item_specification": it.get("item_specification") or "",
                    "qty": float(it.get("qty") or 0),
                    "uom": it.get("uom") or "",
                    "material": it.get("material") or "",
                    "weight_kg": it.get("weight_kg"),
                    "purchase_due_date": "",  # reset — user isi baru
                    "remark": it.get("remark") or "",
                })
        # Create BOM record
        from routers.bom import normalize_so_no as _norm_so
        bom_doc = {
            "id": str(uuid.uuid4()),
            "so_no": _norm_so(payload.so_no),
            "rev_no": 0,
            "bom_no": bom_no_to_create,
            "project_name": payload.project_name.strip(),
            "project_dwg": drawing_no,  # link back
            "customer": payload.customer_code.upper().strip(),
            "class_material": (payload.class_material or (source_bom or {}).get("class_material") or "").strip(),
            "delivery_date": "",
            "bom_date": datetime.now(timezone.utc).date().isoformat(),
            "prepared_by": (payload.prepared_by or user_name or "").strip(),
            "items": copied_items,
            "annotations": {},
            "revision_reason": "",
            "auto_generated": auto_bom,
            "source": "repeat_order" if payload.source_bom_id else "drawing_register",
            "source_bom_id": payload.source_bom_id or None,
            "source_bom_no": (source_bom or {}).get("bom_no") if source_bom else None,
            "is_repeat": bool(payload.source_bom_id),
            "uploaded_by_id": current.get("id"),
            "uploaded_by_name": user_name,
            "uploaded_by_role": current.get("role"),
            "uploaded_at": _now_iso(),
            "original_filename": None,
            # Iter 35 workflow — new BOMs start as draft
            "engineering_status": "draft",
            "signatures": {
                "prepared_by": None,
                "checked_by": None,
                "acknowledged_by": None,
                "approved_by": None,
            },
        }
        await db.boms.insert_one(bom_doc.copy())
        bom_id_final = bom_doc["id"]
        bom_no_final = bom_no_to_create

        # Repeat Order: copy attachments (drawing PDF, customer_ref, nesting, costing) from source BOM
        # Reference-copy: same file_id shared. If user replaces file, sistem create new file.
        if payload.source_bom_id:
            src_attachments = await db.bom_attachments.find(
                {"bom_id": payload.source_bom_id, "deleted_at": {"$exists": False},
                 "category": {"$in": ["drawing", "customer_ref", "nesting", "costing", "costing_prev"]}},
                {"_id": 0},
            ).to_list(length=100)
            for src in src_attachments:
                new_att = {**src}
                new_att["id"] = str(uuid.uuid4())
                new_att["bom_id"] = bom_id_final
                new_att["copied_from"] = src.get("id")
                new_att["copied_at"] = _now_iso()
                # Costing sekarang jadi "costing_prev" di BOM baru (harga lama sebagai referensi)
                if src.get("category") == "costing":
                    new_att["category"] = "costing_prev"
                await db.bom_attachments.insert_one(new_att)
    elif payload.bom_link_mode == "existing":
        # Verify existing BOM
        if not payload.bom_id:
            raise HTTPException(status_code=400, detail="bom_id wajib untuk mode existing")
        b = await db.boms.find_one({"id": payload.bom_id, "deleted_at": {"$exists": False}}, {"bom_no": 1})
        if not b:
            raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
        bom_id_final = payload.bom_id
        bom_no_final = b.get("bom_no", "")

    doc = payload.model_dump()
    doc["drawing_no"] = drawing_no
    doc["customer_code"] = (payload.customer_code or "MKS").upper().strip()
    doc["customer_name"] = (payload.customer_name or "").strip()
    doc["po_customer_no"] = (payload.po_customer_no or "").strip()
    doc["item_name"] = (payload.item_name or "").strip()
    doc["item_qty"] = float(payload.item_qty or 0)
    doc["request_by_sales"] = (payload.request_by_sales or "").strip()
    doc["project_initial"] = (payload.project_initial or "").upper().strip()
    doc["auto_generated"] = auto_generated

    # Auto-save customer_code back to customer master (upsert by name).
    # Bila engineer ketik manual customer_code, sekaligus persist ke customers collection.
    if doc["customer_name"] and doc["customer_code"]:
        try:
            existing_cust = await db.customers.find_one(
                {"name": {"$regex": f"^{re.escape(doc['customer_name'])}$", "$options": "i"},
                 "deleted_at": {"$exists": False}}
            )
            if existing_cust:
                if (existing_cust.get("customer_code") or "").upper() != doc["customer_code"]:
                    await db.customers.update_one(
                        {"id": existing_cust["id"]},
                        {"$set": {
                            "customer_code": doc["customer_code"],
                            "customer_code_updated_by": user_name,
                            "updated_at": _now_iso(),
                        }},
                    )
            else:
                await db.customers.insert_one({
                    "id": str(uuid.uuid4()),
                    "name": doc["customer_name"],
                    "customer_code": doc["customer_code"],
                    "address": "", "pic": "", "phone": "", "email": "", "notes": "",
                    "created_at": _now_iso(),
                    "created_by_name": user_name,
                    "auto_created_from": "drawing_register",
                })
        except Exception:
            pass  # non-fatal — jangan gagalkan create drawing
    doc["year_month"] = meta.get("year_month")
    doc["monthly_running"] = meta.get("monthly_running")
    doc["type_letter"] = meta.get("type_letter")
    doc["type_seq"] = meta.get("type_seq")
    doc["bom_id"] = bom_id_final
    doc["bom_no"] = bom_no_final
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = _now_iso()
    doc["created_by"] = user_name
    doc["updated_at"] = _now_iso()
    doc["updated_by"] = user_name
    doc["file_id"] = None
    doc["filename"] = None
    doc["file_uploaded_at"] = None
    doc["pdf_match_status"] = None
    doc["pdf_match_note"] = None
    # Iter 16 — Digital Approval Workflow untuk Drawing
    # Sequential: draft → pending_eng_head → pending_qc → pending_sales → approved → controlled → released
    doc["approval_status"] = "draft"
    doc["approvals"] = []  # array of {stage, name, role, user_id, username, at, notes}
    # Work-progress tracking (untuk Master Drawing List)
    doc["work_category"] = ""        # simple|moderate|complex — WAJIB dipilih saat upload MKS
    doc["work_completed_at"] = None  # di-set saat Eng Leader (eng_head) approve hasil kerja
    doc["request_received_at"] = None  # = DRF.accepted_at (Riski terima request dari Sales)
    doc["work_started_at"] = None      # = DRF.work_started_at (eng staff klik TERIMA)
    if payload.from_drf_id:
        _drf = await db.drawing_requests.find_one(
            {"id": payload.from_drf_id}, {"_id": 0, "accepted_at": 1, "work_started_at": 1})
        if _drf:
            doc["request_received_at"] = _drf.get("accepted_at")
            doc["work_started_at"] = _drf.get("work_started_at")
    # Iter 19 — Link back to Drawing Request Form (kalau dibuat dari DRF)
    if payload.from_drf_id:
        doc["from_drf_id"] = payload.from_drf_id
    await db.drawings.insert_one(doc.copy())
    # Update DRF: link drawing_id back
    if payload.from_drf_id:
        await db.drawing_requests.update_one(
            {"id": payload.from_drf_id, "deleted_at": {"$exists": False}},
            {"$set": {
                "linked_drawing_id": doc["id"],
                "status": "in_progress",
                "updated_at": _now_iso(),
            }},
        )
    await log_action(current, "drawing_create", "drawings", doc["id"], {"drawing_no": doc["drawing_no"], "bom_no": bom_no_final, "auto": auto_generated, "from_drf_id": payload.from_drf_id or None})
    return _clean(doc)


@router.put("/drawings/{drawing_id}")
async def update_drawing(drawing_id: str, payload: DrawingIn, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not _can_modify_drawing(current, existing):
        assigned_name = existing.get("assigned_to_name", "-")
        raise HTTPException(status_code=403, detail=f"Drawing ini di-assign ke {assigned_name}. Hanya orang tersebut atau Eng Head yang bisa edit.")
    user_name = current.get("username") or current.get("name")
    update = payload.model_dump()
    update["drawing_no"] = update["drawing_no"].strip()
    update["customer_code"] = (update.get("customer_code") or "MKS").upper().strip()
    update["customer_name"] = (update.get("customer_name") or "").strip()
    update["request_by_sales"] = (update.get("request_by_sales") or "").strip()
    update["updated_at"] = _now_iso()
    update["updated_by"] = user_name
    # Iter 20 — Preserve assigned_to fields (hanya bisa diubah via /assign endpoint)
    update.pop("assigned_to_user_id", None)
    update.pop("assigned_to_name", None)
    # Preserve link ke DRF (immutable)
    update.pop("from_drf_id", None)
    await db.drawings.update_one({"id": drawing_id}, {"$set": update})

    # Auto-save customer_code back to customer master
    if update["customer_name"] and update["customer_code"]:
        try:
            existing_cust = await db.customers.find_one(
                {"name": {"$regex": f"^{re.escape(update['customer_name'])}$", "$options": "i"},
                 "deleted_at": {"$exists": False}}
            )
            if existing_cust:
                if (existing_cust.get("customer_code") or "").upper() != update["customer_code"]:
                    await db.customers.update_one(
                        {"id": existing_cust["id"]},
                        {"$set": {"customer_code": update["customer_code"],
                                  "customer_code_updated_by": user_name,
                                  "updated_at": _now_iso()}},
                    )
            else:
                await db.customers.insert_one({
                    "id": str(uuid.uuid4()),
                    "name": update["customer_name"],
                    "customer_code": update["customer_code"],
                    "address": "", "pic": "", "phone": "", "email": "", "notes": "",
                    "created_at": _now_iso(),
                    "created_by_name": user_name,
                    "auto_created_from": "drawing_register",
                })
        except Exception:
            pass
    return {"success": True}


@router.delete("/drawings/{drawing_id}")
async def delete_drawing(drawing_id: str, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    # Hanya boleh hapus saat masih DRAFT / revisi (belum masuk approval / sudah dikembalikan)
    if existing.get("approval_status") not in (None, "", "draft"):
        raise HTTPException(status_code=409, detail="Hanya drawing berstatus DRAFT/revisi yang bisa dihapus. Minta Eng Leader kembalikan untuk revisi dulu.")
    # Hapus semua file GridFS terkait (MKS, customer ref, extras, CAD)
    fids = []
    if existing.get("file_id"):
        fids.append(existing["file_id"])
    if existing.get("customer_ref_file_id"):
        fids.append(existing["customer_ref_file_id"])
    for f in (existing.get("additional_files") or []):
        if f.get("file_id"):
            fids.append(f["file_id"])
    for f in (existing.get("cad_files") or []):
        if f.get("file_id"):
            fids.append(f["file_id"])
    for fid in fids:
        try:
            await _fs().delete(ObjectId(fid))
        except Exception:
            pass
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {"deleted_at": _now_iso(), "deleted_by": current.get("username")}},
    )
    await log_action(current, "drawing_delete", "drawings", drawing_id, {"drawing_no": existing.get("drawing_no")})
    return {"success": True}


class DrawingBasicIn(BaseModel):
    title: Optional[str] = None
    drawing_type: Optional[str] = None
    customer_drawing_no: Optional[str] = None
    project_name: Optional[str] = None


@router.patch("/drawings/{drawing_id}/basic-info")
async def update_drawing_basic(drawing_id: str, payload: DrawingBasicIn, current: dict = Depends(get_current_user)):
    """Edit ringan (judul/type/cust dwg no/project) — hanya saat DRAFT/revisi. Untuk rapikan data new order."""
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not _can_modify_drawing(current, existing):
        raise HTTPException(status_code=403, detail=f"Drawing di-assign ke {existing.get('assigned_to_name','-')}. Hanya orang tsb / Eng Head yang bisa edit.")
    if existing.get("approval_status") not in (None, "", "draft"):
        raise HTTPException(status_code=409, detail="Hanya drawing DRAFT/revisi yang bisa diedit di sini.")
    upd = {k: (v.strip() if isinstance(v, str) else v) for k, v in payload.model_dump().items() if v is not None}
    if not upd:
        return {"success": True, "unchanged": True}
    upd["updated_at"] = _now_iso()
    upd["updated_by"] = current.get("username") or current.get("name")
    await db.drawings.update_one({"id": drawing_id}, {"$set": upd})
    await log_action(current, "drawing_edit_basic", "drawings", drawing_id, upd)
    return {"success": True}


VALID_WORK_CATEGORIES = {"simple", "moderate", "complex"}


class WorkCategoryIn(BaseModel):
    work_category: str


@router.post("/drawings/{drawing_id}/work-category")
async def set_work_category(drawing_id: str, payload: WorkCategoryIn, current: dict = Depends(get_current_user)):
    """Engineer pilih kategori pekerjaan drawing: SIMPLE / MODERATE / COMPLEX.
    Wajib diisi sebelum submit ke Eng Leader. Engineer yang menggambar paling tahu."""
    cat = (payload.work_category or "").strip().lower()
    if cat not in VALID_WORK_CATEGORIES:
        raise HTTPException(status_code=400, detail="Kategori harus salah satu: simple, moderate, complex")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    if not _can_modify_drawing(current, existing):
        raise HTTPException(status_code=403, detail=f"Drawing di-assign ke {existing.get('assigned_to_name','-')}. Hanya orang tsb / Eng Head yang bisa ubah.")
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {"work_category": cat, "updated_at": _now_iso(),
                  "updated_by": current.get("username") or current.get("name")}},
    )
    await log_action(current, "drawing_set_work_category", "drawings", drawing_id, {"work_category": cat})
    return {"success": True, "work_category": cat}


class DrfItemIn(BaseModel):
    item_name: str = ""
    item_qty: float = 0


@router.post("/drawings/{drawing_id}/drf-item")
async def set_drawing_drf_item(drawing_id: str, payload: DrfItemIn, current: dict = Depends(get_current_user)):
    """Engineer pilih Nama Item (dari daftar item DRF) & Qty untuk drawing ini.
    Qty bisa beda/partial per drawing. Dipakai auto-isi qty stamping SO (Sales tetap bisa ubah)."""
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    if not _can_modify_drawing(current, existing):
        raise HTTPException(status_code=403, detail=f"Drawing di-assign ke {existing.get('assigned_to_name','-')}. Hanya orang tsb / Eng Head yang bisa ubah.")
    item_name = (payload.item_name or "").strip()
    try:
        item_qty = float(payload.item_qty or 0)
    except (TypeError, ValueError):
        item_qty = 0
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {"item_name": item_name, "item_qty": item_qty, "updated_at": _now_iso(),
                  "updated_by": current.get("username") or current.get("name")}},
    )
    await log_action(current, "drawing_set_drf_item", "drawings", drawing_id, {"item_name": item_name, "item_qty": item_qty})
    return {"success": True, "item_name": item_name, "item_qty": item_qty}



# =========================================================================
# Iter 22 — Link BOM ke drawing yang sudah ada (untuk Engineering Work Order page)
# =========================================================================
class LinkBomIn(BaseModel):
    bom_link_mode: str = "none"   # none | create_new | existing
    bom_no: str = ""              # untuk create_new
    bom_id: str = ""              # untuk existing


@router.post("/drawings/{drawing_id}/link-bom")
async def drawing_link_bom(drawing_id: str, payload: LinkBomIn, current: dict = Depends(get_current_user)):
    """Link BOM ke drawing yang sudah ada — dipakai di Engineering Work Order page
    ketika Trisna (assignee) menentukan BOM linking setelah drawing di-register oleh Eng Head."""
    drawing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not _can_modify_drawing(current, drawing):
        assigned_name = drawing.get("assigned_to_name", "-")
        raise HTTPException(status_code=403, detail=f"Drawing di-assign ke {assigned_name} — Anda tidak berhak")

    user_name = current.get("username") or current.get("name")
    mode = (payload.bom_link_mode or "none").strip()

    bom_id_final = None
    bom_no_final = ""

    if mode == "none":
        pass  # unlink
    elif mode == "existing":
        if not payload.bom_id:
            raise HTTPException(status_code=400, detail="bom_id wajib untuk mode existing")
        b = await db.boms.find_one({"id": payload.bom_id, "deleted_at": {"$exists": False}}, {"bom_no": 1})
        if not b:
            raise HTTPException(status_code=404, detail="BOM tidak ditemukan")
        bom_id_final = payload.bom_id
        bom_no_final = b.get("bom_no", "")
    elif mode == "create_new":
        bom_no_to_create = (payload.bom_no or "").strip()
        auto_bom = False
        if not bom_no_to_create:
            from routers.bom import _next_bom_no
            bom_info = await _next_bom_no()
            bom_no_to_create = bom_info["bom_no"]
            auto_bom = True
        existing_bom = await db.boms.find_one({"bom_no": bom_no_to_create, "deleted_at": {"$exists": False}})
        if existing_bom:
            raise HTTPException(status_code=409, detail=f"BOM '{bom_no_to_create}' sudah ada")
        bom_doc = {
            "id": str(uuid.uuid4()),
            "so_no": (drawing.get("so_no") or "").strip(),
            "rev_no": 0,
            "bom_no": bom_no_to_create,
            "project_name": (drawing.get("project_name") or "").strip(),
            "project_dwg": drawing.get("drawing_no") or "",
            "customer": (drawing.get("customer_code") or "MKS").upper().strip(),
            "class_material": (drawing.get("class_material") or "").strip(),
            "delivery_date": "",
            "bom_date": datetime.now(timezone.utc).date().isoformat(),
            "prepared_by": (drawing.get("prepared_by") or user_name or "").strip(),
            "items": [],
            "annotations": {},
            "revision_reason": "",
            "auto_generated": auto_bom,
            "source": "work_order_link",
            "source_bom_id": None,
            "source_bom_no": None,
            "is_repeat": False,
            "uploaded_by_id": current.get("id"),
            "uploaded_by_name": user_name,
            "uploaded_by_role": current.get("role"),
            "uploaded_at": _now_iso(),
            "original_filename": None,
            "engineering_status": "draft",
            "signatures": {"prepared_by": None, "checked_by": None,
                           "acknowledged_by": None, "approved_by": None},
        }
        await db.boms.insert_one(bom_doc.copy())
        bom_id_final = bom_doc["id"]
        bom_no_final = bom_no_to_create
    else:
        raise HTTPException(status_code=400, detail=f"Mode tidak valid: {mode}")

    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {"bom_id": bom_id_final, "bom_no": bom_no_final,
                  "updated_at": _now_iso(), "updated_by": user_name}},
    )
    await log_action(current, "drawing_link_bom", "drawings", drawing_id,
                     {"mode": mode, "bom_no": bom_no_final})
    return {"success": True, "bom_id": bom_id_final, "bom_no": bom_no_final}


# =========================================================================
# Iter 20 — Assign engineer & list engineering users
# =========================================================================
ENG_ROLE_LIST = ["eng_staff", "engineering", "eng_head", "eng_leader"]


@router.get("/drawings/my-assignments")
async def list_my_assignments(current: dict = Depends(get_current_user)):
    """Iter 21 — List drawing yang di-assign ke current user (untuk Eng Staff).
    Return drawings dengan assigned_to_user_id == current user_id."""
    if not _can_edit(current):
        return {"items": [], "total": 0}
    docs = await db.drawings.find(
        {"assigned_to_user_id": current["id"], "deleted_at": {"$exists": False}},
        {"_id": 0},
    ).sort("updated_at", -1).to_list(length=200)
    return {"items": docs, "total": len(docs)}


@router.get("/drawings/engineering-users")
async def list_engineering_users(current: dict = Depends(get_current_user)):
    """Return list of engineering users untuk dropdown assign."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    users = await db.users.find(
        {"role": {"$in": ENG_ROLE_LIST}, "active": {"$ne": False}},
        {"_id": 0, "id": 1, "username": 1, "name": 1, "role": 1},
    ).sort("name", 1).to_list(length=200)
    return {"items": users}


class AssignIn(BaseModel):
    assigned_to_user_id: str = ""
    assigned_to_name: str = ""  # opsional — override display name


@router.post("/drawings/{drawing_id}/assign")
async def assign_drawing(
    drawing_id: str,
    payload: AssignIn,
    current: dict = Depends(get_current_user),
):
    """Eng Head/Admin assign drawing ke engineer tertentu.
    Kosongkan assigned_to_user_id untuk hapus assignment."""
    if not _is_eng_head_role(current):
        raise HTTPException(status_code=403, detail="Hanya Eng Head/Admin yang boleh assign")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")

    upd = {"updated_at": _now_iso()}
    if payload.assigned_to_user_id:
        u = await db.users.find_one({"id": payload.assigned_to_user_id})
        if not u:
            raise HTTPException(status_code=404, detail="User tidak ditemukan")
        if u.get("role") not in ENG_ROLE_LIST:
            raise HTTPException(status_code=400, detail="User bukan Engineering")
        upd["assigned_to_user_id"] = u["id"]
        upd["assigned_to_name"] = payload.assigned_to_name or u.get("name") or u.get("username")
        upd["assigned_by"] = current.get("name") or current.get("username")
        upd["assigned_at"] = _now_iso()
    else:
        upd["assigned_to_user_id"] = ""
        upd["assigned_to_name"] = ""
    await db.drawings.update_one({"id": drawing_id}, {"$set": upd})
    await log_action(current, "drawing_assign", "drawings", drawing_id, {
        "drawing_no": existing.get("drawing_no"),
        "assigned_to": upd.get("assigned_to_name"),
    })
    out = await db.drawings.find_one({"id": drawing_id}, {"_id": 0})
    return out


@router.post("/drawings/verify-pdf")
async def verify_pdf(
    drawing_no: str = Form(...),
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    """Verify PDF contents against a drawing_no BEFORE actually uploading. Returns match info."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Hanya PDF yang bisa diverifikasi")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    text, src = _extract_pdf_text_with_source(content)
    result = _check_drawing_no_in_text(drawing_no, text, source=src)
    result["text_extracted_chars"] = len(text)
    return result


@router.post("/drawings/{drawing_id}/upload")
async def upload_drawing_pdf(
    drawing_id: str,
    force: bool = Form(False),
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    """Upload PDF for a registered drawing. If drawing_no mismatch detected → warning
    (but proceed anyway per user preference). Client can call /verify-pdf first.
    """
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not _can_modify_drawing(current, existing):
        assigned_name = existing.get("assigned_to_name", "-")
        raise HTTPException(status_code=403, detail=f"Drawing ini di-assign ke {assigned_name}. Hanya orang tersebut yang bisa upload PDF.")
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Hanya PDF yang boleh diupload")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 100 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File > 100 MB tidak diizinkan")

    # Verify content
    text, src = _extract_pdf_text_with_source(content)
    check = _check_drawing_no_in_text(existing["drawing_no"], text, source=src)

    # STRICT (New Order): nomor drawing sudah di-register oleh sistem → PDF WAJIB mengikuti nomor tsb.
    # Bila nomor DWG di PDF terbaca tapi TIDAK cocok → TOLAK: file tidak disimpan (file lama tetap aman).
    strict = bool(existing.get("auto_generated"))
    if not strict and existing.get("from_drf_id"):
        drf = await db.drawing_requests.find_one({"id": existing["from_drf_id"]}, {"_id": 0, "request_type": 1})
        strict = (drf or {}).get("request_type") == "new_order"
    readable = (check.get("detected_source") not in (None, "none")) and bool(text and text.strip())
    if strict and not check["match"] and readable:
        detected = check.get("detected_no") or ""
        cands = check.get("extracted_candidates") or []
        raise HTTPException(status_code=422, detail={
            "code": "dwg_no_mismatch",
            "message": (
                f"Nomor DWG di dalam PDF tidak sesuai dengan nomor drawing terdaftar "
                f"({existing['drawing_no']}). "
                + (f"Terbaca di PDF: {detected}. " if detected else (f"Kandidat terbaca: {', '.join(cands[:4])}. " if cands else "Nomor terdaftar tidak ditemukan di PDF. "))
                + "File TIDAK disimpan. Silakan revisi PDF agar nomornya sama, lalu upload ulang."
            ),
            "expected": existing["drawing_no"],
            "detected": detected,
            "candidates": cands,
        })

    # Replace old file if any — TAPI jangan hapus bila file lama masih jadi history revisi
    if existing.get("file_id") and not _file_in_revisions(existing, existing["file_id"]):
        try:
            await _fs().delete(ObjectId(existing["file_id"]))
        except Exception:
            pass

    fs = _fs()
    file_id = await fs.upload_from_stream(
        file.filename, content,
        metadata={"content_type": "application/pdf", "drawing_id": drawing_id},
    )

    user_name = current.get("username") or current.get("name")
    update = {
        "file_id": str(file_id),
        "filename": file.filename,
        "file_uploaded_at": _now_iso(),
        "file_uploaded_by": user_name,
        "pdf_match_status": "verified" if check["match"] else "warning",
        "pdf_match_note": check["note"],
        "pdf_extracted_candidates": check["extracted_candidates"],
        "pdf_detected_no": check.get("detected_no") or "",
        "pdf_detected_source": check.get("detected_source") or "native",
        "updated_at": _now_iso(),
        "updated_by": user_name,
    }
    # Auto-status: kalau drawing masih "Draft" dan drawing PDF pertama kali di-upload, promote ke "Issued"
    prev_status = (existing.get("status") or "Draft").strip()
    if prev_status.lower() == "draft":
        update["status"] = "Issued"
        update["status_auto_promoted_at"] = _now_iso()
        update["status_auto_promoted_by"] = user_name
    await db.drawings.update_one({"id": drawing_id}, {"$set": update})
    await log_action(current, "drawing_upload", "drawings", drawing_id, {
        "filename": file.filename, "match": check["match"],
        "status_auto_promoted": prev_status.lower() == "draft",
    })
    return {
        "success": True,
        "match": check["match"],
        "note": check["note"],
        "extracted_candidates": check["extracted_candidates"],
        "detected_no": check.get("detected_no") or "",
        "detected_source": check.get("detected_source") or "native",
        "current_drawing_no": existing["drawing_no"],
        "file_uploaded_at": update["file_uploaded_at"],
        "filename": file.filename,
        "status": update.get("status") or prev_status,
        "status_auto_promoted": prev_status.lower() == "draft",
    }


class RenameDrawingIn(BaseModel):
    new_drawing_no: str


@router.post("/drawings/{drawing_id}/rename")
async def rename_drawing_no(drawing_id: str, payload: RenameDrawingIn, current: dict = Depends(get_current_user)):
    """Ganti nomor drawing agar sama dengan nomor yang terdeteksi/tercetak di PDF (repeat/manual upload).
    Cek keunikan (drawing_no + revision) & update dokumen terkait (project_dwg di BOM)."""
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not _can_modify_drawing(current, existing):
        assigned_name = existing.get("assigned_to_name", "-")
        raise HTTPException(status_code=403, detail=f"Drawing di-assign ke {assigned_name} — Anda tidak berhak")

    new_no = (payload.new_drawing_no or "").strip()
    if not new_no:
        raise HTTPException(status_code=400, detail="Nomor drawing baru wajib diisi")
    if new_no == existing.get("drawing_no"):
        return {"success": True, "drawing_no": new_no, "unchanged": True}

    revision = existing.get("revision") or "Rev-0"
    dup = await db.drawings.find_one({
        "drawing_no": new_no, "revision": revision,
        "id": {"$ne": drawing_id}, "deleted_at": {"$exists": False},
    })
    if dup:
        raise HTTPException(status_code=409, detail=f"Nomor '{new_no}' {revision} sudah dipakai drawing lain")

    old_no = existing.get("drawing_no")
    user_name = current.get("username") or current.get("name")
    now_iso = _now_iso()
    hist_entry = {"from": old_no, "to": new_no, "by": user_name, "at": now_iso}
    set_fields = {
        "drawing_no": new_no, "updated_at": now_iso, "updated_by": user_name,
        "renamed_from": old_no, "renamed_at": now_iso,
    }
    # Re-verify status match terhadap isi PDF (bila sudah ada file) supaya warning konsisten.
    if existing.get("file_id"):
        try:
            existing["drawing_no"] = new_no
            raw = await _target_raw_bytes(existing, "mks", "")
            text = _extract_pdf_text(raw)
            check = _check_drawing_no_in_text(new_no, text)
            set_fields["pdf_match_status"] = "verified" if check["match"] else "warning"
            set_fields["pdf_match_note"] = check["note"]
            set_fields["pdf_extracted_candidates"] = check["extracted_candidates"]
            set_fields["pdf_detected_no"] = check.get("detected_no") or ""
        except Exception:
            pass
    await db.drawings.update_one({"id": drawing_id}, {"$set": set_fields, "$push": {"rename_history": hist_entry}})
    # Sinkronkan project_dwg di BOM terkait bila menunjuk nomor lama.
    if existing.get("bom_id"):
        try:
            await db.boms.update_one(
                {"id": existing["bom_id"], "project_dwg": old_no},
                {"$set": {"project_dwg": new_no, "updated_at": _now_iso()}},
            )
        except Exception:
            pass
    await log_action(current, "drawing_rename", "drawings", drawing_id,
                     {"from": old_no, "to": new_no})
    return {"success": True, "drawing_no": new_no, "from": old_no,
            "rename_history": (existing.get("rename_history") or []) + [hist_entry]}


@router.get("/drawings/{drawing_id}/preview")
async def preview_drawing(drawing_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not doc or not doc.get("file_id"):
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    stream = await _fs().open_download_stream(ObjectId(doc["file_id"]))
    raw = await stream.read()
    return StreamingResponse(
        io.BytesIO(raw),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{doc.get("filename") or doc["drawing_no"]}.pdf"'},
    )


@router.get("/drawings/{drawing_id}/download")
async def download_drawing(drawing_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not doc or not doc.get("file_id"):
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    stream = await _fs().open_download_stream(ObjectId(doc["file_id"]))
    raw = await stream.read()
    # Watermark UNCONTROLLED COPY untuk role non-DC/Admin (mis. Produksi/Store/QC) + footer printed-by.
    is_dc_or_admin = is_doc_control(current) or is_admin_like(current)
    show_watermark = doc.get("approval_status") in ("controlled", "approved") and not is_dc_or_admin
    printed_by = current.get("name") or current.get("username") or ""
    dc_stamp = doc.get("dc_stamp")
    if raw[:5].startswith(b"%PDF") and (show_watermark or printed_by or dc_stamp):
        try:
            raw = _apply_pdf_stamps(
                raw, approvals=[], dc_stamp=dc_stamp,
                watermark_uncontrolled=show_watermark, printed_by=printed_by,
            )
        except Exception:
            pass
    try:
        await log_action(current, "drawing_download", "drawings", drawing_id, {
            "drawing_no": doc.get("drawing_no"), "watermarked": show_watermark,
        })
    except Exception:
        pass
    return StreamingResponse(
        io.BytesIO(raw),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{doc.get("filename") or doc["drawing_no"]}.pdf"'},
    )


# ============ CUSTOMER REFERENCE (opsional) ============

@router.post("/drawings/{drawing_id}/upload-customer-ref")
async def upload_customer_ref(
    drawing_id: str,
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    """Upload PDF referensi dari CUSTOMER (bukan MKS drawing).
    Ini file acuan dari client yang jadi dasar gambar MKS.
    """
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Hanya PDF yang boleh diupload")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 100 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File > 100 MB tidak diizinkan")

    # Delete old customer ref if any
    if existing.get("customer_ref_file_id"):
        try:
            await _fs().delete(ObjectId(existing["customer_ref_file_id"]))
        except Exception:
            pass

    fs = _fs()
    file_id = await fs.upload_from_stream(
        file.filename, content,
        metadata={"content_type": "application/pdf", "drawing_id": drawing_id, "type": "customer_ref"},
    )

    user_name = current.get("username") or current.get("name")
    update = {
        "customer_ref_file_id": str(file_id),
        "customer_ref_filename": file.filename,
        "customer_ref_uploaded_at": _now_iso(),
        "customer_ref_uploaded_by": user_name,
        "updated_at": _now_iso(),
        "updated_by": user_name,
    }
    await db.drawings.update_one({"id": drawing_id}, {"$set": update})
    await log_action(current, "drawing_customer_ref_upload", "drawings", drawing_id, {"filename": file.filename})

    # Auto-baca "No. Drawing Customer" dari isi PDF (native → OCR fallback).
    detected_cust_no, cust_candidates = "", []
    try:
        txt, src = _extract_pdf_text_with_source(content, ocr_fallback=True)
        detected_cust_no, cust_candidates = _detect_customer_dwg_no(txt)
    except Exception:
        src = "none"
    already = (existing.get("customer_drawing_no") or "").strip()
    return {
        "success": True,
        **update,
        "detected_customer_drawing_no": detected_cust_no,
        "customer_drawing_candidates": cust_candidates,
        "detected_source": src if detected_cust_no else "none",
        "existing_customer_drawing_no": already,
        # Frontend pakai flag ini untuk tentukan: prefill popup atau minta input manual.
        "needs_manual_customer_no": (not detected_cust_no) and (not already),
    }


@router.get("/drawings/{drawing_id}/customer-ref/preview")
async def preview_customer_ref(drawing_id: str, current: dict = Depends(get_current_user)):
    """Preview Customer Reference PDF.

    Iter 18: Customer Ref juga dianggap sebagai dokumen controlled MKS — otomatis
    dapat watermark "UNCONTROLLED COPY WHEN PRINTED" + footer info user (kecuali
    yg akses adalah Doc Control atau Admin). Log audit trail print history.
    """
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not doc or not doc.get("customer_ref_file_id"):
        raise HTTPException(status_code=404, detail="Customer reference tidak ditemukan")
    stream = await _fs().open_download_stream(ObjectId(doc["customer_ref_file_id"]))
    raw = await stream.read()

    # Watermark logic: sama seperti PDF stamped — kalau drawing sudah controlled dan user bukan DC/admin, apply watermark
    is_dc_or_admin = is_doc_control(current) or is_admin_like(current)
    show_watermark = (
        doc.get("approval_status") in ("controlled", "approved") and not is_dc_or_admin
    )

    printed_by = current.get("name") or current.get("username") or ""

    # Iter 20 — Customer Ref juga bisa punya DC stamp Salma
    ref_dc_stamp = doc.get("customer_ref_dc_stamp")

    if show_watermark or printed_by or ref_dc_stamp:
        raw = _apply_pdf_stamps(
            raw,
            approvals=[],
            dc_stamp=ref_dc_stamp,   # ← stamp Salma di customer ref
            watermark_uncontrolled=show_watermark,
            printed_by=printed_by,
        )

    # Log audit trail
    try:
        await log_action(current, "customer_ref_preview", "drawings", drawing_id, {
            "drawing_no": doc.get("drawing_no"),
            "filename": doc.get("customer_ref_filename"),
            "watermarked": show_watermark,
        })
    except Exception:
        pass

    return StreamingResponse(
        io.BytesIO(raw),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="customer-ref-{doc["drawing_no"]}.pdf"'},
    )


@router.get("/drawings/{drawing_id}/customer-ref/download")
async def download_customer_ref(drawing_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not doc or not doc.get("customer_ref_file_id"):
        raise HTTPException(status_code=404, detail="Customer reference tidak ditemukan")
    stream = await _fs().open_download_stream(ObjectId(doc["customer_ref_file_id"]))
    raw = await stream.read()
    filename = doc.get("customer_ref_filename") or f"customer-ref-{doc['drawing_no']}.pdf"
    return StreamingResponse(
        io.BytesIO(raw),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/drawings/{drawing_id}/customer-ref")
async def delete_customer_ref(drawing_id: str, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if existing.get("customer_ref_file_id"):
        try:
            await _fs().delete(ObjectId(existing["customer_ref_file_id"]))
        except Exception:
            pass
    await db.drawings.update_one({"id": drawing_id}, {"$set": {
        "customer_ref_file_id": None, "customer_ref_filename": None,
        "customer_ref_uploaded_at": None, "customer_ref_uploaded_by": None,
    }})
    return {"success": True}


@router.delete("/drawings/{drawing_id}/file")
async def delete_drawing_pdf(drawing_id: str, current: dict = Depends(get_current_user)):
    """Delete just the uploaded PDF file, keep the drawing record."""
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if existing.get("file_id"):
        try:
            await _fs().delete(ObjectId(existing["file_id"]))
        except Exception:
            pass
    await db.drawings.update_one({"id": drawing_id}, {"$set": {
        "file_id": None, "filename": None,
        "file_uploaded_at": None, "file_uploaded_by": None,
        "pdf_match_status": None, "pdf_match_note": None, "pdf_extracted_candidates": None,
    }})
    await log_action(current, "drawing_delete_file", "drawings", drawing_id, {})
    return {"success": True}


# ============ MULTI-FILE (Additional Files per Drawing) ============
# User: "kadang dokumen drawing lebih dari 1 file" — support N extra files per drawing
# (misal: gambar rev-1 + rev-2 + detail zoom + BOM sheet). Stored in same GridFS bucket.

DRAWING_EXTRA_MAX_MB = 100
DRAWING_EXTRA_ALLOWED_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".dwg", ".dxf", ".xlsx", ".xls", ".zip"}


@router.post("/drawings/{drawing_id}/extras")
async def upload_extra_file(
    drawing_id: str,
    file: UploadFile = File(...),
    label: str = Form(""),  # optional caption/description
    current: dict = Depends(get_current_user),
):
    """Upload extra/additional file (any type) attached to a drawing. Multiple allowed."""
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in DRAWING_EXTRA_ALLOWED_EXT:
        raise HTTPException(status_code=400, detail=f"Ekstensi {ext} tidak diizinkan. Boleh: {sorted(DRAWING_EXTRA_ALLOWED_EXT)}")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > DRAWING_EXTRA_MAX_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"File > {DRAWING_EXTRA_MAX_MB} MB tidak diizinkan")

    fs = _fs()
    file_id = await fs.upload_from_stream(
        file.filename, content,
        metadata={
            "content_type": file.content_type or "application/octet-stream",
            "drawing_id": drawing_id,
            "type": "extra",
        },
    )

    entry = {
        "id": str(uuid.uuid4()),
        "file_id": str(file_id),
        "filename": file.filename,
        "label": (label or "").strip(),
        "ext": ext,
        "size": len(content),
        "content_type": file.content_type or "application/octet-stream",
        "uploaded_at": _now_iso(),
        "uploaded_by": current.get("username") or current.get("name"),
    }
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$push": {"additional_files": entry}, "$set": {"updated_at": _now_iso()}},
    )
    await log_action(current, "drawing_upload_extra", "drawings", drawing_id, {"filename": file.filename, "ext": ext})
    return {"success": True, "file": entry}


@router.get("/drawings/{drawing_id}/extras/{extra_id}/preview")
async def preview_extra_file(drawing_id: str, extra_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    entry = next((f for f in (doc.get("additional_files") or []) if f.get("id") == extra_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    try:
        stream = await _fs().open_download_stream(ObjectId(entry["file_id"]))
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File gagal dibaca: {e}")
    raw = await stream.read()

    # Iter 20 — Kalau PDF, apply DC stamp (kalau ada) + watermark untuk non-DC
    ctype = (entry.get("content_type") or "").lower()
    is_pdf = ctype == "application/pdf" or (entry.get("ext") or "").lower() == ".pdf"
    if is_pdf:
        is_dc_or_admin = is_doc_control(current) or is_admin_like(current)
        show_watermark = (
            doc.get("approval_status") in ("controlled", "approved") and not is_dc_or_admin
        )
        extra_dc_stamp = entry.get("dc_stamp")
        printed_by = current.get("name") or current.get("username") or ""
        if show_watermark or printed_by or extra_dc_stamp:
            try:
                raw = _apply_pdf_stamps(
                    raw,
                    approvals=[],
                    dc_stamp=extra_dc_stamp,
                    watermark_uncontrolled=show_watermark,
                    printed_by=printed_by,
                )
            except Exception:
                pass  # fallback ke raw kalau gagal parse PDF

    return StreamingResponse(
        io.BytesIO(raw),
        media_type=entry.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{entry.get("filename")}"'},
    )


@router.delete("/drawings/{drawing_id}/extras/{extra_id}")
async def delete_extra_file(drawing_id: str, extra_id: str, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    entry = next((f for f in (doc.get("additional_files") or []) if f.get("id") == extra_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    try:
        await _fs().delete(ObjectId(entry["file_id"]))
    except Exception:
        pass
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$pull": {"additional_files": {"id": extra_id}}, "$set": {"updated_at": _now_iso()}},
    )
    await log_action(current, "drawing_delete_extra", "drawings", drawing_id, {"extra_id": extra_id})
    return {"success": True}


# ============================================================================
# DWG CAD (File Asli) — file sumber gambar engineer (AutoCAD/Inventor/Solidworks/STEP dll).
# Disimpan terpisah dari additional_files (extras/nesting) agar tidak tercampur.
# Tidak dipreview (format native) — hanya list & download.
# ============================================================================
CAD_ALLOWED_EXT = {
    ".dwg", ".dxf", ".ipt", ".iam", ".idw", ".sldprt", ".sldasm", ".slddrw",
    ".step", ".stp", ".iges", ".igs", ".x_t", ".x_b", ".prt", ".catpart",
    ".catproduct", ".3dm", ".f3d", ".sat", ".stl", ".zip", ".rar", ".7z",
}
CAD_MAX_MB = 150


@router.post("/drawings/{drawing_id}/cad-files")
async def upload_cad_file(
    drawing_id: str,
    file: UploadFile = File(...),
    label: str = Form(""),
    current: dict = Depends(get_current_user),
):
    """Upload file CAD asli (AutoCAD/Inventor/dll). Bisa lebih dari 1 file."""
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    existing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in CAD_ALLOWED_EXT:
        raise HTTPException(status_code=400, detail=f"Ekstensi {ext} tidak diizinkan untuk CAD. Boleh: {sorted(CAD_ALLOWED_EXT)}")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > CAD_MAX_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"File > {CAD_MAX_MB} MB tidak diizinkan")

    fs = _fs()
    file_id = await fs.upload_from_stream(
        file.filename, content,
        metadata={"content_type": file.content_type or "application/octet-stream", "drawing_id": drawing_id, "type": "cad"},
    )
    entry = {
        "id": str(uuid.uuid4()),
        "file_id": str(file_id),
        "filename": file.filename,
        "label": (label or "").strip(),
        "ext": ext,
        "size": len(content),
        "content_type": file.content_type or "application/octet-stream",
        "uploaded_at": _now_iso(),
        "uploaded_by": current.get("username") or current.get("name"),
    }
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$push": {"cad_files": entry}, "$set": {"updated_at": _now_iso()}},
    )
    await log_action(current, "drawing_upload_cad", "drawings", drawing_id, {"filename": file.filename, "ext": ext})
    return {"success": True, "file": entry}


@router.get("/drawings/{drawing_id}/cad-files/{cad_id}/download")
async def download_cad_file(drawing_id: str, cad_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    entry = next((f for f in (doc.get("cad_files") or []) if f.get("id") == cad_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail="File CAD tidak ditemukan")
    try:
        stream = await _fs().open_download_stream(ObjectId(entry["file_id"]))
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File gagal dibaca: {e}")
    raw = await stream.read()
    return StreamingResponse(
        io.BytesIO(raw),
        media_type=entry.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{entry.get("filename")}"'},
    )


@router.delete("/drawings/{drawing_id}/cad-files/{cad_id}")
async def delete_cad_file(drawing_id: str, cad_id: str, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Engineering/Admin only")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    entry = next((f for f in (doc.get("cad_files") or []) if f.get("id") == cad_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail="File CAD tidak ditemukan")
    try:
        await _fs().delete(ObjectId(entry["file_id"]))
    except Exception:
        pass
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$pull": {"cad_files": {"id": cad_id}}, "$set": {"updated_at": _now_iso()}},
    )
    await log_action(current, "drawing_delete_cad", "drawings", drawing_id, {"cad_id": cad_id})
    return {"success": True}




# ============================================================================
# Iter 16 — Digital Approval Workflow untuk Drawing
# Strict sequential: draft → pending_eng_head → pending_qc → pending_sales → approved → controlled → released
# ============================================================================
from deps import is_eng_head, is_qc, is_doc_control, is_engineering, is_admin_like, is_super_admin_user, SALES_ROLES, ENGINEERING_ROLES, PRODUCTION_ROLES  # noqa: E402

STAGE_ORDER = ["eng_head", "qc", "sales"]  # sequential approval stages
STAGE_STATUS = {
    "eng_head": "pending_eng_head",
    "qc": "pending_qc",
    "sales": "pending_sales",
}
NEXT_STATUS_AFTER = {
    "pending_eng_head": "pending_qc",
    "pending_qc": "pending_sales",
    "pending_sales": "approved",
}


def _stage_allowed_roles(stage: str) -> tuple:
    if stage == "eng_head":
        return ("eng_leader", "eng_head", "engineering")
    if stage == "qc":
        return ("qc",)
    if stage == "sales":
        return SALES_ROLES
    return tuple()


def _sig_stamp(current: dict, notes: str = "") -> dict:
    return {
        "name": (current.get("name") or current.get("username") or "").strip(),
        "user_id": current.get("id"),
        "username": current.get("username"),
        "role": current.get("role"),
        "at": _now_iso(),
        "notes": (notes or "").strip(),
    }


class ApprovalActionIn(BaseModel):
    notes: str = ""
    # Iter 18 — Digital signature placement (opsional):
    # Approver klik posisi di PDF viewer → koordinat 0..1 relatif ke lebar/tinggi halaman
    stamp_x: Optional[float] = None
    stamp_y: Optional[float] = None
    stamp_page: Optional[int] = 0     # halaman (0-indexed). Default 0 = halaman pertama.
    stamp_size: Optional[str] = "M"    # "S" | "M" | "L" — untuk kolom TTD yang beda-beda ukuran
    # Iter 40 — Posisi berbeda tiap halaman: list [{page,x,y,size?}] (page -1 = semua halaman)
    placements: Optional[List[dict]] = None
    # Iter 20d — Data SO Stamp Produksi (dipakai saat Sales TTD, auto-terisi ke Salma)
    so_stamp_data: Optional[dict] = None


def _norm_placements(raw) -> list:
    """Normalisasi list placement stamp → [{page:int, x:float, y:float, size:str}].
    Abaikan entri tanpa x/y. page default 0."""
    out = []
    for pl in (raw or []):
        if not isinstance(pl, dict):
            continue
        x = pl.get("x"); y = pl.get("y")
        if x is None or y is None:
            continue
        try:
            page = int(pl.get("page")) if pl.get("page") is not None else 0
        except (TypeError, ValueError):
            page = 0
        out.append({
            "page": page,
            "x": float(x),
            "y": float(y),
            "size": (str(pl.get("size")).upper() if pl.get("size") else "M"),
        })
    return out


def _apply_placement_to_stamp(stamp: dict, payload) -> None:
    """Terapkan info posisi dari payload ke dict stamp.
    Prioritas: payload.placements (per-halaman). Fallback: stamp_x/stamp_y/stamp_page legacy.
    Selalu isi x/y/page/size legacy dari placement pertama agar riwayat & backward-compat tetap jalan.
    """
    pls = _norm_placements(getattr(payload, "placements", None)) if payload else []
    if pls:
        stamp["placements"] = pls
        first = pls[0]
        stamp["x"] = first["x"]
        stamp["y"] = first["y"]
        stamp["page"] = first["page"]
        stamp["size"] = first["size"]
        return
    if payload:
        if payload.stamp_x is not None: stamp["x"] = float(payload.stamp_x)
        if payload.stamp_y is not None: stamp["y"] = float(payload.stamp_y)
        if payload.stamp_page is not None: stamp["page"] = int(payload.stamp_page)
        if payload.stamp_size: stamp["size"] = str(payload.stamp_size).upper()


@router.post("/drawings/{drawing_id}/sign-prepared")
async def drawing_sign_prepared(
    drawing_id: str,
    payload: ApprovalActionIn = None,
    current: dict = Depends(get_current_user),
):
    """Engineer TTD 'Prepared By' pada posisi terpilih lalu SIMPAN (status tetap draft).
    Submit ke Eng Leader dilakukan terpisah dari halaman Work Group (boleh partial)."""
    drawing = await db.drawings.find_one({"id": drawing_id})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    status = drawing.get("approval_status", "draft")
    if status != "draft":
        raise HTTPException(status_code=409, detail=f"Drawing sudah '{status}', tidak bisa TTD Prepared ulang")
    if not drawing.get("file_id"):
        raise HTTPException(status_code=400, detail="Upload PDF drawing dulu sebelum TTD")
    # Kategori pekerjaan TIDAK wajib untuk TTD Prepared (boleh TTD langsung setelah upload).
    # Kategori tetap divalidasi saat SUBMIT ke Eng Leader.
    if not is_engineering(current) and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Engineering yang boleh TTD Prepared")
    if not _can_modify_drawing(current, drawing):
        raise HTTPException(status_code=403, detail=f"Drawing ini di-assign ke {drawing.get('assigned_to_name','-')}.")
    stamp = _sig_stamp(current, notes=(payload.notes if payload else ""))
    stamp["stage"] = "submit"
    _apply_placement_to_stamp(stamp, payload)
    if not stamp.get("placements") and stamp.get("x") is None:
        raise HTTPException(status_code=400, detail="Pilih posisi TTD di PDF dulu")
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {
            "prepared_signature": stamp,
            "prepared_signed": True,
            "prepared_by": stamp["name"],
            "prepared_signed_at": _now_iso(),
        }},
    )
    await log_action(current, "drawing_sign_prepared", "drawings", drawing_id, {"drawing_no": drawing.get("drawing_no")})
    return {"success": True, "prepared_signed": True, "signed_by": stamp["name"], "approval_status": "draft"}


@router.post("/drawings/{drawing_id}/submit-for-approval")
async def drawing_submit_for_approval(
    drawing_id: str,
    payload: ApprovalActionIn = None,
    current: dict = Depends(get_current_user),
):
    """Engineer submit drawing draft → pending_eng_head. Approval flow dimulai.

    Engineer (Prepared By) juga langsung TTD di PDF drawing pada posisi yang dipilih.
    """
    drawing = await db.drawings.find_one({"id": drawing_id})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    status = drawing.get("approval_status", "draft")
    if status != "draft":
        raise HTTPException(status_code=409, detail=f"Drawing sudah dalam status '{status}', tidak bisa submit ulang")
    if not drawing.get("file_id"):
        raise HTTPException(status_code=400, detail="Upload PDF drawing terlebih dahulu sebelum submit approval")
    if (drawing.get("work_category") or "").strip().lower() not in ("simple", "moderate", "complex"):
        raise HTTPException(status_code=400, detail="Pilih dulu Kategori Pekerjaan (SIMPLE / MODERATE / COMPLEX) sebelum submit ke Eng Leader.")
    if not is_engineering(current) and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Engineering yang boleh submit drawing untuk approval")
    if not _can_modify_drawing(current, drawing):
        assigned_name = drawing.get("assigned_to_name", "-")
        raise HTTPException(status_code=403, detail=f"Drawing ini di-assign ke {assigned_name}. Hanya orang tersebut yang bisa submit.")

    stamp = _sig_stamp(current, notes=(payload.notes if payload else ""))
    stamp["stage"] = "submit"
    # Iter 22/40 — Prepared By TTD digital di posisi terpilih (dukung per-halaman placements)
    _apply_placement_to_stamp(stamp, payload)
    # Alur baru: TTD Prepared By disimpan lebih dulu (endpoint sign-prepared) lalu submit
    # dari Work Group (bisa partial). Bila submit tidak mengirim posisi, pakai posisi TTD
    # yang sudah disimpan saat Prepared By.
    if not stamp.get("placements") and stamp.get("x") is None:
        prev = drawing.get("prepared_signature") or {}
        if prev.get("placements"):
            stamp["placements"] = prev["placements"]
        for _k in ("x", "y", "page", "size"):
            if prev.get(_k) is not None:
                stamp[_k] = prev[_k]
        if prev.get("notes") and not stamp.get("notes"):
            stamp["notes"] = prev["notes"]

    submitter_is_leader = is_eng_head(current)

    # CATATAN: Submit drawing TIDAK lagi otomatis menyeret BOM.
    # BOM di-submit terpisah dari panel "Submit ke Eng Leader" (checkbox BOM sendiri),
    # setelah Engineer Staff TTD Prepared + Tandai Siap Submit. Lihat bom.py /submit-review.

    # Bila yang mengerjakan/submit adalah Engineering LEADER sendiri → verifikasi Leader
    # otomatis lolos (tidak perlu approval Leader lagi). Langsung ke stage QC.
    if submitter_is_leader:
        auto_check = _sig_stamp(current, notes="Auto-verifikasi: drawing dikerjakan oleh Engineering Leader")
        auto_check["stage"] = "eng_head"
        auto_check["auto"] = True
        _apply_placement_to_stamp(auto_check, payload)
        await db.drawings.update_one(
            {"id": drawing_id},
            {"$set": {"approval_status": "pending_qc", "submitted_at": _now_iso(),
                      "submitted_by": stamp["name"], "prepared_by": stamp["name"],
                      "checked_by": stamp["name"], "eng_head_at": _now_iso()},
             "$push": {"approvals": {"$each": [stamp, auto_check]}}},
        )
        await log_action(current, "drawing_submit_approval", "drawings", drawing_id,
                         {"drawing_no": drawing.get("drawing_no"), "auto_leader_verified": True})
        return {"success": True, "approval_status": "pending_qc", "signed_by": stamp["name"],
                "auto_leader_verified": True}

    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {"approval_status": "pending_eng_head", "submitted_at": _now_iso(),
                  "submitted_by": stamp["name"], "prepared_by": stamp["name"]},
         "$push": {"approvals": stamp}},
    )
    await log_action(current, "drawing_submit_approval", "drawings", drawing_id,
                     {"drawing_no": drawing.get("drawing_no")})
    return {"success": True, "approval_status": "pending_eng_head", "signed_by": stamp["name"]}


@router.post("/drawings/{drawing_id}/approve/{stage}")
async def drawing_approve_stage(
    drawing_id: str, stage: str,
    payload: ApprovalActionIn = None,
    current: dict = Depends(get_current_user),
):
    """Approve stage. Sequential: stage harus sesuai approval_status saat ini.

    Stages valid: eng_head, qc, sales
    """
    if stage not in STAGE_ORDER:
        raise HTTPException(status_code=400, detail=f"Stage '{stage}' tidak valid. Pilih: {STAGE_ORDER}")

    drawing = await db.drawings.find_one({"id": drawing_id})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")

    current_status = drawing.get("approval_status", "draft")
    expected_status = STAGE_STATUS[stage]
    # Sequential enforcement: current_status harus == expected for this stage
    if current_status != expected_status:
        raise HTTPException(
            status_code=409,
            detail=(f"Drawing sedang di status '{current_status}', tidak bisa approve stage '{stage}'. "
                    f"Stage {stage} hanya bisa saat status '{expected_status}'."),
        )

    # Role check — hanya role yang tepat yang boleh TTD stage ini.
    # Super_admin (Susanto) diperbolehkan sebagai emergency override, tapi admin/supervisor
    # biasa TIDAK boleh bypass — mereka juga harus punya role QC/Sales/Eng Head jika mau TTD.
    allowed = _stage_allowed_roles(stage)
    role = current.get("role")
    if role not in allowed and not is_super_admin_user(current):
        raise HTTPException(
            status_code=403,
            detail=f"Role Anda ({role}) tidak boleh approve stage '{stage}'. Butuh role: {allowed}",
        )

    # Iter 22 — Untuk stage 'sales', hanya Sales requester (yg buat DRF) yang boleh TTD.
    # Sales LAIN tidak boleh TTD atas nama sales lain. Super admin bypass.
    if stage == "sales" and role in SALES_ROLES and not is_super_admin_user(current):
        drf_id = drawing.get("from_drf_id")
        my_id = current.get("id")
        my_username = (current.get("username") or "").lower().strip()
        my_name = (current.get("name") or "").lower().strip()
        is_requester = False
        if drf_id:
            drf = await db.drawing_requests.find_one(
                {"id": drf_id}, {"_id": 0, "requested_by": 1}
            )
            rb = (drf or {}).get("requested_by") or {}
            if rb.get("user_id") == my_id:
                is_requester = True
            elif (rb.get("username") or "").lower().strip() == my_username:
                is_requester = True
            elif (rb.get("name") or "").lower().strip() == my_name:
                is_requester = True
        else:
            rbs = (drawing.get("request_by_sales") or "").lower().strip()
            if rbs and (rbs == my_name or rbs == my_username):
                is_requester = True
        if not is_requester:
            requester_name = (drf and (drf.get("requested_by") or {}).get("name")) or drawing.get("request_by_sales") or "Sales lain"
            raise HTTPException(
                status_code=403,
                detail=(f"Anda tidak berhak TTD drawing ini. Drawing ini di-request oleh "
                        f"'{requester_name}' — hanya dia (atau super admin) yang boleh TTD stage Sales."),
            )

    stamp = _sig_stamp(current, notes=(payload.notes if payload else ""))
    stamp["stage"] = stage
    # Iter 18/40 — simpan posisi stamp digital (dukung per-halaman placements) untuk pdf-stamper
    _apply_placement_to_stamp(stamp, payload)
    next_status = NEXT_STATUS_AFTER[current_status]
    update = {
        "approval_status": next_status,
        "updated_at": _now_iso(),
    }
    if next_status == "approved":
        update["approved_at"] = _now_iso()
    # Iter — "Tanggal Selesai Kerja": saat Eng Leader (eng_head) approve hasil kerja engineer.
    # Dianggap SELESAI walaupun QC/Sales belum TTD.
    if stage == "eng_head":
        update["work_completed_at"] = _now_iso()
    # Sales stage: AUTO-ISI so_stamp_draft dari data DRF/drawing (Sales tak perlu ketik ulang).
    # Feature F: po_no diambil dari No. PO Customer (DRF). Feature G: customer = KODE customer (Customer Code Master).
    if stage == "sales":
        _sd = (payload.so_stamp_data if (payload and payload.so_stamp_data) else {}) or {}
        # Ambil qty & tanggal dari Sales Order bila ada
        so_ref = {}
        try:
            if drawing.get("so_no"):
                so_ref = await db.sales_orders.find_one({"so_no": drawing.get("so_no")}, {"_id": 0}) or {}
        except Exception:
            so_ref = {}
        cust_code = (drawing.get("customer_code") or "").strip()
        po_cust = (drawing.get("po_customer_no") or "").strip()
        update["so_stamp_draft"] = {
            "so_no": (_sd.get("so_no") or drawing.get("so_no") or "").strip(),
            "po_no": (_sd.get("po_no") or po_cust or "").strip(),
            "qty": (_sd.get("qty") or (str(drawing.get("item_qty")).rstrip("0").rstrip(".") if drawing.get("item_qty") else "") or str(so_ref.get("qty") or so_ref.get("qty_order") or "")).strip(),
            # Feature G — pakai KODE customer (bukan nama lengkap) untuk kerahasiaan
            "customer": (_sd.get("customer") or cust_code or "").strip(),
            "received_date": (_sd.get("received_date") or (so_ref.get("so_date") or "")).strip(),
            "due_date": (_sd.get("due_date") or (so_ref.get("due_date") or "")).strip(),
            "filled_by": current.get("name") or current.get("username"),
            "filled_at": _now_iso(),
        }
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": update, "$push": {"approvals": stamp}},
    )
    # Iter 19 — Update linked DRF status kalau drawing linked ke DRF
    if next_status == "approved" and drawing.get("from_drf_id"):
        await db.drawing_requests.update_one(
            {"id": drawing["from_drf_id"], "deleted_at": {"$exists": False}},
            {"$set": {"status": "completed", "completed_at": _now_iso(), "updated_at": _now_iso()}},
        )
    await log_action(current, f"drawing_approve_{stage}", "drawings", drawing_id,
                     {"drawing_no": drawing.get("drawing_no"), "next_status": next_status})
    return {"success": True, "approval_status": next_status, "stage": stage, "signed_by": stamp["name"]}


@router.post("/drawings/{drawing_id}/reject/{stage}")
async def drawing_reject_stage(
    drawing_id: str, stage: str,
    payload: ApprovalActionIn,
    current: dict = Depends(get_current_user),
):
    """Reject a stage — kembalikan drawing ke status 'draft' untuk revisi. Wajib notes."""
    if stage not in STAGE_ORDER:
        raise HTTPException(status_code=400, detail=f"Stage '{stage}' tidak valid")
    if not payload.notes or len(payload.notes.strip()) < 5:
        raise HTTPException(status_code=400, detail="Reject wajib menyertakan notes/alasan (min 5 karakter)")

    drawing = await db.drawings.find_one({"id": drawing_id})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")

    current_status = drawing.get("approval_status", "draft")
    expected_status = STAGE_STATUS[stage]
    if current_status != expected_status:
        raise HTTPException(status_code=409, detail=f"Tidak bisa reject stage '{stage}' — status saat ini '{current_status}'")

    allowed = _stage_allowed_roles(stage)
    role = current.get("role")
    if role not in allowed and not is_super_admin_user(current):
        raise HTTPException(status_code=403, detail=f"Role Anda tidak boleh reject stage '{stage}'")

    stamp = _sig_stamp(current, notes=payload.notes)
    stamp["stage"] = f"reject_{stage}"
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {"approval_status": "draft", "rejected_at": _now_iso(),
                  "rejected_stage": stage, "reject_notes": payload.notes.strip()},
         "$push": {"approvals": stamp}},
    )
    await log_action(current, f"drawing_reject_{stage}", "drawings", drawing_id,
                     {"drawing_no": drawing.get("drawing_no"), "notes": payload.notes})
    return {"success": True, "approval_status": "draft", "rejected_stage": stage}


# ============ ECN — Engineering Change Notice (MKS-F-ENG-004) ============
# Eng staff mengajukan revisi drawing yang SUDAH masuk approval/approved via form ECN.
# Eng Leader (Head of Dept) menyetujui → drawing dikembalikan ke 'draft' untuk direvisi.

class EcnIn(BaseModel):
    ecr_no: str = ""
    scope: str = "both"                # drawing | bom | both — bagian yang akan direvisi
    m4: List[str] = []                 # MAN / MACHINE / METHOD / MATERIAL
    item_of_change: List[str] = []     # process/materials/inspection/subcon/design_spec/packing/other
    item_other: str = ""
    change_type: str = "permanent"     # temporary | permanent
    expired_date: str = ""
    current_desc: str
    proposed_desc: str
    purpose: List[str] = []            # customer_request/customer_complaint/quality_improvement/others
    purpose_other: str = ""
    purpose_explanation: str = ""
    effective_date: str = ""
    affected_document: List[str] = []  # drawing_spec / sop_wip / others
    affected_other: str = ""


async def _next_ecn_no() -> str:
    now = datetime.now(timezone.utc)
    prefix = f"ECN-{now.strftime('%y-%m')}"
    cnt = await db.drawings.count_documents({"revision_request.ecn.ecn_no": {"$regex": f"^{prefix}"}})
    return f"{prefix}-{cnt + 1:02d}"


def _is_drawing_owner(d: dict, user: dict) -> bool:
    """True jika `user` adalah engineer yang menggambar drawing ini (designer/assignee).
    Dipakai untuk membatasi siapa yang boleh mengajukan/menjalankan revisi ECN."""
    if not d or not user:
        return False
    uid = user.get("id")
    uname = user.get("name")
    uuser = user.get("username")
    candidates = {
        d.get("assigned_to_user_id"),
        d.get("prepared_by"),
        d.get("assigned_to_name"),
        d.get("created_by_id"),
        d.get("created_by"),
    }
    candidates.discard(None)
    return bool(
        (uid and uid in candidates) or
        (uname and uname in candidates) or
        (uuser and uuser in candidates)
    )


@router.post("/drawings/{drawing_id}/request-revision")
async def request_drawing_revision(drawing_id: str, payload: EcnIn, current: dict = Depends(get_current_user)):
    """Eng staff ajukan revisi (form ECN) untuk drawing yang sudah tidak draft.
    Hanya engineer yang menggambar drawing ini (designer/assignee) yang boleh — kecuali Admin."""
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering/Admin")
    d = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not d:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not _is_drawing_owner(d, current) and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya engineer yang menggambar drawing ini yang boleh mengajukan revisi.")
    status = d.get("approval_status") or "draft"
    if status == "draft":
        raise HTTPException(status_code=400, detail="Drawing masih DRAFT — langsung edit di Work Order, tidak perlu ECN.")
    if (d.get("revision_request") or {}).get("status") == "pending":
        raise HTTPException(status_code=400, detail="Sudah ada pengajuan ECN yang menunggu keputusan Eng Leader.")
    if not (payload.current_desc or "").strip() or not (payload.proposed_desc or "").strip():
        raise HTTPException(status_code=400, detail="Kolom 'Current' dan 'Proposed' wajib diisi.")
    now = _now_iso()
    who = current.get("name") or current.get("username")
    ecn_no = await _next_ecn_no()
    ecn = payload.model_dump()
    ecn.update({
        "ecn_no": ecn_no,
        "to": "Production & QA/QC",
        "originator": who,
        "date_initiated": now,
        "so_no": d.get("so_no"),
        "customer": d.get("customer_name"),
        "part_dwg_no": d.get("drawing_no"),
        "prepared_by": who,
        "prepared_at": now,
        "approvals": {},  # diisi saat keputusan: head_of_dept
    })
    rr = {
        "status": "pending",
        "type": "ecn",
        "scope": (payload.scope or "both") if (payload.scope in ("drawing", "bom", "both")) else "both",
        "ecn": ecn,
        "reason": (payload.purpose_explanation or payload.proposed_desc or "").strip(),
        "requested_by": who,
        "requested_by_id": current.get("id"),
        "requested_at": now,
        "prev_status": status,
    }
    await db.drawings.update_one({"id": drawing_id}, {"$set": {"revision_request": rr, "updated_at": now}})
    await log_action(current, "drawing_request_revision_ecn", "drawings", drawing_id, {"ecn_no": ecn_no})
    return {"success": True, "ecn_no": ecn_no, "revision_request": rr}


class RevisionDecisionIn(BaseModel):
    approve: bool
    notes: str = ""


@router.post("/drawings/{drawing_id}/revision-decision")
async def decide_drawing_revision(drawing_id: str, payload: RevisionDecisionIn, current: dict = Depends(get_current_user)):
    """Eng Leader (Head of Dept) / Admin memutuskan pengajuan ECN.
    APPROVE → drawing kembali ke 'draft' agar bisa direvisi & submit ulang.
    REJECT  → status drawing tidak berubah, ECN ditolak."""
    if not (is_eng_head(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Eng Leader / Admin yang bisa memutuskan ECN")
    d = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not d:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    rr = d.get("revision_request") or {}
    if rr.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Tidak ada pengajuan ECN yang menunggu keputusan")
    now = _now_iso()
    who = current.get("name") or current.get("username")
    rr["decided_by"] = who
    rr["decided_at"] = now
    rr["decision_notes"] = (payload.notes or "").strip()
    ecn = rr.get("ecn") or {}
    ecn.setdefault("approvals", {})["head_of_dept"] = {"name": who, "user_id": current.get("id"), "at": now, "notes": (payload.notes or "").strip()}
    rr["ecn"] = ecn
    if payload.approve:
        # ECN disetujui — TIDAK langsung buka draft. Drawing menunggu staff klik "Lanjut Kerja"
        # (start-revision) di Work Order. approval_status TIDAK diubah sampai revisi dimulai.
        rr["status"] = "approved"
        rr["approved_by"] = who
        rr["approved_at"] = now
        upd = {
            "revision_request": rr,
            "revision_reason": rr.get("reason", ""),
            "revision_approved_at": now,
            "updated_at": now,
        }
    else:
        rr["status"] = "rejected"
        upd = {"revision_request": rr, "updated_at": now}
    await db.drawings.update_one({"id": drawing_id}, {"$set": upd})
    await log_action(current, "drawing_revision_decision", "drawings", drawing_id,
                     {"approve": payload.approve, "ecn_no": ecn.get("ecn_no")})
    return {"success": True, "approved": payload.approve,
            "revision_status": rr["status"], "approval_status": d.get("approval_status")}


def _file_in_revisions(drawing: dict, file_id: str) -> bool:
    """True jika file_id masih direferensikan oleh snapshot history revisi
    (agar file lama TIDAK ikut terhapus saat upload MKS revisi baru)."""
    if not file_id:
        return False
    for rev in (drawing.get("revisions") or []):
        snap = rev.get("snapshot") or {}
        if snap.get("file_id") == file_id or snap.get("customer_ref_file_id") == file_id:
            return True
    return False


@router.post("/drawings/{drawing_id}/start-revision")
async def drawing_start_revision(drawing_id: str, current: dict = Depends(get_current_user)):
    """Staff klik 'Lanjut Kerja' setelah ECN disetujui Eng Leader.
    - Snapshot data lama (PDF MKS, no drawing, rev_no, semua TTD, customer ref) → simpan ke `revisions` (history, TIDAK dihapus).
    - Naikkan rev_no, reset TTD/approvals untuk siklus baru, buka kembali ke 'draft' agar semua menu terbuka."""
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Hanya Engineering/Admin")
    d = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not d:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if not _is_drawing_owner(d, current) and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya engineer yang menggambar drawing ini yang boleh memulai revisi.")
    rr = d.get("revision_request") or {}
    if rr.get("status") != "approved":
        raise HTTPException(status_code=400, detail="Tidak ada revisi yang disetujui & siap dimulai")
    now = _now_iso()
    who = current.get("name") or current.get("username")
    old_rev = int(d.get("rev_no") or 0)
    ecn = rr.get("ecn") or {}

    # Snapshot data lama sebagai history revisi (file lama dipertahankan di storage)
    snapshot = {
        "rev_no": old_rev,
        "drawing_no": d.get("drawing_no"),
        "approval_status": d.get("approval_status"),
        "file_id": d.get("file_id"),
        "filename": d.get("filename"),
        "customer_ref_file_id": d.get("customer_ref_file_id"),
        "customer_ref_filename": d.get("customer_ref_filename"),
        "approvals": d.get("approvals") or [],
        "controlled_at": d.get("controlled_at"),
    }
    rev_entry = {
        "id": str(uuid.uuid4()),
        "type": "ecn_revision",
        "ecn_no": ecn.get("ecn_no"),
        "rev_no": old_rev,
        "started_by": who,
        "started_by_id": current.get("id"),
        "at": now,
        "reason": rr.get("reason", ""),
        "ecn": ecn,
        "snapshot": snapshot,
        "files": [],
    }

    rr["status"] = "in_progress"
    rr["started_by"] = who
    rr["started_at"] = now
    scope = rr.get("scope") or "both"
    do_drawing = scope in ("drawing", "both")
    do_bom = scope in ("bom", "both")

    # ── Revisi BOM (bila scope mencakup BOM): snapshot BOM lama + buka kembali ──
    bom_result = None
    if do_bom and d.get("bom_id"):
        try:
            bom = await db.boms.find_one({"id": d["bom_id"], "deleted_at": {"$exists": False}})
            if bom:
                b_old = int(bom.get("rev_no") or 0)
                _atts = await db.bom_attachments.find(
                    {"bom_id": bom["id"], "deleted_at": {"$exists": False}},
                    {"_id": 0, "id": 1, "category": 1, "filename": 1, "uploaded_at": 1},
                ).to_list(300)
                bsnap = {
                    "rev_no": b_old, "bom_no": bom.get("bom_no"),
                    "engineering_status": bom.get("engineering_status"),
                    "items": bom.get("items") or [], "signatures": bom.get("signatures") or {},
                    "procurement_status": bom.get("procurement_status"),
                    "procurement_signatures": bom.get("procurement_signatures") or {},
                    "attachments": _atts,
                    "snapshot_at": now, "ecn_no": ecn.get("ecn_no"),
                }
                await db.boms.update_one({"id": bom["id"]}, {
                    "$set": {
                        "rev_no": b_old + 1, "engineering_status": "draft",
                        "signatures": {}, "procurement_status": None,
                        "procurement_signatures": {}, "is_revision": True,
                        "revision_reason": rr.get("reason", ""), "updated_at": now,
                    },
                    "$push": {"revisions": bsnap},
                })
                bom_result = {"bom_no": bom.get("bom_no"), "new_rev_no": b_old + 1}
        except Exception:
            pass

    if do_drawing:
        upd = {
            "revision_request": rr,
            "rev_no": old_rev + 1,
            "is_revision": True,
            "approval_status": "draft",
            "approvals": [],  # TTD siklus baru mulai dari kosong (TTD lama tersimpan di snapshot)
            "revision_opened_at": now,
            "revision_reason": rr.get("reason", ""),
            # reset flag controlled agar siklus baru bersih; timestamp lama tersimpan di snapshot
            "controlled_at": None,
            "updated_at": now,
        }
        await db.drawings.update_one({"id": drawing_id},
                                     {"$set": upd, "$push": {"revisions": rev_entry}})
        new_rev = old_rev + 1
    else:
        # BOM-only: drawing TIDAK direset. Catat riwayat & tutup pengajuan.
        rev_entry["scope"] = scope
        await db.drawings.update_one({"id": drawing_id},
                                     {"$set": {"revision_request": rr, "updated_at": now},
                                      "$push": {"revisions": rev_entry}})
        new_rev = old_rev

    await log_action(current, "drawing_start_revision", "drawings", drawing_id,
                     {"ecn_no": ecn.get("ecn_no"), "scope": scope, "new_rev_no": new_rev, "bom": bom_result})
    return {"success": True, "scope": scope, "rev_no": new_rev,
            "approval_status": "draft" if do_drawing else d.get("approval_status"),
            "bom_revised": bom_result}


def _is_production(user: dict) -> bool:
    return (user or {}).get("role") in PRODUCTION_ROLES


@router.get("/drawings/{drawing_id}/ecn-ack-state")
async def ecn_ack_state(drawing_id: str, current: dict = Depends(get_current_user)):
    """State rantai acknowledgment ECN (Produksi -> QA/QC -> Doc Control) untuk sebuah drawing."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    d = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}},
                                   {"_id": 0, "revision_request": 1, "approval_status": 1})
    if not d:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    rr = d.get("revision_request") or {}
    ecn = rr.get("ecn") or {}
    ack = rr.get("ack") or {"stage": "production", "production": None, "qa_qc": None, "doc_control": None}
    is_ifu = d.get("approval_status") in ("controlled", "released")
    available = bool(ecn.get("ecn_no")) and is_ifu
    role = current.get("role")
    can_ack = False
    if available and ack.get("stage") == "production":
        can_ack = _is_production(current) or is_admin_like(current)
    elif available and ack.get("stage") == "qa_qc":
        can_ack = is_qc(current) or is_admin_like(current)
    return {
        "available": available,
        "is_ifu": is_ifu,
        "ecn_no": ecn.get("ecn_no"),
        "ack": ack,
        "stage": ack.get("stage") if available else None,
        "can_ack": can_ack,
        "my_role": role,
    }


@router.post("/drawings/{drawing_id}/ecn-ack")
async def ecn_acknowledge(drawing_id: str, current: dict = Depends(get_current_user)):
    """TTD digital acknowledgment ECN — berurutan: Produksi -> QA/QC -> (otomatis) Doc Control.
    Hanya boleh setelah drawing revisi TERBIT (IFU = controlled/released)."""
    d = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not d:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    rr = d.get("revision_request") or {}
    ecn = rr.get("ecn") or {}
    if not ecn.get("ecn_no"):
        raise HTTPException(status_code=400, detail="Tidak ada ECN aktif pada drawing ini")
    if d.get("approval_status") not in ("controlled", "released"):
        raise HTTPException(status_code=400, detail="Menunggu drawing revisi TERBIT (IFU) sebelum acknowledge.")
    ack = rr.get("ack") or {"stage": "production", "production": None, "qa_qc": None, "doc_control": None}
    stage = ack.get("stage", "production")
    now = _now_iso()

    if stage == "production":
        if not (_is_production(current) or is_admin_like(current)):
            raise HTTPException(status_code=403, detail="Tahap ini menunggu acknowledge dari PRODUKSI.")
        ack["production"] = _sig_stamp(current)
        ack["stage"] = "qa_qc"
        msg = "Produksi acknowledge tercatat. Lanjut ke QA/QC."
    elif stage == "qa_qc":
        if not (is_qc(current) or is_admin_like(current)):
            raise HTTPException(status_code=403, detail="Tahap ini menunggu tanda tangan QA/QC.")
        ack["qa_qc"] = _sig_stamp(current)
        # Otomatis diteruskan ke Doc Control (arsip)
        ack["doc_control"] = {"at": now, "auto": True, "note": "Otomatis diarsipkan ke Document Control setelah QA/QC TTD"}
        ack["stage"] = "done"
        msg = "QA/QC TTD tercatat. ECN otomatis diteruskan ke Document Control."
    elif stage == "done":
        raise HTTPException(status_code=400, detail="Acknowledgment ECN sudah selesai.")
    else:
        raise HTTPException(status_code=400, detail=f"Stage tidak dikenal: {stage}")

    rr["ack"] = ack
    await db.drawings.update_one({"id": drawing_id}, {"$set": {"revision_request": rr, "updated_at": now}})
    await log_action(current, "ecn_acknowledge", "drawings", drawing_id,
                     {"ecn_no": ecn.get("ecn_no"), "stage": stage})
    return {"success": True, "stage": ack["stage"], "ack": ack, "message": msg}


@router.get("/drawings/ecn-pending-ttd")
async def ecn_pending_ttd(current: dict = Depends(get_current_user)):
    """Daftar ECN yang menunggu TTD user saat ini (berdasarkan role: Produksi / QA-QC).
    Muncul di kartu TTD portal masing-masing dept — tidak butuh akses menu Engineering."""
    role = current.get("role")
    prod = _is_production(current) or is_admin_like(current)
    qc = is_qc(current) or is_admin_like(current)
    if not (prod or qc):
        raise HTTPException(status_code=403, detail="Hanya Produksi/QA-QC")
    cursor = db.drawings.find(
        {"approval_status": {"$in": ["controlled", "released"]},
         "revision_request.ecn.ecn_no": {"$exists": True},
         "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "drawing_no": 1, "so_no": 1, "customer_name": 1,
         "approval_status": 1, "rev_no": 1, "revision_request": 1},
    )
    items = []
    async for d in cursor:
        rr = d.get("revision_request") or {}
        ecn = rr.get("ecn") or {}
        ack = rr.get("ack") or {"stage": "production"}
        stage = ack.get("stage", "production")
        if stage == "done":
            continue
        # tentukan apakah user ini yang ditunggu di stage sekarang
        mine = (stage == "production" and prod) or (stage == "qa_qc" and qc)
        if not mine:
            continue
        items.append({
            "drawing_id": d.get("id"),
            "drawing_no": d.get("drawing_no"),
            "so_no": d.get("so_no"),
            "customer": d.get("customer_name"),
            "rev_no": d.get("rev_no"),
            "ecn_no": ecn.get("ecn_no"),
            "stage": stage,
            "stage_label": "Acknowledge Produksi" if stage == "production" else "Tanda Tangan QA/QC",
            "reason": (ecn.get("purpose_explanation") or ecn.get("proposed_desc") or "").strip(),
            "current_desc": ecn.get("current_desc", ""),
            "proposed_desc": ecn.get("proposed_desc", ""),
            "requested_by": rr.get("requested_by"),
            "production_done": bool(ack.get("production")),
        })
    items.sort(key=lambda x: (x.get("ecn_no") or ""))
    return {"items": items, "total": len(items)}


@router.get("/drawings/eng-designers")
async def list_eng_designers(current: dict = Depends(get_current_user)):
    """Daftar user Engineering untuk filter 'Designer' di Master List. Akses: Engineering/Admin."""
    if not (is_engineering(current) or is_admin_like(current)):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    cur = db.users.find(
        {"role": {"$in": list(ENGINEERING_ROLES)}},
        {"_id": 0, "id": 1, "username": 1, "name": 1, "role": 1},
    )
    users = await cur.to_list(200)
    users.sort(key=lambda u: (u.get("name") or u.get("username") or "").lower())
    return {"designers": users}


@router.get("/ecn-register")
async def ecn_register(q: Optional[str] = None, kind: Optional[str] = None,
                       status: Optional[str] = None,
                       date_from: Optional[str] = None, date_to: Optional[str] = None,
                       current: dict = Depends(get_current_user)):
    """Master List ECN & ECR (read-only, record).
    Mengagregasi seluruh ECN dari alur revisi drawing (revision_request + history `revisions`)
    dan entri ECR/ECN lama dari koleksi `ecns`. Drawing hasil revisi tetap di Master List utama."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")

    rows: List[dict] = []

    # 1) ECN dari alur revisi drawing
    cursor = db.drawings.find(
        {"$or": [{"revision_request": {"$exists": True}}, {"revisions.type": "ecn_revision"}],
         "deleted_at": {"$exists": False}},
        {"_id": 0, "id": 1, "drawing_no": 1, "so_no": 1, "customer_name": 1,
         "revision_request": 1, "revisions": 1},
    )
    async for d in cursor:
        by_no: dict = {}

        def _mk(ecn: dict, status: str, at: str, requested_by: str, rev_no, ack: dict = None, dates: dict = None):
            no = (ecn or {}).get("ecn_no")
            if not no:
                return
            dts = dates or {}
            ackd = ack or {}
            by_no[no] = {
                "no": no,
                "kind": "ecn",
                "drawing_id": d.get("id"),
                "drawing_no": d.get("drawing_no"),
                "so_no": d.get("so_no"),
                "customer": d.get("customer_name"),
                "reason": (ecn.get("purpose_explanation") or ecn.get("proposed_desc") or "").strip(),
                "current_desc": ecn.get("current_desc", ""),
                "proposed_desc": ecn.get("proposed_desc", ""),
                "requested_by": requested_by,
                "status": status,
                "rev_no": rev_no,
                "at": at,
                "source": "drawing_revision",
                "ack_stage": ackd.get("stage"),
                "ack_production": bool(ackd.get("production")),
                "ack_qa_qc": bool(ackd.get("qa_qc")),
                "ack_doc_control": bool(ackd.get("doc_control")),
                # Timeline: registrasi -> mulai kerja -> selesai (IFU) -> stamp Doc Control (distribusi)
                "date_reg": dts.get("reg"),
                "date_start": dts.get("start"),
                "date_done": dts.get("done"),
                "date_doco": dts.get("doco"),
            }

        # history dulu
        for rev in (d.get("revisions") or []):
            if rev.get("type") == "ecn_revision":
                snap = rev.get("snapshot") or {}
                _mk(rev.get("ecn") or {}, "completed", rev.get("at"),
                    rev.get("started_by") or "", rev.get("rev_no"),
                    dates={"start": rev.get("at"), "done": snap.get("controlled_at")})
        # lalu revision_request aktif (override status terbaru per ecn_no)
        rr = d.get("revision_request") or {}
        if rr and (rr.get("ecn") or {}).get("ecn_no"):
            ackd = rr.get("ack") or {}
            _mk(rr.get("ecn") or {}, rr.get("status") or "pending",
                rr.get("requested_at"), rr.get("requested_by") or "", None, ackd,
                dates={
                    "reg": rr.get("requested_at"),
                    "start": rr.get("started_at"),
                    "done": rr.get("completed_at"),
                    "doco": (ackd.get("doc_control") or {}).get("at"),
                })
        rows.extend(by_no.values())

    # 2) ECR/ECN lama dari koleksi ecns
    async for e in db.ecns.find({}, {"_id": 0}):
        rows.append({
            "no": e.get("ecn_no"),
            "kind": e.get("kind") or "ecn",
            "drawing_id": None,
            "drawing_no": e.get("drawing_no") or e.get("bom_no"),
            "so_no": e.get("so_no"),
            "customer": e.get("customer_name"),
            "reason": e.get("reason", ""),
            "current_desc": "",
            "proposed_desc": e.get("description", ""),
            "requested_by": (e.get("requested_by") or {}).get("name") if isinstance(e.get("requested_by"), dict) else e.get("requested_by"),
            "status": e.get("status") or "draft",
            "rev_no": None,
            "at": e.get("created_at") or e.get("submitted_at"),
            "source": "manual",
        })

    # filter
    if kind in ("ecn", "ecr"):
        rows = [r for r in rows if r.get("kind") == kind]
    if status:
        rows = [r for r in rows if (r.get("status") or "") == status]
    if date_from:
        rows = [r for r in rows if (r.get("date_reg") or r.get("at") or "") >= date_from]
    if date_to:
        # inklusif sampai akhir hari
        dt_end = date_to + "T23:59:59"
        rows = [r for r in rows if (r.get("date_reg") or r.get("at") or "") <= dt_end]
    if q:
        ql = q.strip().lower()
        rows = [r for r in rows if ql in " ".join(
            str(r.get(k) or "") for k in ("no", "drawing_no", "so_no", "customer", "reason", "requested_by")
        ).lower()]

    rows.sort(key=lambda r: (r.get("at") or ""), reverse=True)
    return {"items": rows, "total": len(rows)}




def _rev_fs() -> AsyncIOMotorGridFSBucket:
    """GridFS bucket khusus file revisi (markup/koreksi dari leader saat reject)."""
    return AsyncIOMotorGridFSBucket(db, bucket_name="revision_files")


@router.post("/drawings/{drawing_id}/reject-with-files/{stage}")
async def drawing_reject_with_files(
    drawing_id: str, stage: str,
    notes: str = Form(...),
    files: List[UploadFile] = File(default=[]),
    current: dict = Depends(get_current_user),
):
    """Fase 3 — Reject bercatatan + unggah banyak file koreksi (markup).
    Drawing kembali ke 'draft'; catatan & file tersimpan di `revisions` agar staff bisa revisi & submit ulang."""
    if stage not in STAGE_ORDER:
        raise HTTPException(status_code=400, detail=f"Stage '{stage}' tidak valid")
    if not notes or len(notes.strip()) < 5:
        raise HTTPException(status_code=400, detail="Reject wajib menyertakan catatan (min 5 karakter)")
    drawing = await db.drawings.find_one({"id": drawing_id})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    current_status = drawing.get("approval_status", "draft")
    if current_status != STAGE_STATUS[stage]:
        raise HTTPException(status_code=409, detail=f"Tidak bisa reject stage '{stage}' — status saat ini '{current_status}'")
    allowed = _stage_allowed_roles(stage)
    if current.get("role") not in allowed and not is_super_admin_user(current):
        raise HTTPException(status_code=403, detail=f"Role Anda tidak boleh reject stage '{stage}'")

    # Simpan file koreksi ke GridFS
    saved = []
    fs = _rev_fs()
    for f in (files or []):
        if not f or not f.filename:
            continue
        data = await f.read()
        if not data:
            continue
        fid = await fs.upload_from_stream(f.filename, data, metadata={"content_type": f.content_type or "application/octet-stream"})
        saved.append({
            "id": str(fid),
            "filename": f.filename,
            "content_type": f.content_type or "application/octet-stream",
            "size": len(data),
            "is_pdf": (f.filename or "").lower().endswith(".pdf"),
        })

    rev_entry = {
        "id": str(uuid.uuid4()),
        "stage": stage,
        "rejected_by_id": current.get("id"),
        "rejected_by_name": current.get("name") or current.get("username"),
        "notes": notes.strip(),
        "files": saved,
        "at": _now_iso(),
        "resolved": False,
    }
    stamp = _sig_stamp(current, notes=notes.strip())
    stamp["stage"] = f"reject_{stage}"
    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {"approval_status": "draft", "rejected_at": _now_iso(),
                  "rejected_stage": stage, "reject_notes": notes.strip(), "revision_reason": notes.strip()},
         "$push": {"approvals": stamp, "revisions": rev_entry}},
    )
    await log_action(current, f"drawing_reject_{stage}", "drawings", drawing_id,
                     {"drawing_no": drawing.get("drawing_no"), "notes": notes.strip(), "files": len(saved)})
    return {"success": True, "approval_status": "draft", "rejected_stage": stage, "revision": rev_entry}


@router.get("/drawings/{drawing_id}/revisions")
async def drawing_revisions(drawing_id: str, current: dict = Depends(get_current_user)):
    """Daftar catatan revisi (reject) + file koreksi untuk sebuah drawing."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    drawing = await db.drawings.find_one({"id": drawing_id}, {"_id": 0, "revisions": 1, "drawing_no": 1})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    return {"drawing_no": drawing.get("drawing_no"), "revisions": drawing.get("revisions", [])}


async def _rev_file_bytes(drawing_id: str, file_id: str) -> tuple:
    drawing = await db.drawings.find_one({"id": drawing_id}, {"_id": 0, "revisions": 1})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    meta = None
    for rev in (drawing.get("revisions") or []):
        for f in (rev.get("files") or []):
            if f.get("id") == file_id:
                meta = f
                break
        if meta:
            break
    if not meta:
        raise HTTPException(status_code=404, detail="File revisi tidak ditemukan")
    try:
        stream = await _rev_fs().open_download_stream(ObjectId(file_id))
        raw = await stream.read()
    except Exception:
        raise HTTPException(status_code=404, detail="File revisi tidak ada di storage")
    return raw, meta


@router.get("/drawings/{drawing_id}/revision-files/{file_id}/page-meta")
async def revision_file_page_meta(drawing_id: str, file_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    from utils.pdf_render import pdf_page_meta
    raw, meta = await _rev_file_bytes(drawing_id, file_id)
    if not meta.get("is_pdf"):
        raise HTTPException(status_code=400, detail="Preview gambar hanya untuk PDF")
    return pdf_page_meta(raw)


@router.get("/drawings/{drawing_id}/revision-files/{file_id}/page-image")
async def revision_file_page_image(drawing_id: str, file_id: str, page: int = 0, scale: float = 2.0,
                                   current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    from utils.pdf_render import pdf_page_png
    raw, meta = await _rev_file_bytes(drawing_id, file_id)
    if not meta.get("is_pdf"):
        raise HTTPException(status_code=400, detail="Preview gambar hanya untuk PDF")
    try:
        png = pdf_page_png(raw, page, scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ditemukan")
    return StreamingResponse(io.BytesIO(png), media_type="image/png",
                             headers={"Cache-Control": "private, max-age=120"})


@router.get("/drawings/{drawing_id}/revision-files/{file_id}/download")
async def revision_file_download(drawing_id: str, file_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    raw, meta = await _rev_file_bytes(drawing_id, file_id)
    return StreamingResponse(io.BytesIO(raw), media_type=meta.get("content_type") or "application/octet-stream",
                             headers={"Content-Disposition": f'inline; filename="{meta.get("filename")}"'})


@router.get("/drawings/pending-my-approval")
async def list_pending_my_approval(current: dict = Depends(get_current_user)):
    """List drawings yang sekarang butuh approval dari user (sesuai role stage)."""
    role = current.get("role")
    if role in ("eng_leader", "eng_head", "engineering"):
        my_status = "pending_eng_head"
    elif role == "qc":
        my_status = "pending_qc"
    elif role in SALES_ROLES:
        my_status = "pending_sales"
    elif is_doc_control(current):
        # Doc Control (Salma) melihat semua yang sudah 'approved' (siap di-stamp)
        my_status = "approved"
    elif is_super_admin_user(current):
        # Super Admin bisa lihat semua stage (emergency override)
        my_status = {"$in": ["pending_eng_head", "pending_qc", "pending_sales", "approved"]}
    else:
        return {"items": [], "total": 0}

    filt = {"approval_status": my_status, "deleted_at": {"$exists": False}}
    docs = await db.drawings.find(filt, {"_id": 0}).sort("updated_at", -1).to_list(length=200)

    # Iter 22 — Sales hanya lihat drawing yang DIA sendiri yang buat DRF-nya (bukan Sales lain).
    # Super admin bypass. Kalau drawing tidak punya from_drf_id (drawing lama tanpa DRF),
    # fallback: bandingkan nama request_by_sales dengan nama user.
    if role in SALES_ROLES and not is_super_admin_user(current):
        my_id = current.get("id")
        my_username = (current.get("username") or "").lower().strip()
        my_name = (current.get("name") or "").lower().strip()

        # Kumpulkan DRF requester untuk semua drawing yg punya from_drf_id
        drf_ids = [d.get("from_drf_id") for d in docs if d.get("from_drf_id")]
        drf_map = {}
        if drf_ids:
            drfs = await db.drawing_requests.find(
                {"id": {"$in": drf_ids}}, {"_id": 0, "id": 1, "requested_by": 1}
            ).to_list(length=len(drf_ids))
            drf_map = {r.get("id"): (r.get("requested_by") or {}) for r in drfs}

        def _mine(d):
            drf_id = d.get("from_drf_id")
            if drf_id and drf_id in drf_map:
                rb = drf_map[drf_id]
                if rb.get("user_id") == my_id:
                    return True
                if (rb.get("username") or "").lower().strip() == my_username:
                    return True
                if (rb.get("name") or "").lower().strip() == my_name:
                    return True
                return False
            # Fallback drawing lama tanpa DRF: cek request_by_sales
            rbs = (d.get("request_by_sales") or "").lower().strip()
            if rbs and (rbs == my_name or rbs == my_username):
                return True
            # Fallback lain: kalau tidak ada info sama sekali, tampilkan (agar tidak hilang saat data lama)
            return not drf_id and not rbs

        docs = [d for d in docs if _mine(d)]

    return {"items": docs, "total": len(docs)}


# ============================================================================
# Iter 17 — Fase 2: Document Control Stamp + Stamped PDF Preview/Print
# ============================================================================
from utils.pdf_stamper import apply_stamps as _apply_pdf_stamps  # noqa: E402


class DCStampIn(BaseModel):
    notes: str = ""
    stamp_x: Optional[float] = None
    stamp_y: Optional[float] = None
    target: Optional[str] = "mks"
    extra_id: Optional[str] = ""
    # Iter 40 — posisi berbeda tiap halaman
    placements: Optional[List[dict]] = None


class SOStampIn(BaseModel):
    """Iter 20b — SO Stamp untuk Production (Salma isi manual)."""
    so_no: str = ""
    po_no: str = ""
    qty: str = ""
    customer: str = ""
    received_date: str = ""
    due_date: str = ""
    stamp_x: Optional[float] = None
    stamp_y: Optional[float] = None
    # Iter 40 — posisi berbeda tiap halaman
    placements: Optional[List[dict]] = None


@router.post("/drawings/{drawing_id}/stamp-controlled")
async def drawing_stamp_controlled(
    drawing_id: str,
    payload: DCStampIn = None,
    current: dict = Depends(get_current_user),
):
    """Iter 20 — Salma stamp SATU dokumen per call (MKS drawing / Customer Ref / Extra).
    Setiap dokumen bisa punya posisi stamp berbeda. Setelah SEMUA dokumen (MKS + customer_ref
    kalau ada + semua extras kalau ada) di-stamp, approval_status jadi 'controlled'.

    payload.target:
      - "mks"          → stamp file_id (drawing MKS utama)
      - "customer_ref" → stamp customer_ref_file_id
      - "extra"        → stamp extra file (butuh extra_id)
    """
    drawing = await db.drawings.find_one({"id": drawing_id})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if drawing.get("approval_status") not in ("approved", "controlled"):
        raise HTTPException(
            status_code=409,
            detail=f"Drawing status='{drawing.get('approval_status')}'. Stamp DC hanya setelah drawing approved.",
        )
    if not is_doc_control(current) and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Document Control atau Admin yang boleh stamp")

    target = (payload.target if payload else "mks") or "mks"
    stamp_common = {
        "name": current.get("name") or current.get("username"),
        "user_id": current.get("id"),
        "username": current.get("username"),
        "role": current.get("role"),
        "at": _now_iso(),
        "notes": (payload.notes if payload else "") or "",
    }
    if payload and payload.stamp_x is not None:
        stamp_common["x"] = float(payload.stamp_x)
        stamp_common["y"] = float(payload.stamp_y or 0.15)
    if payload and payload.placements:
        _pls = _norm_placements(payload.placements)
        if _pls:
            stamp_common["placements"] = _pls
            stamp_common["x"] = _pls[0]["x"]
            stamp_common["y"] = _pls[0]["y"]

    upd = {"updated_at": _now_iso()}

    if target == "mks":
        upd["dc_stamp"] = stamp_common
        upd["controlled_at"] = _now_iso()
    elif target == "customer_ref":
        if not drawing.get("customer_ref_file_id"):
            raise HTTPException(status_code=400, detail="Drawing ini tidak punya Customer Reference")
        upd["customer_ref_dc_stamp"] = stamp_common
        upd["customer_ref_controlled_at"] = _now_iso()
    elif target == "extra":
        eid = (payload.extra_id if payload else "") or ""
        if not eid:
            raise HTTPException(status_code=400, detail="extra_id wajib untuk target=extra")
        extras = drawing.get("additional_files") or []
        idx = next((i for i, x in enumerate(extras) if x.get("id") == eid), -1)
        if idx < 0:
            raise HTTPException(status_code=404, detail="Extra file tidak ditemukan")
        extras[idx]["dc_stamp"] = stamp_common
        extras[idx]["controlled_at"] = _now_iso()
        upd["additional_files"] = extras
    else:
        raise HTTPException(status_code=400, detail=f"Target tidak valid: {target}")

    # Cek apakah SEMUA dokumen sudah di-stamp → set status controlled
    new_dc = upd.get("dc_stamp") or drawing.get("dc_stamp")
    new_ref = upd.get("customer_ref_dc_stamp") or drawing.get("customer_ref_dc_stamp")
    new_extras = upd.get("additional_files") or drawing.get("additional_files") or []

    mks_ok = bool(new_dc)
    ref_ok = (not drawing.get("customer_ref_file_id")) or bool(new_ref)
    extras_ok = all(bool(x.get("dc_stamp")) for x in new_extras)

    if mks_ok and ref_ok and extras_ok:
        upd["approval_status"] = "controlled"
        # Kalau ini drawing hasil revisi ECN, catat tgl selesai revisi (drawing terbit/IFU)
        rr_cur = drawing.get("revision_request") or {}
        if rr_cur.get("status") == "in_progress" and not rr_cur.get("completed_at"):
            rr_cur["completed_at"] = _now_iso()
            upd["revision_request"] = rr_cur

    await db.drawings.update_one({"id": drawing_id}, {"$set": upd})
    await log_action(current, "drawing_stamp_controlled", "drawings", drawing_id,
                     {"drawing_no": drawing.get("drawing_no"), "target": target,
                      "fully_controlled": upd.get("approval_status") == "controlled"})

    updated = await db.drawings.find_one({"id": drawing_id}, {"_id": 0})
    return {
        "success": True,
        "target": target,
        "approval_status": updated.get("approval_status"),
        "dc_stamp": updated.get("dc_stamp"),
        "customer_ref_dc_stamp": updated.get("customer_ref_dc_stamp"),
        "all_stamped": upd.get("approval_status") == "controlled",
    }


@router.post("/drawings/{drawing_id}/stamp-so")
async def drawing_stamp_so(
    drawing_id: str,
    payload: SOStampIn,
    current: dict = Depends(get_current_user),
):
    """Iter 20b — Salma apply SO stamp (kotak merah info SO/PO/Qty/Customer/Due Date)
    untuk print ke Produksi. Hanya boleh setelah drawing controlled."""
    drawing = await db.drawings.find_one({"id": drawing_id})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    if drawing.get("approval_status") != "controlled":
        raise HTTPException(status_code=409, detail="Drawing belum controlled. SO stamp hanya setelah DC stamp selesai.")
    if not is_doc_control(current) and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Document Control yang boleh SO stamp")

    from routers.bom import normalize_so_no as _norm_so
    so_stamp = {
        "so_no": _norm_so(payload.so_no),
        "po_no": payload.po_no.strip(),
        "qty": payload.qty.strip(),
        "customer": payload.customer.strip(),
        "received_date": payload.received_date.strip(),
        "due_date": payload.due_date.strip(),
        "name": current.get("name") or current.get("username"),
        "user_id": current.get("id"),
        "username": current.get("username"),
        "role": current.get("role"),
        "at": _now_iso(),
    }
    if payload.stamp_x is not None:
        so_stamp["x"] = float(payload.stamp_x)
        so_stamp["y"] = float(payload.stamp_y or 0.25)
    if payload.placements:
        _pls = _norm_placements(payload.placements)
        if _pls:
            so_stamp["placements"] = _pls
            so_stamp["x"] = _pls[0]["x"]
            so_stamp["y"] = _pls[0]["y"]

    await db.drawings.update_one(
        {"id": drawing_id},
        {"$set": {"so_stamp": so_stamp, "so_stamped_at": _now_iso(),
                  "approval_status": "released", "released_at": _now_iso()}},
    )
    await log_action(current, "drawing_stamp_so", "drawings", drawing_id,
                     {"drawing_no": drawing.get("drawing_no"), "so_no": payload.so_no})
    return {"success": True, "so_stamp": so_stamp, "approval_status": "released"}


def _placeholder_pdf(title: str = "File Tidak Tersedia", subtitle: str = "") -> bytes:
    """Bangun PDF placeholder 1 halaman yang ramah (dipakai bila file MKS belum ada / rusak),
    agar preview iframe tidak menampilkan error 500/404 melainkan pesan yang jelas."""
    import fitz  # PyMuPDF
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)  # A4 portrait
    ink = (0.09, 0.11, 0.20)
    gray = (0.42, 0.45, 0.52)
    violet = (0.30, 0.16, 0.55)
    # Kotak info di tengah
    box = fitz.Rect(90, 300, 505, 470)
    page.draw_rect(box, color=(0.80, 0.78, 0.90), fill=(0.96, 0.95, 0.99), width=1.2)
    page.insert_text((120, 350), "Drawing (MKS)", fontsize=11, fontname="hebo", color=violet)
    page.insert_text((120, 385), str(title or "File Tidak Tersedia"), fontsize=18, fontname="hebo", color=ink)
    if subtitle:
        # Bungkus teks panjang sederhana per ~60 karakter
        line = ""
        yy = 415
        for word in str(subtitle).split():
            if len(line) + len(word) + 1 > 62:
                page.insert_text((120, yy), line, fontsize=10, fontname="helv", color=gray)
                yy += 16
                line = word
            else:
                line = (line + " " + word).strip()
        if line:
            page.insert_text((120, yy), line, fontsize=10, fontname="helv", color=gray)
    out = doc.tobytes()
    doc.close()
    return out


@router.get("/drawings/{drawing_id}/pdf-stamped")
async def drawing_pdf_stamped(drawing_id: str, current: dict = Depends(get_current_user)):
    """Return PDF drawing dengan overlay stamps (approval signatures + DC stamp bila ada).

    Watermark 'UNCONTROLLED COPY WHEN PRINTED' otomatis muncul bila user BUKAN doc_control/admin
    dan drawing sudah controlled. Print footer 'Printed by: [nama] | tgl | jam' selalu ada.

    Anti-gagal: bila file MKS belum di-upload / rusak / gagal di-stamp, endpoint mengembalikan
    PDF placeholder ramah (status 200) alih-alih 500/404, supaya preview iframe tetap mulus.
    """
    from fastapi.responses import StreamingResponse

    def _placeholder_response(title: str, subtitle: str = ""):
        return StreamingResponse(
            io.BytesIO(_placeholder_pdf(title, subtitle)),
            media_type="application/pdf",
            headers={"Content-Disposition": 'inline; filename="placeholder.pdf"'},
        )

    drawing = await db.drawings.find_one({"id": drawing_id})
    if not drawing:
        return _placeholder_response("Drawing Tidak Ditemukan",
                                     "Data drawing tidak tersedia di sistem.")
    file_id = drawing.get("file_id")
    if not file_id:
        return _placeholder_response("File MKS Belum Diunggah",
                                     "Drawing ini belum memiliki file MKS yang diunggah. "
                                     "Silakan hubungi Engineering untuk melengkapi berkas.")

    # Load original PDF from GridFS (bucket "drawings")
    try:
        stream = await _fs().open_download_stream(ObjectId(file_id))
        content = await stream.read()
    except Exception:
        return _placeholder_response("File MKS Tidak Dapat Dibaca",
                                     "Berkas MKS gagal diambil dari penyimpanan. Mungkin file rusak atau hilang.")

    # Legacy import: file bisa berupa Office (.docx) → konversi ke PDF agar bisa dirender.
    if not (content or b"")[:5].startswith(b"%PDF"):
        try:
            fname = drawing.get("filename") or ""
            ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
            from utils.office_render import is_office_ext, office_to_pdf
            if is_office_ext(ext):
                content = office_to_pdf(content, ext)
        except Exception:
            pass
        if not (content or b"")[:5].startswith(b"%PDF"):
            return _placeholder_response("Format File Tidak Didukung",
                                         "Berkas MKS bukan PDF yang valid dan tidak dapat dikonversi untuk pratinjau.")

    # Uncontrolled watermark logic: only if drawing is controlled AND user is not DC/admin
    is_dc_or_admin = is_doc_control(current) or is_admin_like(current)
    show_uncontrolled_watermark = (
        drawing.get("approval_status") == "controlled" and not is_dc_or_admin
    )

    # Iter 18 — Fetch signature PNG bytes untuk setiap approver yang punya user_id
    signature_bytes_map: dict = {}
    approvals = drawing.get("approvals") or []
    sig_bucket = AsyncIOMotorGridFSBucket(db, bucket_name="signatures")
    for appr in approvals:
        uid = appr.get("user_id")
        if not uid or uid in signature_bytes_map:
            continue
        u = await db.users.find_one({"id": uid}, {"signature_gridfs_id": 1})
        sig_id = (u or {}).get("signature_gridfs_id")
        if not sig_id:
            continue
        try:
            s = await sig_bucket.open_download_stream(ObjectId(sig_id))
            signature_bytes_map[uid] = await s.read()
        except Exception:
            pass

    try:
        stamped = _apply_pdf_stamps(
            content,
            approvals=approvals,
            dc_stamp=drawing.get("dc_stamp"),
            watermark_uncontrolled=show_uncontrolled_watermark,
            printed_by=current.get("name") or current.get("username") or "",
            signature_bytes_map=signature_bytes_map,
            so_stamp=drawing.get("so_stamp"),
        )
    except Exception:
        return _placeholder_response("Gagal Memproses Drawing",
                                     "Terjadi kesalahan saat menyiapkan pratinjau ber-stempel. "
                                     "Berkas MKS mungkin rusak atau tidak valid.")
    # Iter 18 — audit trail print/preview history
    try:
        await log_action(current, "drawing_preview_stamped", "drawings", drawing_id, {
            "drawing_no": drawing.get("drawing_no"),
            "approval_status": drawing.get("approval_status"),
            "watermarked": show_uncontrolled_watermark,
        })
    except Exception:
        pass
    filename = f"{drawing.get('drawing_no', drawing_id)}_stamped.pdf"
    return StreamingResponse(
        io.BytesIO(stamped),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


async def _target_raw_bytes(drawing: dict, target: str, extra_id: str = "") -> bytes:
    """Ambil bytes PDF asli sesuai target (mks / customer_ref / extra) untuk render halaman."""
    if target == "customer_ref":
        fid = drawing.get("customer_ref_file_id")
        if not fid:
            raise HTTPException(status_code=404, detail="Customer reference tidak ada")
    elif target == "extra":
        entry = next((f for f in (drawing.get("additional_files") or []) if f.get("id") == extra_id), None)
        if not entry or not entry.get("file_id"):
            raise HTTPException(status_code=404, detail="Extra file tidak ditemukan")
        fid = entry["file_id"]
    else:  # "mks" (default)
        fid = drawing.get("file_id")
        if not fid:
            raise HTTPException(status_code=404, detail="File PDF drawing belum di-upload")
    stream = await _fs().open_download_stream(ObjectId(fid))
    raw = await stream.read()
    # Legacy import: Eng DWG / Customer bisa berupa Word (.docx) → konversi ke PDF agar bisa dirender.
    if not raw[:5].startswith(b"%PDF"):
        if target == "customer_ref":
            fname = drawing.get("customer_ref_filename") or ""
        elif target == "extra":
            entry = next((f for f in (drawing.get("additional_files") or []) if f.get("id") == extra_id), None)
            fname = (entry or {}).get("filename") or ""
        else:
            fname = drawing.get("filename") or ""
        ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
        from utils.office_render import is_office_ext, office_to_pdf
        if is_office_ext(ext):
            raw = office_to_pdf(raw, ext)
    return raw


async def _build_signature_map(approvals: list) -> dict:
    """Fetch signature PNG bytes untuk tiap approver (dipakai render stamp)."""
    signature_bytes_map: dict = {}
    sig_bucket = AsyncIOMotorGridFSBucket(db, bucket_name="signatures")
    for appr in (approvals or []):
        uid = appr.get("user_id")
        if not uid or uid in signature_bytes_map:
            continue
        u = await db.users.find_one({"id": uid}, {"signature_gridfs_id": 1})
        sig_id = (u or {}).get("signature_gridfs_id")
        if not sig_id:
            continue
        try:
            s = await sig_bucket.open_download_stream(ObjectId(sig_id))
            signature_bytes_map[uid] = await s.read()
        except Exception:
            pass
    return signature_bytes_map


async def _build_stamped_for_target(drawing: dict, target: str, extra_id: str, raw: bytes, current: dict,
                                    hide_so: bool = False) -> bytes:
    """Terapkan overlay stamp yang sesuai untuk preview per target (mks/customer_ref/extra).
    Meniru perilaku endpoint pdf-stamped / customer-ref preview / extras preview.

    hide_so=True → jangan render SO stamp (kotak merah SO) pada target mks. Dipakai di
    Master Drawing List agar preview menampilkan versi DC-stamped yang 'bersih' tanpa SO stamp."""
    is_dc_or_admin = is_doc_control(current) or is_admin_like(current)
    printed_by = current.get("name") or current.get("username") or ""
    try:
        if target == "customer_ref":
            show_wm = drawing.get("approval_status") in ("controlled", "approved") and not is_dc_or_admin
            return _apply_pdf_stamps(raw, approvals=[], dc_stamp=drawing.get("customer_ref_dc_stamp"),
                                     watermark_uncontrolled=show_wm, printed_by=printed_by)
        if target == "extra":
            entry = next((f for f in (drawing.get("additional_files") or []) if f.get("id") == extra_id), None)
            show_wm = drawing.get("approval_status") in ("controlled", "approved") and not is_dc_or_admin
            return _apply_pdf_stamps(raw, approvals=[], dc_stamp=(entry or {}).get("dc_stamp"),
                                     watermark_uncontrolled=show_wm, printed_by=printed_by)
        # default mks — full stamps (approvals + dc + so + watermark)
        show_wm = drawing.get("approval_status") == "controlled" and not is_dc_or_admin
        approvals = drawing.get("approvals") or []
        sig_map = await _build_signature_map(approvals)
        so_stamp = None if hide_so else drawing.get("so_stamp")
        return _apply_pdf_stamps(raw, approvals=approvals, dc_stamp=drawing.get("dc_stamp"),
                                 watermark_uncontrolled=show_wm, printed_by=printed_by,
                                 signature_bytes_map=sig_map, so_stamp=so_stamp)
    except Exception:
        return raw  # fallback ke raw kalau gagal stamp


@router.get("/drawings/{drawing_id}/page-meta")
async def drawing_page_meta(drawing_id: str, target: str = "mks", extra_id: str = "",
                            current: dict = Depends(get_current_user)):
    """Metadata halaman PDF (jumlah halaman + ukuran tiap halaman) — dipakai stamp picker
    multi-halaman & viewer image-based. Bila file belum ada / bukan PDF valid → kembalikan
    {pages:0, message} (HTTP 200) supaya viewer menampilkan empty-state ramah, bukan error."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    drawing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    import fitz  # PyMuPDF
    try:
        raw = await _target_raw_bytes(drawing, target, extra_id)
        doc = fitz.open(stream=raw, filetype="pdf")
    except HTTPException as he:
        return {"pages": 0, "sizes": [], "message": he.detail}
    except Exception:
        return {"pages": 0, "sizes": [],
                "message": "File belum diunggah atau bukan PDF yang valid untuk pratinjau."}
    sizes = [{"w": round(p.rect.width, 2), "h": round(p.rect.height, 2)} for p in doc]
    n = doc.page_count
    doc.close()
    return {"pages": n, "sizes": sizes}


@router.get("/drawings/{drawing_id}/page-image")
async def drawing_page_image(drawing_id: str, request: Request, page: int = 0, target: str = "mks",
                             extra_id: str = "", scale: float = 2.0, stamped: bool = False,
                             hide_so: bool = False,
                             current: dict = Depends(get_current_user)):
    """Render satu halaman PDF menjadi gambar PNG (untuk preview stamp picker & viewer baca-saja).
    stamped=1 → tampilkan versi ber-stamp (approval/DC/SO + watermark) sesuai role.
    hide_so=1 → sembunyikan SO stamp pada target mks (dipakai Master Drawing List).

    Dipercepat: hasil render & PDF ber-stamp di-cache (kunci mengikuti isi + peran),
    plus ETag agar browser tidak mengunduh ulang gambar yang sama (304 Not Modified)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    drawing = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    scale = max(1.0, min(3.0, float(scale or 2.0)))

    from utils.render_cache import STAMP_CACHE, png_etag, render_png_cached, sha

    raw = await _target_raw_bytes(drawing, target, extra_id)
    if stamped:
        # Build PDF ber-stamp SEKALI lalu cache (kunci ikut isi + peran + state drawing).
        skey = ("stamp:" + sha(raw) + f":{target}:{extra_id}:{current.get('role','')}"
                f":{int(bool(hide_so))}:{drawing.get('updated_at','')}")
        src = STAMP_CACHE.get(skey)
        if src is None:
            src = await _build_stamped_for_target(drawing, target, extra_id, raw, current, hide_so=hide_so)
            STAMP_CACHE.set(skey, src)
    else:
        src = raw

    # Validasi cache browser: bila ETag cocok → 304 (tanpa render/kirim ulang).
    etag = png_etag(src, page, scale)
    cache_hdr = {"Cache-Control": "private, max-age=600", "ETag": etag}
    if request.headers.get("if-none-match") == etag:
        return Response(status_code=304, headers=cache_hdr)

    try:
        png = render_png_cached(src, page, scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ada")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"File bukan PDF valid: {e}")
    return StreamingResponse(io.BytesIO(png), media_type="image/png", headers=cache_hdr)


async def _revision_snapshot_bytes(drawing_id: str, rev_id: str, which: str = "mks") -> bytes:
    """Ambil bytes PDF versi lama dari snapshot history revisi (berdasarkan file_id yang disimpan)."""
    d = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}}, {"_id": 0, "revisions": 1})
    if not d:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    rev = next((r for r in (d.get("revisions") or []) if r.get("id") == rev_id), None)
    if not rev:
        raise HTTPException(status_code=404, detail="Riwayat revisi tidak ditemukan")
    snap = rev.get("snapshot") or {}
    fid = snap.get("customer_ref_file_id") if which == "customer" else snap.get("file_id")
    if not fid:
        raise HTTPException(status_code=404, detail="File versi lama tidak tersedia untuk revisi ini")
    try:
        stream = await _fs().open_download_stream(ObjectId(fid))
        return await stream.read()
    except Exception:
        raise HTTPException(status_code=404, detail="File versi lama tidak ditemukan di penyimpanan")


@router.get("/drawings/{drawing_id}/revisions/{rev_id}/page-meta")
async def revision_snapshot_page_meta(drawing_id: str, rev_id: str, which: str = "mks",
                                      current: dict = Depends(get_current_user)):
    """Metadata halaman PDF versi lama (history revisi)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    raw = await _revision_snapshot_bytes(drawing_id, rev_id, which)
    import fitz  # PyMuPDF
    try:
        doc = fitz.open(stream=raw, filetype="pdf")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"File bukan PDF valid: {e}")
    sizes = [{"w": round(p.rect.width, 2), "h": round(p.rect.height, 2)} for p in doc]
    n = doc.page_count
    doc.close()
    return {"pages": n, "sizes": sizes}


@router.get("/drawings/{drawing_id}/revisions/{rev_id}/page-image")
async def revision_snapshot_page_image(drawing_id: str, rev_id: str, request: Request, page: int = 0,
                                       scale: float = 2.0, which: str = "mks",
                                       current: dict = Depends(get_current_user)):
    """Render satu halaman PDF versi lama (history revisi) menjadi PNG (cached + ETag)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    raw = await _revision_snapshot_bytes(drawing_id, rev_id, which)
    scale = max(1.0, min(3.0, float(scale or 2.0)))
    from utils.render_cache import png_etag, render_png_cached
    etag = png_etag(raw, page, scale)
    cache_hdr = {"Cache-Control": "private, max-age=600", "ETag": etag}
    if request.headers.get("if-none-match") == etag:
        return Response(status_code=304, headers=cache_hdr)
    try:
        png = render_png_cached(raw, page, scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ada")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"File bukan PDF valid: {e}")
    return StreamingResponse(io.BytesIO(png), media_type="image/png", headers=cache_hdr)


@router.get("/drawings/{drawing_id}/revisions/{rev_id}/download")
async def revision_snapshot_download(drawing_id: str, rev_id: str, which: str = "mks",
                                     current: dict = Depends(get_current_user)):
    """Unduh PDF versi lama (history revisi)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    raw = await _revision_snapshot_bytes(drawing_id, rev_id, which)
    fname = f"{drawing_id}_rev-{rev_id[:8]}_{which}.pdf"
    return StreamingResponse(io.BytesIO(raw), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


async def _sig_png_by_user(uid: str, name: str = None):
    """Ambil PNG tanda tangan berdasarkan user_id; fallback cari user berdasarkan nama."""
    u = None
    if uid:
        u = await db.users.find_one({"id": uid}, {"signature_gridfs_id": 1})
    if not u and name:
        u = await db.users.find_one({"name": name}, {"signature_gridfs_id": 1})
    sid = (u or {}).get("signature_gridfs_id")
    if not sid:
        return None
    try:
        bucket = AsyncIOMotorGridFSBucket(db, bucket_name="signatures")
        s = await bucket.open_download_stream(ObjectId(sid))
        return await s.read()
    except Exception:
        return None


async def _build_ecn_sheet_bytes(drawing_id: str, current: dict) -> tuple:
    """Bangun bytes PDF Lembar Acknowledgment ECN (Form MKS-F-ENG-004) + stamp TTD digital.
    Dipakai oleh endpoint download (ecn-sheet) maupun viewer image-based (page-meta/page-image).
    Return: (pdf_bytes, filename)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    d = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}})
    if not d:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    rr = d.get("revision_request") or {}
    ecn = rr.get("ecn") or {}
    if not ecn.get("ecn_no"):
        raise HTTPException(status_code=404, detail="Tidak ada ECN pada drawing ini")
    ack = rr.get("ack") or {}

    import fitz  # PyMuPDF
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)  # A4 portrait
    W = 595
    ink = (0.09, 0.11, 0.20)
    gray = (0.42, 0.45, 0.52)
    line_c = (0.75, 0.77, 0.82)

    def txt(x, y, s, size=9, color=ink, bold=False):
        page.insert_text((x, y), str(s or ""), fontsize=size,
                         fontname="helv" if not bold else "hebo", color=color)

    def hline(y, x0=40, x1=W - 40):
        page.draw_line((x0, y), (x1, y), color=line_c, width=0.8)

    # Header
    page.draw_rect(fitz.Rect(40, 36, W - 40, 84), color=ink, fill=(0.93, 0.94, 0.99), width=1)
    txt(52, 58, "ENGINEERING CHANGE NOTICE (ECN)", size=15, bold=True)
    txt(52, 74, "Form MKS-F-ENG-004  ·  Lembar Acknowledgment & Distribusi", size=9, color=gray)
    txt(W - 180, 58, ecn.get("ecn_no"), size=13, bold=True, color=(0.30, 0.16, 0.55))

    y = 104
    def field(label, value, x=52, w=250):
        nonlocal_y = y
        txt(x, nonlocal_y, label.upper(), size=7, color=gray, bold=True)
        txt(x, nonlocal_y + 13, value or "-", size=10)

    # Info grid (2 kolom)
    rows_info = [
        ("No. Drawing", d.get("drawing_no"), "Rev", str(d.get("rev_no") if d.get("rev_no") is not None else "-")),
        ("Nomor SO", d.get("so_no"), "Customer", d.get("customer_name")),
        ("Diajukan oleh", rr.get("requested_by"), "Status", (rr.get("status") or "-")),
    ]
    for l1, v1, l2, v2 in rows_info:
        txt(52, y, l1.upper(), size=7, color=gray, bold=True)
        txt(52, y + 13, v1 or "-", size=10)
        txt(320, y, l2.upper(), size=7, color=gray, bold=True)
        txt(320, y + 13, v2 or "-", size=10)
        y += 34
    hline(y); y += 16

    # Perubahan
    txt(52, y, "DETAIL PERUBAHAN", size=8, color=gray, bold=True); y += 16
    txt(52, y, "Kondisi Saat Ini:", size=8, color=gray); txt(150, y, ecn.get("current_desc") or "-", size=9); y += 16
    txt(52, y, "Menjadi (Usulan):", size=8, color=gray); txt(150, y, ecn.get("proposed_desc") or "-", size=9); y += 16
    txt(52, y, "Alasan/Tujuan:", size=8, color=gray); txt(150, y, ecn.get("purpose_explanation") or "-", size=9); y += 20
    hline(y); y += 16

    # Timeline
    def fmt(iso):
        if not iso:
            return "-"
        try:
            return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%d %b %Y %H:%M")
        except Exception:
            return str(iso)[:16]
    txt(52, y, "TIMELINE", size=8, color=gray, bold=True); y += 15
    tl = [
        ("Registrasi ECN", rr.get("requested_at")),
        ("Mulai Revisi", rr.get("started_at")),
        ("Selesai Revisi (Terbit/IFU)", rr.get("completed_at")),
        ("Distribusi (Doc Control)", (ack.get("doc_control") or {}).get("at")),
    ]
    for lbl, val in tl:
        txt(60, y, f"•  {lbl}", size=9); txt(320, y, fmt(val), size=9, color=gray); y += 15
    y += 6
    hline(y); y += 18

    # Tanda tangan
    txt(52, y, "TANDA TANGAN / ACKNOWLEDGMENT", size=8, color=gray, bold=True); y += 8
    head = (ecn.get("approvals") or {}).get("head_of_dept") or {}
    signers = [
        ("Dibuat (Eng Staff)", {"name": rr.get("requested_by"), "user_id": rr.get("requested_by_id"), "at": rr.get("requested_at")}),
        ("Disetujui (Eng Leader)", head),
        ("Acknowledge Produksi", ack.get("production")),
        ("Tanda Tangan QA/QC", ack.get("qa_qc")),
        ("Distribusi Doc Control", ack.get("doc_control")),
    ]
    n_sign = len(signers)
    gap = 8
    box_w = (W - 80 - gap * (n_sign - 1)) / n_sign
    bx = 40
    by = y
    for title, data in signers:
        rect = fitz.Rect(bx, by, bx + box_w, by + 120)
        page.draw_rect(rect, color=line_c, width=0.8)
        txt(bx + 6, by + 15, title.upper(), size=6, color=gray, bold=True)
        data = data or {}
        png = await _sig_png_by_user(data.get("user_id"), data.get("name"))
        if png:
            try:
                page.insert_image(fitz.Rect(bx + 6, by + 22, bx + box_w - 6, by + 74), stream=png, keep_proportion=True)
            except Exception:
                pass
        elif data.get("auto"):
            txt(bx + 6, by + 52, "(otomatis)", size=7, color=gray)
        page.draw_line((bx + 6, by + 84), (bx + box_w - 6, by + 84), color=line_c, width=0.6)
        nm = data.get("name") or ("Otomatis" if data.get("auto") else "-")
        txt(bx + 6, by + 97, nm, size=7.5, bold=True)
        at = data.get("at")
        txt(bx + 6, by + 110, fmt(at) if at else "Belum TTD", size=6, color=gray)
        bx += box_w + gap

    # Footer
    txt(40, 812, f"Dicetak: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')} oleh {current.get('name') or current.get('username')}", size=7, color=gray)
    txt(W - 200, 812, "Dokumen terkendali — PT MKS", size=7, color=gray)

    out = doc.tobytes()
    doc.close()
    fname = f"ECN_{ecn.get('ecn_no')}_{d.get('drawing_no', drawing_id)}.pdf"
    return out, fname


@router.get("/drawings/{drawing_id}/ecn-sheet")
async def ecn_sheet_pdf(drawing_id: str, current: dict = Depends(get_current_user)):
    """Lembar Acknowledgment ECN (PDF) — Form MKS-F-ENG-004 dengan stamp TTD digital
    (PNG + tanggal/jam) Produksi & QA/QC sebagai bukti resmi distribusi."""
    out, fname = await _build_ecn_sheet_bytes(drawing_id, current)
    return StreamingResponse(io.BytesIO(out), media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="{fname}"'})


@router.get("/drawings/{drawing_id}/ecn-sheet/page-meta")
async def ecn_sheet_page_meta(drawing_id: str, current: dict = Depends(get_current_user)):
    """Metadata halaman Lembar ECN untuk viewer image-based (jumlah halaman + ukuran)."""
    out, _ = await _build_ecn_sheet_bytes(drawing_id, current)
    from utils.pdf_render import pdf_page_meta
    return pdf_page_meta(out)


@router.get("/drawings/{drawing_id}/ecn-sheet/page-image")
async def ecn_sheet_page_image(drawing_id: str, request: Request, page: int = 0, scale: float = 2.0,
                               current: dict = Depends(get_current_user)):
    """Render 1 halaman Lembar ECN sebagai PNG (viewer image-based, cached + ETag)."""
    out, _ = await _build_ecn_sheet_bytes(drawing_id, current)
    scale = max(1.0, min(3.0, float(scale or 2.0)))
    from utils.render_cache import png_etag, render_png_cached
    etag = png_etag(out, page, scale)
    cache_hdr = {"Cache-Control": "private, max-age=300", "ETag": etag}
    if request.headers.get("if-none-match") == etag:
        return Response(status_code=304, headers=cache_hdr)
    try:
        png = render_png_cached(out, page=page, scale=scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ada")
    return StreamingResponse(io.BytesIO(png), media_type="image/png", headers=cache_hdr)


@router.get("/drawings/pending-dc-stamp")
async def list_pending_dc_stamp(current: dict = Depends(get_current_user)):
    """List drawing yg sudah approved tapi belum di-stamp DC (untuk halaman Document Distribution Record)."""
    if not is_doc_control(current) and not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Document Control atau Admin")
    q = {"approval_status": "approved", "deleted_at": {"$exists": False}}
    docs = await db.drawings.find(q, {"_id": 0}).sort([("approved_at", -1)]).to_list(length=200)
    return {"items": [_clean(d) for d in docs], "total": len(docs)}


@router.get("/drawings/my-signature-history")
async def my_signature_history(all: bool = False, current: dict = Depends(get_current_user)):
    """Iter 22 — Riwayat TTD (bukti audit ISO): drawing yang pernah di-TTD.
    Default = milik user login. all=true (khusus Eng Leader/Admin) = semua user."""
    uid = current.get("id")
    uname = (current.get("username") or "").lower().strip()
    if not uid and not uname:
        raise HTTPException(status_code=401, detail="Unauthenticated")

    show_all = bool(all) and (is_eng_head(current) or is_admin_like(current))
    if show_all:
        q = {"deleted_at": {"$exists": False}, "approvals": {"$exists": True, "$ne": []}}
    else:
        q = {
            "deleted_at": {"$exists": False},
            "approvals": {
                "$elemMatch": {
                    "$or": [
                        {"user_id": uid} if uid else {"user_id": "__na__"},
                        {"username": uname} if uname else {"username": "__na__"},
                    ]
                }
            },
        }
    docs = await db.drawings.find(q, {"_id": 0}).sort([("updated_at", -1)]).to_list(length=500)

    history = []
    for d in docs:
        for a in d.get("approvals", []) or []:
            if not show_all:
                match_id = uid and a.get("user_id") == uid
                match_name = uname and (a.get("username") or "").lower().strip() == uname
                if not (match_id or match_name):
                    continue
            history.append({
                "drawing_id": d.get("id"),
                "drawing_no": d.get("drawing_no"),
                "project_name": d.get("project_name") or d.get("title") or "",
                "customer_name": d.get("customer_name") or d.get("customer_code") or "",
                "so_no": d.get("so_no") or "",
                "stage": a.get("stage"),
                "signed_at": a.get("at") or a.get("filled_at"),
                "signed_by": a.get("name"),
                "notes": a.get("notes") or "",
                "drawing_status_now": d.get("approval_status"),
                "position": {"x": a.get("x"), "y": a.get("y"), "page": a.get("page")},
                "has_pdf": bool(d.get("file_id")),
            })
    # sort desc by signed_at
    history.sort(key=lambda x: (x.get("signed_at") or ""), reverse=True)
    return {"items": history, "total": len(history)}


@router.get("/drawings/{drawing_id}")
async def get_drawing(drawing_id: str, current: dict = Depends(get_current_user)):
    """Iter 22 — Fetch single drawing (untuk Engineering Work Order page).
    Route ini di-taruh paling bawah agar tidak menangkap path spesifik seperti
    /drawings/my-assignments, /drawings/pending-my-approval, /drawings/pending-dc-stamp."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await db.drawings.find_one({"id": drawing_id, "deleted_at": {"$exists": False}}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Drawing tidak ditemukan")
    return _clean(doc)
