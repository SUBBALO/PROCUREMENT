"""Finance module.

Fitur:
- Masterlist Rate Karyawan (rate per jam) — hanya Finance/Admin yang bisa lihat & input.
- Daily Production Report (Finance) — salinan read-only data produksi harian
  DITAMBAH kolom biaya (rate/jam x jam kerja) + total per operator & per hari.

CATATAN PENTING: rate disimpan di koleksi terpisah `employee_rates` dan HANYA
diekspos lewat endpoint /finance/*. Endpoint produksi tidak pernah mengembalikan rate,
sehingga operator/role Produksi tidak bisa melihat rate.
"""
import uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel

from db import db
from deps import get_current_user, log_action
from routers.production import _work_hours, _serialize_report, _build_report_filter

router = APIRouter(prefix="/finance", tags=["finance"])

FINANCE_ROLES = ("finance", "admin", "super_admin")


def _can_finance(user: dict) -> bool:
    return (user.get("role") or "") in FINANCE_ROLES


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _f(v) -> float:
    try:
        return float(v)
    except Exception:
        return 0.0


class RateIn(BaseModel):
    rate_per_hour: float = 0


# ---------------------------------------------------------------------------
# Masterlist Rate Karyawan
# ---------------------------------------------------------------------------
@router.get("/employee-rates")
async def list_employee_rates(current: dict = Depends(get_current_user)):
    """Daftar karyawan produksi aktif + rate/jam (Finance/Admin only)."""
    if not _can_finance(current):
        raise HTTPException(status_code=403, detail="Hanya Finance/Admin yang bisa mengakses rate")
    emps = await db.production_employees.find(
        {"active": {"$ne": False}}, {"_id": 0, "id": 1, "name": 1, "designation": 1}
    ).sort("name", 1).to_list(length=10000)
    rates = await db.employee_rates.find({}, {"_id": 0}).to_list(length=10000)
    rate_by_id = {r.get("employee_id"): r for r in rates}
    items = []
    for e in emps:
        r = rate_by_id.get(e.get("id"))
        items.append({
            "employee_id": e.get("id"),
            "name": e.get("name") or "",
            "designation": e.get("designation") or "",
            "rate_per_hour": _f(r.get("rate_per_hour")) if r else 0.0,
            "updated_at": r.get("updated_at") if r else None,
            "updated_by": r.get("updated_by_name") if r else None,
        })
    return {"items": items, "count": len(items)}


@router.put("/employee-rates/{employee_id}")
async def set_employee_rate(employee_id: str, payload: RateIn, current: dict = Depends(get_current_user)):
    """Set/ubah rate per jam seorang karyawan (Finance/Admin only)."""
    if not _can_finance(current):
        raise HTTPException(status_code=403, detail="Hanya Finance/Admin yang bisa input rate")
    emp = await db.production_employees.find_one({"id": employee_id}, {"_id": 0, "name": 1})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    rate = round(_f(payload.rate_per_hour), 2)
    doc = {
        "employee_id": employee_id,
        "employee_name": emp.get("name") or "",
        "rate_per_hour": rate,
        "updated_at": _now_iso(),
        "updated_by": current.get("id"),
        "updated_by_name": current.get("name") or current.get("username") or "",
    }
    existing = await db.employee_rates.find_one({"employee_id": employee_id}, {"_id": 0, "id": 1})
    if existing:
        await db.employee_rates.update_one({"employee_id": employee_id}, {"$set": doc})
    else:
        doc["id"] = str(uuid.uuid4())
        await db.employee_rates.insert_one(doc)
    await log_action(current, "set_employee_rate", "employee_rate", employee_id, {"rate_per_hour": rate})
    return {"ok": True, "employee_id": employee_id, "rate_per_hour": rate}


# ---------------------------------------------------------------------------
# Daily Production Report (Finance view) — salinan produksi + biaya
# ---------------------------------------------------------------------------
async def _compute_finance_daily(month, date, operator, so_no):
    """Ambil baris laporan produksi + hitung biaya (rate/jam x jam kerja) + rekap."""
    emps = await db.production_employees.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(length=10000)
    id_by_name = {(e.get("name") or "").strip().lower(): e.get("id") for e in emps}
    rates = await db.employee_rates.find({}, {"_id": 0}).to_list(length=10000)
    rate_by_id = {r.get("employee_id"): _f(r.get("rate_per_hour")) for r in rates}

    def rate_for(op_name: str) -> float:
        eid = id_by_name.get((op_name or "").strip().lower())
        return rate_by_id.get(eid, 0.0) if eid else 0.0

    filt = _build_report_filter(month, date, operator, so_no)
    rows = await db.production_reports.find(filt, {"_id": 0}).to_list(length=100000)
    rows.sort(key=lambda r: (r.get("report_date") or "", r.get("created_at") or ""), reverse=True)

    items = []
    by_operator = {}
    by_date = {}
    grand_hours = 0.0
    grand_cost = 0.0
    for r in rows:
        base = _serialize_report(r)
        op = base.get("operator_name") or ""
        hrs = _f(base.get("work_hours"))
        rate = rate_for(op)
        cost = round(hrs * rate)
        base["rate_per_hour"] = rate
        base["cost"] = cost
        items.append(base)
        grand_hours += hrs
        grand_cost += cost
        o = by_operator.setdefault(op or "—", {"name": op or "—", "total_hours": 0.0, "total_cost": 0.0, "rows": 0})
        o["total_hours"] = round(o["total_hours"] + hrs, 2)
        o["total_cost"] += cost
        o["rows"] += 1
        d = base.get("report_date") or "—"
        dd = by_date.setdefault(d, {"date": d, "total_hours": 0.0, "total_cost": 0.0, "rows": 0})
        dd["total_hours"] = round(dd["total_hours"] + hrs, 2)
        dd["total_cost"] += cost
        dd["rows"] += 1

    summary_operator = sorted(by_operator.values(), key=lambda x: x["name"].lower())
    summary_date = sorted(by_date.values(), key=lambda x: x["date"], reverse=True)
    return {
        "items": items,
        "count": len(items),
        "total_hours": round(grand_hours, 2),
        "total_cost": round(grand_cost),
        "summary_operator": summary_operator,
        "summary_date": summary_date,
    }


@router.get("/daily-production")
async def finance_daily_production(
    month: Optional[str] = None,
    date: Optional[str] = None,
    operator: Optional[str] = None,
    so_no: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """Salinan read-only laporan produksi harian + kolom biaya (rate/jam x jam kerja).
    Total per operator & per tanggal. Finance/Admin only."""
    if not _can_finance(current):
        raise HTTPException(status_code=403, detail="Hanya Finance/Admin yang bisa mengakses")
    return await _compute_finance_daily(month, date, operator, so_no)


@router.get("/daily-production/export.xlsx")
async def finance_daily_export(
    month: Optional[str] = None,
    date: Optional[str] = None,
    operator: Optional[str] = None,
    so_no: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """Export rekap biaya tenaga kerja ke Excel (Detail + Rekap per Operator) untuk payroll."""
    if not _can_finance(current):
        raise HTTPException(status_code=403, detail="Hanya Finance/Admin yang bisa export")
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await _compute_finance_daily(month, date, operator, so_no)

    head_fill = PatternFill("solid", fgColor="1E293B")
    head_font = Font(bold=True, color="FFFFFF")
    money_fill = PatternFill("solid", fgColor="ECFDF5")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = center
            cell.border = border

    wb = Workbook()

    # --- Sheet 1: Detail ---
    ws = wb.active
    ws.title = "Detail"
    headers = ["Tanggal", "Operator", "SO", "Customer", "Process", "Jam Mulai", "Jam Selesai", "Total Jam Kerja", "Rate/Jam (Rp)", "Biaya (Rp)"]
    ws.append(headers)
    style_header(ws, len(headers))
    for it in data["items"]:
        ws.append([
            it.get("report_date") or "",
            it.get("operator_name") or "",
            it.get("so_no") or "",
            it.get("customer") or "",
            it.get("process") or "",
            it.get("work_start") or "",
            it.get("work_end") or "",
            _f(it.get("work_hours")),
            _f(it.get("rate_per_hour")),
            _f(it.get("cost")),
        ])
    # baris total
    total_row = ["TOTAL", "", "", "", "", "", "", data["total_hours"], "", data["total_cost"]]
    ws.append(total_row)
    trow = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=trow, column=c).font = bold
        ws.cell(row=trow, column=c).fill = money_fill
    for col, w in zip("ABCDEFGHIJ", [12, 20, 12, 22, 22, 11, 12, 11, 15, 16]):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = "#,##0"
    ws.freeze_panes = "A2"

    # --- Sheet 2: Rekap per Operator ---
    ws2 = wb.create_sheet("Rekap per Operator")
    h2 = ["Operator", "Total Jam", "Jumlah Baris", "Total Biaya (Rp)"]
    ws2.append(h2)
    style_header(ws2, len(h2))
    for s in data["summary_operator"]:
        ws2.append([s["name"], s["total_hours"], s["rows"], s["total_cost"]])
    ws2.append(["TOTAL", data["total_hours"], data["count"], data["total_cost"]])
    trow2 = ws2.max_row
    for c in range(1, len(h2) + 1):
        ws2.cell(row=trow2, column=c).font = bold
        ws2.cell(row=trow2, column=c).fill = money_fill
    for col, w in zip("ABCD", [24, 12, 14, 18]):
        ws2.column_dimensions[col].width = w
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "#,##0"
    ws2.freeze_panes = "A2"

    period = date or month or "semua"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"biaya_tenaga_kerja_{period}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
