"""BOM Attachments — Drawing PDF, Nesting PDF, Costing Excel per BOM revision.

Multiple files per category. Preview inline (PDF native; Excel → convert to PDF via LibreOffice).
"""
from __future__ import annotations
import io
import os
import re
import subprocess
import tempfile
import uuid
from datetime import datetime, timezone
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorGridFSBucket

from db import db
from deps import (
    get_current_user, log_action, can_view_costing, is_drawing_preview_only,
    PRICE_ATTACHMENT_CATEGORIES, DRAWING_ATTACHMENT_CATEGORIES,
)
from routers.drawing_register import _normalize_dno

router = APIRouter(tags=["bom-attachments"])


async def _get_attachment_or_404(bom_id: str, attach_id: str) -> dict:
    doc = await db.bom_attachments.find_one({"id": attach_id, "bom_id": bom_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="Attachment tidak ditemukan")
    return doc


def _guard_costing_access(doc: dict, current: dict):
    """403 jika mencoba akses file costing/price tanpa hak."""
    if doc.get("category") in PRICE_ATTACHMENT_CATEGORIES and not can_view_costing(current):
        raise HTTPException(status_code=403, detail="Anda tidak berwenang melihat file Costing/Harga.")


def _guard_drawing_download(doc: dict, current: dict):
    """403 jika role preview-only (QC/DocControl/Store/Produksi) mencoba DOWNLOAD file DWG/Customer."""
    if doc.get("category") in DRAWING_ATTACHMENT_CATEGORIES and is_drawing_preview_only(current):
        raise HTTPException(status_code=403, detail="File Drawing/Customer hanya bisa dipreview (tanpa download) untuk role Anda.")

VALID_CATEGORIES = {"drawing", "customer_ref", "nesting", "nesting_price", "costing", "costing_prev", "revision"}
CATEGORY_LABELS = {
    "drawing": "Drawing PDF (MKS)",
    "customer_ref": "Customer Reference",
    "nesting": "Nesting",
    "nesting_price": "Nesting Price",
    "costing": "Costing (current)",
    "costing_prev": "Costing Sebelumnya",
    "revision": "Revision (dari Engineering Leader)",
}
CATEGORY_ALLOWED_EXT = {
    "drawing": {".pdf", ".doc", ".docx"},
    "customer_ref": {".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx"},
    "nesting": {".pdf", ".xlsx", ".xls", ".doc", ".docx"},
    "nesting_price": {".pdf", ".xlsx", ".xls", ".doc", ".docx"},
    "costing": {".xlsx", ".xls", ".pdf"},
    "costing_prev": {".xlsx", ".xls", ".pdf"},
    "revision": {".pdf", ".jpg", ".jpeg", ".png", ".xlsx", ".xls", ".doc", ".docx"},
}

_gridfs: Optional[AsyncIOMotorGridFSBucket] = None


def _fs() -> AsyncIOMotorGridFSBucket:
    global _gridfs
    if _gridfs is None:
        _gridfs = AsyncIOMotorGridFSBucket(db, bucket_name="bom_attachments")
    return _gridfs


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ext(name: str) -> str:
    return os.path.splitext(name or "")[1].lower()


def _validate_drawing_pdf(target_dno: str, pdf_bytes: bytes) -> dict:
    """Validasi isi PDF terhadap nomor drawing yang terdaftar.

    Cek per halaman (page 1 → last). Return:
      - {status: 'match', page, note}                       → PDF OK
      - {status: 'no_text', note}                           → PDF tanpa text layer (izinkan dgn warning)
      - {status: 'mismatch', candidates, note}              → PDF punya teks tapi drawing_no tidak ditemukan
      - {status: 'no_target', note}                         → BOM tidak punya drawing_no untuk dibandingkan
      - {status: 'pdf_error', note}                         → PDF corrupt / tidak bisa dibaca
    """
    if not target_dno:
        return {"status": "no_target", "note": "BOM tidak memiliki nomor drawing untuk dicocokkan"}
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    except Exception as e:  # corrupt / encrypted / not a real pdf
        return {"status": "pdf_error", "note": f"PDF tidak bisa dibuka: {e}"}

    target = _normalize_dno(target_dno)
    if not target:
        return {"status": "no_target", "note": "Nomor drawing register kosong setelah normalisasi"}

    all_text_chunks = []
    extracted_any = False
    total_pages = len(reader.pages)
    # Safety cap: scan maks 15 halaman (drawing biasanya di halaman awal / kop surat)
    max_pages = min(total_pages, 15)
    for i in range(max_pages):
        try:
            txt = reader.pages[i].extract_text() or ""
        except Exception:
            txt = ""
        if txt.strip():
            extracted_any = True
            all_text_chunks.append(txt)
            if target in _normalize_dno(txt):
                return {
                    "status": "match",
                    "page": i + 1,
                    "note": f"Nomor drawing '{target_dno}' ditemukan di halaman {i + 1}",
                }

    if not extracted_any:
        return {
            "status": "no_text",
            "note": "PDF tidak memiliki text layer (kemungkinan hasil scan/gambar). Validasi otomatis dilewati.",
        }

    # Kumpulkan kandidat nomor drawing-like untuk membantu user (max 8)
    joined = "\n".join(all_text_chunks)
    candidates = set()
    for m in re.findall(r"[A-Za-z0-9][A-Za-z0-9\-/_.]{4,30}[A-Za-z0-9]", joined):
        norm = _normalize_dno(m)
        if 5 <= len(norm) <= 30 and any(c.isalpha() for c in norm) and any(c.isdigit() for c in norm):
            candidates.add(m.strip())
    return {
        "status": "mismatch",
        "candidates": sorted(candidates)[:8],
        "note": f"Nomor drawing '{target_dno}' tidak ditemukan di isi PDF (dicek {max_pages} halaman)",
    }


@router.get("/bom/{bom_id}/attachments")
async def list_attachments(bom_id: str, current: dict = Depends(get_current_user)):
    docs = await db.bom_attachments.find(
        {"bom_id": bom_id, "deleted_at": {"$exists": False}},
        {"_id": 0, "file_id": 0},
    ).sort("uploaded_at", -1).to_list(length=500)
    # Grouped view (backward compat for existing consumers) — all valid categories
    grouped = {k: [] for k in VALID_CATEGORIES}
    show_costing = can_view_costing(current)
    flat = []
    for d in docs:
        cat = d.get("category")
        # RBAC: sembunyikan file costing/price dari role non-privileged.
        if cat in PRICE_ATTACHMENT_CATEGORIES and not show_costing:
            continue
        if cat in grouped:
            grouped[cat].append(d)
        flat.append(d)
    return {
        "bom_id": bom_id,
        "attachments": grouped,   # legacy
        "items": flat,            # flat list — preferred for new code
        "total": len(flat),
        "can_view_costing": show_costing,
        "drawing_preview_only": is_drawing_preview_only(current),
    }


@router.post("/bom/{bom_id}/attachments")
async def upload_attachment(
    bom_id: str,
    category: str = Form(...),
    remark: str = Form(""),
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    if category not in VALID_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Category harus salah satu: {list(VALID_CATEGORIES)}")

    bom = await db.boms.find_one({"id": bom_id, "deleted_at": {"$exists": False}}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM tidak ditemukan")

    ext = _ext(file.filename)
    allowed = CATEGORY_ALLOWED_EXT.get(category, set())
    if allowed and ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Ekstensi {ext} tidak diizinkan untuk kategori {category}. Boleh: {sorted(allowed)}")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File > 50 MB tidak diizinkan")

    # ---- Validasi isi PDF Drawing terhadap nomor drawing yang terdaftar ----
    # Hanya berlaku untuk category='drawing' + .pdf. Jika mismatch → tolak.
    # Jika PDF tidak punya text layer → izinkan dgn warning.
    validation_info = None
    validation_warning = None
    if category == "drawing" and ext == ".pdf":
        target_dno = (bom.get("drawing_no") or bom.get("project_dwg") or "").strip()
        val = _validate_drawing_pdf(target_dno, content)
        validation_info = val
        status = val.get("status")
        if status == "mismatch":
            raise HTTPException(
                status_code=400,
                detail={
                    "message": f"Isi PDF tidak cocok dengan nomor drawing terdaftar. Nomor '{target_dno}' tidak ditemukan di isi PDF.",
                    "target": target_dno,
                    "candidates": val.get("candidates", []),
                    "hint": "Upload ulang PDF drawing yang benar (harus mengandung nomor drawing di kop surat).",
                    "validation": val,
                },
            )
        if status == "pdf_error":
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "PDF tidak bisa dibaca (mungkin corrupt/encrypted).",
                    "hint": "Coba buka & simpan ulang PDF lalu upload lagi.",
                    "validation": val,
                },
            )
        if status == "no_text":
            validation_warning = val.get("note")
        # status 'match' atau 'no_target' → lanjutkan tanpa warning

    fs = _fs()
    file_id = await fs.upload_from_stream(
        file.filename,
        content,
        metadata={"content_type": file.content_type, "bom_id": bom_id, "category": category},
    )

    user_name = current.get("username") or current.get("name")
    doc = {
        "id": str(uuid.uuid4()),
        "bom_id": bom_id,
        "so_no": bom.get("so_no"),
        "revision": bom.get("revision"),
        "category": category,
        "filename": file.filename,
        "file_id": str(file_id),
        "content_type": file.content_type,
        "size_bytes": len(content),
        "remark": remark or "",
        "uploaded_at": _now_iso(),
        "uploaded_by": user_name,
    }
    if validation_info is not None:
        doc["pdf_validation"] = validation_info
    await db.bom_attachments.insert_one(doc.copy())
    await log_action(current, "bom_attachment_upload", "bom", bom_id, {"category": category, "filename": file.filename})
    doc.pop("file_id", None)
    resp = {"success": True, "attachment": doc}
    if validation_warning:
        resp["warning"] = validation_warning
    if validation_info and validation_info.get("status") == "match":
        resp["validation"] = {"status": "match", "page": validation_info.get("page"), "note": validation_info.get("note")}
    return resp


@router.delete("/bom/{bom_id}/attachments/{attach_id}")
async def delete_attachment(bom_id: str, attach_id: str, current: dict = Depends(get_current_user)):
    doc = await db.bom_attachments.find_one({"id": attach_id, "bom_id": bom_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="Attachment tidak ditemukan")
    # Delete GridFS file
    try:
        await _fs().delete(ObjectId(doc["file_id"]))
    except Exception:
        pass
    await db.bom_attachments.update_one(
        {"id": attach_id},
        {"$set": {"deleted_at": _now_iso(), "deleted_by": current.get("username")}},
    )
    await log_action(current, "bom_attachment_delete", "bom", bom_id, {"attach_id": attach_id})
    return {"success": True}


async def _stream_from_gridfs(file_id_str: str):
    fs = _fs()
    stream = await fs.open_download_stream(ObjectId(file_id_str))
    return stream


def _extract_costing_report(xlsx_bytes: bytes) -> dict:
    """Extract structured cost report dari sheet REPORT.
    Label-based scan: cari kata kunci → ambil value dari cell terdekat (right/below)."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel tidak bisa dibaca: {e}")

    # Locate the REPORT sheet (fallback ke sheet pertama)
    ws = None
    for s in wb.worksheets:
        if s.title.strip().upper() == "REPORT":
            ws = s; break
    if ws is None:
        ws = wb.worksheets[0]

    # Build label→cell lookup
    label_map = {}  # canonical_label → (row, col)
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip():
                # Normalise: uppercase, remove extra spaces
                norm = " ".join(v.strip().upper().split())
                label_map.setdefault(norm, (cell.row, cell.column))

    def find_number_near(label_norm):
        """Cari label lalu ambil angka dari cell di sebelah kanan (utk header/text) atau sebelahnya."""
        pos = label_map.get(label_norm)
        if not pos: return None
        r, c = pos
        # Try cells: right up to +6 columns, then +1 row same col
        for dc in range(1, 8):
            try:
                v = ws.cell(row=r, column=c + dc).value
                if v is None or v == "": continue
                if isinstance(v, (int, float)): return float(v)
                if isinstance(v, str) and v.strip():
                    # sometimes labels span merged cells — skip pure text
                    try: return float(str(v).replace(",", "").replace(" ", ""))
                    except Exception: return v.strip()
            except Exception: continue
        # Try same column below
        for dr in range(1, 4):
            try:
                v = ws.cell(row=r + dr, column=c).value
                if isinstance(v, (int, float)): return float(v)
            except Exception: continue
        return None

    def find_str_near(label_norm):
        v = find_number_near(label_norm)
        return v if isinstance(v, str) else (str(int(v)) if isinstance(v, float) and v == int(v) else (str(v) if v is not None else None))

    # Header fields
    header = {
        "project_name": find_str_near("PROJECT NAME"),
        "client": find_str_near("CLIENT"),
        "qty": find_number_near("QTY UNIT (REV)") or find_number_near("QTY"),
        "date": find_str_near("DATE"),
        "drawing_ref": find_str_near("DRAWING REF") or find_str_near("MKS DWG NO"),
        "prepared_by": find_str_near("PREPARED BY"),
        "checked_by": find_str_near("CHECKED BY"),
        "approved_by": find_str_near("APPROVED BY"),
    }
    # Convert numeric header values back if they slipped through
    if isinstance(header["qty"], float) and header["qty"] == int(header["qty"]):
        header["qty"] = int(header["qty"])

    # SECTIONS per Excel real structure (A-E direct/procurement, F=grand, G-J adjustments, K=all, L=per pc)
    section_a = [
        ("RAW MATERIAL (STEEL)", "Raw Material (Steel)"),
        ("RAW MATERIAL (OTHER)", "Raw Material (Other)"),
        ("SCRAP RETURN", "Scrap Return"),
        ("STD PARTS / MECHANICAL PARTS", "Std Parts / Mechanical"),
    ]
    section_b = [
        ("DIRECT LABOUR", "Direct Labour"),
        ("CONSUMABLES", "Consumables"),
    ]
    section_c = [
        ("INDIRECT LABOUR", "Indirect Labour"),
        ("DESIGN / DRW / ENGINNERING", "Design / DRW / Engineering"),
        ("DESIGN / DRW / ENGINEERING", "Design / DRW / Engineering"),
    ]
    section_d = [
        ("CUTTING / SAWING SHEARING / LASER / ETC", "Cutting / Sawing / Shearing / Laser"),
        ("CUTTING / SAWING / SHEARING / LASER / ETC", "Cutting / Sawing / Shearing / Laser"),
        ("ROLLING / BENDING", "Rolling / Bending"),
    ]
    section_e = [
        ("MOB & DEMOB", "Mob & Demob"),
        ("MOB DEMOB", "Mob & Demob"),
    ]

    def to_num(v):
        if isinstance(v, (int, float)): return float(v)
        if isinstance(v, str):
            try: return float(v.replace(",", "").replace(" ", "").replace("Rp", ""))
            except Exception: return 0.0
        return 0.0

    def build_section(spec_list):
        seen = set()
        items = []
        for (key, lbl) in spec_list:
            if lbl in seen: continue
            v = to_num(find_number_near(key))
            if v == 0 and lbl in [i["label"] for i in items]:
                continue
            # dedup same label (some sections have variants)
            items.append({"label": lbl, "value": v})
            seen.add(lbl)
        return items

    a_items = build_section(section_a)
    b_items = build_section(section_b)
    c_items = build_section(section_c)
    d_items = build_section(section_d)
    e_items = build_section(section_e)

    a_total = sum(x["value"] for x in a_items)
    b_total = sum(x["value"] for x in b_items)
    c_total = sum(x["value"] for x in c_items)
    d_total = sum(x["value"] for x in d_items)
    e_total = sum(x["value"] for x in e_items)

    # Get percentage adjustments (G, H, I, J)
    safety_margin = to_num(find_number_near("COST SAFETY MARGIN"))
    profit = to_num(find_number_near("COST PROFIT"))
    marketing = to_num(find_number_near("COST MARKETING FEE"))
    fee_customer = to_num(find_number_near("COST FEE FOR CUSTOMER"))

    total_cost = to_num(find_number_near("TOTAL COST (A+B+C+D+E)")) or (a_total + b_total + c_total + d_total + e_total)
    all_total = to_num(find_number_near("TOTAL ALL COST (F+G+H+I+J) OR SELLING PRICE FOR ALL QTY")) or total_cost
    selling_per_pc = to_num(find_number_near("SELLING PRICE PER PC")) or all_total

    return {
        "sheet_used": ws.title,
        "header": header,
        "section_a": {"title": "A. Procurement All Materials", "items": a_items, "subtotal": a_total},
        "section_b": {"title": "B. Direct Cost", "items": b_items, "subtotal": b_total},
        "section_c": {"title": "C. In-Direct Cost", "items": c_items, "subtotal": c_total},
        "section_d": {"title": "D. Subcontractor", "items": d_items, "subtotal": d_total},
        "section_e": {"title": "E. Miscellaneous", "items": e_items, "subtotal": e_total},
        "adjustments": {
            "safety_margin": safety_margin,
            "profit": profit,
            "marketing_fee": marketing,
            "fee_customer": fee_customer,
        },
        "totals": {
            "total_cost": total_cost,
            "all_total": all_total,
            "selling_price_per_pc": selling_per_pc,
        },
    }


@router.get("/bom/{bom_id}/attachments/{attach_id}/costing-summary")
async def get_costing_summary(bom_id: str, attach_id: str, current: dict = Depends(get_current_user)):
    """Struktur cost report yang diekstrak dari sheet REPORT — tampil rapi di popup, no need to open Excel."""
    doc = await db.bom_attachments.find_one({"id": attach_id, "bom_id": bom_id, "deleted_at": {"$exists": False}})
    if not doc:
        raise HTTPException(status_code=404, detail="Attachment tidak ditemukan")
    if doc.get("category") != "costing":
        raise HTTPException(status_code=400, detail="File ini bukan costing")
    grid_out = await _fs().open_download_stream(ObjectId(doc["file_id"]))
    raw = await grid_out.read()
    return _extract_costing_report(raw)


def _excel_to_html(xlsx_bytes: bytes, orig_name: str) -> str:
    """Convert Excel bytes → styled HTML dengan openpyxl.
    Preserve: column widths, merged cells, multi-sheet tabs. Support A4 Portrait/Landscape view toggle."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from html import escape
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel tidak bisa dibaca: {e}")

    sheets_html = []
    tab_labels = []
    for idx, ws in enumerate(wb.worksheets):
        tab_labels.append(escape(ws.title))
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1

        # Column widths (default Excel unit ≈ 8 chars ≈ 64px)
        col_widths_px = []
        for col_i in range(1, max_col + 1):
            letter = get_column_letter(col_i)
            w = ws.column_dimensions.get(letter)
            width = w.width if w and w.width else 10
            col_widths_px.append(max(30, int(width * 7)))  # ~7px per Excel unit
        colgroup = "<colgroup>" + "".join(f'<col style="width:{w}px">' for w in col_widths_px) + "</colgroup>"

        # Merged cells → build lookup: (row, col) → {rowspan, colspan} + set of hidden cells
        merged_info = {}
        hidden = set()
        for mr in ws.merged_cells.ranges:
            r1, c1, r2, c2 = mr.min_row, mr.min_col, mr.max_row, mr.max_col
            merged_info[(r1, c1)] = {"rowspan": r2 - r1 + 1, "colspan": c2 - c1 + 1}
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    if (rr, cc) != (r1, c1): hidden.add((rr, cc))

        # Build rows
        rows_html = []
        for r in range(1, max_row + 1):
            # Row height
            rh = ws.row_dimensions.get(r)
            row_h_attr = f' style="height:{int(rh.height * 1.33)}px"' if rh and rh.height else ""
            cells = []
            for c in range(1, max_col + 1):
                if (r, c) in hidden: continue
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if val is None: txt = ""
                elif isinstance(val, float):
                    if val == int(val): txt = f"{int(val):,}"
                    else: txt = f"{val:,.4f}".rstrip("0").rstrip(".")
                elif isinstance(val, int): txt = f"{val:,}"
                else: txt = escape(str(val))
                merge = merged_info.get((r, c), {})
                extra = ""
                if merge.get("rowspan", 1) > 1: extra += f' rowspan="{merge["rowspan"]}"'
                if merge.get("colspan", 1) > 1: extra += f' colspan="{merge["colspan"]}"'
                # Alignment
                align = "left"
                if cell.alignment:
                    if cell.alignment.horizontal in ("center",): align = "center"
                    elif cell.alignment.horizontal == "right": align = "right"
                elif isinstance(val, (int, float)): align = "right"
                # Bold detection
                bold = cell.font and cell.font.bold
                style = f'text-align:{align};{"font-weight:700;" if bold else ""}'
                cells.append(f'<td{extra} style="{style}">{txt}</td>')
            rows_html.append(f'<tr{row_h_attr}>' + "".join(cells) + "</tr>")

        sheets_html.append(
            f'<div class="sheet" id="sheet-{idx}" style="display:{"block" if idx==0 else "none"}">'
            f'<table>{colgroup}{"".join(rows_html)}</table></div>'
        )

    tabs_bar = "".join(
        f'<button onclick="show({i})" class="tab {"active" if i==0 else ""}" id="tab-{i}">{lbl}</button>'
        for i, lbl in enumerate(tab_labels)
    )

    css = """
    <style>
      html, body { margin:0; padding:0; background:#e5e7eb; font-family: Calibri, "Segoe UI", Roboto, sans-serif; color:#111827; }
      .toolbar { position: sticky; top: 0; background:#0f172a; color:#fff; padding:6px 10px; font-size:11px; z-index:20; display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
      .toolbar .name { font-family: monospace; opacity:0.85; }
      .toolbar .btn { background:#1e293b; border:1px solid #334155; color:#fff; padding:3px 8px; font-size:10px; cursor:pointer; }
      .toolbar .btn.active { background:#059669; border-color:#10b981; }
      .toolbar .btn:hover { background:#334155; }
      .tabs { background:#fff; border-bottom:1px solid #cbd5e1; padding:2px 6px; overflow-x:auto; white-space:nowrap; position: sticky; top: 32px; z-index: 15; }
      .tab { border:0; background:transparent; padding:5px 10px; font-size:11px; font-weight:600; color:#475569; cursor:pointer; border-bottom:2px solid transparent; }
      .tab.active { color:#0f172a; border-bottom-color:#059669; background:#ecfdf5; }
      .viewport { padding:12px; overflow:auto; }
      .sheet { display:block; margin:0 auto; background:#fff; box-shadow: 0 2px 8px rgba(0,0,0,.08); }
      /* A4 modes */
      .viewport.a4-portrait .sheet { width: 210mm; padding: 10mm; }
      .viewport.a4-landscape .sheet { width: 297mm; padding: 10mm; }
      .viewport.fit .sheet { padding: 8px; }
      table { border-collapse: collapse; table-layout: fixed; font-size:11px; background:#fff; }
      th, td { border:1px solid #d1d5db; padding:2px 5px; vertical-align: middle; overflow:hidden; word-wrap:break-word; }
      td { background:#fff; }
    </style>
    """
    js = """
    <script>
      function show(i){
        document.querySelectorAll('.sheet').forEach((s,idx)=>{ s.style.display = idx===i?'block':'none'; });
        document.querySelectorAll('.tab').forEach((t,idx)=>{ t.classList.toggle('active', idx===i); });
      }
      function setView(mode){
        const vp = document.getElementById('viewport');
        vp.className = 'viewport ' + mode;
        document.querySelectorAll('.view-btn').forEach(b => b.classList.toggle('active', b.dataset.mode === mode));
      }
    </script>
    """
    body = f"""
      <div class="toolbar">
        <span>📊 Excel Preview</span>
        <span class="name">{escape(orig_name)}</span>
        <span style="margin-left:8px; opacity:0.7">{len(tab_labels)} sheet(s)</span>
        <span style="flex:1"></span>
        <button class="btn view-btn active" data-mode="fit" onclick="setView('fit')">🖥 Fit</button>
        <button class="btn view-btn" data-mode="a4-portrait" onclick="setView('a4-portrait')">🖨 A4 Portrait</button>
        <button class="btn view-btn" data-mode="a4-landscape" onclick="setView('a4-landscape')">🖨 A4 Landscape</button>
      </div>
      <div class="tabs">{tabs_bar}</div>
      <div id="viewport" class="viewport fit">
        {"".join(sheets_html)}
      </div>
    """
    return f"<!DOCTYPE html><html><head><meta charset='utf-8'><title>{escape(orig_name)}</title>{css}</head><body>{body}{js}</body></html>"


async def _attachment_pdf_bytes(bom_id: str, attach_id: str, current: dict = None) -> bytes:
    """Ambil bytes attachment & pastikan bentuk PDF.

    - PDF → dipakai apa adanya.
    - Excel (xlsx/xls/xlsm/...) → dikonversi ke PDF via LibreOffice (akurat "sesuai hasil").
    """
    doc = await _get_attachment_or_404(bom_id, attach_id)
    if current is not None:
        _guard_costing_access(doc, current)
    ext = _ext(doc["filename"]).lstrip(".").lower()
    stream = await _stream_from_gridfs(doc["file_id"])
    raw = await stream.read()
    if ext == "pdf":
        return raw
    from utils.office_render import is_office_ext, office_to_pdf
    if is_office_ext(ext):
        return office_to_pdf(raw, ext)
    raise HTTPException(status_code=400, detail=f"Preview gambar belum didukung untuk ekstensi .{ext}")


@router.get("/bom/{bom_id}/attachments/{attach_id}/page-meta")
async def attachment_page_meta(bom_id: str, attach_id: str, current: dict = Depends(get_current_user)):
    """Metadata halaman untuk viewer image-based (PDF & Excel)."""
    from utils.pdf_render import pdf_page_meta
    raw = await _attachment_pdf_bytes(bom_id, attach_id, current)
    return pdf_page_meta(raw)


@router.get("/bom/{bom_id}/attachments/{attach_id}/page-image")
async def attachment_page_image(bom_id: str, attach_id: str, page: int = 0, scale: float = 2.0,
                                current: dict = Depends(get_current_user)):
    """Render satu halaman lampiran (PDF/Excel) menjadi PNG untuk viewer image-based."""
    from utils.pdf_render import pdf_page_png
    raw = await _attachment_pdf_bytes(bom_id, attach_id, current)
    try:
        png = pdf_page_png(raw, page, scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ditemukan")
    return StreamingResponse(io.BytesIO(png), media_type="image/png",
                             headers={"Cache-Control": "private, max-age=300"})



@router.get("/bom/{bom_id}/attachments/{attach_id}/preview")
async def preview_attachment(bom_id: str, attach_id: str, current: dict = Depends(get_current_user)):
    """Inline preview. PDF: native. Excel: convert to PDF first."""
    doc = await _get_attachment_or_404(bom_id, attach_id)
    _guard_costing_access(doc, current)
    ext = _ext(doc["filename"])
    stream = await _stream_from_gridfs(doc["file_id"])
    raw = await stream.read()

    if ext == ".pdf":
        return StreamingResponse(
            io.BytesIO(raw),
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="{doc["filename"]}"'},
        )
    if ext in (".xlsx", ".xls"):
        html = _excel_to_html(raw, doc["filename"])
        return StreamingResponse(
            io.BytesIO(html.encode("utf-8")),
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f'inline; filename="{os.path.splitext(doc["filename"])[0]}.html"'},
        )
    raise HTTPException(status_code=400, detail=f"Preview belum didukung untuk ekstensi {ext}")


@router.get("/bom/{bom_id}/attachments/{attach_id}/download")
async def download_attachment(bom_id: str, attach_id: str, current: dict = Depends(get_current_user)):
    doc = await _get_attachment_or_404(bom_id, attach_id)
    _guard_costing_access(doc, current)      # role non-privileged tidak boleh unduh costing/price
    _guard_drawing_download(doc, current)     # QC/DocControl/Store/Produksi: DWG/Customer preview-only
    stream = await _stream_from_gridfs(doc["file_id"])
    raw = await stream.read()
    return StreamingResponse(
        io.BytesIO(raw),
        media_type=doc.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{doc["filename"]}"'},
    )
