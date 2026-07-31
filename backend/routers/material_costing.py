"""Engineering Costing Reference Database — Material Master (Raw Material).

Bukan inventory/stock. Ini catalog harga referensi Engineering untuk costing manual.
Purchasing input harga → auto-compute weight & harga per Kg (via density table) →
Engineering pakai untuk estimasi project.
"""
from __future__ import annotations
import io
import math
import re
import uuid
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from db import db
from deps import get_current_user, is_admin_like, log_action

router = APIRouter(tags=["material-costing"])


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _clean(d):
    if d:
        d.pop("_id", None)
    return d


def _can_edit(user: dict) -> bool:
    # Full CRUD: super_admin, admin, purchasing, eng_leader/eng_head (Riski)
    role = (user or {}).get("role", "")
    return role in ("purchasing", "admin", "super_admin", "eng_leader", "eng_head")


def _can_view(user: dict) -> bool:
    role = (user or {}).get("role", "")
    return role in (
        "purchasing", "admin", "super_admin", "supervisor", "finance",
        "eng_leader", "eng_head", "eng_staff", "engineering", "sales",
    )


# ============ DENSITY REFERENCE (seed) ============
DEFAULT_DENSITY = {
    # Carbon Steel — Structural
    "ASTM A36": 7.85, "SS400": 7.85, "S235JR": 7.85, "S275JR": 7.85, "S355JR": 7.85,
    "ASTM A572 GRADE 50": 7.85, "ASTM A516 GRADE 70": 7.85, "ASTM A283 GRADE C": 7.85,
    # Carbon Steel — Pipe
    "ASTM A106 GRADE B": 7.85, "ASTM A53 GRADE B": 7.85,
    "API 5L GRADE B": 7.85, "API 5L X42": 7.85, "API 5L X52": 7.85,
    "API 5L X60": 7.85, "API 5L X65": 7.85, "API 5L X70": 7.85,
    # Medium Carbon
    "S45C": 7.85, "C45": 7.85,
    "AISI 1018": 7.87, "AISI 1020": 7.87, "AISI 1045": 7.85,
    "AISI 4140": 7.85, "AISI 4340": 7.85,
    # Wear-resistant / Abrasion (Hardox / AR)
    "HARDOX 400": 7.85, "HARDOX 450": 7.85, "HARDOX 500": 7.85,
    "HARDOX 550": 7.85, "HARDOX 600": 7.85,
    "AR200": 7.85, "AR400": 7.85, "AR500": 7.85,
    # Stainless
    "SS304": 7.93, "SS304L": 7.93, "SS316": 7.98, "SS316L": 7.98,
    "SS310": 7.90, "SS321": 7.90, "SS410": 7.75, "SS420": 7.78, "SS430": 7.70,
    # Legacy SUS aliases (backward compat)
    "SUS304": 7.93, "SUS316": 7.98, "SUS316L": 7.98, "SUS201": 7.80,
    # Aluminium
    "ALUMINIUM 1050": 2.71, "ALUMINIUM 5052": 2.68, "ALUMINIUM 5083": 2.66,
    "ALUMINIUM 6061": 2.70, "ALUMINIUM 6063": 2.70, "ALUMINIUM 7075": 2.81,
    "AL 6061": 2.70, "AL 5052": 2.68, "AL 1100": 2.71, "AL 5083": 2.66,
    # Copper / Brass / Bronze
    "COPPER C110": 8.96, "COPPER C11000": 8.94,
    "BRASS C260": 8.53, "BRONZE C932": 8.80,
    # Cast Iron
    "CAST IRON": 7.30, "DUCTILE IRON": 7.10,
    "FC200": 7.20, "FCD450": 7.10,
    # Galvanized
    "GALVANIZED STEEL": 7.85,
    # Others (legacy)
    "Q235B": 7.85, "Q345B": 7.85, "ST37": 7.85, "ST52": 7.85,
    "TITANIUM GR2": 4.51,
}


class MaterialIn(BaseModel):
    """Input for creating/updating a material entry."""
    category: str = "raw_material"  # raw_material | standard_part | consumable | subcon
    material_type: str = ""  # Plate / Pipe / Angle / etc (raw); or subtype for others
    grade: str = ""          # Material grade OR name for non-raw
    size_description: str = ""  # Free text — e.g. "4' × 8' × 5mm" atau "M12x40" atau "5 Ltr Kaleng"
    # Dimensions (optional — for raw material auto weight compute)
    length_mm: Optional[float] = None
    width_mm: Optional[float] = None
    thickness_mm: Optional[float] = None
    outer_diameter_mm: Optional[float] = None
    wall_thickness_mm: Optional[float] = None
    # Pricing (universal)
    price_per_unit: float = 0     # Harga UTUH (per lembar/batang/piece/kaleng/lumpsum)
    unit: str = "sheet"           # sheet | bar | roll | piece | kaleng | box | lot | m2 | jam
    currency: str = "IDR"         # IDR | USD | SGD | EUR | CNY | JPY | MYR
    exchange_rate: float = 1.0    # currency → IDR conversion (1 for IDR)
    weight_kg: Optional[float] = None  # kalau None → auto-compute (raw material only)
    density_g_cm3: Optional[float] = None  # kalau None → lookup dari grade
    # Markup (untuk ongkir/margin)
    markup_pct: float = 0
    # Supplier
    supplier_name: str = ""
    remark: str = ""
    # === Standard Parts extra fields ===
    catalog_code: str = ""        # e.g. "BLT-M12-40-SS304"
    brand: str = ""               # e.g. "Unbrako", "HILTI"
    moq: Optional[float] = None   # Minimum Order Qty
    # === Consumables & Paint extra fields ===
    pack_size: str = ""           # e.g. "5 Ltr / kaleng", "20 Kg / sak"
    # === Subcon Rate Card extra fields ===
    service_name: str = ""        # e.g. "Sandblasting SA 2.5", "Machining CNC"
    rate_unit: str = ""           # lumpsum | per_item | m2 | jam | kg


def _compute_weight_kg(m: dict, density: float) -> Optional[float]:
    """Compute theoretical weight if dimensions provided. Returns None if insufficient data."""
    if not density or density <= 0:
        return None
    mtype = (m.get("material_type") or "").upper()
    L = m.get("length_mm") or 0
    W = m.get("width_mm") or 0
    T = m.get("thickness_mm") or 0
    OD = m.get("outer_diameter_mm") or 0
    t = m.get("wall_thickness_mm") or 0

    # Convert to cm for density g/cm³ → kg
    if "PLATE" in mtype or "SHEET" in mtype or "PLAT" in mtype:
        if L > 0 and W > 0 and T > 0:
            vol_cm3 = (L / 10) * (W / 10) * (T / 10)
            return round(vol_cm3 * density / 1000, 3)
    if "PIPE" in mtype or "PIPA" in mtype:
        if OD > 0 and t > 0 and L > 0:
            r_out = OD / 2 / 10
            r_in = (OD - 2 * t) / 2 / 10
            vol_cm3 = math.pi * (r_out**2 - r_in**2) * (L / 10)
            return round(vol_cm3 * density / 1000, 3)
    if "ROUND" in mtype or "AS " in mtype:
        if OD > 0 and L > 0:
            r = OD / 2 / 10
            vol_cm3 = math.pi * r * r * (L / 10)
            return round(vol_cm3 * density / 1000, 3)
    if "SQUARE BAR" in mtype:
        if W > 0 and L > 0:
            vol_cm3 = (W / 10) ** 2 * (L / 10)
            return round(vol_cm3 * density / 1000, 3)
    return None


def _enrich(m: dict) -> dict:
    """Compute derived fields: density, weight_kg, price_per_kg, final_price_per_kg.
    For non raw_material categories, skip density/weight logic (just apply markup).
    Also compute IDR-equivalent prices for cross-currency comparison.
    """
    price = float(m.get("price_per_unit") or 0)
    markup = float(m.get("markup_pct") or 0)
    currency = (m.get("currency") or "IDR").upper()
    rate = float(m.get("exchange_rate") or 0)
    if currency == "IDR" or rate <= 0:
        rate = 1.0
    m["currency"] = currency
    m["exchange_rate"] = rate

    m["final_price_per_unit"] = round(price * (1 + markup / 100), 2)
    m["price_per_unit_idr"] = round(price * rate, 2)
    m["final_price_per_unit_idr"] = round(m["final_price_per_unit"] * rate, 2)

    category = (m.get("category") or "raw_material").lower()
    if category != "raw_material":
        # Non-raw: no density/weight math
        m["density_g_cm3"] = m.get("density_g_cm3") or 0
        m["weight_kg"] = m.get("weight_kg") or 0
        m["price_per_kg"] = 0
        m["final_price_per_kg"] = 0
        m["price_per_kg_idr"] = 0
        m["final_price_per_kg_idr"] = 0
        return m

    grade = (m.get("grade") or "").strip().upper()
    density = m.get("density_g_cm3") or DEFAULT_DENSITY.get(grade)
    if not density:
        for k, v in DEFAULT_DENSITY.items():
            if grade.startswith(k.split()[0].upper()):
                density = v
                break
    m["density_g_cm3"] = density or 7.85  # default steel

    # Weight
    if not m.get("weight_kg"):
        wc = _compute_weight_kg(m, m["density_g_cm3"])
        if wc:
            m["weight_kg"] = wc
            m["weight_source"] = "computed"
    else:
        m["weight_source"] = m.get("weight_source") or "manual"

    # Price per Kg (in original currency AND IDR)
    weight = float(m.get("weight_kg") or 0)
    m["price_per_kg"] = round(price / weight, 2) if (weight > 0) else 0
    m["final_price_per_kg"] = round(m["price_per_kg"] * (1 + markup / 100), 2)
    m["price_per_kg_idr"] = round(m["price_per_kg"] * rate, 2)
    m["final_price_per_kg_idr"] = round(m["final_price_per_kg"] * rate, 2)
    return m


# ============ ENDPOINTS ============
@router.get("/material-costing/density-table")
async def get_density_table(current: dict = Depends(get_current_user)):
    """Reference density table (grade → g/cm³). Merged with any DB overrides."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    overrides = await db.material_density.find({}, {"_id": 0}).to_list(length=200)
    ov_map = {o["grade"].upper(): o["density_g_cm3"] for o in overrides}
    merged = dict(DEFAULT_DENSITY)
    merged.update(ov_map)
    return [{"grade": g, "density_g_cm3": d, "source": "override" if g.upper() in ov_map else "default"}
            for g, d in sorted(merged.items())]


@router.post("/material-costing/density-table")
async def upsert_density(payload: dict, current: dict = Depends(get_current_user)):
    """Add/update a density override for a grade."""
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Purchasing/Engineering only")
    grade = (payload.get("grade") or "").strip().upper()
    density = float(payload.get("density_g_cm3") or 0)
    if not grade or density <= 0:
        raise HTTPException(status_code=400, detail="grade & density_g_cm3 wajib")
    await db.material_density.update_one(
        {"grade": grade},
        {"$set": {"grade": grade, "density_g_cm3": density, "updated_at": _now_iso(),
                  "updated_by": current.get("username") or current.get("name")}},
        upsert=True,
    )
    return {"success": True, "grade": grade, "density_g_cm3": density}


@router.get("/material-costing/materials")
async def list_materials(
    q: Optional[str] = None,
    category: Optional[str] = None,
    material_type: Optional[str] = None,
    grade: Optional[str] = None,
    limit: int = 500,
    current: dict = Depends(get_current_user),
):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    filt = {"deleted_at": {"$exists": False}}
    if category and category != "all":
        filt["category"] = category
    if material_type:
        filt["material_type"] = {"$regex": material_type, "$options": "i"}
    if grade:
        filt["grade"] = {"$regex": grade, "$options": "i"}
    if q and q.strip():
        import re
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        filt["$or"] = [
            {"material_type": rx}, {"grade": rx}, {"size_description": rx},
            {"supplier_name": rx}, {"remark": rx},
            {"catalog_code": rx}, {"brand": rx}, {"pack_size": rx},
            {"service_name": rx},
        ]
    docs = await db.materials_costing.find(filt, {"_id": 0}).sort("updated_at", -1).limit(limit).to_list(length=limit)
    return {"items": docs, "total": len(docs)}


@router.get("/material-costing/suppliers")
async def list_suppliers(q: str = "", limit: int = 30, current: dict = Depends(get_current_user)):
    """Distinct supplier names dari materials_costing — untuk autocomplete input."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    match = {"deleted_at": {"$exists": False}, "supplier_name": {"$nin": [None, ""]}}
    if q and q.strip():
        match["supplier_name"] = {"$regex": re.escape(q.strip()), "$options": "i"}
    pipeline = [
        {"$match": match},
        {"$group": {"_id": "$supplier_name", "count": {"$sum": 1}, "last_used": {"$max": "$updated_at"}}},
        {"$project": {"_id": 0, "name": "$_id", "count": 1, "last_used": 1}},
        {"$sort": {"count": -1, "name": 1}},
        {"$limit": limit},
    ]
    items = await db.materials_costing.aggregate(pipeline).to_list(length=limit)
    return {"items": items}


@router.get("/material-costing/price-summary")
async def price_summary(
    category: str = "raw_material",
    current: dict = Depends(get_current_user),
):
    """Rata-rata harga per kg (raw_material) atau per unit (others) — grouped by grade + material_type.
    Menggunakan final_price_*_idr agar konsisten multi-currency."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    pipeline = [
        {"$match": {"category": category, "deleted_at": {"$exists": False}}},
    ]
    price_field = "final_price_per_kg_idr" if category == "raw_material" else "final_price_per_unit_idr"
    pipeline += [
        {"$group": {
            "_id": {"grade": "$grade", "material_type": "$material_type"},
            "count": {"$sum": 1},
            "avg_price": {"$avg": f"${price_field}"},
            "min_price": {"$min": f"${price_field}"},
            "max_price": {"$max": f"${price_field}"},
            "last_updated": {"$max": "$price_last_updated"},
            "last_updated_at": {"$max": "$updated_at"},
        }},
        {"$project": {
            "_id": 0,
            "grade": "$_id.grade",
            "material_type": "$_id.material_type",
            "count": 1,
            "avg_price": 1,
            "min_price": 1,
            "max_price": 1,
            "last_updated": {"$ifNull": ["$last_updated", "$last_updated_at"]},
        }},
        {"$sort": {"grade": 1, "material_type": 1}},
    ]
    rows = await db.materials_costing.aggregate(pipeline).to_list(length=500)
    # Round for cleanliness
    for r in rows:
        for k in ("avg_price", "min_price", "max_price"):
            v = r.get(k)
            r[k] = round(float(v), 0) if v is not None else None
    return {"category": category, "unit": "kg" if category == "raw_material" else "unit", "items": rows}


@router.post("/material-costing/materials")
async def create_material(payload: MaterialIn, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Purchasing/Engineering only")
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = _now_iso()
    doc["created_by"] = current.get("username") or current.get("name")
    doc["updated_at"] = _now_iso()
    doc["updated_by"] = doc["created_by"]
    doc["price_history"] = [{
        "changed_at": _now_iso(), "changed_by": doc["created_by"],
        "price_per_unit": payload.price_per_unit, "markup_pct": payload.markup_pct,
        "supplier_name": payload.supplier_name, "note": "initial",
    }]
    doc["price_last_updated"] = _now_iso()
    _enrich(doc)
    await db.materials_costing.insert_one(doc.copy())
    await log_action(current, "material_costing_create", "material_costing", doc["id"], {})
    return _clean(doc)


@router.put("/material-costing/materials/{material_id}")
async def update_material(material_id: str, payload: MaterialIn, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Purchasing/Engineering only")
    existing = await db.materials_costing.find_one({"id": material_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Material tidak ditemukan")
    new_doc = payload.model_dump()
    new_doc["id"] = material_id
    new_doc["created_at"] = existing.get("created_at")
    new_doc["created_by"] = existing.get("created_by")
    new_doc["updated_at"] = _now_iso()
    new_doc["updated_by"] = current.get("username") or current.get("name")
    # Append to price history if price/markup changed
    history = existing.get("price_history") or []
    price_changed = float(existing.get("price_per_unit") or 0) != float(payload.price_per_unit or 0)
    markup_changed = float(existing.get("markup_pct") or 0) != float(payload.markup_pct or 0)
    if price_changed or markup_changed:
        history.append({
            "changed_at": _now_iso(), "changed_by": new_doc["updated_by"],
            "price_per_unit": payload.price_per_unit, "markup_pct": payload.markup_pct,
            "supplier_name": payload.supplier_name, "note": "update",
        })
        new_doc["price_last_updated"] = _now_iso()
    else:
        # Preserve existing price_last_updated
        new_doc["price_last_updated"] = existing.get("price_last_updated") or existing.get("created_at") or _now_iso()
    new_doc["price_history"] = history
    _enrich(new_doc)
    await db.materials_costing.replace_one({"id": material_id}, new_doc)
    await log_action(current, "material_costing_update", "material_costing", material_id, {})
    return _clean(new_doc)


@router.delete("/material-costing/materials/{material_id}")
async def delete_material(material_id: str, current: dict = Depends(get_current_user)):
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Tidak berwenang menghapus material costing")
    r = await db.materials_costing.update_one(
        {"id": material_id, "deleted_at": {"$exists": False}},
        {"$set": {"deleted_at": _now_iso(), "deleted_by": current.get("username")}},
    )
    if r.modified_count == 0:
        raise HTTPException(status_code=404, detail="Tidak ditemukan")
    await log_action(current, "material_costing_delete", "material_costing", material_id, {})
    return {"success": True}


class PriceUpdateIn(BaseModel):
    price_per_unit: float
    markup_pct: Optional[float] = None
    supplier_name: Optional[str] = None
    currency: Optional[str] = None
    exchange_rate: Optional[float] = None
    note: Optional[str] = ""


@router.post("/material-costing/materials/{material_id}/update-price")
async def update_price(material_id: str, payload: PriceUpdateIn, current: dict = Depends(get_current_user)):
    """Quick price update — only updates price/markup/supplier/currency + appends to history + sets price_last_updated."""
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Purchasing/Engineering only")
    existing = await db.materials_costing.find_one({"id": material_id, "deleted_at": {"$exists": False}})
    if not existing:
        raise HTTPException(status_code=404, detail="Material tidak ditemukan")
    if payload.price_per_unit is None or float(payload.price_per_unit) < 0:
        raise HTTPException(status_code=400, detail="price_per_unit wajib")

    user_name = current.get("username") or current.get("name")
    new_price = float(payload.price_per_unit)
    new_markup = float(payload.markup_pct) if payload.markup_pct is not None else float(existing.get("markup_pct") or 0)
    new_supplier = payload.supplier_name if payload.supplier_name is not None else existing.get("supplier_name", "")
    new_currency = (payload.currency or existing.get("currency") or "IDR").upper()
    new_rate = float(payload.exchange_rate) if payload.exchange_rate is not None else float(existing.get("exchange_rate") or 1)
    if new_currency == "IDR" or new_rate <= 0:
        new_rate = 1.0

    history = existing.get("price_history") or []
    history.append({
        "changed_at": _now_iso(),
        "changed_by": user_name,
        "price_per_unit_old": float(existing.get("price_per_unit") or 0),
        "price_per_unit": new_price,
        "markup_pct_old": float(existing.get("markup_pct") or 0),
        "markup_pct": new_markup,
        "supplier_name": new_supplier,
        "currency_old": existing.get("currency") or "IDR",
        "currency": new_currency,
        "exchange_rate": new_rate,
        "note": (payload.note or "").strip() or "quick update",
    })

    update = {
        "price_per_unit": new_price,
        "markup_pct": new_markup,
        "supplier_name": new_supplier,
        "currency": new_currency,
        "exchange_rate": new_rate,
        "price_last_updated": _now_iso(),
        "updated_at": _now_iso(),
        "updated_by": user_name,
        "price_history": history,
    }
    # Recompute derived fields
    merged = {**existing, **update}
    _enrich(merged)
    for k in ("price_per_kg", "final_price_per_unit", "final_price_per_kg",
              "price_per_unit_idr", "final_price_per_unit_idr",
              "price_per_kg_idr", "final_price_per_kg_idr"):
        if k in merged:
            update[k] = merged[k]

    await db.materials_costing.update_one({"id": material_id}, {"$set": update})
    await log_action(current, "material_costing_price_update", "material_costing", material_id, {
        "old_price": existing.get("price_per_unit"), "new_price": new_price,
        "currency": new_currency,
    })
    return {"success": True, **update}


@router.get("/material-costing/materials/{material_id}/price-history")
async def get_price_history(material_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    doc = await db.materials_costing.find_one({"id": material_id, "deleted_at": {"$exists": False}}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Tidak ditemukan")
    history = doc.get("price_history") or []
    return {
        "material_id": material_id,
        "grade": doc.get("grade"),
        "size_description": doc.get("size_description"),
        "current_price": doc.get("price_per_unit"),
        "current_markup": doc.get("markup_pct"),
        "current_supplier": doc.get("supplier_name"),
        "price_last_updated": doc.get("price_last_updated"),
        "history": sorted(history, key=lambda h: h.get("changed_at", ""), reverse=True),
    }


# ============ EXCEL TEMPLATE (per category) + IMPORT ============
TEMPLATE_HEADERS = {
    "raw_material": [
        "material_type", "grade", "size_description",
        "length_mm", "width_mm", "thickness_mm", "outer_diameter_mm", "wall_thickness_mm",
        "weight_kg", "price_per_unit", "unit", "currency", "exchange_rate",
        "markup_pct", "supplier_name", "remark",
    ],
    "standard_part": [
        "material_type", "grade", "catalog_code", "brand", "moq",
        "size_description", "price_per_unit", "unit", "currency", "exchange_rate",
        "markup_pct", "supplier_name", "remark",
    ],
    "consumable": [
        "material_type", "grade", "brand", "pack_size",
        "size_description", "price_per_unit", "unit", "currency", "exchange_rate",
        "markup_pct", "supplier_name", "remark",
    ],
    "subcon": [
        "material_type", "service_name", "size_description", "rate_unit",
        "price_per_unit", "unit", "currency", "exchange_rate",
        "markup_pct", "supplier_name", "remark",
    ],
}

TEMPLATE_SAMPLES = {
    "raw_material": [
        ["Plate", "ASTM A36", "4' x 8' x 5mm", 2440, 1220, 5, "", "", "", 10000000, "sheet", "IDR", 1, 5, "PT ABC Steel", ""],
        ["Pipe", "ASTM A106 GRADE B", "OD 60.3 x 3.2 x 6M", 6000, "", "", 60.3, 3.2, "", 850000, "bar", "IDR", 1, 0, "PT Pipa Baja", ""],
        ["Round Bar", "AISI 4140", "Dia. 16mm x 6M", 6000, "", "", 16, "", "", 150, "bar", "USD", 16000, 8, "Overseas Supplier", ""],
    ],
    "standard_part": [
        ["Bolt/Baut", "Hex Bolt M12 x 40 SS304", "BLT-M12-40-SS304", "Unbrako", 100, "M12x40 DIN933", 15000, "pcs", "IDR", 1, 10, "PT Baut Jaya", ""],
    ],
    "consumable": [
        ["Cat/Paint", "Epoxy Primer Grey", "Jotun", "5 Ltr", "Ratio mix 4:1", 850000, "kaleng", "IDR", 1, 5, "PT Cat Prima", ""],
    ],
    "subcon": [
        ["Sandblasting", "Sandblast SA 2.5", "Steel struktur", "m2", 75000, "m2", "IDR", 1, 0, "CV Blast Kilat", ""],
    ],
}


def _import_openpyxl():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    return openpyxl, Font, PatternFill, Alignment, get_column_letter


@router.get("/material-costing/materials/template/xlsx")
async def download_template(category: str = Query("raw_material"), current: dict = Depends(get_current_user)):
    """Download blank Excel template for bulk import — headers sesuai kategori."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    if category not in TEMPLATE_HEADERS:
        raise HTTPException(status_code=400, detail=f"Category tidak valid. Pilih: {list(TEMPLATE_HEADERS.keys())}")

    openpyxl, Font, PatternFill, Alignment, get_column_letter = _import_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = category.upper()[:31]

    headers = TEMPLATE_HEADERS[category]
    samples = TEMPLATE_SAMPLES.get(category, [])

    # Instructions row
    ws.cell(row=1, column=1, value="INSTRUKSI: Isi mulai baris 4. Baris 3 = contoh, boleh dihapus. Kolom bertanda * wajib.").font = Font(bold=True, color="B45309")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 22

    # Header row (row 2)
    required = {"material_type", "grade", "size_description", "price_per_unit", "service_name"}
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h + (" *" if h in required else ""))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Sample rows (row 3+)
    for r_idx, sample in enumerate(samples, start=3):
        for c_idx, val in enumerate(sample, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(italic=True, color="94A3B8")
            c.fill = PatternFill("solid", fgColor="F8FAFC")

    # Column widths
    for col_idx, h in enumerate(headers, 1):
        width = max(12, min(28, len(h) + 6))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header
    ws.freeze_panes = "A3"

    # Reference sheet: valid values
    ws2 = wb.create_sheet("REFERENSI")
    ws2.cell(row=1, column=1, value="Referensi Nilai Valid").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="Currency:").font = Font(bold=True)
    for i, c in enumerate(["IDR", "USD", "SGD", "EUR", "CNY", "JPY", "MYR"], start=1):
        ws2.cell(row=3, column=1 + i, value=c)
    ws2.cell(row=4, column=1, value="Unit:").font = Font(bold=True)
    for i, u in enumerate(["sheet", "bar", "roll", "piece", "meter", "kaleng", "box", "pack", "kg", "liter", "sak", "pcs", "set", "lot", "m2", "jam"], start=1):
        ws2.cell(row=4, column=1 + i, value=u)
    if category == "raw_material":
        ws2.cell(row=5, column=1, value="Material Type:").font = Font(bold=True)
        for i, t in enumerate(["Plate", "Pipe", "Round Bar", "Square Bar", "Hollow Square", "Hollow Rect", "Angle L", "Channel U", "H-Beam", "WF", "IWF", "Wire Mesh", "Sheet"], start=1):
            ws2.cell(row=5, column=1 + i, value=t)
        ws2.cell(row=7, column=1, value="Density Table (grade → g/cm³)").font = Font(bold=True)
        overrides = await db.material_density.find({}, {"_id": 0}).to_list(length=500)
        merged = dict(DEFAULT_DENSITY)
        for o in overrides:
            merged[o["grade"].upper()] = o["density_g_cm3"]
        r = 8
        for g, d in sorted(merged.items()):
            ws2.cell(row=r, column=1, value=g)
            ws2.cell(row=r, column=2, value=d)
            r += 1

    # Set column widths for ref sheet
    ws2.column_dimensions["A"].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"template_material_costing_{category}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/material-costing/materials/import/xlsx")
async def import_xlsx(
    category: str = Form("raw_material"),
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    """Bulk import materials from Excel. Format sesuai template."""
    if not _can_edit(current):
        raise HTTPException(status_code=403, detail="Purchasing/Engineering only")
    if category not in TEMPLATE_HEADERS:
        raise HTTPException(status_code=400, detail=f"Category tidak valid")

    openpyxl, *_ = _import_openpyxl()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File Excel tidak valid: {e}")

    # Prefer sheet matching category name, else first sheet
    ws = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == category.lower() or sn.strip().upper() == category.upper():
            ws = wb[sn]
            break
    if ws is None:
        ws = wb.worksheets[0]

    headers = TEMPLATE_HEADERS[category]
    required = {"material_type", "grade", "size_description", "price_per_unit", "service_name"}

    # Detect header row: find row where cells match headers list (case-insensitive, stripped, ignore trailing "*")
    def _clean_header(v):
        s = str(v or "").strip().lower().rstrip("*").strip()
        return s

    header_row = None
    header_map = {}
    for r in range(1, min(10, ws.max_row + 1)):
        cells = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        cleaned = [_clean_header(v) for v in cells]
        if sum(1 for h in headers if h in cleaned) >= max(3, len(headers) // 2):
            header_row = r
            for c_idx, name in enumerate(cleaned, 1):
                if name in headers:
                    header_map[name] = c_idx
            break

    if header_row is None:
        raise HTTPException(status_code=400, detail="Header tidak ditemukan. Pakai template resmi.")

    user_name = current.get("username") or current.get("name")
    created = 0
    errors = []
    docs_to_insert = []

    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        for h, c_idx in header_map.items():
            row[h] = ws.cell(row=r, column=c_idx).value

        # Skip fully empty
        if not any(row.values()):
            continue

        # Skip italic-style sample rows (best-effort: if all values look like defaults from TEMPLATE_SAMPLES)
        # We'll leave that up to user (they should delete sample rows manually)

        # Validate required
        missing = [k for k in required if k in headers and (row.get(k) in (None, ""))]
        if missing:
            errors.append({"row": r, "reason": f"missing: {', '.join(missing)}"})
            continue

        # Coerce numbers
        def num(v, default=None):
            if v in (None, ""): return default
            try: return float(v)
            except (ValueError, TypeError): return default

        payload = {
            "category": category,
            "material_type": str(row.get("material_type") or "").strip(),
            "grade": str(row.get("grade") or "").strip(),
            "size_description": str(row.get("size_description") or "").strip(),
            "length_mm": num(row.get("length_mm")),
            "width_mm": num(row.get("width_mm")),
            "thickness_mm": num(row.get("thickness_mm")),
            "outer_diameter_mm": num(row.get("outer_diameter_mm")),
            "wall_thickness_mm": num(row.get("wall_thickness_mm")),
            "weight_kg": num(row.get("weight_kg")),
            "price_per_unit": num(row.get("price_per_unit"), 0) or 0,
            "unit": str(row.get("unit") or "").strip() or "sheet",
            "currency": (str(row.get("currency") or "IDR").strip().upper() or "IDR"),
            "exchange_rate": num(row.get("exchange_rate"), 1) or 1,
            "markup_pct": num(row.get("markup_pct"), 0) or 0,
            "supplier_name": str(row.get("supplier_name") or "").strip(),
            "remark": str(row.get("remark") or "").strip(),
            # Category-specific
            "catalog_code": str(row.get("catalog_code") or "").strip(),
            "brand": str(row.get("brand") or "").strip(),
            "moq": num(row.get("moq")),
            "pack_size": str(row.get("pack_size") or "").strip(),
            "service_name": str(row.get("service_name") or "").strip(),
            "rate_unit": str(row.get("rate_unit") or "").strip(),
        }

        # For subcon: grade defaults to service_name if empty
        if category == "subcon" and not payload["grade"] and payload["service_name"]:
            payload["grade"] = payload["service_name"]

        payload["id"] = str(uuid.uuid4())
        payload["created_at"] = _now_iso()
        payload["created_by"] = user_name
        payload["updated_at"] = _now_iso()
        payload["updated_by"] = user_name
        payload["price_last_updated"] = _now_iso()
        payload["price_history"] = [{
            "changed_at": _now_iso(), "changed_by": user_name,
            "price_per_unit": payload["price_per_unit"], "markup_pct": payload["markup_pct"],
            "supplier_name": payload["supplier_name"],
            "currency": payload["currency"], "exchange_rate": payload["exchange_rate"],
            "note": f"import xlsx (row {r})",
        }]
        _enrich(payload)
        docs_to_insert.append(payload)

    if docs_to_insert:
        await db.materials_costing.insert_many([d.copy() for d in docs_to_insert])
        created = len(docs_to_insert)

    await log_action(current, "material_costing_import_xlsx", "material_costing", None, {
        "category": category, "created": created, "errors": len(errors),
    })
    return {"success": True, "created": created, "errors": errors, "total_rows_scanned": ws.max_row - header_row}
