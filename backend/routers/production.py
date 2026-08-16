"""Production module — Fase 1: visibilitas SO baru untuk Produksi.

Produksi mendapat daftar Sales Order (SO) sejak SO dibuat (walau drawing belum
di-stamp Doc Control), lengkap dengan penanda apakah drawing/BOM sudah ada, dan
bisa 'acknowledge' (tandai sudah dilihat/disiapkan).
"""
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel

from db import db
from deps import get_current_user, log_action, is_production, is_admin_like, is_qc

router = APIRouter(prefix="/production", tags=["production"])


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _can_view(user: dict) -> bool:
    return is_production(user) or is_admin_like(user)


def _can_qc(user: dict) -> bool:
    """QC staff/head (role 'qc') atau admin-like boleh approve/reject Release Note."""
    return is_qc(user) or is_admin_like(user)


@router.get("/new-so")
async def list_new_so(scope: str = "unack", current: dict = Depends(get_current_user)):
    """Daftar SO untuk Produksi.
    scope: 'unack' (belum di-acknowledge) | 'all'.
    Mengembalikan info SO + apakah drawing/BOM sudah tersedia (konteks kesiapan).
    """
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")

    q = {"deleted_at": {"$exists": False}}
    if scope == "unack":
        q["prod_ack"] = {"$ne": True}

    sos = await db.sales_orders.find(q, {"_id": 0}).sort("created_at", -1).limit(300).to_list(length=300)

    items = []
    for so in sos:
        so_no = so.get("so_no") or ""
        # Konteks kesiapan: apakah sudah ada drawing / BOM untuk SO ini
        has_drawing = False
        has_bom = False
        if so_no:
            has_drawing = (await db.drawings.count_documents({"so_no": so_no, "deleted_at": {"$exists": False}})) > 0
            has_bom = (await db.boms.count_documents({"so_no": so_no, "deleted_at": {"$exists": False}})) > 0
        # Qty total = jumlah qty semua item pada SO
        so_items = so.get("items") or []
        qty_total = 0.0
        for it in so_items:
            try:
                qty_total += float(it.get("qty", 0) or 0)
            except (TypeError, ValueError):
                pass
        # Deskripsi: pakai field description, fallback gabungan nama item
        desc = so.get("description") or ""
        if not desc and so_items:
            desc = ", ".join([str(it.get("name") or "").strip() for it in so_items if it.get("name")])
        # Item list per baris + total qty release (progress "siap X pcs")
        item_list = [{"name": (it.get("name") or "").strip(), "qty": _f(it.get("qty")), "unit": it.get("unit") or ""} for it in so_items]
        released = 0.0
        if so_no:
            for fr in await db.fg_release_notes.find({"so_no": so_no}, {"_id": 0, "qty": 1}).to_list(length=10000):
                released += _f(fr.get("qty"))
        items.append({
            "id": so.get("id"),
            "so_no": so_no,
            "so_date": so.get("so_date") or (so.get("created_at") or "")[:10],
            "customer": so.get("customer") or "",
            "description": desc,
            "items": item_list,
            "qty_total": qty_total,
            "released": released,
            "balance": max(0, qty_total - released),
            "source_quotation_no": so.get("source_quotation_no") or "",
            "created_at": so.get("created_at") or "",
            "created_by_username": so.get("created_by_username") or "",
            "has_drawing": has_drawing,
            "has_bom": has_bom,
            "prod_ack": bool(so.get("prod_ack")),
            "prod_ack_at": so.get("prod_ack_at") or "",
            "prod_ack_by": so.get("prod_ack_by") or "",
            "prod_started": bool(so.get("prod_started")),
            "prod_started_at": so.get("prod_started_at") or "",
            "prod_started_by": so.get("prod_started_by") or "",
        })

    unack_count = await db.sales_orders.count_documents({"deleted_at": {"$exists": False}, "prod_ack": {"$ne": True}})
    return {"items": items, "count": len(items), "unack_count": unack_count, "scope": scope}


@router.get("/so-attachments")
async def so_attachments(so_no: str = "", current: dict = Depends(get_current_user)):
    """Pratinjau lampiran SO untuk Produksi: daftar Drawing (PDF) + BOM (beserta item).
    Dipakai di halaman SO Masuk agar operator bisa cek drawing/BOM tanpa pindah menu."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    so = (so_no or "").strip()
    if not so:
        return {"drawings": [], "boms": []}
    dwg_docs = await db.drawings.find(
        {"so_no": so, "deleted_at": {"$exists": False}}, {"_id": 0}
    ).sort("updated_at", -1).to_list(length=200)
    drawings = [{
        "id": d.get("id"),
        "drawing_no": d.get("drawing_no") or "-",
        "title": d.get("title") or d.get("project_name") or "",
        "revision": d.get("revision") or "",
        "discipline": d.get("discipline") or "",
        "approval_status": d.get("approval_status") or "draft",
        "has_file": bool(d.get("file_id")),
    } for d in dwg_docs]

    bom_docs = await db.boms.find(
        {"so_no": so, "deleted_at": {"$exists": False}}, {"_id": 0}
    ).sort("uploaded_at", 1).to_list(length=100)
    boms = []
    for idx, b in enumerate(bom_docs):
        if not b.get("bom_no"):
            continue
        its = b.get("items") or []
        boms.append({
            "id": b.get("id"),
            "bom_no": b.get("bom_no"),
            "part_no": b.get("part_no") or (idx + 1),
            "project_name": b.get("project_name") or "",
            "items_count": len(its),
            "items": [{
                "item_no": it.get("item_no"),
                "item_name": it.get("item_name") or "",
                "item_specification": it.get("item_specification") or "",
                "qty": it.get("qty") or 0,
                "uom": it.get("uom") or "",
                "material": it.get("material") or "",
                "weight_kg": it.get("weight_kg"),
                "remark": it.get("remark") or "",
            } for it in its],
        })
    return {"so_no": so, "drawings": drawings, "boms": boms}


@router.post("/new-so/{so_id}/ack")
async def ack_new_so(so_id: str, current: dict = Depends(get_current_user)):
    """Tandai SO sudah dilihat/disiapkan Produksi."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa acknowledge")
    so = await db.sales_orders.find_one({"id": so_id, "deleted_at": {"$exists": False}})
    if not so:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    await db.sales_orders.update_one(
        {"id": so_id},
        {"$set": {
            "prod_ack": True,
            "prod_ack_at": _now_iso(),
            "prod_ack_by": current.get("name") or current.get("username") or "",
        }},
    )
    await log_action(current, "prod_ack_so", "sales_order", so_id, {"so_no": so.get("so_no")})
    return {"ok": True}


@router.post("/new-so/{so_id}/unack")
async def unack_new_so(so_id: str, current: dict = Depends(get_current_user)):
    """Batalkan acknowledge (kembalikan ke daftar SO baru)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah")
    so = await db.sales_orders.find_one({"id": so_id, "deleted_at": {"$exists": False}})
    if not so:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    await db.sales_orders.update_one(
        {"id": so_id},
        {"$set": {"prod_ack": False}, "$unset": {"prod_ack_at": "", "prod_ack_by": ""}},
    )
    await log_action(current, "prod_unack_so", "sales_order", so_id, {"so_no": so.get("so_no")})
    return {"ok": True}


class StartWorkIn(BaseModel):
    start_date: str = ""


@router.post("/new-so/{so_id}/start")
async def start_work_so(so_id: str, payload: StartWorkIn = None, current: dict = Depends(get_current_user)):
    """Tandai Mulai Kerja dengan tanggal yang dipilih Produksi."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa menandai mulai kerja")
    so = await db.sales_orders.find_one({"id": so_id, "deleted_at": {"$exists": False}})
    if not so:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    sd = _date_only(payload.start_date) if (payload and payload.start_date) else datetime.now(timezone.utc).strftime("%Y-%m-%d")
    started_at = f"{sd}T00:00:00+00:00"
    await db.sales_orders.update_one(
        {"id": so_id},
        {"$set": {
            "prod_started": True,
            "prod_started_at": started_at,
            "prod_started_by": current.get("name") or current.get("username") or "",
        }},
    )
    await log_action(current, "prod_start_so", "sales_order", so_id, {"so_no": so.get("so_no"), "start_date": sd})
    return {"ok": True}


@router.post("/new-so/{so_id}/unstart")
async def unstart_work_so(so_id: str, current: dict = Depends(get_current_user)):
    """Batalkan status Mulai Kerja."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah")
    so = await db.sales_orders.find_one({"id": so_id, "deleted_at": {"$exists": False}})
    if not so:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    await db.sales_orders.update_one(
        {"id": so_id},
        {"$set": {"prod_started": False}, "$unset": {"prod_started_at": "", "prod_started_by": ""}},
    )
    await log_action(current, "prod_unstart_so", "sales_order", so_id, {"so_no": so.get("so_no")})
    return {"ok": True}



# ==========================================================================
# Daily Production Report
# Satu baris = satu record di koleksi `production_reports`.
# 1 tanggal bisa punya banyak baris (operator/SO/proses berbeda).
# ==========================================================================

class ProductionReportIn(BaseModel):
    report_date: str = ""          # YYYY-MM-DD
    operator_name: str = ""
    so_no: str = ""
    customer: str = ""
    process: str = ""
    qty_ok: float = 0
    qty_ng: float = 0
    work_start: str = ""           # HH:MM
    work_end: str = ""             # HH:MM
    machine_no: str = ""
    remarks: str = ""


def _f(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _work_hours(start: str, end: str) -> float:
    """Durasi kerja (jam) dari HH:MM start–end; lintas tengah malam didukung."""
    try:
        sh, sm = [int(x) for x in (start or "").split(":")[:2]]
        eh, em = [int(x) for x in (end or "").split(":")[:2]]
        mins = (eh * 60 + em) - (sh * 60 + sm)
        if mins < 0:
            mins += 24 * 60
        return round(mins / 60.0, 2)
    except Exception:
        return 0.0


# ===== Shift & pemisahan jam Normal vs Lembur =====
DEFAULT_SHIFT = {"shift1_start": "08:00", "shift1_end": "16:00", "shift2_start": "16:00", "shift2_end": "24:00"}


async def _get_shift_settings() -> dict:
    doc = await db.production_shift_settings.find_one({"_id": "default"}) or {}
    s = {**DEFAULT_SHIFT}
    for k in DEFAULT_SHIFT:
        if doc.get(k):
            s[k] = doc[k]
    return s


def _min_of(t: str) -> int:
    try:
        h, m = [int(x) for x in (t or "").split(":")[:2]]
        return h * 60 + m
    except Exception:
        return 0


def _split_normal_ot(ws: str, we: str, shift: int, s: dict):
    """Bagi jam kerja jadi (normal, lembur). Shift 1 normal 08–16, Shift 2 dari setting.
    Jam kerja di luar jendela normal shift dihitung lembur."""
    a = _min_of(ws); b = _min_of(we)
    if b <= a:
        b += 1440  # lintas tengah malam
    total = b - a
    if total <= 0:
        return (0.0, 0.0)
    if shift == 2:
        ns = _min_of(s.get("shift2_start", "16:00")); ne = _min_of(s.get("shift2_end", "24:00"))
    else:
        ns = _min_of(s.get("shift1_start", "08:00")); ne = _min_of(s.get("shift1_end", "16:00"))
    if ne <= ns:
        ne += 1440
    lo = max(a, ns); hi = min(b, ne)
    normal = max(0, hi - lo)
    ot = total - normal
    return (round(normal / 60.0, 2), round(ot / 60.0, 2))


async def _nightshift_set(operators: set, dates: set) -> set:
    """Kumpulan (operator_name_lower, date) yang berstatus Night Shift (Shift 2)."""
    if not operators or not dates:
        return set()
    emps = await db.production_employees.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(length=5000)
    name_to_id = {(e.get("name") or "").strip().lower(): e.get("id") for e in emps}
    id_to_name = {e.get("id"): (e.get("name") or "").strip().lower() for e in emps}
    att = await db.production_attendance.find(
        {"date": {"$in": list(dates)}, "status": "night_shift"}, {"_id": 0, "employee_id": 1, "date": 1}
    ).to_list(length=100000)
    out = set()
    for a in att:
        nm = id_to_name.get(a.get("employee_id"))
        if nm:
            out.add((nm, a.get("date")))
    return out


class ShiftSettingsIn(BaseModel):
    shift1_start: str = "08:00"
    shift1_end: str = "16:00"
    shift2_start: str = "16:00"
    shift2_end: str = "24:00"


def _serialize_report(r: dict) -> dict:
    ws = r.get("work_start") or ""
    we = r.get("work_end") or ""
    return {
        "id": r.get("id"),
        "report_date": r.get("report_date") or "",
        "operator_name": r.get("operator_name") or "",
        "so_no": r.get("so_no") or "",
        "customer": r.get("customer") or "",
        "process": r.get("process") or "",
        "qty_ok": _f(r.get("qty_ok")),
        "qty_ng": _f(r.get("qty_ng")),
        "work_start": ws,
        "work_end": we,
        "work_hours": _work_hours(ws, we),
        "machine_no": r.get("machine_no") or "",
        "remarks": r.get("remarks") or "",
        "created_by_username": r.get("created_by_username") or "",
        "created_at": r.get("created_at") or "",
        "updated_at": r.get("updated_at") or "",
    }


@router.get("/report-options")
async def report_options(current: dict = Depends(get_current_user)):
    """Data pendukung form: daftar operator (user Produksi + histori), mesin,
    proses, remark yang pernah dipakai, dan daftar SO (untuk auto-isi customer)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")

    # Operator: dari master karyawan produksi (data absensi) + user Produksi + histori
    op_set = set()
    prod_emps = await db.production_employees.find(
        {"active": {"$ne": False}}, {"_id": 0, "name": 1},
    ).to_list(length=5000)
    for e in prod_emps:
        nm = (e.get("name") or "").strip()
        if nm:
            op_set.add(nm)
    prod_users = await db.users.find(
        {"role": {"$in": ["production", "produksi"]}, "active": {"$ne": False}},
        {"_id": 0, "name": 1, "username": 1},
    ).to_list(length=500)
    for u in prod_users:
        nm = (u.get("name") or u.get("username") or "").strip()
        if nm:
            op_set.add(nm)
    for nm in await db.production_reports.distinct("operator_name"):
        if nm and str(nm).strip():
            op_set.add(str(nm).strip())

    machines = sorted({str(m).strip() for m in await db.production_reports.distinct("machine_no") if m and str(m).strip()})
    processes = sorted({str(p).strip() for p in await db.production_reports.distinct("process") if p and str(p).strip()})
    remarks = sorted({str(x).strip() for x in await db.production_reports.distinct("remarks") if x and str(x).strip()})

    # SO options — untuk auto-isi customer saat pilih SO
    sos = await db.sales_orders.find(
        {"deleted_at": {"$exists": False}},
        {"_id": 0, "so_no": 1, "customer": 1},
    ).sort("so_no", 1).to_list(length=5000)
    so_opts = []
    seen = set()
    for s in sos:
        sn = (s.get("so_no") or "").strip()
        if sn and sn not in seen:
            seen.add(sn)
            so_opts.append({"so_no": sn, "customer": s.get("customer") or ""})

    return {
        "operators": sorted(op_set),
        "machines": machines,
        "processes": processes,
        "remarks": remarks,
        "sos": so_opts,
    }


@router.get("/today-summary")
async def today_summary(current: dict = Depends(get_current_user)):
    """Ringkasan 'Hari Ini' untuk panel portal Produksi:
    - berapa karyawan belum diabsen hari ini
    - apakah sudah ada laporan produksi hari ini
    - berapa Release Note ditolak QC (perlu diajukan ulang)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    emps = await db.production_employees.find({"active": {"$ne": False}}, {"_id": 0, "id": 1}).to_list(length=5000)
    total_emp = len(emps)
    emp_ids = {e.get("id") for e in emps}
    att = await db.production_attendance.find({"date": today}, {"_id": 0, "employee_id": 1}).to_list(length=50000)
    att_ids = {a.get("employee_id") for a in att if a.get("employee_id") in emp_ids}
    attendance_done = len(att_ids)
    attendance_missing = max(0, total_emp - attendance_done)

    report_count = await db.production_reports.count_documents({"report_date": today})
    frn_rejected = await db.fg_release_notes.count_documents({"status": "rejected"})

    return {
        "date": today,
        "total_employees": total_emp,
        "attendance_done": attendance_done,
        "attendance_missing": attendance_missing,
        "report_count": report_count,
        "frn_rejected": frn_rejected,
    }


@router.get("/present-operators")
async def present_operators(date: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Nama operator yang HADIR/bekerja pada tanggal tsb (status tidak diblok).
    Karyawan tanpa record dianggap default Hadir. Dipakai untuk auto-isi baris
    Daily Production Report."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    d = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    emps = await db.production_employees.find({"active": {"$ne": False}}, {"_id": 0, "id": 1, "name": 1}).to_list(length=5000)
    emps.sort(key=lambda r: (r.get("name") or "").lower())
    att = await db.production_attendance.find({"date": d}, {"_id": 0, "employee_id": 1, "status": 1}).to_list(length=50000)
    status_map = {a.get("employee_id"): (a.get("status") or "hadir") for a in att}
    names = []
    for e in emps:
        nm = (e.get("name") or "").strip()
        if not nm:
            continue
        st = status_map.get(e.get("id"), "hadir")  # default hadir bila belum diabsen
        if st in ATTEND_BLOCKED:
            continue
        names.append(nm)
    return {"date": d, "operators": names}



@router.get("/reports/masterlist")
async def reports_masterlist(
    month: Optional[str] = None,       # YYYY-MM
    date: Optional[str] = None,        # YYYY-MM-DD (override month)
    operator: Optional[str] = None,
    so_no: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """Masterlist semua baris report dengan filter bulan/tanggal/operator/SO."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")

    filt = _build_report_filter(month, date, operator, so_no)
    rows = await db.production_reports.find(filt, {"_id": 0}).sort("report_date", -1).to_list(length=100000)
    rows.sort(key=lambda r: (r.get("report_date") or "", r.get("created_at") or ""), reverse=True)
    items = [_serialize_report(r) for r in rows]
    total_ok = sum(i["qty_ok"] for i in items)
    total_ng = sum(i["qty_ng"] for i in items)
    total_work_hours = round(sum(i["work_hours"] for i in items), 2)
    return {"items": items, "count": len(items), "total_ok": total_ok, "total_ng": total_ng, "total_work_hours": total_work_hours}


def _build_report_filter(month, date, operator, so_no) -> dict:
    filt: dict = {}
    if date:
        filt["report_date"] = date
    elif month:
        filt["report_date"] = {"$regex": f"^{month}"}
    if operator:
        filt["operator_name"] = {"$regex": operator, "$options": "i"}
    if so_no:
        filt["so_no"] = {"$regex": so_no, "$options": "i"}
    return filt


@router.get("/shift-settings")
async def get_shift_settings(current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    return {"settings": await _get_shift_settings()}


@router.put("/shift-settings")
async def update_shift_settings(payload: ShiftSettingsIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah")
    doc = {**payload.dict(), "_id": "default", "updated_at": _now_iso()}
    await db.production_shift_settings.update_one({"_id": "default"}, {"$set": doc}, upsert=True)
    return {"ok": True, "settings": await _get_shift_settings()}


@router.get("/so-work-summary")
async def so_work_summary(month: Optional[str] = None, q: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Ringkasan kerja per SO dari Daily Production Report:
    berapa hari, berapa jam, siapa saja operator yang mengerjakan 1 SO."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    filt: dict = {"so_no": {"$nin": ["", None]}}
    if month:
        filt["report_date"] = {"$regex": f"^{month}"}
    if q and q.strip():
        filt["so_no"] = {"$regex": q.strip(), "$options": "i"}
    rows = await db.production_reports.find(filt, {"_id": 0}).to_list(length=200000)
    shift_s = await _get_shift_settings()
    all_ops = {(r.get("operator_name") or "").strip().lower() for r in rows if r.get("operator_name")}
    all_dates = {r.get("report_date") or "" for r in rows if r.get("report_date")}
    nightset = await _nightshift_set(all_ops, all_dates)
    agg: dict = {}
    for r in rows:
        so = (r.get("so_no") or "").strip()
        if not so:
            continue
        rd = r.get("report_date") or ""
        op = (r.get("operator_name") or "").strip()
        hrs = _work_hours(r.get("work_start") or "", r.get("work_end") or "")
        shift = 2 if (op.lower(), rd) in nightset else 1
        normal, ot = _split_normal_ot(r.get("work_start") or "", r.get("work_end") or "", shift, shift_s)
        a = agg.setdefault(so, {"so_no": so, "customer": r.get("customer") or "", "dates": set(),
                                "operators": set(), "machines": set(), "total_hours": 0.0,
                                "normal_hours": 0.0, "ot_hours": 0.0, "has_shift2": False, "qty_ok": 0.0, "qty_ng": 0.0})
        if not a["customer"] and r.get("customer"):
            a["customer"] = r.get("customer")
        if rd:
            a["dates"].add(rd)
        if op:
            a["operators"].add(op)
        mc = (r.get("machine_no") or "").strip()
        if mc:
            a["machines"].add(mc)
        a["total_hours"] += hrs
        a["normal_hours"] += normal
        a["ot_hours"] += ot
        if shift == 2:
            a["has_shift2"] = True
        a["qty_ok"] += _f(r.get("qty_ok"))
        a["qty_ng"] += _f(r.get("qty_ng"))
    items = []
    for a in agg.values():
        dts = sorted(a["dates"])
        items.append({
            "so_no": a["so_no"], "customer": a["customer"],
            "total_days": len(dts), "total_hours": round(a["total_hours"], 2),
            "normal_hours": round(a["normal_hours"], 2), "ot_hours": round(a["ot_hours"], 2),
            "has_shift2": a["has_shift2"],
            "operators_count": len(a["operators"]),
            "operators": sorted(a["operators"]),
            "machines": sorted(a["machines"]),
            "first_date": dts[0] if dts else "", "last_date": dts[-1] if dts else "",
            "qty_ok": a["qty_ok"], "qty_ng": a["qty_ng"],
        })
    items.sort(key=lambda x: (x["last_date"] or "", x["so_no"]), reverse=True)
    return {"items": items, "count": len(items),
            "total_hours": round(sum(i["total_hours"] for i in items), 2)}


@router.get("/so-work-summary/{so_no}")
async def so_work_summary_detail(so_no: str, current: dict = Depends(get_current_user)):
    """Detail kerja 1 SO: rincian per tanggal (siapa + jam) & rekap per operator."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    so = (so_no or "").strip()
    rows = await db.production_reports.find({"so_no": so}, {"_id": 0}).to_list(length=200000)
    rows.sort(key=lambda r: (r.get("report_date") or "", r.get("created_at") or ""))
    customer = ""
    by_date: dict = {}
    by_op: dict = {}
    total_hours = 0.0
    total_normal = 0.0
    total_ot = 0.0
    shift_s = await _get_shift_settings()
    all_ops = {(r.get("operator_name") or "").strip().lower() for r in rows if r.get("operator_name")}
    all_dates = {r.get("report_date") or "" for r in rows if r.get("report_date")}
    nightset = await _nightshift_set(all_ops, all_dates)
    for r in rows:
        if not customer and r.get("customer"):
            customer = r.get("customer")
        rd = r.get("report_date") or ""
        op = (r.get("operator_name") or "").strip()
        hrs = _work_hours(r.get("work_start") or "", r.get("work_end") or "")
        shift = 2 if (op.lower(), rd) in nightset else 1
        normal, ot = _split_normal_ot(r.get("work_start") or "", r.get("work_end") or "", shift, shift_s)
        total_hours += hrs
        total_normal += normal
        total_ot += ot
        d = by_date.setdefault(rd, {"date": rd, "hours": 0.0, "normal": 0.0, "ot": 0.0, "operators": set(), "machines": set(), "rows": []})
        d["hours"] += hrs
        d["normal"] += normal
        d["ot"] += ot
        if op:
            d["operators"].add(op)
        mc = (r.get("machine_no") or "").strip()
        if mc:
            d["machines"].add(mc)
        d["rows"].append({
            "operator_name": op, "process": r.get("process") or "",
            "machine_no": mc, "shift": shift,
            "work_start": r.get("work_start") or "", "work_end": r.get("work_end") or "",
            "work_hours": hrs, "normal_hours": normal, "ot_hours": ot,
            "qty_ok": _f(r.get("qty_ok")), "qty_ng": _f(r.get("qty_ng")),
        })
        o = by_op.setdefault(op or "-", {"name": op or "-", "days": set(), "hours": 0.0})
        if rd:
            o["days"].add(rd)
        o["hours"] += hrs
    dates = sorted(by_date.keys())
    by_date_list = [{
        "date": by_date[k]["date"], "hours": round(by_date[k]["hours"], 2),
        "normal": round(by_date[k]["normal"], 2), "ot": round(by_date[k]["ot"], 2),
        "operators": sorted(by_date[k]["operators"]),
        "machines": sorted(by_date[k]["machines"]),
        "rows": by_date[k]["rows"],
    } for k in dates]
    by_op_list = sorted(
        [{"name": o["name"], "days": len(o["days"]), "hours": round(o["hours"], 2)} for o in by_op.values()],
        key=lambda x: -x["hours"],
    )
    # Finished Goods Release Notes untuk SO ini (dengan saldo berjalan)
    frns = await db.fg_release_notes.find({"so_no": so}, {"_id": 0}).to_list(length=100000)
    frns.sort(key=lambda x: (x.get("frn_date") or "", x.get("created_at") or ""))
    # Qty total SO (jumlah semua item)
    so_doc = await db.sales_orders.find_one({"so_no": so, "deleted_at": {"$exists": False}}, {"_id": 0, "items": 1})
    so_qty = 0.0
    for it in ((so_doc or {}).get("items") or []):
        so_qty += _f(it.get("qty"))
    running = 0.0
    finished_goods = []
    for fr in frns:
        q = _f(fr.get("qty"))
        running += q
        finished_goods.append({
            "frn_date": fr.get("frn_date") or "",
            "release_no": fr.get("release_no") or "",
            "description": fr.get("description") or "",
            "qty": q,
            "qc_comment": fr.get("qc_comment") or "",
            "running_total": round(running, 2),
            "balance": round(max(0.0, so_qty - running), 2) if so_qty else None,
        })
    total_released = round(running, 2)
    return {
        "so_no": so, "customer": customer,
        "total_days": len(dates), "total_hours": round(total_hours, 2),
        "total_normal": round(total_normal, 2), "total_ot": round(total_ot, 2),
        "first_date": dates[0] if dates else "", "last_date": dates[-1] if dates else "",
        "by_date": by_date_list, "by_operator": by_op_list,
        "so_qty": so_qty, "total_released": total_released,
        "balance": round(max(0.0, so_qty - total_released), 2) if so_qty else None,
        "is_finished": bool(so_qty and total_released >= so_qty),
        "finished_goods": finished_goods,
    }


@router.get("/reports/masterlist.xlsx")
async def reports_masterlist_xlsx(
    month: Optional[str] = None,
    date: Optional[str] = None,
    operator: Optional[str] = None,
    so_no: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """Export Excel masterlist report (mengikuti filter yang sama)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")

    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    filt = _build_report_filter(month, date, operator, so_no)
    rows = await db.production_reports.find(filt, {"_id": 0}).to_list(length=100000)
    rows.sort(key=lambda r: (r.get("report_date") or "", r.get("created_at") or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Production Report"
    headers = ["Tanggal", "Operator", "SO No", "Customer", "Process",
               "Qty OK", "Qty NG", "Jam Mulai", "Jam Selesai", "Machine No", "Remarks"]
    ws.append(headers)
    for r in rows:
        rr = _serialize_report(r)
        ws.append([
            rr["report_date"], rr["operator_name"], rr["so_no"], rr["customer"], rr["process"],
            rr["qty_ok"], rr["qty_ng"], rr["work_start"], rr["work_end"], rr["machine_no"], rr["remarks"],
        ])
    widths = [12, 20, 16, 24, 20, 9, 9, 11, 11, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    tag = date or month or datetime.now().strftime("%Y%m")
    fname = f"daily_production_report_{tag}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/reports")
async def list_reports(date: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Daily Production Report — daftar baris untuk 1 tanggal (default: hari ini)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    d = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    rows = await db.production_reports.find({"report_date": d}, {"_id": 0}).to_list(length=100000)
    rows.sort(key=lambda r: r.get("created_at") or "")
    items = [_serialize_report(r) for r in rows]
    total_ok = sum(i["qty_ok"] for i in items)
    total_ng = sum(i["qty_ng"] for i in items)
    total_work_hours = round(sum(i["work_hours"] for i in items), 2)
    return {"date": d, "items": items, "count": len(items), "total_ok": total_ok, "total_ng": total_ng, "total_work_hours": total_work_hours}


@router.post("/reports")
async def create_report(payload: ProductionReportIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa input")
    rd = (payload.report_date or "").strip() or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Blok operator yang absen (Tidak Hadir / MC-Sakit) pada tanggal tsb
    op_name = (payload.operator_name or "").strip()
    if op_name:
        emp = await db.production_employees.find_one({"name": {"$regex": f"^{op_name}$", "$options": "i"}, "active": {"$ne": False}})
        if emp:
            att = await db.production_attendance.find_one({"date": rd, "employee_id": emp.get("id")})
            if att and att.get("status") in ATTEND_BLOCKED:
                label = "MC/Sakit" if att.get("status") == "mc_sakit" else "Tidak Hadir"
                raise HTTPException(status_code=400, detail=f"{op_name} berstatus {label} pada {rd}. Tidak bisa input Daily Production untuk operator ini.")
    doc = {
        "id": str(uuid.uuid4()),
        "report_date": rd,
        "operator_name": (payload.operator_name or "").strip(),
        "so_no": (payload.so_no or "").strip(),
        "customer": (payload.customer or "").strip(),
        "process": (payload.process or "").strip(),
        "qty_ok": _f(payload.qty_ok),
        "qty_ng": _f(payload.qty_ng),
        "work_start": (payload.work_start or "").strip(),
        "work_end": (payload.work_end or "").strip(),
        "machine_no": (payload.machine_no or "").strip(),
        "remarks": (payload.remarks or "").strip(),
        "created_by": current.get("id"),
        "created_by_username": current.get("name") or current.get("username") or "",
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    await db.production_reports.insert_one(doc)
    await log_action(current, "create_production_report", "production_report", doc["id"], {"date": rd, "so_no": doc["so_no"]})
    return _serialize_report(doc)


@router.put("/reports/{report_id}")
async def update_report(report_id: str, payload: ProductionReportIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah")
    existing = await db.production_reports.find_one({"id": report_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Report tidak ditemukan")
    updates = {
        "report_date": (payload.report_date or existing.get("report_date") or "").strip(),
        "operator_name": (payload.operator_name or "").strip(),
        "so_no": (payload.so_no or "").strip(),
        "customer": (payload.customer or "").strip(),
        "process": (payload.process or "").strip(),
        "qty_ok": _f(payload.qty_ok),
        "qty_ng": _f(payload.qty_ng),
        "work_start": (payload.work_start or "").strip(),
        "work_end": (payload.work_end or "").strip(),
        "machine_no": (payload.machine_no or "").strip(),
        "remarks": (payload.remarks or "").strip(),
        "updated_at": _now_iso(),
    }
    await db.production_reports.update_one({"id": report_id}, {"$set": updates})
    await log_action(current, "update_production_report", "production_report", report_id, {"so_no": updates["so_no"]})
    merged = {**existing, **updates}
    return _serialize_report(merged)


@router.delete("/reports/{report_id}")
async def delete_report(report_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa menghapus")
    existing = await db.production_reports.find_one({"id": report_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Report tidak ditemukan")
    await db.production_reports.delete_one({"id": report_id})
    await log_action(current, "delete_production_report", "production_report", report_id, {"so_no": existing.get("so_no")})
    return {"ok": True}


# ==========================================================================
# Finished Goods Release Note (FGRN) + Daily Monitoring Job Progress
# ==========================================================================

def _so_qty_total(so: dict) -> float:
    t = 0.0
    for it in so.get("items") or []:
        try:
            t += float(it.get("qty", 0) or 0)
        except (TypeError, ValueError):
            pass
    return t


def _so_desc(so: dict) -> str:
    d = so.get("description") or ""
    if not d:
        d = ", ".join([str(it.get("name") or "").strip() for it in (so.get("items") or []) if it.get("name")])
    return d


def _date_only(s: str) -> str:
    return (s or "")[:10]


def _days_between(start: str, end: str) -> int:
    """Selisih hari kalender antara dua tanggal ISO (YYYY-MM-DD...)."""
    try:
        a = datetime.fromisoformat(_date_only(start))
        b = datetime.fromisoformat(_date_only(end))
        return max(0, (b - a).days)
    except Exception:
        return 0


def _working_days(start: str, end: str, holidays: set = None) -> int:
    """NETWORKDAYS.INTL weekend=11 (hanya Minggu) minus hari libur nasional.
    Jumlah hari kerja dari start s/d end (inklusif), kecualikan Minggu & tanggal libur."""
    holidays = holidays or set()
    try:
        a = datetime.fromisoformat(_date_only(start))
        b = datetime.fromisoformat(_date_only(end))
    except Exception:
        return 0
    if b < a:
        return 0
    cnt = 0
    cur = a
    guard = 0
    while cur <= b and guard < 3000:
        iso = cur.strftime("%Y-%m-%d")
        if cur.weekday() != 6 and iso not in holidays:  # 6 = Minggu
            cnt += 1
        cur += timedelta(days=1)
        guard += 1
    return cnt


class FrnIn(BaseModel):
    frn_date: str = ""
    release_no: str = ""
    so_no: str = ""
    customer: str = ""
    description: str = ""
    qty: float = 0
    qc_comment: str = ""
    item_index: Optional[int] = None   # item ke-berapa dalam SO (untuk SO multi-item)


def _serialize_frn(r: dict) -> dict:
    return {
        "id": r.get("id"),
        "frn_date": r.get("frn_date") or "",
        "release_no": r.get("release_no") or "",
        "so_no": r.get("so_no") or "",
        "customer": r.get("customer") or "",
        "description": r.get("description") or "",
        "qty": _f(r.get("qty")),
        "qc_comment": r.get("qc_comment") or "",
        "item_index": r.get("item_index"),
        "status": r.get("status") or "released",  # data lama tanpa status dianggap released
        "qc_by": r.get("qc_by") or "",
        "qc_at": r.get("qc_at") or "",
        "qc_signature": r.get("qc_signature") or "",
        "ready_for_delivery": bool(r.get("ready_for_delivery")),
        "delivered": bool(r.get("delivered")),
        "created_by_username": r.get("created_by_username") or "",
        "created_at": r.get("created_at") or "",
    }


@router.get("/so-brief")
async def so_brief(started_only: bool = False, current: dict = Depends(get_current_user)):
    """Ringkasan SO untuk dropdown FGRN / Job Progress: so_no, customer, deskripsi, qty total."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    filt = {"deleted_at": {"$exists": False}}
    if started_only:
        filt["prod_started"] = True
    sos = await db.sales_orders.find(filt, {"_id": 0}).sort("so_no", 1).to_list(length=10000)
    # Total qty release per SO untuk hitung balance
    frn_rows = await db.fg_release_notes.find({"status": {"$ne": "rejected"}}, {"_id": 0, "so_no": 1, "qty": 1}).to_list(length=100000)
    released = {}
    for fr in frn_rows:
        sn = fr.get("so_no") or ""
        released[sn] = released.get(sn, 0) + _f(fr.get("qty"))
    items = []
    for so in sos:
        sn = so.get("so_no") or ""
        qt = _so_qty_total(so)
        rel = released.get(sn, 0)
        items.append({
            "so_id": so.get("id"),
            "so_no": sn,
            "customer": so.get("customer") or "",
            "description": _so_desc(so),
            "qty_total": qt,
            "released": rel,
            "balance": max(0, qt - rel),
            "due_date": _date_only(so.get("due_date") or ""),
            "date_received": _date_only(so.get("prod_started_at") or ""),
        })
    return {"items": items}


@router.get("/so-items")
async def so_items(so_no: str = "", current: dict = Depends(get_current_user)):
    """Daftar item dalam 1 SO + sisa balance per item (untuk pilih item saat Release Note)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    so = await db.sales_orders.find_one({"so_no": (so_no or "").strip(), "deleted_at": {"$exists": False}})
    if not so:
        return {"so_no": so_no, "customer": "", "items": []}
    # Qty release per item (non-rejected)
    frns = await db.fg_release_notes.find(
        {"so_no": so_no, "status": {"$ne": "rejected"}}, {"_id": 0, "qty": 1, "item_index": 1}
    ).to_list(length=100000)
    rel_by_idx = {}
    for fr in frns:
        idx = fr.get("item_index")
        idx = int(idx) if idx is not None else -1
        rel_by_idx[idx] = rel_by_idx.get(idx, 0) + _f(fr.get("qty"))
    items = []
    for i, it in enumerate(so.get("items") or []):
        qty = _f(it.get("qty"))
        rel = rel_by_idx.get(i, 0)
        # jika ada release lama tanpa item_index (idx -1), bebankan ke item pertama
        if i == 0 and -1 in rel_by_idx:
            rel += rel_by_idx[-1]
        items.append({
            "index": i,
            "name": str(it.get("name") or f"Item {i + 1}").strip(),
            "qty": qty,
            "released": rel,
            "balance": max(0, qty - rel),
        })
    return {"so_no": so_no, "customer": so.get("customer") or "", "items": items}


async def _validate_frn_qty(so_no: str, new_qty: float, exclude_id: str = None, item_index=None):
    """Tolak jika total qty release (existing + baru) melebihi SO qty."""
    if not so_no:
        return
    so = await db.sales_orders.find_one({"so_no": so_no, "deleted_at": {"$exists": False}})
    if not so:
        return
    so_qty = _so_qty_total(so)
    if so_qty <= 0:
        return
    items = so.get("items") or []
    # Validasi per item bila item_index diberikan & SO multi-item
    if item_index is not None and 0 <= int(item_index) < len(items) and len(items) > 1:
        idx = int(item_index)
        item_qty = _f(items[idx].get("qty"))
        q = {"so_no": so_no, "status": {"$ne": "rejected"}, "item_index": idx}
        if exclude_id:
            q["id"] = {"$ne": exclude_id}
        rows = await db.fg_release_notes.find(q, {"_id": 0, "qty": 1}).to_list(length=100000)
        used = sum(_f(e.get("qty")) for e in rows)
        bal = item_qty - used
        if new_qty > bal:
            nm = str(items[idx].get("name") or f"Item {idx + 1}")
            raise HTTPException(status_code=400, detail=f"Qty melebihi sisa item '{nm}'. Sisa: {bal:g} (qty item {item_qty:g}, sudah rilis {used:g})")
        return
    q = {"so_no": so_no, "status": {"$ne": "rejected"}}
    if exclude_id:
        q["id"] = {"$ne": exclude_id}
    existing = await db.fg_release_notes.find(q, {"_id": 0, "qty": 1}).to_list(length=100000)
    existing_sum = sum(_f(e.get("qty")) for e in existing)
    balance = so_qty - existing_sum
    if new_qty > balance:
        raise HTTPException(status_code=400, detail=f"Qty melebihi sisa SO. Sisa balance: {balance:g} (SO qty {so_qty:g}, sudah rilis {existing_sum:g})")


@router.get("/frn")
async def list_frn(so_no: Optional[str] = None, month: Optional[str] = None,
                   date: Optional[str] = None, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    filt = {}
    if so_no:
        filt["so_no"] = {"$regex": so_no, "$options": "i"}
    if date:
        filt["frn_date"] = date
    elif month:
        filt["frn_date"] = {"$regex": f"^{month}"}
    rows = await db.fg_release_notes.find(filt, {"_id": 0}).to_list(length=100000)
    rows.sort(key=lambda r: (r.get("frn_date") or "", r.get("created_at") or ""), reverse=True)
    return {"items": [_serialize_frn(r) for r in rows]}


@router.get("/frn/pending-qc")
async def frn_pending_qc(current: dict = Depends(get_current_user)):
    """Daftar release note menunggu QC (submitted) untuk approve/tolak. QC-owned."""
    if not _can_qc(current):
        raise HTTPException(status_code=403, detail="Hanya QC/Admin yang bisa mengakses")
    rows = await db.fg_release_notes.find({"status": "submitted"}, {"_id": 0}).to_list(length=10000)
    rows.sort(key=lambda r: (r.get("frn_date") or "", r.get("created_at") or ""))
    return {"items": [_serialize_frn(r) for r in rows], "count": len(rows)}


@router.get("/frn/qc-pending-count")
async def frn_qc_pending_count(current: dict = Depends(get_current_user)):
    """Jumlah release note yang menunggu QC (status submitted). QC/Produksi/Admin."""
    if not (_can_qc(current) or _can_view(current)):
        return {"count": 0}
    n = await db.fg_release_notes.count_documents({"status": "submitted"})
    return {"count": n}


@router.post("/frn")
async def create_frn(payload: FrnIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa input")
    frn_date = (payload.frn_date or "").strip() or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    so_no = (payload.so_no or "").strip()
    customer = (payload.customer or "").strip()
    description = (payload.description or "").strip()
    item_index = payload.item_index
    # Auto-lengkapi customer & deskripsi dari SO; tentukan item
    if so_no:
        so = await db.sales_orders.find_one({"so_no": so_no, "deleted_at": {"$exists": False}})
        if so:
            customer = customer or (so.get("customer") or "")
            so_its = so.get("items") or []
            if item_index is None and len(so_its) == 1:
                item_index = 0  # 1 item → otomatis
            if item_index is not None and 0 <= int(item_index) < len(so_its):
                description = str(so_its[int(item_index)].get("name") or "").strip() or description
            if not description:
                description = _so_desc(so)
    # Auto nomor release note bila kosong
    release_no = (payload.release_no or "").strip()
    if not release_no:
        seq = (await db.fg_release_notes.count_documents({})) + 1
        release_no = f"RN-{datetime.now(timezone.utc).strftime('%Y%m')}-{seq:04d}"
    await _validate_frn_qty(so_no, _f(payload.qty), item_index=item_index)
    doc = {
        "id": str(uuid.uuid4()),
        "frn_date": frn_date,
        "release_no": release_no,
        "so_no": so_no,
        "customer": customer,
        "description": description,
        "qty": _f(payload.qty),
        "qc_comment": (payload.qc_comment or "").strip(),
        "item_index": (int(item_index) if item_index is not None else None),
        "status": "draft",
        "created_by": current.get("id"),
        "created_by_username": current.get("name") or current.get("username") or "",
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    await db.fg_release_notes.insert_one(doc)
    await log_action(current, "create_frn", "fg_release_note", doc["id"], {"so_no": so_no, "qty": doc["qty"]})
    return _serialize_frn(doc)


@router.put("/frn/{frn_id}")
async def update_frn(frn_id: str, payload: FrnIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah")
    existing = await db.fg_release_notes.find_one({"id": frn_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Release note tidak ditemukan")
    await _validate_frn_qty((payload.so_no or "").strip(), _f(payload.qty), exclude_id=frn_id, item_index=payload.item_index if payload.item_index is not None else existing.get("item_index"))
    updates = {
        "frn_date": (payload.frn_date or existing.get("frn_date") or "").strip(),
        "release_no": (payload.release_no or existing.get("release_no") or "").strip(),
        "so_no": (payload.so_no or "").strip(),
        "customer": (payload.customer or "").strip(),
        "description": (payload.description or "").strip(),
        "qty": _f(payload.qty),
        "qc_comment": (payload.qc_comment or "").strip(),
        "item_index": (int(payload.item_index) if payload.item_index is not None else existing.get("item_index")),
        "updated_at": _now_iso(),
    }
    await db.fg_release_notes.update_one({"id": frn_id}, {"$set": updates})
    await log_action(current, "update_frn", "fg_release_note", frn_id, {"so_no": updates["so_no"]})
    return _serialize_frn({**existing, **updates})


@router.delete("/frn/{frn_id}")
async def delete_frn(frn_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa menghapus")
    existing = await db.fg_release_notes.find_one({"id": frn_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Release note tidak ditemukan")
    await db.fg_release_notes.delete_one({"id": frn_id})
    await log_action(current, "delete_frn", "fg_release_note", frn_id, {"so_no": existing.get("so_no")})
    return {"ok": True}


async def _set_frn_status(frn_id: str, status: str, current: dict, extra: dict = None):
    existing = await db.fg_release_notes.find_one({"id": frn_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Release note tidak ditemukan")
    updates = {"status": status, "updated_at": _now_iso(), **(extra or {})}
    await db.fg_release_notes.update_one({"id": frn_id}, {"$set": updates})
    await log_action(current, f"frn_{status}", "fg_release_note", frn_id, {"so_no": existing.get("so_no")})
    return _serialize_frn({**existing, **updates})


class QcActionIn(BaseModel):
    qc_comment: str = ""
    qc_signature: str = ""   # optional base64 data-url gambar tanda tangan


@router.post("/frn/{frn_id}/submit")
async def submit_frn(frn_id: str, current: dict = Depends(get_current_user)):
    """Produksi submit ke QC (draft → submitted)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Tidak diizinkan")
    return await _set_frn_status(frn_id, "submitted", current)


@router.post("/frn/{frn_id}/release")
async def release_frn(frn_id: str, payload: QcActionIn = QcActionIn(), current: dict = Depends(get_current_user)):
    """QC setujui → released (barang jadi siap dikirim → Store dinotifikasi)."""
    if not _can_qc(current):
        raise HTTPException(status_code=403, detail="Hanya QC/Admin yang bisa approve/tolak")
    extra = {
        "qc_by": current.get("name") or current.get("username") or "",
        "qc_at": _now_iso(),
        "ready_for_delivery": True,
    }
    if (payload.qc_comment or "").strip():
        extra["qc_comment"] = payload.qc_comment.strip()
    if (payload.qc_signature or "").strip():
        extra["qc_signature"] = payload.qc_signature.strip()
    return await _set_frn_status(frn_id, "released", current, extra)


@router.post("/frn/{frn_id}/reject")
async def reject_frn(frn_id: str, payload: QcActionIn = QcActionIn(), current: dict = Depends(get_current_user)):
    """QC tolak → rejected (tidak dihitung sbg barang jadi)."""
    if not _can_qc(current):
        raise HTTPException(status_code=403, detail="Hanya QC/Admin yang bisa approve/tolak")
    extra = {
        "qc_by": current.get("name") or current.get("username") or "",
        "qc_at": _now_iso(),
        "ready_for_delivery": False,
    }
    if (payload.qc_comment or "").strip():
        extra["qc_comment"] = payload.qc_comment.strip()
    if (payload.qc_signature or "").strip():
        extra["qc_signature"] = payload.qc_signature.strip()
    return await _set_frn_status(frn_id, "rejected", current, extra)


@router.post("/frn/submit-date")
async def submit_frn_date(payload: dict, current: dict = Depends(get_current_user)):
    """Submit semua draft pada 1 tanggal ke QC."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Tidak diizinkan")
    d = (payload.get("date") or "").strip()
    res = await db.fg_release_notes.update_many({"frn_date": d, "status": "draft"},
                                                {"$set": {"status": "submitted", "updated_at": _now_iso()}})
    return {"ok": True, "count": res.modified_count}


class JobProgressIn(BaseModel):
    pic: str = ""
    remarks: str = ""


@router.get("/job-progress")
async def job_progress(current: dict = Depends(get_current_user)):
    """Daily Monitoring Job Progress — semua SO yang sudah 'Mulai Kerja' (per SO).
    Qty Finished diambil dari total qty Finished Goods Release Note untuk SO tsb.
    Selesai bila total release >= SO Qty; Days dihitung dari Date Received s/d selesai
    (atau s/d hari ini bila belum selesai)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")

    sos = await db.sales_orders.find(
        {"prod_started": True, "deleted_at": {"$exists": False}}, {"_id": 0}
    ).sort("prod_started_at", 1).to_list(length=10000)

    # Agregasi qty release per SO — HANYA yang sudah di-RELEASE oleh QC (barang jadi siap kirim).
    # Data lama tanpa status dianggap released (backward compat).
    frn_rows = await db.fg_release_notes.find({}, {"_id": 0, "so_no": 1, "qty": 1, "frn_date": 1, "status": 1}).to_list(length=100000)
    frn_sum = {}
    frn_last = {}
    for r in frn_rows:
        st = r.get("status") or "released"
        if st != "released":
            continue
        sn = r.get("so_no") or ""
        frn_sum[sn] = frn_sum.get(sn, 0) + _f(r.get("qty"))
        d = _date_only(r.get("frn_date") or "")
        if d and (sn not in frn_last or d > frn_last[sn]):
            frn_last[sn] = d

    jp_docs = await db.job_progress.find({}, {"_id": 0}).to_list(length=10000)
    jp_map = {d.get("so_id"): d for d in jp_docs}

    # Actual working days per SO diambil dari Daily Production Report:
    # jumlah tanggal berbeda di mana SO tsb benar-benar dikerjakan.
    pr_rows = await db.production_reports.find(
        {"so_no": {"$nin": ["", None]}}, {"_id": 0, "so_no": 1, "report_date": 1}
    ).to_list(length=200000)
    work_dates = {}
    for pr in pr_rows:
        sn = pr.get("so_no") or ""
        d = _date_only(pr.get("report_date") or "")
        if sn and d:
            work_dates.setdefault(sn, set()).add(d)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Set tanggal libur nasional untuk perhitungan Working Date Target
    hol_docs = await db.holidays.find({}, {"_id": 0, "date": 1}).to_list(length=5000)
    holidays = {_date_only(h.get("date") or "") for h in hol_docs if h.get("date")}
    items = []
    for so in sos:
        so_id = so.get("id")
        so_no = so.get("so_no") or ""
        so_qty = _so_qty_total(so)
        qty_finished = frn_sum.get(so_no, 0)
        jp = jp_map.get(so_id, {})
        date_received = _date_only(so.get("prod_started_at") or "")
        finished = so_qty > 0 and qty_finished >= so_qty

        finished_at = jp.get("finished_at") or ""
        # Sinkron status selesai
        if finished and not finished_at:
            finished_at = frn_last.get(so_no) or today
            await db.job_progress.update_one({"so_id": so_id}, {"$set": {"finished_at": finished_at, "so_no": so_no}}, upsert=True)
        elif not finished and finished_at:
            finished_at = ""
            await db.job_progress.update_one({"so_id": so_id}, {"$set": {"finished_at": ""}}, upsert=True)

        end_for_days = finished_at if (finished and finished_at) else today
        days = _days_between(date_received, end_for_days) if date_received else 0

        due_date = _date_only(so.get("due_date") or "")   # otomatis dari SO
        pct = (qty_finished / so_qty) if so_qty > 0 else 0
        awd = sorted(work_dates.get(so_no, set()))
        actual_working_days = len(awd)
        # Working Date Target = hari kerja (exclude Minggu) dari mulai kerja s/d Due Date
        working_date_target = _working_days(date_received, due_date, holidays) if (date_received and due_date) else 0
        # Productivity = Working Date Target / Actual Working Day (%)
        productivity = round((working_date_target / actual_working_days) * 100, 1) if actual_working_days > 0 else 0

        # Sisa hari kerja s/d Due Date (kecualikan Minggu & libur) + status kesehatan
        days_remaining = None
        overdue_days = 0
        if due_date:
            if today <= due_date:
                days_remaining = _working_days(today, due_date, holidays)
            else:
                days_remaining = 0
                overdue_days = _days_between(due_date, today)
        if finished:
            health = "finished"
        elif due_date and today > due_date:
            health = "late"
        elif due_date and days_remaining is not None and days_remaining <= 3:
            health = "warning"
        else:
            health = "on_track"

        items.append({
            "so_id": so_id,
            "so_no": so_no,
            "customer": so.get("customer") or "",
            "description": _so_desc(so),
            "so_qty": so_qty,
            "date_received": date_received,
            "due_date": due_date,
            "plan_start": date_received,                       # otomatis = tgl mulai kerja produksi
            "plan_finish": finished_at if (finished and finished_at) else "",  # otomatis saat qty release completed
            "days": days,
            "days_remaining": days_remaining,
            "overdue_days": overdue_days,
            "working_date_target": working_date_target,
            "actual_working_days": actual_working_days,        # dari Daily Production Report
            "actual_working_dates": awd,
            "productivity": productivity,
            "qty_finished": qty_finished,
            "qty_balance": max(0, so_qty - qty_finished),
            "percent": round(pct * 100, 1),
            "pic": jp.get("pic") or "",
            "remarks": jp.get("remarks") or "",
            "finished": finished,
            "finished_at": finished_at,
            "status": "FINISHED" if finished else "PROSES",
            "health": health,
        })
    total = len(items)
    finished_count = sum(1 for i in items if i["finished"])
    late_count = sum(1 for i in items if i["health"] == "late")
    warning_count = sum(1 for i in items if i["health"] == "warning")
    prods = [i["productivity"] for i in items if i["productivity"] > 0]
    avg_productivity = round(sum(prods) / len(prods), 1) if prods else 0
    return {
        "items": items, "count": total, "finished": finished_count,
        "in_progress": total - finished_count, "late": late_count,
        "warning": warning_count, "avg_productivity": avg_productivity,
    }


@router.get("/job-progress/export.xlsx")
async def export_job_progress_xlsx(current: dict = Depends(get_current_user)):
    """Export Daily Monitoring Job Progress ke Excel."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    data = await job_progress(current)
    HEALTH_LABEL = {"finished": "Selesai", "late": "Terlambat", "warning": "Warning", "on_track": "On Track"}
    wb = Workbook(); ws = wb.active; ws.title = "Job Progress"
    headers = ["SO No", "Customer", "Job Description", "Qty SO", "Mulai Kerja", "Due Date",
               "Finished", "Days", "Sisa Hari", "Qty Finished", "Qty Balance", "% Progress",
               "Target Hari Kerja", "Actual Working Day", "Produktivitas %", "Status"]
    ws.append(headers)
    hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="B45309"); center = Alignment(horizontal="center")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c); cell.font = hf; cell.fill = hfill; cell.alignment = center
    for i in data["items"]:
        ws.append([
            i["so_no"], i["customer"], i["description"], i["so_qty"], i["date_received"], i["due_date"],
            i["finished_at"] or "", i["days"], (i["days_remaining"] if i["days_remaining"] is not None else ""),
            i["qty_finished"], i["qty_balance"], i["percent"], i["working_date_target"],
            i["actual_working_days"], i["productivity"], HEALTH_LABEL.get(i["health"], i["status"]),
        ])
    widths = [12, 22, 30, 8, 12, 12, 12, 7, 9, 12, 11, 10, 14, 16, 14, 12]
    for w, col in zip(widths, range(1, len(widths) + 1)):
        ws.column_dimensions[chr(64 + col)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"job_progress_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.put("/job-progress/{so_id}")
async def update_job_progress(so_id: str, payload: JobProgressIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah")
async def update_job_progress(so_id: str, payload: JobProgressIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah")
    so = await db.sales_orders.find_one({"id": so_id, "deleted_at": {"$exists": False}})
    if not so:
        raise HTTPException(status_code=404, detail="SO tidak ditemukan")
    updates = {
        "so_id": so_id,
        "so_no": so.get("so_no") or "",
        "pic": (payload.pic or "").strip(),
        "remarks": (payload.remarks or "").strip(),
        "updated_at": _now_iso(),
    }
    await db.job_progress.update_one({"so_id": so_id}, {"$set": updates}, upsert=True)
    await log_action(current, "update_job_progress", "job_progress", so_id, {"so_no": updates["so_no"]})
    return {"ok": True}



# ==========================================================================
# Master Hari Libur Nasional (untuk perhitungan Working Date Target)
# ==========================================================================

class HolidayIn(BaseModel):
    date: str = ""       # YYYY-MM-DD
    name: str = ""


@router.get("/holidays")
async def list_holidays(year: Optional[str] = None, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    filt = {}
    if year:
        filt["date"] = {"$regex": f"^{year}"}
    rows = await db.holidays.find(filt, {"_id": 0}).to_list(length=5000)
    rows.sort(key=lambda r: r.get("date") or "")
    items = [{"id": r.get("id"), "date": r.get("date") or "", "name": r.get("name") or ""} for r in rows]
    return {"items": items, "count": len(items)}


@router.post("/holidays")
async def create_holiday(payload: HolidayIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa input")
    date = _date_only((payload.date or "").strip())
    if not date:
        raise HTTPException(status_code=400, detail="Tanggal wajib diisi")
    existing = await db.holidays.find_one({"date": date})
    if existing:
        await db.holidays.update_one({"date": date}, {"$set": {"name": (payload.name or "").strip()}})
        return {"id": existing.get("id"), "date": date, "name": (payload.name or "").strip()}
    doc = {
        "id": str(uuid.uuid4()),
        "date": date,
        "name": (payload.name or "").strip(),
        "created_at": _now_iso(),
    }
    await db.holidays.insert_one(doc)
    await log_action(current, "create_holiday", "holiday", doc["id"], {"date": date})
    return {"id": doc["id"], "date": date, "name": doc["name"]}


@router.delete("/holidays/{holiday_id}")
async def delete_holiday(holiday_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa menghapus")
    existing = await db.holidays.find_one({"id": holiday_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Hari libur tidak ditemukan")
    await db.holidays.delete_one({"id": holiday_id})
    await log_action(current, "delete_holiday", "holiday", holiday_id, {"date": existing.get("date")})
    return {"ok": True}



# ==========================================================================
# Absensi Kehadiran Produksi (employee master + attendance harian)
# ==========================================================================

# Status yang dianggap TIDAK bekerja (operator diblok di Daily Production)
ATTEND_BLOCKED = {"mc_sakit", "tidak_hadir"}
ATTEND_STATUSES = ["hadir", "terlambat", "ijin_keluar", "ijin_pulang", "night_shift", "mc_sakit", "tidak_hadir", "insitu"]


class EmployeeIn(BaseModel):
    name: str = ""
    designation: str = ""


class AttendanceEntry(BaseModel):
    employee_id: str = ""
    name: str = ""
    status: str = "hadir"
    out_time: str = ""
    plan_in_time: str = ""
    actual_in_time: str = ""
    home_time: str = ""
    insitu_location: str = ""
    insitu_start: str = ""
    insitu_est_finish: str = ""
    insitu_actual_finish: str = ""
    note: str = ""


class AttendanceBulk(BaseModel):
    date: str = ""
    entries: list[AttendanceEntry] = []


@router.get("/employees")
async def list_employees(current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    rows = await db.production_employees.find({"active": {"$ne": False}}, {"_id": 0}).to_list(length=5000)
    rows.sort(key=lambda r: (r.get("name") or "").lower())
    return {"items": [{"id": r.get("id"), "name": r.get("name") or "", "badge_no": r.get("badge_no") or "", "designation": r.get("designation") or ""} for r in rows]}


@router.post("/employees")
async def create_employee(payload: EmployeeIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa input")
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nama wajib diisi")
    existing = await db.production_employees.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "active": {"$ne": False}})
    if existing:
        return {"id": existing.get("id"), "name": existing.get("name")}
    doc = {"id": str(uuid.uuid4()), "name": name, "designation": (payload.designation or "").strip(), "active": True, "created_at": _now_iso()}
    await db.production_employees.insert_one(doc)
    return {"id": doc["id"], "name": name, "designation": doc["designation"]}


@router.put("/employees/{emp_id}")
async def update_employee(emp_id: str, payload: EmployeeIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah")
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nama wajib diisi")
    existing = await db.production_employees.find_one({"id": emp_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    await db.production_employees.update_one({"id": emp_id}, {"$set": {"name": name, "designation": (payload.designation or "").strip()}})
    # Sinkronkan nama di attendance agar konsisten
    await db.production_attendance.update_many({"employee_id": emp_id}, {"$set": {"name": name}})
    return {"id": emp_id, "name": name, "designation": (payload.designation or "").strip()}


@router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa menghapus")
    await db.production_employees.update_one({"id": emp_id}, {"$set": {"active": False}})
    return {"ok": True}


def _serialize_att(a: dict) -> dict:
    return {
        "employee_id": a.get("employee_id") or "",
        "name": a.get("name") or "",
        "status": a.get("status") or "hadir",
        "out_time": a.get("out_time") or "",
        "plan_in_time": a.get("plan_in_time") or "",
        "actual_in_time": a.get("actual_in_time") or "",
        "home_time": a.get("home_time") or "",
        "insitu_location": a.get("insitu_location") or "",
        "insitu_start": a.get("insitu_start") or "",
        "insitu_est_finish": a.get("insitu_est_finish") or "",
        "insitu_actual_finish": a.get("insitu_actual_finish") or "",
        "note": a.get("note") or "",
    }


@router.get("/attendance/month")
async def attendance_month(month: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Rekap absensi 1 bulan: karyawan x tanggal."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    import calendar as _cal
    m = month or datetime.now(timezone.utc).strftime("%Y-%m")
    try:
        yy, mm = int(m[:4]), int(m[5:7])
        ndays = _cal.monthrange(yy, mm)[1]
    except Exception:
        yy = datetime.now(timezone.utc).year; mm = datetime.now(timezone.utc).month
        ndays = _cal.monthrange(yy, mm)[1]
        m = f"{yy:04d}-{mm:02d}"
    days = [f"{m}-{d:02d}" for d in range(1, ndays + 1)]
    emps = await db.production_employees.find({"active": {"$ne": False}}, {"_id": 0}).to_list(length=5000)
    emps.sort(key=lambda r: (r.get("name") or "").lower())
    att = await db.production_attendance.find({"date": {"$regex": f"^{m}"}}, {"_id": 0}).to_list(length=50000)
    records = {}
    for a in att:
        records.setdefault(a.get("employee_id"), {})[a.get("date")] = a.get("status") or "hadir"
    return {
        "month": m, "days": days,
        "employees": [{"id": e.get("id"), "name": e.get("name") or "", "designation": e.get("designation") or ""} for e in emps],
        "records": records, "statuses": ATTEND_STATUSES,
    }


@router.get("/attendance")
async def get_attendance(date: Optional[str] = None, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    d = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    emps = await db.production_employees.find({"active": {"$ne": False}}, {"_id": 0}).to_list(length=5000)
    emps.sort(key=lambda r: (r.get("name") or "").lower())
    att = await db.production_attendance.find({"date": d}, {"_id": 0}).to_list(length=5000)
    att_map = {a.get("employee_id"): a for a in att}
    items = []
    for e in emps:
        a = att_map.get(e.get("id"))
        base = {"designation": e.get("designation") or "", "badge_no": e.get("badge_no") or ""}
        if a:
            items.append({**_serialize_att(a), **base})
        else:
            items.append({**_serialize_att({}), "employee_id": e.get("id"), "name": e.get("name") or "", "status": "hadir", **base})
    return {"date": d, "items": items, "statuses": ATTEND_STATUSES}


@router.post("/attendance")
async def save_attendance(payload: AttendanceBulk, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa input")
    d = (payload.date or "").strip() or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for e in payload.entries:
        doc = {
            "date": d, "employee_id": e.employee_id, "name": (e.name or "").strip(),
            "status": e.status or "hadir",
            "out_time": e.out_time or "", "plan_in_time": e.plan_in_time or "", "actual_in_time": e.actual_in_time or "",
            "home_time": e.home_time or "",
            "insitu_location": e.insitu_location or "", "insitu_start": e.insitu_start or "",
            "insitu_est_finish": e.insitu_est_finish or "", "insitu_actual_finish": e.insitu_actual_finish or "",
            "note": e.note or "", "updated_at": _now_iso(),
        }
        await db.production_attendance.update_one({"date": d, "employee_id": e.employee_id}, {"$set": doc}, upsert=True)
    await log_action(current, "save_attendance", "attendance", d, {"count": len(payload.entries)})
    return {"ok": True, "count": len(payload.entries)}



# ==========================================================================
# Overtime Request + rekap OT bulanan
# ==========================================================================

class OvertimeIn(BaseModel):
    ot_date: str = ""
    ot_no: str = ""
    name: str = ""
    so_no: str = ""
    customer: str = ""
    ot_start: str = ""
    ot_end: str = ""
    ot_hours: Optional[float] = None   # isi manual jumlah jam lembur (opsional; override jam mulai/selesai)


class OvertimeRulesIn(BaseModel):
    weekday_start: str = "16:00"
    saturday_start: str = "15:00"
    holiday_work_start: str = "08:00"
    holiday_work_end: str = "16:00"
    holiday_break_hours: float = 1
    rounding: str = "floor"          # floor | half | round
    wd_first_mult: float = 1.5       # hari kerja/Sabtu jam ke-1
    wd_rest_mult: float = 2.0        # hari kerja/Sabtu jam ke-2 dst
    hol_normal_hours: int = 7        # Minggu/libur jam ke-1..N
    hol_normal_mult: float = 2.0
    hol_8th_mult: float = 3.0        # jam ke-(N+1)
    hol_extra_mult: float = 4.0      # jam berikutnya


# Aturan lembur default (dari referensi Gaji Trial.xlsx / Depnaker 6-hari kerja)
DEFAULT_OT_RULES = {
    "weekday_start": "16:00",
    "saturday_start": "15:00",
    "holiday_work_start": "08:00",
    "holiday_work_end": "16:00",
    "holiday_break_hours": 1,
    "rounding": "floor",
    "wd_first_mult": 1.5,
    "wd_rest_mult": 2.0,
    "hol_normal_hours": 7,
    "hol_normal_mult": 2.0,
    "hol_8th_mult": 3.0,
    "hol_extra_mult": 4.0,
}


async def _get_ot_rules() -> dict:
    doc = await db.production_overtime_rules.find_one({"_id": "default"}) or {}
    rules = {**DEFAULT_OT_RULES}
    for k in DEFAULT_OT_RULES:
        if doc.get(k) is not None:
            rules[k] = doc[k]
    return rules


def _time_to_min(t: str) -> int:
    try:
        h, m = [int(x) for x in (t or "").split(":")[:2]]
        return h * 60 + m
    except Exception:
        return 0


def _round_hours(mins: int, mode: str) -> float:
    """Bulatkan durasi menit menjadi jam sesuai mode."""
    if mins <= 0:
        return 0.0
    if mode == "half":            # kelipatan 0.5 jam (bulat ke bawah)
        return (mins // 30) * 0.5
    if mode == "round":           # bulat terdekat ke jam penuh
        return float(round(mins / 60.0))
    return float(mins // 60)      # floor: bulat ke bawah ke jam penuh


def _ot_day_type(ot_date: str, holidays: set) -> str:
    iso = _date_only(ot_date)
    try:
        wd = datetime.fromisoformat(iso).weekday()   # Sen=0 .. Sab=5, Min=6
    except Exception:
        wd = 0
    if iso in holidays or wd == 6:
        return "holiday"
    if wd == 5:
        return "saturday"
    return "weekday"


def _calc_overtime(ot_date: str, start: str, end: str, rules: dict, holidays: set, manual_hours=None) -> dict:
    """Hitung jam lembur + rincian pengali (1.5x/2x/3x/4x) + jam tertimbang.
    - manual_hours: jika diisi, dipakai langsung sbg jumlah jam lembur (spt entri manual di Excel).
    - Untuk Minggu/libur, jam istirahat otomatis dikurangi (mis. 08:00-16:00 = 8 jam - 1 = 7 jam)
      HANYA bila jam dihitung dari jam mulai/selesai (bukan input manual).
    """
    day_type = _ot_day_type(ot_date, holidays)
    manual = manual_hours is not None and str(manual_hours) != "" and float(manual_hours) > 0
    if manual:
        ot_hours = float(manual_hours)
        raw_hours = ot_hours
    else:
        mins = _time_to_min(end) - _time_to_min(start)
        if mins < 0:
            mins += 24 * 60  # lewat tengah malam
        raw_hours = round(mins / 60.0, 2)
        ot_hours = _round_hours(mins, rules.get("rounding") or "floor")
        if day_type == "holiday":
            ot_hours = max(0.0, ot_hours - float(rules.get("holiday_break_hours", 1) or 0))
    x15 = x2 = x3 = x4 = 0.0
    if day_type == "holiday":
        n = int(rules.get("hol_normal_hours", 7))
        x2 = min(ot_hours, float(n))
        rem = ot_hours - x2
        x3 = 1.0 if rem >= 1 else rem
        rem = max(0.0, rem - 1)
        x4 = rem
        weighted = (x2 * float(rules.get("hol_normal_mult", 2.0))
                    + x3 * float(rules.get("hol_8th_mult", 3.0))
                    + x4 * float(rules.get("hol_extra_mult", 4.0)))
    else:
        x15 = 1.0 if ot_hours >= 1 else ot_hours
        x2 = max(0.0, ot_hours - 1)
        weighted = (x15 * float(rules.get("wd_first_mult", 1.5))
                    + x2 * float(rules.get("wd_rest_mult", 2.0)))
    label = {"holiday": "Minggu/Libur", "saturday": "Sabtu", "weekday": "Hari Kerja"}[day_type]
    return {
        "day_type": day_type,
        "day_label": label,
        "raw_hours": round(raw_hours, 2),
        "ot_hours": round(ot_hours, 2),
        "manual": manual,
        "x15": round(x15, 2), "x2": round(x2, 2), "x3": round(x3, 2), "x4": round(x4, 2),
        "weighted_hours": round(weighted, 2),
    }


@router.get("/overtime-rules")
async def get_overtime_rules(current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    return {"rules": await _get_ot_rules()}


@router.put("/overtime-rules")
async def update_overtime_rules(payload: OvertimeRulesIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengubah master lembur")
    doc = {**payload.dict(), "_id": "default", "updated_at": _now_iso(),
           "updated_by": current.get("name") or current.get("username") or ""}
    await db.production_overtime_rules.update_one({"_id": "default"}, {"$set": doc}, upsert=True)
    await log_action(current, "update_overtime_rules", "overtime_rules", "default", {})
    return {"ok": True, "rules": await _get_ot_rules()}


@router.get("/overtime")
async def list_overtime(month: Optional[str] = None, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    m = month or datetime.now(timezone.utc).strftime("%Y-%m")
    rules = await _get_ot_rules()
    hol_docs = await db.holidays.find({}, {"_id": 0, "date": 1}).to_list(length=5000)
    holidays = {_date_only(h.get("date") or "") for h in hol_docs if h.get("date")}
    rows = await db.production_overtime.find({"ot_date": {"$regex": f"^{m}"}}, {"_id": 0}).to_list(length=50000)
    rows.sort(key=lambda r: (r.get("ot_date") or "", r.get("created_at") or ""), reverse=True)
    items, summary = [], {}
    for r in rows:
        calc = _calc_overtime(r.get("ot_date") or "", r.get("ot_start") or "", r.get("ot_end") or "", rules, holidays, r.get("ot_hours"))
        items.append({**{k: r.get(k) or "" for k in ["id", "ot_date", "ot_no", "name", "so_no", "customer", "ot_start", "ot_end"]},
                      "hours": calc["ot_hours"], **calc})
        nm = r.get("name") or "-"
        s = summary.setdefault(nm, {"name": nm, "total_hours": 0.0, "x15": 0.0, "x2": 0.0, "x3": 0.0, "x4": 0.0, "weighted_hours": 0.0})
        s["total_hours"] = round(s["total_hours"] + calc["ot_hours"], 2)
        s["x15"] = round(s["x15"] + calc["x15"], 2)
        s["x2"] = round(s["x2"] + calc["x2"], 2)
        s["x3"] = round(s["x3"] + calc["x3"], 2)
        s["x4"] = round(s["x4"] + calc["x4"], 2)
        s["weighted_hours"] = round(s["weighted_hours"] + calc["weighted_hours"], 2)
    summary_list = sorted(summary.values(), key=lambda x: -x["total_hours"])
    return {
        "month": m, "items": items, "summary": summary_list, "rules": rules,
        "total_hours": round(sum(s["total_hours"] for s in summary.values()), 2),
        "total_weighted": round(sum(s["weighted_hours"] for s in summary.values()), 2),
    }


@router.get("/overtime/grid")
async def overtime_grid(month: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Rekap lembur bentuk grid bulanan (karyawan × tanggal) + total per bulan — mirip absensi."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    import calendar as _cal
    m = month or datetime.now(timezone.utc).strftime("%Y-%m")
    yr, mo = int(m[:4]), int(m[5:7])
    ndays = _cal.monthrange(yr, mo)[1]
    days = [f"{m}-{d:02d}" for d in range(1, ndays + 1)]
    rules = await _get_ot_rules()
    hol_docs = await db.holidays.find({}, {"_id": 0, "date": 1}).to_list(length=5000)
    holidays = {_date_only(h.get("date") or "") for h in hol_docs if h.get("date")}
    rows = await db.production_overtime.find({"ot_date": {"$regex": f"^{m}"}}, {"_id": 0}).to_list(length=50000)
    grid = {}
    for r in rows:
        calc = _calc_overtime(r.get("ot_date") or "", r.get("ot_start") or "", r.get("ot_end") or "", rules, holidays, r.get("ot_hours"))
        nm = (r.get("name") or "-").strip() or "-"
        d = _date_only(r.get("ot_date") or "")
        g = grid.setdefault(nm, {})
        g[d] = round(g.get(d, 0) + calc["ot_hours"], 2)
    items = []
    for nm, perdate in grid.items():
        th = round(sum(perdate.values()), 2)
        td = len([1 for v in perdate.values() if v > 0])
        items.append({"name": nm, "per_date": perdate, "total_hours": th, "total_days": td})
    items.sort(key=lambda x: x["name"])
    return {
        "month": m, "days": days, "items": items,
        "grand_total_hours": round(sum(i["total_hours"] for i in items), 2),
        "grand_total_days": sum(i["total_days"] for i in items),
    }


@router.get("/overtime/grid/export.xlsx")
async def export_overtime_grid_xlsx(month: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Export rekap grid lembur bulanan (Nama × tanggal + Total)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    g = await overtime_grid(month, current)
    wb = Workbook(); ws = wb.active; ws.title = "Rekap Lembur"
    hdr = ["Nama"] + [str(int(d[8:])) for d in g["days"]] + ["Total Jam", "Kali"]
    ws.append(hdr)
    hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="B45309"); ctr = Alignment(horizontal="center")
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=1, column=c); cell.font = hf; cell.fill = hfill; cell.alignment = ctr
    for it in g["items"]:
        row = [it["name"]] + [it["per_date"].get(d, "") for d in g["days"]] + [it["total_hours"], it["total_days"]]
        ws.append(row)
    ws.append([])
    ws.append(["TOTAL", *[""] * len(g["days"]), g["grand_total_hours"], g["grand_total_days"]])
    ws.column_dimensions["A"].width = 22
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="rekap_grid_lembur_{g["month"]}.xlsx"'})


@router.get("/overtime/export.xlsx")
async def export_overtime_xlsx(month: Optional[str] = None, current: dict = Depends(get_current_user)):
    """Export rekap lembur bulanan ke Excel (detail per record + rekap per karyawan)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    m = month or datetime.now(timezone.utc).strftime("%Y-%m")
    rules = await _get_ot_rules()
    hol_docs = await db.holidays.find({}, {"_id": 0, "date": 1}).to_list(length=5000)
    holidays = {_date_only(h.get("date") or "") for h in hol_docs if h.get("date")}
    rows = await db.production_overtime.find({"ot_date": {"$regex": f"^{m}"}}, {"_id": 0}).to_list(length=50000)
    rows.sort(key=lambda r: (r.get("ot_date") or "", r.get("created_at") or ""))

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="B45309")
    center = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Detail Lembur"
    headers = ["Tanggal", "No.", "Nama", "SO No", "Customer", "Jenis Hari",
               "Jam Mulai", "Jam Selesai", "Jam Lembur", "1.5x", "2x", "3x", "4x", "Jam Tertimbang"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center

    summary = {}
    for r in rows:
        calc = _calc_overtime(r.get("ot_date") or "", r.get("ot_start") or "", r.get("ot_end") or "", rules, holidays, r.get("ot_hours"))
        ws.append([
            r.get("ot_date") or "", r.get("ot_no") or "", r.get("name") or "", r.get("so_no") or "",
            r.get("customer") or "", calc["day_label"],
            r.get("ot_start") or "", r.get("ot_end") or "",
            calc["ot_hours"], calc["x15"], calc["x2"], calc["x3"], calc["x4"], calc["weighted_hours"],
        ])
        nm = r.get("name") or "-"
        s = summary.setdefault(nm, {"total": 0.0, "x15": 0.0, "x2": 0.0, "x3": 0.0, "x4": 0.0, "w": 0.0})
        s["total"] += calc["ot_hours"]; s["x15"] += calc["x15"]; s["x2"] += calc["x2"]
        s["x3"] += calc["x3"]; s["x4"] += calc["x4"]; s["w"] += calc["weighted_hours"]
    for w, col in zip([12, 16, 22, 14, 24, 14, 11, 11, 11, 8, 8, 8, 8, 14], range(1, 15)):
        ws.column_dimensions[chr(64 + col)].width = w

    ws2 = wb.create_sheet("Rekap per Karyawan")
    h2 = ["Nama", "Total Jam", "1.5x", "2x", "3x", "4x", "Jam Tertimbang"]
    ws2.append(h2)
    for c in range(1, len(h2) + 1):
        cell = ws2.cell(row=1, column=c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
    for nm, s in sorted(summary.items(), key=lambda x: -x[1]["total"]):
        ws2.append([nm, round(s["total"], 2), round(s["x15"], 2), round(s["x2"], 2),
                    round(s["x3"], 2), round(s["x4"], 2), round(s["w"], 2)])
    for w, col in zip([24, 12, 8, 8, 8, 8, 14], range(1, 8)):
        ws2.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="rekap_lembur_{m}.xlsx"'},
    )


@router.post("/overtime/preview")
async def preview_overtime(payload: OvertimeIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa mengakses")
    rules = await _get_ot_rules()
    hol_docs = await db.holidays.find({}, {"_id": 0, "date": 1}).to_list(length=5000)
    holidays = {_date_only(h.get("date") or "") for h in hol_docs if h.get("date")}
    return _calc_overtime(payload.ot_date or "", payload.ot_start or "", payload.ot_end or "", rules, holidays, payload.ot_hours)


@router.post("/overtime")
async def create_overtime(payload: OvertimeIn, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa input")
    od = _date_only((payload.ot_date or "").strip()) or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    so_no = (payload.so_no or "").strip()
    customer = (payload.customer or "").strip()
    if so_no and not customer:
        so = await db.sales_orders.find_one({"so_no": so_no, "deleted_at": {"$exists": False}})
        if so:
            customer = so.get("customer") or ""
    ot_no = (payload.ot_no or "").strip()
    if not ot_no:
        seq = (await db.production_overtime.count_documents({})) + 1
        ot_no = f"OT-{datetime.now(timezone.utc).strftime('%Y%m')}-{seq:04d}"
    doc = {
        "id": str(uuid.uuid4()), "ot_date": od, "ot_no": ot_no, "name": (payload.name or "").strip(),
        "so_no": so_no, "customer": customer, "ot_start": (payload.ot_start or "").strip(),
        "ot_end": (payload.ot_end or "").strip(),
        "ot_hours": (float(payload.ot_hours) if (payload.ot_hours is not None and str(payload.ot_hours) != "") else None),
        "created_by_username": current.get("name") or current.get("username") or "",
        "created_at": _now_iso(),
    }
    await db.production_overtime.insert_one(doc)
    await log_action(current, "create_overtime", "overtime", doc["id"], {"name": doc["name"], "so_no": so_no})
    return {"ok": True, "id": doc["id"], "ot_no": ot_no}


class OvertimeBulkIn(BaseModel):
    ot_date: str = ""
    entries: List[OvertimeIn] = []


@router.post("/overtime/bulk")
async def create_overtime_bulk(payload: OvertimeBulkIn, current: dict = Depends(get_current_user)):
    """Simpan banyak baris OT sekaligus untuk 1 tanggal (OVER TIME REQUEST FORM)."""
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa input")
    od = _date_only((payload.ot_date or "").strip()) or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    ym = od[:7].replace("-", "")
    seq = await db.production_overtime.count_documents({})
    created = []
    for e in payload.entries:
        name = (e.name or "").strip()
        if not name:
            continue
        so_no = (e.so_no or "").strip()
        customer = (e.customer or "").strip()
        if so_no and not customer:
            so = await db.sales_orders.find_one({"so_no": so_no, "deleted_at": {"$exists": False}})
            if so:
                customer = so.get("customer") or ""
        seq += 1
        ot_no = (e.ot_no or "").strip() or f"OT-{ym}-{seq:04d}"
        doc = {
            "id": str(uuid.uuid4()), "ot_date": od, "ot_no": ot_no, "name": name,
            "so_no": so_no, "customer": customer, "ot_start": (e.ot_start or "").strip(),
            "ot_end": (e.ot_end or "").strip(),
            "ot_hours": (float(e.ot_hours) if (e.ot_hours is not None and str(e.ot_hours) != "") else None),
            "created_by_username": current.get("name") or current.get("username") or "",
            "created_at": _now_iso(),
        }
        await db.production_overtime.insert_one(doc)
        created.append(doc["id"])
    await log_action(current, "create_overtime_bulk", "overtime", od, {"count": len(created)})
    return {"ok": True, "count": len(created), "ids": created}


@router.delete("/overtime/{ot_id}")
async def delete_overtime(ot_id: str, current: dict = Depends(get_current_user)):
    if not _can_view(current):
        raise HTTPException(status_code=403, detail="Hanya Produksi/Admin yang bisa menghapus")
    await db.production_overtime.delete_one({"id": ot_id})
    return {"ok": True}

