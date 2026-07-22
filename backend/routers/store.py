"""Store module: receive, issue, stock, FIFO, requests/approvals, manual receive, production issue."""
import io
import uuid
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from db import db
from deps import (
    _now_iso,
    can_see_prices,
    get_current_user,
    log_action,
    require_approve_perm,
    require_store_access,
    require_store_write,
)
from models import (
    BulkIssueRequest,
    BulkReceiveRequest,
    IncomingGoodsRequest,
    ManualReceiveRequest,
    ProductionIssueRequest,
    StoreIssueRequest,
    StoreReceiveRequest,
    StoreRequestCreate,
    StoreRequestReview,
)

router = APIRouter(tags=["store"])


async def _sum_received_for_tx(tx_id: str) -> float:
    agg = await db.store_receipts.aggregate([
        {"$match": {"transaction_id": tx_id}},
        {"$group": {"_id": None, "total": {"$sum": "$qty_received"}}}
    ]).to_list(length=1)
    return float(agg[0]["total"]) if agg else 0.0


@router.get("/store/pending")
async def store_pending(current: dict = Depends(require_store_access)):
    """List item transaksi yang di-flag 'post_to_store' dan belum full-received."""
    txs = await db.transactions.find(
        {"post_to_store": True},
        {"_id": 0},
    ).sort("invoice_date", -1).to_list(length=5000)

    result = []
    for t in txs:
        received = await _sum_received_for_tx(t["id"])
        remaining = float(t.get("qty", 0)) - received
        if remaining > 0:
            result.append({
                "transaction_id": t["id"],
                "invoice_date": t.get("invoice_date"),
                "po_no": t.get("po_no"),
                "invoice_no": t.get("invoice_no"),
                "vendor_name": t.get("vendor_name"),
                "item_name": t.get("item_name"),
                "unit": t.get("unit"),
                "qty_po": float(t.get("qty", 0)),
                "qty_received": received,
                "qty_remaining": remaining,
                "po_date": t.get("po_date"),
            })
    return result


@router.get("/store/pending/grouped")
async def store_pending_grouped(current: dict = Depends(require_store_access)):
    """Grouped view: 1 baris per PO (fallback invoice_no) dengan ringkasan qty."""
    items = await store_pending(current)
    groups: dict = {}
    for it in items:
        key = it.get("po_no") or f"INV:{it.get('invoice_no', '')}" or f"TX:{it.get('transaction_id')}"
        g = groups.setdefault(key, {
            "group_key": key,
            "po_no": it.get("po_no"),
            "invoice_no": it.get("invoice_no"),
            "vendor_name": it.get("vendor_name"),
            "invoice_date": it.get("invoice_date"),
            "po_date": it.get("po_date"),
            "items": [],
            "total_qty_po": 0.0,
            "total_qty_received": 0.0,
            "total_qty_remaining": 0.0,
        })
        g["items"].append(it)
        g["total_qty_po"] += it["qty_po"]
        g["total_qty_received"] += it["qty_received"]
        g["total_qty_remaining"] += it["qty_remaining"]
    return sorted(groups.values(), key=lambda x: (x["invoice_date"] or ""), reverse=True)


@router.post("/store/receive/bulk")
async def store_receive_bulk(payload: BulkReceiveRequest, current: dict = Depends(require_store_access)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    received_docs = []
    touched_tx_ids: set = set()
    for item in payload.items:
        if item.qty_received <= 0:
            continue
        tx = await db.transactions.find_one({"id": item.transaction_id})
        if not tx or not tx.get("post_to_store"):
            raise HTTPException(status_code=400, detail=f"Item {item.transaction_id} tidak valid")
        already = await _sum_received_for_tx(item.transaction_id)
        remaining = float(tx.get("qty", 0)) - already
        if item.qty_received > remaining + 1e-9:
            raise HTTPException(
                status_code=400,
                detail=f"{tx.get('item_name')}: qty terima ({item.qty_received}) > sisa ({remaining})"
            )
        add_to_stock = True if item.add_to_stock is None else bool(item.add_to_stock)
        doc = {
            "id": str(uuid.uuid4()),
            "transaction_id": item.transaction_id,
            "po_no": tx.get("po_no", ""),
            "invoice_no": payload.invoice_no or tx.get("invoice_no", ""),
            "vendor_name": tx.get("vendor_name", ""),
            "item_name": tx.get("item_name", ""),
            "unit": tx.get("unit", "Ea"),
            "unit_price": float(tx.get("unit_price", 0)),
            "do_number": payload.do_number or "",
            "qty_received": float(item.qty_received),
            # If not added to stock, qty_remaining = 0 (barang langsung habis pakai)
            "qty_remaining": float(item.qty_received) if add_to_stock else 0.0,
            "add_to_stock": add_to_stock,
            "receive_date": payload.receive_date,
            "note": item.note or "",
            "source": "po",
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        received_docs.append(doc)
        touched_tx_ids.add(item.transaction_id)
    if not received_docs:
        raise HTTPException(status_code=400, detail="Semua qty kosong / 0")
    await db.store_receipts.insert_many([d.copy() for d in received_docs])

    # Auto-update source transactions with invoice_no + receive_date so purchasing masterlist reflects real receive
    tx_updates: dict = {"receive_date": payload.receive_date}
    if payload.invoice_no:
        tx_updates["invoice_no"] = payload.invoice_no
    if touched_tx_ids:
        await db.transactions.update_many(
            {"id": {"$in": list(touched_tx_ids)}},
            {"$set": tx_updates},
        )

    await log_action(current, "store_receive", "store_receipt", "-", {
        "count": len(received_docs), "po_no": received_docs[0].get("po_no"),
        "do_number": payload.do_number, "invoice_no": payload.invoice_no,
        "vendor": received_docs[0].get("vendor_name"),
    })
    return {"received": len(received_docs)}


@router.post("/store/issue/bulk")
async def store_issue_bulk(payload: BulkIssueRequest, current: dict = Depends(require_store_access)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    created = []
    for it in payload.items:
        if it.qty <= 0 or not it.item_name or not it.taker_name.strip():
            continue
        batches = await db.store_receipts.find(
            {"item_name": it.item_name, "qty_remaining": {"$gt": 0}}
        ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)
        avail = sum(b.get("qty_remaining", 0) for b in batches)
        if it.qty > avail + 1e-9:
            raise HTTPException(status_code=400, detail=f"{it.item_name}: stok tidak cukup (tersedia {avail}, diminta {it.qty})")
        remain = float(it.qty)
        allocations = []
        for b in batches:
            if remain <= 1e-9:
                break
            take = min(float(b["qty_remaining"]), remain)
            allocations.append({
                "receipt_id": b["id"],
                "qty": take,
                "unit_price": float(b.get("unit_price", 0)),
                "vendor_name": b.get("vendor_name", ""),
                "receive_date": b.get("receive_date"),
            })
            await db.store_receipts.update_one({"id": b["id"]}, {"$inc": {"qty_remaining": -take}})
            remain -= take
        total_cost = sum(a["qty"] * a["unit_price"] for a in allocations)
        doc = {
            "id": str(uuid.uuid4()),
            "item_name": it.item_name,
            "unit": batches[0].get("unit", "Ea") if batches else "Ea",
            "qty": float(it.qty),
            "issue_date": it.issue_date,
            "taker_name": it.taker_name.strip(),
            "so_number": it.so_number or "",
            "note": it.note or "",
            "allocations": allocations,
            "total_cost": total_cost,
            "avg_unit_price": (total_cost / it.qty) if it.qty else 0,
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        created.append(doc)
    if not created:
        raise HTTPException(status_code=400, detail="Tidak ada item valid")
    await db.store_issuances.insert_many([d.copy() for d in created])
    await log_action(current, "store_issue", "store_issuance", "-", {
        "count": len(created), "first_item": created[0].get("item_name"), "so_number": created[0].get("so_number"),
    })
    return {"issued": len(created)}


@router.post("/store/receive")
async def store_receive(payload: StoreReceiveRequest, current: dict = Depends(require_store_access)):
    tx = await db.transactions.find_one({"id": payload.transaction_id})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaksi asal tidak ditemukan")
    if not tx.get("post_to_store"):
        raise HTTPException(status_code=400, detail="Item ini tidak di-flag ke Store")
    already = await _sum_received_for_tx(payload.transaction_id)
    remaining = float(tx.get("qty", 0)) - already
    if payload.qty_received <= 0:
        raise HTTPException(status_code=400, detail="Qty terima harus > 0")
    if payload.qty_received > remaining + 1e-9:
        raise HTTPException(
            status_code=400,
            detail=f"Qty terima ({payload.qty_received}) melebihi sisa PO ({remaining}). Over-receive tidak diizinkan."
        )

    doc = {
        "id": str(uuid.uuid4()),
        "transaction_id": payload.transaction_id,
        "po_no": tx.get("po_no", ""),
        "invoice_no": tx.get("invoice_no", ""),
        "vendor_name": tx.get("vendor_name", ""),
        "item_name": tx.get("item_name", ""),
        "unit": tx.get("unit", "Ea"),
        "unit_price": float(tx.get("unit_price", 0)),
        "do_number": payload.do_number or "",
        "qty_received": float(payload.qty_received),
        "qty_remaining": float(payload.qty_received),
        "add_to_stock": True,
        "source": "po",
        "receive_date": payload.receive_date,
        "note": payload.note or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_receipts.insert_one(doc.copy())
    await log_action(current, "store_receive", "store_receipt", doc["id"], {
        "item": doc["item_name"], "qty": doc["qty_received"], "po_no": doc["po_no"],
        "do_number": doc["do_number"],
    })
    doc.pop("_id", None)
    if not can_see_prices(current):
        doc.pop("unit_price", None)
    return doc


@router.get("/store/receipts")
async def store_receipts(current: dict = Depends(require_store_access),
                         item_name: Optional[str] = None,
                         transaction_id: Optional[str] = None):
    filt: dict = {}
    if item_name:
        filt["item_name"] = item_name
    if transaction_id:
        filt["transaction_id"] = transaction_id
    docs = await db.store_receipts.find(filt, {"_id": 0}).sort("receive_date", -1).to_list(length=1000)
    if not can_see_prices(current):
        for d in docs:
            d.pop("unit_price", None)
    return docs


@router.get("/store/stock")
async def store_stock(current: dict = Depends(require_store_access),
                      customer_only: bool = False,
                      exclude_customer: bool = False):
    match: dict = {"qty_remaining": {"$gt": 0}}
    if customer_only:
        match["is_customer_material"] = True
    elif exclude_customer:
        match["$or"] = [{"is_customer_material": False}, {"is_customer_material": {"$exists": False}}]
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {"item": "$item_name", "customer": {"$ifNull": ["$is_customer_material", False]}},
            "qty": {"$sum": "$qty_remaining"},
            "unit": {"$first": "$unit"},
            "last_receive_date": {"$max": "$receive_date"},
            "vendors": {"$addToSet": "$vendor_name"},
            "batches": {"$sum": 1},
        }},
        {"$sort": {"_id.item": 1}},
    ]
    docs = await db.store_receipts.aggregate(pipeline).to_list(length=5000)
    return [{
        "item_name": d["_id"]["item"], "qty": d["qty"], "unit": d["unit"],
        "last_receive_date": d.get("last_receive_date"),
        "vendors": d.get("vendors", []), "batches": d.get("batches", 0),
        "is_customer_material": bool(d["_id"].get("customer")),
    } for d in docs if d["_id"].get("item")]


@router.post("/store/issue")
async def store_issue(payload: StoreIssueRequest, current: dict = Depends(require_store_access)):
    if payload.qty <= 0:
        raise HTTPException(status_code=400, detail="Qty keluar harus > 0")
    if not payload.taker_name.strip():
        raise HTTPException(status_code=400, detail="Nama pengambil wajib diisi")

    batches = await db.store_receipts.find(
        {"item_name": payload.item_name, "qty_remaining": {"$gt": 0}}
    ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)

    total_available = sum(b.get("qty_remaining", 0) for b in batches)
    if payload.qty > total_available + 1e-9:
        raise HTTPException(
            status_code=400,
            detail=f"Stok tidak cukup. Tersedia {total_available}, diminta {payload.qty}."
        )

    remaining_to_take = float(payload.qty)
    allocations = []
    for b in batches:
        if remaining_to_take <= 1e-9:
            break
        take = min(float(b["qty_remaining"]), remaining_to_take)
        allocations.append({
            "receipt_id": b["id"],
            "qty": take,
            "unit_price": float(b.get("unit_price", 0)),
            "vendor_name": b.get("vendor_name", ""),
            "receive_date": b.get("receive_date"),
        })
        await db.store_receipts.update_one({"id": b["id"]}, {"$inc": {"qty_remaining": -take}})
        remaining_to_take -= take

    total_cost = sum(a["qty"] * a["unit_price"] for a in allocations)
    doc = {
        "id": str(uuid.uuid4()),
        "item_name": payload.item_name,
        "unit": batches[0].get("unit", "Ea") if batches else "Ea",
        "qty": float(payload.qty),
        "issue_date": payload.issue_date,
        "taker_name": payload.taker_name.strip(),
        "so_number": payload.so_number or "",
        "note": payload.note or "",
        "allocations": allocations,
        "total_cost": total_cost,
        "avg_unit_price": (total_cost / payload.qty) if payload.qty else 0,
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_issuances.insert_one(doc.copy())
    await log_action(current, "store_issue", "store_issuance", doc["id"], {
        "item": doc["item_name"], "qty": doc["qty"], "so_number": doc["so_number"], "taker": doc["taker_name"],
    })
    doc.pop("_id", None)
    if not can_see_prices(current):
        doc.pop("total_cost", None)
        doc.pop("avg_unit_price", None)
        for a in doc.get("allocations", []):
            a.pop("unit_price", None)
    return doc


@router.get("/store/issuances")
async def list_issuances(
    current: dict = Depends(require_store_access),
    q: Optional[str] = None,
    so_number: Optional[str] = None,
    taker: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"taker_name": {"$regex": q, "$options": "i"}},
            {"so_number": {"$regex": q, "$options": "i"}},
        ]
    if so_number:
        filt["so_number"] = {"$regex": so_number, "$options": "i"}
    if taker:
        filt["taker_name"] = {"$regex": taker, "$options": "i"}
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        filt["issue_date"] = d

    total = await db.store_issuances.count_documents(filt)
    cursor = db.store_issuances.find(filt, {"_id": 0}).sort("issue_date", -1).skip((page - 1) * page_size).limit(page_size)
    items = await cursor.to_list(length=page_size)
    hide_price = not can_see_prices(current)
    if hide_price:
        for d in items:
            d.pop("total_cost", None)
            d.pop("avg_unit_price", None)
            for a in d.get("allocations", []):
                a.pop("unit_price", None)
    return {"total": total, "page": page, "page_size": page_size, "items": items, "prices_visible": not hide_price}


@router.get("/store/report/xlsx")
async def store_report_xlsx(
    current: dict = Depends(require_store_access),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    if not can_see_prices(current):
        raise HTTPException(status_code=403, detail="Tidak berwenang melihat harga di laporan Store")
    filt: dict = {}
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        filt["issue_date"] = d
    issuances = await db.store_issuances.find(filt, {"_id": 0}).sort("issue_date", 1).to_list(length=100000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pengeluaran Stok"
    headers = ["Nomor SO", "Tgl Keluar", "Nama Barang", "Qty", "Unit", "Unit Price (FIFO)", "Total Price", "Pengambil", "Vendor Asal", "Note"]
    ws.append(headers)
    for iss in issuances:
        for a in iss.get("allocations", []):
            ws.append([
                iss.get("so_number", ""), iss.get("issue_date", ""), iss.get("item_name", ""),
                float(a.get("qty", 0)), iss.get("unit", ""), float(a.get("unit_price", 0)),
                float(a.get("qty", 0)) * float(a.get("unit_price", 0)),
                iss.get("taker_name", ""), a.get("vendor_name", ""), iss.get("note", ""),
            ])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"laporan_pengeluaran_stok_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------- Edit/Delete Request (approval workflow) ----------------
@router.post("/store/requests")
async def create_store_request(payload: StoreRequestCreate, current: dict = Depends(require_store_access)):
    if payload.target_type not in ("receipt", "issuance"):
        raise HTTPException(status_code=400, detail="target_type harus 'receipt' atau 'issuance'")
    if payload.action_type not in ("edit", "delete"):
        raise HTTPException(status_code=400, detail="action_type harus 'edit' atau 'delete'")
    if not payload.reason.strip():
        raise HTTPException(status_code=400, detail="Alasan wajib diisi")

    coll = db.store_receipts if payload.target_type == "receipt" else db.store_issuances
    target = await coll.find_one({"id": payload.target_id})
    if not target:
        raise HTTPException(status_code=404, detail="Data yang diajukan tidak ditemukan")

    if current.get("role") == "store" and target.get("created_by") != current["id"]:
        raise HTTPException(status_code=403, detail="Hanya bisa mengajukan koreksi untuk data milik sendiri")

    summary = {
        "item_name": target.get("item_name"),
        "qty": target.get("qty") or target.get("qty_received"),
        "issue_date": target.get("issue_date") or target.get("receive_date"),
        "po_no": target.get("po_no"),
        "so_number": target.get("so_number"),
        "do_number": target.get("do_number"),
        "taker_name": target.get("taker_name"),
    }
    doc = {
        "id": str(uuid.uuid4()),
        "target_type": payload.target_type,
        "target_id": payload.target_id,
        "target_summary": summary,
        "action_type": payload.action_type,
        "reason": payload.reason.strip(),
        "proposed_changes": payload.proposed_changes or {},
        "status": "pending",
        "requested_by": current["id"],
        "requested_by_username": current.get("username", ""),
        "requested_at": _now_iso(),
        "reviewed_by": None,
        "reviewed_by_username": None,
        "reviewed_at": None,
        "review_note": "",
    }
    await db.store_requests.insert_one(doc.copy())
    await log_action(current, "store_request", "store_request", doc["id"], {
        "target_type": payload.target_type, "action": payload.action_type,
        "item": summary.get("item_name"), "reason_preview": payload.reason[:80],
    })
    doc.pop("_id", None)
    return doc


@router.get("/store/requests")
async def list_store_requests(
    current: dict = Depends(get_current_user),
    status: Optional[str] = None,
    mine: bool = False,
):
    filt: dict = {}
    if status:
        filt["status"] = status
    if current.get("role") != "admin" or mine:
        filt["requested_by"] = current["id"]
    docs = await db.store_requests.find(filt, {"_id": 0}).sort("requested_at", -1).to_list(length=500)
    return docs


@router.get("/store/requests/pending-count")
async def pending_count(current: dict = Depends(require_approve_perm)):
    n = await db.store_requests.count_documents({"status": "pending"})
    return {"count": n}


@router.post("/store/requests/{req_id}/review")
async def review_store_request(req_id: str, payload: StoreRequestReview, current: dict = Depends(require_approve_perm)):
    req = await db.store_requests.find_one({"id": req_id})
    if not req:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail=f"Request sudah di-{req['status']}")

    new_status = "approved" if payload.approve else "rejected"

    if payload.approve and req["action_type"] == "delete":
        if req["target_type"] == "issuance":
            iss = await db.store_issuances.find_one({"id": req["target_id"]})
            if iss:
                for a in iss.get("allocations", []):
                    await db.store_receipts.update_one(
                        {"id": a.get("receipt_id")},
                        {"$inc": {"qty_remaining": float(a.get("qty", 0))}}
                    )
                await db.store_issuances.delete_one({"id": req["target_id"]})
        elif req["target_type"] == "receipt":
            rec = await db.store_receipts.find_one({"id": req["target_id"]})
            if rec:
                consumed = float(rec.get("qty_received", 0)) - float(rec.get("qty_remaining", 0))
                if consumed > 1e-9:
                    raise HTTPException(status_code=400, detail=f"Tidak bisa hapus receipt: {consumed} unit sudah dipakai (issuance). Batalkan issuance dulu.")
                await db.store_receipts.delete_one({"id": req["target_id"]})

    await db.store_requests.update_one(
        {"id": req_id},
        {"$set": {
            "status": new_status,
            "reviewed_by": current["id"],
            "reviewed_by_username": current.get("username", ""),
            "reviewed_at": _now_iso(),
            "review_note": payload.review_note or "",
        }}
    )
    await log_action(current, "review_store_request", "store_request", req_id, {
        "decision": new_status, "target_type": req["target_type"], "action": req["action_type"],
        "item": req.get("target_summary", {}).get("item_name"),
    })
    return {"status": new_status}


# ---------------- Manual Store Receipt ----------------
@router.post("/store/receive/manual")
async def store_receive_manual(payload: ManualReceiveRequest, current: dict = Depends(require_store_write)):
    if payload.qty <= 0:
        raise HTTPException(status_code=400, detail="Qty harus > 0")
    if payload.source_type not in ("customer", "supplier"):
        raise HTTPException(status_code=400, detail="source_type harus 'customer' atau 'supplier'")
    if not payload.source_name.strip():
        raise HTTPException(status_code=400, detail="Nama customer/supplier wajib")
    if not payload.item_name.strip():
        raise HTTPException(status_code=400, detail="Nama barang wajib")
    is_customer = payload.source_type == "customer"
    doc = {
        "id": str(uuid.uuid4()),
        "transaction_id": None,
        "source": "manual",
        "source_type": payload.source_type,
        "is_customer_material": is_customer,
        "po_no": payload.po_no or "",
        "invoice_no": "",
        "vendor_name": payload.source_name.strip(),
        "customer_name": payload.source_name.strip() if is_customer else "",
        "item_name": payload.item_name.strip(),
        "unit": payload.unit or "Ea",
        "unit_price": float(payload.unit_price or 0),
        "do_number": payload.do_no or "",
        "so_no": payload.so_no or "",
        "qty_received": float(payload.qty),
        "qty_remaining": float(payload.qty),
        "receive_date": payload.receive_date,
        "mcl_done": bool(payload.mcl_done),
        "mif_done": bool(payload.mif_done),
        "note": payload.remark or "",
        "created_by": current["id"],
        "created_by_username": current.get("username", ""),
        "created_at": _now_iso(),
    }
    await db.store_receipts.insert_one(doc.copy())
    await log_action(current, "store_receive_manual", "store_receipt", doc["id"], {
        "source": payload.source_type, "source_name": doc["vendor_name"],
        "item": doc["item_name"], "qty": doc["qty_received"],
    })
    doc.pop("_id", None)
    if not can_see_prices(current):
        doc.pop("unit_price", None)
    return doc


@router.patch("/store/receipts/{rid}/flags")
async def update_receipt_flags(rid: str, payload: dict, current: dict = Depends(require_store_write)):
    rec = await db.store_receipts.find_one({"id": rid})
    if not rec:
        raise HTTPException(status_code=404, detail="Receipt tidak ditemukan")
    upd: dict = {}
    if "mcl_done" in payload:
        upd["mcl_done"] = bool(payload["mcl_done"])
    if "mif_done" in payload:
        upd["mif_done"] = bool(payload["mif_done"])
    # Admin/store can toggle add_to_stock directly (no request/approval).
    if "add_to_stock" in payload:
        new_val = bool(payload["add_to_stock"])
        consumed = float(rec.get("qty_received", 0)) - float(rec.get("qty_remaining", 0))
        if not new_val:
            # Turning OFF: only allowed if nothing has been consumed (else stock has been used)
            if consumed > 1e-9:
                raise HTTPException(status_code=400, detail=f"Tidak bisa hilangkan dari stok: {consumed} unit sudah dipakai (issuance).")
            upd["add_to_stock"] = False
            upd["qty_remaining"] = 0.0
        else:
            upd["add_to_stock"] = True
            upd["qty_remaining"] = float(rec.get("qty_received", 0))
    if upd:
        await db.store_receipts.update_one({"id": rid}, {"$set": upd})
        await log_action(current, "update_receipt_flags", "store_receipt", rid, upd)
    return {"ok": True, "flags": upd}


# ---------------- Input Incoming Goods (multi-item manual receipt) ----------------
@router.post("/store/incoming")
async def store_incoming(payload: IncomingGoodsRequest, current: dict = Depends(require_store_write)):
    """Multi-item manual receiving. Replaces single-item /store/receive/manual.
    Each item can be flagged add_to_stock=True (masuk stok, tracked via qty_remaining)
    or False (habis pakai, qty_remaining=0 but still logged for Incoming Goods report)."""
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    if payload.source_type not in ("customer", "supplier"):
        raise HTTPException(status_code=400, detail="source_type harus 'customer' atau 'supplier'")
    if not payload.source_name.strip():
        raise HTTPException(status_code=400, detail="Nama customer/supplier wajib")
    is_customer = payload.source_type == "customer"
    docs = []
    for it in payload.items:
        if it.qty <= 0 or not it.item_name.strip():
            continue
        add_stock = True if it.add_to_stock is None else bool(it.add_to_stock)
        docs.append({
            "id": str(uuid.uuid4()),
            "transaction_id": None,
            "source": "manual",
            "source_type": payload.source_type,
            "is_customer_material": is_customer,
            "po_no": payload.po_no or "",
            "invoice_no": "",
            "vendor_name": payload.source_name.strip(),
            "customer_name": payload.source_name.strip() if is_customer else "",
            "item_name": it.item_name.strip(),
            "unit": it.unit or "Ea",
            "unit_price": float(it.unit_price or 0),
            "do_number": payload.do_no or "",
            "so_no": it.so_no or "",
            "qty_received": float(it.qty),
            "qty_remaining": float(it.qty) if add_stock else 0.0,
            "add_to_stock": add_stock,
            "receive_date": payload.receive_date,
            "mcl_done": False,  # set later via /flags (in Incoming Goods report)
            "mif_done": False,
            "note": it.remark or "",
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        })
    if not docs:
        raise HTTPException(status_code=400, detail="Tidak ada item valid")
    await db.store_receipts.insert_many([d.copy() for d in docs])
    await log_action(current, "store_incoming", "store_receipt", "-", {
        "count": len(docs), "source": payload.source_type,
        "source_name": docs[0]["vendor_name"],
    })
    return {"received": len(docs)}


# ---------------- Incoming Goods Report ----------------
@router.get("/store/incoming-report")
async def incoming_report(
    current: dict = Depends(require_store_access),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    source: Optional[str] = None,  # 'po' | 'manual'
    q: Optional[str] = None,
    page: int = 1,
    page_size: int = 100,
):
    """Unified report of ALL incoming goods (from PO purchasing + manual)."""
    filt: dict = {}
    if source in ("po", "manual"):
        filt["source"] = source
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date
        filt["receive_date"] = rng
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"vendor_name": {"$regex": q, "$options": "i"}},
            {"po_no": {"$regex": q, "$options": "i"}},
            {"invoice_no": {"$regex": q, "$options": "i"}},
            {"do_number": {"$regex": q, "$options": "i"}},
        ]
    total = await db.store_receipts.count_documents(filt)
    cursor = db.store_receipts.find(filt, {"_id": 0}).sort("receive_date", -1).skip((page - 1) * page_size).limit(page_size)
    items = await cursor.to_list(length=page_size)
    if not can_see_prices(current):
        for d in items:
            d.pop("unit_price", None)
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.get("/store/incoming-report/xlsx")
async def incoming_report_xlsx(
    current: dict = Depends(require_store_access),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    source: Optional[str] = None,
    q: Optional[str] = None,
):
    """Excel export of Incoming Goods report."""
    filt: dict = {}
    if source in ("po", "manual"):
        filt["source"] = source
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date
        filt["receive_date"] = rng
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"vendor_name": {"$regex": q, "$options": "i"}},
            {"po_no": {"$regex": q, "$options": "i"}},
            {"invoice_no": {"$regex": q, "$options": "i"}},
            {"do_number": {"$regex": q, "$options": "i"}},
        ]
    docs = await db.store_receipts.find(filt, {"_id": 0}).sort("receive_date", -1).to_list(length=100000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Incoming Goods"
    headers = ["Tgl Terima", "Sumber", "Vendor/Customer", "Nama Barang", "Qty", "Unit",
               "PO No", "DO No", "Invoice No", "Ke Stok?", "MCL", "MIF", "Catatan"]
    ws.append(headers)
    for d in docs:
        src_label = "PO" if d.get("source") == "po" else ("Customer" if d.get("is_customer_material") else "Supplier")
        ws.append([
            d.get("receive_date", ""), src_label, d.get("vendor_name", ""),
            d.get("item_name", ""), float(d.get("qty_received", 0)), d.get("unit", ""),
            d.get("po_no", ""), d.get("do_number", ""), d.get("invoice_no", ""),
            "Ya" if d.get("add_to_stock", True) else "Tidak",
            "Ya" if d.get("mcl_done") else "Tidak",
            "Ya" if d.get("mif_done") else "Tidak",
            d.get("note", ""),
        ])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"incoming_goods_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/store/issuances/xlsx")
async def issuances_xlsx(
    current: dict = Depends(require_store_access),
    q: Optional[str] = None,
    so_number: Optional[str] = None,
    taker: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """Excel export of Keluar Barang (issuances). Store role: no prices; admin/finance: with FIFO cost."""
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"item_name": {"$regex": q, "$options": "i"}},
            {"taker_name": {"$regex": q, "$options": "i"}},
            {"so_number": {"$regex": q, "$options": "i"}},
        ]
    if so_number:
        filt["so_number"] = {"$regex": so_number, "$options": "i"}
    if taker:
        filt["taker_name"] = {"$regex": taker, "$options": "i"}
    if start_date or end_date:
        rng: dict = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date
        filt["issue_date"] = rng
    docs = await db.store_issuances.find(filt, {"_id": 0}).sort("issue_date", -1).to_list(length=100000)

    show_prices = can_see_prices(current)
    wb = Workbook()
    ws = wb.active
    ws.title = "Keluar Barang"
    headers = ["Tgl Keluar", "Nama Barang", "Qty", "Unit", "Nomor SO", "Pengambil", "Catatan"]
    if show_prices:
        headers += ["Avg Unit Price (FIFO)", "Total Cost"]
    ws.append(headers)
    for d in docs:
        row = [
            d.get("issue_date", ""), d.get("item_name", ""),
            float(d.get("qty", 0)), d.get("unit", ""),
            d.get("so_number", ""), d.get("taker_name", ""),
            d.get("note", ""),
        ]
        if show_prices:
            row += [float(d.get("avg_unit_price", 0)), float(d.get("total_cost", 0))]
        ws.append(row)
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"keluar_barang_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------- Production Issue (Customer material → Produksi) ----------------
@router.post("/store/issue/production")
async def store_issue_production(payload: ProductionIssueRequest, current: dict = Depends(require_store_write)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Tidak ada item")
    if not payload.taker_name.strip():
        raise HTTPException(status_code=400, detail="Nama penerima produksi wajib")
    created = []
    for it in payload.items:
        if it.qty <= 0 or not it.item_name:
            continue
        batches = await db.store_receipts.find(
            {"item_name": it.item_name, "is_customer_material": True, "qty_remaining": {"$gt": 0}}
        ).sort([("receive_date", 1), ("created_at", 1)]).to_list(length=1000)
        avail = sum(b.get("qty_remaining", 0) for b in batches)
        if it.qty > avail + 1e-9:
            raise HTTPException(status_code=400, detail=f"{it.item_name}: stok Customer tidak cukup (tersedia {avail})")
        remain = float(it.qty)
        allocations = []
        for b in batches:
            if remain <= 1e-9:
                break
            take = min(float(b["qty_remaining"]), remain)
            allocations.append({
                "receipt_id": b["id"], "qty": take,
                "unit_price": float(b.get("unit_price", 0)),
                "vendor_name": b.get("vendor_name", ""),
                "customer_name": b.get("customer_name", ""),
                "receive_date": b.get("receive_date"),
            })
            await db.store_receipts.update_one({"id": b["id"]}, {"$inc": {"qty_remaining": -take}})
            remain -= take
        doc = {
            "id": str(uuid.uuid4()),
            "type": "production",
            "is_customer_material": True,
            "item_name": it.item_name,
            "unit": batches[0].get("unit", "Ea") if batches else "Ea",
            "qty": float(it.qty),
            "issue_date": payload.issue_date,
            "taker_name": payload.taker_name.strip(),
            "so_number": it.so_number or "",
            "note": it.note or "",
            "allocations": allocations,
            "total_cost": 0,
            "avg_unit_price": 0,
            "created_by": current["id"],
            "created_by_username": current.get("username", ""),
            "created_at": _now_iso(),
        }
        created.append(doc)
    if not created:
        raise HTTPException(status_code=400, detail="Tidak ada item valid")
    await db.store_issuances.insert_many([d.copy() for d in created])
    await log_action(current, "store_issue_production", "store_issuance", "-", {
        "count": len(created), "so_number": created[0].get("so_number"),
        "first_item": created[0].get("item_name"), "taker": payload.taker_name,
    })
    return {"issued": len(created)}
