"""Controlled Document Register — modul Document Control (Salma).

Menyimpan dokumen NON-drawing (mis. dokumen ISO: Prosedur, Manual, Instruksi Kerja, Form)
yang diinput manual oleh Document Control lalu di-STAMP DC di sistem (sama seperti drawing:
buka PDF viewer, klik posisi stamp).

Lifecycle status:
    pending    → sudah diupload, MENUNGGU stamp DC
    controlled → sudah di-stamp DC (masuk register aktif)
    obsolete   → versi lama setelah ada revisi baru (otomatis diberi cap OBSOLETE, view-only)

Revisi:
    POST /controlled-documents/{id}/new-revision → buat dokumen baru rev+1 (status pending),
    versi lama otomatis jadi 'obsolete'.

Stamp DC tidak mengubah file asli — cap di-overlay on-the-fly saat preview/print (sama pola drawing).
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from typing import Optional, List

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import Response, StreamingResponse
from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from pydantic import BaseModel

from db import db
from deps import get_current_user, is_admin_like, is_doc_control

router = APIRouter(tags=["controlled_documents"])

_fs: AsyncIOMotorGridFSBucket | None = None


def _bucket() -> AsyncIOMotorGridFSBucket:
    global _fs
    if _fs is None:
        _fs = AsyncIOMotorGridFSBucket(db, bucket_name="controlled_docs")
    return _fs


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _clean(doc: dict) -> dict:
    doc.pop("_id", None)
    return doc


def _can_manage(user: dict) -> bool:
    return is_doc_control(user) or is_admin_like(user)


def _require_manage(user: dict):
    if not _can_manage(user):
        raise HTTPException(status_code=403, detail="Hanya Document Control atau Admin yang boleh mengelola dokumen terkontrol")


def _is_pdf(filename: str, content_type: str) -> bool:
    return "pdf" in (content_type or "").lower() or (filename or "").lower().endswith(".pdf")


def _norm_placements(pls) -> list:
    out = []
    for pl in pls or []:
        if not pl:
            continue
        try:
            out.append({
                "page": int(pl.get("page")) if pl.get("page") is not None else None,
                "x": float(pl.get("x")),
                "y": float(pl.get("y")),
                "size": pl.get("size") or "M",
            })
        except (TypeError, ValueError):
            continue
    return out


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class StampIn(BaseModel):
    notes: str = ""
    stamp_x: Optional[float] = None
    stamp_y: Optional[float] = None
    placements: Optional[List[dict]] = None


# ---------------------------------------------------------------------------
# GridFS helpers
# ---------------------------------------------------------------------------
async def _store_file(file: UploadFile) -> dict:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File kosong")
    fid = await _bucket().upload_from_stream(file.filename or "document.pdf", io.BytesIO(raw))
    return {
        "file_id": str(fid),
        "filename": file.filename or "document.pdf",
        "content_type": file.content_type or "application/octet-stream",
        "size": len(raw),
    }


async def _read_raw(file_id: str) -> bytes:
    try:
        stream = await _bucket().open_download_stream(ObjectId(file_id))
        return await stream.read()
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File tidak bisa dibaca: {e}")


async def _get_doc(doc_id: str) -> dict:
    doc = await db.controlled_documents.find_one({"id": doc_id, "deleted_at": {"$exists": False}}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Dokumen tidak ditemukan")
    return doc


# ---------------------------------------------------------------------------
# CRUD / list
# ---------------------------------------------------------------------------
@router.post("/controlled-documents")
async def create_controlled_document(
    doc_no: str = Form(...),
    title: str = Form(...),
    category: str = Form("iso"),
    doc_type: str = Form(""),
    notes: str = Form(""),
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    """Upload dokumen baru. Untuk kategori ISO → langsung diarsipkan sebagai 'controlled'
    (DC tidak lagi men-stamp dokumen ISO, hanya mengarsipkan ke database)."""
    _require_manage(current)
    stored = await _store_file(file)
    now = _now_iso()
    is_iso = (category or "iso").strip().lower() == "iso"
    doc = {
        "id": str(uuid.uuid4()),
        "doc_no": doc_no.strip(),
        "title": title.strip(),
        "category": (category or "iso").strip().lower(),
        "doc_type": (doc_type or "").strip(),
        "notes": (notes or "").strip(),
        "revision": 0,
        "rev_label": "Rev-0",
        "status": "controlled" if is_iso else "pending",
        "dc_stamp": None,
        "supersedes": None,
        "superseded_by": None,
        "uploaded_by": {"id": current.get("id"), "name": current.get("name") or current.get("username")},
        "created_at": now,
        "updated_at": now,
        "controlled_at": now if is_iso else None,
        "obsoleted_at": None,
        **stored,
    }
    await db.controlled_documents.insert_one(dict(doc))
    return _clean(dict(doc))


@router.get("/controlled-documents")
async def list_controlled_documents(
    category: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    current: dict = Depends(get_current_user),
):
    """List dokumen terkontrol. Filter opsional: category (iso), status (pending/controlled/obsolete), q (search)."""
    query: dict = {"deleted_at": {"$exists": False}}
    if category:
        query["category"] = category.lower()
    if status:
        query["status"] = status
    if q:
        rx = {"$regex": q, "$options": "i"}
        query["$or"] = [{"doc_no": rx}, {"title": rx}, {"doc_type": rx}]
    items = await db.controlled_documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=1000)
    return {"items": items, "total": len(items)}


@router.get("/controlled-documents/doc-types")
async def controlled_document_doc_types(category: str = "iso", current: dict = Depends(get_current_user)):
    """Daftar Tipe Dokumen yang pernah dipakai (untuk dropdown upload — tipe custom 'Lainnya'
    otomatis tersimpan & muncul lagi di upload berikutnya)."""
    query: dict = {"deleted_at": {"$exists": False}}
    if category:
        query["category"] = category.lower()
    vals = await db.controlled_documents.distinct("doc_type", query)
    types = sorted({(v or "").strip() for v in vals if v and v.strip()})
    return {"types": types}


@router.get("/controlled-documents/counts")
async def controlled_document_counts(current: dict = Depends(get_current_user)):
    """Jumlah per status (untuk badge/kartu)."""
    base = {"deleted_at": {"$exists": False}}
    pending = await db.controlled_documents.count_documents({**base, "status": "pending"})
    controlled = await db.controlled_documents.count_documents({**base, "status": "controlled"})
    obsolete = await db.controlled_documents.count_documents({**base, "status": "obsolete"})
    return {"pending": pending, "controlled": controlled, "obsolete": obsolete}


@router.get("/controlled-documents/export")
async def export_controlled_documents(
    category: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    format: str = "xlsx",
    current: dict = Depends(get_current_user),
):
    """Ekspor daftar dokumen terkontrol (untuk audit ISO) ke Excel (.xlsx) atau PDF."""
    fmt = (format or "xlsx").lower()
    if fmt not in ("xlsx", "pdf"):
        raise HTTPException(status_code=400, detail="Format harus 'xlsx' atau 'pdf'")

    query: dict = {"deleted_at": {"$exists": False}}
    if category:
        query["category"] = category.lower()
    if status:
        query["status"] = status
    if q:
        rx = {"$regex": q, "$options": "i"}
        query["$or"] = [{"doc_no": rx}, {"title": rx}, {"doc_type": rx}]
    docs = await db.controlled_documents.find(query, {"_id": 0}).sort("doc_no", 1).to_list(length=5000)

    st_label = {"pending": "Menunggu Stamp DC", "controlled": "Controlled", "obsolete": "Obsolete"}

    def _fmt(iso: str) -> str:
        if not iso:
            return "-"
        try:
            from datetime import timedelta
            d = datetime.fromisoformat(iso.replace("Z", "+00:00")) + timedelta(hours=7)
            return d.strftime("%d %b %Y %H:%M")
        except Exception:
            return (iso or "")[:16].replace("T", " ")

    rows = []
    for d in docs:
        dc = d.get("dc_stamp") or {}
        rows.append({
            "doc_no": d.get("doc_no", "-"),
            "title": d.get("title", "-"),
            "doc_type": d.get("doc_type", "-"),
            "rev_label": d.get("rev_label", "-"),
            "status": st_label.get(d.get("status"), d.get("status", "-")),
            "uploaded_by": (d.get("uploaded_by") or {}).get("name", "-"),
            "created": _fmt(d.get("created_at")),
            "controlled": _fmt(d.get("controlled_at")),
            "stamped_by": dc.get("name", "-") if dc else "-",
            "notes": d.get("notes", ""),
        })

    now = datetime.now(timezone.utc)
    from datetime import timedelta as _td
    wib_now = (now + _td(hours=7)).strftime("%d %b %Y %H:%M")
    exported_by = current.get("name") or current.get("username") or "-"
    scope_status = st_label.get(status, "Semua Status") if status else "Semua Status"
    scope = f"{(category or 'semua').upper()} · {scope_status}"
    meta_line = f"{scope} | Total {len(rows)} dokumen | Diekspor oleh {exported_by} | {wib_now} WIB"
    ts = now.strftime("%Y%m%d_%H%M")

    if fmt == "xlsx":
        data = _build_docs_xlsx(rows, meta_line)
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="ControlledDocumentRegister_{ts}.xlsx"'},
        )
    data = _build_docs_pdf(rows, meta_line)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="ControlledDocumentRegister_{ts}.pdf"'},
    )


@router.get("/controlled-documents/{doc_id}")
async def get_controlled_document(doc_id: str, current: dict = Depends(get_current_user)):
    return await _get_doc(doc_id)


@router.delete("/controlled-documents/{doc_id}")
async def delete_controlled_document(doc_id: str, current: dict = Depends(get_current_user)):
    """Hapus (soft delete). Hanya Admin. Untuk membersihkan dokumen salah upload."""
    if not is_admin_like(current):
        raise HTTPException(status_code=403, detail="Hanya Admin yang boleh menghapus")
    doc = await _get_doc(doc_id)
    await db.controlled_documents.update_one({"id": doc_id}, {"$set": {"deleted_at": _now_iso()}})
    return {"success": True}


# ---------------------------------------------------------------------------
# DC Stamp (pending → controlled)
# ---------------------------------------------------------------------------
@router.post("/controlled-documents/{doc_id}/stamp-controlled")
async def stamp_controlled_document(doc_id: str, payload: StampIn, current: dict = Depends(get_current_user)):
    """Stamp DC pada dokumen (placements dari PDF viewer). Status pending → controlled."""
    _require_manage(current)
    doc = await _get_doc(doc_id)
    if doc.get("status") == "obsolete":
        raise HTTPException(status_code=409, detail="Dokumen sudah OBSOLETE, tidak bisa di-stamp")
    if not _is_pdf(doc.get("filename"), doc.get("content_type")):
        raise HTTPException(status_code=400, detail="Stamp DC hanya untuk file PDF")

    stamp = {
        "name": current.get("name") or current.get("username"),
        "user_id": current.get("id"),
        "username": current.get("username"),
        "role": current.get("role"),
        "at": _now_iso(),
        "notes": payload.notes or "",
        "control_no": doc.get("doc_no"),
    }
    if payload.stamp_x is not None:
        stamp["x"] = float(payload.stamp_x)
        stamp["y"] = float(payload.stamp_y or 0.15)
    pls = _norm_placements(payload.placements)
    if pls:
        stamp["placements"] = pls
        stamp["x"] = pls[0]["x"]
        stamp["y"] = pls[0]["y"]
    if "x" not in stamp:
        raise HTTPException(status_code=400, detail="Posisi stamp wajib (klik pada dokumen)")

    now = _now_iso()
    await db.controlled_documents.update_one(
        {"id": doc_id},
        {"$set": {"dc_stamp": stamp, "status": "controlled", "controlled_at": now, "updated_at": now}},
    )
    return await _get_doc(doc_id)


# ---------------------------------------------------------------------------
# New revision (old → obsolete, new → pending)
# ---------------------------------------------------------------------------
@router.post("/controlled-documents/{doc_id}/new-revision")
async def new_revision_controlled_document(
    doc_id: str,
    notes: str = Form(""),
    file: UploadFile = File(...),
    current: dict = Depends(get_current_user),
):
    """Upload revisi baru. Versi lama otomatis OBSOLETE, versi baru status pending (perlu stamp lagi)."""
    _require_manage(current)
    old = await _get_doc(doc_id)
    if old.get("status") == "obsolete":
        raise HTTPException(status_code=409, detail="Dokumen ini sudah OBSOLETE")

    stored = await _store_file(file)
    now = _now_iso()
    new_rev = int(old.get("revision", 0)) + 1
    new_doc = {
        "id": str(uuid.uuid4()),
        "doc_no": old.get("doc_no"),
        "title": old.get("title"),
        "category": old.get("category", "iso"),
        "doc_type": old.get("doc_type", ""),
        "notes": (notes or "").strip(),
        "revision": new_rev,
        "rev_label": f"Rev-{new_rev}",
        "status": "pending",
        "dc_stamp": None,
        "supersedes": old["id"],
        "superseded_by": None,
        "uploaded_by": {"id": current.get("id"), "name": current.get("name") or current.get("username")},
        "created_at": now,
        "updated_at": now,
        "controlled_at": None,
        "obsoleted_at": None,
        **stored,
    }
    await db.controlled_documents.insert_one(dict(new_doc))
    # versi lama → obsolete
    await db.controlled_documents.update_one(
        {"id": old["id"]},
        {"$set": {
            "status": "obsolete",
            "obsoleted_at": now,
            "superseded_by": new_doc["id"],
            "obsolete_reason": f"Digantikan oleh Rev-{new_rev}",
            "updated_at": now,
        }},
    )
    return _clean(dict(new_doc))


# ---------------------------------------------------------------------------
# Preview (image-based) + download
# ---------------------------------------------------------------------------
def _render_bytes(doc: dict, raw: bytes) -> bytes:
    """Overlay stamp DC (kalau ada) + cap OBSOLETE (kalau status obsolete)."""
    from utils.pdf_stamper import apply_stamps, apply_obsolete
    out = raw
    if doc.get("dc_stamp"):
        out = apply_stamps(out, dc_stamp=doc["dc_stamp"])
    if doc.get("status") == "obsolete":
        out = apply_obsolete(out)
    return out


@router.get("/controlled-documents/{doc_id}/page-meta")
async def controlled_doc_page_meta(doc_id: str, current: dict = Depends(get_current_user)):
    doc = await _get_doc(doc_id)
    if not _is_pdf(doc.get("filename"), doc.get("content_type")):
        raise HTTPException(status_code=400, detail="Preview gambar hanya untuk PDF")
    raw = await _read_raw(doc["file_id"])
    from utils.pdf_render import pdf_page_meta
    return pdf_page_meta(raw)


@router.get("/controlled-documents/{doc_id}/page-image")
async def controlled_doc_page_image(
    doc_id: str, page: int = 0, scale: float = 2.0, stamped: bool = True,
    current: dict = Depends(get_current_user),
):
    doc = await _get_doc(doc_id)
    if not _is_pdf(doc.get("filename"), doc.get("content_type")):
        raise HTTPException(status_code=400, detail="Preview gambar hanya untuk PDF")
    raw = await _read_raw(doc["file_id"])
    if stamped:
        raw = _render_bytes(doc, raw)
    from utils.pdf_render import pdf_page_png
    try:
        png = pdf_page_png(raw, page=page, scale=scale)
    except IndexError:
        raise HTTPException(status_code=404, detail="Halaman tidak ada")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"File bukan PDF valid: {e}")
    return Response(content=png, media_type="image/png", headers={"Cache-Control": "no-store"})


@router.get("/controlled-documents/{doc_id}/download")
async def controlled_doc_download(doc_id: str, current: dict = Depends(get_current_user)):
    """Download versi ter-stamp. Hanya Document Control / Admin."""
    _require_manage(current)
    doc = await _get_doc(doc_id)
    raw = await _read_raw(doc["file_id"])
    if _is_pdf(doc.get("filename"), doc.get("content_type")):
        raw = _render_bytes(doc, raw)
    return StreamingResponse(
        io.BytesIO(raw),
        media_type=doc.get("content_type") or "application/pdf",
        headers={"Content-Disposition": f'inline; filename="{doc.get("doc_no","doc")}-{doc.get("rev_label","")}.pdf"'},
    )


# ---------------------------------------------------------------------------
# Export builders (Excel / PDF) — audit ISO
# ---------------------------------------------------------------------------
_EXPORT_COLS = [
    ("doc_no", "No. Dokumen"),
    ("title", "Judul Dokumen"),
    ("doc_type", "Tipe"),
    ("rev_label", "Revisi"),
    ("status", "Status"),
    ("uploaded_by", "Dibuat Oleh"),
    ("created", "Tgl Dibuat"),
    ("controlled", "Tgl Controlled"),
    ("stamped_by", "Di-stamp Oleh"),
    ("notes", "Catatan"),
]


def _build_docs_xlsx(rows: list, meta_line: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Controlled Documents"
    ws.cell(row=1, column=1, value="Controlled Document Register — Audit ISO").font = Font(bold=True, size=13, color="1E293B")
    ws.cell(row=2, column=1, value=meta_line).font = Font(size=9, italic=True, color="64748B")

    header_row = 4
    header_fill = PatternFill("solid", fgColor="7F1D1D")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, (_, label) in enumerate(_EXPORT_COLS, start=1):
        c = ws.cell(row=header_row, column=ci, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    for ri, row in enumerate(rows, start=header_row + 1):
        for ci, (key, _) in enumerate(_EXPORT_COLS, start=1):
            c = ws.cell(row=ri, column=ci, value=row.get(key, ""))
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
    widths = [20, 34, 14, 9, 18, 18, 16, 16, 18, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_docs_pdf(rows: list, meta_line: str) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=8)
    head_style = ParagraphStyle("head", parent=styles["Normal"], fontSize=7.5, leading=9,
                                textColor=colors.white, fontName="Helvetica-Bold")
    elems = [
        Paragraph("Controlled Document Register — Audit ISO",
                  ParagraphStyle("t", parent=styles["Title"], fontSize=15, spaceAfter=2)),
        Paragraph(meta_line, ParagraphStyle("m", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"))),
        Spacer(1, 6),
    ]
    data = [[Paragraph(label, head_style) for _, label in _EXPORT_COLS]]
    for row in rows:
        data.append([Paragraph(str(row.get(key, "") or "-"), cell_style) for key, _ in _EXPORT_COLS])
    col_widths = [w * mm for w in [26, 46, 18, 12, 24, 24, 22, 22, 24, 40]]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F1D1D")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FEF2F2")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elems.append(table)
    doc.build(elems)
    return buf.getvalue()

